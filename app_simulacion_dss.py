# -*- coding: utf-8 -*-
"""
Dashboard Streamlit · DSS Simulación 2026 — versión mejorada
Autor : JFRodriguez · Hidrólogo / Oceanógrafo Físico · ACP-HIMH

Mejoras respecto a versión anterior
────────────────────────────────────
1. Pestaña "Aportes DSS" dedicada con promedio diario para TODAS las
   probabilidades de excedencia (P90…P5) en cfs / m³/s / hm³/d.
2. Selector de unidad de caudal global (cfs | m³/s | hm³/d) que aplica
   a AP, V y cualquier otra variable de flujo en toda la app.
3. Todas las pestañas presentes:
      · GATÚN Px DSS
      · ALHAJUELA Px DSS
      · Aportes DSS  (nueva — análisis de aportes con promedio diario)
      · Comparativo  (nueva — Gatún vs Alhajuela)
4. Gráficas de "abanico" de probabilidades de excedencia con relleno
   entre P90 y P10 (banda de incertidumbre).
5. Tabla resumen diaria descargable en CSV.

Ejecución (Windows)
────────────────────
    py -m pip install streamlit openpyxl plotly pandas numpy
    py -m streamlit run app_simulacion_dss.py
"""
from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

try:
    import plotly.express as px
    import plotly.graph_objects as go
    PLOTLY_OK = True
except ImportError:
    PLOTLY_OK = False

# ─────────────────────────────────────────────────────────────────────
# Configuración
# ─────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="DSS Simulación 2026",
    page_icon="💧",
    layout="wide",
    initial_sidebar_state="expanded",
)

APP_TITLE   = "DSS Simulación 2026 · Análisis de Embalses"
AUTHOR_NOTE = "JFRodriguez · Hidrólogo / Oceanógrafo Físico · ACP-HIMH"

DSS_DEFAULT_NAMES = [
    "SimulacionDSS_2026.xlsx",
    "SimulacionDSS_2026(2).xlsx",
    "SimulacionDSS_2026(1).xlsx",
]
LAKEHOUSE_DEFAULT_NAMES = [
    "LakeHouse_Data.xlsx",
    "LakeHouse_Data(5).xlsx",
    "LakeHouse_Data(4).xlsx",
]

# Conversiones de caudal desde cfs
CFS_TO_M3S      = 0.028316846592
CFS_TO_HM3_DAY  = CFS_TO_M3S * 86400 / 1_000_000   # ≈ 0.002446575

UNIT_LABELS = {"cfs": "cfs", "m³/s": "m³/s", "hm³/d": "hm³/d"}


def unit_display(unit: str) -> str:
    """Etiqueta visible de unidades.

    Internamente se usa `cfs` para las conversiones, pero en la app
    se muestra también como PCS para mantener el lenguaje operativo local.
    """
    return "PCS (cfs)" if unit == "cfs" else unit

RESERVOIR_CONFIG: Dict[str, Dict] = {
    "gatun": {
        "sheet":        "GATUN Px DSS",
        "token":        "GAT",
        "lkh_col":      "actgatel",
        "lkh_label":    "Nivel obs. LKH (ft)",
        "level_unit":   "ft PLD",
        "name":         "Gatún",
        "color_base":   "#0066cc",
    },
    "alhajuela": {
        "sheet":        "ALHAJUELA Px DSS ",
        "token":        "ALH",
        "lkh_col":      "actmadel",
        "lkh_label":    "Nivel obs. LKH (ft)",
        "level_unit":   "ft PLD",
        "name":         "Alhajuela / Madden",
        "color_base":   "#cc6600",
    },
}

# Paleta para probabilidades de excedencia (P90=azul oscuro … P5=rojo)
EXCEEDANCE_COLORS = {
    90: "#003f88", 80: "#005f99", 70: "#0077b6",
    60: "#0096c7", 50: "#48cae4", 40: "#90e0ef",
    30: "#f4a261", 20: "#e76f51", 10: "#d62828", 5: "#9b2226",
}

PERCENTILE_ORDER = [90, 80, 70, 60, 50, 40, 30, 20, 10, 5]

# ─────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────
def inject_css() -> None:
    st.markdown("""
    <style>
    .main-title{font-size:2.1rem;font-weight:900;color:#003E69;margin-bottom:.05rem}
    .sub-title{color:#5f6b7a;font-size:.98rem;margin-bottom:.9rem}
    div[data-testid="metric-container"]{
        border:1px solid rgba(0,62,105,.15);border-radius:12px;
        padding:.65rem .8rem;background:rgba(248,250,252,.85)}
    .unit-badge{
        display:inline-block;background:#003E69;color:#fff;
        border-radius:8px;padding:2px 10px;font-size:.82rem;font-weight:700}
    </style>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────
# Utilidades de archivo
# ─────────────────────────────────────────────────────────────────────
def find_default_file(candidates: List[str]) -> Optional[Path]:
    base = Path(__file__).resolve().parent
    for name in candidates:
        p = base / name
        if p.exists():
            return p
    return None


@st.cache_data(show_spinner=False)
def bytes_from_path(path_str: str, mtime_ns: int, size_bytes: int) -> bytes:
    """Lee bytes de archivo con invalidación por fecha de modificación y tamaño.

    Esto es importante para Lake House: si el Excel se actualiza con el mismo
    nombre, Streamlit debe volver a leerlo y no quedarse con el valor anterior
    en caché.
    """
    return Path(path_str).read_bytes()


def read_local_file_bytes(path: Optional[Path]) -> Optional[bytes]:
    """Lee un archivo local y fuerza recarga cuando cambia mtime/tamaño."""
    if path is None or not path.exists():
        return None
    stt = path.stat()
    return bytes_from_path(str(path), int(stt.st_mtime_ns), int(stt.st_size))


def file_update_caption(path: Optional[Path]) -> str:
    """Texto corto con fecha de modificación del archivo local."""
    if path is None or not path.exists():
        return "No disponible"
    stt = path.stat()
    ts = pd.to_datetime(stt.st_mtime, unit="s")
    return f"{path.name} · modificado: {ts:%d-%m-%Y %H:%M:%S}"


@st.cache_data(show_spinner=False)
def sheet_names(file_bytes: bytes) -> List[str]:
    return pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl").sheet_names


# ─────────────────────────────────────────────────────────────────────
# Carga de hojas DSS
# ─────────────────────────────────────────────────────────────────────
def _make_unique_headers(raw: List) -> List[str]:
    seen: Dict[str, int] = {}
    result = []
    for i, v in enumerate(raw, 1):
        name = str(v).strip() if v is not None else ""
        if not name or name.lower().startswith("unnamed"):
            name = f"Col_{i}"
        seen[name] = seen.get(name, 0) + 1
        result.append(f"{name}_{seen[name]}" if seen[name] > 1 else name)
    return result


@st.cache_data(show_spinner="Cargando hoja DSS…")
def load_dss_sheet(file_bytes: bytes, wanted: str) -> pd.DataFrame:
    names = sheet_names(file_bytes)
    # flexible match
    target = next((s for s in names if s.strip().lower() == wanted.strip().lower()), None)
    if target is None:
        target = next((s for s in names if wanted.strip().lower() in s.strip().lower()), None)
    if target is None:
        raise ValueError(f"Hoja '{wanted}' no encontrada. Disponibles: {names}")

    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb[target]
        hrow = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        # trim to last non-null header
        last = max((i for i, v in enumerate(hrow) if v is not None), default=-1) + 1
        if last == 0:
            raise ValueError(f"La hoja {target} no tiene encabezados.")
        headers = _make_unique_headers(hrow[:last])
        rows, blanks, started = [], 0, False
        for row in ws.iter_rows(min_row=2, max_col=last, values_only=True):
            fecha = row[0] if row else None
            if fecha is None or str(fecha).strip() == "":
                if started:
                    blanks += 1
                    if blanks >= 24:
                        break
                continue
            started, blanks = True, 0
            rows.append(row)
    finally:
        wb.close()

    df = pd.DataFrame(rows, columns=headers)
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df[df["Fecha"].notna()].sort_values("Fecha").copy()
    for col in df.columns:
        if col != "Fecha":
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df.dropna(axis=1, how="all", inplace=True)
    return df


@st.cache_data(show_spinner="Cargando Lake House…")
def load_lakehouse(file_bytes: bytes) -> pd.DataFrame:
    names = sheet_names(file_bytes)
    preferred = ["daily_input", "AwsDataCatalog"]
    target = next(
        (next((s for s in names if t.lower() in s.lower()), None) for t in preferred),
        None,
    )
    if target is None:
        raise ValueError("Lake House requiere hoja 'daily_input' o 'AwsDataCatalog'.")
    usecols = ["actdate", "actgatel", "actmadel"]
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=target, engine="openpyxl", usecols=usecols)
    df["actdate"] = pd.to_datetime(df["actdate"], errors="coerce")
    df = df[df["actdate"].notna()].copy()
    df["Fecha_dia"] = df["actdate"].dt.floor("D")
    for col in ["actgatel", "actmadel"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df.sort_values("actdate", inplace=True)
    # Último valor válido por día para cada embalse.
    # No se usa simplemente "última fila" porque el último registro del día
    # puede venir vacío para un embalse y eso ocultaría el último observado real.
    daily = df.groupby("Fecha_dia", as_index=False).agg(
        actgatel=("actgatel", "last"),
        actmadel=("actmadel", "last"),
    )
    return daily.sort_values("Fecha_dia")


def latest_lkh_observation(lkh_daily: Optional[pd.DataFrame], cfg: Dict) -> Tuple[Optional[pd.Timestamp], Optional[float]]:
    """Devuelve SIEMPRE el último valor observado disponible en Lake House.

    Se calcula desde toda la hoja Lake House, no desde el filtro de fecha del DSS.
    Así el indicador superior se actualiza con el último dato real aunque el
    usuario esté mirando otro rango de fechas.
    """
    if lkh_daily is None or lkh_daily.empty:
        return None, None
    col = cfg.get("lkh_col")
    if col not in lkh_daily.columns:
        return None, None
    obs = lkh_daily[["Fecha_dia", col]].dropna(subset=[col]).sort_values("Fecha_dia")
    if obs.empty:
        return None, None
    rec = obs.iloc[-1]
    return pd.to_datetime(rec["Fecha_dia"]), float(rec[col])


def closest_np_percentile(
    daily: Optional[pd.DataFrame],
    cfg: Dict,
    obs_date: Optional[pd.Timestamp],
    obs_value: Optional[float],
) -> Optional[Dict]:
    """Identifica el percentil NP más cercano al último nivel observado.

    Usa la fecha del último observado Lake House. Si esa fecha no existe en el
    rango DSS diario, utiliza la fecha DSS más cercana para no dejar el indicador
    vacío. La diferencia se reporta como Observado - DSS.
    """
    if daily is None or daily.empty or obs_date is None or obs_value is None:
        return None

    base = daily.copy()
    if "Fecha_dia" not in base.columns:
        return None
    base["Fecha_dia"] = pd.to_datetime(base["Fecha_dia"], errors="coerce").dt.normalize()
    base = base[base["Fecha_dia"].notna()].sort_values("Fecha_dia")
    if base.empty:
        return None

    token = cfg["token"]
    np_cols = cols_by_prefix(base, "NP", token)
    if not np_cols:
        return None

    obs_day = pd.to_datetime(obs_date).normalize()
    exact = base[base["Fecha_dia"] == obs_day]
    if not exact.empty:
        row = exact.iloc[0]
        exact_date = True
    else:
        idx = (base["Fecha_dia"] - obs_day).abs().idxmin()
        row = base.loc[idx]
        exact_date = False

    candidates = []
    for col in np_cols:
        value = pd.to_numeric(pd.Series([row.get(col, np.nan)]), errors="coerce").iloc[0]
        if pd.notna(value):
            candidates.append((col, float(value), abs(float(obs_value) - float(value))))
    if not candidates:
        return None

    col, dss_value, abs_diff = min(candidates, key=lambda item: item[2])
    percentile = exceedance_from_col(col)
    return {
        "percentile": percentile,
        "label": f"P{percentile}",
        "column": col,
        "dss_value": dss_value,
        "diff": float(obs_value) - dss_value,
        "abs_diff": abs_diff,
        "date": pd.to_datetime(row["Fecha_dia"]),
        "exact_date": exact_date,
    }


def closest_percentile_summary_row(embalse: str, latest_obs: Tuple[Optional[pd.Timestamp], Optional[float]], daily: pd.DataFrame, cfg: Dict) -> Dict:
    """Fila de resumen para mostrar el percentil más cercano en tablas."""
    obs_date, obs_value = latest_obs if latest_obs else (None, None)
    info = closest_np_percentile(daily, cfg, obs_date, obs_value)
    return {
        "Embalse": embalse,
        "Fecha obs. LKH": obs_date.strftime("%d-%m-%Y") if obs_date is not None else "—",
        "Nivel obs. LKH (ft)": round(float(obs_value), 3) if obs_value is not None else np.nan,
        "Percentil más cercano": info["label"] if info else "—",
        "NP DSS usado (ft)": round(info["dss_value"], 3) if info else np.nan,
        "Diferencia Obs-DSS (ft)": round(info["diff"], 3) if info else np.nan,
        "Fecha DSS usada": info["date"].strftime("%d-%m-%Y") if info else "—",
        "Fecha exacta": "Sí" if info and info["exact_date"] else ("No" if info else "—"),
    }


# ─────────────────────────────────────────────────────────────────────
# Extracción de columnas por prefijo / token
# ─────────────────────────────────────────────────────────────────────
def cols_by_prefix(df: pd.DataFrame, prefix: str, token: str) -> List[str]:
    pat = re.compile(rf"^{re.escape(prefix)}(\d+)", re.I)
    out = []
    for col in df.columns:
        m = pat.match(str(col).strip())
        if m and token.upper() in str(col).upper():
            out.append(col)
    # sort descending by number (P90 → P5)
    return sorted(out, key=lambda c: -int(re.search(r"\d+", c).group()))


def exceedance_from_col(col: str) -> int:
    m = re.search(r"(\d+)", str(col))
    return int(m.group()) if m else 50


# ─────────────────────────────────────────────────────────────────────
# Agregación diaria
# ─────────────────────────────────────────────────────────────────────
def to_daily(df: pd.DataFrame, cfg: Dict, level_method: str = "last") -> pd.DataFrame:
    """Horario → diario. AP/V/HP → promedio; NP → last o mean según level_method."""
    if df.empty:
        return pd.DataFrame()
    data = df.copy()
    data["Fecha_dia"] = data["Fecha"].dt.floor("D")
    token = cfg["token"]
    agg: Dict[str, str] = {}
    # NP es nivel, por eso permite último valor del día o promedio diario.
    for c in cols_by_prefix(data, "NP", token):
        agg[c] = level_method

    # HP, AP, V y consumo de esclusajes son variables operativas/caudales,
    # por eso se agregan como promedio diario.
    for prefix in ["HP", "AP", "V", "EG", "EP", "E"]:
        for c in cols_by_prefix(data, prefix, token):
            agg[c] = "mean"
    if "Observado" in data.columns:
        data["Obs_DSS"] = data["Observado"]
        agg["Obs_DSS"] = level_method
    if not agg:
        return pd.DataFrame()
    daily = data.groupby("Fecha_dia", as_index=False).agg(agg)
    return daily.sort_values("Fecha_dia")


def merge_lkh(daily: pd.DataFrame, lkh: pd.DataFrame, cfg: Dict) -> pd.DataFrame:
    if lkh is None or lkh.empty:
        daily[cfg["lkh_label"]] = np.nan
        return daily
    obs = lkh[["Fecha_dia", cfg["lkh_col"]]].rename(columns={cfg["lkh_col"]: cfg["lkh_label"]})
    return daily.merge(obs, on="Fecha_dia", how="left")


# ─────────────────────────────────────────────────────────────────────
# Conversión de unidades
# ─────────────────────────────────────────────────────────────────────
def convert_flow(series: pd.Series, unit: str) -> pd.Series:
    if unit == "m³/s":
        return series * CFS_TO_M3S
    if unit == "hm³/d":
        return series * CFS_TO_HM3_DAY
    return series  # cfs


def apply_unit(df: pd.DataFrame, cols: List[str], unit: str) -> Tuple[pd.DataFrame, List[str]]:
    out = df.copy()
    new_cols = []
    ulabel = unit_display(unit)
    for c in cols:
        nc = f"{c} [{ulabel}]"
        out[nc] = convert_flow(out[c], unit)
        new_cols.append(nc)
    return out, new_cols


def combine_sum_by_exceedance(
    df: pd.DataFrame,
    groups: List[List[str]],
    unit: str,
    label: str = "Consumo total esclusajes",
) -> Tuple[pd.DataFrame, List[str]]:
    """Suma columnas por probabilidad de excedencia y convierte unidad.

    Se usa para Gatún cuando EG y EP representan componentes de consumo.
    La salida queda como consumo total: EG + EP para cada P90, P80, ... P5.
    """
    out = df.copy()
    by_pct: Dict[int, List[str]] = {}
    for cols in groups:
        for col in cols:
            if col in out.columns:
                by_pct.setdefault(exceedance_from_col(col), []).append(col)

    new_cols: List[str] = []
    ulabel = unit_display(unit)
    for pct in sorted(by_pct.keys(), reverse=True):
        src_cols = by_pct[pct]
        total_cfs = out[src_cols].sum(axis=1, min_count=1)
        new_col = f"{label} P{pct} [{ulabel}]"
        out[new_col] = convert_flow(total_cfs, unit)
        new_cols.append(new_col)
    return out, new_cols


# ─────────────────────────────────────────────────────────────────────
# Gráficas
# ─────────────────────────────────────────────────────────────────────
def _safe(v) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    return f"{float(v):,.3f}"


def _add_today_marker(fig, date_values) -> None:
    """Agrega la línea vertical de Hoy sin usar add_vline.

    Plotly puede fallar con add_vline + annotation_text cuando el eje X
    contiene fechas y algunas versiones convierten x a texto internamente.
    Esta versión usa add_shape + add_annotation para evitar el TypeError:
    unsupported operand type(s) for +: 'int' and 'str'.
    """
    fechas = pd.to_datetime(date_values, errors="coerce").dropna()
    if fechas.empty:
        return

    today = pd.Timestamp.today().normalize()
    if fechas.min().normalize() <= today <= fechas.max().normalize():
        x_today = today.to_pydatetime()
        fig.add_shape(
            type="line",
            x0=x_today,
            x1=x_today,
            y0=0,
            y1=1,
            xref="x",
            yref="paper",
            line=dict(width=2, dash="dash", color="rgba(255,140,0,0.85)"),
        )
        fig.add_annotation(
            x=x_today,
            y=1,
            xref="x",
            yref="paper",
            text="Hoy",
            showarrow=False,
            yshift=12,
            font=dict(color="darkorange", size=11),
            bgcolor="rgba(255,255,255,0.75)",
            bordercolor="rgba(255,140,0,0.55)",
            borderwidth=1,
        )


def fan_chart(
    df: pd.DataFrame,
    cols: List[str],
    title: str,
    y_label: str,
    key: str,
    show_band: bool = True,
    highlight_p50: bool = True,
    obs_col: Optional[str] = None,          # columna observado a superponer
    obs_label: str = "Observado",
    show_today_line: bool = True,            # línea vertical del día actual
) -> None:
    """Gráfica de abanico de probabilidades de excedencia con observado y línea de hoy."""
    if df.empty or not cols:
        st.info("Sin datos para graficar.")
        return
    cols = [c for c in cols if c in df.columns]
    if not cols:
        st.info("Columnas no disponibles.")
        return

    all_cols = list(cols) + ([obs_col] if obs_col and obs_col in df.columns else [])
    plot_df = df[["Fecha_dia"] + all_cols].dropna(how="all", subset=cols).copy()
    if len(plot_df) > 4000:
        step = max(1, len(plot_df) // 4000)
        plot_df = plot_df.iloc[::step]

    if not PLOTLY_OK:
        st.line_chart(plot_df.set_index("Fecha_dia")[cols])
        return

    fig = go.Figure()

    # Banda P90-P10
    if show_band:
        p90_col = next((c for c in cols if "90" in c), None)
        p10_col = next((c for c in cols if "10" in c), None)
        if p90_col and p10_col:
            fig.add_trace(go.Scatter(
                x=pd.concat([plot_df["Fecha_dia"], plot_df["Fecha_dia"][::-1]]),
                y=pd.concat([plot_df[p90_col], plot_df[p10_col][::-1]]),
                fill="toself",
                fillcolor="rgba(0,102,204,0.09)",
                line=dict(color="rgba(255,255,255,0)"),
                name="Banda P90-P10",
                showlegend=True,
                hoverinfo="skip",
            ))

    for col in cols:
        exc = exceedance_from_col(col)
        color = EXCEEDANCE_COLORS.get(exc, "#999999")
        width = 2.5 if (highlight_p50 and exc == 50) else 1.2
        dash  = "solid" if exc == 50 else "dot" if exc in (90, 10) else "solid"
        fig.add_trace(go.Scatter(
            x=plot_df["Fecha_dia"],
            y=plot_df[col],
            mode="lines",
            name=f"P{exc}",
            line=dict(color=color, width=width, dash=dash),
        ))

    # Traza observado Lake House
    if obs_col and obs_col in plot_df.columns:
        obs_valid = plot_df[plot_df[obs_col].notna()]
        if not obs_valid.empty:
            fig.add_trace(go.Scatter(
                x=obs_valid["Fecha_dia"],
                y=obs_valid[obs_col],
                mode="lines+markers",
                name=f"🔴 {obs_label}",
                line=dict(color="#e63946", width=3, dash="solid"),
                marker=dict(size=5, color="#e63946", symbol="circle"),
            ))

    # Línea vertical — día actual
    if show_today_line:
        _add_today_marker(fig, plot_df["Fecha_dia"])

    fig.update_layout(
        title=dict(text=title, font=dict(size=15, color="#003E69")),
        height=480,
        hovermode="x unified",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(l=10, r=10, t=60, b=10),
        plot_bgcolor="rgba(250,252,255,1)",
        paper_bgcolor="rgba(250,252,255,0)",
        yaxis=dict(title=y_label, gridcolor="rgba(0,0,0,0.07)"),
        xaxis=dict(title="Fecha", gridcolor="rgba(0,0,0,0.07)"),
    )
    st.plotly_chart(fig, use_container_width=True, key=key)


def lines_chart(
    df: pd.DataFrame,
    y_cols: List[str],
    title: str,
    y_label: str,
    key: str,
    extra_trace: Optional[Tuple[str, str, str]] = None,  # (col, name, color)
    show_today_line: bool = True,
) -> None:
    """Gráfica de líneas estándar + traza extra opcional + línea de hoy."""
    if df.empty:
        st.info("Sin datos.")
        return
    y_cols = [c for c in y_cols if c in df.columns]
    if not y_cols:
        st.info("Selecciona al menos una variable.")
        return

    subset     = [c for c in y_cols if c in df.columns]
    extra_cols = [extra_trace[0]] if extra_trace and extra_trace[0] in df.columns else []
    plot_df    = df[["Fecha_dia"] + subset + extra_cols].dropna(how="all", subset=subset + extra_cols).copy()

    if len(plot_df) > 4000:
        step = max(1, len(plot_df) // 4000)
        plot_df = plot_df.iloc[::step]

    if not PLOTLY_OK:
        st.line_chart(plot_df.set_index("Fecha_dia")[subset])
        return

    long_df = plot_df.melt(id_vars="Fecha_dia", value_vars=subset, var_name="Variable", value_name="Valor")
    fig = px.line(long_df, x="Fecha_dia", y="Valor", color="Variable", title=title)

    if extra_trace and extra_trace[0] in df.columns:
        extra_valid = plot_df[plot_df[extra_trace[0]].notna()]
        fig.add_trace(go.Scatter(
            x=extra_valid["Fecha_dia"],
            y=extra_valid[extra_trace[0]],
            mode="lines+markers",
            name=extra_trace[1],
            line=dict(color=extra_trace[2], width=3, dash="solid"),
            marker=dict(size=5, color=extra_trace[2], symbol="circle"),
        ))

    # Línea vertical — día actual
    if show_today_line:
        _add_today_marker(fig, plot_df["Fecha_dia"])

    fig.update_layout(
        height=480,
        hovermode="x unified",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(l=10, r=10, t=60, b=10),
        plot_bgcolor="rgba(250,252,255,1)",
        paper_bgcolor="rgba(250,252,255,0)",
        yaxis=dict(title=y_label, gridcolor="rgba(0,0,0,0.07)"),
        xaxis=dict(title="Fecha", gridcolor="rgba(0,0,0,0.07)"),
    )
    st.plotly_chart(fig, use_container_width=True, key=key)


# ─────────────────────────────────────────────────────────────────────
# Filtro de fechas
# ─────────────────────────────────────────────────────────────────────
def date_filter(df: pd.DataFrame, key: str) -> pd.DataFrame:
    if df.empty or "Fecha_dia" not in df.columns:
        return df
    valid = df["Fecha_dia"].dropna()
    if valid.empty:
        return df
    mn, mx = valid.min().date(), valid.max().date()
    c1, c2 = st.columns(2)
    s = c1.date_input("Desde", value=mn, min_value=mn, max_value=mx, key=f"{key}_s")
    e = c2.date_input("Hasta", value=mx, min_value=mn, max_value=mx, key=f"{key}_e")
    if s > e:
        st.warning("Fecha inicial > Fecha final. Se usa el periodo completo.")
        return df
    mask = (df["Fecha_dia"].dt.date >= s) & (df["Fecha_dia"].dt.date <= e)
    return df.loc[mask].copy()


# ─────────────────────────────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────────────────────────────
def sidebar() -> Tuple[Optional[bytes], Optional[bytes], str, str, bool]:
    st.sidebar.markdown(f"## 💧 {APP_TITLE.split('·')[0].strip()}")
    st.sidebar.markdown("---")
    st.sidebar.header("📁 Archivos de entrada")

    dss_up  = st.sidebar.file_uploader("Archivo DSS (SimulacionDSS…xlsx)", type=["xlsx", "xlsm"], key="dss_up")
    lkh_up  = st.sidebar.file_uploader("Archivo Lake House (opcional)",    type=["xlsx", "xlsm"], key="lkh_up")

    dss_def = find_default_file(DSS_DEFAULT_NAMES)
    lkh_def = find_default_file(LAKEHOUSE_DEFAULT_NAMES)

    if st.sidebar.button("🔄 Recargar archivos / último Lake House", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    auto_lkh_refresh = st.sidebar.checkbox(
        "Autoactualizar Lake House cada 60 s",
        value=False,
        help="Útil si el Excel LakeHouse_Data.xlsx se actualiza automáticamente en la carpeta del app.",
    )
    if auto_lkh_refresh:
        components.html(
            "<script>setTimeout(function(){window.parent.location.reload();}, 60000);</script>",
            height=0,
        )

    dss_bytes = dss_up.getvalue() if dss_up else read_local_file_bytes(dss_def)
    lkh_bytes = lkh_up.getvalue() if lkh_up else read_local_file_bytes(lkh_def)

    dss_label = (dss_up.name if dss_up else (dss_def.name if dss_def else "—"))
    lkh_label = (lkh_up.name if lkh_up else (lkh_def.name if lkh_def else "No disponible"))

    st.sidebar.caption(f"DSS: **{dss_label}**")
    if lkh_up:
        st.sidebar.caption(f"Lake House: **{lkh_label}** · cargado manualmente")
        st.sidebar.caption("Para actualizar un archivo cargado manualmente, vuelva a seleccionarlo en el cargador.")
    else:
        st.sidebar.caption(f"Lake House: **{file_update_caption(lkh_def)}**")

    st.sidebar.header("⚙️ Ajustes generales")

    flow_unit = st.sidebar.radio(
        "🔁 Unidad de caudal / consumo (AP, V y esclusajes)",
        options=["cfs", "m³/s", "hm³/d"],
        index=0,
        format_func=unit_display,
        help=(
            "PCS/cfs = pies cúbicos por segundo\n"
            "m³/s = metros cúbicos por segundo\n"
            "hm³/d = hectómetros cúbicos por día"
        ),
    )

    level_method_label = st.sidebar.radio(
        "Agregación diaria de niveles (NP)",
        ["Último valor del día", "Promedio diario"],
        index=0,
    )
    level_method = "last" if "Último" in level_method_label else "mean"

    show_raw = st.sidebar.checkbox("Mostrar datos horarios originales", value=False)

    st.sidebar.header("📖 Glosario de variables")
    for var, desc in [
        ("NP", "Nivel proyectado (ft PLD)"),
        ("HP", "Hidrogeneración (MW)"),
        ("AP", "Aportes al embalse (cfs)"),
        ("V",  "Vertidos / Spill (cfs)"),
        ("EG", "Esclusaje Gatún"),
        ("EP", "Esclusaje Panamax"),
        ("E",  "Esclusaje combinado"),
        ("P90…P5", "Probabilidad de excedencia"),
    ]:
        st.sidebar.markdown(f"**{var}**: {desc}")

    st.sidebar.markdown("---")
    st.sidebar.caption(AUTHOR_NOTE)
    return dss_bytes, lkh_bytes, flow_unit, level_method, show_raw


# ─────────────────────────────────────────────────────────────────────
# Métricas de resumen
# ─────────────────────────────────────────────────────────────────────
def show_metrics(
    daily: pd.DataFrame,
    cfg: Dict,
    flow_unit: str,
    latest_lkh: Optional[Tuple[Optional[pd.Timestamp], Optional[float]]] = None,
    daily_full: Optional[pd.DataFrame] = None,
) -> None:
    """Muestra métricas del día actual (hoy) y el último observado disponible."""
    if daily.empty:
        return

    token   = cfg["token"]
    np_cols = cols_by_prefix(daily, "NP", token)
    hp_cols = cols_by_prefix(daily, "HP", token)
    ap_cols = cols_by_prefix(daily, "AP", token)
    np50 = next((c for c in np_cols if "50" in c), np_cols[0] if np_cols else None)
    hp50 = next((c for c in hp_cols if "50" in c), hp_cols[0] if hp_cols else None)
    ap50 = next((c for c in ap_cols if "50" in c), ap_cols[0] if ap_cols else None)

    today = pd.Timestamp.today().normalize()
    srt   = daily.sort_values("Fecha_dia")

    # Fila del día actual (o el más cercano disponible en el rango)
    exact = srt[srt["Fecha_dia"] == today]
    if not exact.empty:
        rec      = exact.iloc[0]
        day_lbl  = f"🟢 Hoy {today.strftime('%d-%m-%Y')}"
    else:
        # día más cercano a hoy dentro del rango
        idx  = (srt["Fecha_dia"] - today).abs().idxmin()
        rec  = srt.loc[idx]
        diff = int((rec["Fecha_dia"] - today).days)
        sign = f"+{diff}d" if diff > 0 else f"{diff}d"
        day_lbl = f"📅 {rec['Fecha_dia'].strftime('%d-%m-%Y')} ({sign})"

    # Último observado Lake House con valor real (no NaN).
    # Se prioriza latest_lkh, calculado desde todo Lake House, no desde el filtro.
    lkh_label = cfg["lkh_label"]
    obs_val = None
    obs_date = None
    if latest_lkh and latest_lkh[0] is not None and latest_lkh[1] is not None:
        obs_date = pd.to_datetime(latest_lkh[0])
        obs_val = float(latest_lkh[1])
    else:
        obs_series = srt[srt[lkh_label].notna()] if lkh_label in srt.columns else pd.DataFrame()
        if not obs_series.empty:
            obs_rec = obs_series.iloc[-1]
            obs_date = pd.to_datetime(obs_rec["Fecha_dia"])
            obs_val = float(obs_rec[lkh_label])

    closest_info = None
    if obs_val is not None and obs_date is not None:
        obs_display = f"{obs_val:,.3f} ft"
        obs_note = f"último observado Lake House: {obs_date:%d-%m-%Y}"
        # Delta: observado vs NP50 del mismo día, buscando primero en todo el diario.
        base_for_delta = daily_full if daily_full is not None and not daily_full.empty else srt
        same_day = base_for_delta[base_for_delta["Fecha_dia"] == obs_date.normalize()] if np50 else pd.DataFrame()
        np50_same = float(same_day.iloc[0].get(np50, np.nan)) if not same_day.empty else np.nan
        delta_obs = f"Δ vs NP50: {obs_val - np50_same:+.3f} ft" if not np.isnan(np50_same) else None
        closest_info = closest_np_percentile(base_for_delta, cfg, obs_date, obs_val)
    else:
        obs_display, obs_note, delta_obs = "—", "sin LKH", None

    # Valores del día de referencia
    np50_val  = float(rec.get(np50, np.nan)) if np50 else np.nan
    hp50_val  = float(rec.get(hp50, np.nan)) if hp50 else np.nan
    ap50_val  = float(rec.get(ap50, np.nan)) if ap50 else np.nan
    ap50_conv = convert_flow(pd.Series([ap50_val]), flow_unit).iloc[0]

    # ── Fila 1: referencia temporal y NP observado
    st.markdown("#### 📊 Valores de referencia")
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric(day_lbl, "Referencia DSS")
    c2.metric(
        f"🔴 Último nivel obs. LKH ({cfg['level_unit']})",
        obs_display,
        delta=delta_obs,
        delta_color="inverse",
        help=obs_note,
    )
    c3.metric(
        f"NP50 DSS ({cfg['level_unit']})",
        f"{np50_val:,.3f}" if not np.isnan(np50_val) else "—",
    )
    c4.metric(
        "HP50 (MW)",
        f"{hp50_val:,.2f}" if not np.isnan(hp50_val) else "—",
    )
    c5.metric(
        f"AP50 ({unit_display(flow_unit)})",
        f"{ap50_conv:,.2f}" if not np.isnan(ap50_conv) else "—",
    )
    if closest_info:
        date_note = "fecha exacta" if closest_info["exact_date"] else f"fecha DSS más cercana: {closest_info['date']:%d-%m-%Y}"
        c6.metric(
            "🎯 Percentil más cercano",
            closest_info["label"],
            delta=f"Obs-DSS: {closest_info['diff']:+.3f} ft",
            delta_color="inverse",
            help=f"{date_note} · NP usado: {closest_info['dss_value']:.3f} ft · columna {closest_info['column']}",
        )
    else:
        c6.metric("🎯 Percentil más cercano", "—")

    # ── Fila 2: abanico de NP para el día de referencia
    if np_cols:
        with st.expander("📐 NP por probabilidad — día de referencia", expanded=False):
            exc_vals = {}
            for col in np_cols:
                exc  = exceedance_from_col(col)
                val  = float(rec.get(col, np.nan))
                exc_vals[f"P{exc}"] = f"{val:,.3f} {cfg['level_unit']}" if not np.isnan(val) else "—"
            n = len(exc_vals)
            cols_row = st.columns(min(n, 10))
            for i, (label, val) in enumerate(exc_vals.items()):
                cols_row[i % min(n, 10)].metric(label, val)


# ─────────────────────────────────────────────────────────────────────
# PESTAÑA EMBALSE (Gatún / Alhajuela)
# ─────────────────────────────────────────────────────────────────────
def tab_reservoir(
    res_key: str,
    dss_bytes: bytes,
    lkh_bytes: Optional[bytes],
    flow_unit: str,
    level_method: str,
    show_raw: bool,
) -> None:
    cfg = RESERVOIR_CONFIG[res_key]
    st.subheader(f"💧 {cfg['name']}")

    try:
        dss_raw = load_dss_sheet(dss_bytes, cfg["sheet"])
    except Exception as exc:
        st.error(f"Error cargando DSS: {exc}")
        return

    lkh_daily = None
    if lkh_bytes:
        try:
            lkh_daily = load_lakehouse(lkh_bytes)
        except Exception as exc:
            st.warning(f"Lake House no disponible: {exc}")

    daily = to_daily(dss_raw, cfg, level_method)
    daily = merge_lkh(daily, lkh_daily, cfg)
    latest_obs_lkh = latest_lkh_observation(lkh_daily, cfg)

    with st.expander("🗓️ Filtro de período", expanded=True):
        filtered = date_filter(daily, f"{res_key}_res")

    show_metrics(filtered, cfg, flow_unit, latest_lkh=latest_obs_lkh, daily_full=daily)
    if latest_obs_lkh and latest_obs_lkh[0] is not None:
        st.caption(f"Lake House actualizado en pantalla con el último valor observado disponible: {latest_obs_lkh[0]:%d-%m-%Y} · {latest_obs_lkh[1]:,.3f} ft")
    st.markdown("---")

    token = cfg["token"]
    np_cols = cols_by_prefix(filtered, "NP", token)
    hp_cols = cols_by_prefix(filtered, "HP", token)
    ap_cols = cols_by_prefix(filtered, "AP", token)
    v_cols  = cols_by_prefix(filtered, "V",  token)
    eg_cols = cols_by_prefix(filtered, "EG", token) if token == "GAT" else []
    ep_cols = cols_by_prefix(filtered, "EP", token) if token == "GAT" else []

    # ── 1. Nivel — NP con observado superpuesto
    st.markdown("### 📈 Nivel proyectado DSS vs observado")
    sel_np = st.multiselect(
        "Series NP",
        options=np_cols,
        default=[c for c in np_cols if any(x in c for x in ["50","90","10"])],
        key=f"{res_key}_np",
    )
    lkh_col_in_df = cfg["lkh_label"] if cfg["lkh_label"] in filtered.columns else None
    # Usar fan_chart para NP también (con observado)
    fan_chart(
        filtered, sel_np,
        f"{cfg['name']} · Nivel diario DSS (ft PLD)",
        "ft PLD", f"{res_key}_np_plot",
        obs_col=lkh_col_in_df,
        obs_label=cfg["lkh_label"],
        show_today_line=True,
    )

    # ── 2. HP
    if hp_cols:
        st.markdown("### ⚡ Hidrogeneración (HP)")
        sel_hp = st.multiselect(
            "Series HP",
            options=hp_cols,
            default=[c for c in hp_cols if "50" in c or "90" in c or "10" in c],
            key=f"{res_key}_hp",
        )
        fan_chart(
            filtered, sel_hp,
            f"{cfg['name']} · Hidrogeneración diaria promedio",
            "MW", f"{res_key}_hp_plot",
            show_today_line=True,
        )

    # ── 3. AP (con conversión de unidades y abanico completo)
    if ap_cols:
        st.markdown(f"### 🌊 Aportes (AP) — promedio diario  "
                    f"<span class='unit-badge'>{unit_display(flow_unit)}</span>",
                    unsafe_allow_html=True)
        sel_ap = st.multiselect(
            "Probabilidades de excedencia AP",
            options=ap_cols,
            default=ap_cols,          # todas por defecto
            key=f"{res_key}_ap",
        )
        ap_df, ap_new_cols = apply_unit(filtered, sel_ap, flow_unit)
        fan_chart(
            ap_df, ap_new_cols,
            f"{cfg['name']} · Aportes diarios promedio ({unit_display(flow_unit)})",
            unit_display(flow_unit), f"{res_key}_ap_plot",
            show_today_line=True,
        )

    # ── 4. V (Vertidos) con conversión
    if v_cols:
        st.markdown(f"### 🚿 Vertidos (V) — promedio diario  "
                    f"<span class='unit-badge'>{unit_display(flow_unit)}</span>",
                    unsafe_allow_html=True)
        sel_v = st.multiselect(
            "Probabilidades de excedencia V",
            options=v_cols,
            default=[c for c in v_cols if any(x in c for x in ["50","90","10"])],
            key=f"{res_key}_v",
        )
        v_df, v_new_cols = apply_unit(filtered, sel_v, flow_unit)
        fan_chart(
            v_df, v_new_cols,
            f"{cfg['name']} · Vertidos diarios promedio ({unit_display(flow_unit)})",
            unit_display(flow_unit), f"{res_key}_v_plot",
            show_today_line=True,
        )

    # ── 5. Esclusajes / consumo total (solo Gatún)
    # EG y EP se suman por probabilidad: total = EG + EP.
    # Se tratan como consumo/caudal para que respondan al selector global de unidad.
    total_esc_df = pd.DataFrame()
    total_esc_cols: List[str] = []
    if eg_cols or ep_cols:
        st.markdown(
            f"### 🚢 Consumo total de esclusajes — EG + EP  "
            f"<span class='unit-badge'>{unit_display(flow_unit)}</span>",
            unsafe_allow_html=True,
        )
        total_esc_df, total_esc_cols = combine_sum_by_exceedance(
            filtered, [eg_cols, ep_cols], flow_unit, label="Total esclusajes"
        )
        default_total = [c for c in total_esc_cols if any(x in c for x in ["P50", "P90", "P10"])]
        sel_total_esc = st.multiselect(
            "Probabilidades de excedencia — consumo total",
            options=total_esc_cols,
            default=default_total or total_esc_cols[:3],
            key=f"{res_key}_total_esc",
        )
        fan_chart(
            total_esc_df, sel_total_esc,
            f"{cfg['name']} · Consumo total de esclusajes EG + EP ({unit_display(flow_unit)})",
            unit_display(flow_unit), f"{res_key}_total_esc_plot",
            show_band=False,
            show_today_line=True,
        )
        with st.expander("Ver componentes originales EG y EP", expanded=False):
            c_eg, c_ep = st.columns(2)
            with c_eg:
                st.caption("EG · componente Gatún")
                st.dataframe(filtered[["Fecha_dia"] + eg_cols], use_container_width=True, height=240)
            with c_ep:
                st.caption("EP · componente Panamax")
                st.dataframe(filtered[["Fecha_dia"] + ep_cols], use_container_width=True, height=240)

    # ── Tabla diaria
    st.markdown("### 📋 Tabla diaria integrada DSS")
    with st.expander("Ver tabla", expanded=False):
        # Agregar columnas convertidas para AP/V y consumo total de esclusajes.
        display = filtered.copy()
        ulabel = unit_display(flow_unit)
        for col in ap_cols:
            display[f"{col} [{ulabel}]"] = convert_flow(display[col], flow_unit)
        for col in v_cols:
            display[f"{col} [{ulabel}]"] = convert_flow(display[col], flow_unit)
        if eg_cols or ep_cols:
            display, total_table_cols = combine_sum_by_exceedance(
                display, [eg_cols, ep_cols], flow_unit, label="Total esclusajes"
            )
        st.dataframe(display, use_container_width=True, height=450)
        csv = display.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            f"⬇️ Descargar CSV — {cfg['name']}",
            data=csv,
            file_name=f"{res_key}_dss_diario.csv",
            mime="text/csv",
            key=f"{res_key}_dl",
        )

    # ── Raw horario
    if show_raw:
        with st.expander("🔍 Datos horarios originales DSS", expanded=False):
            st.dataframe(dss_raw, use_container_width=True, height=380)
            raw_csv = dss_raw.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                "⬇️ Descargar CSV horario",
                data=raw_csv,
                file_name=f"{res_key}_dss_horario.csv",
                mime="text/csv",
                key=f"{res_key}_raw_dl",
            )


# ─────────────────────────────────────────────────────────────────────
# PESTAÑA APORTES DSS  (análisis enfocado AP — todas las probabilidades)
# ─────────────────────────────────────────────────────────────────────
def tab_aportes(
    dss_bytes: bytes,
    flow_unit: str,
    level_method: str,
) -> None:
    st.subheader("🌊 Aportes DSS — Análisis por probabilidad de excedencia")
    st.caption(
        "Visualiza el promedio diario de Aportes (AP) para **todas** las probabilidades de "
        "excedencia (P5 a P90) de Gatún y Alhajuela, con conversión de unidades en tiempo real."
    )

    try:
        gat_raw = load_dss_sheet(dss_bytes, "GATUN Px DSS")
        alh_raw = load_dss_sheet(dss_bytes, "ALHAJUELA Px DSS ")
    except Exception as exc:
        st.error(f"Error cargando hojas: {exc}")
        return

    gat_daily = to_daily(gat_raw, RESERVOIR_CONFIG["gatun"], level_method)
    alh_daily = to_daily(alh_raw, RESERVOIR_CONFIG["alhajuela"], level_method)

    # Filtro de fechas común
    if not gat_daily.empty:
        mn = max(gat_daily["Fecha_dia"].min(), alh_daily["Fecha_dia"].min()) if not alh_daily.empty else gat_daily["Fecha_dia"].min()
        mx = min(gat_daily["Fecha_dia"].max(), alh_daily["Fecha_dia"].max()) if not alh_daily.empty else gat_daily["Fecha_dia"].max()
        c1, c2 = st.columns(2)
        s = c1.date_input("Desde", value=mn.date(), min_value=mn.date(), max_value=mx.date(), key="ap_s")
        e = c2.date_input("Hasta", value=mx.date(), min_value=mn.date(), max_value=mx.date(), key="ap_e")
        mask_g = (gat_daily["Fecha_dia"].dt.date >= s) & (gat_daily["Fecha_dia"].dt.date <= e)
        mask_a = (alh_daily["Fecha_dia"].dt.date >= s) & (alh_daily["Fecha_dia"].dt.date <= e)
        gat_f = gat_daily.loc[mask_g].copy()
        alh_f = alh_daily.loc[mask_a].copy()
    else:
        gat_f, alh_f = gat_daily, alh_daily

    gat_ap = cols_by_prefix(gat_f, "AP", "GAT")
    alh_ap = cols_by_prefix(alh_f, "AP", "ALH")

    # ── Gatún AP abanico
    st.markdown(f"#### Gatún · AP promedio diario [{unit_display(flow_unit)}]")
    pct_options_g = [exceedance_from_col(c) for c in gat_ap]
    sel_pct_g = st.multiselect(
        "Probabilidades de excedencia — Gatún",
        options=pct_options_g,
        default=pct_options_g,
        format_func=lambda x: f"P{x}",
        key="ap_gat_pct",
    )
    sel_g_cols = [c for c in gat_ap if exceedance_from_col(c) in sel_pct_g]
    gat_ap_df, gat_ap_conv = apply_unit(gat_f, sel_g_cols, flow_unit)
    fan_chart(
        gat_ap_df, gat_ap_conv,
        f"Gatún · Aportes promedio diario ({unit_display(flow_unit)})",
        unit_display(flow_unit), "ap_gat_fan",
    )

    # ── Alhajuela AP abanico
    st.markdown(f"#### Alhajuela · AP promedio diario [{unit_display(flow_unit)}]")
    pct_options_a = [exceedance_from_col(c) for c in alh_ap]
    sel_pct_a = st.multiselect(
        "Probabilidades de excedencia — Alhajuela",
        options=pct_options_a,
        default=pct_options_a,
        format_func=lambda x: f"P{x}",
        key="ap_alh_pct",
    )
    sel_a_cols = [c for c in alh_ap if exceedance_from_col(c) in sel_pct_a]
    alh_ap_df, alh_ap_conv = apply_unit(alh_f, sel_a_cols, flow_unit)
    fan_chart(
        alh_ap_df, alh_ap_conv,
        f"Alhajuela · Aportes promedio diario ({unit_display(flow_unit)})",
        unit_display(flow_unit), "ap_alh_fan",
    )

    # ── Tabla estadísticas AP diario
    st.markdown("#### 📊 Estadísticas descriptivas — AP diario")

    def ap_stats(daily: pd.DataFrame, ap_cols_list: List[str], unit: str, label: str) -> pd.DataFrame:
        stats_rows = []
        ulabel = unit_display(unit)
        for col in ap_cols_list:
            exc = exceedance_from_col(col)
            series = convert_flow(daily[col].dropna(), unit)
            stats_rows.append({
                "Embalse": label,
                "Prob. excedencia": f"P{exc}",
                f"Mín ({ulabel})": round(series.min(), 3),
                f"Media ({ulabel})": round(series.mean(), 3),
                f"Mediana ({ulabel})": round(series.median(), 3),
                f"Máx ({ulabel})": round(series.max(), 3),
                f"Desv.Std ({ulabel})": round(series.std(), 3),
            })
        return pd.DataFrame(stats_rows)

    stats_gat = ap_stats(gat_f, gat_ap, flow_unit, "Gatún")
    stats_alh = ap_stats(alh_f, alh_ap, flow_unit, "Alhajuela")
    stats_all = pd.concat([stats_gat, stats_alh], ignore_index=True)

    st.dataframe(stats_all, use_container_width=True, hide_index=True)
    csv_stats = stats_all.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        "⬇️ Descargar estadísticas CSV",
        data=csv_stats,
        file_name="ap_dss_estadisticas.csv",
        mime="text/csv",
        key="ap_stats_dl",
    )

    # ── Boxplot por probabilidad
    if PLOTLY_OK and not gat_f.empty:
        st.markdown("#### 📦 Distribución AP por probabilidad de excedencia")
        box_rows = []
        for col in gat_ap:
            exc = exceedance_from_col(col)
            series = convert_flow(gat_f[col].dropna(), flow_unit)
            for v in series:
                box_rows.append({"Embalse":"Gatún", "P": f"P{exc}", "valor": v})
        for col in alh_ap:
            exc = exceedance_from_col(col)
            series = convert_flow(alh_f[col].dropna(), flow_unit)
            for v in series:
                box_rows.append({"Embalse":"Alhajuela", "P": f"P{exc}", "valor": v})
        if box_rows:
            bdf = pd.DataFrame(box_rows)
            # Ordenar P correctamente
            p_order = [f"P{p}" for p in PERCENTILE_ORDER if f"P{p}" in bdf["P"].unique()]
            bdf["P"] = pd.Categorical(bdf["P"], categories=p_order, ordered=True)
            fig = px.box(
                bdf.sort_values("P"), x="P", y="valor", color="Embalse",
                title=f"Distribución AP diario por probabilidad de excedencia ({unit_display(flow_unit)})",
                color_discrete_map={"Gatún":"#0066cc","Alhajuela":"#cc6600"},
            )
            fig.update_layout(height=430, yaxis_title=unit_display(flow_unit), plot_bgcolor="rgba(250,252,255,1)")
            st.plotly_chart(fig, use_container_width=True, key="ap_boxplot")


# ─────────────────────────────────────────────────────────────────────
# PESTAÑA COMPARATIVO
# ─────────────────────────────────────────────────────────────────────
def tab_comparativo(
    dss_bytes: bytes,
    lkh_bytes: Optional[bytes],
    flow_unit: str,
    level_method: str,
) -> None:
    st.subheader("🔀 Comparativo Gatún vs Alhajuela")

    try:
        gat_raw = load_dss_sheet(dss_bytes, "GATUN Px DSS")
        alh_raw = load_dss_sheet(dss_bytes, "ALHAJUELA Px DSS ")
    except Exception as exc:
        st.error(f"Error: {exc}")
        return

    lkh_daily = None
    if lkh_bytes:
        try:
            lkh_daily = load_lakehouse(lkh_bytes)
        except Exception:
            pass

    gat_d = to_daily(gat_raw, RESERVOIR_CONFIG["gatun"],     level_method)
    alh_d = to_daily(alh_raw, RESERVOIR_CONFIG["alhajuela"], level_method)
    gat_d = merge_lkh(gat_d, lkh_daily, RESERVOIR_CONFIG["gatun"])
    alh_d = merge_lkh(alh_d, lkh_daily, RESERVOIR_CONFIG["alhajuela"])

    latest_gat = latest_lkh_observation(lkh_daily, RESERVOIR_CONFIG["gatun"])
    latest_alh = latest_lkh_observation(lkh_daily, RESERVOIR_CONFIG["alhajuela"])
    closest_table = pd.DataFrame([
        closest_percentile_summary_row("Gatún", latest_gat, gat_d, RESERVOIR_CONFIG["gatun"]),
        closest_percentile_summary_row("Alhajuela / Madden", latest_alh, alh_d, RESERVOIR_CONFIG["alhajuela"]),
    ])
    st.markdown("#### 🎯 Percentil NP más cercano al último nivel observado Lake House")
    st.dataframe(closest_table, use_container_width=True, hide_index=True)

    with st.expander("🗓️ Filtro de período", expanded=True):
        mn = max(gat_d["Fecha_dia"].min(), alh_d["Fecha_dia"].min()) if not gat_d.empty and not alh_d.empty else pd.Timestamp("2026-01-01")
        mx = min(gat_d["Fecha_dia"].max(), alh_d["Fecha_dia"].max()) if not gat_d.empty and not alh_d.empty else pd.Timestamp("2026-12-31")
        c1, c2 = st.columns(2)
        s = c1.date_input("Desde", value=mn.date(), min_value=mn.date(), max_value=mx.date(), key="cmp_s")
        e = c2.date_input("Hasta", value=mx.date(), min_value=mn.date(), max_value=mx.date(), key="cmp_e")
        mask_g = (gat_d["Fecha_dia"].dt.date >= s) & (gat_d["Fecha_dia"].dt.date <= e)
        mask_a = (alh_d["Fecha_dia"].dt.date >= s) & (alh_d["Fecha_dia"].dt.date <= e)
        gat_f = gat_d.loc[mask_g].copy()
        alh_f = alh_d.loc[mask_a].copy()

    # Variable a comparar
    var_choice = st.selectbox("Variable a comparar", ["NP (Nivel)", "AP (Aportes)", "V (Vertidos)", "HP (Hidrogeneración)"], key="cmp_var")
    prefix = var_choice.split()[0]

    gat_cols = cols_by_prefix(gat_f, prefix, "GAT")
    alh_cols = cols_by_prefix(alh_f, prefix, "ALH")

    pct_choice = st.multiselect(
        "Probabilidades de excedencia",
        options=PERCENTILE_ORDER,
        default=[50, 90, 10],
        format_func=lambda x: f"P{x}",
        key="cmp_pct",
    )

    sel_gat = [c for c in gat_cols if exceedance_from_col(c) in pct_choice]
    sel_alh = [c for c in alh_cols if exceedance_from_col(c) in pct_choice]

    is_flow = prefix in ("AP", "V")
    if is_flow:
        gat_plot, sel_gat = apply_unit(gat_f, sel_gat, flow_unit)
        alh_plot, sel_alh = apply_unit(alh_f, sel_alh, flow_unit)
        y_lbl = unit_display(flow_unit)
    else:
        gat_plot, alh_plot = gat_f.copy(), alh_f.copy()
        y_lbl = "ft PLD" if prefix == "NP" else "MW"

    if not PLOTLY_OK:
        st.info("Instala plotly para ver gráficas comparativas.")
        return

    obs_gat_col = RESERVOIR_CONFIG["gatun"]["lkh_label"] if prefix == "NP" and RESERVOIR_CONFIG["gatun"]["lkh_label"] in gat_plot.columns else None
    obs_alh_col = RESERVOIR_CONFIG["alhajuela"]["lkh_label"] if prefix == "NP" and RESERVOIR_CONFIG["alhajuela"]["lkh_label"] in alh_plot.columns else None

    c_gat, c_alh = st.columns(2)
    with c_gat:
        st.caption("**Gatún**")
        fan_chart(
            gat_plot, sel_gat, f"Gatún · {var_choice}", y_lbl, "cmp_gat_fan",
            obs_col=obs_gat_col, obs_label="Nivel observado LKH Gatún"
        )
    with c_alh:
        st.caption("**Alhajuela**")
        fan_chart(
            alh_plot, sel_alh, f"Alhajuela · {var_choice}", y_lbl, "cmp_alh_fan",
            obs_col=obs_alh_col, obs_label="Nivel observado LKH Alhajuela"
        )

    # Gráfica superpuesta P50. Para NP también se agrega el nivel observado Lake House.
    st.markdown("#### P50 superpuesto — Gatún vs Alhajuela" + (" + observado Lake House" if prefix == "NP" else ""))
    gat50 = next((c for c in gat_cols if "50" in c), None)
    alh50 = next((c for c in alh_cols if "50" in c), None)
    if gat50 and alh50:
        fig = go.Figure()
        gat_y = convert_flow(gat_f[gat50], flow_unit) if is_flow else gat_f[gat50]
        alh_y = convert_flow(alh_f[alh50], flow_unit) if is_flow else alh_f[alh50]
        fig.add_trace(go.Scatter(x=gat_f["Fecha_dia"], y=gat_y, mode="lines", name="Gatún P50",
                                  line=dict(color="#0066cc", width=2.2)))
        fig.add_trace(go.Scatter(x=alh_f["Fecha_dia"], y=alh_y, mode="lines", name="Alhajuela P50",
                                  line=dict(color="#cc6600", width=2.2, dash="dash")))
        if prefix == "NP":
            gat_obs_col = RESERVOIR_CONFIG["gatun"]["lkh_label"]
            alh_obs_col = RESERVOIR_CONFIG["alhajuela"]["lkh_label"]
            if gat_obs_col in gat_f.columns:
                gat_obs = gat_f[["Fecha_dia", gat_obs_col]].dropna()
                if not gat_obs.empty:
                    fig.add_trace(go.Scatter(
                        x=gat_obs["Fecha_dia"], y=gat_obs[gat_obs_col],
                        mode="lines+markers", name="Gatún observado LKH",
                        line=dict(color="#e63946", width=3), marker=dict(size=5),
                    ))
            if alh_obs_col in alh_f.columns:
                alh_obs = alh_f[["Fecha_dia", alh_obs_col]].dropna()
                if not alh_obs.empty:
                    fig.add_trace(go.Scatter(
                        x=alh_obs["Fecha_dia"], y=alh_obs[alh_obs_col],
                        mode="lines+markers", name="Alhajuela observado LKH",
                        line=dict(color="#6a4c93", width=3, dash="dot"), marker=dict(size=5),
                    ))
        fig.update_layout(
            height=420, hovermode="x unified", yaxis_title=y_lbl,
            plot_bgcolor="rgba(250,252,255,1)",
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        )
        st.plotly_chart(fig, use_container_width=True, key="cmp_p50_overlay")


# ─────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────
def main() -> None:
    inject_css()
    st.markdown(f"<div class='main-title'>💧 {APP_TITLE}</div>", unsafe_allow_html=True)
    st.markdown("<div class='sub-title'>Simulación DSS 2026 · Manejo de embalses Gatún y Alhajuela — ACP HIMH</div>",
                unsafe_allow_html=True)

    dss_bytes, lkh_bytes, flow_unit, level_method, show_raw = sidebar()

    if dss_bytes is None:
        st.warning(
            "⚠️ Coloque `SimulacionDSS_2026.xlsx` en la misma carpeta que este script, "
            "o cárguelo desde el panel lateral."
        )
        return

    st.sidebar.success("✅ Archivo DSS cargado")

    tabs = st.tabs([
        "🌊 GATÚN Px DSS",
        "🏔️ ALHAJUELA Px DSS",
        "📥 Aportes DSS",
        "🔀 Comparativo",
    ])

    with tabs[0]:
        tab_reservoir("gatun",     dss_bytes, lkh_bytes, flow_unit, level_method, show_raw)

    with tabs[1]:
        tab_reservoir("alhajuela", dss_bytes, lkh_bytes, flow_unit, level_method, show_raw)

    with tabs[2]:
        tab_aportes(dss_bytes, flow_unit, level_method)

    with tabs[3]:
        tab_comparativo(dss_bytes, lkh_bytes, flow_unit, level_method)


if __name__ == "__main__":
    main()
