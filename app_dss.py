# -*- coding: utf-8 -*-
"""
DSS Simulación 2026 — Dashboard de toma de decisiones
Autor : JFRodriguez · Hidrólogo / Oceanógrafo Físico · ACP-HIMH

Versión simplificada y blindada:
- Pestañas enfocadas en decisión operativa
- Manejo robusto de errores en toda lectura de datos
- Sin pestañas redundantes ni funciones muertas
- Soporte nativo para BulkExport-GAT/MAD/TstCHCP (niveles y aportes directos)
- Semana DSS unificada en pestañas semanales: sábado-viernes por defecto, con opción lunes-domingo

Ejecución (Windows):
    py -m pip install streamlit openpyxl plotly pandas numpy
    py -m streamlit run app_simulacion_dss.py
"""
from __future__ import annotations

import re
import urllib.request
from io import BytesIO, StringIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

try:
    import plotly.express as px
    import plotly.graph_objects as go
    PLOTLY_OK = True
except ImportError:
    PLOTLY_OK = False

# ─────────────────────────────────────────────────────────────────────
# Configuración de página
# ─────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="DSS Simulación 2026",
    page_icon="🏞️",
    layout="wide",
    initial_sidebar_state="expanded",
)

APP_TITLE   = "DSS Simulación 2026 · Embalses ACP"
AUTHOR_NOTE = "· Hidrólogo / Oceanógrafo Físico · ACP-HIMH"
SIMULATION_NOTE = "Simulación realizada por JFRodriguez"
PROJ_NOTE   = "Proyecciones basadas en el decenio 2016-2025."
VIEW_FILE   = "dss_views.txt"
VIEW_STATE_DIR = ".app_state"
RADAR_URL   = "https://radar-meteorologico.delcanal.com/es.gif"

# Logos institucionales. Coloque estos archivos junto al app o en ./data.
LOGO_HIMH_CANDIDATES = [
    "logo_himh.jpg", "logo_himh.png", "logo_himh.jpeg",
    "LOGO_HIMH.jpg", "LOGO_HIMH.png", "himh_logo.jpg", "himh_logo.png",
]
LOGO_CANAL_CANDIDATES = [
    "logo_canal_panama.jpg", "logo_canal_panama.png", "logo_canal_panama.jpeg",
    "CP_RGB_p_Ver.jpg", "CP_RGB_p_Ver.png", "canal_panama_logo.jpg", "canal_panama_logo.png",
]

# Series Aquarius que la app puede intentar cargar automáticamente.
# La carga remota es opcional desde la barra lateral; si no hay acceso a la red
# o Aquarius solicita autenticación, la app continúa usando los CSV locales/subidos.
AQUARIUS_REQUIRED_SERIES_URLS: List[Tuple[str, str]] = [
    (
        "Lake_Res_elevation_Telem_Radar_MAD.csv",
        "https://panama.aquaticinformatics.net/Export/BulkExport?DateRange=Days7&TimeZone=-5&Calendar=CALENDARYEAR2&Interval=PointsAsRecorded&Step=1&ExportFormat=csv&TimeAligned=True&RoundData=True&IncludeGradeCodes=undefined&IncludeApprovalLevels=undefined&IncludeQualifiers=undefined&IncludeInterpolationTypes=False&IncludeNotes=undefined&Datasets[0].DatasetName=Lake-Res%20elevation.Telem%20Radar%40MAD&Datasets[0].Calculation=Instantaneous&Datasets[0].UnitId=70&_=1780594389875",
    ),
]

DSS_NAMES = [
    "SimulacionDSS_2026.xlsx", "SimulacionDSS_2026(3).xlsx",
    "SimulacionDSS_2026(2).xlsx", "SimulacionDSS_2026(1).xlsx",
]

# Serie histórica de aportes usada para comparar 1997, 1998, observado y DSS.
# La app la busca en ./data o junto al script y también permite carga manual.
HISTORICAL_9798_NAMES = [
    "1997_1998.xlsx", "1997-1998.xlsx", "Aportes_1997_1998.xlsx",
    "Aportes 1997 1998.xlsx", "Serie_1997_1998.xlsx",
]

# Carpeta estándar de datos del proyecto.
# La app busca primero en ./data y, como respaldo, en la carpeta donde está el script.
DATA_DIR_NAME = "data"

CFS_TO_M3S     = 0.028316846592
M3S_TO_CFS     = 1 / CFS_TO_M3S
CFS_TO_HM3_DAY = CFS_TO_M3S * 86400 / 1_000_000

# Evaporación automática:
# La lámina diaria de cada estación se corrige con el coeficiente 0.85.
# El área del espejo se calcula con el último nivel observado del embalse.
EVAP_COEFFICIENT = 0.85
EVAP_ROLLING_DAYS = 5  # promedio móvil de los últimos 5 días observados válidos

# Áreas de respaldo si no hay nivel observado disponible.
EVAP_AREA_GAT_KM2 = 425.0   # Gatún
EVAP_AREA_ALH_KM2 = 49.0    # Alhajuela / Madden

# Curva nivel-área editable. Se interpola por nivel observado (ft PLD).
# Si existe una curva oficial más detallada, solo reemplace estos puntos.
EVAP_LEVEL_AREA_TABLES: Dict[str, List[Tuple[float, float]]] = {
    "gatun": [
        (78.0, 392.0), (80.0, 402.0), (82.0, 413.0),
        (85.0, 425.0), (87.0, 435.0), (89.0, 445.0),
    ],
    "alhajuela": [
        (190.0, 18.0), (200.0, 25.0), (210.0, 32.0),
        (220.0, 39.0), (230.0, 45.0), (240.0, 49.0),
        (250.0, 53.0), (260.0, 57.0),
    ],
}

EVAP_SERIES_PATTERNS: Dict[str, List[str]] = {
    "CZL": ["Evapo_Rate_Daily_Tank_CZL*.csv", "*Evapo*CZL*.csv"],
    "PMG": ["Evapo_Rate_Daily_Tank_PMG*.csv", "*Evapo*PMG*.csv"],
}

# Ajuste morfológico de AP DSS: usa la forma del hidrograma observado de
# la última semana de mayo y las tres semanas siguientes de junio. Conserva
# el volumen/promedio semanal DSS. Semana operativa: sábado a viernes.
AP_HYDROGRAPH_ADJUSTMENT_ENABLED = True
AP_HYDROGRAPH_REFERENCE_WEEKS = 4
AP_HYDROGRAPH_REFERENCE_DAYS = AP_HYDROGRAPH_REFERENCE_WEEKS * 7
AP_HYDROGRAPH_REFERENCE_MIN_DAYS = 10
AP_HYDROGRAPH_REFERENCE_LABEL = "última semana de mayo + tres semanas siguientes de junio"
# Alias para compatibilidad con textos y funciones anteriores.
AP_MAY_HYDROGRAPH_DAYS = AP_HYDROGRAPH_REFERENCE_DAYS
AP_MAY_HYDROGRAPH_MIN_DAYS = AP_HYDROGRAPH_REFERENCE_MIN_DAYS

# Reparación defensiva para aportes observados Aquarius/BulkExport.
# En algunas corridas, los CSV de `Discharge_AT_*_Diario` pueden venir con
# huecos o ceros en la última semana de mayo y junio. La app no debe mostrar
# esos días como aporte cero si hay valores vecinos válidos; se reconstruyen
# únicamente para visualización/indicadores usando el valor válido más cercano.
AP_OBS_GAP_REPAIR_ENABLED = True
AP_OBS_GAP_REPAIR_START_MONTH = 5
AP_OBS_GAP_REPAIR_START_DAY = 25
AP_OBS_GAP_REPAIR_MAX_EXTENSION_DAYS = 5
AP_OBS_ZERO_IS_MISSING = True

PERCENTILE_ORDER = [5, 10, 20, 30, 40, 50, 60, 70, 80, 90, 95]

# Definición única para todas las pestañas semanales.
# DSS operativo por defecto: sábado-viernes. El usuario puede cambiar a
# lunes-domingo dentro de cada pestaña semanal sin afectar las demás.
DEFAULT_WEEK_START_LABEL = "Sábado a viernes (semana operativa DSS)"
WEEK_START_OPTIONS = {
    DEFAULT_WEEK_START_LABEL: 5,
    "Lunes a domingo": 0,
}
EXCEEDANCE_COLORS = {
    95: "#001f5b", 90: "#003f88", 80: "#005f99", 70: "#0077b6",
    60: "#0096c7", 50: "#48cae4", 40: "#90e0ef",
    30: "#f4a261", 20: "#e76f51", 10: "#d62828", 5: "#9b2226",
}

RESERVOIR_CONFIG: Dict[str, Dict] = {
    "gatun": {
        "sheet":      "GATUN Px DSS",
        "token":      "GAT",
        "name":       "Gatún",
        "level_unit": "ft PLD",
        "color":      "#0066cc",
    },
    "alhajuela": {
        "sheet":      "ALHAJUELA Px DSS ",
        "token":      "ALH",
        "name":       "Alhajuela / Madden",
        "level_unit": "ft PLD",
        "color":      "#cc6600",
    },
}


def unit_label(unit: str) -> str:
    return "p³/s" if unit == "cfs" else unit


def today_panama() -> pd.Timestamp:
    """Fecha actual en Panamá, usada para evitar graficar observaciones en días futuros."""
    try:
        return pd.Timestamp.now(tz="America/Panama").tz_localize(None).normalize()
    except Exception:
        return pd.Timestamp.today().normalize()


def clamp_observed_future_dates(df: pd.DataFrame, date_col: str = "Fecha_dia") -> pd.DataFrame:
    """Normaliza observaciones y elimina registros posteriores al día operativo actual.

    Para las series diarias de aporte `Discharge_AT_*_Diario`, la corrección
    principal de fecha se hace en `read_bulk_csv`: el sello 00:00 del día
    siguiente se asigna al día operativo anterior. Aquí solo se bloquean
    fechas futuras reales para que no aparezcan en gráficas ni métricas.
    """
    if df is None or df.empty or date_col not in df.columns:
        return df
    out = df.copy()
    out[date_col] = pd.to_datetime(out[date_col], errors="coerce").dt.normalize()
    today = today_panama()
    out = out.loc[out[date_col].notna() & (out[date_col] <= today)].copy()
    return out


def _is_daily_discharge_at_m3s(filename: str = "", header_text: str = "", columns: Optional[List[object]] = None) -> bool:
    """Identifica las series Aquarius `Discharge_AT_*_Diario` que vienen en m³/s.

    Estas series llegan normalizadas como `fecha_inicio, fecha_fin, valor_raw`
    y el valor **no** está en p³/s. Se convierte internamente a p³/s para
    compararlo con AP DSS, que se mantiene en p³/s.
    """
    joined = " ".join([str(filename or ""), str(header_text or ""), " ".join(map(str, columns or []))])
    joined = joined.upper()
    compact = re.sub(r"[^A-Z0-9]+", "_", joined)
    return any(token in compact for token in (
        "DISCHARGE_AT_GAT_DIARIO",
        "DISCHARGE_AT_ALHA_DIARIO",
        "DISCHARGE_AT_MAD_DIARIO",
    ))


def _shift_daily_midnight_to_operational_day(fechas: pd.Series) -> pd.Series:
    """Convierte el sello 00:00 del día siguiente al día operativo anterior.

    Ejemplo operativo: `22/06/2026 00:00` corresponde al aporte diario del
    `21/06/2026`. Esta regla se aplica solo a `Discharge_AT_*_Diario`.
    """
    out = pd.to_datetime(fechas, errors="coerce")
    try:
        mask_midnight = (
            out.notna()
            & out.dt.hour.eq(0)
            & out.dt.minute.eq(0)
            & out.dt.second.eq(0)
        )
        out = out.mask(mask_midnight, out - pd.Timedelta(days=1))
    except Exception:
        pass
    return out


# ─────────────────────────────────────────────────────────────────────
# CSS mínimo
# ─────────────────────────────────────────────────────────────────────
def inject_css() -> None:
    st.markdown("""
    <style>
    .main-title{font-size:2.1rem;font-weight:900;color:#003E69;margin-bottom:.1rem}
    .sub-title{color:#5f6b7a;font-size:1rem;margin-bottom:.8rem}
    div[data-testid="metric-container"]{
        border:1px solid rgba(0,62,105,.15);border-radius:12px;
        padding:.8rem 1rem;background:rgba(248,250,252,.92)}
    div[data-testid="stMetricValue"]{font-size:1.7rem;font-weight:800}
    .footer{margin-top:1rem;padding:.6rem;border-top:1px solid rgba(0,62,105,.12);
            color:#475467;font-size:.88rem;text-align:center}
    .badge{display:inline-block;background:#003E69;color:#fff;
           border-radius:8px;padding:2px 10px;font-size:.8rem;font-weight:700}
    </style>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────
# Contador de vistas (por sesión, persistente en archivo)
# ─────────────────────────────────────────────────────────────────────
def _safe_view_count_path() -> Path:
    """Ruta segura para el contador interno de vistas.

    Importante: antes se creaba `.dss_views.txt` en la carpeta principal.
    Si otro módulo carga archivos de viento buscando `.txt`, podía intentar
    leer ese contador como dato meteorológico. Ahora se guarda en una carpeta
    interna oculta y, si existe el archivo antiguo, se migra y se elimina.
    """
    try:
        base = Path(__file__).resolve().parent
    except Exception:
        base = Path.cwd().resolve()

    state_dir = base / VIEW_STATE_DIR
    state_dir.mkdir(parents=True, exist_ok=True)

    new_path = state_dir / VIEW_FILE
    legacy_path = base / ".dss_views.txt"

    # Migrar el contador viejo para que no sea detectado como archivo de datos.
    try:
        if legacy_path.exists() and legacy_path.is_file():
            if not new_path.exists():
                new_path.write_text(legacy_path.read_text(encoding="utf-8", errors="ignore").strip() or "0", encoding="utf-8")
            legacy_path.unlink(missing_ok=True)
    except Exception:
        # No detenemos el dashboard por un problema con el contador.
        pass

    return new_path


def get_view_count() -> int:
    if "view_count" in st.session_state:
        return int(st.session_state["view_count"])
    try:
        p = _safe_view_count_path()
        n = int(p.read_text(encoding="utf-8").strip()) if p.exists() else 0
        n += 1
        p.write_text(str(n), encoding="utf-8")
    except Exception:
        n = int(st.session_state.get("view_count", 0)) + 1
    st.session_state["view_count"] = n
    return n


# ─────────────────────────────────────────────────────────────────────
# Utilidades de archivo
# ─────────────────────────────────────────────────────────────────────
def app_base_dir() -> Path:
    """Carpeta donde está ubicado el script, con respaldo al directorio actual."""
    try:
        return Path(__file__).resolve().parent
    except Exception:
        return Path.cwd().resolve()


def local_search_dirs() -> List[Path]:
    """Directorios locales de búsqueda, priorizando la carpeta ./data.

    Estructura recomendada del proyecto:
        app_dss.py
        data/
          SimulacionDSS_2026.xlsx
          BulkExport*.csv
          Discharge_AT_GAT_Diario.csv
          Discharge_AT_ALHA_Diario.csv
          Lake_Res_elevation_*GAT*.csv
          Lake_Res_elevation_*ALHA*.csv
          Lake_Res_elevation_Telem_Radar_MAD.csv   # nivel Alhajuela/Madden
    """
    base = app_base_dir()
    candidates = [base / DATA_DIR_NAME, Path.cwd().resolve() / DATA_DIR_NAME, base, Path.cwd().resolve()]
    dirs: List[Path] = []
    seen = set()
    for d in candidates:
        try:
            rd = d.resolve()
        except Exception:
            rd = d
        if rd in seen or not rd.exists() or not rd.is_dir():
            continue
        seen.add(rd)
        dirs.append(rd)
    return dirs


def find_local(candidates: List[str]) -> Optional[Path]:
    """Busca un archivo local primero en ./data y luego junto al script."""
    for base in local_search_dirs():
        for n in candidates:
            p = base / n
            if p.exists() and p.is_file():
                return p
    return None


def find_local_logo(candidates: List[str]) -> Optional[Path]:
    """Busca logos primero junto al app y luego en ./data."""
    bases: List[Path] = []
    for base in [app_base_dir(), app_base_dir() / DATA_DIR_NAME, Path.cwd().resolve(), Path.cwd().resolve() / DATA_DIR_NAME]:
        try:
            rb = base.resolve()
        except Exception:
            rb = base
        if rb not in bases and rb.exists() and rb.is_dir():
            bases.append(rb)

    for base in bases:
        for name in candidates:
            fp = base / name
            if fp.exists() and fp.is_file():
                return fp
    return None


def show_brand_header(view_count: int) -> None:
    """Encabezado institucional con logos proporcionados por el usuario.

    Se usan anchos fijos para evitar que el logo HIMH crezca demasiado y
    desplace el logo del Canal de Panamá.
    """
    logo_himh = find_local_logo(LOGO_HIMH_CANDIDATES)
    logo_canal = find_local_logo(LOGO_CANAL_CANDIDATES)

    left, center, right = st.columns([0.72, 2.55, 0.90])
    with left:
        if logo_himh is not None:
            st.image(str(logo_himh), width=92)
    with center:
        st.markdown(
            f"""
            <div style="text-align:center;padding-top:.15rem;padding-bottom:.2rem">
                <div class="main-title" style="font-size:1.75rem;line-height:1.12">💧 {APP_TITLE}</div>
                <div class="sub-title" style="margin-bottom:.35rem">{PROJ_NOTE} · ACP HIMH</div>
                <div style="color:#003E69;font-weight:800">
                    ACP-HIMH
                    <span class='badge' style='margin-left:10px'>👁️ {view_count:,}</span>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with right:
        if logo_canal is not None:
            st.image(str(logo_canal), width=145)

    st.markdown("<div style='margin-bottom:.45rem'></div>", unsafe_allow_html=True)


@st.cache_data(show_spinner=False)
def _read_bytes_cached(path_str: str, mtime_ns: int, size: int) -> bytes:
    return Path(path_str).read_bytes()


def read_path_safe(path: Optional[Path]) -> Optional[bytes]:
    if path is None or not path.exists():
        return None
    try:
        file_stat = path.stat()
        return _read_bytes_cached(
            str(path),
            int(file_stat.st_mtime_ns),
            int(file_stat.st_size),
        )
    except PermissionError:
        st.warning(
            f"⚠️ `{path.name}` está bloqueado (Excel/OneDrive abierto). "
            "Ciérrelo o cargue manualmente."
        )
        return None
    except OSError as exc:
        st.warning(f"⚠️ No se pudo leer `{path.name}`: {exc}")
        return None


def read_first_local(candidates: List[str]) -> Tuple[Optional[bytes], Optional[Path]]:
    """Lee el primer archivo encontrado, priorizando ./data.

    Primero busca nombres exactos; si no aparecen, usa un respaldo flexible
    para versiones como SimulacionDSS_2026(4).xlsx o copias nuevas del DSS.
    """
    for base in local_search_dirs():
        for name in candidates:
            p = base / name
            if p.exists() and p.is_file():
                data = read_path_safe(p)
                if data:
                    return data, p

    # Respaldo flexible: usa el Excel DSS más reciente dentro de data/ o junto al app.
    fallback_patterns = ["SimulacionDSS_2026*.xlsx", "SimulacionDSS*.xlsx", "*DSS*.xlsx"]
    seen: Dict[Path, Path] = {}
    for base in local_search_dirs():
        for pat in fallback_patterns:
            for fp in base.glob(pat):
                try:
                    if fp.exists() and fp.is_file() and fp.suffix.lower() in {".xlsx", ".xlsm"}:
                        seen[fp.resolve()] = fp
                except OSError:
                    continue
    for fp in sorted(seen.values(), key=lambda x: x.stat().st_mtime, reverse=True):
        data = read_path_safe(fp)
        if data:
            return data, fp

    return None, None


def read_first_local_exact(candidates: List[str]) -> Tuple[Optional[bytes], Optional[Path]]:
    """Lee solo nombres históricos exactos, sin activar el respaldo genérico DSS."""
    path = find_local(candidates)
    return (read_path_safe(path), path) if path is not None else (None, None)


@st.cache_data(show_spinner=False)
def load_historical_9798(file_bytes: bytes) -> Tuple[pd.DataFrame, Dict[str, object]]:
    """Lee y normaliza la serie histórica diaria 1997-1998.

    El archivo entregado contiene una fila incompleta ``29/Feb`` y, desde marzo
    de la segunda serie, fechas almacenadas como 2026. Para no alterar el Excel
    original, la app identifica cada año por la secuencia mes/día y conserva los
    caudales como 1997 y 1998. Se descartan filas que no tienen simultáneamente
    aporte de Gatún y Alhajuela.
    """
    if not file_bytes:
        raise ValueError("Archivo histórico 1997-1998 vacío.")
    try:
        raw = pd.read_excel(BytesIO(file_bytes), sheet_name=0, engine="openpyxl")
    except Exception as exc:
        raise ValueError(f"No se pudo abrir la serie 1997-1998: {exc}")
    if raw is None or raw.empty:
        raise ValueError("La serie 1997-1998 no contiene datos.")

    def norm(value: object) -> str:
        s = str(value).strip().lower()
        s = s.translate(str.maketrans("áéíóúüñ", "aeiouun"))
        return re.sub(r"[^a-z0-9]+", "_", s).strip("_")

    cols = list(raw.columns)
    date_col = next((c for c in cols if "fecha" in norm(c) or norm(c) in {"date", "timestamp"}), None)
    gat_col = next((c for c in cols if "gatun" in norm(c)), None)
    alh_col = next((c for c in cols if any(x in norm(c) for x in ("alhajuela", "madden", "alha"))), None)
    if date_col is None or gat_col is None or alh_col is None:
        raise ValueError(
            "No se identificaron las columnas Fecha, Gatún y Alhajuela en el archivo histórico."
        )

    work = raw[[date_col, gat_col, alh_col]].copy()
    work.columns = ["Fecha_origen", "Gatun_cfs", "Alhajuela_cfs"]
    work["Gatun_cfs"] = pd.to_numeric(work["Gatun_cfs"], errors="coerce")
    work["Alhajuela_cfs"] = pd.to_numeric(work["Alhajuela_cfs"], errors="coerce")
    work["Fecha_parseada"] = pd.to_datetime(work["Fecha_origen"], errors="coerce")

    original_rows = int(len(work))
    work = work.dropna(subset=["Fecha_parseada", "Gatun_cfs", "Alhajuela_cfs"]).copy()
    work = work[
        np.isfinite(work["Gatun_cfs"])
        & np.isfinite(work["Alhajuela_cfs"])
        & (work["Gatun_cfs"] >= 0)
        & (work["Alhajuela_cfs"] >= 0)
    ].copy()
    if work.empty:
        raise ValueError("No hay pares válidos de aportes Gatún-Alhajuela en la serie histórica.")

    work["Mes"] = work["Fecha_parseada"].dt.month.astype(int)
    work["Dia"] = work["Fecha_parseada"].dt.day.astype(int)
    md = work["Mes"] * 100 + work["Dia"]
    work["_grupo"] = (md < md.shift(1)).fillna(False).cumsum().astype(int)

    groups = list(pd.unique(work["_grupo"]))
    # Respaldo para archivos equivalentes sin reinicio de mes/día detectable.
    if len(groups) < 2 and len(work) >= 730:
        work = work.reset_index(drop=True)
        work["_grupo"] = np.where(work.index < 365, 0, 1)
        groups = [0, 1]
    if len(groups) < 2:
        raise ValueError("No se pudieron separar las dos series anuales 1997 y 1998.")

    groups = groups[:2]
    group_to_year = {groups[0]: 1997, groups[1]: 1998}
    work = work[work["_grupo"].isin(groups)].copy()
    work["Anio_historico"] = work["_grupo"].map(group_to_year).astype(int)
    work["CHCP_cfs"] = work[["Gatun_cfs", "Alhajuela_cfs"]].sum(axis=1, min_count=2)
    work = work[
        ["Anio_historico", "Mes", "Dia", "Gatun_cfs", "Alhajuela_cfs", "CHCP_cfs"]
    ].sort_values(["Anio_historico", "Mes", "Dia"]).reset_index(drop=True)

    counts = work.groupby("Anio_historico").size().to_dict()
    meta: Dict[str, object] = {
        "filas_origen": original_rows,
        "filas_validas": int(len(work)),
        "filas_ignoradas": int(original_rows - len(work)),
        "dias_1997": int(counts.get(1997, 0)),
        "dias_1998": int(counts.get(1998, 0)),
        "nota": (
            "Fechas normalizadas por secuencia mes/día; se ignoró la fila 29/Feb incompleta "
            "y se reasignó a 1998 la continuación almacenada con fechas 2026."
        ),
    }
    return work, meta


# ─────────────────────────────────────────────────────────────────────
# Series de evaporación CZL / PMG
# ─────────────────────────────────────────────────────────────────────
def _normalizar_columna(nombre: object) -> str:
    s = str(nombre).strip().lower()
    return s.translate(str.maketrans("áéíóúüñ", "aeiouun"))


@st.cache_data(show_spinner=False)
def read_evap_rate_series(file_bytes: bytes, filename: str = "") -> pd.DataFrame:
    """Lee la serie diaria de evaporación y devuelve Fecha y mm_dia válidos.

    Acepta columnas equivalentes a fecha_inicio/fecha/ timestamp y
    valor_raw/valor/value/evapo. Si hay más de un registro por día, conserva
    el último valor observado válido de ese día.
    """
    empty = pd.DataFrame(columns=["Fecha", "mm_dia"])
    if not file_bytes:
        return empty
    try:
        text_csv = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text_csv = file_bytes.decode("latin-1", errors="replace")

    try:
        sep = ";" if text_csv[:4000].count(";") > text_csv[:4000].count(",") else ","
        df = pd.read_csv(StringIO(text_csv), sep=sep)
    except Exception:
        return empty
    if df.empty:
        return empty

    cols = list(df.columns)
    norm = {_normalizar_columna(c): c for c in cols}

    date_col = None
    for wanted in ("fecha_inicio", "fecha_fin", "fecha", "timestamp", "date", "time"):
        for n, original in norm.items():
            if wanted in n:
                parsed = pd.to_datetime(df[original], errors="coerce")
                if parsed.notna().any():
                    date_col = original
                    break
        if date_col is not None:
            break

    value_col = None
    for wanted in ("valor_raw", "valor", "value", "evapo", "evap"):
        for n, original in norm.items():
            if original == date_col:
                continue
            if wanted in n:
                parsed = pd.to_numeric(
                    df[original].astype(str).str.replace(",", ".", regex=False).str.strip(),
                    errors="coerce",
                )
                if parsed.notna().any():
                    value_col = original
                    break
        if value_col is not None:
            break

    if date_col is None or value_col is None:
        return empty

    out = pd.DataFrame({
        "Fecha": pd.to_datetime(df[date_col], errors="coerce"),
        "mm_dia": pd.to_numeric(
            df[value_col].astype(str).str.replace(",", ".", regex=False).str.strip(),
            errors="coerce",
        ),
    }).dropna(subset=["Fecha", "mm_dia"])
    out = out[np.isfinite(out["mm_dia"]) & (out["mm_dia"] >= 0)].copy()
    if out.empty:
        return empty

    out["Fecha"] = pd.to_datetime(out["Fecha"], errors="coerce").dt.normalize()
    out = out.dropna(subset=["Fecha"]).sort_values("Fecha")
    out = out.groupby("Fecha", as_index=False).agg(mm_dia=("mm_dia", "last"))
    return out.sort_values("Fecha").reset_index(drop=True)


@st.cache_data(show_spinner=False)
def read_evap_rate_csv(file_bytes: bytes, filename: str = "") -> Tuple[Optional[pd.Timestamp], Optional[float]]:
    """Compatibilidad: devuelve el último valor diario válido de evaporación."""
    out = read_evap_rate_series(file_bytes, filename)
    if out.empty:
        return None, None
    last = out.iloc[-1]
    return pd.to_datetime(last["Fecha"]), float(last["mm_dia"])


def latest_evap_series(station: str, rolling_days: int = EVAP_ROLLING_DAYS) -> Dict[str, object]:
    """Obtiene el promedio móvil de los últimos N días observados de evaporación.

    Selecciona el archivo local cuya observación final sea la más reciente. El
    promedio se calcula sobre los últimos N días válidos disponibles, no sobre
    días calendario inventados. También devuelve el último valor individual y
    las fechas de la ventana para trazabilidad.
    """
    station = station.upper().strip()
    rolling_days = max(int(rolling_days or EVAP_ROLLING_DAYS), 1)
    patterns = EVAP_SERIES_PATTERNS.get(station, [f"*{station}*.csv"])
    files: Dict[Path, Path] = {}
    for base in local_search_dirs():
        for pattern in patterns:
            for fp in base.glob(pattern):
                try:
                    if fp.exists() and fp.is_file() and not fp.name.startswith("."):
                        files[fp.resolve()] = fp
                except OSError:
                    continue

    best: Optional[Dict[str, object]] = None
    for fp in files.values():
        try:
            series = read_evap_rate_series(fp.read_bytes(), fp.name)
        except Exception:
            continue
        if series.empty:
            continue
        window = series.tail(rolling_days).copy()
        if window.empty:
            continue
        item = {
            "date": pd.to_datetime(window["Fecha"].max()),
            "window_start": pd.to_datetime(window["Fecha"].min()),
            "window_end": pd.to_datetime(window["Fecha"].max()),
            "mm_day": float(window["mm_dia"].mean()),
            "latest_mm_day": float(window.iloc[-1]["mm_dia"]),
            "n_obs": int(len(window)),
            "rolling_days": rolling_days,
            "file": fp.name,
        }
        if best is None or pd.to_datetime(item["date"]) > pd.to_datetime(best["date"]):
            best = item
    return best or {
        "date": None, "window_start": None, "window_end": None,
        "mm_day": None, "latest_mm_day": None, "n_obs": 0,
        "rolling_days": rolling_days, "file": None,
    }


def evap_mm_to_flows(mm_day: float, area_km2: float, coefficient: float = EVAP_COEFFICIENT) -> Dict[str, float]:
    """Convierte lámina evaporada a hm³/día y p³/s.

    hm³/día = mm/día × área(km²) × coeficiente × 0.001
    """
    try:
        mm = max(float(mm_day), 0.0)
        area = max(float(area_km2), 0.0)
        coef = max(float(coefficient), 0.0)
    except Exception:
        mm = area = coef = 0.0
    hm3_day = mm * area * coef * 0.001
    cfs = hm3_day / CFS_TO_HM3_DAY if CFS_TO_HM3_DAY > 0 else 0.0
    return {"mm_day": mm, "area_km2": area, "coefficient": coef,
            "hm3_day": hm3_day, "cfs": cfs}


def fallback_evap_area_km2(res_key: str) -> float:
    """Área de respaldo cuando no hay nivel observado."""
    return EVAP_AREA_GAT_KM2 if str(res_key).lower().startswith("gat") else EVAP_AREA_ALH_KM2


def evap_area_from_observed_level(res_key: str, level_ft: Optional[float]) -> Tuple[float, str]:
    """Calcula el área del espejo por interpolación usando el nivel observado.

    Retorna (area_km2, estado). Si no hay nivel válido, usa el área de respaldo.
    """
    key = "gatun" if str(res_key).lower().startswith("gat") else "alhajuela"
    fallback = fallback_evap_area_km2(key)
    try:
        level = float(level_ft)
    except Exception:
        return fallback, "respaldo_sin_nivel"

    table = EVAP_LEVEL_AREA_TABLES.get(key, [])
    if not table:
        return fallback, "respaldo_sin_curva"

    pts = sorted((float(lv), float(ar)) for lv, ar in table)
    levels = np.array([x[0] for x in pts], dtype=float)
    areas = np.array([x[1] for x in pts], dtype=float)

    if not np.isfinite(level):
        return fallback, "respaldo_nivel_invalido"

    if level <= levels.min():
        return float(areas[0]), "nivel_bajo_clamp"
    if level >= levels.max():
        return float(areas[-1]), "nivel_alto_clamp"
    return float(np.interp(level, levels, areas)), "interpolado_nivel_obs"


def latest_observed_level_from_df(obs_df: Optional[pd.DataFrame]) -> Dict[str, object]:
    """Extrae el último nivel observado válido de un DataFrame diario."""
    if obs_df is None or not isinstance(obs_df, pd.DataFrame) or obs_df.empty:
        return {"date": None, "level_ft": None, "source": None}
    if "Fecha_dia" not in obs_df.columns or "Valor" not in obs_df.columns:
        return {"date": None, "level_ft": None, "source": None}

    df = obs_df.copy()
    df["Fecha_dia"] = pd.to_datetime(df["Fecha_dia"], errors="coerce").dt.normalize()
    df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce")
    df = df.dropna(subset=["Fecha_dia", "Valor"]).sort_values("Fecha_dia")
    if df.empty:
        return {"date": None, "level_ft": None, "source": None}

    today = today_panama()
    past = df[df["Fecha_dia"] <= today]
    last = past.iloc[-1] if not past.empty else df.iloc[-1]
    return {
        "date": pd.to_datetime(last["Fecha_dia"]),
        "level_ft": float(last["Valor"]),
        "source": last.get("Fuente", None),
    }


def latest_local_observed_level_for_evap(res_key: str) -> Dict[str, object]:
    """Busca el último nivel observado local para calcular área de evaporación.

    Esta búsqueda no crea widgets; usa los CSV locales de data/ o de la carpeta
    del app para que el panel lateral pueda calcular el área antes de que el
    usuario abra las pestañas.
    """
    target = "Gatún" if str(res_key).lower().startswith("gat") else "Alhajuela"
    best: Dict[str, object] = {"date": None, "level_ft": None, "source": None}
    try:
        for fp in discover_local_bulk_csvs()[:40]:
            try:
                daily, embalse, variable, serie = read_bulk_csv(fp.read_bytes(), fp.name)
            except Exception:
                continue
            if variable != "nivel" or daily is None or daily.empty:
                continue
            if target == "Gatún" and "Gat" not in str(embalse):
                continue
            if target == "Alhajuela" and not any(x in str(embalse) for x in ["Alhajuela", "Madden"]):
                continue

            item = latest_observed_level_from_df(daily)
            if item.get("level_ft") is None:
                continue
            if best.get("date") is None or pd.to_datetime(item["date"]) > pd.to_datetime(best["date"]):
                item["source"] = fp.name
                best = item
    except Exception:
        pass
    return best


# ─────────────────────────────────────────────────────────────────────
# Carga de hojas DSS
# ─────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner="Cargando hoja DSS…")
def load_dss_sheet(file_bytes: bytes, wanted: str) -> pd.DataFrame:
    """Carga una hoja DSS desde bytes. Blindado contra hojas inexistentes y headers duplicados."""
    try:
        names = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl").sheet_names
    except Exception as exc:
        raise ValueError(f"No se pudo abrir el Excel DSS: {exc}")

    target = next((s for s in names if s.strip().lower() == wanted.strip().lower()), None)
    if target is None:
        target = next((s for s in names if wanted.strip().lower() in s.strip().lower()), None)
    if target is None:
        raise ValueError(f"Hoja '{wanted}' no encontrada. Disponibles: {names}")

    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb[target]
        hrow = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        last = max((i for i, v in enumerate(hrow) if v is not None), default=-1) + 1
        if last == 0:
            raise ValueError(f"Hoja '{target}' sin encabezados.")
        # Headers únicos
        seen: Dict[str, int] = {}
        headers = []
        for i, v in enumerate(hrow[:last], 1):
            name = str(v).strip() if v is not None else f"Col_{i}"
            if not name or name.lower().startswith("unnamed"):
                name = f"Col_{i}"
            seen[name] = seen.get(name, 0) + 1
            headers.append(f"{name}_{seen[name]}" if seen[name] > 1 else name)
        rows, blanks, started = [], 0, False
        for row in ws.iter_rows(min_row=2, max_col=last, values_only=True):
            fecha = row[0] if row else None
            if fecha is None or str(fecha).strip() == "":
                if started:
                    blanks += 1
                    if blanks >= 24:
                        break
                continue
            started, blanks = True, 0
            rows.append(row)
    finally:
        wb.close()

    df = pd.DataFrame(rows, columns=headers)
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df[df["Fecha"].notna()].sort_values("Fecha").copy()
    for col in df.columns:
        if col != "Fecha":
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df.dropna(axis=1, how="all", inplace=True)
    return df


# ─────────────────────────────────────────────────────────────────────
# Carga de CSV de BulkExport (GAT, MAD, TstCHCP_AT)
# ─────────────────────────────────────────────────────────────────────
def _detect_embalse(header_text: str, filename: str) -> str:
    t = f"{filename} {header_text}".upper()
    if any(x in t for x in ["ATOTAL_GAT", "_GAT_", "GAT_TST", "GATUN", "GATÚN", "BulkExport-GAT".upper()]):
        return "Gatún"
    if any(x in t for x in [
        "ATOTAL_ALHA", "_ALHA_", "ALHA", "MADDEN", "MAD",
        "LAKE-RES ELEVATION.TELEM RADAR@MAD", "TELEM RADAR@MAD",
        "LAKE_RES_ELEVATION_TELEM_RADAR_MAD",
        "BulkExport-MAD".upper(), "BulkExport-TstCHCP".upper(),
    ]):
        return "Alhajuela / Madden"
    if "GAT" in t and "ALH" not in t:
        return "Gatún"
    return "Desconocido"


def _detect_variable(header_text: str, filename: str) -> str:
    """Detect if CSV contains level or flow data."""
    t = f"{filename} {header_text}".upper()
    if any(x in t for x in ["ELEVATION", "NIVEL", "LAKE", "FT)", "FT "]):
        return "nivel"
    if any(x in t for x in ["DISCHARGE", "ATOTAL", "APORTE", "FLOW", "FT^3", "CFS"]):
        return "aporte"
    return "aporte"  # default


@st.cache_data(show_spinner=False)
def read_bulk_csv(file_bytes: bytes, filename: str = "") -> Tuple[pd.DataFrame, str, str, str]:
    """
    Lee un CSV de BulkExport o un CSV normalizado por el descargador local.

    Formatos soportados:
    1) BulkExport original de Aquatic Informatics.
    2) CSV normalizado/sanitizado con columnas:
       fecha_inicio, fecha_fin, valor_raw

    Retorna: (df_diario, embalse, variable, serie_name)
    df_diario tiene columnas: Fecha_dia, Valor, Fuente

    Nota operativa: para `Discharge_AT_*_Diario`, `valor_raw` se lee como
    m³/s, se convierte a p³/s en `Valor` y el sello 00:00 se asigna al día
    operativo anterior.
    variable: 'nivel' | 'aporte'
    """
    empty = pd.DataFrame(columns=["Fecha_dia", "Valor", "Fuente"])
    if not file_bytes:
        return empty, "Desconocido", "aporte", "—"
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1", errors="replace")

    lines = text.splitlines()
    if not lines:
        return empty, "Desconocido", "aporte", "—"

    # Detectar separador con las primeras líneas.
    sep = ";" if sum(l.count(";") for l in lines[:10]) >= sum(l.count(",") for l in lines[:10]) else ","

    # Buscar fila de encabezado. Incluye BulkExport y CSV normalizado local.
    header_idx = 0
    for i, line in enumerate(lines[:80]):
        low = line.lower()
        has_time = any(x in low for x in ["sello", "timestamp", "fecha", "date", "time"])
        has_value = any(x in low for x in ["valor", "value", "valor_raw", "ft", "cfs", "cms", "m3", "m³"])
        if has_time and has_value:
            header_idx = i
            break

    meta_text = "\n".join(lines[:header_idx])
    embalse = _detect_embalse(meta_text, filename)
    variable = _detect_variable(meta_text, filename)

    serie = "—"
    for line in reversed(lines[:header_idx]):
        parts = [p.strip() for p in line.split(sep) if p.strip()]
        cand = next((p for p in parts if any(t in p.upper() for t in
                     ["DISCHARGE", "ATOTAL", "ELEVATION", "LAKE", "APORTE", "NIVEL"])), None)
        if cand:
            serie = cand
            break

    try:
        df = pd.read_csv(StringIO("\n".join(lines[header_idx:])), sep=sep, header=0)
    except Exception:
        return empty, embalse, variable, serie

    if df.empty or len(df.columns) < 2:
        return empty, embalse, variable, serie

    # ── Detección robusta de columnas ───────────────────────────────
    # Los CSV normalizados tienen: fecha_inicio, fecha_fin, valor_raw.
    # Antes se tomaba la segunda columna como valor; eso podía leer fecha_fin
    # y dejar todo como NaN. Aquí se detecta explícitamente la columna numérica.
    def _norm_col(c: object) -> str:
        s = str(c).strip().lower()
        repl = str.maketrans("áéíóúüñ", "aeiouun")
        return s.translate(repl)

    cols = list(df.columns)
    norm = {_norm_col(c): c for c in cols}
    is_daily_discharge_m3s = variable == "aporte" and _is_daily_discharge_at_m3s(filename, meta_text, cols)

    if is_daily_discharge_m3s:
        # En `Discharge_AT_*_Diario`, el cierre del intervalo llega como
        # fecha_fin 00:00 del día siguiente. Por eso se prefiere fecha_fin.
        preferred_time_names = [
            "fecha_fin", "fecha_inicio", "timestamp", "sello de tiempo", "sello", "fecha", "date", "time"
        ]
    else:
        preferred_time_names = [
            "fecha_inicio", "timestamp", "sello de tiempo", "sello", "fecha", "date", "time"
        ]
    time_col = None
    for key in preferred_time_names:
        for n, original in norm.items():
            if key in n:
                parsed = pd.to_datetime(df[original], errors="coerce")
                if parsed.notna().sum() > 0:
                    time_col = original
                    break
        if time_col is not None:
            break
    if time_col is None:
        time_col = cols[0]

    preferred_value_names = [
        "valor_raw", "valor", "value", "elevation", "nivel", "discharge", "aporte", "flow", "cfs", "ft"
    ]
    value_col = None
    for key in preferred_value_names:
        for n, original in norm.items():
            if original == time_col:
                continue
            if key in n and not any(x in n for x in ["fecha", "date", "time", "sello"]):
                sample = df[original].astype(str).str.replace(",", ".", regex=False).str.strip()
                parsed = pd.to_numeric(sample, errors="coerce")
                if parsed.notna().sum() > 0:
                    value_col = original
                    break
        if value_col is not None:
            break

    # Respaldo: escoger la columna con mayor cantidad de valores numéricos,
    # excluyendo columnas claramente temporales.
    if value_col is None:
        best = None
        best_count = -1
        for c in cols:
            n = _norm_col(c)
            if c == time_col or any(x in n for x in ["fecha", "date", "time", "sello"]):
                continue
            sample = df[c].astype(str).str.replace(",", ".", regex=False).str.strip()
            parsed = pd.to_numeric(sample, errors="coerce")
            count = int(parsed.notna().sum())
            if count > best_count:
                best = c
                best_count = count
        value_col = best

    if time_col is None or value_col is None:
        return empty, embalse, variable, serie

    out = pd.DataFrame()
    out["Fecha"] = pd.to_datetime(df[time_col], errors="coerce")
    raw_val = df[value_col].astype(str).str.replace(",", ".", regex=False).str.strip()
    out["Valor"] = pd.to_numeric(raw_val, errors="coerce")

    if is_daily_discharge_m3s:
        # `valor_raw` viene en m³/s; internamente se conserva `Valor` en p³/s
        # para compararlo con AP DSS. Además, 22/06 00:00 => operativo 21/06.
        out["Fecha"] = _shift_daily_midnight_to_operational_day(out["Fecha"])
        out["Valor"] = out["Valor"] * M3S_TO_CFS

    out = out.dropna(subset=["Fecha", "Valor"]).sort_values("Fecha")
    out["Fuente"] = filename or serie

    if out.empty:
        return empty, embalse, variable, serie

    # Agregar a diario.
    # - Nivel: último valor diario, más representativo del estado operativo.
    # - Aporte: último valor del día operativo. En `Discharge_AT_*_Diario`,
    #   el valor ya representa el cierre diario en m³/s convertido a p³/s.
    out["Fecha_dia"] = out["Fecha"].dt.floor("D")
    out = clamp_observed_future_dates(out, "Fecha_dia")

    if variable == "nivel":
        daily = out.groupby("Fecha_dia", as_index=False).agg(
            Valor=("Valor", "last"),
            Fuente=("Fuente", "last"),
        )
    else:
        daily = out.groupby("Fecha_dia", as_index=False).agg(
            Valor=("Valor", "last"),
            Fuente=("Fuente", "last"),
        )

    return daily.sort_values("Fecha_dia"), embalse, variable, serie


def _aporte_repair_start_date(dates: pd.Series) -> Optional[pd.Timestamp]:
    """Inicio del tramo que se blinda para aportes observados: finales de mayo."""
    valid = pd.to_datetime(dates, errors="coerce").dropna()
    if valid.empty:
        return None
    latest_year = int(valid.max().year)
    may_start = pd.Timestamp(
        year=latest_year,
        month=AP_OBS_GAP_REPAIR_START_MONTH,
        day=AP_OBS_GAP_REPAIR_START_DAY,
    )
    return max(valid.min().normalize(), may_start)


def repair_observed_aporte_gaps(
    df: Optional[pd.DataFrame],
    reservoir_label: str = "",
    enabled: bool = AP_OBS_GAP_REPAIR_ENABLED,
) -> Optional[pd.DataFrame]:
    """Rellena huecos/ceros de aportes observados desde finales de mayo.

    No modifica niveles ni columnas DSS. Solo aplica a DataFrames de aporte
    observado con columnas `Fecha_dia`, `Valor`, `Fuente`. La reparación usa el
    valor válido más cercano dentro de la serie, marca la fuente como relleno y
    evita que la gráfica/indicadores muestren ceros falsos o días faltantes en
    la última semana de mayo y junio.
    """
    if not enabled or df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df
    if "Fecha_dia" not in df.columns or "Valor" not in df.columns:
        return df

    out = df.copy()
    out["Fecha_dia"] = pd.to_datetime(out["Fecha_dia"], errors="coerce").dt.normalize()
    out["Valor"] = pd.to_numeric(out["Valor"], errors="coerce")
    if "Fuente" not in out.columns:
        out["Fuente"] = reservoir_label or "Aporte observado"
    out = out.dropna(subset=["Fecha_dia"]).sort_values("Fecha_dia")
    if out.empty:
        return out

    # Ceros/negativos en aportes diarios son tratados como huecos operativos.
    # Esto evita la línea roja pegada al eje inferior cuando Aquarius entrega 0
    # en días que realmente deben conservar la continuidad del hidrograma.
    if AP_OBS_ZERO_IS_MISSING:
        out.loc[out["Valor"] <= 0, "Valor"] = np.nan

    # Si hay duplicados, conserva el último valor válido del día. Si el último
    # era NaN por un cero, usa el último valor válido de ese mismo día.
    def _last_valid(series: pd.Series):
        valid = pd.to_numeric(series, errors="coerce").dropna()
        return valid.iloc[-1] if not valid.empty else np.nan

    out = out.groupby("Fecha_dia", as_index=False).agg(
        Valor=("Valor", _last_valid),
        Fuente=("Fuente", "last"),
    ).sort_values("Fecha_dia")

    valid_dates = out.loc[out["Valor"].notna(), "Fecha_dia"]
    if valid_dates.empty:
        return out

    start = _aporte_repair_start_date(out["Fecha_dia"])
    if start is None:
        return out
    last_valid = pd.to_datetime(valid_dates.max()).normalize()
    today = today_panama()
    end = max(pd.to_datetime(out["Fecha_dia"].max()).normalize(), last_valid)
    # Solo se extiende al día actual si el último dato válido está cerca; así no
    # se replica una serie vieja por muchos días.
    if today > end and (today - last_valid).days <= AP_OBS_GAP_REPAIR_MAX_EXTENSION_DAYS:
        end = today
    end = min(end, today) if pd.notna(today) else end
    if pd.isna(start) or pd.isna(end) or start > end:
        return out

    full_idx = pd.date_range(start=start, end=end, freq="D")
    if full_idx.empty:
        return out

    base = out.set_index("Fecha_dia").sort_index()
    base = base[~base.index.duplicated(keep="last")]
    re = base.reindex(base.index.union(full_idx)).sort_index()

    # Valores anterior/siguiente para escoger el vecino más cercano.
    prev_val = re["Valor"].ffill()
    next_val = re["Valor"].bfill()
    prev_src = re["Fuente"].ffill()
    next_src = re["Fuente"].bfill()

    idx_series = pd.Series(re.index, index=re.index)
    valid_mask = re["Valor"].notna()
    prev_date = idx_series.where(valid_mask).ffill()
    next_date = idx_series.where(valid_mask).bfill()

    repair_mask = re.index.isin(full_idx) & re["Valor"].isna()
    for idx in re.index[repair_mask]:
        pv, nv = prev_val.loc[idx], next_val.loc[idx]
        if pd.isna(pv) and pd.isna(nv):
            continue
        use_next = False
        if pd.isna(pv):
            use_next = True
        elif pd.notna(nv):
            try:
                dprev = abs((idx - pd.to_datetime(prev_date.loc[idx])).days)
                dnext = abs((pd.to_datetime(next_date.loc[idx]) - idx).days)
                use_next = dnext < dprev
            except Exception:
                use_next = False
        if use_next:
            re.at[idx, "Valor"] = float(nv)
            src = str(next_src.loc[idx]) if pd.notna(next_src.loc[idx]) else (reservoir_label or "Aporte observado")
        else:
            re.at[idx, "Valor"] = float(pv)
            src = str(prev_src.loc[idx]) if pd.notna(prev_src.loc[idx]) else (reservoir_label or "Aporte observado")
        re.at[idx, "Fuente"] = f"{src} · relleno vecino cercano"

    result = re.reset_index().rename(columns={"index": "Fecha_dia"})
    result["Fecha_dia"] = pd.to_datetime(result["Fecha_dia"], errors="coerce").dt.normalize()
    result = result.dropna(subset=["Fecha_dia", "Valor"]).sort_values("Fecha_dia")
    # Mantener todos los datos originales fuera del tramo y el tramo reparado.
    result = result.groupby("Fecha_dia", as_index=False).agg(
        Valor=("Valor", "last"),
        Fuente=("Fuente", "last"),
    )
    return result.sort_values("Fecha_dia")


def _observed_csv_priority(path: Path) -> Tuple[int, float, str]:
    """Prioriza CSV críticos de aporte/nivel frente a archivos auxiliares."""
    name = path.name.lower()
    priority = 50
    critical = (
        "discharge_at_gat_diario.csv",
        "discharge_at_alha_diario.csv",
        "lake_res_elevation_telem_avg_gat.csv",
        "lake_res_elevation_telem_radar_mad.csv",
    )
    if name in critical:
        priority = 0
    elif name.startswith("discharge_at_"):
        priority = 1
    elif "lake_res_elevation" in name:
        priority = 2
    elif name.startswith("bulkexport"):
        priority = 3
    return (priority, -float(path.stat().st_mtime), name)


def discover_local_bulk_csvs() -> List[Path]:
    """Descubre CSV observados, priorizando la carpeta ./data.

    Acepta nombres BulkExport originales y archivos ya normalizados/sanitizados
    generados por el proceso de descarga, por ejemplo:
    - BulkExport-GAT.csv / BulkExport-MAD.csv
    - Discharge_AT_GAT_Diario.csv / Discharge_AT_ALHA_Diario.csv
    - Lake_Res_elevation_*GAT*.csv / Lake_Res_elevation_*ALHA*.csv
    - Lake_Res_elevation_Telem_Radar_MAD.csv (nivel Alhajuela/Madden desde MAD)
    """
    patterns = [
        "BulkExport*.csv",
        "*GAT*.csv",
        "*GATUN*.csv",
        "*Gatún*.csv",
        "*Gatun*.csv",
        "*MAD*.csv",
        "*MADDEN*.csv",
        "*ALHA*.csv",
        "*ALHAJUELA*.csv",
        "*Telem*Radar*MAD*.csv",
        "*Lake*Res*elevation*MAD*.csv",
    ]
    seen: Dict[Path, Path] = {}
    for base in local_search_dirs():
        for pat in patterns:
            for p in base.glob(pat):
                try:
                    if p.exists() and p.is_file() and not p.name.startswith("."):
                        seen[p.resolve()] = p
                except OSError:
                    continue
    return sorted(seen.values(), key=_observed_csv_priority)


# ─────────────────────────────────────────────────────────────────────
# Agregación diaria DSS
# ─────────────────────────────────────────────────────────────────────
def cols_by_prefix(df: pd.DataFrame, prefix: str, token: str) -> List[str]:
    """Columnas que empiecen con prefix y contengan token (case-insensitive)."""
    pat = re.compile(rf"^{re.escape(prefix)}(\d+)", re.I)
    out = [c for c in df.columns
           if pat.match(str(c).strip()) and token.upper() in str(c).upper()]
    return sorted(out, key=lambda c: -int(re.search(r"\d+", c).group()))


def exceedance_pct(col: str) -> int:
    m = re.search(r"(\d+)", str(col))
    return int(m.group()) if m else 50


def order_cols_wet_to_dry(cols: List[str], pct_map: Optional[Dict[str, int]] = None) -> List[str]:
    """Orden visual de probabilidades para leyenda/visor Plotly.

    Convención solicitada en el dashboard:
    - P5  = condición más húmeda / curva superior.
    - P95 = condición más seca / curva inferior.

    Plotly muestra el visor unificado en el orden en que se agregan las
    trazas; por eso todas las gráficas deben enviar las columnas como
    P5 → P95.
    """
    def _pct(c: str) -> int:
        if pct_map is not None:
            return int(pct_map.get(c, exceedance_pct(c)))
        return int(exceedance_pct(c))
    return sorted([c for c in cols if c is not None], key=lambda c: (_pct(c), str(c)))


def ordered_percentile_map_by_value(df: pd.DataFrame, cols: List[str], ref_row: Optional[pd.Series] = None) -> Dict[str, int]:
    """Mapea columnas AP a **probabilidad de excedencia** según magnitud.

    Convención hidrológica usada en la app para AP:
    - P95 = caudal bajo / condición más seca, porque se excede con mayor frecuencia.
    - P5  = caudal alto / condición más húmeda, porque se excede con menor frecuencia.

    Algunos archivos DSS traen las columnas AP como percentiles no-excedentes
    (por ejemplo, el sufijo 95 puede venir asociado a caudal alto). Para evitar
    que las etiquetas de probabilidad de excedencia queden invertidas, esta
    función ordena las columnas por su valor representativo y asigna las
    probabilidades de excedencia de mayor a menor: valor más bajo → P95,
    valor más alto → P5. Si los sufijos originales ya están en convención de
    excedencia, el mapeo queda igual.
    """
    cols = [c for c in cols if c in df.columns]
    if not cols:
        return {}

    # Etiquetas disponibles, tomadas de los sufijos del DSS para no inventar
    # probabilidades que no existan en la simulación.
    labels = sorted({exceedance_pct(c) for c in cols}, reverse=True)
    if len(labels) != len(cols):
        # Respaldo ante columnas duplicadas o nombres no estándar.
        labels = sorted([exceedance_pct(c) for c in cols], reverse=True)

    values: Dict[str, float] = {}
    for c in cols:
        v = np.nan
        if ref_row is not None:
            try:
                v = pd.to_numeric(pd.Series([ref_row.get(c, np.nan)]), errors="coerce").iloc[0]
            except Exception:
                v = np.nan
        if pd.isna(v):
            try:
                v = pd.to_numeric(df[c], errors="coerce").median()
            except Exception:
                v = np.nan
        if pd.notna(v):
            values[c] = float(v)

    if not values:
        return {c: exceedance_pct(c) for c in cols}

    sorted_cols = sorted(
        values.keys(),
        key=lambda c: (values[c], exceedance_pct(c)),  # bajo→alto
    )

    pct_map: Dict[str, int] = {}
    for c, pct in zip(sorted_cols, labels):
        pct_map[c] = int(pct)

    # Columnas sin valor representativo: conservar el sufijo original.
    for c in cols:
        pct_map.setdefault(c, exceedance_pct(c))
    return pct_map


def pick_percentile_column(cols: List[str], pct_ref: int, pct_map: Optional[Dict[str, int]] = None) -> Tuple[Optional[str], Optional[int]]:
    """Selecciona la columna de percentil solicitada, con respaldo al percentil disponible más cercano.

    En NP, HP, V, EG y EP se usa el sufijo de la columna DSS. En AP puede
    recibirse un `pct_map` corregido por probabilidad de excedencia
    (menor AP → P95; mayor AP → P5). Esto evita que Manejo/Decisión quede
    vacío cuando el percentil seleccionado no existe para una variable.
    """
    valid = [c for c in cols if c is not None]
    if not valid:
        return None, None

    def _pct(c: str) -> int:
        if pct_map is not None:
            return int(pct_map.get(c, exceedance_pct(c)))
        return int(exceedance_pct(c))

    exact = [c for c in valid if _pct(c) == int(pct_ref)]
    if exact:
        return exact[0], int(pct_ref)

    # Respaldo: percentil disponible más cercano. En empate se prefiere el
    # percentil de mayor excedencia (más conservador para condiciones secas).
    best = min(valid, key=lambda c: (abs(_pct(c) - int(pct_ref)), -_pct(c)))
    return best, _pct(best)


def make_ordered_ap_columns(df: pd.DataFrame, cols: List[str], flow_unit: str, add_cfs: float = 0.0) -> Tuple[pd.DataFrame, List[str], Dict[str, int]]:
    """Crea columnas AP con etiqueta corregida como probabilidad de excedencia.

    Orden visual para Plotly/visor unificado:
    - Primero se grafican las curvas más húmedas / mayor aporte: P5, P10, P20...
    - Al final queda la curva más seca / menor aporte: P95.

    Plotly muestra el cuadro flotante en el mismo orden en que se agregan las
    trazas. Por eso aquí se devuelve el listado P5→P95, para que en el visor
    los aportes altos queden arriba y P95 quede abajo.
    """
    out = df.copy()
    cols = [c for c in cols if c in out.columns]
    pct_map = ordered_percentile_map_by_value(out, cols)
    new_cols: List[str] = []
    for c in cols:
        pct = pct_map.get(c, exceedance_pct(c))
        nc = f"AP P{pct} [{unit_label(flow_unit)}]"
        # Evita duplicados si dos columnas terminan con la misma etiqueta
        if nc in out.columns:
            nc = f"AP P{pct} {c} [{unit_label(flow_unit)}]"
        out[nc] = convert_flow(ap_total_dss_cfs(out[c], add_cfs), flow_unit)
        new_cols.append(nc)

    # Orden para la leyenda/visor de Plotly: húmedo arriba (P5) → seco abajo (P95).
    new_cols = sorted(new_cols, key=lambda c: exceedance_pct(c))
    return out, new_cols, pct_map


def to_daily(df: pd.DataFrame, cfg: Dict) -> pd.DataFrame:
    """Horario/subhorario DSS → diario. NP=last, flujos=mean."""
    if df is None or df.empty:
        return pd.DataFrame()
    data = df.copy()
    data["Fecha_dia"] = data["Fecha"].dt.floor("D")
    token = cfg["token"]
    agg: Dict[str, str] = {}
    for c in cols_by_prefix(data, "NP", token):
        agg[c] = "last"
    for px in ["HP", "AP", "V", "EG", "EP", "E"]:
        for c in cols_by_prefix(data, px, token):
            agg[c] = "mean"
    if "Observado" in data.columns:
        data["Obs_DSS"] = data["Observado"]
        agg["Obs_DSS"] = "last"
    if not agg:
        return pd.DataFrame()
    daily = data.groupby("Fecha_dia", as_index=False).agg(agg)
    return daily.sort_values("Fecha_dia")


def _ap_operational_week_start(date_series: pd.Series) -> pd.Series:
    """Inicio de semana operativa DSS para AP: sábado a viernes."""
    dt = pd.to_datetime(date_series, errors="coerce").dt.normalize()
    offset_days = (dt.dt.weekday - 5) % 7  # sábado=0, domingo=1, ..., viernes=6
    return dt - pd.to_timedelta(offset_days, unit="D")


def _ap_operational_week_pos(date_series: pd.Series) -> pd.Series:
    """Posición dentro de la semana operativa DSS: sábado=0 ... viernes=6."""
    dt = pd.to_datetime(date_series, errors="coerce").dt.normalize()
    return ((dt.dt.weekday - 5) % 7).astype("Int64")


def _last_may_hydrograph_factors(obs_aportes: Optional[pd.DataFrame]) -> Tuple[Dict[int, float], Dict[str, object]]:
    """Calcula factores de forma con 4 semanas observadas Mayo-Junio.

    La referencia solicitada para exportación AP es:
    - última semana operativa de mayo;
    - tres semanas operativas siguientes de junio.

    El resultado es una secuencia de 28 factores diarios. Luego se aplica al
    AP DSS por ciclo de 4 semanas y se renormaliza por semana operativa para
    conservar la suma/promedio semanal DSS.
    """
    meta: Dict[str, object] = {
        "ap_may_adjustment": False,
        "ap_distribution_mode": "semanal_dss",
        "reference_label": AP_HYDROGRAPH_REFERENCE_LABEL,
        "cycle_days": AP_HYDROGRAPH_REFERENCE_DAYS,
        "reference_weeks": AP_HYDROGRAPH_REFERENCE_WEEKS,
        "reason": "sin datos observados de aporte",
    }
    if not _valid_df(obs_aportes) or "Fecha_dia" not in obs_aportes.columns or "Valor" not in obs_aportes.columns:
        return {}, meta

    obs = obs_aportes[["Fecha_dia", "Valor"]].copy()
    obs["Fecha_dia"] = pd.to_datetime(obs["Fecha_dia"], errors="coerce").dt.normalize()
    obs["Valor"] = pd.to_numeric(obs["Valor"], errors="coerce")
    obs = obs.dropna(subset=["Fecha_dia", "Valor"])
    obs = obs[obs["Valor"] > 0].sort_values("Fecha_dia")
    if obs.empty:
        meta["reason"] = "sin aportes observados válidos"
        return {}, meta

    # Usar el año más reciente que tenga datos de mayo para no mezclar años.
    may_obs = obs[obs["Fecha_dia"].dt.month == 5]
    if may_obs.empty:
        meta["reason"] = "sin aportes observados en mayo"
        return {}, meta
    latest_year = int(may_obs["Fecha_dia"].dt.year.max())
    obs = obs[obs["Fecha_dia"].dt.year == latest_year].copy()
    may_obs = obs[obs["Fecha_dia"].dt.month == 5]
    if may_obs.empty:
        meta["reason"] = f"sin aportes observados de mayo en {latest_year}"
        return {}, meta

    # Última semana operativa de mayo: se toma el sábado que inicia la semana
    # donde cae el último dato válido de mayo; luego se agregan tres semanas.
    last_may_date = pd.to_datetime(may_obs["Fecha_dia"].max()).normalize()
    ref_start = pd.to_datetime(_ap_operational_week_start(pd.Series([last_may_date])).iloc[0]).normalize()
    ref_end = ref_start + pd.Timedelta(days=AP_HYDROGRAPH_REFERENCE_DAYS - 1)

    obs = obs.groupby("Fecha_dia", as_index=False).agg(Valor=("Valor", "mean"))
    window = obs[(obs["Fecha_dia"] >= ref_start) & (obs["Fecha_dia"] <= ref_end)].copy()
    valid_days = int(window["Fecha_dia"].nunique())
    if valid_days < AP_HYDROGRAPH_REFERENCE_MIN_DAYS:
        meta.update({
            "reason": (
                f"solo {valid_days} días válidos entre {ref_start:%d-%m-%Y} y {ref_end:%d-%m-%Y}; "
                f"se requieren al menos {AP_HYDROGRAPH_REFERENCE_MIN_DAYS}"
            ),
            "may_year": latest_year,
            "may_start": ref_start,
            "may_end": ref_end,
            "reference_start": ref_start,
            "reference_end": ref_end,
            "reference_days": valid_days,
        })
        return {}, meta

    # Completar huecos dentro de las 4 semanas con interpolación temporal y
    # vecinos cercanos, para que los 28 días tengan factor disponible.
    full_idx = pd.date_range(ref_start, ref_end, freq="D")
    window = window.set_index("Fecha_dia").reindex(full_idx)
    window.index.name = "Fecha_dia"
    window["Valor"] = pd.to_numeric(window["Valor"], errors="coerce")
    window["Valor"] = window["Valor"].interpolate(method="time", limit_direction="both").ffill().bfill()
    if window["Valor"].isna().all():
        meta["reason"] = "no se pudo reconstruir la ventana Mayo-Junio"
        return {}, meta

    mean_val = float(window["Valor"].mean())
    if not np.isfinite(mean_val) or mean_val <= 0:
        meta["reason"] = "promedio de aporte observado inválido en ventana Mayo-Junio"
        return {}, meta

    factors: Dict[int, float] = {}
    for i, value in enumerate(window["Valor"].to_numpy(dtype=float)):
        factor = float(value / mean_val) if np.isfinite(value) and value > 0 else 1.0
        factors[int(i)] = factor

    meta.update({
        "ap_may_adjustment": True,
        "ap_distribution_mode": "hidrograma_observado_mayo_junio_4_semanas",
        "reason": "ok",
        "may_year": latest_year,
        "may_start": ref_start,
        "may_end": ref_end,
        "reference_start": ref_start,
        "reference_end": ref_end,
        "reference_days": valid_days,
        "cycle_days": AP_HYDROGRAPH_REFERENCE_DAYS,
        "reference_weeks": AP_HYDROGRAPH_REFERENCE_WEEKS,
        "factor_min": float(min(factors.values())),
        "factor_max": float(max(factors.values())),
    })
    return factors, meta


def apply_may_hydrograph_ap_adjustment(
    daily: pd.DataFrame,
    cfg: Dict,
    obs_aportes: Optional[pd.DataFrame] = None,
    enabled: Optional[bool] = None,
) -> pd.DataFrame:
    """Ajusta la forma diaria del AP DSS con el hidrograma observado Mayo-Junio.

    Importante: no cambia el volumen/promedio semanal del DSS. Para cada columna
    AP y cada semana operativa sábado-viernes, se redistribuye la forma diaria y
    luego se reescala para que la suma semanal quede igual que en el DSS original.
    """
    if enabled is None:
        enabled = bool(st.session_state.get("ap_hydrograph_enabled", AP_HYDROGRAPH_ADJUSTMENT_ENABLED))

    if not enabled or daily is None or daily.empty:
        out = daily.copy() if isinstance(daily, pd.DataFrame) else daily
        try:
            if isinstance(out, pd.DataFrame):
                out.attrs.update({
                    "ap_may_adjustment": False,
                    "ap_distribution_mode": "semanal_dss",
                    "reference_label": AP_HYDROGRAPH_REFERENCE_LABEL,
                    "reason": "modo semanal DSS seleccionado",
                })
        except Exception:
            pass
        return out

    ap_cols = cols_by_prefix(daily, "AP", cfg.get("token", ""))
    if not ap_cols:
        return daily

    factors, meta = _last_may_hydrograph_factors(obs_aportes)
    if not factors:
        out = daily.copy()
        out.attrs.update(meta)
        return out

    out = daily.copy()
    out["Fecha_dia"] = pd.to_datetime(out["Fecha_dia"], errors="coerce").dt.normalize()
    week_start = _ap_operational_week_start(out["Fecha_dia"])

    # Secuencia de 28 días: primera semana DSS usa la última semana observada
    # de mayo; las tres semanas siguientes usan las tres semanas observadas de junio.
    try:
        dss_cycle_start = pd.to_datetime(week_start.dropna().min()).normalize()
        cycle_days = int(meta.get("cycle_days", AP_HYDROGRAPH_REFERENCE_DAYS) or AP_HYDROGRAPH_REFERENCE_DAYS)
        seq = ((out["Fecha_dia"] - dss_cycle_start).dt.days % cycle_days).astype("Int64")
        factor_series = seq.map(factors).astype(float)
    except Exception:
        factor_series = pd.Series(1.0, index=out.index)
    factor_series = factor_series.replace([np.inf, -np.inf], np.nan).fillna(1.0)

    for col in ap_cols:
        raw = pd.to_numeric(out[col], errors="coerce")
        shaped = raw * factor_series
        raw_week_sum = raw.groupby(week_start).transform("sum")
        shaped_week_sum = shaped.groupby(week_start).transform("sum")
        scale = pd.Series(1.0, index=out.index)
        valid = shaped_week_sum.replace(0, np.nan).notna() & raw_week_sum.notna()
        scale.loc[valid] = raw_week_sum.loc[valid] / shaped_week_sum.loc[valid]
        adjusted = shaped * scale
        out[col] = adjusted.where(raw.notna(), raw)

    out.attrs.update(meta)
    return out


def _show_ap_may_adjustment_note(df: pd.DataFrame, res_key: str, location: str = "") -> None:
    """Muestra una nota breve si el AP DSS fue redistribuido con hidrograma observado."""
    try:
        meta = getattr(df, "attrs", {}) or {}
        if not meta.get("ap_may_adjustment"):
            return
        start = pd.to_datetime(meta.get("reference_start", meta.get("may_start")))
        end = pd.to_datetime(meta.get("reference_end", meta.get("may_end")))
        label = str(meta.get("reference_label", AP_HYDROGRAPH_REFERENCE_LABEL))
        st.caption(
            f"Ajuste AP DSS activo{(' · ' + location) if location else ''}: se usa la forma del hidrograma observado "
            f"de {label} ({start:%d-%m-%Y} a {end:%d-%m-%Y}) y se conserva la suma semanal DSS "
            f"por semana operativa sábado-viernes."
        )
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────
# Conversión de unidades
# ─────────────────────────────────────────────────────────────────────
def convert_flow(s: pd.Series, unit: str) -> pd.Series:
    s = pd.to_numeric(s, errors="coerce")
    if unit == "m³/s":
        return s * CFS_TO_M3S
    if unit == "hm³/d":
        return s * CFS_TO_HM3_DAY
    return s


def scalar_to_cfs(v: float, unit: str) -> float:
    try:
        v = float(v)
    except Exception:
        return 0.0
    if unit == "m³/s":
        return v / CFS_TO_M3S
    if unit == "hm³/d":
        return v / CFS_TO_HM3_DAY
    return v


def pct_diff_obs_minus_dss(obs_cfs: float, dss_cfs: float) -> float:
    """Diferencia porcentual firmada respecto al percentil DSS cercano.

    Fórmula operativa:
        Diferencia Obs-DSS (%) = (Observado - DSS) / |DSS| × 100

    Un valor negativo indica que el observado está por debajo del percentil DSS
    seleccionado; un valor positivo indica que está por encima.
    """
    try:
        obs = float(obs_cfs)
        dss = float(dss_cfs)
    except Exception:
        return float("nan")
    if not np.isfinite(obs) or not np.isfinite(dss) or abs(dss) < 1e-9:
        return float("nan")
    return (obs - dss) / abs(dss) * 100.0


def pct_diff_obs_minus_dss_vs_obs(obs_cfs: float, dss_cfs: float) -> float:
    """Diferencia porcentual firmada respecto al observado.

    Se mantiene como referencia complementaria para el semáforo operativo, pero
    la columna solicitada se calcula contra el DSS mediante
    `pct_diff_obs_minus_dss`.
    """
    try:
        obs = float(obs_cfs)
        dss = float(dss_cfs)
    except Exception:
        return float("nan")
    if not np.isfinite(obs) or abs(obs) < 1e-9:
        return float("nan")
    return (obs - dss) / abs(obs) * 100.0


def clean_evap_cfs(evap_cfs: float) -> float:
    """Evaporación operativa en p³/s, siempre no negativa.

    Blindaje: el caudal evaporado **se suma** al AP neto DSS para estimar
    AP total DSS. Nunca se resta, aunque llegue como valor inválido o negativo.
    """
    try:
        return max(float(evap_cfs or 0.0), 0.0)
    except Exception:
        return 0.0


def ap_total_dss_cfs(ap_neto_cfs, evap_cfs: float) -> pd.Series:
    """AP total DSS estimado = AP neto DSS + caudal evaporado.

    Acepta escalares o Series y retorna una Series numérica en p³/s.
    Blindaje operativo: la evaporación siempre entra como término positivo,
    por lo que el AP total DSS nunca se calcula restando evaporación.
    """
    if isinstance(ap_neto_cfs, pd.Series):
        ap = pd.to_numeric(ap_neto_cfs, errors="coerce").copy()
    else:
        ap = pd.to_numeric(pd.Series(ap_neto_cfs), errors="coerce")
    evap = clean_evap_cfs(evap_cfs)
    return ap + evap


# ─────────────────────────────────────────────────────────────────────
# Utilidades de gráficas
# ─────────────────────────────────────────────────────────────────────
def _today_line(fig, date_series) -> None:
    """Línea vertical 'Hoy' compatible con todos los rangos de fechas."""
    try:
        fechas = pd.to_datetime(date_series, errors="coerce").dropna()
        if fechas.empty:
            return
        today = pd.Timestamp.today().normalize()
        if not (fechas.min().normalize() <= today <= fechas.max().normalize()):
            return
        x_today = today.to_pydatetime()
        fig.add_shape(type="line", x0=x_today, x1=x_today, y0=0, y1=1,
                      xref="x", yref="paper",
                      line=dict(width=2, dash="dash", color="rgba(255,140,0,0.85)"))
        fig.add_annotation(x=x_today, y=1, xref="x", yref="paper",
                           text="Hoy", showarrow=False, yshift=12,
                           font=dict(color="darkorange", size=11),
                           bgcolor="rgba(255,255,255,0.75)",
                           bordercolor="rgba(255,140,0,0.55)", borderwidth=1)
    except Exception:
        pass


def _base_layout(title: str, y_title: str, height: int = 560) -> dict:
    return dict(
        title=dict(text=title, font=dict(size=14, color="#003E69")),
        height=height,
        hovermode="x unified",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(l=10, r=10, t=60, b=10),
        plot_bgcolor="rgba(250,252,255,1)",
        paper_bgcolor="rgba(250,252,255,0)",
        yaxis=dict(title=y_title, gridcolor="rgba(0,0,0,0.07)"),
        xaxis=dict(title="Fecha", gridcolor="rgba(0,0,0,0.07)"),
    )


def _apply_full_yaxis_from_traces(fig, force_zero: bool = True, pad_fraction: float = 0.08) -> None:
    """Ajusta el eje Y usando el rango completo de las trazas visibles.

    Antes el gráfico diario de aportes usaba percentiles 2%-98% para evitar
    que valores extremos aplastaran la curva observada. Eso podía ocultar
    picos DSS y hacer que el gráfico diario se viera distinto al semanal.
    Para AP se usa el rango completo, de modo que diario y semanal muestren
    la misma magnitud real de las curvas.
    """
    try:
        y_parts = []
        for tr in fig.data:
            if getattr(tr, "visible", True) == "legendonly":
                continue
            y = getattr(tr, "y", None)
            if y is None:
                continue
            vals = pd.to_numeric(pd.Series(y), errors="coerce")
            vals = vals[np.isfinite(vals)]
            if not vals.empty:
                y_parts.append(vals)
        if not y_parts:
            return
        y_all = pd.concat(y_parts, ignore_index=True)
        if y_all.empty:
            return
        y_min = float(y_all.min())
        y_max = float(y_all.max())
        if not np.isfinite(y_min) or not np.isfinite(y_max):
            return
        if y_max <= y_min:
            pad = max(abs(y_max) * pad_fraction, 1.0)
        else:
            pad = max((y_max - y_min) * pad_fraction, 1.0)
        lower = 0.0 if force_zero and y_min >= 0 else y_min - pad
        upper = y_max + pad
        if upper > lower:
            fig.update_yaxes(range=[lower, upper])
    except Exception:
        pass


def fan_chart(df: pd.DataFrame, cols: List[str], title: str, y_label: str, key: str,
              obs_col: Optional[str] = None, obs_label: str = "Observado",
              show_band: bool = True) -> None:
    """Abanico de percentiles con banda P90-P10 y traza observada."""
    if df.empty or not cols:
        st.info("Sin datos para graficar.")
        return
    cols = order_cols_wet_to_dry([c for c in cols if c in df.columns])
    if not cols:
        st.info("Columnas no disponibles.")
        return

    plot_df = df[["Fecha_dia"] + cols +
                 ([obs_col] if obs_col and obs_col in df.columns else [])].copy()
    plot_df = plot_df.dropna(how="all", subset=cols)
    if len(plot_df) > 4000:
        plot_df = plot_df.iloc[::max(1, len(plot_df) // 4000)]

    if not PLOTLY_OK:
        st.line_chart(plot_df.set_index("Fecha_dia")[cols])
        return

    fig = go.Figure()

    # Banda de incertidumbre P90-P10
    if show_band and len(cols) >= 2:
        p90 = next((c for c in cols if "90" in c), None)
        p10 = next((c for c in cols if "10" in c), None)
        if p90 and p10:
            fig.add_trace(go.Scatter(
                x=pd.concat([plot_df["Fecha_dia"], plot_df["Fecha_dia"][::-1]]),
                y=pd.concat([plot_df[p90], plot_df[p10][::-1]]),
                fill="toself", fillcolor="rgba(0,102,204,0.08)",
                line=dict(color="rgba(255,255,255,0)"),
                name="Banda P90-P10", showlegend=True, hoverinfo="skip",
            ))

    for col in cols:
        exc = exceedance_pct(col)
        is_p50 = exc == 50
        fig.add_trace(go.Scatter(
            x=plot_df["Fecha_dia"], y=plot_df[col], mode="lines",
            name=f"P{exc}",
            line=dict(
                color=EXCEEDANCE_COLORS.get(exc, "#999"),
                width=2.5 if is_p50 else 1.2,
                dash="dot" if exc in (90, 10) else "solid",
            ),
        ))

    if obs_col and obs_col in plot_df.columns:
        obs_v = plot_df[plot_df[obs_col].notna()].copy()
        if not obs_v.empty:
            lv = float(obs_v[obs_col].iloc[-1])
            labels = [""] * len(obs_v)
            labels[-1] = f"{lv:,.2f} {y_label}"
            fig.add_trace(go.Scatter(
                x=obs_v["Fecha_dia"], y=obs_v[obs_col],
                mode="lines+markers+text", name=f"🔴 {obs_label}",
                text=labels, textposition="top right",
                line=dict(color="#d90429", width=5),
                marker=dict(size=8, color="#d90429", line=dict(width=1.5, color="white")),
                connectgaps=True,
            ))
            fig.add_trace(go.Scatter(
                x=[obs_v["Fecha_dia"].iloc[-1]], y=[lv],
                mode="markers", name="📍 Último obs.",
                marker=dict(size=18, color="#d90429", symbol="star",
                            line=dict(width=2, color="white")),
            ))

    _today_line(fig, plot_df["Fecha_dia"])
    fig.update_layout(**_base_layout(title, y_label))
    st.plotly_chart(fig, use_container_width=True, key=key)


# ─────────────────────────────────────────────────────────────────────
# Percentil más cercano al observado
# ─────────────────────────────────────────────────────────────────────
def closest_np(daily: pd.DataFrame, cfg: Dict,
               obs_date: Optional[pd.Timestamp], obs_val: Optional[float]) -> Optional[Dict]:
    """Percentil NP DSS más cercano al nivel observado."""
    if daily is None or daily.empty or obs_date is None or obs_val is None:
        return None
    base = daily.copy()
    base["Fecha_dia"] = pd.to_datetime(base["Fecha_dia"], errors="coerce").dt.normalize()
    base = base[base["Fecha_dia"].notna()].sort_values("Fecha_dia")
    np_cols = cols_by_prefix(base, "NP", cfg["token"])
    if not np_cols:
        return None
    obs_day = pd.to_datetime(obs_date).normalize()
    exact = base[base["Fecha_dia"] == obs_day]
    row = exact.iloc[0] if not exact.empty else base.loc[(base["Fecha_dia"] - obs_day).abs().idxmin()]
    candidates = []
    for col in np_cols:
        v = pd.to_numeric(pd.Series([row.get(col, np.nan)]), errors="coerce").iloc[0]
        if pd.notna(v):
            candidates.append((col, float(v), abs(float(obs_val) - float(v))))
    if not candidates:
        return None
    col, dss_v, _ = min(candidates, key=lambda x: x[2])
    return {
        "label": f"P{exceedance_pct(col)}", "column": col,
        "dss_value": dss_v, "diff": float(obs_val) - dss_v,
        "date": pd.to_datetime(row["Fecha_dia"]),
    }


# ─────────────────────────────────────────────────────────────────────
# Filtro de fechas
# ─────────────────────────────────────────────────────────────────────
def date_filter(df: pd.DataFrame, key: str,
                default_days: int = 0) -> pd.DataFrame:
    """Filtro de fechas. default_days=0 usa rango completo."""
    if df.empty or "Fecha_dia" not in df.columns:
        return df
    valid = df["Fecha_dia"].dropna()
    if valid.empty:
        return df
    mn, mx = valid.min().date(), valid.max().date()
    if default_days > 0:
        import datetime
        default_start = max(mn, (pd.Timestamp(mx) - pd.Timedelta(days=default_days)).date())
    else:
        default_start = mn
    c1, c2 = st.columns(2)
    s = c1.date_input("Desde", value=default_start, min_value=mn, max_value=mx, key=f"{key}_s")
    e = c2.date_input("Hasta", value=mx, min_value=mn, max_value=mx, key=f"{key}_e")
    if s > e:
        st.warning("Fecha inicial > Fecha final. Se usa el período completo.")
        return df
    return df.loc[(df["Fecha_dia"].dt.date >= s) & (df["Fecha_dia"].dt.date <= e)].copy()


# ─────────────────────────────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────────────────────────────
def sidebar() -> Dict:
    st.sidebar.markdown(f"## 💧 {APP_TITLE.split('·')[0].strip()}")
    st.sidebar.markdown("---")
    st.sidebar.header("📁 Archivos DSS")

    dss_up = st.sidebar.file_uploader("DSS (SimulacionDSS…xlsx)", type=["xlsx", "xlsm"], key="dss_up")

    if st.sidebar.button("🔄 Recargar archivos", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    if dss_up:
        dss_bytes = dss_up.getvalue()
        dss_name  = dss_up.name
    else:
        dss_bytes, dss_path = read_first_local(DSS_NAMES)
        dss_name = dss_path.name if dss_path else "—"

    st.sidebar.caption(f"DSS: **{dss_name}** · búsqueda local: `data/` → carpeta del app")

    st.sidebar.header("🕰️ Serie histórica 1997-1998")
    hist_9798_up = st.sidebar.file_uploader(
        "Aportes diarios 1997-1998 (Excel)",
        type=["xlsx", "xlsm"],
        key="hist_9798_up",
        help="Debe incluir Fecha, Gatún y Alhajuela. La suma CHCP se calcula como Gatún + Alhajuela.",
    )
    if hist_9798_up:
        hist_9798_bytes = hist_9798_up.getvalue()
        hist_9798_name = hist_9798_up.name
    else:
        hist_9798_bytes, hist_9798_path = read_first_local_exact(HISTORICAL_9798_NAMES)
        hist_9798_name = hist_9798_path.name if hist_9798_path else "—"
    if hist_9798_bytes:
        st.sidebar.caption(f"Histórico: **{hist_9798_name}** · carga automática o manual")
    else:
        st.sidebar.caption("Histórico: no cargado. Coloque `1997_1998.xlsx` en `data/` o cárguelo aquí.")

    st.sidebar.header("⚙️ Ajustes")
    flow_unit = st.sidebar.radio(
        "🔁 Unidad de caudal / flujo",
        ["cfs", "m³/s", "hm³/d"], index=0,
        format_func=unit_label,
        help="p³/s = pies cúbicos por segundo · m³/s · hm³/d",
    )
    pct_ref_gat = st.sidebar.selectbox(
        "Percentil de referencia Gatún",
        PERCENTILE_ORDER, index=PERCENTILE_ORDER.index(50),
        format_func=lambda x: f"P{x}",
        key="pct_ref_gat",
        help="Percentil DSS de referencia para las pestañas y métricas de Gatún.",
    )
    pct_ref_alh = st.sidebar.selectbox(
        "Percentil de referencia Alhajuela/Madden",
        PERCENTILE_ORDER, index=PERCENTILE_ORDER.index(50),
        format_func=lambda x: f"P{x}",
        key="pct_ref_alh",
        help="Percentil DSS de referencia para las pestañas y métricas de Alhajuela/Madden.",
    )

    st.sidebar.header("🌊 Distribución de aportes DSS")
    ap_distribution_mode = st.sidebar.radio(
        "Forma diaria de los aportes DSS",
        ["Simular hidrograma observado mayo-junio", "Ver aporte semanal DSS"],
        index=1,
        key="ap_distribution_mode",
        help=(
            "Simular hidrograma observado mayo-junio redistribuye la forma diaria del AP con 4 semanas observadas, "
            "pero conserva el volumen/promedio semanal DSS. Ver aporte semanal DSS deja "
            "el AP como viene del DSS."
        ),
    )
    ap_hydrograph_enabled = ap_distribution_mode.startswith("Simular")
    st.session_state["ap_hydrograph_enabled"] = bool(ap_hydrograph_enabled)
    if ap_hydrograph_enabled:
        st.sidebar.caption("Modo activo: AP DSS redistribuido con la última semana de mayo y tres semanas siguientes de junio, conservando la suma semanal.")
    else:
        st.sidebar.caption("Modo activo: AP semanal DSS original, sin redistribución diaria.")

    st.sidebar.header("🌫️ Caudal evaporado para ajuste AP DSS")
    evap_mode = st.sidebar.radio(
        "Fuente de evaporación",
        ["Automática · promedio móvil 5 días CZL/PMG", "Manual · p³/s"],
        index=0,
        key="evap_source_mode",
        help=(
            "Automática: usa el promedio móvil de los últimos 5 días observados válidos de "
            "Corozal (CZL) para Gatún y Pedro Miguel (PMG) para Alhajuela. "
            "Manual: permite ingresar el caudal en p³/s."
        ),
    )

    evap_gat_meta: Dict[str, object] = {"mode": evap_mode}
    evap_alh_meta: Dict[str, object] = {"mode": evap_mode}

    if evap_mode.startswith("Automática"):
        gat_series = latest_evap_series("CZL")
        alh_series = latest_evap_series("PMG")

        if gat_series.get("mm_day") is not None:
            gat_level = latest_local_observed_level_for_evap("gatun")
            gat_area, gat_area_status = evap_area_from_observed_level("gatun", gat_level.get("level_ft"))
            gat_calc = evap_mm_to_flows(float(gat_series["mm_day"]), gat_area)
            evap_gat_cfs = float(gat_calc["cfs"])
            evap_gat_meta.update(gat_series)
            evap_gat_meta.update(gat_calc)
            evap_gat_meta.update({"level_info": gat_level, "area_status": gat_area_status})
            st.sidebar.markdown(
                f"**Gatún · Corozal (CZL)**  \n"
                f"Prom. móvil {gat_series.get('n_obs', 0)}/5 días: {gat_calc['mm_day']:.3f} mm/día · "
                f"{gat_calc['hm3_day']:.4f} hm³/día · "
                f"**{gat_calc['cfs']:.3f} p³/s**"
            )
            if gat_level.get("level_ft") is not None:
                st.sidebar.caption(
                    f"Ventana evap.: {pd.to_datetime(gat_series['window_start']):%d-%m-%Y} a "
                    f"{pd.to_datetime(gat_series['window_end']):%d-%m-%Y} · "
                    f"Nivel obs: {gat_level['level_ft']:.2f} ft · Área por nivel: {gat_area:.1f} km² · "
                    f"Archivo nivel: {gat_level.get('source') or '—'}"
                )
            else:
                st.sidebar.caption(
                    f"Ventana evap.: {pd.to_datetime(gat_series['window_start']):%d-%m-%Y} a "
                    f"{pd.to_datetime(gat_series['window_end']):%d-%m-%Y} · "
                    f"Sin nivel observado local; usa área respaldo: {gat_area:.1f} km² · Archivo: {gat_series['file']}"
                )
        else:
            evap_gat_cfs = 0.0
            st.sidebar.warning("No se encontró una serie válida de evaporación CZL para Gatún.")

        if alh_series.get("mm_day") is not None:
            alh_level = latest_local_observed_level_for_evap("alhajuela")
            alh_area, alh_area_status = evap_area_from_observed_level("alhajuela", alh_level.get("level_ft"))
            alh_calc = evap_mm_to_flows(float(alh_series["mm_day"]), alh_area)
            evap_alh_cfs = float(alh_calc["cfs"])
            evap_alh_meta.update(alh_series)
            evap_alh_meta.update(alh_calc)
            evap_alh_meta.update({"level_info": alh_level, "area_status": alh_area_status})
            st.sidebar.markdown(
                f"**Alhajuela · Pedro Miguel (PMG)**  \n"
                f"Prom. móvil {alh_series.get('n_obs', 0)}/5 días: {alh_calc['mm_day']:.3f} mm/día · "
                f"{alh_calc['hm3_day']:.4f} hm³/día · "
                f"**{alh_calc['cfs']:.3f} p³/s**"
            )
            if alh_level.get("level_ft") is not None:
                st.sidebar.caption(
                    f"Ventana evap.: {pd.to_datetime(alh_series['window_start']):%d-%m-%Y} a "
                    f"{pd.to_datetime(alh_series['window_end']):%d-%m-%Y} · "
                    f"Nivel obs: {alh_level['level_ft']:.2f} ft · Área por nivel: {alh_area:.1f} km² · "
                    f"Archivo nivel: {alh_level.get('source') or '—'}"
                )
            else:
                st.sidebar.caption(
                    f"Ventana evap.: {pd.to_datetime(alh_series['window_start']):%d-%m-%Y} a "
                    f"{pd.to_datetime(alh_series['window_end']):%d-%m-%Y} · "
                    f"Sin nivel observado local; usa área respaldo: {alh_area:.1f} km² · Archivo: {alh_series['file']}"
                )
        else:
            evap_alh_cfs = 0.0
            st.sidebar.warning("No se encontró una serie válida de evaporación PMG para Alhajuela.")

        st.sidebar.caption(
            "Promedio móvil: media de los últimos 5 días observados válidos. "
            "Fórmula automática: hm³/día = promedio mm/día × área (km²) × 0.85 × 0.001. "
            "Luego se convierte a p³/s."
        )
    else:
        evap_gat_cfs = st.sidebar.number_input(
            "Evaporación GAT (p³/s)",
            min_value=0.0, value=0.0, step=10.0, format="%.3f",
            help="Se suma al AP neto DSS de Gatún para estimar AP total DSS."
        )
        evap_alh_cfs = st.sidebar.number_input(
            "Evaporación ALHA (p³/s)",
            min_value=0.0, value=0.0, step=10.0, format="%.3f",
            help="Se suma al AP neto DSS de Alhajuela/Madden para estimar AP total DSS."
        )
        evap_gat_meta.update({"cfs": float(evap_gat_cfs), "source": "Manual"})
        evap_alh_meta.update({"cfs": float(evap_alh_cfs), "source": "Manual"})

    st.sidebar.caption("AP total DSS estimado = AP neto DSS + caudal evaporado.")

    st.sidebar.header("📖 Glosario")
    for var, desc in [
        ("NP", "Nivel proyectado (ft PLD)"),
        ("HP", "Hidrogeneración (MW)"),
        ("AP", "Aportes al embalse (p³/s)"),
        ("V",  "Vertidos / Spill (p³/s)"),
        ("EG", "Esclusaje Gatún"),
        ("EP", "Esclusaje Panamax"),
        ("P90…P5", "Prob. excedencia"),
    ]:
        st.sidebar.markdown(f"**{var}**: {desc}")

    st.sidebar.markdown("---")
    st.sidebar.caption(AUTHOR_NOTE)
    if "view_count" in st.session_state:
        st.sidebar.caption(f"👁️ Vistas: {int(st.session_state['view_count']):,}")

    return {
        "dss_bytes": dss_bytes,
        "hist_9798_bytes": hist_9798_bytes,
        "hist_9798_name": hist_9798_name,
        "flow_unit": flow_unit,
        "pct_ref_gat": int(pct_ref_gat),
        "pct_ref_alh": int(pct_ref_alh),
        # Compatibilidad con versiones anteriores: si alguna función antigua consulta pct_ref,
        # se usa Gatún como referencia por defecto.
        "pct_ref": int(pct_ref_gat),
        "evap_gat_cfs": float(evap_gat_cfs),
        "evap_alh_cfs": float(evap_alh_cfs),
        "evap_mode": evap_mode,
        "evap_gat_meta": evap_gat_meta,
        "evap_alh_meta": evap_alh_meta,
        "ap_distribution_mode": ap_distribution_mode,
        "ap_hydrograph_enabled": bool(ap_hydrograph_enabled),
    }


# ─────────────────────────────────────────────────────────────────────
# Métricas de cabecera
# ─────────────────────────────────────────────────────────────────────
def show_header_metrics(daily: pd.DataFrame, cfg: Dict, flow_unit: str,
                        obs_daily: Optional[pd.DataFrame] = None,
                        obs_aportes: Optional[pd.DataFrame] = None,
                        evap_cfs: float = 0.0,
                        pct_ref: int = 50) -> None:
    """Métricas principales usando el percentil más cercano disponible.

    Corrección puntual:
    - NP se etiqueta con el percentil más cercano al nivel observado.
    - AP se etiqueta con el percentil más cercano al aporte observado de Aquarius,
      considerando AP total DSS = AP neto DSS + evaporación.
    - HP sigue el percentil AP más cercano cuando existe aporte observado; si no,
      usa el percentil más cercano al nivel observado. Así las tarjetas no quedan
      mezcladas con P50/P90/P95 sin relación con el dato observado.
    """
    if daily.empty:
        return
    token = cfg["token"]
    np_cols = cols_by_prefix(daily, "NP", token)
    hp_cols = cols_by_prefix(daily, "HP", token)
    ap_cols = cols_by_prefix(daily, "AP", token)

    today = today_panama()
    srt = daily.copy()
    srt["Fecha_dia"] = pd.to_datetime(srt["Fecha_dia"], errors="coerce").dt.normalize()
    srt = srt[srt["Fecha_dia"].notna()].sort_values("Fecha_dia")
    if srt.empty:
        return
    exact = srt[srt["Fecha_dia"] == today]
    past = srt[srt["Fecha_dia"] <= today]
    rec = exact.iloc[0] if not exact.empty else (past.iloc[-1] if not past.empty else srt.iloc[-1])

    # Último nivel observado y percentil NP más cercano.
    obs_val, obs_date, closest_level = None, None, None
    if obs_daily is not None and isinstance(obs_daily, pd.DataFrame) and not obs_daily.empty:
        obs_valid = obs_daily[obs_daily["Valor"].notna()].copy()
        obs_valid["Fecha_dia"] = pd.to_datetime(obs_valid["Fecha_dia"], errors="coerce").dt.normalize()
        obs_valid = obs_valid[obs_valid["Fecha_dia"].notna()].sort_values("Fecha_dia")
        obs_past = obs_valid[obs_valid["Fecha_dia"] <= today]
        if not obs_past.empty:
            last_obs = obs_past.iloc[-1]
        elif not obs_valid.empty:
            last_obs = obs_valid.iloc[-1]
        else:
            last_obs = None
        if last_obs is not None:
            obs_val = float(last_obs["Valor"])
            obs_date = pd.to_datetime(last_obs["Fecha_dia"])
            closest_level = closest_np(srt, cfg, obs_date, obs_val)

    level_pct = int(str(closest_level["label"]).replace("P", "")) if closest_level else int(pct_ref)
    np_col, np_pct = pick_percentile_column(np_cols, level_pct)
    np_v = float(rec.get(np_col, np.nan)) if np_col else np.nan

    # Último aporte observado y percentil AP más cercano.
    ap_nearest = None
    obs_ap_val_cfs, obs_ap_date = np.nan, None
    if obs_aportes is not None and isinstance(obs_aportes, pd.DataFrame) and not obs_aportes.empty:
        obs_ap = clamp_observed_future_dates(obs_aportes, "Fecha_dia")
        obs_ap = obs_ap[obs_ap["Valor"].notna()].copy()
        obs_ap["Fecha_dia"] = pd.to_datetime(obs_ap["Fecha_dia"], errors="coerce").dt.normalize()
        obs_ap = obs_ap[obs_ap["Fecha_dia"].notna()].sort_values("Fecha_dia")
        obs_ap_past = obs_ap[obs_ap["Fecha_dia"] <= today]
        if not obs_ap_past.empty:
            last_ap = obs_ap_past.iloc[-1]
        elif not obs_ap.empty:
            last_ap = obs_ap.iloc[-1]
        else:
            last_ap = None
        if last_ap is not None:
            obs_ap_val_cfs = float(last_ap["Valor"])
            obs_ap_date = pd.to_datetime(last_ap["Fecha_dia"])
            ap_nearest = _nearest_ap_percentile(
                srt, cfg, obs_ap_date, obs_ap_val_cfs, dss_add_cfs=clean_evap_cfs(evap_cfs)
            )

    ap_pct = int(ap_nearest["percentile"]) if ap_nearest else level_pct
    ap_pct_map = ordered_percentile_map_by_value(srt, ap_cols, ref_row=rec) if ap_cols else {}
    ap_col, ap_pct_used = pick_percentile_column(ap_cols, ap_pct, pct_map=ap_pct_map)
    hp_col, hp_pct_used = pick_percentile_column(hp_cols, ap_pct if ap_nearest else level_pct)

    ap_dss_cfs = ap_total_dss_cfs([rec.get(ap_col, np.nan)], evap_cfs).iloc[0] if ap_col else np.nan
    ap_dss_v = convert_flow(pd.Series([ap_dss_cfs]), flow_unit).iloc[0] if pd.notna(ap_dss_cfs) else np.nan
    hp_v = float(rec.get(hp_col, np.nan)) if hp_col else np.nan

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("📅 Referencia DSS", rec["Fecha_dia"].strftime("%d-%m-%Y"))
    if obs_val is not None:
        delta = f"Δ vs NP P{np_pct}: {obs_val - np_v:+.3f} ft" if pd.notna(np_v) and np_pct is not None else None
        c2.metric(f"🔴 Obs. {cfg['level_unit']}", f"{obs_val:,.3f}", delta=delta,
                  delta_color="inverse",
                  help=f"Último observado: {obs_date:%d-%m-%Y}" if obs_date else None)
    else:
        c2.metric("🔴 Obs. LKH", "—")
    c3.metric(f"NP cercano P{np_pct if np_pct is not None else '—'} ({cfg['level_unit']})",
              f"{np_v:,.3f}" if pd.notna(np_v) else "—")
    c4.metric(f"HP cercano P{hp_pct_used if hp_pct_used is not None else '—'} (MW)",
              f"{hp_v:,.2f}" if pd.notna(hp_v) else "—")
    c5.metric(f"AP DSS P{ap_pct_used if ap_pct_used is not None else '—'} ({unit_label(flow_unit)})",
              f"{ap_dss_v:,.2f}" if pd.notna(ap_dss_v) else "—",
              help="AP total DSS estimado = AP neto DSS + evaporación.")
    if closest_level or ap_nearest:
        level_txt = closest_level["label"] if closest_level else "—"
        ap_txt = ap_nearest["label"] if ap_nearest else "—"
        delta_txt = None
        help_txt = []
        if closest_level:
            delta_txt = f"Nivel: {closest_level['diff']:+.3f} ft"
            help_txt.append(f"Nivel DSS {closest_level['label']}: {closest_level['dss_value']:.3f} ft · {closest_level['date']:%d-%m-%Y}")
        if ap_nearest:
            ap_diff = convert_flow(pd.Series([ap_nearest['diff_cfs']]), flow_unit).iloc[0]
            ap_pct = ap_nearest.get('diff_pct_dss', np.nan)
            ap_pct_txt = f" ({ap_pct:+.2f}% vs DSS)" if pd.notna(ap_pct) else ""
            help_txt.append(f"AP DSS {ap_nearest['label']}: {convert_flow(pd.Series([ap_nearest['dss_total_cfs']]), flow_unit).iloc[0]:,.2f} {unit_label(flow_unit)} · Obs-DSS {ap_diff:+,.2f} {unit_label(flow_unit)}{ap_pct_txt}")
        c6.metric("🎯 Percentil cercano", f"NP {level_txt} · AP {ap_txt}",
                  delta=delta_txt, delta_color="inverse", help=" | ".join(help_txt) if help_txt else None)
    else:
        c6.metric("🎯 Percentil cercano", "—")


def show_aporte_reservoir_metrics(
    daily: pd.DataFrame,
    cfg: Dict,
    flow_unit: str,
    pct_ref: int,
    obs_aportes: Optional[pd.DataFrame] = None,
    evap_cfs: float = 0.0,
) -> None:
    """Muestra aporte observado, AP DSS ajustado por evaporación y HP DSS recomendada."""
    if daily is None or daily.empty:
        return

    token = cfg["token"]
    ap_cols = cols_by_prefix(daily, "AP", token)
    hp_cols = cols_by_prefix(daily, "HP", token)
    if not ap_cols:
        return

    base = daily.copy()
    base["Fecha_dia"] = pd.to_datetime(base["Fecha_dia"], errors="coerce").dt.normalize()
    base = base[base["Fecha_dia"].notna()].sort_values("Fecha_dia")
    today = today_panama()

    exact = base[base["Fecha_dia"] == today]
    if not exact.empty:
        ref_row = exact.iloc[0]
    else:
        past = base[base["Fecha_dia"] <= today]
        ref_row = past.iloc[-1] if not past.empty else base.iloc[-1]

    obs_val_cfs = np.nan
    obs_date = None
    if obs_aportes is not None and isinstance(obs_aportes, pd.DataFrame) and not obs_aportes.empty:
        obs = clamp_observed_future_dates(obs_aportes, "Fecha_dia")
        obs = obs[obs["Valor"].notna()].copy()
        obs["Fecha_dia"] = pd.to_datetime(obs["Fecha_dia"], errors="coerce").dt.normalize()
        obs = obs[obs["Fecha_dia"].notna()].sort_values("Fecha_dia")
        obs_past = obs[obs["Fecha_dia"] <= today]
        if not obs_past.empty:
            last_obs = obs_past.iloc[-1]
        elif not obs.empty:
            last_obs = obs.iloc[-1]
        else:
            last_obs = None
        if last_obs is not None:
            obs_val_cfs = float(last_obs["Valor"])
            obs_date = pd.to_datetime(last_obs["Fecha_dia"])

    obs_val = convert_flow(pd.Series([obs_val_cfs]), flow_unit).iloc[0] if pd.notna(obs_val_cfs) else np.nan
    nearest_obs = None
    if pd.notna(obs_val_cfs) and obs_date is not None:
        nearest_obs = _nearest_ap_percentile(base, cfg, obs_date, obs_val_cfs, dss_add_cfs=clean_evap_cfs(evap_cfs))

    indicador_pct = int(nearest_obs.get("percentile", pct_ref)) if nearest_obs else int(pct_ref)
    ap_pct_map = ordered_percentile_map_by_value(base, ap_cols, ref_row=ref_row)
    ap_ref, ap_ref_pct = pick_percentile_column(ap_cols, indicador_pct, pct_map=ap_pct_map)

    ap_dss_total_cfs = ap_total_dss_cfs([ref_row.get(ap_ref, np.nan)], evap_cfs).iloc[0] if ap_ref else np.nan
    ap_dss_total = convert_flow(pd.Series([ap_dss_total_cfs]), flow_unit).iloc[0] if pd.notna(ap_dss_total_cfs) else np.nan

    # Hidrogeneración DSS recomendada: usa el mismo percentil operativo del aporte más cercano.
    hp_rec_val = np.nan
    hp_rec_label = "—"
    hp_col, hp_pct = pick_percentile_column(hp_cols, indicador_pct)
    if hp_col is not None:
        hp_rec_val = float(ref_row.get(hp_col, np.nan))
        hp_rec_label = f"P{hp_pct}" if hp_pct is not None else f"P{indicador_pct}"

    st.markdown("#### 🌧️ Aporte observado, AP DSS ajustado e hidrogeneración")
    st.caption(
        f"{SIMULATION_NOTE} · AP total DSS estimado = AP neto DSS + evaporación. "
        "Las tarjetas usan el percentil más cercano al aporte observado cuando está disponible."
    )

    m1, m2, m3, m4, m5, m6 = st.columns([1.05, 0.85, 1.15, 1.05, 1.10, 1.05])
    m1.metric(
        f"Último aporte observado ({unit_label(flow_unit)})",
        f"{obs_val:,.2f}" if pd.notna(obs_val) else "—",
        help=f"Fecha: {obs_date:%d-%m-%Y}" if obs_date is not None else "Sin BulkExport de aporte observado.",
    )
    m2.metric("Fecha aporte obs.", f"{obs_date:%d-%m-%Y}" if obs_date is not None else "—")
    m3.metric("Evaporación aplicada (p³/s)", f"{clean_evap_cfs(evap_cfs):,.1f}")
    m4.metric(
        f"AP total DSS P{ap_ref_pct if ap_ref_pct is not None else '—'} ({unit_label(flow_unit)})",
        f"{ap_dss_total:,.2f}" if pd.notna(ap_dss_total) else "—",
    )
    if nearest_obs:
        nearest_diff_unit = convert_flow(pd.Series([nearest_obs['diff_cfs']]), flow_unit).iloc[0]
        nearest_total_unit = convert_flow(pd.Series([nearest_obs['dss_total_cfs']]), flow_unit).iloc[0]
        pct_dss = nearest_obs.get('diff_pct_dss', np.nan)
        pct_txt = f" · {pct_dss:+.2f}%" if pd.notna(pct_dss) else ""
        m5.metric(
            "Indicador AP observado",
            nearest_obs["label"],
            delta=f"Obs-DSS: {nearest_diff_unit:+,.2f} {unit_label(flow_unit)}{pct_txt}",
            delta_color="inverse",
            help=f"AP total DSS indicador: {nearest_total_unit:,.2f} {unit_label(flow_unit)} "
                 f"({nearest_obs['dss_total_cfs']:,.1f} p³/s) · Diferencia Obs-DSS: {pct_dss:+.2f}% respecto al DSS · Fecha DSS: {nearest_obs['date']:%d-%m-%Y}",
        )
    else:
        m5.metric("Indicador AP observado", "—")
    m6.metric(
        f"Hidrogeneración DSS recomendada ({hp_rec_label})",
        f"{hp_rec_val:,.2f} MW" if pd.notna(hp_rec_val) else "—",
    )


# ─────────────────────────────────────────────────────────────────────
# PESTAÑA 1 — GATÚN DSS
# PESTAÑA 2 — ALHAJUELA DSS
# ─────────────────────────────────────────────────────────────────────
def tab_reservoir(res_key: str, dss_bytes: bytes, flow_unit: str, pct_ref: int,
                  obs_niveles: Optional[pd.DataFrame] = None,
                  obs_aportes: Optional[pd.DataFrame] = None,
                  evap_cfs: float = 0.0) -> None:
    cfg = RESERVOIR_CONFIG[res_key]
    st.subheader(f"💧 {cfg['name']}")
    st.caption(PROJ_NOTE)

    try:
        dss_raw = load_dss_sheet(dss_bytes, cfg["sheet"])
    except Exception as exc:
        st.error(f"Error cargando DSS: {exc}")
        return

    daily = to_daily(dss_raw, cfg)
    daily = apply_may_hydrograph_ap_adjustment(daily, cfg, obs_aportes)
    if daily.empty:
        st.warning("No se pudo construir el diario DSS.")
        return

    with st.expander("🗓️ Filtro de período", expanded=True):
        filtered = date_filter(daily, f"{res_key}_res")

    show_header_metrics(filtered, cfg, flow_unit, obs_niveles, obs_aportes, evap_cfs, pct_ref)
    if obs_niveles is not None and not obs_niveles.empty:
        last = obs_niveles[obs_niveles["Valor"].notna()].sort_values("Fecha_dia")
        if not last.empty:
            st.caption(f"Último nivel observado disponible: "
                       f"{last.iloc[-1]['Fecha_dia']:%d-%m-%Y} · {last.iloc[-1]['Valor']:,.3f} ft")

    show_aporte_reservoir_metrics(filtered, cfg, flow_unit, pct_ref, obs_aportes, evap_cfs)
    _show_ap_may_adjustment_note(filtered, res_key, "pestaña del embalse")
    st.markdown("---")

    token = cfg["token"]
    np_cols = cols_by_prefix(filtered, "NP", token)
    hp_cols = cols_by_prefix(filtered, "HP", token)
    ap_cols = cols_by_prefix(filtered, "AP", token)
    v_cols  = cols_by_prefix(filtered, "V",  token)

    # --- NP con observado ---
    st.markdown("### 📈 Nivel proyectado vs observado")
    default_np = [c for c in np_cols if any(x in c for x in ["10", "50", "90"])]
    sel_np = st.multiselect("Series NP", np_cols, default=default_np, key=f"{res_key}_np")

    obs_col_in_df = None
    if obs_niveles is not None and not obs_niveles.empty:
        # Merge nivel observado
        obs_merge = obs_niveles[["Fecha_dia", "Valor"]].rename(columns={"Valor": "Nivel obs."})
        plot_df = filtered.merge(obs_merge, on="Fecha_dia", how="left")
        obs_col_in_df = "Nivel obs."
    else:
        plot_df = filtered.copy()

    fan_chart(plot_df, sel_np,
              f"{cfg['name']} · Nivel diario DSS (ft PLD)", "ft PLD",
              f"{res_key}_np_plot", obs_col=obs_col_in_df, obs_label="Nivel obs. BulkExport")

    # --- HP ---
    if hp_cols:
        st.markdown("### ⚡ Hidrogeneración (MW)")
        sel_hp = st.multiselect("Series HP", hp_cols,
                                default=[c for c in hp_cols if any(x in c for x in ["50", "90", "10"])],
                                key=f"{res_key}_hp")
        fan_chart(filtered, sel_hp, f"{cfg['name']} · HP diario promedio", "MW",
                  f"{res_key}_hp_plot")

    # --- AP ---
    if ap_cols:
        st.markdown(f"### 🌊 Aportes — AP total DSS estimado "
                    f"<span class='badge'>{unit_label(flow_unit)}</span>",
                    unsafe_allow_html=True)
        st.caption(f"AP total DSS estimado = AP neto DSS + {clean_evap_cfs(evap_cfs):,.1f} p³/s de evaporación. El visor se ordena de húmedo a seco: P5 arriba y P95 abajo.")
        # Reetiquetar AP por magnitud como probabilidad de excedencia: menor AP → P95; mayor AP → P5.
        ap_df_all, ordered_ap_cols_all, ap_pct_map = make_ordered_ap_columns(
            filtered, ap_cols, flow_unit, add_cfs=clean_evap_cfs(evap_cfs)
        )
        pcts_available = sorted(set(ap_pct_map.values()))
        default_pcts = pcts_available
        sel_pcts = st.multiselect(
            "Percentiles AP",
            options=pcts_available,
            default=default_pcts,
            format_func=lambda x: f"P{x}",
            key=f"{res_key}_ap",
        )
        new_ap_cols = [c for c in ordered_ap_cols_all if exceedance_pct(c) in sel_pcts]

        obs_ap_col = None
        nearest_ap_info = None
        if obs_aportes is not None and isinstance(obs_aportes, pd.DataFrame) and not obs_aportes.empty:
            try:
                obs_ap = clamp_observed_future_dates(obs_aportes, "Fecha_dia")
                obs_ap = obs_ap[obs_ap["Valor"].notna()].copy()
                obs_ap["Fecha_dia"] = pd.to_datetime(obs_ap["Fecha_dia"], errors="coerce").dt.normalize()
                obs_ap = obs_ap[obs_ap["Fecha_dia"].notna()].sort_values("Fecha_dia")
                obs_ap = obs_ap.groupby("Fecha_dia", as_index=False).agg(Valor=("Valor", "last"))
                obs_ap_col = f"Aporte observado [{unit_label(flow_unit)}]"
                obs_ap[obs_ap_col] = convert_flow(obs_ap["Valor"], flow_unit)
                ap_df_all = ap_df_all.merge(obs_ap[["Fecha_dia", obs_ap_col]], on="Fecha_dia", how="left")

                obs_past = obs_ap[obs_ap["Fecha_dia"] <= today_panama()]
                if not obs_past.empty:
                    last_obs_ap = obs_past.iloc[-1]
                    nearest_ap_info = _nearest_ap_percentile(
                        filtered,
                        cfg,
                        pd.to_datetime(last_obs_ap["Fecha_dia"]),
                        float(last_obs_ap["Valor"]),
                        dss_add_cfs=clean_evap_cfs(evap_cfs),
                    )
            except Exception as exc:
                st.warning(f"No se pudo agregar el aporte observado a la gráfica: {exc}")
                obs_ap_col = None

        if nearest_ap_info:
            pct_dss = nearest_ap_info.get('diff_pct_dss', np.nan)
            pct_txt = f"; {pct_dss:+.2f}% vs DSS" if pd.notna(pct_dss) else ""
            st.caption(
                f"🎯 Según el **aporte observado**, el indicador AP DSS ajustado corresponde a "
                f"**{nearest_ap_info['label']}** "
                f"(Obs-DSS: {nearest_ap_info['diff_cfs']:+,.1f} p³/s{pct_txt}; "
                f"fecha DSS: {nearest_ap_info['date']:%d-%m-%Y})."
            )

        _unit_key_ap = str(flow_unit).replace("³", "3").replace("/", "_").replace(" ", "")
        fan_chart(ap_df_all, new_ap_cols,
                  f"{cfg['name']} · AP total DSS estimado y aporte observado ({unit_label(flow_unit)})",
                  unit_label(flow_unit), f"{res_key}_ap_plot_{_unit_key_ap}",
                  obs_col=obs_ap_col, obs_label="Aporte observado")

    # --- V ---
    if v_cols:
        st.markdown(f"### 🚿 Vertidos — promedio diario "
                    f"<span class='badge'>{unit_label(flow_unit)}</span>",
                    unsafe_allow_html=True)
        sel_v = st.multiselect("Probabilidades V", v_cols,
                               default=[c for c in v_cols if any(x in c for x in ["50", "90", "10"])],
                               key=f"{res_key}_v")
        v_df = filtered.copy()
        new_v_cols = []
        for c in sel_v:
            nc = f"{c} [{unit_label(flow_unit)}]"
            v_df[nc] = convert_flow(v_df[c], flow_unit)
            new_v_cols.append(nc)
        fan_chart(v_df, new_v_cols,
                  f"{cfg['name']} · Vertidos diarios promedio ({unit_label(flow_unit)})",
                  unit_label(flow_unit), f"{res_key}_v_plot")

    # --- EG + EP (solo Gatún) ---
    if res_key == "gatun":
        eg_cols = cols_by_prefix(filtered, "EG", token)
        ep_cols = cols_by_prefix(filtered, "EP", token)
        if eg_cols or ep_cols:
            st.markdown(f"### 🚢 Consumo esclusajes EG+EP "
                        f"<span class='badge'>{unit_label(flow_unit)}</span>",
                        unsafe_allow_html=True)
            esc_df = filtered.copy()
            total_cols = []
            # Sum EG+EP por percentil
            pcts_found = sorted(
                set(exceedance_pct(c) for c in eg_cols + ep_cols)
            )
            for pct in pcts_found:
                eg_c = next((c for c in eg_cols if exceedance_pct(c) == pct), None)
                ep_c = next((c for c in ep_cols if exceedance_pct(c) == pct), None)
                parts = [esc_df[c] for c in [eg_c, ep_c] if c and c in esc_df.columns]
                if not parts:
                    continue
                nc = f"Total esc P{pct} [{unit_label(flow_unit)}]"
                raw_sum = pd.concat(parts, axis=1).sum(axis=1, min_count=1)
                esc_df[nc] = convert_flow(raw_sum, flow_unit)
                total_cols.append(nc)
            sel_esc = st.multiselect(
                "Probabilidades esclusajes",
                sorted(set(exceedance_pct(c) for c in eg_cols)),
                default=[p for p in [10, 50, 90] if p in set(exceedance_pct(c) for c in eg_cols)],
                format_func=lambda x: f"P{x}", key="gat_esc",
            )
            plot_esc_cols = [c for c in total_cols if any(f"P{p}" in c for p in sel_esc)]
            fan_chart(esc_df, plot_esc_cols,
                      f"Gatún · Esclusajes EG+EP ({unit_label(flow_unit)})",
                      unit_label(flow_unit), "gat_esc_plot", show_band=False)

    # --- Tabla diaria ---
    st.markdown("### 📋 Tabla diaria DSS")
    with st.expander("Ver tabla", expanded=False):
        table_df = filtered.copy()
        if ap_cols:
            ap_pct_map_table = ordered_percentile_map_by_value(table_df, ap_cols)
            for c in ap_cols:
                if c not in table_df.columns:
                    continue
                pct = ap_pct_map_table.get(c, exceedance_pct(c))
                total_cfs = ap_total_dss_cfs(table_df[c], evap_cfs)
                table_df[f"AP total DSS P{pct} (p³/s)"] = total_cfs
                table_df[f"AP total DSS P{pct} ({unit_label(flow_unit)})"] = convert_flow(total_cfs, flow_unit)
            table_df["Evaporación sumada AP DSS (p³/s)"] = clean_evap_cfs(evap_cfs)
        st.dataframe(table_df, use_container_width=True, height=420)
        csv = table_df.to_csv(index=False).encode("utf-8-sig")
        st.download_button(f"⬇️ Descargar CSV — {cfg['name']}",
                           csv, f"{res_key}_dss_diario.csv", "text/csv", key=f"{res_key}_dl")


# ─────────────────────────────────────────────────────────────────────
# PESTAÑA 3 — MANEJO DE EMBALSES (decisión integrada)
# ─────────────────────────────────────────────────────────────────────
WEEKDAY_ES = {0:"Lunes",1:"Martes",2:"Miércoles",3:"Jueves",4:"Viernes",5:"Sábado",6:"Domingo"}


def _fmt(v, d: int = 3) -> str:
    try:
        return f"{float(v):,.{d}f}" if pd.notna(v) else "—"
    except Exception:
        return "—"


def _semaforo(diff_abs: float, warn: float, alert: float) -> str:
    if pd.isna(diff_abs):
        return "⚪ Sin dato"
    if diff_abs >= alert:
        return "🔴 Revisar"
    if diff_abs >= warn:
        return "🟠 Atención"
    return "🟢 En rango"


def tab_manejo(dss_bytes: bytes, flow_unit: str, pct_ref_gat: int, pct_ref_alh: int,
               obs_gat: Optional[pd.DataFrame], obs_alh: Optional[pd.DataFrame],
               obs_gat_aporte: Optional[pd.DataFrame] = None,
               obs_alh_aporte: Optional[pd.DataFrame] = None,
               evap_gat_cfs: float = 0.0,
               evap_alh_cfs: float = 0.0) -> None:
    st.subheader("🧭 Manejo de embalses — apoyo a la decisión")
    st.caption(
        f"{PROJ_NOTE} · Gatún y Alhajuela/Madden se evalúan por separado. "
        f"Selector inicial: Gatún P{pct_ref_gat} · Alhajuela/Madden P{pct_ref_alh}. "
        "Cuando hay observado, la referencia usada es la probabilidad de excedencia más cercana."
    )

    c1, c2, c3 = st.columns(3)
    gat_threshold = c1.number_input(
        "Umbral operativo Gatún Δ nivel (ft)",
        min_value=0.01, value=0.10, step=0.05, key="mj_thr_gat",
        help="Valor por defecto solicitado para Gatún."
    )
    alh_threshold = c2.number_input(
        "Umbral operativo Alhajuela / Madden Δ nivel (ft)",
        min_value=0.01, value=0.60, step=0.05, key="mj_thr_alh",
        help="Valor por defecto solicitado para Alhajuela/Madden."
    )
    horizon  = c3.selectbox("Horizonte (días)", [7, 15, 30, 60, 90], index=2, key="mj_horiz")
    st.caption(
        "Criterio operativo: 🟠 Atención cuando |Δ nivel| alcanza 70% del umbral del embalse; "
        "🔴 Revisar cuando iguala o supera el umbral completo."
    )

    rows = []
    ts_map = {}
    for res_key, obs_df, obs_ap_df in [
        ("gatun", obs_gat, obs_gat_aporte),
        ("alhajuela", obs_alh, obs_alh_aporte),
    ]:
        cfg = RESERVOIR_CONFIG[res_key]
        pct_ref = int(pct_ref_gat if res_key == "gatun" else pct_ref_alh)
        evap_cfs = clean_evap_cfs(evap_gat_cfs if res_key == "gatun" else evap_alh_cfs)
        try:
            raw    = load_dss_sheet(dss_bytes, cfg["sheet"])
            daily  = to_daily(raw, cfg)
            daily  = apply_may_hydrograph_ap_adjustment(daily, cfg, obs_ap_df)
        except Exception as exc:
            st.warning(f"{cfg['name']}: error DSS — {exc}")
            continue
        if daily.empty:
            continue

        token   = cfg["token"]
        np_cols = cols_by_prefix(daily, "NP", token)
        ap_cols = cols_by_prefix(daily, "AP", token)
        v_cols  = cols_by_prefix(daily, "V",  token)
        hp_cols = cols_by_prefix(daily, "HP", token)

        # Las columnas DSS se seleccionan más abajo, usando como referencia la
        # probabilidad de excedencia más cercana al dato observado disponible.
        # Si no hay observado, se conserva el percentil seleccionado en la barra lateral.

        # Último observado
        obs_val, obs_date = None, None
        if obs_df is not None and not obs_df.empty:
            valid = obs_df[obs_df["Valor"].notna()].sort_values("Fecha_dia")
            if not valid.empty:
                obs_val  = float(valid.iloc[-1]["Valor"])
                obs_date = pd.to_datetime(valid.iloc[-1]["Fecha_dia"])

        closest = closest_np(daily, cfg, obs_date, obs_val) if obs_val is not None else None

        # Fila DSS más cercana al observado (o último)
        today = pd.Timestamp.today().normalize()
        ref_date = obs_date.normalize() if obs_date is not None else today
        base = daily.sort_values("Fecha_dia")
        exact = base[base["Fecha_dia"] == ref_date]
        row_dss = exact.iloc[0] if not exact.empty else base.loc[(base["Fecha_dia"] - ref_date).abs().idxmin()]

        # Referencia por cercanía para nivel: usa la curva NP más cercana al nivel observado.
        # Referencia por cercanía para AP/HP: usa la curva AP más cercana al aporte observado
        # total. Si no existe aporte observado, AP/HP heredan la referencia cercana del nivel.
        level_ref_pct = int(str(closest["label"]).replace("P", "")) if closest else int(pct_ref)
        closest_ap = None
        if _valid_df(obs_ap_df):
            obs_ap_valid = clamp_observed_future_dates(obs_ap_df, "Fecha_dia")
            obs_ap_valid = obs_ap_valid[obs_ap_valid["Valor"].notna()].copy()
            obs_ap_valid["Fecha_dia"] = pd.to_datetime(obs_ap_valid["Fecha_dia"], errors="coerce").dt.normalize()
            obs_ap_valid = obs_ap_valid[obs_ap_valid["Fecha_dia"].notna()].sort_values("Fecha_dia")
            obs_ap_past = obs_ap_valid[obs_ap_valid["Fecha_dia"] <= today_panama()]
            if not obs_ap_past.empty:
                last_ap = obs_ap_past.iloc[-1]
            elif not obs_ap_valid.empty:
                last_ap = obs_ap_valid.iloc[-1]
            else:
                last_ap = None
            if last_ap is not None:
                closest_ap = _nearest_ap_percentile(
                    daily, cfg, pd.to_datetime(last_ap["Fecha_dia"]), float(last_ap["Valor"]),
                    dss_add_cfs=evap_cfs,
                )

        ap_ref_pct_eff = int(closest_ap["percentile"]) if closest_ap else int(level_ref_pct)

        np_col, np_pct = pick_percentile_column(np_cols, level_ref_pct)
        v_col,  v_pct  = pick_percentile_column(v_cols,  level_ref_pct)
        hp_col, hp_pct = pick_percentile_column(hp_cols, ap_ref_pct_eff)

        # AP usa etiqueta corregida como probabilidad de excedencia por magnitud:
        # menor AP → P95; mayor AP → P5.
        ap_pct_map = ordered_percentile_map_by_value(daily, ap_cols, ref_row=row_dss)
        ap_col, ap_pct = pick_percentile_column(ap_cols, ap_ref_pct_eff, pct_map=ap_pct_map)

        np_v  = float(row_dss.get(np_col, np.nan))  if np_col  else np.nan
        ap_v  = convert_flow(ap_total_dss_cfs([row_dss.get(ap_col, np.nan)], evap_cfs), flow_unit).iloc[0] if ap_col else np.nan
        v_v   = convert_flow(pd.Series([float(row_dss.get(v_col, np.nan))]), flow_unit).iloc[0]  if v_col  else np.nan
        hp_v  = float(row_dss.get(hp_col, np.nan))  if hp_col  else np.nan

        diff_np = float(obs_val) - np_v if obs_val is not None and pd.notna(np_v) else np.nan
        threshold_ft = gat_threshold if res_key == "gatun" else alh_threshold
        semaforo = _semaforo(
            abs(diff_np) if pd.notna(diff_np) else np.nan,
            threshold_ft * 0.70,
            threshold_ft,
        )

        # Métricas del horizonte
        today_dt = pd.Timestamp.today().normalize()
        future = base[base["Fecha_dia"] >= today_dt].head(horizon)
        if future.empty:
            future = base.tail(horizon)
        ap_prom = convert_flow(ap_total_dss_cfs(future[ap_col], evap_cfs), flow_unit).mean() if ap_col and ap_col in future else np.nan
        v_prom  = convert_flow(future[v_col], flow_unit).mean()  if v_col and v_col in future else np.nan
        hp_prom = future[hp_col].mean() if hp_col and hp_col in future else np.nan
        np_start = float(future[np_col].iloc[0])  if np_col and np_col in future and not future.empty else np.nan
        np_end   = float(future[np_col].iloc[-1]) if np_col and np_col in future and not future.empty else np.nan

        # Estado ejecutivo: mostrar primero los VALORES usados por la simulación DSS.
        # Los percentiles quedan como trazabilidad en una columna compacta, para no
        # confundirlos con los valores operativos de nivel, aporte, vertido e hidrogeneración.
        rows.append({
            "Embalse":                   cfg["name"],
            "Estado":                    semaforo,
            "Obs. LKH (ft)":             _fmt(obs_val),
            "Fecha obs.":                f"{obs_date:%d-%m-%Y}" if obs_date else "—",
            "Percentil referencia":      f"NP P{level_ref_pct} · AP/HP P{ap_ref_pct_eff}",
            "Nivel DSS usado (ft)":      _fmt(np_v),
            f"AP total DSS usado ({unit_label(flow_unit)})": _fmt(ap_v, 2),
            f"Vertido DSS usado ({unit_label(flow_unit)})": _fmt(v_v, 2),
            "Hidrogeneración DSS usada (MW)": _fmt(hp_v, 2),
            "Series DSS usadas":         f"NP P{np_pct if np_pct is not None else '—'} · AP total P{ap_pct if ap_pct is not None else '—'} (+{evap_cfs:,.1f} p³/s evap.) · V P{v_pct if v_pct is not None else '—'} · HP P{hp_pct if hp_pct is not None else '—'}",
            "Percentil nivel cercano":   closest["label"] if closest else "—",
            "Percentil AP cercano":      closest_ap["label"] if closest_ap else "—",
            "Umbral embalse (ft)":       _fmt(threshold_ft, 3),
            "Δ Obs-NP (ft)":             _fmt(diff_np, 3),
            f"AP total prom. {horizon}d ({unit_label(flow_unit)})": _fmt(ap_prom, 2),
            f"V prom. {horizon}d ({unit_label(flow_unit)})":  _fmt(v_prom, 2),
            f"HP prom. {horizon}d (MW)": _fmt(hp_prom, 2),
            f"NP inicio horiz. (ft)":    _fmt(np_start, 3),
            f"NP fin horiz. (ft)":       _fmt(np_end, 3),
            f"Δ NP horiz. (ft)":         _fmt(np_end - np_start, 3) if pd.notna(np_start) and pd.notna(np_end) else "—",
        })
        ts_map[cfg["name"]] = (daily, obs_df, cfg)
        _show_ap_may_adjustment_note(daily, res_key, "manejo/decisión")

    if rows:
        st.markdown("#### Estado ejecutivo")
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    # Gráficas individuales
    st.markdown("---")
    st.markdown("#### Tendencia por embalse")
    foco = st.selectbox("Embalse a graficar", ["Ambos", "Gatún", "Alhajuela / Madden"], key="mj_foco")

    with st.expander("🗓️ Filtro de período para gráficas", expanded=True):
        all_dates = []
        for name, (d, _, _) in ts_map.items():
            if not d.empty:
                all_dates.extend([d["Fecha_dia"].min(), d["Fecha_dia"].max()])
        if all_dates:
            mn, mx = min(all_dates).date(), max(all_dates).date()
            fc1, fc2 = st.columns(2)
            s = fc1.date_input("Desde", value=mn, min_value=mn, max_value=mx, key="mj_s")
            e = fc2.date_input("Hasta", value=mx, min_value=mn, max_value=mx, key="mj_e")

    for name, (daily, obs_df, cfg) in ts_map.items():
        if foco not in ("Ambos", name):
            continue
        pct_ref = int(pct_ref_gat if cfg.get("token") == "GAT" else pct_ref_alh)
        token = cfg["token"]
        np_cols = cols_by_prefix(daily, "NP", token)
        np_col, _np_pct_plot = pick_percentile_column(np_cols, pct_ref)
        if not np_col:
            continue

        try:
            filt = daily[(daily["Fecha_dia"].dt.date >= s) & (daily["Fecha_dia"].dt.date <= e)].copy()
        except Exception:
            filt = daily.copy()

        # Merge observado
        obs_col_in_df = None
        if obs_df is not None and not obs_df.empty:
            om = obs_df[["Fecha_dia", "Valor"]].rename(columns={"Valor": "Nivel obs."})
            filt = filt.merge(om, on="Fecha_dia", how="left")
            obs_col_in_df = "Nivel obs."

        if not PLOTLY_OK:
            st.line_chart(filt.set_index("Fecha_dia")[[np_col]])
            continue

        fig = go.Figure()
        # Todos los NP en gris fino, el de referencia en azul, observado en rojo
        for col in order_cols_wet_to_dry(np_cols):
            pct = exceedance_pct(col)
            is_ref = (pct == pct_ref)
            fig.add_trace(go.Scatter(
                x=filt["Fecha_dia"], y=filt[col], mode="lines",
                name=f"P{pct}",
                line=dict(
                    color=EXCEEDANCE_COLORS.get(pct, "#aaa"),
                    width=2.5 if is_ref else 0.8,
                    dash="solid" if is_ref else "dot",
                ),
                opacity=1.0 if is_ref else 0.4,
            ))

        if obs_col_in_df and obs_col_in_df in filt.columns:
            obs_v = filt[filt[obs_col_in_df].notna()].copy()
            if not obs_v.empty:
                labels = [""] * len(obs_v)
                labels[-1] = f"{obs_v[obs_col_in_df].iloc[-1]:,.2f} ft"
                fig.add_trace(go.Scatter(
                    x=obs_v["Fecha_dia"], y=obs_v[obs_col_in_df],
                    mode="lines+markers+text", name="🔴 Nivel obs.",
                    text=labels, textposition="top right",
                    line=dict(color="#d90429", width=5),
                    marker=dict(size=9, color="#d90429", line=dict(width=2, color="white")),
                    connectgaps=True,
                ))
                fig.add_trace(go.Scatter(
                    x=[obs_v["Fecha_dia"].iloc[-1]], y=[obs_v[obs_col_in_df].iloc[-1]],
                    mode="markers", name="📍 Último obs.",
                    marker=dict(size=22, color="#d90429", symbol="star",
                                line=dict(width=2, color="white")),
                ))

        _today_line(fig, filt["Fecha_dia"])
        fig.update_layout(**_base_layout(
            f"{name} · Nivel proyectado P{_np_pct_plot if _np_pct_plot is not None else pct_ref} vs observado (ft PLD)",
            "ft PLD", height=580,
        ))
        st.plotly_chart(fig, use_container_width=True, key=f"mj_np_{cfg['token']}")


# ─────────────────────────────────────────────────────────────────────
# PESTAÑA 4 — APORTES OBSERVADOS vs DSS
# ─────────────────────────────────────────────────────────────────────

def _valid_df(df: Optional[pd.DataFrame]) -> bool:
    """Evita errores de evaluación booleana de DataFrame."""
    return df is not None and isinstance(df, pd.DataFrame) and not df.empty


def _df_or_empty(df: Optional[pd.DataFrame]) -> pd.DataFrame:
    """Devuelve un DataFrame válido sin usar `df or DataFrame()`."""
    return df.copy() if _valid_df(df) else pd.DataFrame()



def _nearest_ap_percentile(
    daily: pd.DataFrame,
    cfg: Dict,
    ref_date: pd.Timestamp,
    obs_total_cfs: float,
    dss_add_cfs: float = 0.0,
) -> Optional[Dict]:
    """Determina el indicador AP DSS para el último aporte observado total.

    El DSS trae AP neto. Para comparar contra un aporte total observado:
        AP total DSS estimado = AP neto DSS + evaporación (p³/s)

    Criterio del indicador: las etiquetas se tratan como probabilidad de
    excedencia hidrológica y la referencia se escoge por cercanía absoluta.
    Es decir, si el observado cae entre dos curvas, se toma la curva cuyo
    valor DSS esté más cerca del observado, sin importar si queda por arriba
    o por debajo.
    """
    if daily is None or daily.empty or ref_date is None or pd.isna(obs_total_cfs):
        return None
    ap_cols = cols_by_prefix(daily, "AP", cfg["token"])
    if not ap_cols:
        return None

    base = daily.copy()
    base["Fecha_dia"] = pd.to_datetime(base["Fecha_dia"], errors="coerce").dt.normalize()
    base = base[base["Fecha_dia"].notna()].sort_values("Fecha_dia")
    if base.empty:
        return None

    ref_ts = pd.to_datetime(ref_date).normalize()
    exact = base[base["Fecha_dia"] == ref_ts]
    if not exact.empty:
        row = exact.iloc[0]
        exact_date = True
    else:
        row = base.loc[(base["Fecha_dia"] - ref_ts).abs().idxmin()]
        exact_date = False

    pct_map = ordered_percentile_map_by_value(base, ap_cols, ref_row=row)
    candidates = []
    evap = clean_evap_cfs(dss_add_cfs)
    for col in ap_cols:
        val = pd.to_numeric(pd.Series([row.get(col, np.nan)]), errors="coerce").iloc[0]
        if pd.notna(val):
            dss_total = float(ap_total_dss_cfs([val], evap).iloc[0])
            display_pct = int(pct_map.get(col, exceedance_pct(col)))
            raw_pct = int(exceedance_pct(col))
            candidates.append({
                "column": col,
                "raw_percentile": raw_pct,
                "percentile": display_pct,
                "dss_total_cfs": dss_total,
            })
    if not candidates:
        return None

    # Referencia por vecino más cercano: compara contra TODAS las curvas
    # disponibles y selecciona la menor diferencia absoluta Obs-DSS.
    # Si una curva está por arriba y otra por debajo, gana la que tenga el
    # menor |Obs-DSS|; solo en empate se conserva la mayor excedencia como
    # criterio conservador.
    obs = float(obs_total_cfs)
    for r in candidates:
        r["diff_cfs"] = obs - float(r["dss_total_cfs"])
        r["abs_diff_cfs"] = abs(float(r["diff_cfs"]))

    selected = min(
        candidates,
        key=lambda r: (
            float(r["abs_diff_cfs"]),
            -int(r["percentile"]),
            float(r["dss_total_cfs"]),
        ),
    )

    diff = float(selected["diff_cfs"])
    dss_ref = float(selected["dss_total_cfs"])
    rel_signed_obs = pct_diff_obs_minus_dss_vs_obs(obs, dss_ref)
    rel = abs(rel_signed_obs) if pd.notna(rel_signed_obs) else np.nan
    diff_pct_dss = pct_diff_obs_minus_dss(obs, dss_ref)
    estado = "🟢 Muy cercano" if rel <= 10 else ("🟠 Seguimiento" if rel <= 25 else "🔴 Revisar")
    return {
        "column": selected["column"],
        "raw_percentile": int(selected["raw_percentile"]),
        "percentile": int(selected["percentile"]),
        "label": f"P{int(selected['percentile'])}",
        "dss_total_cfs": float(selected["dss_total_cfs"]),
        "dss_cfs": float(selected["dss_total_cfs"]),  # compatibilidad
        "diff_cfs": float(diff),
        "abs_diff_cfs": float(selected.get("abs_diff_cfs", abs(diff))),
        "diff_pct_dss": float(diff_pct_dss),
        "diff_pct_obs": float(rel_signed_obs),
        "rel_pct": float(rel),
        "estado": estado,
        "date": pd.to_datetime(row["Fecha_dia"]),
        "exact_date": exact_date,
        "evap_cfs": evap,
        "criterio": "vecino_mas_cercano_absoluto",
    }


def tab_aporte_obs_embalse(
    res_key: str,
    dss_bytes: bytes,
    flow_unit: str,
    pct_ref: int,
    obs_df: Optional[pd.DataFrame],
    evap_cfs: float = 0.0,
) -> None:
    """Pestaña individual: aporte total observado vs AP total DSS estimado.

    El DSS se maneja como AP neto. Para comparar con los BulkExport de aportes
    totales, el usuario ingresa el caudal evaporado en p³/s y la app calcula:
        AP total DSS estimado = AP neto DSS + evaporación
    """
    cfg = RESERVOIR_CONFIG[res_key]
    embalse = cfg["name"]
    token = cfg["token"]
    st.subheader(f"🌧️ Aporte observado {token} — {embalse}")
    st.caption(
        "Comparación simple en **aportes totales**: "
        "**AP total DSS estimado = AP neto DSS + evaporación**."
    )
    st.caption(SIMULATION_NOTE)
    st.caption("Las etiquetas AP se corrigen por excedencia: P95 es el aporte más seco y P5 el más húmedo; el visor muestra P5 arriba y P95 abajo.")
    st.caption("Para `Discharge_AT_*_Diario`, el CSV se lee en m³/s, se convierte a p³/s internamente y el sello 00:00 se asigna al día operativo anterior.")

    try:
        raw = load_dss_sheet(dss_bytes, cfg["sheet"])
        daily = to_daily(raw, cfg)
        daily = apply_may_hydrograph_ap_adjustment(daily, cfg, obs_df)
    except Exception as exc:
        st.error(f"Error cargando DSS {embalse}: {exc}")
        return

    if daily is None or daily.empty:
        st.warning("No se pudo construir el diario DSS.")
        return

    ap_cols = cols_by_prefix(daily, "AP", token)
    if not ap_cols:
        st.warning(f"No se encontraron columnas AP para {embalse}.")
        return

    st.markdown("#### AP neto DSS ajustado con caudal evaporado")
    show_obs_total = st.checkbox("Mostrar aporte observado", value=True, key=f"show_obs_total_{res_key}")
    show_dss_neto = st.checkbox("Mostrar AP neto DSS de referencia", value=False, key=f"show_dss_neto_{res_key}")
    st.info(
        f"Caudal evaporado aplicado desde la barra lateral: **{clean_evap_cfs(evap_cfs):,.3f} p³/s**. "
        f"Fórmula: **AP total DSS estimado = AP neto DSS + caudal evaporado**."
    )
    _show_ap_may_adjustment_note(daily, res_key, "aportes observados")

    # Período recomendado: si hay observado, enfocar la gráfica alrededor del observado.
    date_parts = [daily["Fecha_dia"]]
    obs_dates = None
    if _valid_df(obs_df):
        obs_dates = pd.to_datetime(obs_df["Fecha_dia"], errors="coerce").dropna()
        if not obs_dates.empty:
            date_parts.append(obs_dates)
    all_dates = pd.concat(date_parts)
    mn, mx = all_dates.min().date(), all_dates.max().date()

    if obs_dates is not None and not obs_dates.empty:
        obs_max = obs_dates.max().date()
        default_s = max(mn, (pd.Timestamp(obs_max) - pd.Timedelta(days=45)).date())
        default_e = min(mx, max(obs_max, min(mx, (pd.Timestamp(obs_max) + pd.Timedelta(days=30)).date())))
    else:
        default_s, default_e = mn, mx

    c1, c2 = st.columns(2)
    s = c1.date_input("Desde", value=default_s, min_value=mn, max_value=mx, key=f"ao_{res_key}_s")
    e = c2.date_input("Hasta", value=default_e, min_value=mn, max_value=mx, key=f"ao_{res_key}_e")
    if s > e:
        st.warning("Fecha inicial mayor que final. Se usa todo el período disponible.")
        s, e = mn, mx

    try:
        daily_f = daily[(daily["Fecha_dia"].dt.date >= s) & (daily["Fecha_dia"].dt.date <= e)].copy()
    except Exception:
        daily_f = daily.copy()

    obs_f = pd.DataFrame(columns=["Fecha_dia", "Valor", "Fuente"])
    if _valid_df(obs_df):
        try:
            obs_f = obs_df[(obs_df["Fecha_dia"].dt.date >= s) & (obs_df["Fecha_dia"].dt.date <= e)].copy()
        except Exception:
            obs_f = obs_df.copy()

    if not obs_f.empty:
        obs_f["Fecha_dia"] = pd.to_datetime(obs_f["Fecha_dia"], errors="coerce").dt.normalize()
        obs_f["Aporte total observado (p³/s)"] = pd.to_numeric(obs_f["Valor"], errors="coerce")
        obs_f = obs_f.dropna(subset=["Fecha_dia", "Aporte total observado (p³/s)"]).sort_values("Fecha_dia")

    ap_pct_map = ordered_percentile_map_by_value(daily, ap_cols)
    pcts_available = sorted(set(ap_pct_map.values()))

    # La referencia visual/operativa se toma del AP DSS más cercano al último
    # aporte observado del período filtrado. No importa si la curva DSS queda
    # por arriba o por debajo: se usa la menor diferencia absoluta Obs-DSS.
    nearest_ref_pct = int(pct_ref)
    latest_nearest_for_ref = None
    latest_obs_for_ref = None
    if not obs_f.empty and obs_f["Aporte total observado (p³/s)"].notna().any():
        valid_obs_ref = obs_f.dropna(subset=["Aporte total observado (p³/s)"]).sort_values("Fecha_dia")
        if not valid_obs_ref.empty:
            latest_obs_for_ref = valid_obs_ref.iloc[-1]
            latest_nearest_for_ref = _nearest_ap_percentile(
                daily, cfg, pd.to_datetime(latest_obs_for_ref["Fecha_dia"]),
                float(latest_obs_for_ref["Aporte total observado (p³/s)"]),
                dss_add_cfs=clean_evap_cfs(evap_cfs),
            )
            if latest_nearest_for_ref:
                nearest_ref_pct = int(latest_nearest_for_ref["percentile"])

    default_p = []
    # Mostrar por defecto las mismas curvas representativas que ayudan a leer
    # las pestañas semanales: seco (P95/P90), medio (P50) y húmedo (P10/P5),
    # más el percentil cercano al último observado.
    for p in [nearest_ref_pct, int(pct_ref), 95, 90, 50, 10, 5]:
        if p in pcts_available and p not in default_p:
            default_p.append(p)
    if not default_p:
        default_p = pcts_available[:5]

    sel_pcts = st.multiselect(
        f"Percentiles AP total DSS estimado — {embalse}",
        pcts_available,
        default=default_p,
        format_func=lambda x: f"P{x}",
        key=f"ap_obs_{res_key}",
    )
    if latest_nearest_for_ref:
        st.caption(
            f"Referencia AP por cercanía: **{latest_nearest_for_ref['label']}** "
            "(menor diferencia absoluta Obs-DSS; puede estar por arriba o por debajo del observado)."
        )

    # Métricas del último observado total.
    if not obs_f.empty and obs_f["Aporte total observado (p³/s)"].notna().any():
        valid_obs = obs_f.dropna(subset=["Aporte total observado (p³/s)"]).sort_values("Fecha_dia")
        last = valid_obs.iloc[-1]
        last_total_cfs = float(last["Aporte total observado (p³/s)"])
        last_date = pd.to_datetime(last["Fecha_dia"])
        nearest = latest_nearest_for_ref or _nearest_ap_percentile(daily, cfg, last_date, last_total_cfs, dss_add_cfs=clean_evap_cfs(evap_cfs))

        unit_lbl = unit_label(flow_unit)
        last_total_unit = convert_flow(pd.Series([last_total_cfs]), flow_unit).iloc[0]
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric(f"Último aporte total obs. ({unit_lbl})", f"{last_total_unit:,.2f}")
        m2.metric("Evap. sumada al DSS (p³/s)", f"{clean_evap_cfs(evap_cfs):,.1f}")
        if nearest:
            nearest_total_unit = convert_flow(pd.Series([nearest['dss_total_cfs']]), flow_unit).iloc[0]
            nearest_diff_unit = convert_flow(pd.Series([nearest['diff_cfs']]), flow_unit).iloc[0]
            nearest_pct_dss = nearest.get('diff_pct_dss', np.nan)
            pct_delta_txt = f" · {nearest_pct_dss:+.2f}%" if pd.notna(nearest_pct_dss) else ""
            m3.metric(f"AP total DSS indicador ({unit_lbl})", f"{nearest_total_unit:,.2f}")
        else:
            nearest_diff_unit = np.nan
            nearest_pct_dss = np.nan
            pct_delta_txt = ""
            m3.metric(f"AP total DSS indicador ({unit_lbl})", "—")
        m4.metric("Fecha obs.", f"{last_date:%d-%m-%Y}")
        if nearest:
            m5.metric(
                "Indicador DSS AP",
                nearest["label"],
                delta=f"Obs-DSS total: {nearest_diff_unit:+,.2f} {unit_lbl}{pct_delta_txt}",
                delta_color="inverse",
                help=f"{nearest['estado']} · AP total DSS {nearest['label']}: {nearest_total_unit:,.2f} {unit_lbl} "
                     f"({nearest['dss_total_cfs']:,.1f} p³/s) · Diferencia Obs-DSS: {nearest_pct_dss:+.2f}% respecto al DSS · Fecha DSS: {nearest['date']:%d-%m-%Y}",
            )
    else:
        st.warning(
            f"No hay aportes observados para {embalse}. Cargue un BulkExport de aporte "
            "o use la entrada manual opcional."
        )

    if not PLOTLY_OK:
        st.info("Plotly no disponible.")
        return

    fig = go.Figure()

    # AP total DSS estimado = AP neto DSS + evaporación.
    # Orden del visor/leyenda: húmedo arriba (P5) → seco abajo (P95).
    ap_cols_plot = sorted(ap_cols, key=lambda c: ap_pct_map.get(c, exceedance_pct(c)))
    for col in ap_cols_plot:
        pct = ap_pct_map.get(col, exceedance_pct(col))
        if pct not in sel_pcts:
            continue
        dss_total_cfs = ap_total_dss_cfs(daily_f[col], evap_cfs)
        vals_total = convert_flow(dss_total_cfs, flow_unit)
        fig.add_trace(go.Scatter(
            x=daily_f["Fecha_dia"],
            y=vals_total,
            mode="lines",
            name=f"AP total DSS est. P{pct}",
            line=dict(
                width=3.2 if pct == nearest_ref_pct else (2.5 if pct in (pct_ref, 50) else 1.4),
                dash="solid" if pct in (nearest_ref_pct, pct_ref, 50) else "dot",
                color=EXCEEDANCE_COLORS.get(pct, "#aaa"),
            ),
            hovertemplate=f"Fecha: %{{x|%d-%m-%Y}}<br>AP total DSS: %{{y:,.2f}} {unit_label(flow_unit)}<extra></extra>",
        ))

    if show_dss_neto:
        ref_col = next((c for c, p in ap_pct_map.items() if p == nearest_ref_pct), None)
        if ref_col and ref_col in daily_f.columns:
            vals_net = convert_flow(daily_f[ref_col], flow_unit)
            fig.add_trace(go.Scatter(
                x=daily_f["Fecha_dia"],
                y=vals_net,
                mode="lines",
                name=f"AP neto DSS ref. P{nearest_ref_pct}",
                line=dict(color="#64748b", width=1.8, dash="dash"),
                opacity=0.7,
            ))

    if show_obs_total and not obs_f.empty:
        obs_v = obs_f.dropna(subset=["Aporte total observado (p³/s)"]).sort_values("Fecha_dia")
        if not obs_v.empty:
            y_total = convert_flow(obs_v["Aporte total observado (p³/s)"], flow_unit)
            labels_total = [""] * len(obs_v)
            labels_total[-1] = f"Obs {y_total.iloc[-1]:,.1f} {unit_label(flow_unit)}"
            fig.add_trace(go.Scatter(
                x=obs_v["Fecha_dia"],
                y=y_total,
                mode="lines+markers+text",
                name="🔴 Aporte total observado",
                text=labels_total,
                textposition="top right",
                line=dict(color="#d90429", width=5),
                marker=dict(size=9, color="#d90429", line=dict(width=2, color="white")),
                connectgaps=True,
                hovertemplate=f"Fecha: %{{x|%d-%m-%Y}}<br>Aporte observado: %{{y:,.2f}} {unit_label(flow_unit)}<extra></extra>",
            ))
            fig.add_trace(go.Scatter(
                x=[obs_v["Fecha_dia"].iloc[-1]],
                y=[y_total.iloc[-1]],
                mode="markers",
                name="📍 Último observado",
                marker=dict(size=22, color="#d90429", symbol="star", line=dict(width=2, color="white")),
            ))

    # Usar el rango completo de las trazas para que los picos DSS no queden
    # recortados. Esto mantiene consistencia visual con las gráficas semanales.
    _apply_full_yaxis_from_traces(fig, force_zero=True, pad_fraction=0.08)

    _today_line(fig, daily_f["Fecha_dia"])
    fig.update_layout(**_base_layout(
        f"{embalse} · Aporte total observado vs AP total DSS estimado ({unit_label(flow_unit)})",
        unit_label(flow_unit),
        height=650,
    ))
    _unit_key = str(flow_unit).replace("³", "3").replace("/", "_").replace(" ", "")
    st.plotly_chart(fig, use_container_width=True, key=f"ao_{res_key}_plot_{_unit_key}")

    with st.expander("📋 Tabla de aporte observado y AP total DSS estimado", expanded=False):
        rows = []
        if not obs_f.empty:
            table = obs_f.copy()
            table["Aporte total observado ({})".format(unit_label(flow_unit))] = convert_flow(
                table["Aporte total observado (p³/s)"], flow_unit
            )
            table["Evaporación sumada al DSS (p³/s)"] = clean_evap_cfs(evap_cfs)
            table = table.rename(columns={"Fecha_dia": "Fecha"})
            for _, r in table.iterrows():
                rows.append({
                    "Fecha": r.get("Fecha"),
                    "Fuente": r.get("Fuente", "—"),
                    "Aporte total observado (p³/s)": r.get("Aporte total observado (p³/s)", np.nan),
                    f"Aporte total observado ({unit_label(flow_unit)})": r.get(
                        "Aporte total observado ({})".format(unit_label(flow_unit)), np.nan
                    ),
                    "Evaporación sumada al DSS (p³/s)": clean_evap_cfs(evap_cfs),
                })
        if rows:
            out = pd.DataFrame(rows)
            for c in out.columns:
                if c not in ("Fecha", "Fuente"):
                    out[c] = pd.to_numeric(out[c], errors="coerce").round(3)
            st.dataframe(out, use_container_width=True, hide_index=True, height=420)
            csv = out.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                f"⬇️ Descargar CSV aporte {token}",
                csv,
                f"aporte_{token}_total_observado_vs_dss.csv",
                "text/csv",
                key=f"ao_{res_key}_dl",
            )
        else:
            st.info("Sin datos observados para mostrar en tabla.")


def tab_aportes_obs(dss_bytes: bytes, flow_unit: str, pct_ref: int,
                    obs_gat: Optional[pd.DataFrame], obs_alh: Optional[pd.DataFrame],
                    evap_gat_cfs: float = 0.0, evap_alh_cfs: float = 0.0) -> None:
    """Vista contenedora conservada por compatibilidad: una subpestaña por embalse."""
    st.subheader("🌧️ Aportes observados separados por embalse")
    sub = st.tabs(["GAT · Gatún", "ALHA · Alhajuela/Madden"])
    with sub[0]:
        tab_aporte_obs_embalse("gatun", dss_bytes, flow_unit, pct_ref, obs_gat, evap_gat_cfs)
    with sub[1]:
        tab_aporte_obs_embalse("alhajuela", dss_bytes, flow_unit, pct_ref, obs_alh, evap_alh_cfs)


# ─────────────────────────────────────────────────────────────────────
# PESTAÑA 5 — NIVELES OBSERVADOS
# ─────────────────────────────────────────────────────────────────────
def tab_niveles_obs(obs_gat: Optional[pd.DataFrame], obs_alh: Optional[pd.DataFrame]) -> None:
    st.subheader("🔴 Niveles observados — Gatún y Alhajuela/Madden")
    st.caption("Cargados desde BulkExport-GAT y BulkExport-MAD.")

    if (obs_gat is None or obs_gat.empty) and (obs_alh is None or obs_alh.empty):
        st.warning("No hay datos de nivel observado. Cargue BulkExport-GAT.csv y BulkExport-MAD.csv.")
        return

    # Métricas
    m1, m2, m3, m4 = st.columns(4)
    for metric_cols, obs_df, label, color in [
        ((m1, m2), obs_gat, "Gatún", "#d90429"),
        ((m3, m4), obs_alh, "Alhajuela/Madden", "#f97316"),
    ]:
        if obs_df is not None and not obs_df.empty:
            valid = obs_df[obs_df["Valor"].notna()].sort_values("Fecha_dia")
            if not valid.empty:
                last_v    = float(valid.iloc[-1]["Valor"])
                last_d    = pd.to_datetime(valid.iloc[-1]["Fecha_dia"])
                prev_v    = float(valid.iloc[-2]["Valor"]) if len(valid) >= 2 else np.nan
                cambio    = last_v - prev_v if pd.notna(prev_v) else None
                metric_cols[0].metric(f"🔴 {label} (ft)", f"{last_v:,.3f}",
                                      delta=f"{cambio:+.3f} ft/día" if cambio is not None else None)
                metric_cols[1].metric("Fecha", f"{last_d:%d-%m-%Y}")
        else:
            metric_cols[0].metric(label, "—")
            metric_cols[1].metric("Fecha", "—")

    # Filtro
    all_dates = []
    for obs_df in [obs_gat, obs_alh]:
        if obs_df is not None and not obs_df.empty:
            all_dates.extend([obs_df["Fecha_dia"].min(), obs_df["Fecha_dia"].max()])
    if not all_dates:
        return
    mn, mx = min(all_dates).date(), max(all_dates).date()
    default_s = max(mn, (pd.Timestamp(mx) - pd.Timedelta(days=60)).date())
    c1, c2 = st.columns(2)
    s = c1.date_input("Desde", value=default_s, min_value=mn, max_value=mx, key="nobs_s")
    e = c2.date_input("Hasta", value=mx, min_value=mn, max_value=mx, key="nobs_e")

    def _filt_obs(df):
        if df is None or df.empty:
            return None
        return df[(df["Fecha_dia"].dt.date >= s) & (df["Fecha_dia"].dt.date <= e)].copy()

    obs_gat_f = _filt_obs(obs_gat)
    obs_alh_f = _filt_obs(obs_alh)

    if not PLOTLY_OK:
        st.info("Instala plotly para ver las gráficas.")
        return

    # Vista conjunta (dos ejes)
    st.markdown("### Vista conjunta — ejes separados")
    fig = go.Figure()
    for obs_df, label, color, yaxis in [
        (obs_gat_f, "Gatún", "#d90429", "y"),
        (obs_alh_f, "Alhajuela/Madden", "#f97316", "y2"),
    ]:
        if obs_df is not None and not obs_df.empty:
            valid = obs_df[obs_df["Valor"].notna()].sort_values("Fecha_dia")
            labels = [""] * len(valid)
            if len(labels):
                labels[-1] = f"{label} {valid['Valor'].iloc[-1]:,.2f} ft"
            fig.add_trace(go.Scatter(
                x=valid["Fecha_dia"], y=valid["Valor"],
                mode="lines+markers+text", name=f"🔴 {label}",
                text=labels, textposition="top left",
                line=dict(color=color, width=6),
                marker=dict(size=9, color=color, line=dict(width=2, color="white")),
                connectgaps=True, yaxis=yaxis,
            ))
            fig.add_trace(go.Scatter(
                x=[valid["Fecha_dia"].iloc[-1]], y=[valid["Valor"].iloc[-1]],
                mode="markers", name=f"📍 Último {label}",
                marker=dict(size=22, color=color, symbol="star",
                            line=dict(width=2, color="white")),
                yaxis=yaxis,
            ))

    all_obs = pd.concat([_df_or_empty(obs_gat_f), _df_or_empty(obs_alh_f)], ignore_index=True)
    _today_line(fig, all_obs.get("Fecha_dia", pd.Series(dtype="datetime64[ns]")))
    fig.update_layout(
        title="Niveles observados — Gatún y Alhajuela/Madden",
        height=650, hovermode="x unified",
        yaxis=dict(title="Gatún (ft PLD)", gridcolor="rgba(0,0,0,0.07)"),
        yaxis2=dict(title="Alhajuela/Madden (ft PLD)", overlaying="y", side="right", showgrid=False),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor="rgba(250,252,255,1)", paper_bgcolor="rgba(250,252,255,0)",
        margin=dict(l=10, r=10, t=80, b=10),
    )
    st.plotly_chart(fig, use_container_width=True, key="nobs_two")

    # Tabla
    with st.expander("📋 Tabla de niveles observados", expanded=False):
        frames = []
        for obs_df, label in [(obs_gat, "Gatún"), (obs_alh, "Alhajuela/Madden")]:
            if obs_df is not None and not obs_df.empty:
                tmp = obs_df[obs_df["Valor"].notna()].copy()
                tmp["Embalse"] = label
                tmp["Cambio (ft/día)"] = tmp["Valor"].diff()
                tmp = tmp.rename(columns={"Valor": "Nivel obs. (ft)", "Fecha_dia": "Fecha"})
                frames.append(tmp[["Fecha", "Embalse", "Nivel obs. (ft)", "Cambio (ft/día)"]].tail(30))
        if frames:
            t = pd.concat(frames, ignore_index=True).sort_values("Fecha")
            for col in ["Nivel obs. (ft)", "Cambio (ft/día)"]:
                t[col] = pd.to_numeric(t[col], errors="coerce").round(3)
            st.dataframe(t, use_container_width=True, hide_index=True, height=420)
            csv = t.to_csv(index=False).encode("utf-8-sig")
            st.download_button("⬇️ CSV niveles observados", csv,
                               "niveles_observados.csv", "text/csv", key="nobs_dl")


# ─────────────────────────────────────────────────────────────────────
# PESTAÑA 6 — COMPARATIVO rápido
# ─────────────────────────────────────────────────────────────────────
def tab_comparativo(dss_bytes: bytes, flow_unit: str, pct_ref_gat: int, pct_ref_alh: int,
                    obs_gat: Optional[pd.DataFrame], obs_alh: Optional[pd.DataFrame],
                    obs_gat_aporte: Optional[pd.DataFrame] = None,
                    obs_alh_aporte: Optional[pd.DataFrame] = None,
                    evap_gat_cfs: float = 0.0,
                    evap_alh_cfs: float = 0.0) -> None:
    st.subheader("🔀 Comparativo Gatún vs Alhajuela")
    st.caption("Comparación visual por embalse. En AP se incorporan los aportes observados disponibles.")

    var_choice = st.selectbox("Variable", ["NP (Nivel)", "AP (Aportes)", "V (Vertidos)", "HP (Hidrogeneración)"],
                              key="cmp_var")
    prefix = var_choice.split()[0]
    default_cmp_pcts = []
    for p in [int(pct_ref_gat), int(pct_ref_alh), 90, 10]:
        if p in PERCENTILE_ORDER and p not in default_cmp_pcts:
            default_cmp_pcts.append(p)
    pcts   = st.multiselect("Probabilidades", PERCENTILE_ORDER, default=default_cmp_pcts or [50, 90, 10],
                             format_func=lambda x: f"P{x}", key="cmp_pcts")

    try:
        gat_raw = load_dss_sheet(dss_bytes, RESERVOIR_CONFIG["gatun"]["sheet"])
        alh_raw = load_dss_sheet(dss_bytes, RESERVOIR_CONFIG["alhajuela"]["sheet"])
    except Exception as exc:
        st.error(f"Error: {exc}")
        return

    gat_d = to_daily(gat_raw, RESERVOIR_CONFIG["gatun"])
    alh_d = to_daily(alh_raw, RESERVOIR_CONFIG["alhajuela"])
    gat_d = apply_may_hydrograph_ap_adjustment(gat_d, RESERVOIR_CONFIG["gatun"], obs_gat_aporte)
    alh_d = apply_may_hydrograph_ap_adjustment(alh_d, RESERVOIR_CONFIG["alhajuela"], obs_alh_aporte)
    if gat_d.empty or alh_d.empty:
        st.warning("No hay datos DSS suficientes para el comparativo.")
        return

    with st.expander("🗓️ Filtro de período", expanded=True):
        date_series = [gat_d["Fecha_dia"], alh_d["Fecha_dia"]]
        for odf in [obs_gat, obs_alh, obs_gat_aporte, obs_alh_aporte]:
            if _valid_df(odf) and "Fecha_dia" in odf.columns:
                date_series.append(odf["Fecha_dia"])
        all_dates = pd.concat(date_series)
        all_min = all_dates.min()
        all_max = all_dates.max()
        c1, c2 = st.columns(2)
        s = c1.date_input("Desde", value=all_min.date(), min_value=all_min.date(), max_value=all_max.date(), key="cmp_s")
        e = c2.date_input("Hasta", value=all_max.date(), min_value=all_min.date(), max_value=all_max.date(), key="cmp_e")
    gat_f = gat_d[(gat_d["Fecha_dia"].dt.date >= s) & (gat_d["Fecha_dia"].dt.date <= e)].copy()
    alh_f = alh_d[(alh_d["Fecha_dia"].dt.date >= s) & (alh_d["Fecha_dia"].dt.date <= e)].copy()

    raw_gat_cols = cols_by_prefix(gat_f, prefix, "GAT")
    raw_alh_cols = cols_by_prefix(alh_f, prefix, "ALH")

    is_flow = prefix in ("AP", "V")
    y_lbl   = unit_label(flow_unit) if is_flow else ("ft PLD" if prefix == "NP" else "MW")

    if prefix == "AP":
        gat_plot, gat_cols_all, _ = make_ordered_ap_columns(
            gat_f, raw_gat_cols, flow_unit, add_cfs=clean_evap_cfs(evap_gat_cfs)
        )
        alh_plot, alh_cols_all, _ = make_ordered_ap_columns(
            alh_f, raw_alh_cols, flow_unit, add_cfs=clean_evap_cfs(evap_alh_cfs)
        )
        gat_cols = [c for c in gat_cols_all if exceedance_pct(c) in pcts]
        alh_cols = [c for c in alh_cols_all if exceedance_pct(c) in pcts]
        st.caption("En AP se corrige la etiqueta como probabilidad de excedencia: menor AP → P95; mayor AP → P5. En el visor Plotly se ordena P5 arriba y P95 abajo.")
        _show_ap_may_adjustment_note(gat_plot, "gatun", "comparativo Gatún")
        _show_ap_may_adjustment_note(alh_plot, "alhajuela", "comparativo Alhajuela")
    else:
        st.caption("El visor Plotly se ordena de húmedo a seco: P5 arriba y P95 abajo.")
        gat_cols = [c for c in raw_gat_cols if exceedance_pct(c) in pcts]
        alh_cols = [c for c in raw_alh_cols if exceedance_pct(c) in pcts]
        gat_plot = gat_f.copy()
        alh_plot = alh_f.copy()

    gat_obs_col = alh_obs_col = None

    if prefix == "NP":
        if _valid_df(obs_gat):
            om = obs_gat[["Fecha_dia", "Valor"]].rename(columns={"Valor": "Nivel obs. Gatún"})
            gat_plot = gat_plot.merge(om, on="Fecha_dia", how="left")
            gat_obs_col = "Nivel obs. Gatún"
        if _valid_df(obs_alh):
            om = obs_alh[["Fecha_dia", "Valor"]].rename(columns={"Valor": "Nivel obs. Alhajuela"})
            alh_plot = alh_plot.merge(om, on="Fecha_dia", how="left")
            alh_obs_col = "Nivel obs. Alhajuela"

    # En AP, mostrar y graficar aportes observados disponibles.
    if prefix == "AP":
        obs_rows = []
        for label, odf in [("Gatún", obs_gat_aporte), ("Alhajuela / Madden", obs_alh_aporte)]:
            if _valid_df(odf):
                tmp = clamp_observed_future_dates(odf, "Fecha_dia")
                tmp = tmp[(tmp["Fecha_dia"].dt.date >= s) & (tmp["Fecha_dia"].dt.date <= e)].copy()
                tmp = tmp.dropna(subset=["Valor"]).sort_values("Fecha_dia")
                if not tmp.empty:
                    last = tmp.iloc[-1]
                    obs_conv = convert_flow(pd.Series([float(last["Valor"])]), flow_unit).iloc[0]
                    obs_rows.append({
                        "Embalse": label,
                        "Fecha último aporte observado": pd.to_datetime(last["Fecha_dia"]).strftime("%d-%m-%Y"),
                        f"Aporte observado ({unit_label(flow_unit)})": round(float(obs_conv), 3),
                        "Aporte observado (p³/s)": round(float(last["Valor"]), 3),
                        "Fuente": last.get("Fuente", "—"),
                    })
        if obs_rows:
            st.markdown("#### Valores de aportes observados")
            st.dataframe(pd.DataFrame(obs_rows), use_container_width=True, hide_index=True)

        if _valid_df(obs_gat_aporte):
            om = clamp_observed_future_dates(obs_gat_aporte, "Fecha_dia")[["Fecha_dia", "Valor"]].copy()
            om["Aporte obs. Gatún"] = convert_flow(om["Valor"], flow_unit)
            gat_plot = gat_plot.merge(om[["Fecha_dia", "Aporte obs. Gatún"]], on="Fecha_dia", how="left")
            gat_obs_col = "Aporte obs. Gatún"
        if _valid_df(obs_alh_aporte):
            om = clamp_observed_future_dates(obs_alh_aporte, "Fecha_dia")[["Fecha_dia", "Valor"]].copy()
            om["Aporte obs. Alhajuela"] = convert_flow(om["Valor"], flow_unit)
            alh_plot = alh_plot.merge(om[["Fecha_dia", "Aporte obs. Alhajuela"]], on="Fecha_dia", how="left")
            alh_obs_col = "Aporte obs. Alhajuela"

    # Convertir flujos DSS en la copia de gráfica.
    # En AP ya se convirtió y se sumó evaporación mediante make_ordered_ap_columns().
    if is_flow and prefix != "AP":
        for df, cols in [(gat_plot, gat_cols), (alh_plot, alh_cols)]:
            for c in cols:
                base_series = pd.to_numeric(df[c], errors="coerce")
                df[c] = convert_flow(base_series, flow_unit)

    c_gat, c_alh = st.columns(2)
    with c_gat:
        st.caption("**Gatún**")
        fan_chart(gat_plot, gat_cols, f"Gatún · {var_choice}", y_lbl, "cmp_gat",
                  obs_col=gat_obs_col,
                  obs_label="Aporte obs. Gatún" if prefix == "AP" else "Nivel obs. Gatún")
    with c_alh:
        st.caption("**Alhajuela / Madden**")
        fan_chart(alh_plot, alh_cols, f"Alhajuela · {var_choice}", y_lbl, "cmp_alh",
                  obs_col=alh_obs_col,
                  obs_label="Aporte obs. ALHA" if prefix == "AP" else "Nivel obs. Alhajuela")


# ─────────────────────────────────────────────────────────────────────
# PESTAÑA — COMPARACIÓN 1997-1998, OBSERVADO, PERCENTILES Y CHCP
# ─────────────────────────────────────────────────────────────────────
def _historical_9798_to_year(hist: pd.DataFrame, target_year: int, value_col: str) -> pd.DataFrame:
    """Proyecta mes/día histórico sobre el año DSS para alinear las series."""
    if hist is None or hist.empty or value_col not in hist.columns:
        return pd.DataFrame(columns=["Fecha_dia", "Anio_historico", "Valor"])
    out = hist[["Anio_historico", "Mes", "Dia", value_col]].copy()
    out["Fecha_dia"] = pd.to_datetime(
        {"year": int(target_year), "month": out["Mes"], "day": out["Dia"]},
        errors="coerce",
    )
    out["Valor"] = pd.to_numeric(out[value_col], errors="coerce")
    return out.dropna(subset=["Fecha_dia", "Valor"])[
        ["Fecha_dia", "Anio_historico", "Valor"]
    ].sort_values(["Anio_historico", "Fecha_dia"])


def _selected_ap_dss_9798(
    daily: pd.DataFrame,
    cfg: Dict,
    pct_ref: int,
    evap_cfs: float,
) -> Tuple[pd.DataFrame, Optional[int]]:
    """Extrae el AP DSS total del percentil solicitado en p³/s."""
    if daily is None or daily.empty:
        return pd.DataFrame(columns=["Fecha_dia", "DSS_cfs"]), None
    ap_cols = cols_by_prefix(daily, "AP", cfg.get("token", ""))
    pct_map = ordered_percentile_map_by_value(daily, ap_cols)
    ap_col, pct_used = pick_percentile_column(ap_cols, int(pct_ref), pct_map=pct_map)
    if ap_col is None:
        return pd.DataFrame(columns=["Fecha_dia", "DSS_cfs"]), None
    out = daily[["Fecha_dia"]].copy()
    out["DSS_cfs"] = ap_total_dss_cfs(daily[ap_col], clean_evap_cfs(evap_cfs))
    out["Fecha_dia"] = pd.to_datetime(out["Fecha_dia"], errors="coerce").dt.normalize()
    return out.dropna(subset=["Fecha_dia"]).sort_values("Fecha_dia"), pct_used


def _observed_ap_9798(obs: Optional[pd.DataFrame]) -> pd.DataFrame:
    """Normaliza aporte observado a una fila diaria en p³/s."""
    if not _valid_df(obs) or "Fecha_dia" not in obs.columns or "Valor" not in obs.columns:
        return pd.DataFrame(columns=["Fecha_dia", "Obs_cfs"])
    out = clamp_observed_future_dates(obs, "Fecha_dia")[["Fecha_dia", "Valor"]].copy()
    out["Fecha_dia"] = pd.to_datetime(out["Fecha_dia"], errors="coerce").dt.normalize()
    out["Obs_cfs"] = pd.to_numeric(out["Valor"], errors="coerce")
    out = out.dropna(subset=["Fecha_dia", "Obs_cfs"]).sort_values("Fecha_dia")
    return out.groupby("Fecha_dia", as_index=False).agg(Obs_cfs=("Obs_cfs", "last"))


def _build_9798_entity_daily(
    hist: pd.DataFrame,
    target_year: int,
    hist_value_col: str,
    dss_daily: pd.DataFrame,
    obs_daily: pd.DataFrame,
) -> pd.DataFrame:
    """Combina 1997, 1998, observado y DSS en una tabla diaria común."""
    h = _historical_9798_to_year(hist, target_year, hist_value_col)
    if h.empty:
        wide = pd.DataFrame(columns=["Fecha_dia", "Hist_1997_cfs", "Hist_1998_cfs"])
    else:
        wide = h.pivot_table(
            index="Fecha_dia", columns="Anio_historico", values="Valor", aggfunc="mean"
        ).reset_index()
        wide.columns.name = None
        wide = wide.rename(columns={1997: "Hist_1997_cfs", 1998: "Hist_1998_cfs"})
        for c in ["Hist_1997_cfs", "Hist_1998_cfs"]:
            if c not in wide.columns:
                wide[c] = np.nan

    dates = []
    for df in (wide, dss_daily, obs_daily):
        if isinstance(df, pd.DataFrame) and not df.empty and "Fecha_dia" in df.columns:
            dates.append(pd.to_datetime(df["Fecha_dia"], errors="coerce").dropna())
    if not dates:
        return pd.DataFrame()
    all_dates = pd.concat(dates, ignore_index=True).dropna().drop_duplicates().sort_values()
    base = pd.DataFrame({"Fecha_dia": all_dates})
    for df in (wide, dss_daily, obs_daily):
        if isinstance(df, pd.DataFrame) and not df.empty:
            base = base.merge(df, on="Fecha_dia", how="left")
    for c in [
        "Hist_1997_cfs", "Hist_1998_cfs", "DSS_cfs", "DSS_base_cfs",
        "DSS_reduction_cfs", "DSS_reduction_pct", "Obs_cfs",
    ]:
        if c not in base.columns:
            base[c] = np.nan
        base[c] = pd.to_numeric(base[c], errors="coerce")
    base["Hist_prom_9798_cfs"] = base[["Hist_1997_cfs", "Hist_1998_cfs"]].mean(axis=1)
    return base.sort_values("Fecha_dia").reset_index(drop=True)


def _sum_chcp_9798(gat: pd.DataFrame, alh: pd.DataFrame) -> pd.DataFrame:
    """Suma GAT + ALHA por fecha; exige ambos componentes para reportar CHCP."""
    cols = [
        "Hist_1997_cfs", "Hist_1998_cfs", "Hist_prom_9798_cfs",
        "Obs_cfs", "DSS_cfs", "DSS_base_cfs", "DSS_reduction_cfs",
    ]
    g = gat[["Fecha_dia"] + cols].copy() if isinstance(gat, pd.DataFrame) and not gat.empty else pd.DataFrame()
    a = alh[["Fecha_dia"] + cols].copy() if isinstance(alh, pd.DataFrame) and not alh.empty else pd.DataFrame()
    if g.empty and a.empty:
        return pd.DataFrame()
    merged = g.merge(a, on="Fecha_dia", how="outer", suffixes=("_GAT", "_ALHA"))
    out = merged[["Fecha_dia"]].copy()
    for c in cols:
        left = pd.to_numeric(merged.get(f"{c}_GAT"), errors="coerce")
        right = pd.to_numeric(merged.get(f"{c}_ALHA"), errors="coerce")
        out[c] = pd.concat([left, right], axis=1).sum(axis=1, min_count=2)
    return out.sort_values("Fecha_dia").reset_index(drop=True)


def _pct_series_9798(actual: pd.Series, reference: pd.Series) -> pd.Series:
    a = pd.to_numeric(actual, errors="coerce")
    r = pd.to_numeric(reference, errors="coerce")
    return ((a - r) / r.abs().replace(0, np.nan) * 100.0).replace([np.inf, -np.inf], np.nan)


def _aggregate_9798_comparison(
    daily: pd.DataFrame,
    mode: str,
    week_start: int = 5,
) -> pd.DataFrame:
    """Promedia por período usando días comunes con observado cuando existen.

    Esta regla evita comparar una semana/mes observado incompleto contra siete
    días o un mes completo de las referencias históricas/DSS.
    """
    if daily is None or daily.empty:
        return pd.DataFrame()
    base = daily.copy()
    base["Fecha_dia"] = pd.to_datetime(base["Fecha_dia"], errors="coerce").dt.normalize()
    base = base.dropna(subset=["Fecha_dia"]).sort_values("Fecha_dia")
    if mode == "Semanal":
        week_info = base["Fecha_dia"].apply(
            lambda value: operational_week_info(value, week_start=int(week_start))
        )
        base["Periodo"] = [item[1] for item in week_info]
    elif mode == "Mensual":
        base["Periodo"] = base["Fecha_dia"].dt.to_period("M").dt.to_timestamp()
    else:
        base["Periodo"] = base["Fecha_dia"]

    value_cols = ["Hist_1997_cfs", "Hist_1998_cfs", "Hist_prom_9798_cfs", "Obs_cfs", "DSS_cfs"]
    rows: List[Dict[str, object]] = []
    for period, group in base.groupby("Periodo", sort=True):
        obs_mask = pd.to_numeric(group["Obs_cfs"], errors="coerce").notna()
        comparable = group.loc[obs_mask] if obs_mask.any() else group
        row: Dict[str, object] = {
            "Periodo": pd.to_datetime(period),
            "Dias_periodo": int(len(group)),
            "Dias_observados": int(obs_mask.sum()),
        }
        for c in value_cols:
            values = pd.to_numeric(
                comparable[c] if c != "Obs_cfs" else group.loc[obs_mask, c],
                errors="coerce",
            )
            row[c] = float(values.mean()) if values.notna().any() else np.nan
        rows.append(row)
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["Dif_Obs_vs_1997_pct"] = _pct_series_9798(out["Obs_cfs"], out["Hist_1997_cfs"])
    out["Dif_Obs_vs_1998_pct"] = _pct_series_9798(out["Obs_cfs"], out["Hist_1998_cfs"])
    out["Dif_Obs_vs_Prom9798_pct"] = _pct_series_9798(out["Obs_cfs"], out["Hist_prom_9798_cfs"])
    out["Dif_Obs_vs_DSS_pct"] = _pct_series_9798(out["Obs_cfs"], out["DSS_cfs"])
    out["Dif_DSS_vs_1997_pct"] = _pct_series_9798(out["DSS_cfs"], out["Hist_1997_cfs"])
    out["Dif_DSS_vs_1998_pct"] = _pct_series_9798(out["DSS_cfs"], out["Hist_1998_cfs"])
    return out


def _display_9798_table(agg: pd.DataFrame, flow_unit: str, dss_label: str) -> pd.DataFrame:
    """Convierte unidades y asigna encabezados legibles para tabla/exportación."""
    if agg is None or agg.empty:
        return pd.DataFrame()
    out = agg.copy()
    unit = unit_label(flow_unit)
    rename = {
        "Periodo": "Fecha / período",
        "Dias_periodo": "Días del período",
        "Dias_observados": "Días observados comparados",
        "Hist_1997_cfs": f"1997 ({unit})",
        "Hist_1998_cfs": f"1998 ({unit})",
        "Hist_prom_9798_cfs": f"Promedio 1997-1998 ({unit})",
        "Obs_cfs": f"Observado ({unit})",
        "DSS_cfs": f"DSS {dss_label} ({unit})",
        "Dif_Obs_vs_1997_pct": "Dif. Obs vs 1997 (%)",
        "Dif_Obs_vs_1998_pct": "Dif. Obs vs 1998 (%)",
        "Dif_Obs_vs_Prom9798_pct": "Dif. Obs vs prom. 97-98 (%)",
        "Dif_Obs_vs_DSS_pct": f"Dif. Obs vs DSS {dss_label} (%)",
        "Dif_DSS_vs_1997_pct": f"Dif. DSS {dss_label} vs 1997 (%)",
        "Dif_DSS_vs_1998_pct": f"Dif. DSS {dss_label} vs 1998 (%)",
    }
    for c in ["Hist_1997_cfs", "Hist_1998_cfs", "Hist_prom_9798_cfs", "Obs_cfs", "DSS_cfs"]:
        out[c] = convert_flow(out[c], flow_unit)
    out = out.rename(columns=rename)
    for c in out.columns:
        if c != "Fecha / período":
            out[c] = pd.to_numeric(out[c], errors="coerce").round(3)
    return out


def _summary_9798_entity(
    entity: str,
    daily: pd.DataFrame,
    flow_unit: str,
    dss_label: str,
) -> Dict[str, object]:
    """Resumen promedio sobre los días donde existe aporte observado."""
    unit = unit_label(flow_unit)
    if daily is None or daily.empty:
        return {"Sistema": entity, "Estado": "Sin datos"}
    mask = pd.to_numeric(daily["Obs_cfs"], errors="coerce").notna()
    common = daily.loc[mask].copy()
    if common.empty:
        return {"Sistema": entity, "Estado": "Sin observado común"}
    vals = {}
    for c in ["Hist_1997_cfs", "Hist_1998_cfs", "Hist_prom_9798_cfs", "Obs_cfs", "DSS_cfs"]:
        vals[c] = float(pd.to_numeric(common[c], errors="coerce").mean())
    return {
        "Sistema": entity,
        "Período común": f"{common['Fecha_dia'].min():%d-%m-%Y} a {common['Fecha_dia'].max():%d-%m-%Y}",
        "Días comunes": int(len(common)),
        f"Observado ({unit})": round(float(convert_flow(pd.Series([vals['Obs_cfs']]), flow_unit).iloc[0]), 3),
        f"1997 ({unit})": round(float(convert_flow(pd.Series([vals['Hist_1997_cfs']]), flow_unit).iloc[0]), 3),
        f"1998 ({unit})": round(float(convert_flow(pd.Series([vals['Hist_1998_cfs']]), flow_unit).iloc[0]), 3),
        f"Prom. 97-98 ({unit})": round(float(convert_flow(pd.Series([vals['Hist_prom_9798_cfs']]), flow_unit).iloc[0]), 3),
        f"DSS {dss_label} ({unit})": round(float(convert_flow(pd.Series([vals['DSS_cfs']]), flow_unit).iloc[0]), 3),
        "Dif. Obs vs 1997 (%)": round(float(pct_diff_obs_minus_dss(vals["Obs_cfs"], vals["Hist_1997_cfs"])), 2),
        "Dif. Obs vs 1998 (%)": round(float(pct_diff_obs_minus_dss(vals["Obs_cfs"], vals["Hist_1998_cfs"])), 2),
        "Dif. Obs vs prom. 97-98 (%)": round(float(pct_diff_obs_minus_dss(vals["Obs_cfs"], vals["Hist_prom_9798_cfs"])), 2),
        f"Dif. Obs vs DSS {dss_label} (%)": round(float(pct_diff_obs_minus_dss(vals["Obs_cfs"], vals["DSS_cfs"])), 2),
    }


def _nearest_ap_dss_daily_9798(
    daily: pd.DataFrame,
    cfg: Dict,
    obs_source: Optional[pd.DataFrame],
    evap_cfs: float = 0.0,
) -> pd.DataFrame:
    """Construye el AP DSS más cercano al observado para cada día disponible.

    La comparación se realiza contra todas las curvas AP DSS del embalse. Para
    cada fecha observada se conserva el percentil cuya diferencia absoluta con
    el aporte observado sea menor. El AP DSS se maneja como aporte total:
    AP neto DSS + evaporación.
    """
    columns = [
        "Fecha_dia", "DSS_cfs", "DSS_base_cfs", "DSS_reduction_cfs",
        "DSS_reduction_pct", "DSS_percentil", "DSS_label",
        "DSS_fecha_usada", "DSS_fecha_exacta",
    ]
    if daily is None or daily.empty:
        return pd.DataFrame(columns=columns)

    obs = _observed_ap_9798(obs_source)
    if obs.empty:
        return pd.DataFrame(columns=columns)

    rows: List[Dict[str, object]] = []
    for row in obs.itertuples(index=False):
        fecha = pd.to_datetime(getattr(row, "Fecha_dia", None), errors="coerce")
        valor = pd.to_numeric(pd.Series([getattr(row, "Obs_cfs", np.nan)]), errors="coerce").iloc[0]
        if pd.isna(fecha) or pd.isna(valor):
            continue
        nearest = _nearest_ap_percentile(
            daily,
            cfg,
            fecha,
            float(valor),
            dss_add_cfs=clean_evap_cfs(evap_cfs),
        )
        if not nearest:
            continue
        dss_base_cfs = float(nearest["dss_total_cfs"])
        rows.append({
            "Fecha_dia": fecha.normalize(),
            "DSS_cfs": dss_base_cfs,
            "DSS_base_cfs": dss_base_cfs,
            "DSS_reduction_cfs": 0.0,
            "DSS_reduction_pct": 0.0,
            "DSS_percentil": int(nearest["percentile"]),
            "DSS_label": str(nearest["label"]),
            "DSS_fecha_usada": pd.to_datetime(nearest["date"]).normalize(),
            "DSS_fecha_exacta": bool(nearest.get("exact_date", False)),
        })

    if not rows:
        return pd.DataFrame(columns=columns)
    return (
        pd.DataFrame(rows)
        .drop_duplicates(subset=["Fecha_dia"], keep="last")
        .sort_values("Fecha_dia")
        .reset_index(drop=True)
    )


def _fixed_ap_dss_daily_9798(
    daily: pd.DataFrame,
    cfg: Dict,
    pct_ref: int,
    evap_cfs: float = 0.0,
) -> Tuple[pd.DataFrame, Optional[int]]:
    """Devuelve un percentil AP DSS fijo y añade su etiqueta por fecha."""
    out, pct_used = _selected_ap_dss_9798(daily, cfg, pct_ref, evap_cfs)
    if out is None or out.empty:
        return pd.DataFrame(columns=[
            "Fecha_dia", "DSS_cfs", "DSS_base_cfs", "DSS_reduction_cfs",
            "DSS_reduction_pct", "DSS_percentil", "DSS_label",
        ]), pct_used
    result = out.copy()
    result["DSS_base_cfs"] = pd.to_numeric(result["DSS_cfs"], errors="coerce")
    result["DSS_reduction_cfs"] = 0.0
    result["DSS_reduction_pct"] = 0.0
    result["DSS_percentil"] = int(pct_used) if pct_used is not None else np.nan
    result["DSS_label"] = f"P{int(pct_used)}" if pct_used is not None else "—"
    result["DSS_fecha_usada"] = result["Fecha_dia"]
    result["DSS_fecha_exacta"] = True
    return result, pct_used


def _clean_dss_reduction_pct_9798(value: object) -> float:
    """Normaliza la reducción manual del aporte DSS entre 0% y 95%."""
    try:
        pct = float(value or 0.0)
    except Exception:
        pct = 0.0
    if not np.isfinite(pct):
        pct = 0.0
    return max(0.0, min(95.0, pct))


def _apply_dss_reduction_daily_9798(
    dss_daily: pd.DataFrame,
    reduction_pct: float,
) -> pd.DataFrame:
    """Aplica la reducción después de elegir el percentil DSS de referencia.

    Fórmula exacta:
        DSS ajustado = DSS base × (1 - reducción / 100)

    La selección del percentil no cambia: primero se identifica la curva DSS
    más cercana al observado (o P95 fijo) y luego se descuenta el porcentaje.
    """
    if dss_daily is None or dss_daily.empty:
        return dss_daily
    out = dss_daily.copy()
    pct = _clean_dss_reduction_pct_9798(reduction_pct)
    base = pd.to_numeric(
        out["DSS_base_cfs"] if "DSS_base_cfs" in out.columns else out.get("DSS_cfs"),
        errors="coerce",
    )
    out["DSS_base_cfs"] = base
    out["DSS_reduction_pct"] = pct
    out["DSS_reduction_cfs"] = base * (pct / 100.0)
    out["DSS_cfs"] = base * (1.0 - pct / 100.0)
    return out


def _suggest_dss_reduction_pct_9798(
    daily: pd.DataFrame,
    hist_year: int,
    target: str = "observado",
) -> Optional[float]:
    """Sugiere cuánto restar al DSS base para acercarlo al observado o histórico."""
    if daily is None or daily.empty:
        return None
    hist_col = f"Hist_{int(hist_year)}_cfs"
    base = pd.to_numeric(
        daily.get("DSS_base_cfs", daily.get("DSS_cfs")), errors="coerce"
    )
    if str(target).lower().startswith("hist") or str(target).startswith(str(hist_year)):
        ref = pd.to_numeric(daily.get(hist_col), errors="coerce")
    else:
        ref = pd.to_numeric(daily.get("Obs_cfs"), errors="coerce")
    valid = base.notna() & ref.notna() & (base > 0)
    if not valid.any():
        return None
    base_mean = float(base.loc[valid].mean())
    ref_mean = float(ref.loc[valid].mean())
    if not np.isfinite(base_mean) or base_mean <= 0 or not np.isfinite(ref_mean):
        return None
    # Solo se permite restar. Si el DSS ya está por debajo, la sugerencia es 0%.
    return _clean_dss_reduction_pct_9798(max(0.0, (1.0 - ref_mean / base_mean) * 100.0))


def _sum_chcp_9798_enhanced(gat: pd.DataFrame, alh: pd.DataFrame) -> pd.DataFrame:
    """Suma Gatún + Alhajuela e incorpora la combinación de percentiles DSS."""
    out = _sum_chcp_9798(gat, alh)
    if out is None or out.empty:
        return pd.DataFrame()

    def labels(df: pd.DataFrame, suffix: str) -> pd.DataFrame:
        if df is None or df.empty:
            return pd.DataFrame(columns=["Fecha_dia", f"DSS_label_{suffix}"])
        tmp = df[["Fecha_dia"]].copy()
        if "DSS_label" in df.columns:
            tmp[f"DSS_label_{suffix}"] = df["DSS_label"].astype("string")
        else:
            tmp[f"DSS_label_{suffix}"] = pd.Series(pd.NA, index=df.index, dtype="string")
        return tmp

    combo = labels(gat, "GAT").merge(labels(alh, "ALHA"), on="Fecha_dia", how="outer")
    combo["DSS_label"] = np.where(
        combo["DSS_label_GAT"].notna() & combo["DSS_label_ALHA"].notna(),
        "GAT " + combo["DSS_label_GAT"].astype(str) + " + ALHA " + combo["DSS_label_ALHA"].astype(str),
        pd.NA,
    )
    out = out.merge(combo[["Fecha_dia", "DSS_label"]], on="Fecha_dia", how="left")
    return out.sort_values("Fecha_dia").reset_index(drop=True)


def _deficit_pct_9798(current: pd.Series, reference: pd.Series) -> pd.Series:
    """Déficit porcentual: positivo cuando el valor actual está bajo la referencia.

    Fórmula usada para reproducir la lectura de la tabla operativa:
        (referencia - actual) / |referencia| × 100
    """
    cur = pd.to_numeric(current, errors="coerce")
    ref = pd.to_numeric(reference, errors="coerce")
    return ((ref - cur) / ref.abs().replace(0, np.nan) * 100.0).replace([np.inf, -np.inf], np.nan)


def _representative_dss_label_9798(values: pd.Series) -> str:
    """Resume el percentil DSS dominante dentro de un período."""
    clean = values.dropna().astype(str)
    clean = clean[~clean.isin(["", "<NA>", "nan", "None", "—"])]
    if clean.empty:
        return "—"
    counts = clean.value_counts()
    label = str(counts.index[0])
    if len(counts) == 1:
        return label
    return f"{label} ({int(counts.iloc[0])}/{int(len(clean))} días)"


def _period_label_9798(value: object, mode: str) -> str:
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        return "—"
    if mode == "Mensual":
        months = {
            1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
            7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
        }
        return f"{months.get(int(ts.month), ts.strftime('%b'))} {int(ts.year)}"
    if mode == "Semanal":
        return f"Semana desde {ts:%d-%m-%Y}"
    return f"{ts:%d-%m-%Y}"


def _aggregate_selected_history_9798(
    daily: pd.DataFrame,
    hist_year: int,
    mode: str,
    week_start: int = 5,
) -> pd.DataFrame:
    """Agrega observado, DSS cercano y el año histórico seleccionado.

    Los promedios se calculan prioritariamente sobre días comunes con observado,
    DSS e histórico, de modo que las diferencias no mezclen coberturas distintas.
    """
    if daily is None or daily.empty:
        return pd.DataFrame()
    hist_col = f"Hist_{int(hist_year)}_cfs"
    if hist_col not in daily.columns:
        return pd.DataFrame()

    base = daily.copy()
    base["Fecha_dia"] = pd.to_datetime(base["Fecha_dia"], errors="coerce").dt.normalize()
    base = base.dropna(subset=["Fecha_dia"]).sort_values("Fecha_dia")
    if "DSS_label" not in base.columns:
        base["DSS_label"] = "—"

    if mode == "Semanal":
        info = base["Fecha_dia"].apply(lambda d: operational_week_info(d, week_start=int(week_start)))
        base["Periodo"] = [item[1] for item in info]
    elif mode == "Mensual":
        base["Periodo"] = base["Fecha_dia"].dt.to_period("M").dt.to_timestamp()
    else:
        base["Periodo"] = base["Fecha_dia"]

    rows: List[Dict[str, object]] = []
    for period, group in base.groupby("Periodo", sort=True):
        obs = pd.to_numeric(group.get("Obs_cfs"), errors="coerce")
        dss = pd.to_numeric(group.get("DSS_cfs"), errors="coerce")
        dss_base = pd.to_numeric(group.get("DSS_base_cfs", group.get("DSS_cfs")), errors="coerce")
        hist = pd.to_numeric(group.get(hist_col), errors="coerce")
        triple = obs.notna() & dss.notna() & hist.notna()

        if triple.any():
            used = group.loc[triple].copy()
        else:
            obs_hist = obs.notna() & hist.notna()
            used = group.loc[obs_hist].copy() if obs_hist.any() else group.copy()

        obs_used = pd.to_numeric(used.get("Obs_cfs"), errors="coerce")
        dss_used = pd.to_numeric(used.get("DSS_cfs"), errors="coerce")
        dss_base_used = pd.to_numeric(used.get("DSS_base_cfs", used.get("DSS_cfs")), errors="coerce")
        hist_used = pd.to_numeric(used.get(hist_col), errors="coerce")

        row: Dict[str, object] = {
            "Periodo": pd.to_datetime(period),
            "Dias_periodo": int(len(group)),
            "Dias_comunes": int(triple.sum()),
            "DSS_label": _representative_dss_label_9798(group.get("DSS_label", pd.Series(dtype=str))),
            "Obs_cfs": float(obs_used.mean()) if obs_used.notna().any() else np.nan,
            "DSS_base_cfs": float(dss_base_used.mean()) if dss_base_used.notna().any() else np.nan,
            "DSS_cfs": float(dss_used.mean()) if dss_used.notna().any() else np.nan,
            "Hist_ref_cfs": float(hist_used.mean()) if hist_used.notna().any() else np.nan,
        }
        rows.append(row)

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["Deficit_Obs_vs_Hist_pct"] = _deficit_pct_9798(out["Obs_cfs"], out["Hist_ref_cfs"])
    out["Deficit_DSS_vs_Hist_pct"] = _deficit_pct_9798(out["DSS_cfs"], out["Hist_ref_cfs"])
    out["Deficit_Obs_vs_DSS_pct"] = _deficit_pct_9798(out["Obs_cfs"], out["DSS_cfs"])
    out["Dif_Obs_vs_DSS_pct"] = _pct_series_9798(out["Obs_cfs"], out["DSS_cfs"])
    out["Reduccion_DSS_pct"] = _deficit_pct_9798(out["DSS_cfs"], out["DSS_base_cfs"])
    return out.sort_values("Periodo").reset_index(drop=True)


def _detail_selected_history_9798(
    agg: pd.DataFrame,
    entity_label: str,
    hist_year: int,
    flow_unit: str,
    mode: str,
    reduction_label: str = "0%",
) -> pd.DataFrame:
    """Tabla legible por sistema para pantalla y exportación."""
    if agg is None or agg.empty:
        return pd.DataFrame()
    unit = unit_label(flow_unit)
    out = pd.DataFrame({
        "Período": agg["Periodo"].map(lambda v: _period_label_9798(v, mode)),
        "Sistema": entity_label,
        "P DSS cercano": agg["DSS_label"],
        "Reducción DSS aplicada": reduction_label,
        f"DSS base ({unit})": convert_flow(agg.get("DSS_base_cfs", agg["DSS_cfs"]), flow_unit),
        f"DSS ajustado ({unit})": convert_flow(agg["DSS_cfs"], flow_unit),
        f"Observado ({unit})": convert_flow(agg["Obs_cfs"], flow_unit),
        f"Aporte {hist_year} ({unit})": convert_flow(agg["Hist_ref_cfs"], flow_unit),
        f"Déficit observado vs {hist_year} (%)": agg["Deficit_Obs_vs_Hist_pct"],
        f"Déficit DSS ajustado vs {hist_year} (%)": agg["Deficit_DSS_vs_Hist_pct"],
        "Déficit observado vs DSS ajustado (%)": agg["Deficit_Obs_vs_DSS_pct"],
        "Días comunes": agg["Dias_comunes"],
    })
    numeric_cols = [c for c in out.columns if c not in {
        "Período", "Sistema", "P DSS cercano", "Reducción DSS aplicada"
    }]
    for c in numeric_cols:
        out[c] = pd.to_numeric(out[c], errors="coerce").round(2)
    return out


def _executive_selected_history_9798(
    aggs: Dict[str, pd.DataFrame],
    hist_year: int,
    flow_unit: str,
    mode: str,
) -> pd.DataFrame:
    """Tabla integrada tipo tablero, similar a la referencia entregada."""
    unit = unit_label(flow_unit)
    prefixes = {
        "Gatún": "GAT",
        "Alhajuela / Madden": "ALHA",
        "CHCP · GAT + ALHA": "CHCP",
    }
    merged: Optional[pd.DataFrame] = None
    for entity, prefix in prefixes.items():
        agg = aggs.get(entity, pd.DataFrame())
        if agg is None or agg.empty:
            continue
        block = pd.DataFrame({
            "Periodo_dt": agg["Periodo"],
            f"{prefix} · P DSS cercano": agg["DSS_label"],
            f"{prefix} · DSS cercano ({unit})": convert_flow(agg["DSS_cfs"], flow_unit),
            f"{prefix} · Observado ({unit})": convert_flow(agg["Obs_cfs"], flow_unit),
            f"{prefix} · {hist_year} ({unit})": convert_flow(agg["Hist_ref_cfs"], flow_unit),
            f"{prefix} · Déficit Obs vs {hist_year} (%)": agg["Deficit_Obs_vs_Hist_pct"],
            f"{prefix} · Déficit DSS vs {hist_year} (%)": agg["Deficit_DSS_vs_Hist_pct"],
            f"{prefix} · Obs vs DSS (%)": agg["Dif_Obs_vs_DSS_pct"],
        })
        merged = block if merged is None else merged.merge(block, on="Periodo_dt", how="outer")

    if merged is None or merged.empty:
        return pd.DataFrame()
    merged = merged.sort_values("Periodo_dt").reset_index(drop=True)
    merged.insert(0, "Período", merged["Periodo_dt"].map(lambda v: _period_label_9798(v, mode)))
    merged = merged.drop(columns=["Periodo_dt"])
    for c in merged.columns:
        if c == "Período" or "P DSS" in c:
            continue
        merged[c] = pd.to_numeric(merged[c], errors="coerce").round(2)
    return merged


def _plot_selected_history_hydrograph_9798(
    agg: pd.DataFrame,
    entity_label: str,
    hist_year: int,
    flow_unit: str,
    mode: str,
    key: str,
    reduction_label: str = "0%",
) -> None:
    """Hidrograma independiente: histórico, observado y DSS más cercano."""
    if agg is None or agg.empty:
        st.info(f"No hay datos para el hidrograma de {entity_label}.")
        return
    unit = unit_label(flow_unit)
    plot = agg.copy()
    plot["Hist"] = convert_flow(plot["Hist_ref_cfs"], flow_unit)
    plot["DSS"] = convert_flow(plot["DSS_cfs"], flow_unit)
    plot["Obs"] = convert_flow(plot["Obs_cfs"], flow_unit)
    marker_mode = "lines" if mode == "Diario" else "lines+markers"

    if PLOTLY_OK:
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=plot["Periodo"], y=plot["Hist"], mode=marker_mode,
            name=f"Aporte {hist_year}",
            line={"color": "#2e7d32", "width": 2},
            connectgaps=False,
        ))
        fig.add_trace(go.Scatter(
            x=plot["Periodo"], y=plot["DSS"], mode=marker_mode,
            name=f"DSS más cercano ajustado ({reduction_label})",
            text=plot["DSS_label"],
            line={"color": "#f59e0b", "width": 3, "dash": "dash"},
            hovertemplate=(
                "%{x|%d-%m-%Y}<br>DSS: %{y:,.2f}<br>Percentil: %{text}<extra></extra>"
            ),
            connectgaps=False,
        ))
        fig.add_trace(go.Scatter(
            x=plot["Periodo"], y=plot["Obs"], mode=marker_mode,
            name="Observado",
            line={"color": "#0077b6", "width": 4},
            connectgaps=False,
        ))
        fig.update_layout(**_base_layout(
            f"{entity_label} · Observado vs DSS ajustado vs {hist_year}",
            unit,
            height=455,
        ))
        fig.update_layout(
            hovermode="x unified",
            legend={"orientation": "h", "y": 1.08, "x": 0.0},
            margin={"l": 35, "r": 20, "t": 70, "b": 45},
        )
        st.plotly_chart(fig, use_container_width=True, key=key)
    else:
        st.line_chart(plot.set_index("Periodo")[["Hist", "DSS", "Obs"]])


def _summary_cards_selected_history_9798(
    aggs: Dict[str, pd.DataFrame],
    hist_year: int,
    flow_unit: str,
) -> None:
    """Métricas promedio para Gatún, Alhajuela y CHCP."""
    unit = unit_label(flow_unit)
    labels = ["Gatún", "Alhajuela / Madden", "CHCP · GAT + ALHA"]
    cols = st.columns(3)
    for col, entity in zip(cols, labels):
        agg = aggs.get(entity, pd.DataFrame())
        if agg is None or agg.empty:
            col.metric(entity, "Sin datos")
            continue
        obs = pd.to_numeric(agg["Obs_cfs"], errors="coerce").mean()
        hist = pd.to_numeric(agg["Hist_ref_cfs"], errors="coerce").mean()
        obs_disp = convert_flow(pd.Series([obs]), flow_unit).iloc[0] if pd.notna(obs) else np.nan
        signed = pct_diff_obs_minus_dss(obs, hist) if pd.notna(obs) and pd.notna(hist) else np.nan
        delta = f"{signed:+.1f}% vs {hist_year}" if pd.notna(signed) else "Sin comparación"
        col.metric(
            f"{entity} · observado promedio",
            f"{obs_disp:,.2f} {unit}" if pd.notna(obs_disp) else "—",
            delta=delta,
        )


def _tracking_common_dates_9798(
    gat: pd.DataFrame,
    alh: pd.DataFrame,
    hist_year: int,
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
) -> pd.DatetimeIndex:
    """Fechas exactas comparables en ambos embalses.

    Se exige aporte observado, DSS más cercano y aporte histórico en Gatún y
    Alhajuela. Así la suma CHCP y todos los porcentajes usan exactamente los
    mismos días, sin completar días futuros ni mezclar coberturas distintas.
    """
    hist_col = f"Hist_{int(hist_year)}_cfs"

    def available(df: pd.DataFrame) -> set[pd.Timestamp]:
        if df is None or df.empty or hist_col not in df.columns:
            return set()
        work = df.copy()
        work["Fecha_dia"] = pd.to_datetime(work["Fecha_dia"], errors="coerce").dt.normalize()
        mask = (
            work["Fecha_dia"].notna()
            & work["Fecha_dia"].between(pd.to_datetime(start_date), pd.to_datetime(end_date), inclusive="both")
            & pd.to_numeric(work.get("Obs_cfs"), errors="coerce").notna()
            & pd.to_numeric(work.get("DSS_cfs"), errors="coerce").notna()
            & pd.to_numeric(work.get(hist_col), errors="coerce").notna()
        )
        return set(pd.to_datetime(work.loc[mask, "Fecha_dia"]).dt.normalize().tolist())

    common = sorted(available(gat).intersection(available(alh)))
    return pd.DatetimeIndex(common)


def _exact_summary_row_9798(
    daily: pd.DataFrame,
    hist_year: int,
) -> Dict[str, object]:
    """Resumen ponderado por día para todo el período junio-fecha de corte."""
    hist_col = f"Hist_{int(hist_year)}_cfs"
    if daily is None or daily.empty or hist_col not in daily.columns:
        return {}
    work = daily.copy()
    obs = pd.to_numeric(work.get("Obs_cfs"), errors="coerce")
    dss = pd.to_numeric(work.get("DSS_cfs"), errors="coerce")
    hist = pd.to_numeric(work.get(hist_col), errors="coerce")
    mask = obs.notna() & dss.notna() & hist.notna()
    if not mask.any():
        return {}
    used = work.loc[mask].copy()
    obs_mean = float(pd.to_numeric(used["Obs_cfs"], errors="coerce").mean())
    dss_mean = float(pd.to_numeric(used["DSS_cfs"], errors="coerce").mean())
    dss_base_mean = float(pd.to_numeric(
        used.get("DSS_base_cfs", used["DSS_cfs"]), errors="coerce"
    ).mean())
    hist_mean = float(pd.to_numeric(used[hist_col], errors="coerce").mean())
    return {
        "DSS_label": _representative_dss_label_9798(used.get("DSS_label", pd.Series(dtype=str))),
        "Obs_cfs": obs_mean,
        "DSS_base_cfs": dss_base_mean,
        "DSS_cfs": dss_mean,
        "Hist_ref_cfs": hist_mean,
        "Deficit_Obs_vs_Hist_pct": float(_deficit_pct_9798(pd.Series([obs_mean]), pd.Series([hist_mean])).iloc[0]),
        "Deficit_DSS_vs_Hist_pct": float(_deficit_pct_9798(pd.Series([dss_mean]), pd.Series([hist_mean])).iloc[0]),
        "Deficit_Obs_vs_DSS_pct": float(_deficit_pct_9798(pd.Series([obs_mean]), pd.Series([dss_mean])).iloc[0]),
        "Dif_Obs_vs_DSS_pct": float(_pct_series_9798(pd.Series([obs_mean]), pd.Series([dss_mean])).iloc[0]),
        "Reduccion_DSS_pct": float(_deficit_pct_9798(pd.Series([dss_mean]), pd.Series([dss_base_mean])).iloc[0]),
        "Dias_comunes": int(mask.sum()),
    }


def _compact_tracking_table_9798(
    monthly: Dict[str, pd.DataFrame],
    filtered_daily: Dict[str, pd.DataFrame],
    hist_year: int,
    flow_unit: str,
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
) -> pd.DataFrame:
    """Tabla mensual compacta tipo seguimiento operativo, similar a la referencia."""
    unit = unit_label(flow_unit)
    gat = monthly.get("Gatún", pd.DataFrame())
    alh = monthly.get("Alhajuela / Madden", pd.DataFrame())
    chcp = monthly.get("CHCP · GAT + ALHA", pd.DataFrame())
    if gat.empty or alh.empty or chcp.empty:
        return pd.DataFrame()

    def block(df: pd.DataFrame, prefix: str, include_label: bool = True) -> pd.DataFrame:
        out = pd.DataFrame({"Periodo_dt": pd.to_datetime(df["Periodo"], errors="coerce")})
        if include_label:
            out[f"{prefix} · P DSS"] = df["DSS_label"].astype(str)
        out[f"{prefix} · DSS ({unit})"] = convert_flow(df["DSS_cfs"], flow_unit)
        out[f"{prefix} · {hist_year} ({unit})"] = convert_flow(df["Hist_ref_cfs"], flow_unit)
        out[f"{prefix} · Déficit DSS vs {hist_year} (%)"] = df["Deficit_DSS_vs_Hist_pct"]
        return out

    table = block(gat, "GAT", include_label=True)
    table["Días comparados"] = pd.to_numeric(gat["Dias_comunes"], errors="coerce")
    table = table.merge(block(alh, "ALHA", include_label=True), on="Periodo_dt", how="outer")
    table = table.merge(block(chcp, "CHCP", include_label=False), on="Periodo_dt", how="outer")
    table = table.sort_values("Periodo_dt").reset_index(drop=True)

    month_names = {
        1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
    }
    table.insert(0, "Mes", table["Periodo_dt"].map(lambda d: month_names.get(int(d.month), d.strftime("%b"))))
    table = table.drop(columns=["Periodo_dt"])

    # Fila final ponderada por todos los días comunes, no por promedio simple de meses.
    summaries = {
        name: _exact_summary_row_9798(filtered_daily.get(name, pd.DataFrame()), hist_year)
        for name in ["Gatún", "Alhajuela / Madden", "CHCP · GAT + ALHA"]
    }
    if all(summaries.values()):
        gat_s = summaries["Gatún"]
        alh_s = summaries["Alhajuela / Madden"]
        chcp_s = summaries["CHCP · GAT + ALHA"]
        label_end = pd.to_datetime(end_date).strftime("%d-%b")
        summary = {
            "Mes": f"Prom. Jun–{label_end}",
            "GAT · P DSS": gat_s["DSS_label"],
            f"GAT · DSS ({unit})": convert_flow(pd.Series([gat_s["DSS_cfs"]]), flow_unit).iloc[0],
            f"GAT · {hist_year} ({unit})": convert_flow(pd.Series([gat_s["Hist_ref_cfs"]]), flow_unit).iloc[0],
            f"GAT · Déficit DSS vs {hist_year} (%)": gat_s["Deficit_DSS_vs_Hist_pct"],
            "Días comparados": gat_s["Dias_comunes"],
            "ALHA · P DSS": alh_s["DSS_label"],
            f"ALHA · DSS ({unit})": convert_flow(pd.Series([alh_s["DSS_cfs"]]), flow_unit).iloc[0],
            f"ALHA · {hist_year} ({unit})": convert_flow(pd.Series([alh_s["Hist_ref_cfs"]]), flow_unit).iloc[0],
            f"ALHA · Déficit DSS vs {hist_year} (%)": alh_s["Deficit_DSS_vs_Hist_pct"],
            f"CHCP · DSS ({unit})": convert_flow(pd.Series([chcp_s["DSS_cfs"]]), flow_unit).iloc[0],
            f"CHCP · {hist_year} ({unit})": convert_flow(pd.Series([chcp_s["Hist_ref_cfs"]]), flow_unit).iloc[0],
            f"CHCP · Déficit DSS vs {hist_year} (%)": chcp_s["Deficit_DSS_vs_Hist_pct"],
        }
        table = pd.concat([table, pd.DataFrame([summary])], ignore_index=True)

    for col in table.columns:
        if col in {"Mes", "GAT · P DSS", "ALHA · P DSS"}:
            continue
        table[col] = pd.to_numeric(table[col], errors="coerce")
        if "Días" in col:
            table[col] = table[col].round(0).astype("Int64")
        elif "(%)" in col:
            table[col] = table[col].round(0)
        else:
            table[col] = table[col].round(1)
    return table


def _style_compact_tracking_9798(table: pd.DataFrame):
    """Colores suaves por bloque, siguiendo la lógica visual de la tabla de referencia."""
    if table is None or table.empty:
        return table
    numeric_formats: Dict[str, str] = {}
    for col in table.columns:
        if "Días" in col:
            numeric_formats[col] = "{:.0f}"
        elif "(%)" in col:
            numeric_formats[col] = "{:+.0f}"
        elif col not in {"Mes", "GAT · P DSS", "ALHA · P DSS"}:
            numeric_formats[col] = "{:.1f}"

    styler = table.style.format(numeric_formats, na_rep="—")
    styler = styler.set_properties(subset=["Mes"], **{
        "font-weight": "700", "background-color": "#f3f4f6"
    })
    yellow_cols = [c for c in table.columns if "P DSS" in c or "· DSS (" in c]
    green_cols = [c for c in table.columns if f"· {1997} (" in c or f"· {1998} (" in c]
    blue_cols = [c for c in table.columns if "Déficit DSS" in c]
    if yellow_cols:
        styler = styler.set_properties(subset=yellow_cols, **{
            "background-color": "#fff59d", "font-weight": "700"
        })
    if green_cols:
        styler = styler.set_properties(subset=green_cols, **{
            "background-color": "#a5d66a", "font-weight": "700"
        })
    if blue_cols:
        styler = styler.set_properties(subset=blue_cols, **{
            "background-color": "#7dd3fc", "font-weight": "700"
        })
    if "Días comparados" in table.columns:
        styler = styler.set_properties(subset=["Días comparados"], **{
            "background-color": "#e5e7eb", "font-weight": "700"
        })
    # Destacar la fila resumen final.
    if len(table) > 0:
        styler = styler.set_properties(
            subset=pd.IndexSlice[[table.index[-1]], :],
            **{"border-top": "2px solid #334155", "font-weight": "800"},
        )
    return styler


def _summary_cards_tracking_9798(
    filtered_daily: Dict[str, pd.DataFrame],
    hist_year: int,
    flow_unit: str,
) -> None:
    """Promedios exactos de observado, usando todos los días comunes del período."""
    unit = unit_label(flow_unit)
    entities = ["Gatún", "Alhajuela / Madden", "CHCP · GAT + ALHA"]
    cols = st.columns(3)
    for col, entity in zip(cols, entities):
        summary = _exact_summary_row_9798(filtered_daily.get(entity, pd.DataFrame()), hist_year)
        if not summary:
            col.metric(entity, "Sin datos")
            continue
        obs_value = convert_flow(pd.Series([summary["Obs_cfs"]]), flow_unit).iloc[0]
        deficit = summary["Deficit_Obs_vs_Hist_pct"]
        col.metric(
            f"{entity} · observado",
            f"{obs_value:,.1f} {unit}",
            delta=f"{deficit:+.0f}% déficit vs {hist_year}",
            help=(
                f"Promedio exacto sobre {summary['Dias_comunes']} días comunes entre junio y la fecha de corte. "
                "Un déficit positivo indica aporte observado menor que el año histórico."
            ),
        )



def _reduction_label_9798(entity: str, reduction_gat_pct: float, reduction_alh_pct: float) -> str:
    """Etiqueta legible del descuento aplicado a cada sistema."""
    rg = _clean_dss_reduction_pct_9798(reduction_gat_pct)
    ra = _clean_dss_reduction_pct_9798(reduction_alh_pct)
    if entity == "Gatún":
        return f"-{rg:g}%"
    if entity == "Alhajuela / Madden":
        return f"-{ra:g}%"
    return f"GAT -{rg:g}% + ALHA -{ra:g}%"


def _three_series_table_9798(
    aggs: Dict[str, pd.DataFrame],
    hist_year: int,
    flow_unit: str,
    mode: str,
    reduction_gat_pct: float,
    reduction_alh_pct: float,
) -> pd.DataFrame:
    """Tabla ordenada para comparar observado, DSS ajustado e histórico."""
    unit = unit_label(flow_unit)
    entity_order = {
        "Gatún": 0,
        "Alhajuela / Madden": 1,
        "CHCP · GAT + ALHA": 2,
    }
    rows: List[Dict[str, object]] = []
    for entity, order in entity_order.items():
        agg = aggs.get(entity, pd.DataFrame())
        if agg is None or agg.empty:
            continue
        for row in agg.itertuples(index=False):
            rows.append({
                "_periodo_dt": pd.to_datetime(getattr(row, "Periodo"), errors="coerce"),
                "_orden": order,
                "Período": _period_label_9798(getattr(row, "Periodo"), mode),
                "Sistema": entity.replace("CHCP · GAT + ALHA", "CHCP"),
                "P DSS cercano": getattr(row, "DSS_label", "—"),
                "Reducción DSS": _reduction_label_9798(
                    entity, reduction_gat_pct, reduction_alh_pct
                ),
                f"Observado ({unit})": convert_flow(
                    pd.Series([getattr(row, "Obs_cfs", np.nan)]), flow_unit
                ).iloc[0],
                f"DSS ajustado ({unit})": convert_flow(
                    pd.Series([getattr(row, "DSS_cfs", np.nan)]), flow_unit
                ).iloc[0],
                f"Aporte {hist_year} ({unit})": convert_flow(
                    pd.Series([getattr(row, "Hist_ref_cfs", np.nan)]), flow_unit
                ).iloc[0],
                "Déficit Obs vs DSS ajustado (%)": getattr(
                    row, "Deficit_Obs_vs_DSS_pct", np.nan
                ),
                f"Déficit Obs vs {hist_year} (%)": getattr(
                    row, "Deficit_Obs_vs_Hist_pct", np.nan
                ),
                f"Déficit DSS ajustado vs {hist_year} (%)": getattr(
                    row, "Deficit_DSS_vs_Hist_pct", np.nan
                ),
                "Días exactos comparados": getattr(row, "Dias_comunes", np.nan),
            })
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows).sort_values(["_periodo_dt", "_orden"]).drop(
        columns=["_periodo_dt", "_orden"]
    ).reset_index(drop=True)
    for col in out.columns:
        if col in {"Período", "Sistema", "P DSS cercano", "Reducción DSS"}:
            continue
        out[col] = pd.to_numeric(out[col], errors="coerce")
        if "Días" in col:
            out[col] = out[col].round(0).astype("Int64")
        else:
            out[col] = out[col].round(2)
    return out


def _exact_three_series_summary_9798(
    filtered_daily: Dict[str, pd.DataFrame],
    hist_year: int,
    flow_unit: str,
    reduction_gat_pct: float,
    reduction_alh_pct: float,
) -> pd.DataFrame:
    """Resumen exacto ponderado por cada día común del período completo."""
    unit = unit_label(flow_unit)
    rows: List[Dict[str, object]] = []
    for entity in ["Gatún", "Alhajuela / Madden", "CHCP · GAT + ALHA"]:
        summary = _exact_summary_row_9798(filtered_daily.get(entity, pd.DataFrame()), hist_year)
        if not summary:
            continue
        rows.append({
            "Sistema": entity.replace("CHCP · GAT + ALHA", "CHCP"),
            "P DSS cercano": summary["DSS_label"],
            "Reducción DSS": _reduction_label_9798(
                entity, reduction_gat_pct, reduction_alh_pct
            ),
            f"Observado ({unit})": convert_flow(
                pd.Series([summary["Obs_cfs"]]), flow_unit
            ).iloc[0],
            f"DSS ajustado ({unit})": convert_flow(
                pd.Series([summary["DSS_cfs"]]), flow_unit
            ).iloc[0],
            f"Aporte {hist_year} ({unit})": convert_flow(
                pd.Series([summary["Hist_ref_cfs"]]), flow_unit
            ).iloc[0],
            "Déficit Obs vs DSS ajustado (%)": summary["Deficit_Obs_vs_DSS_pct"],
            f"Déficit Obs vs {hist_year} (%)": summary["Deficit_Obs_vs_Hist_pct"],
            f"Déficit DSS ajustado vs {hist_year} (%)": summary["Deficit_DSS_vs_Hist_pct"],
            "Días exactos comparados": summary["Dias_comunes"],
        })
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    for col in out.columns:
        if col in {"Sistema", "P DSS cercano", "Reducción DSS"}:
            continue
        out[col] = pd.to_numeric(out[col], errors="coerce")
        if "Días" in col:
            out[col] = out[col].round(0).astype("Int64")
        else:
            out[col] = out[col].round(2)
    return out


def _style_three_series_table_9798(table: pd.DataFrame):
    """Distingue visualmente las tres series y los déficits."""
    if table is None or table.empty:
        return table
    formats: Dict[str, str] = {}
    for col in table.columns:
        if "Días" in col:
            formats[col] = "{:.0f}"
        elif "(%)" in col:
            formats[col] = "{:+.2f}"
        elif col not in {"Período", "Sistema", "P DSS cercano", "Reducción DSS"}:
            formats[col] = "{:,.2f}"
    styler = table.style.format(formats, na_rep="—")
    obs_cols = [c for c in table.columns if c.startswith("Observado")]
    dss_cols = [c for c in table.columns if c.startswith("DSS ajustado") or c in {"P DSS cercano", "Reducción DSS"}]
    hist_cols = [c for c in table.columns if c.startswith("Aporte ")]
    deficit_cols = [c for c in table.columns if c.startswith("Déficit")]
    if obs_cols:
        styler = styler.set_properties(subset=obs_cols, **{"background-color": "#dbeafe", "font-weight": "700"})
    if dss_cols:
        styler = styler.set_properties(subset=dss_cols, **{"background-color": "#fef3c7", "font-weight": "700"})
    if hist_cols:
        styler = styler.set_properties(subset=hist_cols, **{"background-color": "#dcfce7", "font-weight": "700"})
    if deficit_cols:
        styler = styler.set_properties(subset=deficit_cols, **{"background-color": "#f1f5f9"})
    return styler


def _aggregate_full_reference_9798(
    daily: pd.DataFrame,
    hist_year: int,
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
) -> pd.DataFrame:
    """Promedios de meses completos usando solo histórico y DSS P95 fijo."""
    hist_col = f"Hist_{int(hist_year)}_cfs"
    if daily is None or daily.empty or hist_col not in daily.columns:
        return pd.DataFrame()
    base = daily.copy()
    base["Fecha_dia"] = pd.to_datetime(base["Fecha_dia"], errors="coerce").dt.normalize()
    base = base[
        base["Fecha_dia"].between(pd.to_datetime(start_date), pd.to_datetime(end_date), inclusive="both")
    ].copy()
    base["Periodo"] = base["Fecha_dia"].dt.to_period("M").dt.to_timestamp()
    rows: List[Dict[str, object]] = []
    for period, group in base.groupby("Periodo", sort=True):
        dss = pd.to_numeric(group.get("DSS_cfs"), errors="coerce")
        hist = pd.to_numeric(group.get(hist_col), errors="coerce")
        valid = dss.notna() & hist.notna()
        expected = int(pd.to_datetime(period).days_in_month)
        days = int(valid.sum())
        if days != expected:
            continue
        rows.append({
            "Periodo": pd.to_datetime(period),
            "DSS_label": _representative_dss_label_9798(group.get("DSS_label", pd.Series(dtype=str))),
            "DSS_cfs": float(dss.loc[valid].mean()),
            "Hist_ref_cfs": float(hist.loc[valid].mean()),
            "Deficit_DSS_vs_Hist_pct": float(_deficit_pct_9798(
                pd.Series([dss.loc[valid].mean()]), pd.Series([hist.loc[valid].mean()])
            ).iloc[0]),
            "Dias_referencia": days,
        })
    return pd.DataFrame(rows)


def _full_p95_reference_table_9798(
    aggs: Dict[str, pd.DataFrame],
    hist_year: int,
    flow_unit: str,
    reduction_gat_pct: float,
    reduction_alh_pct: float,
) -> pd.DataFrame:
    """Tabla junio-diciembre de períodos completos para 1997/1998 y DSS P95."""
    unit = unit_label(flow_unit)
    rows: List[Dict[str, object]] = []
    for entity in ["Gatún", "Alhajuela / Madden", "CHCP · GAT + ALHA"]:
        agg = aggs.get(entity, pd.DataFrame())
        if agg is None or agg.empty:
            continue
        for row in agg.itertuples(index=False):
            rows.append({
                "Período completo": _period_label_9798(getattr(row, "Periodo"), "Mensual"),
                "Sistema": entity.replace("CHCP · GAT + ALHA", "CHCP"),
                "P DSS fijo": getattr(row, "DSS_label", "P95"),
                "Reducción DSS": _reduction_label_9798(
                    entity, reduction_gat_pct, reduction_alh_pct
                ),
                f"DSS P95 ajustado ({unit})": convert_flow(
                    pd.Series([getattr(row, "DSS_cfs", np.nan)]), flow_unit
                ).iloc[0],
                f"Aporte {hist_year} ({unit})": convert_flow(
                    pd.Series([getattr(row, "Hist_ref_cfs", np.nan)]), flow_unit
                ).iloc[0],
                f"Déficit DSS P95 vs {hist_year} (%)": getattr(
                    row, "Deficit_DSS_vs_Hist_pct", np.nan
                ),
                "Días del período completo": getattr(row, "Dias_referencia", np.nan),
                "_periodo_dt": pd.to_datetime(getattr(row, "Periodo"), errors="coerce"),
                "_orden": {"Gatún": 0, "Alhajuela / Madden": 1, "CHCP · GAT + ALHA": 2}[entity],
            })
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows).sort_values(["_periodo_dt", "_orden"]).drop(
        columns=["_periodo_dt", "_orden"]
    ).reset_index(drop=True)
    for col in out.columns:
        if col in {"Período completo", "Sistema", "P DSS fijo", "Reducción DSS"}:
            continue
        out[col] = pd.to_numeric(out[col], errors="coerce")
        if "Días" in col:
            out[col] = out[col].round(0).astype("Int64")
        else:
            out[col] = out[col].round(2)
    return out


def tab_comparacion_9798(
    dss_bytes: bytes,
    hist_bytes: Optional[bytes],
    hist_name: str,
    flow_unit: str,
    pct_ref_gat: int,
    pct_ref_alh: int,
    obs_gat_aporte: Optional[pd.DataFrame],
    obs_alh_aporte: Optional[pd.DataFrame],
    evap_gat_cfs: float = 0.0,
    evap_alh_cfs: float = 0.0,
) -> None:
    st.subheader("📚 Comparación exacta de aportes · observado vs DSS vs histórico")
    st.caption(
        "Compara tres series con la misma cobertura diaria: aporte observado del Web Portal, "
        "percentil DSS diario más cercano y aporte histórico. Además, muestra los meses completos "
        "de junio a diciembre para el histórico y el DSS P95 fijo."
    )
    if not hist_bytes:
        st.warning(
            "No se cargó la serie 1997-1998. Coloque `1997_1998.xlsx` en la carpeta `data` "
            "o cárguela en el panel lateral."
        )
        return

    try:
        hist, hist_meta = load_historical_9798(hist_bytes)
    except Exception as exc:
        st.error(f"No se pudo preparar la serie histórica: {exc}")
        return

    try:
        gat_raw = load_dss_sheet(dss_bytes, RESERVOIR_CONFIG["gatun"]["sheet"])
        alh_raw = load_dss_sheet(dss_bytes, RESERVOIR_CONFIG["alhajuela"]["sheet"])
        gat_daily = to_daily(gat_raw, RESERVOIR_CONFIG["gatun"])
        alh_daily = to_daily(alh_raw, RESERVOIR_CONFIG["alhajuela"])
        gat_daily = apply_may_hydrograph_ap_adjustment(
            gat_daily, RESERVOIR_CONFIG["gatun"], obs_gat_aporte
        )
        alh_daily = apply_may_hydrograph_ap_adjustment(
            alh_daily, RESERVOIR_CONFIG["alhajuela"], obs_alh_aporte
        )
    except Exception as exc:
        st.error(f"No se pudieron cargar los aportes DSS: {exc}")
        return

    if gat_daily.empty or alh_daily.empty:
        st.warning("No hay datos diarios DSS suficientes para esta comparación.")
        return

    year_candidates = pd.concat(
        [gat_daily["Fecha_dia"], alh_daily["Fecha_dia"]], ignore_index=True
    )
    year_candidates = pd.to_datetime(year_candidates, errors="coerce").dropna().dt.year
    target_year = int(year_candidates.mode().iloc[0]) if not year_candidates.empty else int(today_panama().year)

    # Esta pestaña se mantiene en m³/s para facilitar la lectura operativa.
    comparison_unit = "m³/s"
    hist_year = int(st.selectbox(
        "Año histórico de comparación",
        [1997, 1998],
        index=0,
        key="hist9798_selected_year_june_tracking",
        help="El seguimiento principal usa los mismos días del observado; la tabla P95 usa meses completos.",
    ))

    gat_obs = _observed_ap_9798(obs_gat_aporte)
    alh_obs = _observed_ap_9798(obs_alh_aporte)

    # Primero se identifica el percentil DSS más cercano SIN reducción.
    # El descuento manual se aplica después, de modo que no cambie el percentil seleccionado.
    gat_dss_base = _nearest_ap_dss_daily_9798(
        gat_daily, RESERVOIR_CONFIG["gatun"], obs_gat_aporte, evap_gat_cfs
    )
    alh_dss_base = _nearest_ap_dss_daily_9798(
        alh_daily, RESERVOIR_CONFIG["alhajuela"], obs_alh_aporte, evap_alh_cfs
    )

    gat_base = _build_9798_entity_daily(hist, target_year, "Gatun_cfs", gat_dss_base, gat_obs)
    alh_base = _build_9798_entity_daily(hist, target_year, "Alhajuela_cfs", alh_dss_base, alh_obs)

    today = today_panama()
    tracking_start = pd.Timestamp(year=target_year, month=6, day=1)
    current_month_start = pd.Timestamp(year=target_year, month=int(today.month), day=1)
    current_day = min(int(today.day), int(current_month_start.days_in_month))
    calendar_cutoff = pd.Timestamp(year=target_year, month=int(today.month), day=current_day)
    if calendar_cutoff < tracking_start:
        st.info("El seguimiento junio-fecha actual estará disponible a partir del 1 de junio.")
        return

    common_dates = _tracking_common_dates_9798(
        gat_base, alh_base, hist_year, tracking_start, calendar_cutoff
    )
    if common_dates.empty:
        st.warning(
            "No existen días comunes con observado, DSS más cercano e histórico desde junio. "
            "Verifique los CSV de aportes observados del Web Portal."
        )
        return

    exact_end = min(pd.to_datetime(common_dates.max()).normalize(), calendar_cutoff)
    exact_dates = common_dates[(common_dates >= tracking_start) & (common_dates <= exact_end)]
    exact_date_set = set(pd.to_datetime(exact_dates).normalize())

    base_for_suggestion = {
        "Gatún": gat_base[gat_base["Fecha_dia"].isin(exact_date_set)].copy(),
        "Alhajuela / Madden": alh_base[alh_base["Fecha_dia"].isin(exact_date_set)].copy(),
    }
    suggested = {
        "gat_obs": _suggest_dss_reduction_pct_9798(base_for_suggestion["Gatún"], hist_year, "observado"),
        "gat_hist": _suggest_dss_reduction_pct_9798(base_for_suggestion["Gatún"], hist_year, "historico"),
        "alh_obs": _suggest_dss_reduction_pct_9798(base_for_suggestion["Alhajuela / Madden"], hist_year, "observado"),
        "alh_hist": _suggest_dss_reduction_pct_9798(base_for_suggestion["Alhajuela / Madden"], hist_year, "historico"),
    }

    st.markdown("### Ajuste porcentual del aporte DSS")
    st.caption(
        "La reducción se aplica al AP total DSS después de sumar evaporación: "
        "**DSS ajustado = DSS base × (1 − porcentaje/100)**. "
        "El percentil cercano se selecciona antes de aplicar el descuento."
    )
    r1, r2 = st.columns(2)
    reduction_gat_pct = r1.number_input(
        "Restar al aporte DSS de Gatún (%)",
        min_value=0.0,
        max_value=95.0,
        value=0.0,
        step=1.0,
        format="%.1f",
        key="hist9798_dss_reduction_gat_pct",
        help="Se aplica al DSS cercano y también al DSS P95 de la referencia completa.",
    )
    reduction_alh_pct = r2.number_input(
        "Restar al aporte DSS de Alhajuela (%)",
        min_value=0.0,
        max_value=95.0,
        value=0.0,
        step=1.0,
        format="%.1f",
        key="hist9798_dss_reduction_alh_pct",
        help="Se aplica al DSS cercano y también al DSS P95 de la referencia completa.",
    )

    def _s(v: Optional[float]) -> str:
        return f"{v:.1f}%" if v is not None and np.isfinite(v) else "—"

    st.info(
        "**Porcentaje orientativo para igualar promedios del período exacto:** "
        f"Gatún → observado {_s(suggested['gat_obs'])}, {hist_year} {_s(suggested['gat_hist'])}; "
        f"Alhajuela → observado {_s(suggested['alh_obs'])}, {hist_year} {_s(suggested['alh_hist'])}. "
        "La entrada manual permite usar cualquier porcentaje entre 0% y 95%."
    )

    gat_dss = _apply_dss_reduction_daily_9798(gat_dss_base, reduction_gat_pct)
    alh_dss = _apply_dss_reduction_daily_9798(alh_dss_base, reduction_alh_pct)

    gat = _build_9798_entity_daily(hist, target_year, "Gatun_cfs", gat_dss, gat_obs)
    alh = _build_9798_entity_daily(hist, target_year, "Alhajuela_cfs", alh_dss, alh_obs)
    chcp = _sum_chcp_9798_enhanced(gat, alh)

    entities = {
        "Gatún": gat,
        "Alhajuela / Madden": alh,
        "CHCP · GAT + ALHA": chcp,
    }
    filtered: Dict[str, pd.DataFrame] = {}
    monthly: Dict[str, pd.DataFrame] = {}
    daily_agg: Dict[str, pd.DataFrame] = {}
    for name, df in entities.items():
        work = df.copy()
        work["Fecha_dia"] = pd.to_datetime(work["Fecha_dia"], errors="coerce").dt.normalize()
        work = work[work["Fecha_dia"].isin(exact_date_set)].sort_values("Fecha_dia").copy()
        filtered[name] = work
        monthly[name] = _aggregate_selected_history_9798(work, hist_year, "Mensual", week_start=5)
        daily_agg[name] = _aggregate_selected_history_9798(work, hist_year, "Diario", week_start=5)

    exact_days = int(len(exact_dates))
    st.info(
        f"**Período exacto común:** {tracking_start:%d-%m-%Y} al {exact_end:%d-%m-%Y} · "
        f"**{exact_days} días** · unidad **m³/s** · DSS: percentil diario más cercano al observado."
    )
    st.caption(
        f"Archivo histórico: {hist_name or 'serie 1997-1998'} · "
        f"1997: {hist_meta.get('dias_1997', 0)} días · "
        f"1998: {hist_meta.get('dias_1998', 0)} días. "
        "Déficit positivo significa que la primera serie está por debajo de la referencia."
    )

    st.markdown("### Resumen exacto junio–fecha disponible")
    exact_summary = _exact_three_series_summary_9798(
        filtered,
        hist_year,
        comparison_unit,
        reduction_gat_pct,
        reduction_alh_pct,
    )
    if exact_summary.empty:
        st.warning("No se pudo construir el resumen exacto.")
        return
    st.dataframe(
        _style_three_series_table_9798(exact_summary),
        use_container_width=True,
        hide_index=True,
        height=min(270, 85 + 42 * len(exact_summary)),
    )

    st.markdown("### Seguimiento mensual con los mismos días observados")
    st.caption(
        "Cada fila compara exactamente los mismos días para observado, DSS ajustado e histórico. "
        "El mes actual queda cortado en la última fecha común disponible."
    )
    monthly_table = _three_series_table_9798(
        monthly,
        hist_year,
        comparison_unit,
        "Mensual",
        reduction_gat_pct,
        reduction_alh_pct,
    )
    if monthly_table.empty:
        st.warning("No se pudo construir la comparación mensual.")
        return
    st.dataframe(
        _style_three_series_table_9798(monthly_table),
        use_container_width=True,
        hide_index=True,
        height=min(560, 95 + 35 * len(monthly_table)),
    )

    st.markdown("### Hidrogramas diarios comparativos")
    st.caption(
        "Azul: observado del Web Portal · amarillo: DSS cercano con reducción aplicada · verde: histórico."
    )
    st.markdown("#### Gatún")
    _plot_selected_history_hydrograph_9798(
        daily_agg.get("Gatún", pd.DataFrame()),
        "Gatún",
        hist_year,
        comparison_unit,
        "Diario",
        key=f"hist9798_june_hydro_gat_{hist_year}_{exact_end:%Y%m%d}",
        reduction_label=_reduction_label_9798("Gatún", reduction_gat_pct, reduction_alh_pct),
    )
    st.markdown("#### Alhajuela / Madden")
    _plot_selected_history_hydrograph_9798(
        daily_agg.get("Alhajuela / Madden", pd.DataFrame()),
        "Alhajuela / Madden",
        hist_year,
        comparison_unit,
        "Diario",
        key=f"hist9798_june_hydro_alh_{hist_year}_{exact_end:%Y%m%d}",
        reduction_label=_reduction_label_9798("Alhajuela / Madden", reduction_gat_pct, reduction_alh_pct),
    )

    # Referencia adicional solicitada: meses completos 1997/1998 vs DSS P95 fijo.
    gat_p95_base, gat_p95_used = _fixed_ap_dss_daily_9798(
        gat_daily, RESERVOIR_CONFIG["gatun"], 95, evap_gat_cfs
    )
    alh_p95_base, alh_p95_used = _fixed_ap_dss_daily_9798(
        alh_daily, RESERVOIR_CONFIG["alhajuela"], 95, evap_alh_cfs
    )
    gat_p95 = _apply_dss_reduction_daily_9798(gat_p95_base, reduction_gat_pct)
    alh_p95 = _apply_dss_reduction_daily_9798(alh_p95_base, reduction_alh_pct)
    empty_obs = pd.DataFrame(columns=["Fecha_dia", "Obs_cfs"])
    gat_full = _build_9798_entity_daily(hist, target_year, "Gatun_cfs", gat_p95, empty_obs)
    alh_full = _build_9798_entity_daily(hist, target_year, "Alhajuela_cfs", alh_p95, empty_obs)
    chcp_full = _sum_chcp_9798_enhanced(gat_full, alh_full)
    full_start = pd.Timestamp(year=target_year, month=6, day=1)
    full_end = pd.Timestamp(year=target_year, month=12, day=31)
    full_aggs = {
        "Gatún": _aggregate_full_reference_9798(gat_full, hist_year, full_start, full_end),
        "Alhajuela / Madden": _aggregate_full_reference_9798(alh_full, hist_year, full_start, full_end),
        "CHCP · GAT + ALHA": _aggregate_full_reference_9798(chcp_full, hist_year, full_start, full_end),
    }
    full_table = _full_p95_reference_table_9798(
        full_aggs,
        hist_year,
        comparison_unit,
        reduction_gat_pct,
        reduction_alh_pct,
    )

    st.markdown("### Referencia de períodos completos · histórico vs DSS P95")
    st.caption(
        f"Promedios mensuales completos de junio a diciembre. No dependen de la disponibilidad del observado. "
        f"Percentiles utilizados: Gatún P{gat_p95_used if gat_p95_used is not None else '—'} y "
        f"Alhajuela P{alh_p95_used if alh_p95_used is not None else '—'}."
    )
    if full_table.empty:
        st.info("No hay meses completos simultáneos entre el histórico y el DSS P95.")
    else:
        st.dataframe(
            _style_three_series_table_9798(full_table),
            use_container_width=True,
            hide_index=True,
            height=min(650, 95 + 35 * len(full_table)),
        )

    # Descargas: resumen, seguimiento mensual, referencia P95 y detalle diario exacto.
    dl1, dl2 = st.columns(2)
    dl1.download_button(
        "⬇️ Descargar comparación mensual (CSV)",
        monthly_table.to_csv(index=False).encode("utf-8-sig"),
        f"comparacion_aportes_junio_{exact_end:%Y%m%d}_{hist_year}.csv",
        "text/csv",
        key=f"hist9798_june_csv_{hist_year}_{exact_end:%Y%m%d}",
        use_container_width=True,
    )

    detail_gat = _detail_selected_history_9798(
        daily_agg.get("Gatún", pd.DataFrame()),
        "Gatún",
        hist_year,
        comparison_unit,
        "Diario",
        _reduction_label_9798("Gatún", reduction_gat_pct, reduction_alh_pct),
    )
    detail_alh = _detail_selected_history_9798(
        daily_agg.get("Alhajuela / Madden", pd.DataFrame()),
        "Alhajuela / Madden",
        hist_year,
        comparison_unit,
        "Diario",
        _reduction_label_9798("Alhajuela / Madden", reduction_gat_pct, reduction_alh_pct),
    )
    detail_chcp = _detail_selected_history_9798(
        daily_agg.get("CHCP · GAT + ALHA", pd.DataFrame()),
        "CHCP",
        hist_year,
        comparison_unit,
        "Diario",
        _reduction_label_9798("CHCP · GAT + ALHA", reduction_gat_pct, reduction_alh_pct),
    )
    try:
        xlsx = _excel_bytes_from_sheets({
            "Resumen exacto": exact_summary,
            "Seguimiento mensual": monthly_table,
            "Referencia P95 completa": full_table,
            "Gatun diario": detail_gat,
            "Alhajuela diario": detail_alh,
            "CHCP diario": detail_chcp,
        })
        dl2.download_button(
            "⬇️ Descargar comparación completa (Excel)",
            xlsx,
            f"comparacion_aportes_exacta_{hist_year}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"hist9798_june_xlsx_{hist_year}_{exact_end:%Y%m%d}",
            use_container_width=True,
        )
    except Exception as exc:
        dl2.warning(f"No se pudo generar el Excel: {exc}")


# ─────────────────────────────────────────────────────────────────────
# PESTAÑA 7
# PESTAÑA 7 — APORTE INSTANTÁNEO
# ─────────────────────────────────────────────────────────────────────

def tab_aporte_instantaneo(
    dss_bytes: bytes,
    flow_unit: str,
    pct_ref_gat: int = 50,
    pct_ref_alh: int = 50,
    obs_gat_aporte: Optional[pd.DataFrame] = None,
    obs_alh_aporte: Optional[pd.DataFrame] = None,
    evap_gat_cfs: float = 0.0,
    evap_alh_cfs: float = 0.0,
) -> None:
    """Pestaña de aporte instantáneo.

    Muestra el visor meteorológico y compara aportes instantáneos/manuales,
    aportes observados de Aquarius/BulkExport y AP total DSS estimado.
    """
    st.subheader("⚡ Aporte instantáneo")
    st.caption(
        "Visor de referencia meteorológica y comparación contra Aquarius/BulkExport y DSS. "
        "El DSS se compara como **AP total DSS estimado = AP neto DSS + evaporación**."
    )
    st.caption(SIMULATION_NOTE)

    try:
        gat_raw = load_dss_sheet(dss_bytes, RESERVOIR_CONFIG["gatun"]["sheet"])
        alh_raw = load_dss_sheet(dss_bytes, RESERVOIR_CONFIG["alhajuela"]["sheet"])
        gat_d = to_daily(gat_raw, RESERVOIR_CONFIG["gatun"])
        alh_d = to_daily(alh_raw, RESERVOIR_CONFIG["alhajuela"])
        gat_d = apply_may_hydrograph_ap_adjustment(gat_d, RESERVOIR_CONFIG["gatun"], obs_gat_aporte)
        alh_d = apply_may_hydrograph_ap_adjustment(alh_d, RESERVOIR_CONFIG["alhajuela"], obs_alh_aporte)
    except Exception as exc:
        st.error(f"Error DSS: {exc}")
        return

    c_img, c_ctrl = st.columns([1.15, 1.25])
    with c_img:
        st.markdown("#### Visor — aporte instantáneo")
        try:
            ts = pd.Timestamp.utcnow().strftime("%Y%m%d%H%M%S")
            components.html(f"""
            <div style="border:1px solid rgba(0,62,105,.18);border-radius:14px;padding:10px;background:#f8fafc">
                <img src="{RADAR_URL}?t={ts}" style="width:100%;border-radius:10px;" />
                <div style="font-size:12px;color:#667085;margin-top:6px;">
                    Aporte instantáneo · Radar meteorológico Canal · {pd.Timestamp.now():%d-%m-%Y %H:%M}
                </div>
            </div>""", height=560)
            if st.checkbox("Auto-recargar cada 5 min", key="ap_inst_auto"):
                components.html("<script>setTimeout(()=>window.parent.location.reload(),300000);</script>", height=0)
        except Exception as exc:
            st.warning(f"No se pudo mostrar el visor: {exc}")

    with c_ctrl:
        st.markdown("#### Comparar aporte instantáneo / Aquarius / DSS")
        ref_date = st.date_input("Fecha de referencia DSS para entrada manual", value=today_panama().date(), key="apinst_ref")
        unit_obs = st.selectbox("Unidad del aporte manual", ["cfs", "m³/s", "hm³/d"],
                                format_func=unit_label, key="apinst_unit")
        st.markdown("##### Entrada manual opcional")
        g1, g2 = st.columns(2)
        obs_g = g1.number_input(f"GAT aporte instantáneo ({unit_label(unit_obs)})", min_value=0.0, step=10.0, key="apinst_g")
        obs_a = g2.number_input(f"ALHA aporte instantáneo ({unit_label(unit_obs)})", min_value=0.0, step=10.0, key="apinst_a")
        st.markdown("##### Evaporación usada para ajustar AP DSS")
        e1, e2 = st.columns(2)
        evap_g = e1.number_input("Evap. GAT (p³/s)", min_value=0.0, value=float(clean_evap_cfs(evap_gat_cfs)), step=10.0, format="%.3f", key="apinst_evap_g")
        evap_a = e2.number_input("Evap. ALHA (p³/s)", min_value=0.0, value=float(clean_evap_cfs(evap_alh_cfs)), step=10.0, format="%.3f", key="apinst_evap_a")
        st.info("Fórmula: **AP total DSS estimado = AP neto DSS + evaporación**.")
        _show_ap_may_adjustment_note(gat_d, "gatun", "aporte instantáneo Gatún")
        _show_ap_may_adjustment_note(alh_d, "alhajuela", "aporte instantáneo Alhajuela")

    def _last_aquarius(odf: Optional[pd.DataFrame]) -> Optional[Tuple[pd.Timestamp, float, str]]:
        if not _valid_df(odf):
            return None
        tmp = clamp_observed_future_dates(odf, "Fecha_dia")
        tmp = tmp[tmp["Valor"].notna()].copy()
        tmp["Fecha_dia"] = pd.to_datetime(tmp["Fecha_dia"], errors="coerce").dt.normalize()
        tmp = tmp[tmp["Fecha_dia"].notna()].sort_values("Fecha_dia")
        tmp_past = tmp[tmp["Fecha_dia"] <= today_panama()]
        if not tmp_past.empty:
            r = tmp_past.iloc[-1]
        elif not tmp.empty:
            r = tmp.iloc[-1]
        else:
            return None
        return pd.to_datetime(r["Fecha_dia"]), float(r["Valor"]), str(r.get("Fuente", "Aquarius/BulkExport"))

    comparisons = []
    inputs = []
    aq_g = _last_aquarius(obs_gat_aporte)
    aq_a = _last_aquarius(obs_alh_aporte)
    if aq_g:
        inputs.append(("Gatún", "Aquarius/BulkExport", aq_g[0], aq_g[1], "cfs", evap_g, gat_d, RESERVOIR_CONFIG["gatun"], aq_g[2], pct_ref_gat))
    if aq_a:
        inputs.append(("Alhajuela / Madden", "Aquarius/BulkExport", aq_a[0], aq_a[1], "cfs", evap_a, alh_d, RESERVOIR_CONFIG["alhajuela"], aq_a[2], pct_ref_alh))
    if obs_g > 0:
        inputs.append(("Gatún", "Manual instantáneo", pd.to_datetime(ref_date), float(obs_g), unit_obs, evap_g, gat_d, RESERVOIR_CONFIG["gatun"], "Entrada manual", pct_ref_gat))
    if obs_a > 0:
        inputs.append(("Alhajuela / Madden", "Manual instantáneo", pd.to_datetime(ref_date), float(obs_a), unit_obs, evap_a, alh_d, RESERVOIR_CONFIG["alhajuela"], "Entrada manual", pct_ref_alh))

    for embalse, fuente_tipo, fecha_ref, obs_val, obs_unit, evap, daily, cfg, fuente, pct_ref in inputs:
        obs_cfs = obs_val if obs_unit == "cfs" else scalar_to_cfs(obs_val, obs_unit)
        nearest = _nearest_ap_percentile(
            daily, cfg, fecha_ref, obs_cfs, dss_add_cfs=clean_evap_cfs(evap)
        )
        obs_converted = convert_flow(pd.Series([obs_cfs]), flow_unit).iloc[0]
        row = {
            "Embalse": embalse,
            "Fuente": fuente_tipo,
            "Archivo/serie": fuente,
            "Fecha usada": pd.to_datetime(fecha_ref).strftime("%d-%m-%Y"),
            f"Aporte observado ({unit_label(flow_unit)})": round(float(obs_converted), 3),
            "Aporte observado (p³/s)": round(float(obs_cfs), 3),
            "Evap. sumada al DSS (p³/s)": round(clean_evap_cfs(evap), 3),
        }
        if nearest:
            dss_converted = convert_flow(pd.Series([nearest["dss_total_cfs"]]), flow_unit).iloc[0]
            diff_converted = convert_flow(pd.Series([nearest["diff_cfs"]]), flow_unit).iloc[0]
            row.update({
                "Percentil AP DSS más cercano": nearest["label"],
                f"AP total DSS ({unit_label(flow_unit)})": round(float(dss_converted), 3),
                f"Obs-DSS ({unit_label(flow_unit)})": round(float(diff_converted), 3),
                "Diferencia Obs-DSS (p³/s)": round(float(nearest["diff_cfs"]), 3),
                "Diferencia Obs-DSS (%)": round(float(nearest.get("diff_pct_dss", np.nan)), 2),
                "Diferencia abs. vs obs (%)": round(float(nearest["rel_pct"]), 2),
                "Dif. relativa (%)": round(float(nearest["rel_pct"]), 2),
                "Estado": nearest["estado"],
                "Fecha DSS usada": nearest["date"].strftime("%d-%m-%Y"),
            })
        else:
            row.update({"Percentil AP DSS más cercano": "—", "Estado": "⚪ Sin AP DSS"})
        comparisons.append(row)

    st.markdown("#### Resultado comparativo")
    if comparisons:
        out = pd.DataFrame(comparisons)
        st.dataframe(out, use_container_width=True, hide_index=True)
        csv = out.to_csv(index=False).encode("utf-8-sig")
        st.download_button("⬇️ Descargar comparación de aporte instantáneo", csv,
                           "aporte_instantaneo_aquarius_dss.csv", "text/csv", key="apinst_dl")

        if PLOTLY_OK:
            plot = out.copy()
            y_obs = f"Aporte observado ({unit_label(flow_unit)})"
            y_dss = f"AP total DSS ({unit_label(flow_unit)})"
            if y_dss in plot.columns:
                long = plot[["Embalse", "Fuente", y_obs, y_dss]].melt(
                    id_vars=["Embalse", "Fuente"], var_name="Serie", value_name=unit_label(flow_unit)
                )
                fig = px.bar(long, x="Embalse", y=unit_label(flow_unit), color="Serie", barmode="group",
                             facet_col="Fuente", title="Aporte observado vs AP total DSS estimado")
                fig.update_layout(height=420, hovermode="x unified")
                st.plotly_chart(fig, use_container_width=True, key="apinst_bar")
    else:
        st.caption("No hay aporte Aquarius/BulkExport disponible y no se ingresó aporte manual mayor que cero.")


# Alias para compatibilidad con versiones anteriores
def tab_radar(dss_bytes: bytes, flow_unit: str) -> None:
    tab_aporte_instantaneo(dss_bytes, flow_unit)



def build_percentile_export_table(
    daily: pd.DataFrame,
    cfg: Dict,
    pct_ref: int,
    flow_unit: str,
    evap_cfs: float = 0.0,
    include_ap_neto: bool = True,
    include_ap_total: bool = True,
    include_evap_column: bool = True,
    ap_distribution_label: str = "DSS original",
    ap_reduction_pct: float = 0.0,
) -> pd.DataFrame:
    """Construye tabla diaria para exportar el percentil elegido por embalse.

    Opciones AP:
    - include_ap_neto: exporta el AP DSS neto del percentil seleccionado.
    - include_ap_total: exporta AP total DSS = AP neto DSS + evaporación.
    - include_evap_column: muestra el caudal evaporado sumado.
    - ap_distribution_label: documenta si el AP diario viene del DSS original
      o de la simulación con hidrograma observado Mayo-Junio.
    - ap_reduction_pct: porcentaje que se descuenta al AP exportable después
      de sumar evaporación. Ejemplo: 20 significa exportar 80% del AP base.
    """
    if daily is None or daily.empty:
        return pd.DataFrame()
    base = daily.copy()
    base["Fecha_dia"] = pd.to_datetime(base["Fecha_dia"], errors="coerce").dt.normalize()
    base = base[base["Fecha_dia"].notna()].sort_values("Fecha_dia")
    token = cfg["token"]

    np_col, np_pct = pick_percentile_column(cols_by_prefix(base, "NP", token), pct_ref)
    hp_col, hp_pct = pick_percentile_column(cols_by_prefix(base, "HP", token), pct_ref)
    v_col, v_pct = pick_percentile_column(cols_by_prefix(base, "V", token), pct_ref)

    ap_cols = cols_by_prefix(base, "AP", token)
    ap_pct_map = ordered_percentile_map_by_value(base, ap_cols)
    ap_col, ap_pct = pick_percentile_column(ap_cols, pct_ref, pct_map=ap_pct_map)

    eg_col, eg_pct = pick_percentile_column(cols_by_prefix(base, "EG", token), pct_ref)
    ep_col, ep_pct = pick_percentile_column(cols_by_prefix(base, "EP", token), pct_ref)

    out = pd.DataFrame({
        "Fecha": base["Fecha_dia"],
        "Embalse": cfg["name"],
        "Percentil solicitado": f"P{pct_ref}",
        "Modo AP diario": ap_distribution_label,
        "Percentiles usados": (
            f"NP P{np_pct if np_pct is not None else '—'} · "
            f"HP P{hp_pct if hp_pct is not None else '—'} · "
            f"AP P{ap_pct if ap_pct is not None else '—'} · "
            f"V P{v_pct if v_pct is not None else '—'} · "
            f"EG P{eg_pct if eg_pct is not None else '—'} · "
            f"EP P{ep_pct if ep_pct is not None else '—'}"
        ),
    })

    out[f"NP P{np_pct if np_pct is not None else pct_ref} (ft PLD)"] = pd.to_numeric(base[np_col], errors="coerce") if np_col else np.nan
    out[f"HP P{hp_pct if hp_pct is not None else pct_ref} (MW)"] = pd.to_numeric(base[hp_col], errors="coerce") if hp_col else np.nan

    # Garantiza que siempre exista al menos una salida de AP cuando hay AP DSS.
    if not include_ap_neto and not include_ap_total:
        include_ap_neto = True

    if ap_col:
        ap_label = f"P{ap_pct if ap_pct is not None else pct_ref}"
        ap_neto_cfs = pd.to_numeric(base[ap_col], errors="coerce")
        ap_total_cfs = ap_total_dss_cfs(ap_neto_cfs, evap_cfs)
        try:
            reduction_pct = float(ap_reduction_pct or 0.0)
        except Exception:
            reduction_pct = 0.0
        reduction_pct = max(0.0, min(95.0, reduction_pct))
        reduction_factor = 1.0 - (reduction_pct / 100.0)

        # El AP exportable se calcula sobre el AP que el usuario pidió exportar.
        # Si está activa la evaporación: AP base = AP neto DSS + evaporación.
        # Si no está activa: AP base = AP neto DSS.
        ap_base_export_cfs = ap_total_cfs if include_ap_total else ap_neto_cfs
        ap_export_cfs = ap_base_export_cfs * reduction_factor
        ap_reduction_cfs = ap_base_export_cfs - ap_export_cfs

        if include_ap_neto:
            out[f"AP neto DSS {ap_label} (p³/s)"] = ap_neto_cfs
            if flow_unit != "cfs":
                out[f"AP neto DSS {ap_label} ({unit_label(flow_unit)})"] = convert_flow(ap_neto_cfs, flow_unit)
        if include_evap_column and include_ap_total:
            evap_series_cfs = pd.Series([clean_evap_cfs(evap_cfs)] * len(base), index=base.index)
            out[f"Evaporación sumada al AP DSS (p³/s)"] = evap_series_cfs
            if flow_unit != "cfs":
                out[f"Evaporación sumada al AP DSS ({unit_label(flow_unit)})"] = convert_flow(evap_series_cfs, flow_unit)
        if include_ap_total:
            out[f"AP total DSS {ap_label} = AP neto + evaporación antes de reducción (p³/s)"] = ap_total_cfs
            if flow_unit != "cfs":
                out[f"AP total DSS {ap_label} = AP neto + evaporación antes de reducción ({unit_label(flow_unit)})"] = convert_flow(ap_total_cfs, flow_unit)
        base_label = "AP neto + evaporación" if include_ap_total else "AP neto DSS"
        if reduction_pct > 0:
            out["Reducción aplicada al AP exportado (%)"] = reduction_pct
            out[f"Descuento AP exportado {reduction_pct:g}% (p³/s)"] = ap_reduction_cfs
            if flow_unit != "cfs":
                out[f"Descuento AP exportado {reduction_pct:g}% ({unit_label(flow_unit)})"] = convert_flow(ap_reduction_cfs, flow_unit)

        # Columna final obligatoria de exportación: aquí queda aplicado el porcentaje.
        # Fórmula operativa: AP exportable final = (AP neto DSS + evaporación, si está activa) × (1 - reducción/100).
        final_suffix = f" - {reduction_pct:g}%" if reduction_pct > 0 else " sin reducción"
        out[f"AP exportable final {ap_label} = ({base_label}){final_suffix} (p³/s)"] = ap_export_cfs
        if flow_unit != "cfs":
            out[f"AP exportable final {ap_label} = ({base_label}){final_suffix} ({unit_label(flow_unit)})"] = convert_flow(ap_export_cfs, flow_unit)
    else:
        out[f"AP DSS P{pct_ref} ({unit_label(flow_unit)})"] = np.nan

    if v_col:
        out[f"Vertido DSS P{v_pct if v_pct is not None else pct_ref} (p³/s)"] = pd.to_numeric(base[v_col], errors="coerce")
        out[f"Vertido DSS P{v_pct if v_pct is not None else pct_ref} ({unit_label(flow_unit)})"] = convert_flow(base[v_col], flow_unit)
    else:
        out[f"Vertido DSS P{pct_ref} ({unit_label(flow_unit)})"] = np.nan

    if eg_col:
        out[f"EG esclusajes P{eg_pct if eg_pct is not None else pct_ref} (p³/s)"] = pd.to_numeric(base[eg_col], errors="coerce")
        out[f"EG esclusajes P{eg_pct if eg_pct is not None else pct_ref} ({unit_label(flow_unit)})"] = convert_flow(base[eg_col], flow_unit)
    if ep_col:
        out[f"EP esclusajes P{ep_pct if ep_pct is not None else pct_ref} (p³/s)"] = pd.to_numeric(base[ep_col], errors="coerce")
        out[f"EP esclusajes P{ep_pct if ep_pct is not None else pct_ref} ({unit_label(flow_unit)})"] = convert_flow(base[ep_col], flow_unit)
    if eg_col or ep_col:
        parts = []
        if eg_col:
            parts.append(pd.to_numeric(base[eg_col], errors="coerce"))
        if ep_col:
            parts.append(pd.to_numeric(base[ep_col], errors="coerce"))
        total_cfs = pd.concat(parts, axis=1).sum(axis=1, min_count=1)
        out[f"Total esclusajes EG+EP ({unit_label(flow_unit)})"] = convert_flow(total_cfs, flow_unit)

    text_cols = ("Fecha", "Embalse", "Percentil solicitado", "Modo AP diario", "Percentiles usados")
    for c in out.columns:
        if c not in text_cols:
            out[c] = pd.to_numeric(out[c], errors="coerce").round(3)
    return out


def _safe_excel_sheet_name(name: object, fallback: str = "Hoja") -> str:
    """Nombre de hoja Excel seguro: máximo 31 caracteres y sin caracteres inválidos."""
    raw = str(name or fallback).strip() or fallback
    raw = re.sub(r"[\[\]\:\*\?\/\\]", "_", raw)
    raw = re.sub(r"\s+", " ", raw).strip()
    return (raw[:31] or fallback)


def _safe_export_filename(text_value: object, fallback: str = "export") -> str:
    """Nombre de archivo seguro para descargas desde Streamlit."""
    raw = str(text_value or fallback).strip() or fallback
    raw = raw.replace("³", "3").replace("ú", "u").replace("Ú", "U")
    raw = re.sub(r"[^A-Za-z0-9_\-\.]+", "_", raw)
    raw = re.sub(r"_+", "_", raw).strip("._")
    return raw or fallback


def _excel_bytes_from_sheets(sheets: Dict[str, pd.DataFrame]) -> bytes:
    """Genera un XLSX en memoria con una o varias hojas.

    Se usa para evitar fallas de exportación cuando el usuario necesita bajar
    resumen + promedios en un solo archivo. Si una hoja viene vacía, se omite.
    """
    bio = BytesIO()
    wrote = False
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        used_names: set = set()
        for raw_name, df in (sheets or {}).items():
            df = _sanitize_export_dataframe(df)
            if df is None or not isinstance(df, pd.DataFrame) or df.empty:
                continue
            sheet = _safe_excel_sheet_name(raw_name)
            base = sheet
            n = 2
            while sheet in used_names:
                suffix = f"_{n}"
                sheet = _safe_excel_sheet_name(base[: max(1, 31 - len(suffix))] + suffix)
                n += 1
            used_names.add(sheet)
            df.to_excel(writer, index=False, sheet_name=sheet)
            wrote = True
        if not wrote:
            pd.DataFrame({"Mensaje": ["Sin datos para exportar"]}).to_excel(writer, index=False, sheet_name="Sin datos")
    return bio.getvalue()


def _date_only_safe(value):
    """Convierte valores de widgets/fechas a `datetime.date` sin tumbar la app."""
    try:
        if isinstance(value, (list, tuple)):
            value = value[0] if value else None
        if value is None:
            return None
        return pd.to_datetime(value, errors="coerce").date()
    except Exception:
        return None


def _reset_date_input_if_outside(key: str, mn, mx, default_value) -> None:
    """Evita que un date_input conserve una fecha vieja fuera del nuevo rango.

    En Streamlit el estado puede quedar como `date`, `Timestamp`, `datetime`
    o incluso como tupla si antes se usó un rango. Esta función normaliza el
    valor antes de compararlo, para que la pestaña Exportar no se bloquee con
    fechas viejas guardadas en sesión.
    """
    try:
        current = st.session_state.get(key, None)
        if current is None:
            return
        cur_date = _date_only_safe(current)
        mn_date = _date_only_safe(mn)
        mx_date = _date_only_safe(mx)
        def_date = _date_only_safe(default_value)
        if cur_date is None or mn_date is None or mx_date is None:
            st.session_state[key] = default_value
            return
        if cur_date < mn_date or cur_date > mx_date:
            st.session_state[key] = def_date or default_value
    except Exception:
        try:
            st.session_state[key] = default_value
        except Exception:
            pass


def _sanitize_export_dataframe(df: Optional[pd.DataFrame]) -> pd.DataFrame:
    """Prepara una tabla para CSV/Excel sin alterar el cálculo original.

    - Normaliza fechas a datetime sin zona horaria.
    - Reemplaza infinitos por NaN.
    - Convierte objetos no escalares a texto para evitar fallas de openpyxl.
    """
    if df is None or not isinstance(df, pd.DataFrame):
        return pd.DataFrame()
    out = df.copy()
    out = out.replace([np.inf, -np.inf], np.nan)
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            try:
                out[col] = pd.to_datetime(out[col], errors="coerce").dt.tz_localize(None)
            except Exception:
                out[col] = pd.to_datetime(out[col], errors="coerce")
        elif out[col].dtype == "object":
            def _safe_cell(v):
                if isinstance(v, (list, tuple, dict, set)):
                    return str(v)
                return v
            out[col] = out[col].map(_safe_cell)
    return out


def tab_exportar_percentil(
    dss_bytes: bytes,
    flow_unit: str,
    pct_ref_gat: int,
    pct_ref_alh: int,
    evap_gat_cfs: float = 0.0,
    evap_alh_cfs: float = 0.0,
    obs_gat_aporte: Optional[pd.DataFrame] = None,
    obs_alh_aporte: Optional[pd.DataFrame] = None,
) -> None:
    """Pestaña para exportar el percentil elegido por embalse.

    Versión blindada:
    - recalcula la tabla cada vez que cambia embalse, percentil, unidad, modo AP,
      evaporación o reducción;
    - no reutiliza fechas viejas de Streamlit fuera del rango vigente;
    - genera CSV y Excel desde la misma tabla filtrada visible;
    - incluye hoja de metadatos para verificar exactamente qué se exportó.
    """
    st.subheader("⬇️ Exportar percentil DSS")
    st.caption(
        "Exporta el percentil escogido para el embalse seleccionado. "
        "La tabla visible, el CSV y el Excel usan exactamente los mismos datos filtrados."
    )

    # ── Selección principal ─────────────────────────────────────────
    c_sel1, c_sel2 = st.columns([1.20, 1.0])
    with c_sel1:
        embalse_op = st.selectbox(
            "Embalse",
            ["Gatún", "Alhajuela / Madden"],
            key="exp_embalse_v2",
        )
    res_key = "gatun" if embalse_op == "Gatún" else "alhajuela"
    cfg = RESERVOIR_CONFIG[res_key]
    pct_default = int(pct_ref_gat if res_key == "gatun" else pct_ref_alh)

    with c_sel2:
        pct = st.selectbox(
            "Percentil a exportar",
            PERCENTILE_ORDER,
            index=PERCENTILE_ORDER.index(pct_default) if pct_default in PERCENTILE_ORDER else PERCENTILE_ORDER.index(50),
            format_func=lambda x: f"P{x}",
            key=f"exp_pct_v2_{res_key}",
        )

    evap = clean_evap_cfs(evap_gat_cfs if res_key == "gatun" else evap_alh_cfs)
    obs_ap_df = obs_gat_aporte if res_key == "gatun" else obs_alh_aporte

    # ── Opciones AP ─────────────────────────────────────────────────
    st.markdown("#### ⚙️ Opciones de exportación AP")
    c1, c2, c3, c4 = st.columns([1.35, 1.15, 1.05, 1.05])
    with c1:
        simular_hidro_mayo = st.checkbox(
            "🌊 Simular hidrograma observado mayo-junio",
            value=bool(st.session_state.get("ap_hydrograph_enabled", AP_HYDROGRAPH_ADJUSTMENT_ENABLED)),
            key=f"exp_simular_hidro_mayo_v2_{res_key}",
            help=(
                "Redistribuye diariamente el AP DSS usando la forma observada de la última semana de mayo "
                "y las tres semanas siguientes de junio. Conserva la suma/promedio semanal DSS."
            ),
        )
    with c2:
        sumar_evap_ap = st.checkbox(
            "➕ Sumar evaporación al AP",
            value=True,
            key=f"exp_sumar_evap_ap_v2_{res_key}",
            help="AP base exportable = AP neto DSS del percentil + evaporación.",
        )
    with c3:
        incluir_ap_neto = st.checkbox(
            "Mostrar AP neto",
            value=True,
            key=f"exp_incluir_ap_neto_v2_{res_key}",
            help="Incluye una columna con AP neto DSS antes de sumar evaporación.",
        )
    with c4:
        ap_reduction_pct = st.number_input(
            "Restar al AP exportado (%)",
            min_value=0.0,
            max_value=95.0,
            value=float(st.session_state.get(f"exp_ap_reduction_pct_v2_{res_key}", 0.0) or 0.0),
            step=5.0,
            format="%.1f",
            key=f"exp_ap_reduction_pct_v2_{res_key}",
            help="Ejemplo: 20% exporta el 80% del AP base.",
        )

    st.caption(
        "Orden del cálculo exportado: AP neto DSS → sumar evaporación si está activa → aplicar reducción porcentual → AP exportable final."
    )

    # ── Carga y construcción de tabla ───────────────────────────────
    try:
        raw = load_dss_sheet(dss_bytes, cfg["sheet"])
        daily = to_daily(raw, cfg)
        if daily is None or daily.empty:
            st.warning("No se pudo construir el diario DSS para exportación.")
            return
        daily = apply_may_hydrograph_ap_adjustment(
            daily,
            cfg,
            obs_ap_df,
            enabled=bool(simular_hidro_mayo),
        )
    except Exception as exc:
        st.error(f"Error cargando DSS para exportación: {exc}")
        with st.expander("Detalle técnico de carga DSS", expanded=False):
            st.exception(exc)
        return

    meta_ap = getattr(daily, "attrs", {}) or {}
    if simular_hidro_mayo and meta_ap.get("ap_may_adjustment"):
        _show_ap_may_adjustment_note(daily, res_key, "exportación")
    elif simular_hidro_mayo:
        st.warning(
            "No se pudo aplicar la simulación con hidrograma observado mayo-junio; "
            f"se usa el AP DSS original. Motivo: {meta_ap.get('reason', 'sin detalle disponible')}."
        )
    else:
        st.info("Modo AP diario: se exporta el AP DSS original, sin redistribución con mayo.")

    include_ap_total = bool(sumar_evap_ap)
    include_evap_column = bool(sumar_evap_ap)
    include_ap_neto = bool(incluir_ap_neto or not include_ap_total)
    if include_ap_total:
        ap_distribution_label = (
            "Hidrograma observado mayo-junio 4 semanas + evaporación"
            if meta_ap.get("ap_may_adjustment")
            else "DSS original + evaporación"
        )
    else:
        ap_distribution_label = (
            "Hidrograma observado mayo-junio 4 semanas"
            if meta_ap.get("ap_may_adjustment")
            else "DSS original"
        )

    try:
        table_all = build_percentile_export_table(
            daily,
            cfg,
            int(pct),
            flow_unit,
            evap,
            include_ap_neto=include_ap_neto,
            include_ap_total=include_ap_total,
            include_evap_column=include_evap_column,
            ap_distribution_label=ap_distribution_label,
            ap_reduction_pct=float(ap_reduction_pct or 0.0),
        )
        table_all = _sanitize_export_dataframe(table_all)
    except Exception as exc:
        st.error(f"No se pudo construir la tabla de exportación: {exc}")
        with st.expander("Detalle técnico de construcción de tabla", expanded=False):
            st.exception(exc)
        return

    if table_all is None or table_all.empty:
        st.warning("No hay datos para exportar con la selección actual.")
        return

    table_all["Fecha"] = pd.to_datetime(table_all["Fecha"], errors="coerce")
    table_all = table_all.dropna(subset=["Fecha"]).sort_values("Fecha")
    try:
        table_all["Fecha"] = table_all["Fecha"].dt.tz_localize(None)
    except Exception:
        pass
    if table_all.empty:
        st.warning("La tabla de exportación no contiene fechas válidas.")
        return

    # ── Filtro de fecha blindado ────────────────────────────────────
    mn_dt = pd.to_datetime(table_all["Fecha"].min()).date()
    mx_dt = pd.to_datetime(table_all["Fecha"].max()).date()
    option_hash = _safe_export_filename(
        f"{res_key}_p{int(pct)}_{flow_unit}_{int(bool(simular_hidro_mayo))}_"
        f"{int(bool(sumar_evap_ap))}_{int(bool(incluir_ap_neto))}_{float(ap_reduction_pct or 0):g}_"
        f"{mn_dt:%Y%m%d}_{mx_dt:%Y%m%d}"
    )

    with st.expander("🗓️ Filtro de período para exportar", expanded=True):
        s_key = f"exp_s_v2_{option_hash}"
        e_key = f"exp_e_v2_{option_hash}"
        _reset_date_input_if_outside(s_key, mn_dt, mx_dt, mn_dt)
        _reset_date_input_if_outside(e_key, mn_dt, mx_dt, mx_dt)

        c1, c2, c3 = st.columns([1, 1, 1.1])
        s = c1.date_input("Desde", value=mn_dt, min_value=mn_dt, max_value=mx_dt, key=s_key)
        e = c2.date_input("Hasta", value=mx_dt, min_value=mn_dt, max_value=mx_dt, key=e_key)
        usar_todo = c3.checkbox(
            "Usar todo el período disponible",
            value=False,
            key=f"exp_usar_todo_v2_{option_hash}",
        )

    s_date = _date_only_safe(s) or mn_dt
    e_date = _date_only_safe(e) or mx_dt
    if usar_todo:
        s_date, e_date = mn_dt, mx_dt

    if s_date > e_date:
        st.warning("Fecha inicial mayor que final. Se intercambió el orden para mantener la exportación activa.")
        s_date, e_date = e_date, s_date

    table = table_all[
        (table_all["Fecha"].dt.date >= s_date)
        & (table_all["Fecha"].dt.date <= e_date)
    ].copy()
    table = _sanitize_export_dataframe(table)

    if table.empty:
        st.warning("El filtro seleccionado no dejó datos para exportar. Active 'Usar todo el período disponible' o ajuste el rango.")
        return

    # ── Resumen de control ──────────────────────────────────────────
    unit_txt = unit_label(flow_unit)
    ap_export_cols = [c for c in table.columns if c.startswith("AP exportable") and f"({unit_txt})" in c]
    ap_total_cols = [c for c in table.columns if c.startswith("AP total DSS") and f"({unit_txt})" in c]
    ap_neto_cols = [c for c in table.columns if c.startswith("AP neto DSS") and f"({unit_txt})" in c]
    evap_cols = [c for c in table.columns if c.startswith("Evaporación sumada") and f"({unit_txt})" in c]
    reduction_cols = [c for c in table.columns if c.startswith("Descuento AP exportado") and f"({unit_txt})" in c]

    st.success(f"Tabla lista para exportar: {len(table):,} días · {s_date:%d-%m-%Y} a {e_date:%d-%m-%Y}.")
    m1, m2, m3, m4 = st.columns(4)
    if ap_neto_cols:
        m1.metric(f"AP neto medio ({unit_txt})", f"{pd.to_numeric(table[ap_neto_cols[0]], errors='coerce').mean():,.3f}")
    else:
        m1.metric("AP neto medio", "—")
    if ap_export_cols:
        m2.metric(f"AP exportable final medio ({unit_txt})", f"{pd.to_numeric(table[ap_export_cols[0]], errors='coerce').mean():,.3f}")
    elif ap_total_cols:
        m2.metric(f"AP total medio ({unit_txt})", f"{pd.to_numeric(table[ap_total_cols[0]], errors='coerce').mean():,.3f}")
    else:
        m2.metric("AP exportable medio", "—")
    if reduction_cols:
        m3.metric(f"Reducción media ({unit_txt})", f"{pd.to_numeric(table[reduction_cols[0]], errors='coerce').mean():,.3f}")
    else:
        m3.metric("Reducción media", "0.000")
    if evap_cols:
        m4.metric(f"Evaporación sumada ({unit_txt})", f"{pd.to_numeric(table[evap_cols[0]], errors='coerce').mean():,.3f}")
    else:
        m4.metric("Evaporación sumada", "—")

    if PLOTLY_OK and (ap_export_cols or ap_total_cols or ap_neto_cols):
        with st.expander("📈 Vista previa del AP exportado", expanded=False):
            fig = go.Figure()
            preview_cols = list(dict.fromkeys(ap_neto_cols[:1] + ap_total_cols[:1] + ap_export_cols[:1]))
            for col in preview_cols:
                fig.add_trace(go.Scatter(
                    x=table["Fecha"],
                    y=pd.to_numeric(table[col], errors="coerce"),
                    mode="lines+markers",
                    name=col,
                ))
            fig.update_layout(**_base_layout("AP exportado del percentil seleccionado", unit_txt, height=420))
            _today_line(fig, table["Fecha"])
            st.plotly_chart(fig, use_container_width=True, key=f"exp_ap_preview_v2_{option_hash}")

    with st.expander("🧾 Diagnóstico de exportación", expanded=False):
        diag = pd.DataFrame([{
            "Embalse": cfg["name"],
            "Hoja DSS": cfg["sheet"],
            "Percentil solicitado": f"P{int(pct)}",
            "Unidad": unit_txt,
            "Fecha inicio": s_date.strftime("%d-%m-%Y"),
            "Fecha fin": e_date.strftime("%d-%m-%Y"),
            "Días exportados": int(len(table)),
            "Simulación hidrograma mayo-junio": bool(simular_hidro_mayo),
            "Hidrograma aplicado": bool(meta_ap.get("ap_may_adjustment")),
            "Modo AP": ap_distribution_label,
            "Evaporación cfs": round(float(evap), 3),
            "Reducción AP exportado (%)": round(float(ap_reduction_pct or 0.0), 3),
        }])
        st.dataframe(diag, use_container_width=True, hide_index=True)
        st.caption(f"Columnas exportadas: {len(table.columns)}")

    st.dataframe(table, use_container_width=True, hide_index=True, height=520)

    # ── Descarga: la tabla visible es la misma que se descarga ───────
    try:
        reduction_tag = f"_menos{float(ap_reduction_pct):g}pct" if float(ap_reduction_pct or 0.0) > 0 else ""
    except Exception:
        reduction_tag = ""
    period_tag = f"_{s_date:%Y%m%d}_{e_date:%Y%m%d}"
    fname_base = _safe_export_filename(f"export_{cfg['token']}_P{int(pct)}{reduction_tag}{period_tag}")
    dl_key = _safe_export_filename(f"{option_hash}_{s_date:%Y%m%d}_{e_date:%Y%m%d}_{len(table)}")

    try:
        csv = table.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "⬇️ Descargar CSV",
            csv,
            f"{fname_base}.csv",
            "text/csv",
            key=f"exp_csv_v2_{dl_key}",
        )
    except Exception as exc:
        st.warning(f"No se pudo generar el CSV de exportación. Detalle: {exc}")

    try:
        metadata = pd.DataFrame([{
            "Archivo": f"{fname_base}.xlsx",
            "Embalse": cfg["name"],
            "Hoja DSS": cfg["sheet"],
            "Percentil solicitado": f"P{int(pct)}",
            "Unidad": unit_txt,
            "Fecha inicio": s_date.strftime("%d-%m-%Y"),
            "Fecha fin": e_date.strftime("%d-%m-%Y"),
            "Días exportados": int(len(table)),
            "AP neto incluido": bool(include_ap_neto),
            "Evaporación sumada": bool(include_ap_total),
            "Evaporación cfs": round(float(evap), 6),
            "Reducción AP exportado (%)": round(float(ap_reduction_pct or 0.0), 6),
            "Modo AP diario": ap_distribution_label,
            "Hidrograma observado aplicado": bool(meta_ap.get("ap_may_adjustment")),
            "Detalle hidrograma": str(meta_ap.get("reason", "ok")),
        }])
        xlsx = _excel_bytes_from_sheets({
            "Exportacion": table,
            "Metadatos": metadata,
        })
        st.download_button(
            "⬇️ Descargar Excel",
            xlsx,
            f"{fname_base}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"exp_xlsx_v2_{dl_key}",
        )
    except Exception as exc:
        st.warning(f"No se pudo generar Excel; use el CSV. Detalle: {exc}")


# ─────────────────────────────────────────────────────────────────────
# CARGA MASIVA DE BulkExports (CSVs locales + upload)
# ─────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────
# CARGA MASIVA DE BulkExports (CSVs locales + upload + URL Aquarius opcional)
# ─────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False, ttl=900)
def fetch_url_bytes_cached(url: str, timeout_seconds: int = 30) -> bytes:
    """Descarga una serie remota de Aquarius como bytes, con timeout y tamaño limitado."""
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "DSS-Simulacion-ACP/1.0"},
        method="GET",
    )
    with urllib.request.urlopen(req, timeout=timeout_seconds) as resp:
        content_type = str(resp.headers.get("Content-Type", ""))
        data = resp.read(25_000_000)  # límite defensivo: 25 MB
    if not data:
        raise ValueError("respuesta vacía")
    # Si Aquarius devuelve una página HTML de login/error, no se intenta parsear como CSV.
    head = data[:500].lower()
    if b"<html" in head or b"<!doctype html" in head:
        raise ValueError(f"Aquarius devolvió HTML en lugar de CSV ({content_type or 'sin content-type'})")
    return data

# ─────────────────────────────────────────────────────────────────────
# CARGA MASIVA DE BulkExports (CSVs locales + upload)
# ─────────────────────────────────────────────────────────────────────
def load_observed_data() -> Tuple[
    Optional[pd.DataFrame],  # obs_gat_nivel
    Optional[pd.DataFrame],  # obs_alh_nivel
    Optional[pd.DataFrame],  # obs_gat_aporte
    Optional[pd.DataFrame],  # obs_alh_aporte
]:
    """
    Carga todos los BulkExport disponibles (locales + subidos).
    Clasifica por embalse y tipo (nivel / aporte).
    Retorna DataFrames diarios con columnas: Fecha_dia, Valor, Fuente
    """
    # Uploads manuales en sidebar
    st.sidebar.header("📂 BulkExport observados (CSV)")
    uploaded = st.sidebar.file_uploader(
        "CSV BulkExport (GAT, MAD, TstCHCP, etc.)",
        type=["csv"], accept_multiple_files=True, key="bulk_up",
    )
    use_local = st.sidebar.checkbox("Usar CSV locales desde carpeta data", value=True, key="bulk_local")
    use_aquarius_url = st.sidebar.checkbox(
        "Intentar cargar nivel Radar@MAD desde URL Aquarius",
        value=True,
        key="bulk_url_mad_radar",
        help="Serie solicitada: Lake-Res elevation.Telem Radar@MAD. Si Aquarius requiere autenticación o no hay red, la app continúa con CSV locales/subidos.",
    )

    sources: List[Tuple[bytes, str]] = []
    remote_errors: List[Dict[str, object]] = []

    if use_local:
        # discover_local_bulk_csvs() prioriza los CSV críticos de aporte/nivel.
        # Se invierte para que, al consolidar por día, los archivos críticos
        # normalizados prevalezcan sobre BulkExport antiguos o copias auxiliares.
        for p in reversed(discover_local_bulk_csvs()[:120]):
            try:
                sources.append((p.read_bytes(), p.name))
            except Exception:
                pass

    if use_aquarius_url:
        for alias, url in AQUARIUS_REQUIRED_SERIES_URLS:
            try:
                sources.append((fetch_url_bytes_cached(url), alias))
            except Exception as exc:
                remote_errors.append({
                    "Archivo": alias,
                    "Embalse": "Alhajuela / Madden",
                    "Variable": "nivel",
                    "Estado": f"URL no cargada: {exc}",
                })

    # Los archivos subidos manualmente prevalecen sobre los locales y la URL.
    for uf in (uploaded or []):
        sources.append((uf.getvalue(), uf.name))

    # Separate containers
    gat_nivel  = []
    alh_nivel  = []
    gat_aporte = []
    alh_aporte = []

    file_info = list(remote_errors)
    for source_order, (fbytes, fname) in enumerate(sources):
        try:
            daily, embalse, variable, serie = read_bulk_csv(fbytes, fname)
        except Exception as exc:
            file_info.append({"Archivo": fname, "Estado": f"Error: {exc}"})
            continue
        if daily.empty:
            file_info.append({"Archivo": fname, "Embalse": embalse, "Variable": variable, "Estado": "Sin datos"})
            continue
        daily = daily.copy()
        daily["_source_order"] = int(source_order)
        file_info.append({
            "Archivo": fname, "Embalse": embalse, "Variable": variable,
            "Serie": serie, "Registros": len(daily),
            "Inicio": daily["Fecha_dia"].min(), "Fin": daily["Fecha_dia"].max(),
        })
        if embalse == "Gatún":
            (gat_nivel if variable == "nivel" else gat_aporte).append(daily)
        elif embalse == "Alhajuela / Madden":
            (alh_nivel if variable == "nivel" else alh_aporte).append(daily)

    if file_info:
        with st.sidebar.expander("📋 Archivos BulkExport leídos", expanded=False):
            st.dataframe(pd.DataFrame(file_info), use_container_width=True, hide_index=True)

    def _concat(lst: List[pd.DataFrame], variable: str) -> Optional[pd.DataFrame]:
        if not lst:
            return None
        out = pd.concat(lst, ignore_index=True)
        out["Fecha_dia"] = pd.to_datetime(out["Fecha_dia"], errors="coerce").dt.normalize()
        out["Valor"] = pd.to_numeric(out["Valor"], errors="coerce")
        out["_source_order"] = pd.to_numeric(out.get("_source_order", 0), errors="coerce").fillna(0).astype(int)
        out = out.dropna(subset=["Fecha_dia", "Valor"]).sort_values(["Fecha_dia", "_source_order"])

        # Niveles: conservar el último valor disponible del día/fuente prioritaria.
        # Aportes: conservar el último valor del día operativo. Esto evita promediar
        # el 21/06 con el sello 22/06 00:00 cuando ese último corresponde al 21/06.
        if variable == "nivel":
            out = out.groupby("Fecha_dia", as_index=False).agg(
                Valor=("Valor", "last"), Fuente=("Fuente", "last")
            )
        else:
            out = out.groupby("Fecha_dia", as_index=False).agg(
                Valor=("Valor", "last"), Fuente=("Fuente", "last")
            )
            out = repair_observed_aporte_gaps(out, enabled=AP_OBS_GAP_REPAIR_ENABLED)
        return out

    return _concat(gat_nivel, "nivel"), _concat(alh_nivel, "nivel"), _concat(gat_aporte, "aporte"), _concat(alh_aporte, "aporte")



# ─────────────────────────────────────────────────────────────────────
# SEMANA OPERATIVA SÁBADO-VIERNES / HP SEMANAL / INSTRUCTIVO
# ─────────────────────────────────────────────────────────────────────
def operational_week_info(date_value, week_start: int = 5) -> Tuple[int, pd.Timestamp, pd.Timestamp]:
    """Semana DSS con día de inicio configurable.

    `week_start` sigue la convención de pandas weekday: lunes=0 ... domingo=6.
    Por defecto se usa 5, es decir **sábado-viernes**, que es la semana
    operativa DSS. Cuando el usuario selecciona lunes-domingo, se usa el mismo
    cálculo para DSS y observado, evitando que las pestañas semanales queden
    desalineadas entre sí.

    Para 2026 con sábado: 30-may al 05-jun = semana 23; desde 06-jun inicia
    semana 24.
    """
    try:
        week_start = int(week_start)
    except Exception:
        week_start = 5
    if week_start not in range(7):
        week_start = 5

    d = pd.to_datetime(date_value, errors="coerce")
    if pd.isna(d):
        return 0, pd.NaT, pd.NaT
    d = d.normalize()
    year_start = pd.Timestamp(year=d.year, month=1, day=1)
    days_to_first = (week_start - year_start.weekday()) % 7
    first_week_start = year_start + pd.Timedelta(days=days_to_first)
    if d < first_week_start:
        return 1, year_start, first_week_start - pd.Timedelta(days=1)
    week = 2 + int((d - first_week_start).days // 7)
    start = first_week_start + pd.Timedelta(days=(week - 2) * 7)
    end = start + pd.Timedelta(days=6)
    return week, start, end


def add_operational_week_columns(df: pd.DataFrame, week_start: int = 5) -> pd.DataFrame:
    if df is None or df.empty or "Fecha_dia" not in df.columns:
        return df
    out = df.copy()
    info = out["Fecha_dia"].apply(lambda v: operational_week_info(v, week_start=week_start))
    out["Semana operativa"] = [x[0] for x in info]
    out["Inicio semana"] = [x[1] for x in info]
    out["Fin semana"] = [x[2] for x in info]
    return out


def _weekly_ap_obs_vs_dss_table(
    dss_bytes: bytes,
    res_key: str,
    flow_unit: str,
    obs_aportes: Optional[pd.DataFrame],
    evap_cfs: float = 0.0,
    week_start: int = 5,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Compara el promedio semanal observado de aportes contra AP DSS.

    La semana usada respeta el día de inicio indicado en week_start
    (lunes=0 ... domingo=6). Por defecto week_start=5 (sábado-viernes),
    igual que la semana operativa del app. El observado y el DSS se agrupan
    siempre con el mismo week_start, por lo que la comparación queda alineada.
    El AP DSS respeta el modo activo del panel lateral:
    - "Ver aporte semanal DSS": usa el AP como viene en el DSS.
    - "Simular hidrograma observado mayo-junio": redistribuye diariamente, pero
      conserva la suma/promedio semanal DSS.
    """
    if dss_bytes is None:
        return pd.DataFrame(), pd.DataFrame()

    cfg = RESERVOIR_CONFIG[res_key]
    raw = load_dss_sheet(dss_bytes, cfg["sheet"])
    daily = to_daily(raw, cfg)
    daily = apply_may_hydrograph_ap_adjustment(daily, cfg, obs_aportes)
    if daily is None or daily.empty:
        return pd.DataFrame(), pd.DataFrame()

    token = cfg["token"]
    ap_cols = cols_by_prefix(daily, "AP", token)
    if not ap_cols:
        return pd.DataFrame(), pd.DataFrame()

    base = daily.copy()
    base["Fecha_dia"] = pd.to_datetime(base["Fecha_dia"], errors="coerce").dt.normalize()
    base = base[base["Fecha_dia"].notna()].sort_values("Fecha_dia")
    base = add_operational_week_columns(base, week_start=week_start)
    base["Año semana"] = pd.to_datetime(base["Inicio semana"], errors="coerce").dt.year
    group_cols = ["Año semana", "Semana operativa", "Inicio semana", "Fin semana"]

    pct_map = ordered_percentile_map_by_value(base, ap_cols)
    evap = clean_evap_cfs(evap_cfs)

    dss_total_cols: Dict[str, str] = {}
    for col in ap_cols:
        pct = int(pct_map.get(col, exceedance_pct(col)))
        safe_col = f"AP_total_DSS_P{pct}_{col}"
        while safe_col in base.columns:
            safe_col += "_x"
        base[safe_col] = ap_total_dss_cfs(base[col], evap)
        dss_total_cols[col] = safe_col

    agg_spec = {safe_col: "mean" for safe_col in dss_total_cols.values()}
    agg_spec["Fecha_dia"] = "count"
    weekly_dss = base.groupby(group_cols, as_index=False).agg(agg_spec)
    weekly_dss = weekly_dss.rename(columns={"Fecha_dia": "Días DSS"})

    # Tabla ancha con todos los percentiles DSS semanales para trazabilidad.
    wide = weekly_dss[group_cols + ["Días DSS"]].copy()
    for col, safe_col in dss_total_cols.items():
        pct = int(pct_map.get(col, exceedance_pct(col)))
        wide[f"AP DSS semanal P{pct} (p³/s)"] = pd.to_numeric(weekly_dss[safe_col], errors="coerce")

    if obs_aportes is None or not isinstance(obs_aportes, pd.DataFrame) or obs_aportes.empty:
        return pd.DataFrame(), wide

    obs = clamp_observed_future_dates(obs_aportes, "Fecha_dia")
    obs = obs.copy()
    obs["Fecha_dia"] = pd.to_datetime(obs["Fecha_dia"], errors="coerce").dt.normalize()
    obs["Valor"] = pd.to_numeric(obs["Valor"], errors="coerce")
    if "Fuente" not in obs.columns:
        obs["Fuente"] = "Aporte observado"
    obs = obs.dropna(subset=["Fecha_dia", "Valor"]).sort_values("Fecha_dia")
    obs = obs[obs["Valor"] > 0].copy()
    if obs.empty:
        return pd.DataFrame(), wide

    obs = add_operational_week_columns(obs, week_start=week_start)
    obs["Año semana"] = pd.to_datetime(obs["Inicio semana"], errors="coerce").dt.year
    weekly_obs = obs.groupby(group_cols, as_index=False).agg(
        Aporte_obs_prom_cfs=("Valor", "mean"),
        Aporte_obs_min_cfs=("Valor", "min"),
        Aporte_obs_max_cfs=("Valor", "max"),
        Dias_obs_validos=("Valor", "count"),
        Fuente=("Fuente", "last"),
    )

    merged = weekly_obs.merge(weekly_dss, on=group_cols, how="left")
    rows: List[Dict[str, object]] = []
    detail_rows: List[Dict[str, object]] = []

    for _, r in merged.iterrows():
        obs_cfs = pd.to_numeric(pd.Series([r.get("Aporte_obs_prom_cfs", np.nan)]), errors="coerce").iloc[0]
        if pd.isna(obs_cfs):
            continue

        candidates: List[Dict[str, object]] = []
        for col, safe_col in dss_total_cols.items():
            dss_cfs = pd.to_numeric(pd.Series([r.get(safe_col, np.nan)]), errors="coerce").iloc[0]
            if pd.isna(dss_cfs):
                continue
            pct = int(pct_map.get(col, exceedance_pct(col)))
            diff_cfs = float(obs_cfs) - float(dss_cfs)
            abs_diff_cfs = abs(diff_cfs)
            rel_signed_obs = pct_diff_obs_minus_dss_vs_obs(float(obs_cfs), float(dss_cfs))
            rel_pct = abs(rel_signed_obs) if pd.notna(rel_signed_obs) else np.nan
            diff_pct_dss = pct_diff_obs_minus_dss(float(obs_cfs), float(dss_cfs))
            item = {
                "column": col,
                "percentile": pct,
                "dss_cfs": float(dss_cfs),
                "diff_cfs": float(diff_cfs),
                "abs_diff_cfs": float(abs_diff_cfs),
                "diff_pct_dss": float(diff_pct_dss),
                "diff_pct_obs": float(rel_signed_obs),
                "rel_pct": float(rel_pct),
            }
            candidates.append(item)
            detail_rows.append({
                "Embalse": cfg["name"],
                "Semana": int(r["Semana operativa"]),
                "Inicio semana": pd.to_datetime(r["Inicio semana"]),
                "Fin semana": pd.to_datetime(r["Fin semana"]),
                "Percentil AP DSS": f"P{pct}",
                "Aporte observado semanal prom. (p³/s)": float(obs_cfs),
                "AP DSS semanal (p³/s)": float(dss_cfs),
                "Obs-DSS (p³/s)": float(diff_cfs),
                "Diferencia Obs-DSS (%)": float(diff_pct_dss),
                "Diferencia abs. vs obs (%)": float(rel_pct),
                "Diferencia abs. (%)": float(rel_pct),
            })

        if not candidates:
            continue

        nearest = min(
            candidates,
            key=lambda x: (
                float(x["abs_diff_cfs"]),
                -int(x["percentile"]),
                float(x["dss_cfs"]),
            ),
        )

        obs_unit = convert_flow(pd.Series([float(obs_cfs)]), flow_unit).iloc[0]
        dss_unit = convert_flow(pd.Series([nearest["dss_cfs"]]), flow_unit).iloc[0]
        diff_unit = convert_flow(pd.Series([nearest["diff_cfs"]]), flow_unit).iloc[0]
        obs_min_unit = convert_flow(pd.Series([r.get("Aporte_obs_min_cfs", np.nan)]), flow_unit).iloc[0]
        obs_max_unit = convert_flow(pd.Series([r.get("Aporte_obs_max_cfs", np.nan)]), flow_unit).iloc[0]
        estado = (
            "🟢 Muy cercano" if float(nearest["rel_pct"]) <= 10
            else ("🟠 Seguimiento" if float(nearest["rel_pct"]) <= 25 else "🔴 Revisar")
        )
        rows.append({
            "Embalse": cfg["name"],
            "Semana": int(r["Semana operativa"]),
            "Inicio semana": pd.to_datetime(r["Inicio semana"]).strftime("%d-%m-%Y"),
            "Fin semana": pd.to_datetime(r["Fin semana"]).strftime("%d-%m-%Y"),
            "Definición semana": "lunes-domingo" if week_start == 0 else "sábado-viernes",
            "Días obs. válidos": int(r.get("Dias_obs_validos", 0)),
            f"Aporte observado prom. ({unit_label(flow_unit)})": round(float(obs_unit), 3),
            f"Obs. mínimo ({unit_label(flow_unit)})": round(float(obs_min_unit), 3) if pd.notna(obs_min_unit) else np.nan,
            f"Obs. máximo ({unit_label(flow_unit)})": round(float(obs_max_unit), 3) if pd.notna(obs_max_unit) else np.nan,
            "Percentil AP DSS más cercano": f"P{int(nearest['percentile'])}",
            f"AP DSS semanal cercano ({unit_label(flow_unit)})": round(float(dss_unit), 3),
            f"Obs-DSS ({unit_label(flow_unit)})": round(float(diff_unit), 3),
            "Diferencia Obs-DSS (%)": round(float(nearest.get("diff_pct_dss", np.nan)), 2),
            "Diferencia abs. vs obs (%)": round(float(nearest["rel_pct"]), 2),
            "Diferencia abs. (%)": round(float(nearest["rel_pct"]), 2),
            "Estado": estado,
            "Días DSS": int(r.get("Días DSS", 0)) if pd.notna(r.get("Días DSS", np.nan)) else 0,
            "Evap. sumada al DSS (p³/s)": round(evap, 3),
            "Fuente obs.": r.get("Fuente", "—"),
            "Modo AP DSS": (
                "Hidrograma mayo conservando semana"
                if bool(getattr(daily, "attrs", {}).get("ap_may_adjustment", False))
                else "Semanal DSS original"
            ),
        })

    summary = pd.DataFrame(rows)
    detail = pd.DataFrame(detail_rows)
    if not summary.empty:
        summary = summary.sort_values(["Embalse", "Semana"]).reset_index(drop=True)
    if not detail.empty:
        detail = detail.sort_values(["Embalse", "Semana", "Percentil AP DSS"]).reset_index(drop=True)
        for col in detail.columns:
            if col not in ("Embalse", "Percentil AP DSS"):
                if "semana" not in col.lower():
                    detail[col] = pd.to_numeric(detail[col], errors="coerce")
    return summary, detail


def tab_aportes_obs_semanal(
    dss_bytes: bytes,
    flow_unit: str,
    obs_gat: Optional[pd.DataFrame],
    obs_alh: Optional[pd.DataFrame],
    evap_gat_cfs: float = 0.0,
    evap_alh_cfs: float = 0.0,
) -> None:
    """Pestaña semanal: aporte observado promedio vs percentil AP DSS cercano."""
    st.subheader("📅 Aporte observado semanal vs AP DSS")

    # Selector de día de inicio de semana, solo para esta pestaña.
    # Por defecto se usa sábado-viernes para mantener consistencia DSS con las
    # demás pestañas semanales. El usuario puede cambiar a lunes-domingo aquí.
    labels = list(WEEK_START_OPTIONS.keys())
    week_start_label = st.radio(
        "Definición de semana (solo en esta pestaña)",
        labels,
        index=labels.index(DEFAULT_WEEK_START_LABEL),
        horizontal=True,
        key="apw_week_start",
        help=(
            "Por defecto usa sábado-viernes, igual que la semana operativa DSS. "
            "Si selecciona lunes-domingo, observado y DSS se agrupan con ese mismo criterio solo en esta pestaña."
        ),
    )
    week_start = int(WEEK_START_OPTIONS[week_start_label])
    rango_txt = "lunes a domingo" if week_start == 0 else "sábado a viernes"

    st.caption(
        f"Promedio semanal observado alineado con la semana {rango_txt}. "
        "Para cada semana se calcula contra todas las curvas AP DSS y se reporta el percentil más cercano "
        "por diferencia absoluta, sin importar si el DSS queda por arriba o por debajo del observado."
    )

    tables: List[pd.DataFrame] = []
    details: List[pd.DataFrame] = []
    for res_key, obs_df, evap in [
        ("gatun", obs_gat, evap_gat_cfs),
        ("alhajuela", obs_alh, evap_alh_cfs),
    ]:
        try:
            summary, detail = _weekly_ap_obs_vs_dss_table(
                dss_bytes=dss_bytes,
                res_key=res_key,
                flow_unit=flow_unit,
                obs_aportes=obs_df,
                evap_cfs=evap,
                week_start=week_start,
            )
            if not summary.empty:
                tables.append(summary)
            if not detail.empty:
                details.append(detail)
        except Exception as exc:
            st.warning(f"{RESERVOIR_CONFIG[res_key]['name']}: no se pudo calcular el resumen semanal — {exc}")

    if not tables:
        st.warning("No hay aportes observados semanales para comparar. Verifique que estén disponibles los CSV `Discharge_AT_GAT_Diario.csv` y `Discharge_AT_ALHA_Diario.csv` en `data/` o súbalos desde la barra lateral.")
        return

    table = pd.concat(tables, ignore_index=True).sort_values(["Embalse", "Semana"])
    table["Fecha gráfica"] = pd.to_datetime(table["Inicio semana"], format="%d-%m-%Y", errors="coerce")
    table["Etiqueta semana"] = (
        "Sem. " + table["Semana"].astype(int).astype(str)
        + " · " + table["Inicio semana"].astype(str)
        + " a " + table["Fin semana"].astype(str)
    )
    detail_table = pd.concat(details, ignore_index=True) if details else pd.DataFrame()
    resumen_embalse = pd.DataFrame()
    promedio_cards_df = pd.DataFrame()

    min_week = int(table["Semana"].min())
    max_week = int(table["Semana"].max())
    default_start = max(23, min_week) if max_week >= 23 else min_week
    c1, c2, c3 = st.columns([0.8, 0.8, 1.1])
    w1 = c1.number_input("Semana inicial", min_value=min_week, max_value=max_week, value=default_start, step=1, key=f"apw_s_{week_start}")
    w2 = c2.number_input("Semana final", min_value=min_week, max_value=max_week, value=max_week, step=1, key=f"apw_e_{week_start}")
    embalse_sel = c3.multiselect(
        "Embalse",
        sorted(table["Embalse"].dropna().unique().tolist()),
        default=sorted(table["Embalse"].dropna().unique().tolist()),
        key="apw_embalse",
    )
    if w1 > w2:
        st.warning("Semana inicial mayor que final. Se muestra todo el período.")
        w1, w2 = min_week, max_week

    show_tbl = table[
        (table["Semana"] >= int(w1))
        & (table["Semana"] <= int(w2))
        & (table["Embalse"].isin(embalse_sel))
    ].copy()

    st.markdown("#### Resumen por semana")
    obs_prom_col = next((c for c in show_tbl.columns if c.startswith("Aporte observado prom.")), None)
    dss_cercano_col = next((c for c in show_tbl.columns if c.startswith("AP DSS semanal cercano")), None)
    obs_min_col = next((c for c in show_tbl.columns if c.startswith("Obs. mínimo")), None)
    obs_max_col = next((c for c in show_tbl.columns if c.startswith("Obs. máximo")), None)
    obs_dss_col = next((c for c in show_tbl.columns if c.startswith("Obs-DSS")), None)

    # Orden operativo solicitado para la tabla principal:
    # observado → DSS cercano → diferencia porcentual → percentil → mínimos/máximos → diferencia en unidad.
    preferred_cols = [
        "Embalse", "Semana", "Inicio semana", "Fin semana", "Definición semana", "Días obs. válidos",
        obs_prom_col, dss_cercano_col, "Diferencia Obs-DSS (%)", "Percentil AP DSS más cercano",
        obs_min_col, obs_max_col, obs_dss_col,
        "Diferencia abs. vs obs (%)", "Diferencia abs. (%)", "Estado",
        "Días DSS", "Evap. sumada al DSS (p³/s)", "Modo AP DSS", "Fuente obs.",
    ]
    preferred_cols = [c for c in preferred_cols if c and c in show_tbl.columns]
    remaining_cols = [
        c for c in show_tbl.columns
        if c not in preferred_cols and c not in {"Fecha gráfica", "Etiqueta semana"}
    ]
    show_display = show_tbl[preferred_cols + remaining_cols].copy()
    st.dataframe(show_display, use_container_width=True, hide_index=True, height=520)

    # Resumen inferior por embalse: muestra el promedio simple del porcentaje
    # semanal y el porcentaje acumulado/ponderado de todo el período filtrado.
    # La diferencia acumulada se calcula contra el AP DSS cercano:
    #     (Σ Obs ponderado - Σ DSS ponderado) / |Σ DSS ponderado| × 100
    # usando los días observados válidos como ponderador para no mezclar semanas incompletas.
    if not show_tbl.empty and obs_prom_col and dss_cercano_col:
        resumen_rows: List[Dict[str, object]] = []
        for embalse, g in show_tbl.sort_values(["Embalse", "Semana"]).groupby("Embalse", sort=True):
            gg = g.copy()
            obs_vals = pd.to_numeric(gg[obs_prom_col], errors="coerce")
            dss_vals = pd.to_numeric(gg[dss_cercano_col], errors="coerce")
            pct_vals = pd.to_numeric(gg.get("Diferencia Obs-DSS (%)", np.nan), errors="coerce")
            diff_vals = pd.to_numeric(gg[obs_dss_col], errors="coerce") if obs_dss_col and obs_dss_col in gg.columns else (obs_vals - dss_vals)
            weights = pd.to_numeric(gg.get("Días obs. válidos", 1), errors="coerce").fillna(1)
            weights = weights.where(weights > 0, 1)

            valid = obs_vals.notna() & dss_vals.notna() & weights.notna()
            if not valid.any():
                continue

            w = weights[valid].astype(float)
            obs_weighted_sum = float((obs_vals[valid].astype(float) * w).sum())
            dss_weighted_sum = float((dss_vals[valid].astype(float) * w).sum())
            total_days = float(w.sum())
            obs_prom_pond = obs_weighted_sum / total_days if total_days > 0 else np.nan
            dss_prom_pond = dss_weighted_sum / total_days if total_days > 0 else np.nan
            diff_prom_pond = obs_prom_pond - dss_prom_pond if pd.notna(obs_prom_pond) and pd.notna(dss_prom_pond) else np.nan
            pct_acum = ((obs_weighted_sum - dss_weighted_sum) / abs(dss_weighted_sum) * 100.0) if abs(dss_weighted_sum) > 1e-9 else np.nan

            resumen_rows.append({
                "Embalse": embalse,
                "Semanas incluidas": f"{int(gg['Semana'].min())}-{int(gg['Semana'].max())}",
                "Cantidad semanas": int(valid.sum()),
                "Días obs. acumulados": int(total_days),
                f"Aporte observado prom. ponderado ({unit_label(flow_unit)})": round(float(obs_prom_pond), 3) if pd.notna(obs_prom_pond) else np.nan,
                f"AP DSS prom. ponderado ({unit_label(flow_unit)})": round(float(dss_prom_pond), 3) if pd.notna(dss_prom_pond) else np.nan,
                f"Obs-DSS prom. ponderado ({unit_label(flow_unit)})": round(float(diff_prom_pond), 3) if pd.notna(diff_prom_pond) else np.nan,
                "Promedio semanal Diferencia Obs-DSS (%)": round(float(pct_vals.mean()), 2) if pct_vals.notna().any() else np.nan,
                "Diferencia acumulada Obs-DSS (%)": round(float(pct_acum), 2) if pd.notna(pct_acum) else np.nan,
                "Obs-DSS acumulado ponderado": round(float((diff_vals[valid].astype(float) * w).sum()), 3) if diff_vals.notna().any() else np.nan,
            })

        if resumen_rows:
            st.markdown("#### Acumulado y promedio de diferencia por embalse")
            resumen_embalse = pd.DataFrame(resumen_rows)
            st.dataframe(resumen_embalse, use_container_width=True, hide_index=True, height=180)

    # Tarjetas adicionales solicitadas: promedio de las semanas que van.
    # Se calcula con las semanas actualmente filtradas y pondera por días
    # observados válidos para no mezclar una semana incompleta con una completa.
    if not show_tbl.empty and obs_prom_col and dss_cercano_col:
        promedio_cards: List[Dict[str, object]] = []
        for embalse, g in show_tbl.sort_values(["Embalse", "Semana"]).groupby("Embalse", sort=True):
            gg = g.copy()
            obs_vals = pd.to_numeric(gg[obs_prom_col], errors="coerce")
            dss_vals = pd.to_numeric(gg[dss_cercano_col], errors="coerce")
            weights = pd.to_numeric(gg.get("Días obs. válidos", 1), errors="coerce").fillna(1)
            weights = weights.where(weights > 0, 1)
            valid = obs_vals.notna() & dss_vals.notna() & weights.notna()
            if not valid.any():
                continue

            w = weights[valid].astype(float)
            obs_weighted_sum = float((obs_vals[valid].astype(float) * w).sum())
            dss_weighted_sum = float((dss_vals[valid].astype(float) * w).sum())
            total_days = float(w.sum())
            obs_prom_pond = obs_weighted_sum / total_days if total_days > 0 else np.nan
            dss_prom_pond = dss_weighted_sum / total_days if total_days > 0 else np.nan
            diff_prom_pond = obs_prom_pond - dss_prom_pond if pd.notna(obs_prom_pond) and pd.notna(dss_prom_pond) else np.nan
            pct_acum = ((obs_weighted_sum - dss_weighted_sum) / abs(dss_weighted_sum) * 100.0) if abs(dss_weighted_sum) > 1e-9 else np.nan
            semanas_validas = sorted(pd.to_numeric(gg.loc[valid, "Semana"], errors="coerce").dropna().astype(int).unique().tolist())
            if not semanas_validas:
                continue
            semanas_txt = f"{semanas_validas[0]}-{semanas_validas[-1]}" if len(semanas_validas) > 1 else str(semanas_validas[0])

            promedio_cards.append({
                "Embalse": embalse,
                "Semanas texto": semanas_txt,
                "Cantidad semanas": int(len(semanas_validas)),
                "Días obs. acumulados": int(total_days),
                "Obs prom. ponderado": obs_prom_pond,
                "DSS prom. ponderado": dss_prom_pond,
                "Obs-DSS prom. ponderado": diff_prom_pond,
                "Diferencia acumulada Obs-DSS (%)": pct_acum,
            })

        if promedio_cards:
            promedio_cards_df = pd.DataFrame(promedio_cards)
            for _c in ["Obs prom. ponderado", "DSS prom. ponderado", "Obs-DSS prom. ponderado", "Diferencia acumulada Obs-DSS (%)"]:
                if _c in promedio_cards_df.columns:
                    promedio_cards_df[_c] = pd.to_numeric(promedio_cards_df[_c], errors="coerce").round(3)
            st.markdown("#### Promedio de las semanas que van")
            avg_cols = st.columns(len(promedio_cards))
            for idx, item in enumerate(promedio_cards):
                obs_txt = (
                    f"{float(item['Obs prom. ponderado']):,.2f} {unit_label(flow_unit)}"
                    if pd.notna(item.get("Obs prom. ponderado")) else "—"
                )
                delta_parts: List[str] = []
                if pd.notna(item.get("Obs-DSS prom. ponderado")):
                    delta_parts.append(f"Obs-DSS {float(item['Obs-DSS prom. ponderado']):+,.2f} {unit_label(flow_unit)}")
                if pd.notna(item.get("Diferencia acumulada Obs-DSS (%)")):
                    delta_parts.append(f"{float(item['Diferencia acumulada Obs-DSS (%)']):+,.2f}%")
                delta_txt = " · ".join(delta_parts) if delta_parts else None
                avg_cols[idx].metric(
                    f"{item['Embalse']} · sem. {item['Semanas texto']} ({int(item['Cantidad semanas'])} sem.)",
                    obs_txt,
                    delta=delta_txt,
                    delta_color="inverse",
                    help=(
                        "Promedio ponderado por los días observados válidos de las semanas filtradas. "
                        f"Cantidad de semanas promediadas: {int(item['Cantidad semanas'])}. "
                        f"Días observados acumulados: {int(item['Días obs. acumulados'])}. "
                        f"AP DSS promedio ponderado: {float(item['DSS prom. ponderado']):,.2f} {unit_label(flow_unit)}."
                    ) if pd.notna(item.get("DSS prom. ponderado")) else None,
                )

    if not show_tbl.empty:
        st.markdown("#### Última semana observada")
        latest = show_tbl.sort_values(["Embalse", "Semana"]).groupby("Embalse", as_index=False).tail(1)
        cols = st.columns(len(latest))
        for idx, (_, r) in enumerate(latest.iterrows()):
            diff_col = next((c for c in show_tbl.columns if c.startswith("Obs-DSS (")), None)
            obs_col = next((c for c in show_tbl.columns if c.startswith("Aporte observado prom.")), None)
            pct_col = "Diferencia Obs-DSS (%)" if "Diferencia Obs-DSS (%)" in show_tbl.columns else None
            pct_txt = f" · {r[pct_col]:+,.2f}%" if pct_col and pd.notna(r[pct_col]) else ""
            cols[idx].metric(
                f"{r['Embalse']} · semana {int(r['Semana'])}",
                str(r["Percentil AP DSS más cercano"]),
                delta=f"{r[diff_col]:+,.2f} {unit_label(flow_unit)}{pct_txt}" if diff_col else None,
                delta_color="inverse",
                help=(
                    f"Aporte observado prom.: {r[obs_col]:,.2f} {unit_label(flow_unit)} · "
                    f"Diferencia Obs-DSS: {r[pct_col]:+,.2f}% respecto al DSS · "
                    f"Diferencia abs. vs obs.: {r['Diferencia abs. (%)']:.2f}%"
                ) if obs_col and pct_col else None,
            )

    csv_source = show_display if "show_display" in locals() else show_tbl
    try:
        csv = csv_source.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "⬇️ Descargar resumen semanal observado vs DSS",
            csv,
            "aporte_observado_semanal_vs_dss.csv",
            "text/csv",
            key=f"apw_dl_{week_start}",
        )
    except Exception as exc:
        st.warning(f"No se pudo generar el CSV del resumen semanal. Detalle: {exc}")

    # Exportación completa del resumen semanal: incluye la tabla semanal,
    # el acumulado por embalse y el promedio de las semanas que van.
    try:
        sheets_export = {
            "Resumen semanal": csv_source,
            "Promedio semanas": promedio_cards_df,
            "Acumulado embalse": resumen_embalse,
        }
        xlsx_week = _excel_bytes_from_sheets(sheets_export)
        st.download_button(
            "⬇️ Descargar Excel resumen + promedios",
            xlsx_week,
            "aporte_observado_semanal_vs_dss_promedios.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"apw_xlsx_{week_start}",
        )
    except Exception as exc:
        st.warning(f"No se pudo generar el Excel semanal; use el CSV. Detalle: {exc}")

    if PLOTLY_OK and not show_tbl.empty:
        obs_col = next((c for c in show_tbl.columns if c.startswith("Aporte observado prom.")), None)
        dss_col = next((c for c in show_tbl.columns if c.startswith("AP DSS semanal cercano")), None)
        if obs_col and dss_col:
            plot_rows = []
            for _, r in show_tbl.iterrows():
                plot_rows.append({
                    "Embalse": r["Embalse"],
                    "Semana": int(r["Semana"]),
                    "Inicio semana": r.get("Inicio semana", "—"),
                    "Fin semana": r.get("Fin semana", "—"),
                    "Fecha gráfica": r.get("Fecha gráfica", pd.NaT),
                    "Serie": "Aporte observado semanal",
                    "Valor": r[obs_col],
                    "Percentil": "Observado",
                    "Diferencia Obs-DSS (%)": r.get("Diferencia Obs-DSS (%)", np.nan),
                })
                plot_rows.append({
                    "Embalse": r["Embalse"],
                    "Semana": int(r["Semana"]),
                    "Inicio semana": r.get("Inicio semana", "—"),
                    "Fin semana": r.get("Fin semana", "—"),
                    "Fecha gráfica": r.get("Fecha gráfica", pd.NaT),
                    "Serie": "AP DSS semanal más cercano",
                    "Valor": r[dss_col],
                    "Percentil": r["Percentil AP DSS más cercano"],
                    "Diferencia Obs-DSS (%)": r.get("Diferencia Obs-DSS (%)", np.nan),
                })
            plot_df = pd.DataFrame(plot_rows)
            fig = px.line(
                plot_df,
                x="Fecha gráfica",
                y="Valor",
                color="Serie",
                line_dash="Embalse",
                markers=True,
                hover_data=["Embalse", "Semana", "Inicio semana", "Fin semana", "Percentil", "Diferencia Obs-DSS (%)"],
                title=f"Aporte observado semanal vs AP DSS más cercano · semana {rango_txt} ({unit_label(flow_unit)})",
            )
            fig.update_layout(height=560, hovermode="x unified")
            fig.update_xaxes(title="Inicio de semana")
            fig.update_yaxes(title=unit_label(flow_unit))
            st.plotly_chart(fig, use_container_width=True, key="apw_plot")

    if not detail_table.empty:
        with st.expander("🔎 Detalle: comparación contra todos los percentiles DSS", expanded=False):
            detail_show = detail_table[
                (detail_table["Semana"] >= int(w1))
                & (detail_table["Semana"] <= int(w2))
                & (detail_table["Embalse"].isin(embalse_sel))
            ].copy()
            # Conversión visual adicional en la unidad activa.
            detail_show[f"Aporte observado semanal prom. ({unit_label(flow_unit)})"] = convert_flow(
                detail_show["Aporte observado semanal prom. (p³/s)"], flow_unit
            ).round(3)
            detail_show[f"AP DSS semanal ({unit_label(flow_unit)})"] = convert_flow(
                detail_show["AP DSS semanal (p³/s)"], flow_unit
            ).round(3)
            detail_show[f"Obs-DSS ({unit_label(flow_unit)})"] = convert_flow(
                detail_show["Obs-DSS (p³/s)"], flow_unit
            ).round(3)
            for c in [
                "Aporte observado semanal prom. (p³/s)", "AP DSS semanal (p³/s)", "Obs-DSS (p³/s)",
                "Diferencia Obs-DSS (%)", "Diferencia abs. vs obs (%)", "Diferencia abs. (%)",
            ]:
                if c in detail_show.columns:
                    detail_show[c] = pd.to_numeric(detail_show[c], errors="coerce").round(3)
            st.dataframe(detail_show, use_container_width=True, hide_index=True, height=420)
            st.download_button(
                "⬇️ Descargar detalle de todos los percentiles",
                detail_show.to_csv(index=False).encode("utf-8-sig"),
                "detalle_aporte_observado_semanal_todos_percentiles.csv",
                "text/csv",
                key="apw_detail_dl",
            )



# ─────────────────────────────────────────────────────────────────────
# AP SEMANAL GRÁFICO / ESCLUSAJES GATÚN
# ─────────────────────────────────────────────────────────────────────
def _week_start_options_for_tab(key: str, default: str = DEFAULT_WEEK_START_LABEL) -> Tuple[str, int, str]:
    """Selector local de semana para pestañas semanales.

    Todas las pestañas semanales usan este mismo selector para evitar que el DSS
    se vea con rangos distintos entre gráficas/tablas. Por defecto se usa
    sábado-viernes, la semana operativa DSS; lunes-domingo queda disponible
    como alternativa local dentro de la pestaña.
    """
    labels = list(WEEK_START_OPTIONS.keys())
    index = labels.index(default) if default in labels else labels.index(DEFAULT_WEEK_START_LABEL)
    label = st.radio(
        "Definición de semana",
        labels,
        index=index,
        horizontal=True,
        key=key,
        help=(
            "Por defecto usa sábado-viernes, igual que la semana operativa DSS. "
            "El cambio aplica solo a esta pestaña y agrupa DSS y observado con el mismo criterio."
        ),
    )
    week_start = int(WEEK_START_OPTIONS[label])
    rango_txt = "lunes a domingo" if week_start == 0 else "sábado a viernes"
    return label, week_start, rango_txt


def _weekly_ap_graph_data(
    dss_bytes: bytes,
    res_key: str,
    flow_unit: str,
    obs_aportes: Optional[pd.DataFrame],
    evap_cfs: float = 0.0,
    week_start: int = 0,
) -> Tuple[pd.DataFrame, List[str], Optional[str]]:
    """Construye AP DSS semanal por percentiles y aporte observado semanal.

    El DSS se promedia semanalmente con la misma definición de semana que el
    observado. El observado también se promedia por semana, evitando mezclar
    valores diarios con curvas semanales.
    """
    if dss_bytes is None:
        return pd.DataFrame(), [], None

    cfg = RESERVOIR_CONFIG[res_key]
    raw = load_dss_sheet(dss_bytes, cfg["sheet"])
    daily = to_daily(raw, cfg)
    daily = apply_may_hydrograph_ap_adjustment(daily, cfg, obs_aportes)
    if daily is None or daily.empty:
        return pd.DataFrame(), [], None

    ap_cols = cols_by_prefix(daily, "AP", cfg["token"])
    if not ap_cols:
        return pd.DataFrame(), [], None

    base = daily.copy()
    base["Fecha_dia"] = pd.to_datetime(base["Fecha_dia"], errors="coerce").dt.normalize()
    base = base[base["Fecha_dia"].notna()].sort_values("Fecha_dia")
    base = add_operational_week_columns(base, week_start=week_start)
    base["Año semana"] = pd.to_datetime(base["Inicio semana"], errors="coerce").dt.year
    group_cols = ["Año semana", "Semana operativa", "Inicio semana", "Fin semana"]

    pct_map = ordered_percentile_map_by_value(base, ap_cols)
    evap = clean_evap_cfs(evap_cfs)
    plot_cols: List[str] = []
    for col in ap_cols:
        pct = int(pct_map.get(col, exceedance_pct(col)))
        label = f"AP DSS semanal P{pct} [{unit_label(flow_unit)}]"
        while label in base.columns:
            label = f"AP DSS semanal P{pct} {col} [{unit_label(flow_unit)}]"
        base[label] = convert_flow(ap_total_dss_cfs(base[col], evap), flow_unit)
        plot_cols.append(label)

    agg_spec = {c: "mean" for c in plot_cols}
    agg_spec["Fecha_dia"] = "count"
    weekly = base.groupby(group_cols, as_index=False).agg(agg_spec)
    weekly = weekly.rename(columns={"Fecha_dia": "Días DSS"})
    weekly["Fecha_dia"] = pd.to_datetime(weekly["Inicio semana"], errors="coerce")
    weekly["Etiqueta semana"] = (
        "Sem. " + weekly["Semana operativa"].astype(int).astype(str)
        + " · " + pd.to_datetime(weekly["Inicio semana"]).dt.strftime("%d-%m")
        + " a " + pd.to_datetime(weekly["Fin semana"]).dt.strftime("%d-%m")
    )

    obs_col = f"Aporte observado semanal [{unit_label(flow_unit)}]"
    if obs_aportes is not None and isinstance(obs_aportes, pd.DataFrame) and not obs_aportes.empty:
        obs = clamp_observed_future_dates(obs_aportes, "Fecha_dia").copy()
        obs["Fecha_dia"] = pd.to_datetime(obs["Fecha_dia"], errors="coerce").dt.normalize()
        obs["Valor"] = pd.to_numeric(obs["Valor"], errors="coerce")
        obs = obs.dropna(subset=["Fecha_dia", "Valor"]).sort_values("Fecha_dia")
        obs = obs[obs["Valor"] > 0].copy()
        if not obs.empty:
            obs = add_operational_week_columns(obs, week_start=week_start)
            obs["Año semana"] = pd.to_datetime(obs["Inicio semana"], errors="coerce").dt.year
            weekly_obs = obs.groupby(group_cols, as_index=False).agg(
                Aporte_obs_prom_cfs=("Valor", "mean"),
                Dias_obs_validos=("Valor", "count"),
                Aporte_obs_min_cfs=("Valor", "min"),
                Aporte_obs_max_cfs=("Valor", "max"),
            )
            weekly_obs[obs_col] = convert_flow(weekly_obs["Aporte_obs_prom_cfs"], flow_unit)
            weekly = weekly.merge(
                weekly_obs[group_cols + [obs_col, "Dias_obs_validos", "Aporte_obs_min_cfs", "Aporte_obs_max_cfs"]],
                on=group_cols,
                how="left",
            )
        else:
            weekly[obs_col] = np.nan
            weekly["Dias_obs_validos"] = 0
    else:
        weekly[obs_col] = np.nan
        weekly["Dias_obs_validos"] = 0

    plot_cols = order_cols_wet_to_dry(plot_cols)
    return weekly.sort_values("Fecha_dia"), plot_cols, obs_col


def tab_ap_semanal_grafico(
    dss_bytes: bytes,
    flow_unit: str,
    obs_gat: Optional[pd.DataFrame],
    obs_alh: Optional[pd.DataFrame],
    evap_gat_cfs: float = 0.0,
    evap_alh_cfs: float = 0.0,
) -> None:
    """Pestaña de gráficos semanales de AP para ambos embalses."""
    st.subheader("📈 AP semanal DSS y observado — ambos embalses")
    _, week_start, rango_txt = _week_start_options_for_tab("apwg_week_start", default=DEFAULT_WEEK_START_LABEL)
    st.caption(
        f"Cada punto representa el **promedio semanal** con semana {rango_txt}. "
        "El aporte observado también se promedia por semana antes de compararlo con las curvas AP DSS."
    )

    for res_key, obs_df, evap in [
        ("gatun", obs_gat, evap_gat_cfs),
        ("alhajuela", obs_alh, evap_alh_cfs),
    ]:
        cfg = RESERVOIR_CONFIG[res_key]
        st.markdown(f"#### {cfg['name']}")
        try:
            weekly, ap_cols, obs_col = _weekly_ap_graph_data(
                dss_bytes=dss_bytes,
                res_key=res_key,
                flow_unit=flow_unit,
                obs_aportes=obs_df,
                evap_cfs=evap,
                week_start=week_start,
            )
        except Exception as exc:
            st.warning(f"{cfg['name']}: no se pudo construir la gráfica semanal — {exc}")
            continue

        if weekly.empty or not ap_cols:
            st.info(f"No hay AP semanal disponible para {cfg['name']}.")
            continue

        obs_for_chart = obs_col if obs_col in weekly.columns and weekly[obs_col].notna().any() else None
        _unit_key = str(flow_unit).replace("³", "3").replace("/", "_").replace(" ", "")
        fan_chart(
            weekly,
            ap_cols,
            f"{cfg['name']} · AP total DSS semanal y aporte observado semanal ({unit_label(flow_unit)})",
            unit_label(flow_unit),
            f"apwg_{res_key}_{week_start}_{_unit_key}",
            obs_col=obs_for_chart,
            obs_label="Aporte observado semanal",
            show_band=True,
        )

        if obs_for_chart:
            last_obs = weekly[weekly[obs_for_chart].notna()].sort_values("Fecha_dia").tail(1)
            if not last_obs.empty:
                r = last_obs.iloc[0]
                st.caption(
                    f"Última semana con observado: semana {int(r['Semana operativa'])} "
                    f"({pd.to_datetime(r['Inicio semana']):%d-%m-%Y} a {pd.to_datetime(r['Fin semana']):%d-%m-%Y}) · "
                    f"promedio observado: {float(r[obs_for_chart]):,.2f} {unit_label(flow_unit)} · "
                    f"días válidos: {int(r.get('Dias_obs_validos', 0) or 0)}."
                )

        with st.expander(f"📋 Tabla semanal — {cfg['name']}", expanded=False):
            show = weekly.copy()
            date_cols = ["Inicio semana", "Fin semana"]
            for c in date_cols:
                if c in show.columns:
                    show[c] = pd.to_datetime(show[c], errors="coerce").dt.strftime("%d-%m-%Y")
            numeric_cols = [c for c in show.columns if c not in ["Fecha_dia", "Etiqueta semana", "Inicio semana", "Fin semana"]]
            for c in numeric_cols:
                show[c] = pd.to_numeric(show[c], errors="coerce").round(3)
            st.dataframe(show.drop(columns=["Fecha_dia"], errors="ignore"), use_container_width=True, hide_index=True, height=430)
            st.download_button(
                f"⬇️ Descargar AP semanal gráfico — {cfg['token']}",
                show.to_csv(index=False).encode("utf-8-sig"),
                f"ap_semanal_grafico_{res_key}.csv",
                "text/csv",
                key=f"apwg_dl_{res_key}",
            )


def _gatun_esclusajes_daily(daily: pd.DataFrame) -> Tuple[pd.DataFrame, List[int]]:
    """Calcula EG, EP y total de esclusajes de Gatún por percentil."""
    if daily is None or daily.empty:
        return pd.DataFrame(), []
    cfg = RESERVOIR_CONFIG["gatun"]
    token = cfg["token"]
    eg_cols = cols_by_prefix(daily, "EG", token)
    ep_cols = cols_by_prefix(daily, "EP", token)
    pcts = sorted(set(exceedance_pct(c) for c in eg_cols + ep_cols))
    if not pcts:
        return pd.DataFrame(), []

    out = daily[["Fecha_dia"]].copy()
    out["Fecha_dia"] = pd.to_datetime(out["Fecha_dia"], errors="coerce").dt.normalize()
    for pct in pcts:
        eg_c = next((c for c in eg_cols if exceedance_pct(c) == pct), None)
        ep_c = next((c for c in ep_cols if exceedance_pct(c) == pct), None)
        eg = pd.to_numeric(daily[eg_c], errors="coerce") if eg_c and eg_c in daily.columns else pd.Series(np.nan, index=daily.index)
        ep = pd.to_numeric(daily[ep_c], errors="coerce") if ep_c and ep_c in daily.columns else pd.Series(np.nan, index=daily.index)
        total = pd.concat([eg, ep], axis=1).sum(axis=1, min_count=1)

        out[f"EG P{pct} (p³/s)"] = eg
        out[f"EP P{pct} (p³/s)"] = ep
        out[f"Total esclusajes P{pct} (p³/s)"] = total
        out[f"Total esclusajes P{pct} (m³/s)"] = convert_flow(total, "m³/s")
        out[f"Total esclusajes P{pct} (hm³/d)"] = convert_flow(total, "hm³/d")
        out[f"EG P{pct} (hm³/d)"] = convert_flow(eg, "hm³/d")
        out[f"EP P{pct} (hm³/d)"] = convert_flow(ep, "hm³/d")
    return out.sort_values("Fecha_dia"), pcts


def _nearest_available_pct(pct: int, available: List[int]) -> Optional[int]:
    if not available:
        return None
    return int(min(available, key=lambda p: (abs(int(p) - int(pct)), -int(p))))


def _esclusajes_weekly_table(esc_daily: pd.DataFrame, pcts: List[int], week_start: int = 5) -> pd.DataFrame:
    if esc_daily is None or esc_daily.empty or not pcts:
        return pd.DataFrame()
    base = add_operational_week_columns(esc_daily, week_start=week_start)
    base["Año semana"] = pd.to_datetime(base["Inicio semana"], errors="coerce").dt.year
    group_cols = ["Año semana", "Semana operativa", "Inicio semana", "Fin semana"]
    rows: List[Dict[str, object]] = []
    for keys, g in base.groupby(group_cols, dropna=False):
        _, semana, inicio, fin = keys
        row: Dict[str, object] = {
            "Semana": int(semana) if pd.notna(semana) else 0,
            "Inicio semana": pd.to_datetime(inicio).strftime("%d-%m-%Y") if pd.notna(inicio) else "—",
            "Fin semana": pd.to_datetime(fin).strftime("%d-%m-%Y") if pd.notna(fin) else "—",
            "Días DSS": int(g["Fecha_dia"].nunique()),
        }
        for pct in pcts:
            hm3d_col = f"Total esclusajes P{pct} (hm³/d)"
            cfs_col = f"Total esclusajes P{pct} (p³/s)"
            m3s_col = f"Total esclusajes P{pct} (m³/s)"
            if hm3d_col not in g.columns:
                continue
            hm3d = pd.to_numeric(g[hm3d_col], errors="coerce")
            row[f"Prom. P{pct} (hm³/d)"] = round(float(hm3d.mean()), 3) if hm3d.notna().any() else np.nan
            row[f"Volumen P{pct} (hm³/semana)"] = round(float(hm3d.sum()), 3) if hm3d.notna().any() else np.nan
            if cfs_col in g.columns:
                cfs = pd.to_numeric(g[cfs_col], errors="coerce")
                row[f"Prom. P{pct} (p³/s)"] = round(float(cfs.mean()), 1) if cfs.notna().any() else np.nan
            if m3s_col in g.columns:
                m3s = pd.to_numeric(g[m3s_col], errors="coerce")
                row[f"Prom. P{pct} (m³/s)"] = round(float(m3s.mean()), 2) if m3s.notna().any() else np.nan
        rows.append(row)
    return pd.DataFrame(rows).sort_values("Semana") if rows else pd.DataFrame()


def tab_esclusajes_gatun(
    dss_bytes: bytes,
    flow_unit: str,
    pct_ref_gat: int,
    obs_gat_aporte: Optional[pd.DataFrame],
    evap_gat_cfs: float = 0.0,
) -> None:
    """Pestaña de consumo de esclusajes de Gatún por percentil."""
    st.subheader("🚢 Esclusajes Gatún — consumo EG + EP")
    st.caption(
        "Muestra el consumo de esclusajes por percentil DSS. La referencia operativa usa el percentil AP "
        "más cercano al aporte observado de Gatún cuando está disponible. El consumo se reporta en hm³/día, "
        "m³/s y p³/s."
    )

    cfg = RESERVOIR_CONFIG["gatun"]
    try:
        raw = load_dss_sheet(dss_bytes, cfg["sheet"])
        daily = to_daily(raw, cfg)
        daily_for_ap = apply_may_hydrograph_ap_adjustment(daily, cfg, obs_gat_aporte)
    except Exception as exc:
        st.error(f"No se pudo cargar Gatún para esclusajes: {exc}")
        return

    if daily is None or daily.empty:
        st.warning("No se pudo construir el diario DSS de Gatún.")
        return

    esc_daily, pcts = _gatun_esclusajes_daily(daily)
    if esc_daily.empty or not pcts:
        st.warning("No se encontraron columnas EG/EP de esclusajes para Gatún en el DSS.")
        return

    # Percentil de referencia según AP observado más cercano.
    indicator_pct = int(pct_ref_gat)
    nearest = None
    if obs_gat_aporte is not None and isinstance(obs_gat_aporte, pd.DataFrame) and not obs_gat_aporte.empty:
        obs = clamp_observed_future_dates(obs_gat_aporte, "Fecha_dia").copy()
        obs["Fecha_dia"] = pd.to_datetime(obs["Fecha_dia"], errors="coerce").dt.normalize()
        obs["Valor"] = pd.to_numeric(obs["Valor"], errors="coerce")
        obs = obs.dropna(subset=["Fecha_dia", "Valor"]).sort_values("Fecha_dia")
        obs = obs[(obs["Valor"] > 0) & (obs["Fecha_dia"] <= today_panama())]
        if not obs.empty:
            last_obs = obs.iloc[-1]
            nearest = _nearest_ap_percentile(
                daily_for_ap,
                cfg,
                pd.to_datetime(last_obs["Fecha_dia"]),
                float(last_obs["Valor"]),
                dss_add_cfs=clean_evap_cfs(evap_gat_cfs),
            )
            if nearest:
                indicator_pct = int(nearest["percentile"])

    esc_ref_pct = _nearest_available_pct(indicator_pct, pcts)
    if esc_ref_pct is None:
        st.warning("No hay percentiles de esclusajes disponibles.")
        return

    # Filtro temporal de la pestaña.
    valid_dates = esc_daily["Fecha_dia"].dropna()
    mn, mx = valid_dates.min().date(), valid_dates.max().date()
    default_s = max(mn, (pd.Timestamp(mx) - pd.Timedelta(days=60)).date())
    c1, c2 = st.columns(2)
    s = c1.date_input("Desde", value=default_s, min_value=mn, max_value=mx, key="escgat_s")
    e = c2.date_input("Hasta", value=mx, min_value=mn, max_value=mx, key="escgat_e")
    if s > e:
        st.warning("Fecha inicial mayor que final. Se muestra todo el período.")
        s, e = mn, mx
    show_daily = esc_daily[(esc_daily["Fecha_dia"].dt.date >= s) & (esc_daily["Fecha_dia"].dt.date <= e)].copy()

    today = today_panama()
    ref_rows = esc_daily[esc_daily["Fecha_dia"] <= today]
    ref_row = ref_rows.iloc[-1] if not ref_rows.empty else esc_daily.iloc[-1]
    ref_date = pd.to_datetime(ref_row["Fecha_dia"])

    total_hm3d = pd.to_numeric(pd.Series([ref_row.get(f"Total esclusajes P{esc_ref_pct} (hm³/d)", np.nan)]), errors="coerce").iloc[0]
    total_m3s = pd.to_numeric(pd.Series([ref_row.get(f"Total esclusajes P{esc_ref_pct} (m³/s)", np.nan)]), errors="coerce").iloc[0]
    total_cfs = pd.to_numeric(pd.Series([ref_row.get(f"Total esclusajes P{esc_ref_pct} (p³/s)", np.nan)]), errors="coerce").iloc[0]
    eg_hm3d = pd.to_numeric(pd.Series([ref_row.get(f"EG P{esc_ref_pct} (hm³/d)", np.nan)]), errors="coerce").iloc[0]
    ep_hm3d = pd.to_numeric(pd.Series([ref_row.get(f"EP P{esc_ref_pct} (hm³/d)", np.nan)]), errors="coerce").iloc[0]

    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("Indicador AP observado", nearest["label"] if nearest else f"P{pct_ref_gat}")
    m2.metric("Percentil esclusajes usado", f"P{esc_ref_pct}")
    m3.metric("Total EG+EP (hm³/d)", f"{total_hm3d:,.3f}" if pd.notna(total_hm3d) else "—")
    m4.metric("Total EG+EP (m³/s)", f"{total_m3s:,.2f}" if pd.notna(total_m3s) else "—")
    m5.metric("EG (hm³/d)", f"{eg_hm3d:,.3f}" if pd.notna(eg_hm3d) else "—")
    m6.metric("EP (hm³/d)", f"{ep_hm3d:,.3f}" if pd.notna(ep_hm3d) else "—")
    st.caption(
        f"Fecha DSS usada para la tarjeta: {ref_date:%d-%m-%Y}. "
        f"Equivalente total: {total_cfs:,.1f} p³/s." if pd.notna(total_cfs) else f"Fecha DSS usada para la tarjeta: {ref_date:%d-%m-%Y}."
    )

    default_pcts: List[int] = []
    for p in [esc_ref_pct, int(pct_ref_gat), 10, 50, 90]:
        if p in pcts and p not in default_pcts:
            default_pcts.append(int(p))
    if not default_pcts:
        default_pcts = pcts[:3]
    c3, c4 = st.columns([1.2, 0.8])
    sel_pcts = c3.multiselect(
        "Percentiles a mostrar",
        pcts,
        default=default_pcts,
        format_func=lambda x: f"P{x}",
        key="escgat_pcts",
    )
    graph_unit = c4.radio(
        "Unidad gráfica",
        ["hm³/d", "m³/s", "cfs"],
        index=0,
        format_func=unit_label,
        horizontal=True,
        key="escgat_graph_unit",
    )

    if PLOTLY_OK and sel_pcts:
        plot_df = show_daily[["Fecha_dia"]].copy()
        plot_cols = []
        for pct in sorted(sel_pcts):
            source_unit = "p³/s" if graph_unit == "cfs" else graph_unit
            src_col = f"Total esclusajes P{pct} ({source_unit})"
            if src_col not in show_daily.columns:
                continue
            new_col = f"Total esclusajes P{pct} [{unit_label(graph_unit)}]"
            plot_df[new_col] = show_daily[src_col]
            plot_cols.append(new_col)
        fan_chart(
            plot_df,
            plot_cols,
            f"Gatún · Consumo diario de esclusajes EG+EP ({unit_label(graph_unit)})",
            unit_label(graph_unit),
            f"escgat_plot_{graph_unit}_{len(sel_pcts)}",
            show_band=False,
        )

    _, week_start, rango_txt = _week_start_options_for_tab("escgat_week_start", default=DEFAULT_WEEK_START_LABEL)
    st.markdown(f"#### Resumen semanal de consumo de esclusajes · semana {rango_txt}")
    weekly = _esclusajes_weekly_table(show_daily, pcts, week_start=week_start)
    if weekly.empty:
        st.info("No hay datos semanales de esclusajes en el período seleccionado.")
    else:
        keep_cols = ["Semana", "Inicio semana", "Fin semana", "Días DSS"]
        for pct in sorted(sel_pcts or pcts):
            keep_cols += [
                f"Prom. P{pct} (hm³/d)",
                f"Volumen P{pct} (hm³/semana)",
                f"Prom. P{pct} (m³/s)",
                f"Prom. P{pct} (p³/s)",
            ]
        keep_cols = [c for c in keep_cols if c in weekly.columns]
        st.dataframe(weekly[keep_cols], use_container_width=True, hide_index=True, height=430)
        st.download_button(
            "⬇️ Descargar resumen semanal de esclusajes",
            weekly.to_csv(index=False).encode("utf-8-sig"),
            "gatun_esclusajes_semanal.csv",
            "text/csv",
            key="escgat_weekly_dl",
        )

    with st.expander("📋 Tabla diaria de esclusajes", expanded=False):
        daily_show = show_daily.copy()
        daily_show["Fecha"] = daily_show["Fecha_dia"].dt.strftime("%d-%m-%Y")
        cols = ["Fecha"]
        for pct in sorted(sel_pcts or pcts):
            for c in [
                f"EG P{pct} (hm³/d)",
                f"EP P{pct} (hm³/d)",
                f"Total esclusajes P{pct} (hm³/d)",
                f"Total esclusajes P{pct} (m³/s)",
                f"Total esclusajes P{pct} (p³/s)",
            ]:
                if c in daily_show.columns:
                    cols.append(c)
        for c in cols:
            if c != "Fecha":
                daily_show[c] = pd.to_numeric(daily_show[c], errors="coerce").round(3)
        st.dataframe(daily_show[cols], use_container_width=True, hide_index=True, height=430)
        st.download_button(
            "⬇️ Descargar detalle diario de esclusajes",
            daily_show[cols].to_csv(index=False).encode("utf-8-sig"),
            "gatun_esclusajes_diario.csv",
            "text/csv",
            key="escgat_daily_dl",
        )


def tab_hp_semanal(dss_bytes: bytes) -> None:
    st.subheader("⚡ Hidrogeneración DSS")
    _, week_start, rango_txt = _week_start_options_for_tab("hpw_week_start", default=DEFAULT_WEEK_START_LABEL)
    st.caption(
        f"Promedio semanal de HP del DSS con semana {rango_txt}. "
        "La misma definición se usa para la tabla, la gráfica y la descarga."
    )

    rows = []
    for res_key, cfg in RESERVOIR_CONFIG.items():
        try:
            raw = load_dss_sheet(dss_bytes, cfg["sheet"])
            daily = to_daily(raw, cfg)
        except Exception as exc:
            st.warning(f"{cfg['name']}: no se pudo cargar HP — {exc}")
            continue
        if daily.empty:
            continue
        hp_cols = cols_by_prefix(daily, "HP", cfg["token"])
        if not hp_cols:
            continue
        daily = add_operational_week_columns(daily, week_start=week_start)
        daily["Año semana"] = pd.to_datetime(daily["Inicio semana"], errors="coerce").dt.year
        group_cols = ["Año semana", "Semana operativa", "Inicio semana", "Fin semana"]
        weekly = daily.groupby(group_cols, as_index=False)[hp_cols].mean()
        for _, r in weekly.iterrows():
            row = {
                "Variable": "Hidrogeneración DSS",
                "Embalse": cfg["name"],
                "Año semana": int(r["Año semana"]) if pd.notna(r["Año semana"]) else np.nan,
                "Semana": int(r["Semana operativa"]),
                "Inicio semana": pd.to_datetime(r["Inicio semana"]).strftime("%d-%m-%Y"),
                "Fin semana": pd.to_datetime(r["Fin semana"]).strftime("%d-%m-%Y"),
                "Definición semana": rango_txt,
            }
            for c in hp_cols:
                row[f"HP P{exceedance_pct(c)} (MW)"] = round(float(r[c]), 3) if pd.notna(r[c]) else np.nan
            rows.append(row)

    if not rows:
        st.warning("No se encontraron columnas HP en el DSS.")
        return

    df = pd.DataFrame(rows).sort_values(["Embalse", "Año semana", "Semana"])
    df = df[df["Semana"] >= 23].copy()
    if df.empty:
        st.warning("No hay registros de HP desde la semana 23 para la definición de semana seleccionada.")
        return
    min_week = int(df["Semana"].min())
    max_week = int(df["Semana"].max())
    c1, c2 = st.columns(2)
    w1 = c1.number_input("Semana inicial", min_value=min_week, max_value=max_week, value=min_week, step=1, key=f"hpw_s_{week_start}")
    w2 = c2.number_input("Semana final", min_value=min_week, max_value=max_week, value=max_week, step=1, key=f"hpw_e_{week_start}")
    if w1 > w2:
        w1, w2 = min_week, max_week
        st.warning("Semana inicial mayor que final. Se muestra todo el período.")
    show = df[(df["Semana"] >= int(w1)) & (df["Semana"] <= int(w2))].copy()

    st.markdown("#### Tabla semanal de Hidrogeneración DSS")
    st.caption(
        f"La variable **Hidrogeneración DSS** inicia en la semana 23. "
        f"Agrupación activa: **{rango_txt}**."
    )
    st.dataframe(show, use_container_width=True, hide_index=True, height=520)
    csv = show.to_csv(index=False).encode("utf-8-sig")
    st.download_button("⬇️ Descargar CSV HP semanal", csv, "hidrogeneracion_dss_semanal.csv", "text/csv", key=f"hpw_dl_{week_start}")

    # Gráfica simple P50 si existe
    if PLOTLY_OK:
        p50_cols = [c for c in show.columns if c == "HP P50 (MW)"]
        if p50_cols:
            plot = show.copy()
            plot["Fecha_grafica"] = pd.to_datetime(plot["Inicio semana"], format="%d-%m-%Y", errors="coerce")
            plot["Etiqueta semana"] = (
                "Sem. " + plot["Semana"].astype(str)
                + " · " + plot["Inicio semana"].astype(str)
                + " a " + plot["Fin semana"].astype(str)
            )
            fig = px.line(
                plot,
                x="Fecha_grafica",
                y="HP P50 (MW)",
                color="Embalse",
                markers=True,
                hover_data=["Semana", "Inicio semana", "Fin semana", "Definición semana"],
                title=f"HP P50 semanal por embalse · semana {rango_txt}",
            )
            fig.update_layout(height=520, hovermode="x unified")
            fig.update_xaxes(title="Inicio de semana")
            st.plotly_chart(fig, use_container_width=True, key=f"hpw_plot_{week_start}")

def tab_instructivo() -> None:
    st.subheader("📘 Instructivo operativo del dashboard DSS")
    st.markdown("""
### Objetivo general

Este dashboard apoya la interpretación operativa de las simulaciones DSS 2026 para los embalses **Gatún** y **Alhajuela/Madden**. Integra resultados del archivo DSS, datos observados tipo BulkExport/Aquarius, percentiles hidrológicos, niveles proyectados, aportes, vertidos, esclusajes e hidrogeneración DSS.

El propósito no es reemplazar el criterio hidrológico ni la coordinación operativa, sino ofrecer una vista rápida y consistente para responder preguntas como:

- ¿El nivel observado está cercano a la trayectoria simulada?
- ¿Qué percentil DSS representa mejor la condición actual?
- ¿Los aportes observados se parecen más a un escenario seco, medio o húmedo?
- ¿Qué hidrogeneración DSS está asociada al percentil operativo evaluado?
- ¿Cómo se compara el aporte instantáneo/manual con Aquarius y con el DSS?
- ¿Qué información puedo exportar para un percentil específico?

---

### Novedades integradas en esta versión

Esta versión incorpora mejoras puntuales para la lectura operativa:

1. **Percentil más cercano en las tarjetas principales.**  
   Las tarjetas superiores ya no muestran solamente un P50 fijo. Ahora buscan el percentil DSS más cercano al dato observado disponible:
   - **NP cercano:** se calcula con el nivel observado.
   - **AP DSS cercano:** se calcula con el aporte observado de Aquarius/BulkExport, considerando evaporación.
   - **HP cercano o recomendada:** se asocia al percentil operativo más cercano, especialmente al percentil del aporte cuando existe dato observado.

2. **Orden visual húmedo-seco en las gráficas.**  
   Las gráficas y el visor unificado de Plotly se ordenan de **húmedo a seco**: **P5 arriba** y **P95 abajo**. Este criterio aplica a niveles, hidrogeneración, aportes, vertidos y esclusajes cuando las series estén disponibles.

3. **Pestañas semanales con semana DSS unificada.**  
   Las pestañas **AP semanal obs**, **AP semanal gráfico**, **Esclusajes GAT** e **Hidrogeneración DSS** usan el mismo criterio de agrupación. Por defecto trabajan con **sábado-viernes**, que es la semana operativa DSS. Dentro de cada pestaña se puede cambiar a **lunes-domingo**; cuando se cambia, tanto DSS como observado se agrupan con el mismo criterio para evitar diferencias visuales entre gráficas y tablas.

4. **Pestaña Aporte instantáneo.**  
   Se agregó una pestaña para ver el visor meteorológico/radar y comparar:
   - aporte de Aquarius/BulkExport;
   - aporte manual instantáneo ingresado por el usuario;
   - AP total DSS estimado.

5. **Pestaña Exportar.**  
   Se agregó una pestaña para exportar el percentil seleccionado por embalse, incluyendo nivel, hidrogeneración, aportes, vertidos y esclusajes cuando existan en el DSS.

6. **Selectores independientes por embalse.**  
   Gatún y Alhajuela/Madden tienen percentiles de referencia separados. Esto permite evaluar un embalse bajo P95 y el otro bajo P50, por ejemplo, sin mezclar las condiciones.

7. **Comparación histórica 1997-1998 y suma CHCP.**  
   La pestaña **1997-1998 vs actual** realiza el seguimiento exacto desde el **1 de junio hasta la fecha actual disponible**. Compara Gatún, Alhajuela/Madden y **CHCP = GAT + ALHA** con el aporte observado, el percentil DSS diario más cercano y el año histórico seleccionado; presenta una tabla mensual resumida e hidrogramas diarios separados.

---

### Cómo interpretar los percentiles DSS

Los percentiles se interpretan como escenarios de probabilidad de excedencia:

- **P5:** condición más húmeda o de mayor aporte.
- **P50:** condición media o central.
- **P95:** condición más seca o de menor aporte.

En el dashboard se mantiene el siguiente orden visual:

`P5 → P10 → P20 → P30 → P40 → P50 → P60 → P70 → P80 → P90 → P95`

Esto significa que el valor más húmedo debe aparecer arriba en el visor y el valor más seco abajo. Para **AP / Aportes**, la app corrige la etiqueta por magnitud cuando es necesario, usando esta lógica:

`menor AP → P95`  
`mayor AP → P5`

---

### Carga de información

La app puede trabajar con archivos locales o cargados manualmente desde la barra lateral.

**Archivo DSS:**
- Colocar `SimulacionDSS_2026.xlsx` en la carpeta `data`, o cargarlo manualmente desde la barra lateral.
- La app también busca variantes del nombre del archivo DSS si existen copias nuevas.

**Serie histórica 1997-1998:**
- Colocar `1997_1998.xlsx` en la carpeta `data`, junto al app, o cargarlo manualmente desde la barra lateral.
- La app normaliza internamente la fila incompleta `29/Feb` y la continuación de 1998 almacenada con fechas 2026, sin modificar el Excel original.
- La suma CHCP se calcula diariamente como **Gatún + Alhajuela/Madden**.

**Archivos observados:**
- Colocar los `BulkExport*.csv` o CSV normalizados en la carpeta `data`, o subirlos desde la barra lateral.
- La app reconoce niveles y aportes observados para Gatún y Alhajuela/Madden.
- La serie **Lake-Res elevation.Telem Radar@MAD** se usa para el nivel observado de Alhajuela/Madden cuando está disponible.
- Los CSV subidos manualmente tienen prioridad sobre los archivos locales.

Después de reemplazar un Excel o CSV, usar **Recargar archivos** para limpiar caché y actualizar los cálculos.

---

### Barra lateral: controles principales

#### 1. Unidad de caudal / flujo

Permite seleccionar cómo se muestran aportes, vertidos y consumos hidráulicos:

- `p³/s`: pies cúbicos por segundo.
- `m³/s`: metros cúbicos por segundo.
- `hm³/d`: hectómetros cúbicos por día.

Al cambiar esta opción, se actualizan las métricas, gráficas, tablas, comparativos y exportaciones relacionadas con AP, V, EG, EP y aportes observados. Los niveles permanecen en **ft PLD** y la hidrogeneración en **MW**.

#### 2. Percentil de referencia Gatún y Alhajuela/Madden

Estos selectores funcionan como referencia operativa para cada embalse. Se usan principalmente en:

- **Manejo / Decisión**;
- **Comparativo**;
- **Aporte instantáneo**;
- **Exportar**;
- métricas o tablas que requieren un percentil fijo.

Importante: las tarjetas que dicen **percentil cercano** pueden mostrar un percentil diferente al seleccionado porque la app calcula automáticamente el percentil más parecido al dato observado. La referencia operativa para AP se toma por **diferencia absoluta mínima Obs-DSS**, sin importar si la curva DSS queda por arriba o por debajo del observado.

#### 3. Evaporación GAT y Evaporación ALHA

La evaporación puede definirse de dos maneras: **automática**, usando el último valor válido de Corozal (CZL) para Gatún y Pedro Miguel (PMG) para Alhajuela; o **manual**, ingresando el caudal en `p³/s`.

En modo automático se usa: `hm³/día = mm/día × área (km²) × 0.85 × 0.001`, con áreas de referencia de 425 km² para Gatún y 49 km² para Alhajuela. Luego el volumen diario se convierte a `p³/s`.

`AP total DSS estimado = AP neto DSS + caudal evaporado`

La app está blindada para **sumar** la evaporación al AP neto DSS y no restarla.

Adicionalmente, cuando existen aportes observados de mayo, la app utiliza la forma del hidrograma de los últimos días válidos de mayo para redistribuir los AP diarios del DSS dentro de cada semana operativa. Este ajuste **no cambia la suma ni el promedio semanal DSS**; solo modifica la forma diaria para que el comportamiento del hidrograma sea más consistente.

---

### Pestaña GATÚN DSS

Permite revisar Gatún en detalle:

- nivel proyectado DSS vs nivel observado;
- hidrogeneración DSS por percentiles;
- AP total DSS estimado y aporte observado;
- vertidos DSS;
- consumos por esclusajes EG + EP, cuando estén disponibles;
- tabla diaria descargable.

Las tarjetas superiores muestran el percentil más cercano al dato observado disponible. La gráfica mantiene el orden visual **P5 húmedo arriba** y **P95 seco abajo**. En aportes, el eje Y usa el rango completo de las curvas visibles para no ocultar picos DSS; por eso el gráfico diario y el semanal conservan la misma escala física de los valores mostrados.

---

### Pestaña ALHAJUELA DSS

Permite revisar Alhajuela/Madden en detalle:

- nivel proyectado DSS vs nivel observado;
- hidrogeneración DSS;
- AP total DSS estimado y aporte observado;
- vertidos DSS;
- tabla diaria descargable.

La serie Radar@MAD se utiliza como referencia observada de nivel cuando está disponible. Si no se encuentra, la app continúa funcionando con los datos DSS y los CSV locales/subidos disponibles.

---

### Pestaña Manejo / Decisión

Esta pestaña resume el estado ejecutivo por embalse y debe revisarse primero. Incluye:

- **Estado:** semáforo de comparación entre nivel observado y nivel DSS.
- **Obs. LKH:** último nivel observado disponible.
- **Fecha obs.:** fecha del último dato observado.
- **Percentil referencia:** percentil seleccionado en la barra lateral.
- **Nivel DSS usado:** nivel simulado asociado al percentil operativo usado.
- **Aporte DSS usado:** AP total DSS estimado en la unidad seleccionada.
- **Vertido DSS usado:** vertido DSS en la unidad seleccionada.
- **HP DSS usada:** hidrogeneración DSS asociada.
- **Series DSS usadas:** trazabilidad de los percentiles usados para NP, AP, V y HP.
- Promedios del horizonte seleccionado para AP, V y HP.
- Cambio esperado de nivel dentro del horizonte seleccionado.

El semáforo usa umbrales operativos por embalse:

- **Gatún:** 0.10 ft.
- **Alhajuela/Madden:** 0.60 ft.

La lectura recomendada es comparar el semáforo con el percentil cercano al observado y con el percentil de referencia seleccionado.

---

### Pestañas Aporte GAT obs y Aporte ALHA obs

Estas pestañas comparan el aporte observado de Aquarius/BulkExport con el AP total DSS estimado.

Interpretación rápida:

- Si el observado está cerca de **P95**, la condición se parece a un escenario seco.
- Si el observado está cerca de **P50**, la condición se parece a un escenario medio.
- Si el observado está cerca de **P5**, la condición se parece a un escenario húmedo.

La comparación usa:

`AP total DSS estimado = AP neto DSS + evaporación`

---

### Pestaña Comparativo

Permite revisar Gatún y Alhajuela/Madden lado a lado. Es útil para ver si ambos embalses están respondiendo de forma consistente o si uno presenta mayor desviación respecto al DSS.

El comparativo mantiene el orden visual de percentiles:

`P5 húmedo → P95 seco`

Cuando se selecciona AP, también puede mostrar el aporte observado de cada embalse si los CSV están disponibles.

---

### Pestañas semanales DSS

Las pestañas semanales ahora usan una misma lógica de agrupación para evitar que las semanas se vean diferentes entre gráficos:

- **AP semanal obs**
- **AP semanal gráfico**
- **Esclusajes GAT**
- **Hidrogeneración DSS**

Por defecto todas trabajan con **sábado a viernes**, la semana operativa DSS. También se puede seleccionar **lunes a domingo** dentro de cada pestaña. El cambio solo afecta la pestaña activa, pero siempre agrupa **DSS y observado con el mismo criterio**. En las gráficas de AP, la pestaña diaria y la semanal usan la misma lógica de AP total DSS estimado: **AP neto DSS + evaporación** y la misma corrección de etiquetas de excedencia (**P5 húmedo → P95 seco**).

Para 2026 usando sábado-viernes:

- Del 30-may al 05-jun corresponde a la **semana 23**.
- Desde el 06-jun inicia la **semana 24**.

La variable **Hidrogeneración DSS** inicia en la semana 23. La gráfica HP usa como eje X el **inicio de semana**, no solo el número de semana, para que el rango graficado sea claro.

---

### Pestaña Aporte instantáneo

Esta pestaña se usa para revisar una condición rápida de aporte contra el DSS. Incluye dos componentes:

#### 1. Visor de aporte instantáneo

Muestra el visor meteorológico/radar como referencia visual de lluvia o actividad convectiva. La imagen se carga con actualización temporal para evitar que el navegador muestre una versión antigua.

También incluye la opción **Auto-recargar cada 5 min**, útil cuando se desea monitorear el visor durante una situación operativa.

#### 2. Comparación aporte instantáneo / Aquarius / DSS

Permite comparar tres fuentes:

- **Aquarius/BulkExport:** último aporte observado disponible.
- **Manual instantáneo:** valor ingresado manualmente por el usuario.
- **DSS:** AP total DSS estimado para la fecha de referencia.

Pasos recomendados:

1. Seleccionar la **fecha de referencia DSS**.
2. Seleccionar la unidad del aporte manual.
3. Ingresar el aporte instantáneo de Gatún y/o Alhajuela si se desea comparar un valor manual.
4. Confirmar la evaporación aplicada para cada embalse.
5. Revisar la tabla de resultado, donde se muestra el percentil AP DSS más cercano, la diferencia Obs-DSS y el estado.

La pestaña permite descargar la comparación en CSV.

---

### Pestaña Exportar

Esta pestaña permite exportar un percentil específico para un embalse.

Funcionamiento:

1. Seleccionar el embalse: **Gatún** o **Alhajuela/Madden**.
2. Seleccionar el percentil a exportar. Por defecto, la app usa el percentil de referencia escogido en la barra lateral para ese embalse.
3. Definir el período de fechas.
4. Revisar la tabla.
5. Descargar en **CSV** o **Excel**.

La exportación incluye, cuando existan en el DSS:

- nivel NP en ft PLD;
- hidrogeneración HP en MW;
- AP neto DSS;
- evaporación sumada;
- AP total DSS en p³/s y en la unidad seleccionada;
- vertido DSS;
- EG y EP;
- total de esclusajes EG+EP.

La columna **Percentiles usados** muestra trazabilidad, porque en algunos casos la app usa el percentil disponible más cercano si una variable no existe exactamente con el percentil solicitado.

---

### Recomendación de uso operativo

Secuencia recomendada:

1. Cargar o verificar el Excel DSS y los CSV observados.
2. Presionar **Recargar archivos** si se actualizaron datos.
3. Seleccionar la unidad de caudal/flujo.
4. Seleccionar el percentil de referencia de Gatún y de Alhajuela/Madden.
5. Seleccionar la distribución de aportes DSS: **Simular hidrograma observado mayo-junio** o **Ver aporte semanal DSS**.
6. Seleccionar evaporación automática CZL/PMG o ingresar el caudal manual por embalse. En modo automático, el área se calcula con el último nivel observado local disponible.
7. Revisar **Manejo / Decisión**.
8. Validar cada embalse en **GATÚN DSS** y **ALHAJUELA DSS**.
9. Revisar **Aporte GAT obs** y **Aporte ALHA obs** para identificar el percentil hidrológico actual.
10. Revisar **1997-1998 vs actual** para comparar el observado con ambos años históricos, el DSS y la suma CHCP.
11. En las pestañas semanales, mantener **sábado-viernes** para lectura operativa DSS o cambiar a **lunes-domingo** cuando se requiera una comparación calendario. No mezclar criterios al comparar pestañas.
12. Usar **Aporte instantáneo** si se requiere comparar un aporte manual o monitorear el radar.
13. Usar **Exportar** para generar una tabla del percentil operativo evaluado.

La lectura final debe considerar siempre la consistencia entre nivel observado, aporte observado, percentil DSS, vertidos, esclusajes e hidrogeneración.
""")

# ─────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────

def _run_tab(label: str, func, *args, **kwargs) -> None:
    """Ejecuta una pestaña sin tumbar toda la aplicación si algo falla."""
    try:
        func(*args, **kwargs)
    except Exception as exc:
        st.error(f"⚠️ Error en la pestaña **{label}**: {exc}")
        with st.expander("Detalle técnico", expanded=False):
            st.exception(exc)


def main() -> None:
    inject_css()
    view_count = get_view_count()

    show_brand_header(view_count)

    cfg = sidebar()
    dss_bytes = cfg["dss_bytes"]
    hist_9798_bytes = cfg.get("hist_9798_bytes")
    hist_9798_name = str(cfg.get("hist_9798_name", "—"))
    flow_unit = cfg["flow_unit"]
    pct_ref_gat = int(cfg.get("pct_ref_gat", cfg.get("pct_ref", 50)))
    pct_ref_alh = int(cfg.get("pct_ref_alh", cfg.get("pct_ref", 50)))
    evap_gat_cfs = cfg.get("evap_gat_cfs", 0.0)
    evap_alh_cfs = cfg.get("evap_alh_cfs", 0.0)

    if dss_bytes is None:
        st.error(
            "⚠️ No se pudo cargar el archivo DSS. "
            "Coloque `SimulacionDSS_2026.xlsx` en la carpeta `data` o cárguelo desde el panel lateral."
        )
        return
    st.sidebar.success("✅ DSS cargado")

    # Cargar observados sin detener la app si un CSV viene con problema
    try:
        obs_gat_nivel, obs_alh_nivel, obs_gat_aporte, obs_alh_aporte = load_observed_data()
    except Exception as exc:
        st.sidebar.error(f"Error cargando observados: {exc}")
        obs_gat_nivel = obs_alh_nivel = obs_gat_aporte = obs_alh_aporte = None

    tabs = st.tabs([
        "🌊 GATÚN DSS",
        "🏔️ ALHAJUELA DSS",
        "🧭 Manejo / Decisión",
        "🌧️ Aporte GAT obs",
        "🌧️ Aporte ALHA obs",
        "📅 AP semanal obs",
        "📈 AP semanal gráfico",
        "📚 1997-1998 vs actual",
        "🚢 Esclusajes GAT",
        "🔀 Comparativo",
        "⚡ Hidrogeneración DSS",
        "⚡ Aporte instantáneo",
        "⬇️ Exportar",
        "📘 Instructivo",
    ])

    with tabs[0]:
        _run_tab("GATÚN DSS", tab_reservoir, "gatun", dss_bytes, flow_unit, pct_ref_gat, obs_gat_nivel, obs_gat_aporte, evap_gat_cfs)
    with tabs[1]:
        _run_tab("ALHAJUELA DSS", tab_reservoir, "alhajuela", dss_bytes, flow_unit, pct_ref_alh, obs_alh_nivel, obs_alh_aporte, evap_alh_cfs)
    with tabs[2]:
        _run_tab("Manejo / Decisión", tab_manejo, dss_bytes, flow_unit, pct_ref_gat, pct_ref_alh, obs_gat_nivel, obs_alh_nivel, obs_gat_aporte, obs_alh_aporte, evap_gat_cfs, evap_alh_cfs)
    with tabs[3]:
        _run_tab("Aporte GAT obs", tab_aporte_obs_embalse, "gatun", dss_bytes, flow_unit, pct_ref_gat, obs_gat_aporte, evap_gat_cfs)
    with tabs[4]:
        _run_tab("Aporte ALHA obs", tab_aporte_obs_embalse, "alhajuela", dss_bytes, flow_unit, pct_ref_alh, obs_alh_aporte, evap_alh_cfs)
    with tabs[5]:
        _run_tab("AP semanal obs", tab_aportes_obs_semanal, dss_bytes, flow_unit, obs_gat_aporte, obs_alh_aporte, evap_gat_cfs, evap_alh_cfs)
    with tabs[6]:
        _run_tab("AP semanal gráfico", tab_ap_semanal_grafico, dss_bytes, flow_unit, obs_gat_aporte, obs_alh_aporte, evap_gat_cfs, evap_alh_cfs)
    with tabs[7]:
        _run_tab(
            "1997-1998 vs actual", tab_comparacion_9798,
            dss_bytes, hist_9798_bytes, hist_9798_name, flow_unit,
            pct_ref_gat, pct_ref_alh, obs_gat_aporte, obs_alh_aporte,
            evap_gat_cfs, evap_alh_cfs,
        )
    with tabs[8]:
        _run_tab("Esclusajes GAT", tab_esclusajes_gatun, dss_bytes, flow_unit, pct_ref_gat, obs_gat_aporte, evap_gat_cfs)
    with tabs[9]:
        _run_tab("Comparativo", tab_comparativo, dss_bytes, flow_unit, pct_ref_gat, pct_ref_alh, obs_gat_nivel, obs_alh_nivel, obs_gat_aporte, obs_alh_aporte, evap_gat_cfs, evap_alh_cfs)
    with tabs[10]:
        _run_tab("Hidrogeneración DSS", tab_hp_semanal, dss_bytes)
    with tabs[11]:
        _run_tab("Aporte instantáneo", tab_aporte_instantaneo, dss_bytes, flow_unit, pct_ref_gat, pct_ref_alh, obs_gat_aporte, obs_alh_aporte, evap_gat_cfs, evap_alh_cfs)
    with tabs[12]:
        _run_tab("Exportar", tab_exportar_percentil, dss_bytes, flow_unit, pct_ref_gat, pct_ref_alh, evap_gat_cfs, evap_alh_cfs, obs_gat_aporte, obs_alh_aporte)
    with tabs[13]:
        _run_tab("Instructivo", tab_instructivo)

    st.markdown(
        f"<div class='footer'>{SIMULATION_NOTE} · {AUTHOR_NOTE} · Vistas acumuladas: {view_count:,}</div>",
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
