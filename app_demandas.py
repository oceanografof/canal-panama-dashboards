"""
💧 Dashboard Demandas de Agua por Embalse — Canal de Panamá
Creado para HIMH por JFRodriguez
pip install streamlit pandas numpy plotly openpyxl pillow pyxlsb
streamlit run app_demandas.py
"""
import streamlit as st, pandas as pd, numpy as np, datetime, io, base64, os, tempfile, glob
import plotly.graph_objects as go
from plotly.subplots import make_subplots

st.set_page_config(page_title="💧 Demandas — Canal de Panamá", page_icon="💧", layout="wide")


# ── Período LakeHouse seguro y disponible en cualquier sección ───────────────
def _dias_lkh_seguros(default: int = 5) -> int:
    """Devuelve 1, 5, 7, 10 o 30 días sin depender de variables creadas más adelante."""
    try:
        info = globals().get("_info_balance_lkh") or {}
        valor = int(info.get("n_dias", st.session_state.get("dias_op", default) or default))
    except Exception:
        valor = int(default)
    return valor if valor in (1, 5, 7, 10, 30) else int(default)


# Valor seguro disponible desde el inicio para cualquier bloque del dashboard.
# Se actualiza nuevamente cuando se carga el LakeHouse.
_dias_lkh_balance = _dias_lkh_seguros()


# ── Contador consecutivo persistente (una vez por sesión) ─────────────────────
def _counter_state_file() -> str:
    """Ubicación persistente del contador consecutivo del app."""
    base_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in globals() else os.getcwd()
    candidate_dirs = [
        os.path.join(base_dir, ".app_state"),
        os.path.join(os.getcwd(), ".app_state"),
        os.path.join(tempfile.gettempdir(), "app_demandas_state"),
    ]

    for folder in candidate_dirs:
        try:
            os.makedirs(folder, exist_ok=True)
            test_file = os.path.join(folder, ".write_test")
            with open(test_file, "w", encoding="utf-8") as fh:
                fh.write("ok")
            try:
                os.remove(test_file)
            except Exception:
                pass
            return os.path.join(folder, "contador_consecutivo.txt")
        except Exception:
            continue

    return os.path.join(tempfile.gettempdir(), "contador_consecutivo_app_demandas.txt")


def get_consecutive_counter() -> int:
    """Incrementa una sola vez por sesión y conserva el consecutivo entre ejecuciones."""
    session_key = "_contador_consecutivo_app_demandas"
    if session_key in st.session_state:
        return int(st.session_state[session_key])

    counter_file = _counter_state_file()

    try:
        with open(counter_file, "r", encoding="utf-8") as fh:
            raw_value = fh.read().strip()
        current_value = int(raw_value) if raw_value.isdigit() else 0
    except Exception:
        current_value = 0

    next_value = current_value + 1

    try:
        with open(counter_file, "w", encoding="utf-8") as fh:
            fh.write(str(next_value))
    except Exception:
        pass

    st.session_state[session_key] = next_value
    return next_value

# ── Ajuste visual general: métricas más legibles sin alterar la lógica ─────────
st.markdown("""
<style>
/* KPI superiores del dashboard */
div[data-testid="stMetric"] {
    background: rgba(15, 23, 42, 0.025);
    border: 1px solid rgba(148, 163, 184, 0.20);
    border-radius: 14px;
    padding: 0.75rem 0.90rem;
}
div[data-testid="stMetricLabel"] p {
    font-size: 0.95rem !important;
    font-weight: 700 !important;
}
div[data-testid="stMetricValue"] {
    font-size: 2.05rem !important;
    line-height: 1.15 !important;
}
/* Tarjetas internas de Datos Lake House */
.lkh-card {
    border: 1px solid rgba(148, 163, 184, 0.28);
    border-radius: 16px;
    padding: 16px 18px;
    background: linear-gradient(180deg, rgba(255,255,255,0.05), rgba(148,163,184,0.04));
    min-height: 105px;
    box-shadow: 0 1px 5px rgba(15, 23, 42, 0.08);
}
.lkh-card .label {
    font-size: 0.88rem;
    font-weight: 700;
    color: rgba(148, 163, 184, 0.95);
    margin-bottom: 4px;
}
.lkh-card .value {
    font-size: 1.85rem;
    font-weight: 800;
    line-height: 1.15;
    color: inherit;
}
.lkh-card .sub {
    font-size: 0.82rem;
    color: rgba(148, 163, 184, 0.95);
    margin-top: 5px;
}

/* Visor principal limpio: tarjetas por embalse */
.flow-card {
    border: 1px solid rgba(148, 163, 184, 0.28);
    border-radius: 15px;
    padding: 14px 16px;
    min-height: 132px;
    margin-bottom: 12px;
    background: rgba(255,255,255,0.025);
    box-shadow: 0 2px 8px rgba(15, 23, 42, 0.06);
}
.flow-card-alh { border-top: 4px solid #3498db; }
.flow-card-gat { border-top: 4px solid #1a5276; }
.flow-card .flow-label {
    font-size: 0.84rem;
    font-weight: 750;
    color: rgba(100, 116, 139, 0.98);
    margin-bottom: 5px;
}
.flow-card .flow-value {
    font-size: 1.62rem;
    font-weight: 800;
    line-height: 1.12;
    color: inherit;
}
.flow-card .flow-conv {
    font-size: 0.82rem;
    margin-top: 7px;
    color: rgba(100, 116, 139, 0.95);
}
.flow-card .flow-source {
    display: inline-block;
    font-size: 0.72rem;
    margin-top: 7px;
    padding: 2px 7px;
    border-radius: 999px;
    background: rgba(148, 163, 184, 0.14);
    color: rgba(100, 116, 139, 0.98);
}
.flow-card .flow-compare {
    font-size: 0.76rem;
    margin-top: 5px;
    color: rgba(100, 116, 139, 0.95);
}
</style>
""", unsafe_allow_html=True)

# ═══ CONSTANTES ═══
CFS2HM3  = 1 / 408.68          # hm³/d por cada cfs (promedio diario)
CFS2M3S  = 1 / 35.3147         # m³/s por cada cfs
M3S2CFS  = 35.3147
HM3D2M3S = 1e6 / 86400         # m³/s por cada hm³/d
MCF_TO_CFS = 1_000_000.0 / 86400.0  # MCF/MPC por día → cfs
MPC_TO_HM3 = 0.028316846592  # 1 MPC (millón de pies³) → hm³
HM3_TO_MPC = 1 / MPC_TO_HM3   # 1 hm³ → MPC
ACREFT_PER_HM3 = 810.7132
HM3_TO_MGAL = 264.172052358  # 1 hm³ → millones de galones US
MGAL_TO_HM3 = 1 / HM3_TO_MGAL
EVAP_COEF = 0.85             # coeficiente operativo aplicado a lámina evaporada × área
EED_MGALD = 55.0             # 1 EED = 55 millones de galones por día
EED_M3S = EED_MGALD * 3.785 * 1000.0 / (24.0 * 3600.0)
EED_HM3D = EED_M3S / HM3D2M3S
EED_CFS = EED_M3S * M3S2CFS
ZZ_FLUSH_M3S = 333.5          # caudal instantáneo de referencia; 2 h ≈ 2.40 hm³
AHORA    = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

COL = {"alhajuela":"#3498db","gatun":"#1a5276","esclusas":"#2980b9",
       "potable":"#27ae60","fugas":"#e67e22","vertidos":"#9b59b6",
       "generacion":"#f39c12","flush":"#1abc9c","pnx":"#2c3e50",
       "npx":"#16a085","total":"#c0392b","tambor":"#d35400",
       "evap":"#e74c3c","gatgen":"#f1c40f"}

# ── Curvas hipsométricas (nivel ft → área km²) ──────────────────────────────
_NV_GAT  = np.array([55, 60, 65, 70, 75, 79, 82, 84, 85, 86, 87, 88, 89])
_AR_GAT  = np.array([220,255,285,318,350,375,394,408,414,420,425,430,436])

_NV_ALH  = np.array([180,190,200,210,220,228,235,240,245,248,252,255])
_AR_ALH  = np.array([ 10, 14, 19, 25, 31, 36, 40, 43, 46, 47, 49, 51])

# ── Curvas hipsométricas Daily — embebidas (Elevacion_Volumen_Area.xlsx) ─────────
# Gatun:     1,599 pts @ 0.01 ft · rango 77.01-92.99 ft
# Alhajuela: 10,001 pts @ 0.01 ft · rango 160.00-260.00 ft
import zlib as _zlib
_B64_GAT_DAILY = (
    b"eNprLt7n3AzEqyUOg+lmIvmDSc9Qdz8yH8SGYXQ+tcVH7SBO/Mre43hpcuWopWYk2QGiqYn/Jp2h"
    b"upm0wIPVnTB3YaOJFSNVPT4zqGk3rf1BrBxy3ONTr8px0dn23ylnoTenycLI9lGCn9+kDr4UcJYi"
    b"vOc4dfFyu3NUwRO3kodrdM7jpInBaYsJ06TgAKkLOGlysPVE6mFQXiAVC9QTponBv74SpknFT3Mu"
    b"UQ2ff0we3hV1mSBNDF5yETdNCu73uIKTJgdX7qceTjG7Sjb2W4ubJgVbqlzDoMnByrMxaUowr9B1"
    b"quAfHZThR/8xaVLw2bIbOGly8fa31MOLUm6ShXtu46ZJweVBt3DSxODEk4RpcrCPw22yMffSOxj4"
    b"q/RdrOKkqqGFedRWN5j8+2DSXef53+85d+XdR8HYxGDi6HKk8GmlltrmEzIbxgbRJU+p5w8uaPgS"
    b"wifmPCZKHTX0jdo1vOzCl59JzR8gu5DNIQWTkq+xuXkUDzwGAG3LDdQ="
)
_AR_GAT_DAILY = np.frombuffer(
    _zlib.decompress(__import__('base64').b64decode(_B64_GAT_DAILY)),
    dtype='<f4').astype(float).copy()
_NV_GAT_DAILY = np.round(np.arange(1599) * 0.01 + 77.01, 6)
_DAILY_GAT_LOADED = True

_B64_ALH_DAILY = (
    b"eNoMl3dcj+sbx6kUaRmV0SCKSIMopHqXQogGKqKEkpGZktHU3nvvvfee9jg49jy2c4yffXDM3/PH"
    b"9bqv9fl8X9/nee7r/tyyQ0UYLS/CODURJuiIMMVIBO0lIsxeLcICNxEsdomw/LAIqyNE2JAqgkeR"
    b"CHvqRTjUI8Kx8yLE3xIh65kIpR9EaPwtQq+UKBfGinJbQ5Tns0T5ZCrKYGtR5NaKouIhyoz9oiwI"
    b"FGVZrCjrskTZUS7K4RZRYo+LkndZlIb7opx4KcqNL6K8EBPjxwgx5FTFmKwlhuFcMZZbiuFqJ8YB"
    b"FzGidohRcFCM9lAxLiWJ8TxfjF81Yih0iaF9VgzLG2JseCKGzzsx4n+KUSk5hJOKQ3g4eQjf9Yag"
    b"YDKEmcuGYO04BM8tQwjdO4Qi/yH0Rw/hQcYQfpYOYXzzEOYNDMHx0hB87g0h7cUQ2j4P4baoON/k"
    b"xBmvIo7RdHHWG4oTYCFOoa04pzaI83K7ONIHxdELFWd1kjgH88XJrRHnZKc4r86IM+KGOAZPBNw7"
    b"cUJ+ilMtKcE1RQl+TpZAfaYE1iYSHFgmQb6jBOe3SPDvXgkmBEiwNEYC70wJCsokuNgswfcBCaZc"
    b"lmDVfQkCX0pQ90WCv8SGIj1yKPNVh7JNaygZc4dyznIo3+yGMt11KOt2DiXabyi9YUN5nzyUyYVD"
    b"WV03lPDuoXSfE3I3h6L+bCiOH4YS83sox6WG8d/YYehMGcZm/WFkMYyr1sOQXDcMtg7D13sYDUHD"
    b"eBU3jMk5w3CuHEZq2zD+PDmM4VeHYflwGAH/G0b3N4FHQpLZ8pLsVpOkRkeS10aSaFpJ4r5GkuJN"
    b"kjzdLcmko5JsjJKkIF2SxyVC3CSJW79QvyjJ87uSTH0hiednSapFh/NWbjizVIbjPX04HYbD+WUx"
    b"HHO74YS5DOePHcMZ5Tccx7Dh5CYP53nBcGbUDWd/93C6zg1nyK3hLH82nJQPw3nwezia0lLsGSdF"
    b"9xQpJGZLYWsmRc4KKV6sk2K2pxSBB6S4GCzF+AQptuZK0VIlhViHFHanpSi4JsX7R1KYvpUi7ocU"
    b"D4dJo6coTeBkaa7qSaNuIi28T2nOOkqj5C6N1z5pBgKkUYiVxjNLmp5yaUa1SuNxQpruPwX/gTRb"
    b"X0vT9580ihIy7Bwtw6mJMijryOBtJMPFJTJMWSNDwCYZ7uyWQf+oDLFRMrxIl2FhqQx5TTL86JfB"
    b"4ZIMTfdkGPFS4PgiwwUxWaaNlCVMVZbnWrJYzJOleJEsYqtk2bRRlpNesqgfluVYhCz/pMqypFiW"
    b"qgZZZPpk2f2HLNfuyGLwjyyZ/8ryW0SOTXJynFWWQ3u6HEmGcnyzkMPFTo7TLkJupxzJfnL8DJNj"
    b"c4ocfxTKMadejtweOYZekGPPbTnuPZdj0Sc5GgaPQFl2BGFKI/ikOQIXgxH8sXAE82xHULphBPI7"
    b"RhB0cATvQ4Va8gguFYzApG4ENd0jUDk/gthbIxj0fAS7P47g8aCR2MuM5NT4kczVHEnVnJGoLhxJ"
    b"os1IJDaMxG/7SN75jmRz6EjuJI1kZYHQWzsSo+6RNJ4bybRbIyl4NpJxH0eSMGgUkjKjCBo/ih9T"
    b"R+E9ZxTvzEfhaTOKZ+tH4bJ9FHd9R7EmdBRXk0axomAUF2pHsaR7FKfPjWLhrVEMPBsFH0fRN2g0"
    b"JjKj6R0/GmNNYZ0jxAtH02czGjaM5vj20VgcHM3p0NFYJY/mj4LR2NSN5lr3aBzOj+b+rdFsfD6a"
    b"vz+OZvtgeT7IyOOjJM9vTXmOGcgjbSFPsq08yi7yFO+QR9tPntYweUxT5DlXKI99vTx/9ciz9YI8"
    b"n27L4/+3PMP/lSdNRIHJcgrUKytgPF2BC4YKOFkq8I+dAgdcFRD3UiD5kALqEQo0pypgUazAjQYF"
    b"3PsU+PqHAhF3FVB6oUDtZwUQU+T6CEU8VBX5oaVI3DxF1Bcr0rFKkRVuijzbpYjfEUVGRClSlq6I"
    b"SakiN5oU2TmgiPhlRXLuK2LwSpHLXxXxFB+D2Ogx5Ewcg6HOGK4ajcHLagySDmMo2TwGs71jeOA/"
    b"hkMxYxibNYbW8jGsah3DpxNjSLgyBr2HY/jzf2PY/X0MI4aNpV5hLLaTx/JRbyzJJmMxWD6W205j"
    b"OeQxFlXvsQwEjWVL/Fgkc8dSUzUWu46xfD09luzrYzF/MpYX78YS92ssc6TG8dfYcYRMGceM2eO4"
    b"bjaOIyvHMWX9OC5vG8dB33FMDh3HxaRx+BYIft04LnWPw++80HN7HFefj8P/k4AVGc9d2fGEKY/H"
    b"YPp4nhmOJ8lyPOb24/noOp5Cr/HYHR6PWOR4mtPGs6VkPGOaxnOufzyHLo1H5/54Hr8cT8rX8SwR"
    b"V+LnKCUaJirhrqOE0gIlrlgpEeqghPEWJT7tVaIyQAm3WCXGZQu1CiUi25QwP6XEj6tKtDxSwuut"
    b"EtN+KvFUUpncMco4aSijoK/MnygTs0IZK2dlJLYpc8JHmcBjypgmKfMrX5nuWmUOdSsz/7wy328p"
    b"0/lciD8pYySiwk9ZFXqVVQiYroL5XBWGLFLhjL0KURtVWLFLhVFHVLgZqUJWugqupSpoNKvwekCF"
    b"hssq+PylgslrFcS/qXBRQpVkeVXWT1JFQ0+Vt8aqtC1TJdBJlaUeqih4q/IwSJXKeFUO5KpiVq2K"
    b"TKcqd8+oUnpDlX1PhdwHVWQHTeC+9ASqxk/AT3MCVgYTGGsxgRe2E2h3mUDEzgmsPTQBrYgJ/Eqd"
    b"wOXiCRQ0TmB//wQWX5rA+PsTePtyAse/TiBVfCLbR08EtYko6k7k9YKJHF86kXTHiexyn8ii/RNR"
    b"DZrIl7iJXMqZSFnVRPw7JuJ0ZiKzbkxE+ulE/nkvYH5PJEdaDd/xaqzSVEPPQE3Yy2q8tFUT5qga"
    b"xTvVCDqkhmuEGqZpaqiWqPGrUY2/+tXouaQm7Bk1jr5Sw+U/NZCYxCT5SQyZNIm/dSdxzngS1csm"
    b"Ee80iX0ek3D0nsSC4EmoJUxCIm8S/6uexNXOSbSfnUTezUmEPpvEzo+TWD14Msayk9FQnozs9Ml8"
    b"NZzMY8vJnLefTMvGyeTtmkzkkcl4R01mY8ZkrMsmM69lMlNOTGbUlckMfjiZN/+bzL3vAmaYOh2K"
    b"6pSrq5M2S50w1PFZoY6HszqO29Sx8lXHKFQd7WR1JhSqM6penSG96vx3QZ3Xd9R5+I861z+rc1ZM"
    b"g56RGjRN0KBcW4M8Iw1SrDSIdtAgeIsGfvs02BuowbY4DTblaOBcpcHqDg1WntHA6oYGC59qYPxB"
    b"g7mDpjBbZgq6SlPQmjYFTcMpaFhOYZL9FCZsnILKrikoHZnC+KgpjM0QrGwKY1oEOyHYFSF+OIVx"
    b"b4T6jykoS05FdcxUJmpMZbL+VKaYTWXayqlor5/KzO1TMTg4lflhUyFlKpZFU1nWMBXbvqk4XpyK"
    b"y72puL+citfXqXiLa3J0tCahaprE6WqSYaxJ0TJNapw0afPQ5Li3pqA9NLmdoMmzPE3e12jyq0uT"
    b"4ec1GXNbE42/NdH/VxMz0WmsHDGNDarT8JoxjSPzpxGzZBq5a6ZRt3ka/XuncSVgGk9jp/E5expD"
    b"q6YxvmMa2memYXZjGqufTsPzwzSODppOksx0ypWm0zttOtcNp/PKcjoiq6Yz1m06uruns/jodFyi"
    b"p+OTOZ348ulUtE7nxMnp/HV1Ol8fTWfku+nM+DWdJVJabBqnhf9ULbLmaNG+UIsbtlp8dNFihJcW"
    b"Ooe1sI7UYnu6FpGlWlQ2a3H+uBav/tRi+EMttN5osfyHFl6SM4gfM4NGjRnc0J/Bf2YzULKZgemG"
    b"GWzaMYMwvxlUh8/gSuoMvhTPQLlpBuYDM/C8LOD+mkHb6xk8+DYDiWHa6Chq46CuTcAsbSrQ5toK"
    b"bX45azN1uzZ2B7U5EqZNeYo214u0GdyozYx+bZwuaRN6X5vmV9o8+U+bEUN1MFHQYedkHXJm6vCH"
    b"qQ4/rXWY4azD+m06xPrq0Beqw4dkHSYX6bCmQYfwPh26Lurw7p6Qe6WDw386REno0i+vy+dJumjN"
    b"1GWjqS5p1rpcWqeL+DZdFvjqsj9Ul+pkXZ4V6qLSoMuaPl3iLupy7p4uYq90Mf5PF18JPZrk9Xg3"
    b"SU/g0cPDVI9iaz0er9NDdZse63z1yAjV43ayHopFeqxu0COlT48bF/VQuC/Er/RI/U+PWxIzGasw"
    b"k7WTZ5I9cyYPTWcyacVMtjjPpGLbTN74zmRm2Ey8U2bSWTRTeD4zseyfSfSlmVy9P5Nxr2fi+m0m"
    b"ZUNn8U5hFobqs4TnPIuzzGLUylmsWz+Lku1C7eAs5ofP4ljqLP4sniW8r1lsHZhF8+VZiDyYxYr/"
    b"zSLr+yxeDtPHcIw+oRr6XNfXR91cn302+pzYoM/onfpsPqRPS4Q+Eun6OJTqU9Gsz8/j+qy4ok/B"
    b"Q30+v9FnyU99cobP5sPY2SyaOlv4DmfzfqHg280m23U2n7xms/TIbAqiZvNfxmxsymcL3/NsRE7N"
    b"Zu212TQ/no3M+9l4/J7Ncek5KCvNwWfaHK4ZzkF30RyiVs3hH7c5WOyZQ4H/HAbFzmF99hy6KucI"
    b"+2sOvmfmcPvGHAyfzSHt4xz+G2yAo5wBHSoGKM0wEPapAQ+XGGDmYEDxFgOG7jdge5ABl+IN0M8z"
    b"IK3GgJ9dBmw8b8CZ2wZo/2NA0mcDvosZsnGUIWcnGqKna0i6sSGDlxviudaQq1sNWeBjSMkxQ0Yk"
    b"G+JXaMjzekNs+gzpvmiI5n1Dkl8ZIvLNEK+hc7mnMJcl6nNpmzUXDbO5JK+ci9iGuezdMZcnfnOx"
    b"i5jL8bS56JfOpbh5Loon5hJ+ZS7fH85l59u5PPo5F3upeZweN4/5mvOoM5jHZMt5pNvPQ9ZtHiG7"
    b"5/H96Dx2x8zjn6x5bKicx832eVifETA35mHybB5tH+ehJzKfSrn5qKvOJ3fGfMYZzSfZaj5yjvOJ"
    b"cp/PUO/5BAfPZ3DifI7kz+d77Xx8e+bz5cJ8vO/O598X89n3dT4fxY3YK2/Ep0lG7JtpxL+mRniv"
    b"MOKrsxEHtxvx46ARR8ONEEkz4liJEcOajYg5bsTIK0akPjRC6a0R+T+NmCK1gJpxC9DXXECnwQLM"
    b"LBdwzn4Btm4LhLvXAtz8F/A6ZgH7sxfwu3IB4R0LGHV2ATk3F6D5fAHNnxaAqDF/jDDGcYIxz7WN"
    b"2bvAmMHLjIl1MkZ5qzFVB4yZf8yYc0nGOBUa87LemIN9xgy/ZEzWfWO0XxvT+80Y22EmPFU04YCG"
    b"CZKzTcgxN0HP1oRTLiY4eZnw9rAJIVEmjMs0oa7chEVtJtw/ZcK+6yZIPTWh8IMJRoNNuSZryg4V"
    b"U8RnmJI735S5VqZcdTBlp7spw7xNKQo2xTTRlHv5pvjUmSLfa0rDH6asuGfK/16aEvmfKZpD4Ywg"
    b"+d3VBdmhDyVmsMgG/t4AYTtB8zCcj4QdGSBbDg2tsOoUfL0GWU/A9AM8HWRGuKwZ2ipmXNUyw3e+"
    b"GapWZpxyMGO7uxmjvc3oDDbDLdGM4QVmNNWZ4dxrhvhFM+rumeH0ygyxb2bUDjXHSdGcIRrm1Oub"
    b"42xuzjBbc1pczHHzMmfEEXN6o8zZkWnO+ApzzrWZ43vaHM0b5tx+ak74R3PmiSzkldxCslQXYq29"
    b"kEELFtKwdCGbnRYyZutCLhxYiP+xhegnL+SfwoVkNyzEtn8hQy8vpOevhez730Km/1jIY0kL0sda"
    b"YDPVgmEGFvRbWHDQ3oJZbha83m1Bib8FLrEWjM+x4HqVBfGdFiw7Z8HQ2xac+NuCgM8WGA+x5Pso"
    b"S9rVLPHWs2S2qSUfrS1pdLZkz3ZLZvpZ8iHckoY0S/aWWqLfYsnnE5a0XrXk4GNLjN5b8vu3JQMy"
    b"izimvAgrrUXIzF/E1SWLSHNYhLP7IiZ5L+JF8CJqExfhXbCIBfWLEOtbxIWLi0i6L/S8XoTG90W8"
    b"HbaYtjGLCZyymOVzFqNosZjHdoup3rgY392LsfBfzIjYxfyVvZjKKiHXuRjLc4sZfXsxT/5eTMPn"
    b"xQQMWcLK0UuYMGkJ7/SW0G+6hIQVS3Bbv4TZO5YgcWgJtyOWUJW+hKNlS7BtXYLGqSV8u7aEi0+W"
    b"CN/tEnwGW2EtZ8VkVSu+zbDispEVpUutOOJkxeqtVmj7WCEeasWDZCvaiqyIb7Ri24AVFn9aofpQ"
    b"wLyx4vpPK+qklhI1fike05ZiMXcpaouX8nv1Uu5tXkrHvqWkBS3lQMJSVucvZXbdUuHbX8q/fyzl"
    b"+r2ltLxaSuq3pfgMW4bTmGUYTVmG6pxliFgs45ndMs5sXEbV7mXE+S9jf6zQk7MMk+plqHctQ/L8"
    b"Mt7fXsbNf5bR82UZJeLLiZFfjvfk5ayftZxFZsvRtVnOWJfliHot53+Hl3MrajnHM5dTU7Gc9Pbl"
    b"hJxZzp6bQv/z5Sz9dzlzxazRGGXNaDVrRPWseW9izUNray47W9O33Zo6P2vyIqyJT7cmsMyafa3W"
    b"bDlljcN1a6yeWrPgo7Uwb1cwecQKxkxYgbTOCgYbr+DzshW8XruCR54ruOm7gothKziZuoKukhU0"
    b"Nq+g8sQKCq+uIOvxCpLfryBm0ErCZFcSqLKSQzNWcsBoJXuWrmSH00q2bl3JZp+VuIauZH3KStYW"
    b"r8ShaSWrj6/E/spK7B6txPbdSmx+r2SljA0rlW2w0RJsvg22VjbYOdqwysOGNQdscDxmw7pkGzYU"
    b"2eDWaMOWARu2/WmD10Mb9r21wfeXDUekbQlWsiVyui3x82xJW2JLroMtpe621Hrb0hpiS1+SLWcL"
    b"bbnSYMu9flueX7bl/QNbvr+xRfyXLSOl7VBRsmPadDvmzLPDfIkdNg52rHe3Y7u3HQdD7AhPsiOt"
    b"0I7SBjta++04fdmOmw/s+OeNHf/9tGO4tL2gE+zRmW4P8+yxX2KPu4M9fu72xHrbUxhiT2uSPRcK"
    b"7XncYM+Xfnuk/7Rn8kN75r21x+aXPVulVxGgtIr06auon7eKc0tW8cRhFT/dV6F4YBV6x1axNHkV"
    b"W4qEvsZVZA+sov3PVVx/uIoPb1ch+3sVWjKrsVJejYfWakLnr6bYajUnHFfz1GM1Yj6rmRy6moUp"
    b"q9lcLNSbVlN2fDXnrqzm9aPVyL5fzcxBa1gluwZflTVkz1hDv9Eani9dw/C1a9DzXMNq3zUcCltD"
    b"YeoazpWs4X3zGsaeXIPptTVsfbKG+A9r6BjswBM5B6QnOGCg44CrsQNRyx1oXSfktzkg4+fAvAgH"
    b"3NMdSCpzoL/VgTenHFC64YDVMwd8PzlQJurIzZGOSKg5YqDniLupI2krHDm73pFvOxzROuzI+ihH"
    b"4jIdOV7hyOd2RzTPOuJ8y5H4vx059dmR70Oc0JV3YstkJzJnOXHFzImhtk6YuDpxYJcTNUedeB7j"
    b"hGqOE2uqnYjrcuLceSfE7jph/NIJn/+caBy6ljeKa5k2ZS1b5qylwGItf9mvZfymtTjuXUtK4Fqu"
    b"xa9lVP5abOrWEte7lssX1yL711pB0wrxj7X8OXwdI8evw27aOpLnruPm4nWMc1gnzOV15Hmv42nI"
    b"OqYmr2N70TrqGtfx78A65l9Zh/+jdZx+tw7pQc7YyzqTqeLM4xnOTFvgzJ5lznSudUZsmzPLDzqT"
    b"Eu7MwzShVubM/lZn+k45M/yGM2ueOVP4yZm3ousxGrWecLX13NRbz2TWs2flevo3rEfOaz0bjqyn"
    b"Jno9v7LWY121ntzO9bw7tx7urCfxxXqefV2PwdANRChu4L7GBnTnbCDYYgO37DegtWkDAXs3cCNQ"
    b"8BM2EJgv5Os2oN23gZBLG7j31wb032wg6ucGnkq5sEDJhZTpLryd58ISKxcKHV346eGCg48LjaEu"
    b"yKS64FniwqlmFyaddMH/mgv3n7gw76MLaSKufBnhyqqJrjTpujLK1JW9K1y5ut6VWTtdSTrsyr9R"
    b"rqzJcqWt0pVxna4cPufKw9uumL9wpeSrK5JDN7JTcSNXNTZiOGcj2RYbEV21ka2bNnJp70bmBAm5"
    b"hI2IFWxke/1GrvVtxOjyRoofbETm7UZ8fm3kkbQby5TdaNFyY6KRG1FL3fjq5MYmTzcu+7qxINyN"
    b"ijQ3FMvcCGl14+MpN1xvCLVnbpj860aN2CZURm8iZtImfs/chJfZJh7ZbMLWdRMndm1ijv8mymM3"
    b"oZS7idiaTYj2bOLAH5t4dW8TLq83cf37JqyGb6Zv3GbmTNtM1dzNTFqymQyHzYz02EzEgc2IhG7m"
    b"YMpmPhZvZkfzZp6f2IzLtc3cebKZVR83c1lkC0tHbuH0xC2Y622hz3QLRiu30L5hCwZeW2g+soVZ"
    b"MVuoz96CbvUWaru2oHNBWO8K66st1H3bgp6kOw1j3QW97E6LoTuGi93pXOOOsbs7/d7uWBxz52yy"
    b"O9bF7lxpcsfhhDv3r7rj9sSdFx/c2SXiIbxPD45M9EBcz4MYUw8UVnqQu8GDqV4e1B/xYH6MByez"
    b"PVhR7cGdLg82X/Dg/V0PDr3yYNh3D5Ilt6I2biu1mltZMHcr5xdvxdFhK/+4b+XAga2CTthKSspW"
    b"NEq20tK8lUUnt3Lr2lY8n27l+8etRIt6MmGUJ41qnljO9OQOnuyw8UTE1ZOUXZ5M9/ekP9aT1bme"
    b"vK7xJKjHk7EXPam778ni/3ny8IcnPlLbkFPaRvn0bZjN38Y9q23sdxJyntuo8N2GRfg2YX9u41DZ"
    b"Nsa0baPp9DZsbm7jzfNtRH7ehqb4dk7Lb2eL+nbEZm+ncOF2zO2388RtO0F7tzMpaDsnEoR6wXYk"
    b"GrZT3r+dpX9u583D7cS/247+oB3ckt3BIdUdTNDZwUnjHXha70Bu/Q6ad+xg3eEdiETvoDxrByur"
    b"dvC1cwd553ew+O4O3r3cQfq3HZhL7uT12J2kaAqCfu5OXi0WfIedmHns5H8HdpIRuhPL1J18LNlJ"
    b"XstOlp/ayffrOyl/thOHf3ciMcSLltFewrz1QkHfi1PmXhyw82Kqmxe393gRGejFggQv3uZ7UVDv"
    b"xap+L4b96UXXQy92v/NCfdAu7sjuIlZ1FxY6u/huvIt66114rN+F6s5d3Di8i5joXVhm7+JX1S5a"
    b"u3ax68IuNO/t4smrXWR/34XD8N3CrN3NxWm7CZ+3Gwur3Qx22k3P1t34+e7GMHw3n9N201S2mz1t"
    b"u9E7s5t3N3dT+/duvL7sRkdiD28V9lCnsYfdc/Ywy3IP/67aQ+vmPfju34NRyB5+J+3heNEe4fzc"
    b"w9ITe5C7tofrT/aQ8XEPLqJ7BY22l9dqe2mcuZeDZnsxs92L5Ma9XNm9l4yAvbjF70Urfy+f6/bS"
    b"17eXiMt7WfVwLxPe7eXV7720yu4jSHUfK3T2oWSyjxfW+2hZv4/gnfuwPbKPCTH7eJu9j57qfcR0"
    b"72P9H/vQvr+P36/3cfnHPvKl9rNHaT8LtfajaLSfF0v307V2P3Hb9rPJbz+GkfuRytzPo4r9tHTs"
    b"J/Lcflzv7Mfg5X5kvu3n2TBvusZ6k6Tpzfa53ixc4o2yozefPby55ONNeZg3QWneOJd5Y9jmzagz"
    b"3ry56c25v70p+SLUJA7goniABVMOMM7gAF8tD3Bj9QGatxwgyfsAe48dwC7lALNKDjCq5QCfTh7g"
    b"+vUDtDw7QNq/Bzg4xAdneR9M1X2YNNsHcQsfXtr7cHGTD437fEgL9uFwkg9uRT4safJB54QPCtd8"
    b"+PnEh6cffbgg6kvTKF+yJvkSMsuXnea+ONj5gpsvWnt9UQjyZXCiL68LfLnV4MuJAV/qrviS/diX"
    b"iA++HBA5yOaRB7FTOwgzD6JrdhBV24PIbDzIr90HeRNwkAfxB7mUf5D++oM09B+k6M+DpDw6SPj7"
    b"gxwa7IfXCD/cJvqxRs+PpfhhauPHbFc/pu32Y0KAHwrxfkjl+yFa78e3Pj8+XPbj5UM/Hr/z4+6g"
    b"Q1yVO8QfEw5xSveQMJcP0bnyEM0uh6jbdYgq/0OUxh2iMO8QeXWHyOo7RMblQ6Q+PETyu0MkDjpM"
    b"gtxh4iYIpnuYWFPBVgrmIsS7DhPvL9TjDpOUd5iUusOk9R0m8/Jhch4epuDdYYoHHaFc7gg1E47Q"
    b"qHuENtMjdK88wnGXI5zbdYTL/ke4GXeEv/KO8LzuCG/6jvD58hF+PzzC0PdHGDH4KONHHEV94lF0"
    b"9Y4yj6NY2hzFxvUozruPsjXgKN7xRwWdcJS4+qNk9x+l8s+jdDw6ypn3R3k9yJ+PQ/35LueP6Fh/"
    b"hk/0Z7SmP0p6/qjP9UcbfwyW+AvP1B8rR3/sXP1x3uqP+25/dvv64xfgz7Fwf+Lj/clM96ck35+G"
    b"cn+66/050+7PtX5/Hp715/Wf/ny97c+Qx/6MeOmPygd/pn/zx1AkAAvJAOxGBuAyLoCdagEcmhZA"
    b"xMwA0ucFUGoWQItVACdtA7jmFMCTjQF88AxAZG8AI/0CUAsKYGZkAGaJAdhmBuBWGMC+ygBCGgNI"
    b"6RTwxwNoPx/AuasB3LsbwJsnAQx+HcCoTwGo/wjAUCyQpVKBrB8dyG6lQEImB5KmFUiVfiB9RoFc"
    b"WxjIi2WB/LQPZKRzIBqbA5m/IxCb/YFsORzIoZBA4qMDKU0OpDs7kKvFgbysDmRQSyCKPYFonwrE"
    b"4mIg624Esu+vQKKeB1L4JpDOz0Lvr0BeiwchJhuEsmIQc1SDWDElCA+dIAIMgsg0CaJ5URCXVgTx"
    b"Yk0Qoi5Cn0cQhruCsPMJYqd/EOFhQRTFBdGXFsTdvCC+lAUxqj4InfYglvULXGeDCPkziILbQfQ+"
    b"CuLeiyC+vQ9C8VsQs0WCsZMMZvfIYGLHBVOjFsyFacG8mhmM5PxgNM2DWbI0GA+7YMLWBlPmFsyZ"
    b"bcH8szeYYYeEenAwVlHBbEsKJiormOqiYC5VBfOuKZiR3cHonwxm9R/B+F4PJvN+MD3Pgnn0v2DE"
    b"Pgcz5ZeAFQ9hp0wI8QohNKuEcFsjhF/aIagZhLDIJIRti0KIWxFCy5oQ7m4IYZBHCBq7QljmE8Ie"
    b"/xDSwkLoiQvhWVqIsL9DmFUeglN9CAHtIZT1h3D5bAhf/wxhwp0QFj8OYddLAfMhhP5vIbwUOcao"
    b"4ccEnX6MzeOPETvpGO3Tj/F41jGkjI5hsPAYrsuOEWV/jJZ1x3i0ScjvOIbh/mO4HRb6Q47REX2M"
    b"58kCT84xTEqOsa3mGGktxzjZc4wPp46heukYy28e4+CDY5T9fYwbb48h9vWYcO8LxWVoKLFyofSM"
    b"CeV/E0JR0gxlqV4ofnNDqSCUO0tCkbQNZZ5TKNs2hpLpGcqFPaH8PBiKdlAoGyJDiU8MZSAzlE+F"
    b"oWhUhQp3/1AiukLpPhHK2wuhqF0PZdX9UMKehdL1PyH3byiTfoWyRjyMSJkwehXC+KgSxtQpYTjr"
    b"hBFvEMZpkzB+LApDb2WYcKcOI9sljKseYQzbHYaJbxjeAWFUhYfxJD6McRlh2BSEEV4RRl9DGF87"
    b"wtA9HobH+TDyroZx+24YI56GYfU6jKBPYXT9CONfsXB0pMPxkA+nQDmce+rhKGiHYzMnnEjjcE5Z"
    b"hjN4RThGa8I5sCGcBvdw3niFM80nnC3+AiYsnL/iwhmfHs6a/HCSysP5sz4cmY5wlg6EE34unNNX"
    b"whlyNxzzJ+EEvAqn92M4v74LnGIR+ElF0DE6gm9KEcxVj8B3RgTtsyP4ukCILYXYWohXC/X1Ecx3"
    b"j+CQVwTdByL4dTQC07AIAuMiOJEWgXh+BIvLI4ioj+CP9gjkBiKwOxdBypUIbt+JQOlJBC6vIij6"
    b"GME/3yPQEotkl1QkzaMjhd+OxEQ9kuAZkZydHYmMcST2lpFkWEfyYHUkGhsi2e4eSYNXJP8diMTU"
    b"P5LQsEguxkUinx6Jc34kReWRvK6PRL8jkkMDkZw8F4n01UhW340k90kk/7yKRO9TJH4/IjkhFoWM"
    b"dBQO8lHCM4/itXoUBtpRBM6J4oJxFIqLonBbEUXNmii+bYjCwiOK+F1R3PeJQjMgCu/wKAbio5DN"
    b"iGJdQRTlFVF8bohiYWcUCcejeHA+ihnXojh0L4pzT6MY878o3P+NouVnFEPEo7GXiaZIIVr41qIx"
    b"nxJNok40TwyimWUaTfDiaK6tjEbdMRpv12hOb41mzJ5oPA9G0xUYjXRkNBsSo2nIjEasKJo1VdGU"
    b"N0XzoyuaFSejKfwjms/Xo7H6K5rs59G8fxONxZdoMn5H80YiBnO5GNLGxAh7LAYzzRhS9QR/ruCb"
    b"CXmrGN7YxmCxNoZMtxg+bIth8b4Ycg/F8Dk4BuvoGIqTY/iZHcOqkhiqa2IY0hqDc28MLadjkLkc"
    b"g/utGPoexjD2RQx738dw4b8YNERi8ZeM5c7IWPTHxwqzJZYX02NZqB9LrlEs3xbGsnp5LPWrYpFe"
    b"H8vWLbGc3BnLxAOxHD4q4EJjmRMXS2JaLO/yYrEuj6WyPpZhHbFsGYjlxLlY1K7GEnA3lgdPYjF+"
    b"HUv2p1h+/Ihl7ZA4OqTjGKcQh69KnDBT45irE0eGQRzfTOJwWhwn6Js4lBzjOOwax4OtcZjuiaPg"
    b"YByiQXFsjozjTGIc07LiiCqK421VHLbNcTR3xzHmVByHLgqYG3GYP4ij9O84hr+Lw+trHFcHxWM4"
    b"LJ7sEfGIjIvHXS1eOEvimTkrntT58fw0j8dtWTxn7ePRdY4nZbOQ2xHPJu94zh8R+kLjyYiNZ3Ba"
    b"PFvz4vmzLJ659fHkt8cjORDPnnPx3L0Sz8K78VQ/iUfxdTz+n+J5+SMe+yEJ9EgnMFUhgUSVBH5q"
    b"JAhnaAJXDRIwNk2gfHEC8jYJ+Dsm8No1AQfPBE7sSUDXL4HsoAQkoxLwTkrgSVYCNsUCV3UCWi0J"
    b"pPckIHE6gf2XEnh6MwHbhwn0/5OAzvsEcv5LQFokkcOSibwemci68YlcmJSIkVaioB0SUVqQSLRF"
    b"Ij+XJ7JzdSIP1idi457IgFci+j6JFPsnohieSHh8Ij/ShZ6CROGOkIh9YyKnOxOZdyKR6guJTLye"
    b"SNL9RIY+T+TQm0TefU5k8+9EbkskYS2XxPExSRhOTKJaM4lJM5NIm5eEjHkSwUuT+GaXxO51Sfy9"
    b"KYn1O5K4vj+JZUcEzLEk5sUm0ZCaxLS8JPLLkhhXn0RCe5LwrJMIOpfEjytJ7L+bJOilJNxfJ/Hw"
    b"UxJOP5O4NiQZa5lkzigkY6aaTPeUZAx1k2k0TBa0YjLlS5KZbJtMnlMySm7JpG5LZvS+ZOIOJSMV"
    b"kkx4dDLiKckE5SQzqDSZw7XJfG9Nxqcvmc9nktn3ZzIfbiez+3Eyb18ms/NjMv/7nsx2sRReSaWw"
    b"TT6Fl8opeGoIq7awGgh5EyG/OIXXK1PY4ZjCG9cUvDxTeLcnhT1+KXwMSmF/VApfklI4mJ3Cj+IU"
    b"jtakMLg1heDeFIaeSSHycgoyt1NIeJSCwssUMj6koPI9hQLRVKZIpVI5OhVd5VSa1VOZr51K35xU"
    b"LE1SOb8oFZuVqdx0SGW9aypPt6aybU8qHw6m4hOUyqCoVEKTUpHJTiWlOBWVmlRKWlLR7k2l9XQq"
    b"JpdTOXNL4HiUyp0Xqbh9SOV/31LxFk1jkFQaEaPTGK2cRo56GpraaTTOScPYJI2zi9JYtTKNhw5p"
    b"bHNN48vWNIL2pCHrl0ZmUBpTo9JoSkqD7DQuFqextiaNf1rS8O5NQ+xMGvGX01C9nUb1ozSMXqZx"
    b"7kMajt+FHtF0DkilIy6fTopyOuoa6TRrp2NhkM51k3Q2L07n88p0jjmmo7gxnTLPdObuTeecXzrr"
    b"gtN5E5XO0eR0RuSkU1SSzpzadM62prO2T6idSSfgz3Tk7wi4x+kYvUrn0sd03H6k80UsgyjpDCYo"
    b"ZAhaMAOrKRk80Mlgn2EGw8ggd0kG+rYZnHPKwMUtgy/bMojel8Hkwxl0hGRgE5PBPykZ+OdmoFiW"
    b"QW1dBovaBY7+DA6cy0D2agZldzMwfZrB7dcZ7Pk3g+G/MigWz8RYNpNbipnsmZCJlGYmJXqZMC+T"
    b"u2aZeC/NZIR9JlXrMlm8OZMnOzI56p3J+KOZtIZmYheXydu0TKLyM5lakcnJhkw2dmYy6EQm2Rcy"
    b"mX9d4L4v8DzPRP5tJo1fMrEZlMW7oVnEjshCe1wWf6hlsX16FlL6WVQZZbHUIotXy7OIWp2F1gah"
    b"7p7Fjl1ZyPhmURuQhU1EFh8SskjKzGJOURa3qrI42JyFUk8WvaeycL2UhditLEofZmH1Ios377OI"
    b"/5bFbNFsbg/P5vDobCYqZ3NKPRtP7WxkDbJpMsnGcXE2v1dmU+SYjdXGbN56ZpOyNxujQ9k8Ds4W"
    b"9m42OinZ3MgROEqzmVyXzYW2bPb1ZzP+XDbHr2Sz7W42o59m0/U6m83/ZiPzK5tW8RxcZHMYNiaH"
    b"xgk5rNPMQXxmDnXzcnAyz2HIshxq7QXfWfC3CPmdOaw9kIOEfw4NYTmsj89BMiOHloIc3CpzkG3K"
    b"oasrh60nc1C4mMPxGznsepCDyj85XHiXw8H/cpgqkssNyVyCR+UySymXx5NziZ+Ri+mcXN4a55K7"
    b"KBfrlbn8csilxjWX9Z65yO7Npdcvl13BuUyMzuVKcq4wp3LRL83leW0uaW25WPXn8uOsgLmSi+vd"
    b"XEY9zeXU61x8/81F61cuD8XzSJTNY9GYPL5PyKNWMw+3mXkozs/jgnkeAcvymL0qj1fOeeRtyWOV"
    b"Vx7DffLo98/jQHgeMxLyeJKRR0ZhHiur8pBozqOnO4/9p4TapTzhHMoj62Ee9i/ykPqQx4lveRwW"
    b"zWe2VD5vRudTqpyPi0Y+Y3XyhfMvn2jTfCyX5DPINp9Op3z2u+Wjsz2fl/vyKT4s9B7LZ3xsPjdT"
    b"80nIyxe0Rj6SDfmc6sgn8Hg+Jhfy+X4tn7b7AvZ5PjPf5vPuSz7VgwrYNqwAzZEF/D2ugOJJBbhp"
    b"FaA2u4BHCwrIsyxgw4oCVBwKuO9SQPbWApz3FKDsJ8RBQhxVwPpkoZ5TwIMSob+2ANe2Aib1F/D0"
    b"bAElVwrwuFvAtKcFvH5dQO2/Bez+VYC+RCFfZAvpGFPI4YmFMK2QIbMKOTO/kKiFhdgsL0R+dSG3"
    b"1xeS416I265CpvoW8r+AQhojCvFJLMQ4qxCx4kLOVRcS11LImt5ClM8U8vRyIZW3C9nzuJB5rwoZ"
    b"/KmQsz8KiR9ShKNMERMVi3ihWkTD1CL89Iown1eEtHkRN5YWkWtfxFbnImZtKeLnziJOHygi3r+I"
    b"deFFaCQU8T6jiK7CIkKrirBrLkKlp4iXp4pouVRE4K0irB8VMe5lEX9/KKLpexEBYsVYSxczXqGY"
    b"FyrFtEwpJli3GNu5xUwwK+atVTHddsVErStm7eZipu0s5j/vYs4cLRbuxcV4xBdjmFHM0MJiblUW"
    b"U9ZUjG93MVanBM5Lxby+KeAfFhP7ohiXD8XM/F6MmFgJN6RKKJcvwU+lBOspJUzULeGTYQmnKSHT"
    b"qoSddiWYrStBYXMJL3eU0OtdQuLREjzCSlgQX8LIjBL+Liihq7KEhKYS3LuF/KkSRl0q4cXNEkEr"
    b"l5DyooQdH0pY+L2E8WKlfJQq5Zx8KQUqpfhOKcVGt5Rpc0sRMSvljlUpjXalwv8sZfPmUox3ljLm"
    b"QCnvj5ZyPqyU4vhSjmaU4lhYin5VKTLNpbzoLuXEqVJyL5Xid6uU1Y9KmflSqH0s5eX3Uk6JlVEg"
    b"XcZRhTLWqpZhOLUMeb0yPs4t47JZGTVLy4i0LxPeZxmLtpSh4VWGmE8ZT/zLGAgvIz+hDP/MMjYU"
    b"lWFSXYZKSxm/esr463QZvZfLyL0t1B+XCfe/MvhUhtpPAS9ezjOZck4rllM+oZxIzXJ2zCxnxfxy"
    b"Zi4sZ/Tycr6sKufO+nK63cvJ31VOsG85HoHlLIssRzdJ6Mku57/icu7XlDPQWk5pXzlRZ8vZc6Wc"
    b"NXfLWfC0nEn/K2fY53Le/irnukQFXXIVFIytIFytgl3TK3DQr8BkQQVTLCuQWVHB5zUV/OVSIdy7"
    b"KqjdU0GaXwUBwRV4Rldgl1KBUW4FGmUVyNZX8F97BU8GKrh4voK2awLv/Qqin1fg87YCt68VrBhc"
    b"yTzJSjRGVTJCqZKfkyt5OaOSG3MqGTCppHZxJVk2lYQ7VeLtVsmm7ZXY7K/E5EglM0IrGR9XybD0"
    b"Sr7kV/K8opLrjZWc6Kqk8WQlBRcrSbhZScDDSna/qMT1g4D9XomZWBUzpauYpFDFaNUqhkyt4rNu"
    b"Ff/MreK2WRXnl1bRbV9FrXMV+VuqSPKq4phPFb4BVWyPqBLullXYZVVhWVzFvJoqtFurUOurQuFs"
    b"FZJXqvh9p4pPT6p48bqKv/6t4uqvKs5KVNMjV03T2Goq1KrJm15Nin410QuqCbas5uCKanY7VLPV"
    b"tRoXz2oc9laz4lA1i0KqMYmpxjC1Gt28ajTLq5nYUM24zmpGnahG+o9qxG9UM+hBNd/+rubTu2re"
    b"/FfNPyI1PBlew/3RNdxSruGqRg2XdGo4Z1jDKWrot6oRZkEN7etqaNlcQ8POGmoP1FDlX0N5eA0l"
    b"CTUUZtaQX1RDTnUNWS01ZPTWkHamhtQ/a0i+U0PSkxoSX9eQ8G8N8b8Ek6glTk6wsYKpCTa9lnh9"
    b"wRbUkmBZS+KKWpIcakl2rSXFs5a0vbWkH6olM6SW7Jha8lJrKcirpbi8ltKGWio6a6k5UUv9H7U0"
    b"3ail7UEtXf/U0ve+VjjDajkjWscfUnVcka/jhkodd6fU8Ui3judz63htVsf7pXV8sa/jp3MdYu51"
    b"SO6qY4RvHWMC61CNrEMjqY4Z2XXMLqnDqLaOhW11LOuvw/5cHc5X69h8r46dz+o48KYO/y91hA+q"
    b"J3FYPVkj6ykZX0/d5Ho6ZtRzck49l0zqubO4nmc29bxzque7Wz3iO+oZ6V2P8tF6NMPqmR1fj1lG"
    b"PdaF9aytqse9uZ59PfUEnK4n9rLAe7ueisf1tL2q59Sneq79rOexeAPvZRv4PaYBGbUGlKc3oKXf"
    b"gNGCBpZaNrB2RQPbHBrwc20g0rOBrL0NVB9qoDukgYsxDTxIFfB5DYhUNDC6sQGNrgYMTzZgdbGB"
    b"dTcb8HrYQOCLBpI/NFD2vYFOsUYuSjfyWKGRz6qNSGo2ojKzkZnzG1m0sJG1yxvZtbqRkA2NZHg0"
    b"Ure7kZMHG7kb1Mj7qEYkUhpRzm1Ev6wRq/pGXDsa8TneSOyFRkquN9LzVyPX/27kf+8aGfJN6BVt"
    b"Yo5UE9byTbirNOE/pYl03SYa5jZx3qyJZ0ub+G3fxNj1Tcxyb2L5LqHPt4nAwCayI5toS2rianYT"
    b"b0qakKxrQr29CQaacD7fhO+1JlLuC1zPm7j0tonXX5sYJtKMxvBmFo5uxlW5maMazWTrNNNp2Mwd"
    b"mvlq1YyifTNznJtZtaWZfV7NJPo00xjQzJWIZj4kNjMyu5mZJc3Y1Tazt62ZpP5mms81c+OqgL/X"
    b"zNjnzcx/28y6r80cGdxCvmQLx0e18EyphaEaLUzXacHasIXdtJBs1UKbXQv31rUwaEsLk71aWOzT"
    b"wo6AFuIjWmhJbOFullAraUG9toWlbQKuv4WUcy10XW3hyb0Whj1vQfdtC6u/tnB4cCtFkq2cH9XK"
    b"R6VWxmm0YqbTiqdhKwm00mnVyhO7VqScW5m9pZX1Xq2E+rRSF9DKnYhWRJNahT3RypqSVgJqW6lq"
    b"a+VGfyuDz7eidU3I328l8HkrNW+F/q+tDBFpQ294G86j24hQbqNFo40nOm3IzW1jgVkbnkvbSLVv"
    b"46RzGx+3tDFxVxsrfNs4HNhGZWQbd5LaGJbThmFpG+51baS0t3FqoI1/z7ehfr2NVX+1EfK3wPuu"
    b"jef/taEo2s5iqXZ85dupUGnn3pR2pPXaMZnXzm7zdgqXtXN9VTtDN7Qzz6OdHbvbyTvYztWgdiSi"
    b"hVxKOztz2ykoa+dGfTuSnQL2RDt7/2in7EY79x+0M/JFO4s+tHP4ezsNYh38I92BimIH9hM6iNDs"
    b"oH9mB1/nd6Bt0cEW6w6y13RwzaUDKc8OzPd24Heog8aQDl7HdDA5rQPn/A6SKzq42NjB0O4OBKHM"
    b"wUsdNN3q4M2jDqa+6mDjpw6yfnZwU7yTkXKdLBvbSZhaJ8end/JLv5N5xp14L+qkfmUnbxw7mebW"
    b"yZbtnRTu7+ThkU6Uwzpxiu8kNaOT64UCR3UnK1s6ient5I8znQy/0onVXYHzaSen/9eJ+JdOLAd1"
    b"ETKsi5MjuxBT6mKhuhBrC7FBF0PowtKqi1C7Ls6s62LYli6svLqI8uniYkAXcpFd2CZ1kZTdxY2S"
    b"LsbWdbGuvYvcgS4en+8S3l0XW//qourvLt6960L/Wxc+ot10S3UzWKEbS9VuIqd2c1mvG/n53Tgt"
    b"7CZveTfPV3ej5dLNnq3dtO3p5pdfNxYh3UTFdHM1tZtx+d24VnRT3tjN+65uDE9143+pm7O3uhnx"
    b"WOB51U3Bp25e/ewWdH8PR+R6OD22hxGTenDS6qFodo9wl+th7uIegmx6uOjUw9hNPWza0UOtdw/f"
    b"jvZgGd5DQkIPf2X2MK24B++aHkE79SDb38O6cz2UX+3h870ezJ/3EP9W6Pvag5ZILweH93JmdC+K"
    b"Kr1sntJLk24vovN6sTXvJX9ZL+9W9WK6oZc4j14e7u5F168X/+BeLkf3MjG1lz15vRwv7xXmtIDv"
    b"6qX1ZC9DL/XidKuXqke9/H7Zi82nXgp/9vJZvI8lcn1kje3jrVof5lp9pMzu46VxH8aL+0iw6eNv"
    b"pz7mb+ojbkcfz7z7mOffR2x4H08TBD9L8IuFfI3Q0yb09/fxz7k+TK71kXy/j9fPBc53fWT+18cH"
    b"kX6spPopkO/nm0o/tlP7qdDrR3R+P2sX9tO0vB/pNf24u/TTt7WfsXv72Xuonz9C+pkS209gWj/3"
    b"8vsxqOwnvqmf1939LDrdT+Hlfn7d7sfpST8tr/sZ+bmfnb/7OTd0gCkjBwgeP8CjyQMYaw+QaTDA"
    b"V9MBVlsN0GQ3wEjnAXZtGeCi1wAzfAeIDBzgZeQAS5IHKMsZYGjZAO71A5zuELhODBD6xwB/3xhg"
    b"0UOh/mKAYR8H8PwxwPkhx9GSPU7MmOO8nXgcm+nHadQ/jrzxcXwWHefOyuMscDpOnttxRHccZ4v3"
    b"cc4ePc6M8OPEJxznU+ZxHIuP011znIn/Z7ico+zKmijedr+2+7VtO9g71sS27XyZOJnYtu1MnInZ"
    b"tm3b5nf/qHXPqdr7V9Vrvb7n3I9/cPDXH1RF/sG45D94n/MHhmV/sKf+D0o6/mCUdDBeqwRDTzcY"
    b"u0yDUWwXjFEewXgTGAyDIcHY81cwyqcGY+z8YPy3Ihimm4JxaGcw6g4GY9rpYPy8Egz7e8E4+zwY"
    b"Xe+DsfhHMGLDg+GXGIy7WcFQKQnG5tpgFLQJXMkQfFAKgbl2CI6LQ9BmE4KFbiGI9Q9BwKAQPBwd"
    b"As0pIdg9NwRVy0IwbUMIQraHwONACG6dDIHK5RBsuxOC8qchmPwuBMHfhFpYCG7Hh0AtMwS7ikJQ"
    b"Ux2CWa0hiOoLQaBiKJ5qhsLQOBRHrUPR6RKKlX6hyBJum6NHhQr36VC4zAnFraWh0Fgfir3bQtG8"
    b"LxRLToQi/WIoRt4WNE9C4fo2FHe/hkInNBSH40LRnR6KtYWhKKoKxdSWUET1hmKAQhjeaoTBzigM"
    b"N6zCoOkShkO+YegeGIb1I8NQNjEMs2eHIWlJGEasC8OPrWHw2ReGF8fDYHkxDFdvCZ4nYTjyJgyS"
    b"X8OwJSQM9bFhWJYehoKCMEyvCkNicxhG9obhj3w4gjTC8cEwHK5W4XjiHA5L33DcGBgOvZHhODMx"
    b"XDhLw3FoSTgk14Vjx9ZwtO8Nx8bj4ai7EI4Vt8JR+jgcC9+EI+9LOGaGhCM9NhyT08ORWBCOsVXh"
    b"iG4OF3qFI0w+AkM0IvDHMAK0isBP5wgM8I3At4ERCBoZgS8TIxA4OwKfl0QgYF0EPm2NgP8+YX9c"
    b"2F8UnreE+pMIfH0TgX5fI/A9RPDHReBXegQGFUYgpCoCw1oiENEbgVEKkYjViMR4o0gkW0Viqksk"
    b"Mn0jMQeRKBgZicWTIlExOxKrl0aicV0kNm+LRNe+SOw5EQnZS5E4ejsSak8jceFtJIy+ReJOaCRs"
    b"4yPxIiMSXkWR+FwdCbRGIrwvEmMVo5CmGYU5xlEotY7CGtcotPpFYRejIDc6CqcmR0F/bhRuL4uC"
    b"w4YovN0ehaADUQg9GYVxl6OQeScKi55FofZdFLZ+j4J0eBROJkTBICsK94uj4FYbhS9tURgmGY0k"
    b"pWjM1Y5GlTgaW2yjIeMejTMB0TAZHI1nY6LhNzUaofOiMXFFNAo2RmPtzmj0HIzG8dPRMLwajSf3"
    b"BM2LaIR9iMaUn9EojYjG/5KiIZsTjQul0bCuj8b7jmgMlY5BqkoMlurGoN00BkfsY2DoGYPnQTHo"
    b"PzQG8WNjsGB6DJoXxODgqhgYbI7Bs91C7UgMEs7GYPF1wfcgBsdfxsDsUwze/o7B8OgYZKfEYF1e"
    b"DGQrYnC1MQau3TEIlo3FdLVY1OnH4oBFLIycYvHaOxbDBsQiZ3gsNk2IhfKsWNxdHAv/tbGI3xKL"
    b"ZXtj0XcsFpcuxML1VizCHsdi3ptYtH+JxZmQWDjExeJPeixmFcaipSoWp1piYd8Xi98KcZitGYc2"
    b"ozicsY6Dk2scwvzisIBx6BkVh8uT4+A1Nw5xy+KwckMc5HfE4d6BOAw8FYfsy3HYejcOus/j8PZ9"
    b"HMb9iENteByOJcbBPlvglMRhcV0cZDricFcqHlCJR75OPHabxsPEPh7fPOIxOyge3UPicX1sPIKm"
    b"xyN7QTx2roqHeHM8vu+Ox9wj8ZA4F4+71+Mx6GE8Sl7G49CneNj9iUdUdDxWp8ZDPT8ebyriMbkp"
    b"Hu3d8bgml4AB6gkoNEjAQcsEODgnINYnARsGJgj/wwnC/1UC5s1OgOzSBDxdl4Bx2xLQsi8B104k"
    b"QPjYRdntBJx4mgDPdwnI/JaAPWEJsE0QGJkJ2FycAHFtAkLaErBKMhHayon4pp2IxSaJULVLxAf3"
    b"RMwNTITCkES8/isRM6YlQmZBIv5dmYip/0uE1O5EPD+ciClnEyF5PRHPHgjrl0L+UyJe/E7EtGhB"
    b"n5qIV3mJmFUhcJoS8a47EfPlkqCqnoTPBklYapkEbeck/PZJwtqBSRCPTELUxCRsnZ0E26VJSF2X"
    b"hP3bkuC5PwmFJ5Jw5lIScCcJ9U+TcPtdEsZ9T4JEeBJeJSRhXlYSNEoEVm0SNrQnwVIqGSnKyTio"
    b"kww/02RU2iXjmkcy/gpKhsTQZLwZm4zF05OhvzAZUauSsWtzMtz3JKPkSDKunEvG6BuC7lEy3r9K"
    b"xvLPyTAJTkZiTDIOpSUjqCAZjZXJeNScjNm9ydBSSEGERgp2G6XA2zoF1S4puOuXgulMgfroFIRN"
    b"FmpzU+CzPAW1G1LwcEcK5hxMge7pFMRdScGheykY+CIFHR9S8OZnClZFpsA6OQV5OSm4XJaCiQ0p"
    b"UO0SODKp2CtKRZB+KtrMU/HGMRWrvVNhPyAVJcNTcXtCKmbNSoX+klSkrE3Fma2p+GtfKpRPpCL8"
    b"YioO3E4Fn6ZC4l0qvn1LxfawVPgnCKzMVHwoTsWm2lR4tqeiUTINr5XTsE4nDW6maaizS8NLjzSs"
    b"DUqD69A01I9Nw6vpaVi/MA0eq9PQtDkN7/ak4X9H0+B7Pg0dN9Lw5VEadr5Ow4AvaZAMSUNwbBoO"
    b"padhZGEaVKvTEN+ShvN9aZiqmA4jrXTkGafjnk06lrmlwykgHQ2D0vFhTDq2T03HwPnpkF2ZjqhN"
    b"6TizKx1TD6dDfDYdRdfS8fRBOta9TIfvp3T0/k5HSHQ6jqemY1J+Oowr01HclI7nPenYJJ8hnIcZ"
    b"kDXKQIxVBi66ZGCuXwbsmYHGURn4OjkDB+dmYNzyDBhuzEDxjgy8PJiBraczMPhqBtTuZyDzRQYe"
    b"/pchfMdmoF9UBhRTMpCSm4E75RlY05iBgO4MyMtlIlktE3cMMrHGMhNBzplQ9M1E2sBMPBiZiY2T"
    b"MsE5mVBflom89Zn4d3smdh7IxOhTmTC+komqu5n48jwTxz5kYubPTDhGZqI7KRPROZm4USYwGzIx"
    b"oCsTGrJZKBRl4Z1+Fg5aZGGaUxYcfLLQPSALcSOycHdiFjbNzsLwpVkwWp+F2m1Z+L0/CxdPZmHF"
    b"5Sz0v5sFzedZKHufhS8/snA6IguLk7IQkJMF9bIslNQL+c4snJHJxlJRNvrpZ0PLIhuVjtn45Z2N"
    b"ywOysXZENoZOzIbJ7Gy0LMlGzLpsPNiWjZ37szHlZDZcLmdD/m428p9l49P7bJz9kY2VEdkYkpQN"
    b"05xsdJRmI6k+Gy86s3FIJgcLRDlCnxzoW+SgyTEHsd45eDogBwdH5GD+RKE2OweGS3PQui4HSdty"
    b"8HJ/jvB9lYPll3Mw7G4OrJ7nQOpDDvJ+5OBbRA6uJeVgW04OppXlwLchB7pdOWiRyUWyKBdv9XNx"
    b"ziIXG51yMcEnF54Dc6E1MhdNE4X67Fy8X5qLC+tz8ff2XEw7kAv/U7kwupKLnru5yHuei18fcnHv"
    b"Zy4ORuZiWXIuRuXmwqU8F5qNuWjpykWGbB6+quXhtkEe9lvmYblzHsb45sEDedAdlYeuSXkomJOH"
    b"0GV5eLYhD2d25OHvg3mYfToPg67mweF+HtT/zUPbf3nI/ZWHkKg8PE/Jw7m8PGyvyMPCpjyM6smD"
    b"p3w+jDTyIWOUj2qrfKS45OO7Xz4eMR9nRudj25R8LJqXj79W5MNvUz4sduVD9XA+2s7ko/BaPmIe"
    b"5OPjy3zc+5SPU38EfUw+lqTlY0JBPvpX5cOxJR96fQJfsQANmgXINS5AlE0BProV4GFAAc4PLsDe"
    b"vwqwbloB5iwowOhVBQjcXACHPQUwOFoAhfMFaLtRgLJHBUh9XYDQLwX4ECJ44wpwKaMAh4oKsKWm"
    b"AMvbCjBDshCjlAvRT6cQLqaFMLMvhKZnIWT6FaJ1aCEqxhUia0YhYhcV4teaQrzbUojHewtx7Xgh"
    b"Tl0sxL7bhfj7aSFWvivEvO+FmBReiBGJheifXQjP0kLY1RfCpLMQWjJFUBQVoU+vCC3mRah2LEKh"
    b"dxEyBhQhfkQRwiYW4fvsIuE3UIQX64vwYHsRbhwowoVTRThxpQgH7hVh14si/P1fEdb9KsLyqCIs"
    b"SCnCrLwiTK4owtimIozoKcIg+WL01yiGn1ExPK2L4eJaDHv/YlgNKobZmGIYTS2G3vxiaK0shtr/"
    b"iqG8uxgKR4ohc64YkjeK0fuwGF2vitHxuRitwcVoji1GY3ox6guLUVtdjOrWYlRJlKBSqQQV2iUo"
    b"NylBmV0JSj2ECBJiqBDjhJghxCKhtkbQbBFir6A/XoKqiyWovl2C2qclqHtXgobvJWgKL0FrYgna"
    b"s0vQVVqCnvoSSHSVQFq2FPJqpVAyKIXIshSazqXQ9S2FIUphMqoUlpNLYTe3FE7LS+G+sRS+O0sR"
    b"dKgUPFOKYddKMeZBKSa9LMWMT6WY96cUS2NKsSatFP8rKMWOqlLsaynFsb5SnFUsw1WtMtwVl+GZ"
    b"bRneupfhS2AZgoeUIWZsGVKnlyF/YRkqVpeh6e8ydP9TBrnjZVC/WAaj22WwfloGt3dlCPxehqHh"
    b"ZRifWIZZ2WVYVlqGTfVl2NNZhuMy5bgsKsd9/XK8tijHd6dyRPmUI31gOUpGlqNpUjkk5pZDtLwc"
    b"xhvL4bCzHP6HyjHsTDmmXCvH4gfl2PSyHPs+lePsn3LcjSnHm7Ry/C4oR2JVOQpbBEZfOaSVKqCt"
    b"XQFrkwr42FVgmEcFpgdVYOXQCuwcV4FTMypwZ1EF3q6pQOiWCqTvrUDV8Qr0XqyA5h3B96wC/u8r"
    b"MOZHBeZHVAj3+QocyakQzoEKvGmoQFhXBbJlK9GgVgl5w0oYW1XCw6USw/0qMYeV2DS6EkenVOLO"
    b"vEp8XFGJuE2VKN1ViZ7DldA5VwmnG5UY/KgSs14L2i+VOBFSiQdxlfieUYm0okrU11RCsb0SFlJV"
    b"CFKpwhTdKqwzq8JRhyrc96rC9/5VyBheheYJVVCbXQWHpVUYsr4K87dXCWdaFS6fqsK7K1VIuFeF"
    b"mhdVUPpYBZvfVRgUXYV5qVXYlV+Fa5VV+NRchdRegaNQDS2tariLqzHWthpr3KtxIrAaz4dUI2ps"
    b"NSqnV0NpUTUc1lRj5JZqrNhbjaPHq/H0olC/XY3qp9VQfV8Nlx+CP6Ia65KqcTanGu/KqpHaUI32"
    b"rmoYytWgn3oN5hnWYK9VDR661CDCrwbVrIH6mBp4Tq3B1Pk12L6yBrf/V4Pg3TWoOFID0XmhdrMG"
    b"0x7XYOebGtz7WoPw0BrUxtdAO6sGASU1mF9Xg0MdNfhXuhYpqrXo1quFpUUtRjnVYoNPLa4OrMXv"
    b"kbWonFQLzbm1CFxei0Uba3F8Zy3eH6pF3plaKFyvhfvDWsx8VYsDn2vxMrgWGbG1kMmohUtRLabX"
    b"1GJ/m5CXrEOmch3kdOvgYVaH2Q51OOJVh/f961A4vA6iiXXCN3cdli2tE86xOvzeXof6A3UQn67D"
    b"qKt12Hq/Do/+rUPKxzrI/BEYMXWYn1aH0wV1+FFVh9qWOphI1OMvpXrs1K7HC5N65NrVQ+RZjwH9"
    b"6rF2WD1uj69Hwsx6SC+ph9e6eizZVo9L++sRcbIeXZfr4XKvHvNe1OPcf/UI/VWPjqh6OKXWY26+"
    b"kKsUcs316OwVdIoNWKjVgEviBkTbNkDCowHeQQ1YMbQBt8c1IGVGA5QXNwBrG7B5awNe7GtA0YkG"
    b"GFxuwLi7DTj0vAHfPzSg5WcDnKMasCilAdfzGpBc0QCV5gYM7m3AToVGvNdsRK1xI2xtGzHPvRFX"
    b"AhuRNKQRquMaMXRGI/YsasSXNY1o2dIIt32NWHGiEQ8uNSL/TiOMnzdi6odGnPvZiNjIRiilCJ68"
    b"RuyraMSPpkZ09zTCX6EJmzWb8Na4CfU2TXBxb8LKwCY8GdKEsrFNsJ7RhEWLmnB3TRMKtjTBbF8T"
    b"5p5owq1LTci90wST502Y/aEJN34K+0hhn9KEOXlCvaIJ+U1NMO9twgKFZtzXbEaJcbPwdzRjuXsz"
    b"ngU2o2ZIM1zHNWPDjGa8X9SM9jXNCNzajF37mvH7RDNkLzdjxN1mnHjejIQPzdD91YwZUc24ldKM"
    b"4rxm2Fc2Y01zM971NqNDoQUDtFpwQNyCKNsWaHi0YFpQC24ObUHJuBY4zWzBxsUt+LK2BVLbWjBy"
    b"fwvOnmxB5uUWWNxrwcoXLXj7Xwu6f7VgSHQLTqa2IC2/BeZVLVjR0oJ3fS3oVWzFcO1WnDVpRY5d"
    b"K+w8W7GhXyu+DWuF4oRWTJrVittLWlG1rhW+21ux70Ar4k+1wvhqK5bfb8X7f1uFb81WjP3Tihsx"
    b"rahMa4VfYSsOVrciubUVFpJtWKfchh86bRCZtWGWQxueebWhq38bRo5ow9WJbaiY3Qb/ZW04sqEN"
    b"mTva4HCoDdvPtCH6WhvED9uw5pXg/9wGjZA2LIhrw7uMNsgVt2FarcBqb0OPVDvGqbbjnl47Ws3b"
    b"McKpHTd82lE/sB2DR7Xj8uR2VM9tB1a048KmdlTuaseAI+04f64dFTfa0f+xsH4j5L8K+bB2XEwQ"
    b"9FntYGk7rtQLnM52DJPtwE21DrQYdGC0VQfuu3Sg268DkwZ14PmYDshM68DsBR34sKoDan93YNk/"
    b"Hfh1rAOGFzuw6XYHYp92wPZ9B/b+6EB2RAd8kjtwJrcD1eUdGNbUgbs9HeiR78R0zU68M+6Eum0n"
    b"Vrl3IiKwE9ZDO/HPuE7kzuhE4OJOXF7biZatnZi4vxOvT3ZCdEXQ3utE5ItO2H/sxKHfnSiN7sSQ"
    b"tE7cL+iEdHUnFrR24rdEFyyUu/CPThcKTLtAhy7c9eqC9IAuLBzRheCJXbCZ04VDy7pQvqELI3d2"
    b"4fmhLojOdmHd9S4kPuyC9+suXP7Shc6QLsyO78KvzC5YlXThcF0Xqjq6MFamG29F3dA36MYOy24U"
    b"OHdjqF83nrFbeK93439Tu5E1vxsDV3Xj4eZuqPzTjQ3HupFxoRsDbgu5p93C2dGNTT8EXUQ3BiV3"
    b"42luNzQrurGtqRuFPd0YqdCDt5o9MBb3YL9tD2rcezA1qAc/h/bAfnwPzs3sQc/iHixd14OEbT0I"
    b"OtCDh6d6oHG1Bzvu96Ds3x5M+NSD73964BDbgwvpPZAs6sGamh5ktvVgqFQv3qr0wkyvFyfMe9Hp"
    b"2ItlPr1IGdiLQaN68XpyL0znCbUVveje1IsVu3uRcaQXw8/34r+bvbB90ouLb3sh970Xm8N7UZrY"
    b"iyk5vQgt64VvYy8edfdCX74PRzT60GXUh1U2fch168PYwD78GtIHr3F9eDijD/qL+3BsbR96t/Zh"
    b"/f4+FJ/sw9QrfYi614cB//bhzcc+2Pzpw7WYPqil92F/YR/aq/uwuq0PDRISdJKT4FJlCd5Rl2C2"
    b"jgT1jCQ4wUyCJ60lGOEgQRk3CQ70luD2AAn+N0CCjYMl6DJSgivGSvDhJAkWTpegyVwJzlgkwUvL"
    b"JZi0RoLqmyQ4eqsEj+ySYOg+CUodkeCAkxLccU6Cny5LsO2GBL3uSXD9Ywm+fCHB6jcSdPgowWXf"
    b"JPjgtwSLwiRoHiPBuYkSvJEmzJUtQcNCCU4rE3pUSzC1QYLabRKc2C3Bs5KSTJCTpLqKJMdqSPKk"
    b"riRjjCSpai7J0TaSPO4oySg3SSr7SHJUoCSPDRT2Q4T9KGE/TthPlmT0DEmqzJPkmMWSPLFCkrFr"
    b"Jan2P0mO2ybJM7slmbhfklpHJTnplCQvnJdk2hVJ6t+S5Iz7krz2RJI5/0rS9J0k53+S5L3vkiz5"
    b"I0nbCEkuj5Xk8yRJ1qZL0j1XkhuLJPmhXJIdNZIMapLk7nZJ/u6RpKy0FIcrSPGYqhTjNKWoqS/F"
    b"KWIpXrWQYo6tFM2dpbjYQ4pPfKVYEyRFd0px8zApfh4txb7xUhw0VYqHZ0kxer4UNZYK3lVSvLZe"
    b"ivmbpWi9Q4or/5Hiq4NSbDkmxcAzUvznohRDr0lR5Y4UJzyU4uVnUsx9JUWrD1Jc8UWKr39KsS1E"
    b"iv2ipLg/XopRKQI3S4rT8qV4q0SKpZVSdK6X4qYWKX7tlKK0hDRHykrzrJI0M9Wkaa4jzeWG0nxt"
    b"Ks0OK2nCQZpHXKWZ4CVNwwBpLhggzWeDpdk0QppBY6V5YJI0Y6ZLU3euNOcukubj5dJsXCPNwE3S"
    b"3L9VmrG7pKm3X5rzjwi+k9JsOSfN/lekefimNBPvSVP8RJpL/xX6vZVm90dpDv0uzTN/pJkVLk2b"
    b"WGmuT5Lm13RpyudKc0KRNG+US7OiRppeTdLc0y7NqB6BLy3DhQoy/FdVhp2aMhyqL8OzYhnmWcjQ"
    b"0U6GfzvLMNhDhup+MpzdT4ZPKMPWYTIcNEaGpyfIMGeqDB1mC7oFMgxZKkPN1TKct0GGL/6WYdcO"
    b"GY7YK8NLh2RYfFyGHmdl+M8lGcZdl6H4rgxXPpLh5+cyVHgjw6n/yfDhVxm2/BLYYcIM0TIsTJCh"
    b"e5oM92QLngIZmpbJcE21DL83yFC1TZinW+gjKcseOVmOUZHlDQ1Z1ujKsr+xLE+ayzLPRpauTrLc"
    b"4y7LBB9ZmgfJcgNk+WeoLLVHy3LJeFn+N0WWCrNkOWO+LJ8tEVgrZfnXelne3izLxu2yHPyPLC8e"
    b"lGX5MVkGnJHliYsC95osPe7Icv9DWaY9k6XDa1nu+CDLuC+ytPgly82hsoyIkqVxgizXpgr9smSp"
    b"WyDLFaWy/F4lS80GWS5uleWnLlmqSspxgZwcPyjLUUlDjnN15fjWSI4K5nKcZSPH145ylHOX40wf"
    b"Ob4KlKMs5DhjqLAeJazHC+spwnqmoJkvaJYI+pVylF8vx9mbBc52OSr+I8d5BwX+MTmqnJHjwoty"
    b"/HxNjup35Lj0oRy/P5Ojzms5rvogxz9f5Gj4S44bQuWE+eVoliDHLalyjM+So22BHHeXyjGtSo6u"
    b"DXI81CrH/C45+knK87ScPCuU5QkNeV7VlWejkTxHmcvzvo08exzlOcVdni995KkYJM8FkOeXofLU"
    b"GS3PNePlGT5Fnhaz5LljvjxTl8jTbZU8j66XZ/FmefbfIc8r/8iz6aA8/zouz8dn5Cl9SZ5zrsvz"
    b"0x15aj+S59rn8ox8LU/r/+T5z1d5Zv+Sp1+YPM9Fy7MuQZgjTZ6PsgVfoTznlgm9q+Wp1yjPjW3y"
    b"jO+Wp5OUAo/IK7BERYHUVOAtPQV2GytwuoUCP9gqUMtZges8FBjrq0DHfoKWCiwdpsDBYxR4d4IC"
    b"JaYpcM5sBX5doEDDZQrcslqBaRsU6L1Fged2KrBxrwLHH1bgyxMKFJ1T4OrLCoy+IbDuKfDYYwVW"
    b"vlDgiLcKfPJRgQrfFbjsjwLDwxVoF6vAQ0kKLEtX4LBcBT4uEuoVQr1WgRFNCrTvUODRXsEvrchR"
    b"iop8LlKkqrYi1xgoMs5EkW5Wijxrr8gmF0VO9lLkB39F6g9Q5LbBiswaoch+YxV5a5IiJWcoctFc"
    b"RYYuUqTdCkUeW6vImk2KwhmgyLe7FalzQJFbjgqeU4rsf0GRd64qUva2Ipc9UGTUU0W6vlLkufeK"
    b"bP2syBk/Ffk9RJEWUYo8GK/IyhRF/pWlyDf5itQrVeT2KkXm1ytycKsin3QpUiSpxI1ySkxXVmI/"
    b"DSXe1VWivLESV5srMdFGiX5OSrzhrkRpXyUuD1JiHJToNUyJ10YrUXKCEpdOVWLMLCG3QIlXlwq5"
    b"1UJugxJj/1ai904lXt8reA8rccUJJSacFXiXlXj7htDjnhLXPlZi6gsl9n+rxIcflSj6rsTNf5SY"
    b"G67EobFK/DdJiboZStydq8SyIiWOq1Dip1olmjcr8WiHEpt6lThLRpkhisp0UVPmZW1lShoqc4Wp"
    b"MpOtlNnfQZmPXJWp4a0s3CGUWTJAmWOHKPPTSGVajVPmycnKbJ+hzIXzlBmzWJl+K5V5b50yRZuV"
    b"uXW7Mov3CPqDyvx8TJnWZ5R55qIyu68pc9kdZSY9VOaA58p89lqZev8pc99XZdb9UubMMGWGRyvT"
    b"K1GZt9OUqZojsAqVWVqmzIk1yvzZqEzndmVe7VGmvLQKNymosFBVhWO1VPhVX4UOJiq8bKlCWXsV"
    b"bnRRYYGnUPNX4bf+KnQarMKrI1SoMFaFf09SYcl0FU6aq8I/i1TovkKFt9eqCHcHFe7apsKa3Sqc"
    b"fUCF0UdVGHhahc8uqNDomgqP3lZh5wMVrnimwsxXKhz5QYWfvwh9fwnsUBUqR6twe4IKq1MFf7YK"
    b"YwpU2L9MhS+rVWjWqMIzbSqU6hHmllJlibwqp6iqMlxTlf76qnwmVqXYUpWn7FQp6aLKjZ6Cxk+V"
    b"U/urMnKQKoNGqPLfv1RpPkmV56erUm6uKrcuUmX1clXOXavKxE2qHLpNlZ92q9LpgCpvH1Wl1mlV"
    b"Hrqgyq6rqlxzW5WFD4SezwTeK1X2/6DKt19UaftLlddCVakercoDCarsSBW02aosKlAV7m2qjKlW"
    b"JRtV+V+bwO1R5V0pEfUURDyhKqKUlohb9EWsFYu4yFLETDsRx7mIGOopYpC/iG/7i2g/WMTbI0TU"
    b"HSt4JokoM0PE7XNFbFwk4vIVIuavFXHq/0SM3Sbi0D0ifjsgovcxEV+cFtHqoojXr4mofUfE4w8F"
    b"73MRd7wWseWDiKu/iljyS8Q5YSKmRYs4NlHE8DQRkSPi50IRPctFfF4jonWTiLfaRdTvFfGstBqV"
    b"FdV4QKTGPi01bjVQY5OJGldbqbHMXo3zXdWY5aXGKQFqjB+gxlFD1BgyUo0Dx6nxy2Q1+sxU4+t5"
    b"anRaosZHK9VosV6NNzer0WCHGi/8o0b1Q2o8flyNCmfVuP+S0OO6GnfcVWP7IzVueqHG+jdCr49q"
    b"rPimxiV/1FgYrsa5sWrMTlLj9Aw1puaqcWKx0LtCjWPq1BjVrMbhnWoM7VPjIFl1/lJS5wB1dX7T"
    b"UWegkTo/manT10ad7x3V6eWuzjc+6nQLUudLqNNlmDqfj1an4wR1Pp2qTvvZ6ny8QJ12y9T5aLU6"
    b"bTaq8+EWdVrvUueDfcLziDrvn1Sn1XnheUV43hLy94X8U+H5UtC/F/Sf1Wn7U/CHCJwodT6JV6dD"
    b"qjqfZanTqUCdL0rV6VqtzlcN6vRoU+fbbnV6S2nwP3kN+qlq8LOmBoP0NfhdrEFYavCPnQaHuGgw"
    b"3FODI/01GNNfg+MGazBxhAYnj9Vg+iQNzpyhwdy5GlywWIMlKzS4bJ0Gq/+nwXXbNdi0R4N/H9Rg"
    b"1zEN7j6jIZytGjx0XYPKdzV46pEGtV9o8PIbDYo/avDONw3a/NHgs3ANusVq8H2SBgMyNPgjV4OD"
    b"izUYWaHBsXUaTGnW4IxODeb3aXCJrCarlTS5QV2T7Tqa3GWkSRlzTR610aSGkyYvuWvSxFeT94M0"
    b"6UhNvhmmSf8xmvw5QZPDpmkydrYmJy/UZM4yTS5aI7A2anLjVk127dLkvv2aVD6qyXOnNGl0QZP3"
    b"rgqM25p8+0CTQc80GfxKk2M+aDL1iybn/NJkWagm10VrsiNBk3vTBG+OJi8UCv3LNfm4RpPuTZr8"
    b"0q7Jwb1CX2ktTlPUYqFIi6u0tdhqoMV/TLWobK3FCw5aNHPT4jNvLXoHavHnQC2OGqrF1FFanD9e"
    b"i9VTtPj3LC1KL9DiyaVaNFitxQcbtOi+RYtfd2px+D4tJh/W4ryTgvacFrdc0aLsLS2eua9F8VMt"
    b"Pn2pRd/3WvzzWYvjfmoxJ0SLK6K02Bavxf2pWtTI1uKtAi06lWnxc7UWhzVqMaVNiwt7tFgvpc1d"
    b"CtpUEWnzqpY27Q20+Z+JNodYaTPZXpsLXLXZ4KXNPQHaFA3U5o0h2nQapc0v47Q5aoo2M2dqc/l8"
    b"bXYs0ebhVdrU36DNR39r03enNsP2anPqYW2WndDm3+e0KX9Fm5dvatPuvjY/PdHmiJeC/502V3zW"
    b"ZtcPbR4P0aY4Spv/xmtzQKo2E7KE/gXabC7V5oFqbeEOps3HbdoM6NFmtJQO5yjosF5Vh3u1dKht"
    b"oMNHJjr0t9JhtL1Qc9UR5tbh/gAd6g7U4ZMhOgwapcP4cTpcOEWHrTN1eGS+DsVLdfhqlQ4Hb9Bh"
    b"+t86XLVTh317dXjusA5tT+rw6zkdjruiw+KbOtx6X4eqT3V496UOfd7rMOqzDuf91GFLiA6PRunQ"
    b"NEGH71J1ODJbh3kFOvxfmQ6VanR4u1GH3u2CvkeH86V12aagyxMiXVpq6/KzgS7HmeqyzEqXOx10"
    b"qe2my6feukSgLtMH6nLtUF3KjdbljfG69Jqqy+hZuly4QJddS3V5drUuHTbq8vcWXU7fpcuGfbo8"
    b"ckSX5qcE7nldTriqy8pbutz3QJdGz3T57pWu8PvWZckXXe76pUu9MF2+itblyERdFqXpckeOLnWL"
    b"dPmyXJcjanVZ2CTkOoRcn6CT0eNIJT0Wq+lxl44eDYz0+NZMj2Ns9FjuqMd97noU++rxY5AeJ1KP"
    b"tcP0eGSMHq0m6vHnND3OnKPH1oV6PLtcj85r9RixSY+Lt+lRco8ebxzQo/8xPaac1uP6i3oUXdfj"
    b"0zt6HPpIj0XP9bjnjR6NP+rx0zc9Tvmjx6ZwPZ6OFTjJeozK0OOyPD3KlejxXqUeUa/H3BY97ujS"
    b"o6GkPj/K6XOKij6bNfR5Tk+f7mJ9xlnoc42dPlVc9PnMU194D+qzor8+Dw/Wp+1IfYaO1efiyfqU"
    b"manP+/P0OWiJPotW6nPven1a/K3P3zv0uWCvPqUO6/PuCX3ynD4LL+tz3019Wt7XZ/ATfS56qU/Z"
    b"9/rC+1ufw37qsyxEn0ei9OmQoM+oVH2uytanqFCfL8v0Ob5Gn42N+jzfrk+fXn2mSxtwm6IBjdUM"
    b"+F3bgPMMDShtZsAH1gYc7mjASjcDnvAxEM4dAybBgJuHGdBgjAG/TjDg3GkGlJpjwIcLDThiuQFr"
    b"1hjwzCYDem0zYPpuA+44YECzYwYMPm3A5RcNqHrdgK/vGHDKIwN2PjfgrTcGHPzRgOXfhB5/DOgR"
    b"YcC0WAPuTDagRaYBw/IMuLrEgJpVBvxYb8DZrcJs3QZ8ImnIsfKGbFEx5DVNQ0LfkOViQ560NKS3"
    b"vSGzXAy518uQ9gGGjBtgyM1DDCkeZcjgcYZcNcWQWrMM+Xm+IRcsNaTiakO+3mDIaVsMKbHLkI/3"
    b"GXLcEUO2nzTk7fOGHHHVkA23DHnlgSH5zJCVrwx57oMhg74aCvcToWeYIX1jDJmfaMgj6Yb0yDVk"
    b"dpEhD1QY0rXOkOnNwiydhnSSMGKqrBH3KBvRQcOIybpG3GVsRHsLIybZGnGnsxHtPIW1n5DvL6wH"
    b"C+sRwnqsoJ9sxJQZgneeER2XGDFtpZHw+zCiy99GzNxhxIN7jehx2Ii5J4x49JwRfa4YsfCmEU/d"
    b"N2LgUyOWvTTi+fdGxBcj1v40Eu6yRhwWbcSWBCPeTTPi2Bwjdhca8Wm5EafWGlG22YhvOow4r8+I"
    b"IlljflUy5gp1Y+rrGjPUyJj/Mzempa0xE5yMucfDmK5+xszpZ8zjg4wZOMKYlX8Z88okYw6fYcy2"
    b"ucZ8tNiYU1YaU269Md9vNubiHcbU2WvMkEMC64Qxrc8ZM+WyMQ/cNKb3fWOWPjHmxZfGHPZe8H82"
    b"5uOfxpweakylaGN+STDmqjRjinOMGVso9C83pnutMYuajHm+w5hD+4zZLiPmUyUxZ6mLKdIV86eR"
    b"mBvMxbSyFTPNScwjHmIG+YlZ10/MO4PEnDRCTPmxYn6eJObqGWKazRMzebGYh1aKGbBezNrNYt7d"
    b"IebkvWIqHhbz2wkx158T0/qKmBk3xTx5X0w+FbP1pZjP3os594uY2r/EjAgVc1e0mJ6JYpaniXkj"
    b"R8wJRUKvCoFRK8zULKZdp5g5fWKekzXhCGUT9qmb8L2uCVcam9DcwoRptiY86WzCwZ4m7PIz4ev+"
    b"Jlw22ISmI02YOtaEJyYLtZlCbZ4J3ywx4fJVgm+DiXC+mPD0ThPhPBeYh03430kTrj1vQturJsy7"
    b"ZcJLD0w47pkJFV6b8OcHE279akKP3yasDDPhvRgTzkoyoU6GCWNzTXio2IQDK03YWWfCty0mXN1l"
    b"QhtJU+bLmfKqiiknaZpSpG/KcLEp91qaMsjelK0upnztZcqVAaa0GShoh5jy2ihTThlvSo2ppsLZ"
    b"YspDC0zJZabsWW3KTxtNuWmrKd12m7JqvykfHTXlwtOmNL1oyqxrprx0x5QTH5lS7YUpo96Y8vBH"
    b"Uw7+bkrJYFN+izDltjhT+qaYsjlT6JtvyjWlpnSsNmV5gykftgmsHlOaS5sxV8GM10VmnKFtRn1D"
    b"M6aamvG8tRknOJpR3d2McT5mPBFkxtE0o9JwM0aMMePhiWYcOt2MsnPNGLzIjPtXmHHQOjNKbTbj"
    b"r+1m/OcfMw48ZEaJE2b8edaMey4L+5vC/r6wfyLsXwr798L+i6D/KehDzchoM0onmvFPmsDLMeOQ"
    b"IjPKVZgxrNaMR5rNOLLTjMoS5oyRNedJZXOO0zCnpp45k43NedHCnNPszGnoYs5sT3Pe9Dfn/AHm"
    b"tBxiztKR5nw8zpwrp5jTZZY5G+ab891Sc/692pwBG83Zs8WcP3eZc/9+cw47ak7l0+aMvWDOs9fM"
    b"OeWOwHxkztzn5rz7xpxLPprT8bs56/4IjAhzbokzZ78Uc0pmmTM035zHSoW5qs2p02jOzDZz3u4x"
    b"52JpCzooWrBeZMEP2hbcbmhBmFlQ3saCMY4WPOduwem+FjTtZ8FSWvD5cAtu+MuC/pMsKDHDgqFz"
    b"LXhisQUnrbSg0XoLFm624JMdFly/14J+hwXNSQuGnbPgqSsWnHLLgiYPBM5TC/77yoKbP1iw/1cL"
    b"yv22YFyYBS/FWHBekgXtMizYkGvBz8UW3FdpwdH1FtRptWBulwUfSVpynbwlA1QtKa1lyRh9S140"
    b"seQ8K0s6OFiy2dWS370teTjQkhNgSeNhliwdbcnXEyy5fZolh86xpMYiS2Ytt+TDtZZc/z9LBm23"
    b"pPw/lkw6aMmbxy25/KwlvS9bUuKmJaPvWfLyE0vhzLak23tLdn+2ZPhPS54PFXpGW9Ip0ZIdaZYM"
    b"ybHk2SJLzq0QcnWWbG8Wcp1CTsKK8+Ss6KxixU4NK4brWfGC2IoLLa3oZm/FXhcrRntZ8WqAFZcN"
    b"tKLPUCvKjLZi4ngr3plqxbWzrdh/oRVVl1sxa40Vn26y4tZtVhy+x4p6B61YesyK789Y8cAlK066"
    b"YUWre1ZsemzFP/9a8dw7oc9nK3r+tKJ0qBWToqx4L8GKG9OsODjHitpFViwpF/y1VjzYbMWpnVa0"
    b"k7Bmh6w1I5WteU3Dmqv0rNlfbE01S2sW2FnzjYs193tZC9/k1rQbaM3OIdaMHmXNm+OtuW6qNTnb"
    b"mjoLrVm2zJqf11jzxCZrzt1mTY891pQ9aM30Y9Z8dsaauy5Zc8INa1rfE/o9Fhj/WvP2O2tu/Gwt"
    b"3ImsaRRqzbooa/5JsOalNGGOHGsOLLKmdoU1K2qt+a3Zmmc7rblUwoZBcjbUULFhqYYNv+jZ8LTY"
    b"hostbRhob0N1VyHvJeQDbHhmoA2XDrVhv9E21Jpgw4qpNvwx24YXFtpw1XIbcq0NDf5nw7ptNgzd"
    b"Y8MbB2246bgNR521ocVlG3bcsGH8PRs+emLD3S9tOOW9DV2+2FDulw1zQm34PtqGJxKF3ulCj1wb"
    b"6hTbsLZCYNXZ8FaLDf/usuE4SVvaydtSUtWWmZq2fKtvy+MmtlxsZcsBDrbUd7Nlg7ctowJteR+2"
    b"3DXMltPG2NJjoi1VptuydI4tfy6y5dUVtty0zpZjN9vSfoctZfbaMveQLT+dsOX5c7Zce8WWI2/Z"
    b"0vqBLSWe2TLrlS3/+2DLs19tuea3LUeEC7VYYY5kW+ZkCL48W14oseX6KluOaRCYbbaU67FloZQd"
    b"fyjYCe9gO27RtuNkQzt6mNlRzcaO1Y52jHC34yNfO+7vZ8f5g+w4YIQdxWPt2DnJjukz7Phhnh3P"
    b"L7HjhlV2HLfBji5b7Ki6y45V++wYecSOj0/Z8dAFOy65Zschd+xo9ciO0i/sWPjGjr8+2vH2dzvu"
    b"Cbbj3EiBHW9Hk1Q79mbZMbfAjt/L7Hizxo67muw4p8OO/fvsaCprzz4le+ar2/Onrj3vGNtzr4U9"
    b"F9rZc5CLPa297CkXYM+yAfYMH2LPp6PseWy8PVdPtedfs+3pttCemsvt2bzGnqmb7Plxmz2v7rHn"
    b"joP2nHvcnjhrT6vLAuOmPSvu2TP6iT1fvrTn2ff2/N8Xe077Zc/AMHuaxthTKsmeJen2jMi154ti"
    b"e56utOemekHTas+gbnuaSTlQVsGBFaoOjNFy4BsDB140deB2awfOc3TgEHcH2vs6UNTPgU10YMZw"
    b"B37/y4H3JjnwyAwHrpnnwElLHBiwyoGmGwTWFgdW7XRgwj4H/nfEgTdPOXD/BQeuuObA8Xcc6PvI"
    b"gSYvBN1bB1Z/dGDidwd+CnbgnUgHHo534NpUB07JdmD/QgdalztQtdaBLU0OzO5wYHCfA5/LOvK8"
    b"siN3ajhysZ4jx4gd6WPpSBN7R8q7OrLOy5HpAY78NdCRT4c68txoR+6YIGinOfKvOY70W+RI8xWO"
    b"VF7nyJb/OTJ3uyPD/3Hkm0OOvH7CkQfPOXLdFUfOuOXIIQ8c6frMkYavHSnzn8D+6sjM344MCXfk"
    b"q1hHXksW9JmOXJ/vyFmljhxe7UjPRkeatjtSqdeRrdJOLFR0YqyaEz/rOPGhkRPPmjtxl60TVzo7"
    b"caqnEwf7O9FtgBPFQ5yoNMqJbeOcWDzFiQmznPh9gROfLXPilTVOPLjJiRu3OXH+HieOPejEoONO"
    b"dDjrRP3LTsLvwInN95xY9ETwvXTij/dO/PeLE6//cuKxMCdui3HisiShX4YTh+Y50bvEidZVTtRu"
    b"cKJ0mxObuoU5pZyZqODMXyJnvtZ25h1DZ54xc+Y/Ns5c7+TM+R7OnODnTPZ3pudgZ1qOdKb2OGfK"
    b"THFmy0xnls53ZtpSZ4avdubnjc58ttWZ13c78+QBZ+4+JjDOOHPBJWfhfHDm0HvO9HviTIeXzjR+"
    b"70y1L86U/OXMplCBE+3M9ERnRqU783uuMEuxM+9XOvNSvTOPtjpzZ7cz10m5cJGCC6eKXDhK24X9"
    b"DV2E94ELbWxcaOjkQpGHC6X8XNjaz4WVg1yYO8KFSWNdGDbZhV9nuvD1fBc+WOrCq6tdeGqjC/dv"
    b"deHW3S5ce0DgHnPh9DMuHHvJhYNvuDDgngvdngjsly40eu9CjS8uVPjlwt5QFzZHC/xEFxakuzAt"
    b"14WxxS4MqXThl3oXvml14ZNuF96WcuUlBVeeErnyoLYrdxq68n9mrlxl48pFTq6c5eHKSX6uHN3f"
    b"lYMHuzJopCu9xrnSaYorrWe5UrzAlbrLXCla40r5Ta6U2ObKjt2ubDzgyqpjriw+48qcS65MveEq"
    b"nAuujHjiyj8vXfntvSs/fHHlq1+ufBrmyvsxrryZ5MorGa48l+fKEyWuPFzlyr0NwkxtrtzS48qN"
    b"0m5crejGZWpuXKTjxrlGbpxh7sbJtm4c7+zG0Z5uHO7vxsED3DhgiBsDR7nRd7wbPae60XW2G50W"
    b"utFuuRut1rrR/H9uFG93o+E/btQ75EbtE27UOOdG0RU3Kt9yo8IDN8o+c6P0azdK/OfG3q9u7Prt"
    b"xo5wN7bFurEl2Y1NmW5syHdjXakba6rdWNXoxsp2N5b3urFMxp2lSu4sVndnka47C43dWWDhznw7"
    b"IVzcmeflztwAIQa6M2eoEKOFmODO7GlCzBFikRArhNw6ITYLsUPQ7hXisOA9KcR5gXNVYN4W2A+F"
    b"Hs+FXm+Enh/dWfbdneXB7qyMdGdVvDtrUt1Zl+3OhkJ3Npa7s7nWnW3N7uzodGeXhAd75DwooepB"
    b"KS0Pyhp4UMHUg8rWHlR19BC+Hzyo5etBvX4eNBzkQfEID5qP9aDVZA/azfSg43wPui71oOdqD/pu"
    b"9GDgVg8O2O3BQQc8OPyYB0ef8eC4Sx6cfMODM+55cM4TDy586cFl7z24+osHN/zy4N9hHtwZ48G9"
    b"SR48lOHBE3kePFviwctVHrzZ4MH7bR582uPBV9Ke/KDoya9qnvyj48kII0/GmXsy1daTOc6eLPL0"
    b"ZKW/JxsGeLJ9iCf7RnlSboInRdM8qTPHk+JFnrRa4UmndZ703OzJwB2eHLTXk6MOe3LiSU/OPO/J"
    b"RVc9ueq2Jzc99OTO55488MaTJz968tJ3T94O9uSTSE++iffkl1RPBmd7MqbQk2nlnsyvFXo3e7Kp"
    b"05M9El6Ul/eihqoXjbS8aG3gRVdTL/pbe3Gwoxf/cvfiNF8vLuznxTWDvLh1hBf3jfXiyclevDLT"
    b"iw/me/HVUi9+We3F0I1eTNzqxdzdXqw44MWWY16UPOtF1cteNLgpsO970f2pF/u98uLID16c8lXg"
    b"/vbi2nAv7oj14pFkL17M9OK9fIFZ6sVv1V6MbPRiWrsXS3q92CjjTQllb6pqeNNIz5v2Ym/6Wnpz"
    b"iL03J7p6c763N9cFenMXvHlimDevjfHm04ne/Djdm2FzvZm62JslK73ZvN6b0lu8qbXLmxb7vel+"
    b"1Js47c3xF70577rAuOvNPY+9efpfb+EO681Xn73586c340O9WRDtzYZEb0plCP48b1qWeNOrSpih"
    b"wZtT2ry5tMebW6R9eFTRh9fUfPhcx4ffjXwYb+7DAlsfNjn7UNbLh3oBPrQf6MPAoT4cM9qH8yb4"
    b"cMM0Hx6Y48NLi3z4dIUPv63zYdxmHxbu8GHLXh8qHvGh8Skful7wIa/5CN+LPlz+yIc7X/jwzFsf"
    b"Pvjkw48/fBgdIvSKEjwJgifdh+JcH3oU+3BopQ9n1vtwbasP93f78IqUL/9V8OUfkS/TtX1ZY+hL"
    b"aXNf6tv60sXZl4M8fTnd35drB/hy/xBfXh3ly1fjfRk61Zc5s33ZtNCXiit8abbOlz6bfTlmhy8X"
    b"7vXltsO+PHPSl4/P+/LHVV+m3vZl7UNfyr7wpfFbX3p98uWoH4I2xJfbo3x5LsGXz9KEOXJ8mVUk"
    b"cCt8qVzvS6tWX+Fe5cvJUn5creDHgyI/3tL240dDPyaY+bHSxo/Szn409vSjt78fxw7w47Ihftw7"
    b"yo/Xx/vx/VQ/xs32Y8VCQbfCj+J1fvTb7McJO/y4aq8fDx32452Tfvx63o9pV/3YcNuPKo/8aPvC"
    b"j3zrxzmf/Lj1hx/Ph/jxVZQfoxL8WJYmsHL9aFLsx4BKP06p9+OGVj+e7PbjMyl/hin4s0jkTwkd"
    b"fxob+dPf3J9TbP250dmfpz39+cLfn5ED/Fk2xF/4LvSnxQR/Dpjmz9lz/LltkT8vr/Dnh3X+TN7s"
    b"z8Yd/lTf50+XI/4cfcqfKy7488g1fz6648/QR/4seeFP6XcC47M/8dOf80L9uTvan7cS/fk93Z85"
    b"uf7sKRbmqPJnUIM/Z7X5c3uPP69JB/CLYgCz1ALYrRNAY+MABlkEcLZdAHe6BPCmVwB/BAQwb2AA"
    b"JYYF0HxMADkxgAunB/DA3AA+XBzAsJUBrFgfQKUtAXTaFcC/9gdw3dEAnjkdwHcXA5h6PYDtdwNo"
    b"+ERgvwzg3PcB/OdLAO//ErxhAayMCaBKcgBdMwM4MT+Am0sDeKU6gF8bA5jfHkDpvgDaygZypHIg"
    b"12gE8qxeID+IA5lpGche+0BauAVymE8gVwYF8jQD+W54IDP+CmTPJKE2U6jND+SqpYE8s1rwbQxk"
    b"9tZASu4JpM3BQI4+HsgNZwN56XIgv90MZNH9QCo+C6Tr60BO/i+Q278F8u6fQIZHBLIuLpA6qYEM"
    b"yg7kgsJAHikP5KvaQKY1C3N0CjzJII6RD+Im1SD+v+HyDu/xauN4jdoNam9i1x4xf/Pz23vvHVut"
    b"2iMae0tKBZWgIjRSxKhRYjRRrxgNWjsoQa1arVGE9/nr3Oe+v+Ncz7mec+6TUUtEQQMRj5qJqNVG"
    b"RL+OIhK7i1jUR8QuiYirShHlDCI62EQ4PCKSQiKyBog4M0zEq9Eimk0UoU0S8dUsEWsWCFopIh6v"
    b"EFE3XYR0g0j410Us3yYib7eIez+LqHFU8DguYuBpESnnRey/LKLkhojP7oro80jwfi5i6WshXyrk"
    b"y4mJqyKmbw0xA+uKSW0s5kC8mHvtxNTsIkacIGZofzEr5GKOasQ8Nomp7xSj9IsZExWTMVhM4Qgx"
    b"L8eKiZ8ixpwsZtpcMdmLxVxYJqbsajFd1okJZolZmCNm3w4xd/YK2ofESAvEjCwUs6ZIzIkLYl5d"
    b"E9Pqthj7fTEzn4jJ/VfM9bdiqpWR0K+ihGGfSVhdS8LxBhJeNpPQqo0ER0cJs7tL2NVHwi2JhJoq"
    b"CTKDhDE2Ces9EuE/l/BxgIQuwyVExkhInSjhSJKEZ7MkNF8owZYqYWaawE+XULJBQq1sCcrtEib8"
    b"JGHzAQmXj0qo/D/B/4yEEb9LWHtF0Lwpocw9Cd0fSxj4QsLKNxJOfJDwtryUTlWlRGpKWVZPyrEm"
    b"Ul61lNL+CynBrlJSe0nJF0l5iZR2OikBi1S4L4VcQMjFBNxQATdKyjfjBe5UKW9mSOk4X0p0qVR4"
    b"60s58Z2U0vVSum2WMmirlDW7pBTtl1L+iJQ+v0oZeUpK5jkply9J+eyGFO5ImfRQyrZnUm6/klK/"
    b"VIq5nIw5lWUcqC7jWR0ZbRrLCMbLWN5ORmFnGR97ykjoL2OEXMZGjYyrJhk1nDI0fhnJURl7B8v4"
    b"e4SMVuME3hQZ3ybLODVXRtklMvotlzF2tYycdTJuZ8lo+KMM+04Zi/bJKDgk412BjJ4nZYw8K2PT"
    b"RRk3imXUK5FhfSBj4VMZ+S8FzDvBv6ycUZXkZMfJuVVbTqNGclwt5KS2lVPYSU65nnLE/eRMksnZ"
    b"qZbz2CinjUNOzCcnPSIX7nM5NUfIMY6VM3+ynPyv5ZTOkdNnsZzxy+TkrpLzaK3AyZKTmCNn3Q45"
    b"V/fKqXtIjr1ATkqhnJNFcipclKMolpN8W07efTlvnshJeCln7DtBoww8rgjt42BwbchqCLeaQ9O2"
    b"EBSecEILxqW+IHxiHGpYZoQiO3zmA0MEFg2CE19CxbGgmgxzvoaCOVB2MciXwYxVcGQtfNgI4hyY"
    b"vgPy9sL7POhfANMK4WARvLsgzIuF+W1hfl+YPwHRS0h6B4fKKCitqEASpyC5toKjDRWUaaGAtgpm"
    b"d1JwrIeCT/sp0MgULFArOGlUUNWhwORTkBpRcG6QglojFLjGKlg1WcHVrxU0nqsgvFjBhmUK7qxS"
    b"0GadgmFZCrbmKHi6Q0H3fQomHFKwv0DB+0IF0rOC10UFx4sVVClRYH6gYPlTBRdeKmj4XtAqqySz"
    b"kpJ7cUo61FEyupGSn1ooedNWibizklk9lfyvn5JqciU2jZKVJiXXHEpa+JUMjir5cbCS5yOUQt+u"
    b"ZPoUJfnJSirNU2JeoiRtuYBdLWDXKxm6Scn2H5W83Kmk/35B97CSwmNKapxS4jmnZN0lJXevK+l4"
    b"R8n4h0rhfakU3u1K9KVKlpdTcbWyivgaKobXVbGrsYq38Sph81Us6qLi9wQVjUQqBqJim1bFS7MK"
    b"iUvF/ICKszEVDYaqGDBKxdbxQm2qCulMFQvmqzi/VEXjFSoGrVGR+72K/zarUGxTkbJbxeWfBb+j"
    b"KkYeV7HvtIqyv6swXlGx6qaK23dVdHqsYvILFQVvVFT/qML/qZpNVdU8r6lGVF/NgqZq/milpnkH"
    b"NV92U7Ovt5ryEjVWpZoMvZr7VjUJHjUzQ2p+G6Cm0XA1Q8ao+WmimjLT1ZhnC7iFah6kqum1Us2c"
    b"DDXnMtU026JmRK6aA3vUVMpT485Xk3VCzYvf1MguqEm5pub6LTUd7quZ+kRN4b9q6r1TM7iMhj0V"
    b"NZSP0+CorSGzoYbnzTXI22r4ppOGP3to6NJPQ7JMQ5FaQzOThtEODUd8GqpHNUQGa8gdoeGTcRqs"
    b"UzRsSBb4czWwRMO3yzXcWa0hYb2GeZs0XPpRQ/tdGqbu13DqsIYmv2oYdUrD0XMaPr+sYeANDXvv"
    b"aKj0SIP/uYatrzV8LNVgK68lq4qWVzW06OppyWii5UlLLfIvtKzoquWvXlr6i7WkKLTc0mlJsGpZ"
    b"4NZSHNTSdYCW2cO0XBqtpcNELclJWv6YpaXdQi1JqVrOpWlpnaFlaqaW37K1tMzVMnmPljMHtcTn"
    b"a5l0Qsvp34T4ghBfE/K3hPi+gHki4P8V8O+0TCmjo6iijtZxOpJq6zjfUEe7FjqS2+q40ElHx546"
    b"ZvfTcUWmo6tGx3yTjhsOHQl+HUujOu4M1tF/pI5vx+l4OEUnHDI61szT8WKJDv23OjK/0/F2vQ77"
    b"Zh05W3WU260j8LOOn47oqHZcx6DTOg6f11Hvio4xN3UU3tUR/1jHtBc6/nijo/NHwfNTPbeq6un/"
    b"uZ4V9fU8aapH21pPZgc977vp8fTRs1Oip6pKz2CDnqM2PY28eiaG9ZwbqKfjl3rmf6Xn9iQ94q/1"
    b"rJ6j559FeizL9OSs0lNhnZ7ELD2Hc/Q02Klnwj6Bd0hP52N6Fp3Uc++sHsUlPeuvC34lerwP9ex5"
    b"pufz13pGl+o5Xc5A+yoG5tUwUFLXgLyJgXUtDbxvb8DX1cC+XgbqiA2MUxg4pzPQxWpgqdvAo6AB"
    b"/QAD2cMMVBxjYNBEA8eSDLScbWDmQgN/phqQrjSwPsPAh0wDoS0G8nINNN5rICnPwLV8A/0LDaQX"
    b"GYQz20Cg2MDB2wYaPTAw7amB4pcG4U1jIKOskdJKRsLVjRypY6RZYyMz4o3camdE0cVIVoKRT0VG"
    b"hmDkhNbIFxYji11G/g4YsSQa2TnUSK3RRiZMMHJpmpG+s4xkLDDyIcVILM1IQbqRNplGFmYbebxd"
    b"4OwxsuugkTr5RiafMFL8mxHpBSMbrxmpcNvI8PtGfntipNtLI2nvjPxXxkS4kon8OBNt6phY1MjE"
    b"0xYmHO1M7O9sokmCiVn9TfwlN2HUmthpNlHXZWJawMSfMRPqoSZ+HGWi5gQTk6aZKJ5pggUmslNM"
    b"xKWZmJBu4toGE/JsEz9sF3J7hNxBIfeLgDthYstvJmpcELjXTNy4ZUJ138TWJyZqvxQ83pkoKWNG"
    b"X8nMrjgzDeuYmdXIzMMWZuztzBzobCY+wcyi/mb+kZsJaM0cM5vp5DKTFjDzIWZmyFAzZ0eZ6TvB"
    b"zMZpZqrNMjNhgZmbKWa0aYJuupnGmWbmZpt5ut2Mb4+ZgoNmOuebWX3CTNkiMyMvmLl0zQy3zWy9"
    b"b6beU2EdL838/c6Mt6yFgkoWOle3sKaOhU8bW/gq3kJxOwvaLhZ+SrDQTGRhMRZeay0MtFg467Ig"
    b"ClrITrRQZ5iFmaMtPJlgIZBk4X+zLPRcaGFDqoW4lRaSMiw8yLTg3mLhWK6FbnstrM+zUK3AwtRC"
    b"C/eLhNpFoVZsoXuJhe8fWKj+zML0VxYevbfgL2elsLKV3jWsbK5rpU4TK3NaWvmnvZXErlbO9bIi"
    b"F1vZobDSXG8l1Wrlo9vK6JCVmwOsWIZbOTLGSpdJVtZPt1J9jpUZi6w8+8ZKbJWV82utQrNkZXeO"
    b"lVY7raTts1LhsJVJx6w8OGnFf87KqUtWxDes5N4RPB5ZWf7cSvk3ViZ+sHK/vA1/VRuna9qQ1rex"
    b"s6mNVq1trOxgo3J3G0l9bDyR2IipbPxhsKG12zjotdE5YmPDIBt1RthYMNbG+8k2xiTbuDPXhmeJ"
    b"jVPLbci+s/HTehvtNtvI2Gqj5m4bc3+28faIjdHHbZSctuH93caZKzb408a+ezY6/W0j8x8b9d7a"
    b"WPqJnbIV7Uz+zM6TWnYGNrRzpbkdS1s7xzvZEfW0s6ufnXZyO+s1duqY7Sxx2ikXsDMlZufZEDtD"
    b"Rtm5Md6Oe5qdMzPtKBfYOZhip0eanR/T7cRn2lmTbefzXDuL9gh+eXam5tt5ccLOcKHBLLlgF84W"
    b"O3/ctmN8YOfXp3Ykr+zse2+nazkHWyo7hD7GQXpdh7C3DlJaOqj0hYOZXR286+VggtjBU4WDoXoH"
    b"JVYHIY+DSyEHtoEOTg93oP7KwS+THPT/2sHeOQ66LXawdZmDtqsdZK5z0GSTg9U/Oqi9y0HqfgdV"
    b"jziY96uDsqcdfH3ewdvLDibedPDiroNRjx08fOFg8H+C10cHkQpOiqs58dZycqGBE3tzJ0VtnBg7"
    b"OSns4UTdz8kxmRO5xslhkxOR08kBv5PeMSd7hjjpMcrJzvFOukxzsm2mk44LnOSkOGmX5uSHdCet"
    b"M51kZTuFu9dJ5h4nLfKcfJ/vpFmhk3VFTppcdLK22EnjEicZD5w0euYk/ZUwlgpjOReNqrhIryGM"
    b"9YSxiYvGrVxkfOGiSTcXa3u7aCpxsV7pornBxQabi3ivi41hF60Gudj0pYu2Y11kT3bxRbKLrXNd"
    b"dFriIne5i27fudi93kXCZhf7trrot9tF3s8upEdd/HLchfKMi+O/u9BddXH6TxeWv1yc/9uF618X"
    b"V966CJZxc7OimwFxbu7VdjO8kZsnLdyMbefmVWc3UxPcfOjvZhZuKujcLLa4qe52syLopv4AN2uH"
    b"uYkf4+aHiW46Tnezc7abXovcHPzGjXyVm+Nr3Riz3JzLcePZ6eb6PsHrsJsHx9yMPuXm5Tk30y67"
    b"KXPTzYK7buIeu0l74abxf24yP7ppX8HDjmoeetfycLiBB1VzD2faeHB28nCth4cB/Tw8knkYq/Hw"
    b"1uRhltNDlYCH5TEPDYd6yBzlocMED7uneRDN8nBsgQdTqoeLaR4iGR7uZ3r4aovAzfUwe6+Haoc8"
    b"rCzw0Oykhy1nPXS/5CHvugf1HQ9nH3qEns/DndceRn3w8Ka8l9lVvXz2uZfV9b3EN/OyrbWXPh29"
    b"FHT3Yurr5YrUy0C1l6dGL1MdXir4vSyLemkyxEv2SC89x3s5OtWLcaaXy/MFbIqXZyu8JKV7qZzp"
    b"JS1b0M31sn2Pl/55Xk7ke3EVerld5GX0RS/vi70sLPFS76GXrGdeur32cqRU8C7v42oVH0Nr+nhV"
    b"z8ecpj4+b+1jQwcfXbr7ONTHh1EqYFQ+hhl9vLH7mO/zUSfqI2uwj54jfRSM82Gf6uPWDB9fzffx"
    b"SYqPb1b4aJ7uI3eDD2m2j6LtPiJ7fDw96GNGvo8ahT6+L/LR7aKPX4oFfomPkgc+xj/zUf61j7RS"
    b"H23L+9lfxY+upp+r9fyMaOqntJWf1A5+WnT3s7uPH6XUz0WVn6FGP+/sfpb6/DSL+tk1WKiN9HNp"
    b"nJ9hU/28nyHw5vuJT/GzZ4Ufbbpf6AH8jM72UzbXT9oeP+3z/BzK92Mt9HOnyM/ki36qXffzfYmf"
    b"ng/9nHjmJ/Taz/NSP/PKB2hYNUBuzQDK+gGuNA0wqnWAch0DrO4eoFPfAPnSAF51gL+NAWY7AtT3"
    b"B9geFfBDAlwdGWDM+AAVpgXImBmgx4IAhSkBImkBXqUHWJIZoOWWAAdyA9j2BrifFyC5IEDdkwG2"
    b"nQ2guhSg+HqA8XcCVH0UYOPzAP3fBDj/IcDwT4OUrRZkzedBejQIcqpZkAFtgrzvGGRFjyCd+gU5"
    b"LgsS0QR5bQryjTNI+0CQgliQ4NAgL0cFSZ0QpF1SkPxZQQILhVyqkFsp4NYKuI1BQjkCd0eQ5fuC"
    b"dDwc5H/HgsROCR7ngqy6HKTbzSCn7wYZ8lhYyz9B1v0XpO8nIf6oEGLMZyGq1g7xQ8MQihYhbrQN"
    b"MbVziLoJIXb1D2EmxCNtiPmWEC3dIY4GQwQHhPhvWIiVY0L0mBTi7PQQI+eEqLJY0FkWQrk6xO11"
    b"IZI3hWiyNcSBXSG8P4d4dSTEiuMhup8ROL+HGHU1RLVbIXL+CqF7EuKvf0PMexeiddmw0DOFSawe"
    b"pkzdMN83DiNtGeZm+zBfdw3TpHeYQ+IwQWWY9/owGbYwIm+Y4nCYpEFhGo8IkzdWqE8R6slh1s4L"
    b"I14a5sa3YZLXhGm2IczRH8LEtocpuydM1sEwqvww906EmV8Upv3FMKeKw4woCRP3MMyOZ2Hsr8P8"
    b"WxpmZfkIfapGuFozwvT6EZo1i5DfOsLAjhEq9oiQ0zeCSRbhmTrCt6YIvZwRrvgFbCxC86ERjo2K"
    b"MGRChKpJEXJnRbAvjPAqNcKalRGkayOUbIwwPydCh50Rzu6LMP5whAa/Rjh8KsKA8xEqXYmw7abA"
    b"uxfh9eMIGf9EkL+N8NcnUZZUjNI9Lsrl2lG+bhSlZXyUwnZRRneJUrdXlDxRlERFlMr6KLnWKG5P"
    b"lPehKBsHRtF/GeX5V1FWT44iTY5yb26UpUuiJHwbpfi7KHO+j9Lhhyi/b4sy7SdB+2CUU79EGXci"
    b"SqOiKAUXoowojlK7RPB5EGXQsyhxr6PsK40SKR+jUtUYO2vG8NePUb5ZTDhzY7g7xvikR4wtfWM4"
    b"ZDFK1TF+MMWwOWO888fYFIthGRrjv1ExNk6IYU6K8WaWEC8U4m+EeKUQrxXiLAGTEyNrZwzrfoF7"
    b"OMbmX2PYTwua5wX9KzFcf8Yo85fg+3cM378xPn0nrKdMIqFKiVSpnsjeOokMaJxI9ZaJ5LVPZFjX"
    b"ROr2TqRAnMgYZSJNDImctCUyyZtI60gi5wclMmNEIp3HJXJtSiILZyTSe34id5cm8n8csDEl"
)
_AR_ALH_DAILY = np.frombuffer(
    _zlib.decompress(__import__('base64').b64decode(_B64_ALH_DAILY)),
    dtype='<f4').astype(float).copy()
_NV_ALH_DAILY = np.round(np.arange(10001) * 0.01 + 160.0, 6)
_DAILY_ALH_LOADED = True
del _zlib

# ── Tablas de hidrogeneración Madden — datos oficiales ACP ─────────────────────
# Fuente: Tablas_Hidrogeneracion_Madden_Alhajuela.xlsx
# Tabla de agua:  CFS/MW al máximo disponible por nivel Alhajuela (ft)
# Modesto:        Qmax(cfs) / Pmax(MW) por nivel (col3/col7)
# Bill Shaw:      Qmax_3turb(cfs) / (3 × Pmax_MW) por nivel (col3/(3×col7))
# Rango: 190–255 ft  |  66 puntos  |  cada 1 ft
_MAD_NV   = np.array([190.0, 191.0, 192.0, 193.0, 194.0, 195.0, 196.0, 197.0, 198.0, 199.0, 200.0, 201.0, 202.0, 203.0, 204.0, 205.0, 206.0, 207.0, 208.0, 209.0, 210.0, 211.0, 212.0, 213.0, 214.0, 215.0, 216.0, 217.0, 218.0, 219.0, 220.0, 221.0, 222.0, 223.0, 224.0, 225.0, 226.0, 227.0, 228.0, 229.0, 230.0, 231.0, 232.0, 233.0, 234.0, 235.0, 236.0, 237.0, 238.0, 239.0, 240.0, 241.0, 242.0, 243.0, 244.0, 245.0, 246.0, 247.0, 248.0, 249.0, 250.0, 251.0, 252.0, 253.0, 254.0, 255.0], dtype=float)
_MAD_AGUA = np.array([139.3878, 140.0000, 138.6538, 137.5472, 138.1818, 136.9643, 135.7895, 134.6552, 135.3333, 133.0081, 134.7619, 133.5938, 132.4615, 131.2121, 130.1493, 128.9706, 129.4286, 128.3099, 127.2222, 127.4324, 126.2667, 122.9139, 124.1558, 122.8205, 121.7722, 120.7500, 119.8765, 118.1818, 117.5904, 115.9763, 115.0877, 114.2197, 114.6591, 114.1573, 113.3333, 110.4972, 109.9454, 108.8649, 109.5745, 108.8421, 106.4921, 106.2500, 105.4639, 104.5918, 103.8384, 103.1000, 102.3762, 101.7647, 101.0680, 100.3846, 98.3732, 98.1905, 97.5472, 96.9159, 95.0698, 94.9074, 94.3119, 93.7273, 93.1532, 92.5893, 90.9333, 90.8850, 90.3509, 89.8261, 88.3117, 88.1897], dtype=float)
_MAD_MOD  = np.array([146.9276, 144.0574, 142.0333, 140.0748, 137.5725, 135.7696, 134.0240, 132.3337, 130.2335, 128.4650, 126.7762, 125.3381, 123.9447, 122.5949, 121.2870, 120.0200, 118.5589, 117.3899, 116.2575, 115.0038, 113.9601, 113.0078, 111.9716, 111.0252, 110.1094, 109.2237, 108.3672, 107.5355, 106.7394, 105.9785, 105.2403, 104.5281, 103.8798, 103.2256, 102.5959, 101.9301, 101.3411, 100.7754, 100.3145, 99.8017, 99.2151, 98.6403, 98.1795, 97.7403, 97.3223, 96.9255, 96.5496, 96.1945, 95.8600, 95.5460, 95.0929, 94.6497, 94.3819, 94.1342, 93.7259, 93.3269, 93.1241, 92.9410, 92.7779, 92.6346, 92.2934, 91.9609, 91.8616, 91.7824, 91.4822, 91.1904], dtype=float)
_MAD_SHAW = np.array([147.7679, 146.3908, 145.0405, 143.7163, 142.4175, 141.1433, 139.8930, 138.6660, 137.4616, 136.2792, 135.1183, 133.9781, 132.8583, 131.7582, 130.6773, 129.6152, 128.5713, 127.5452, 126.5364, 125.5445, 124.5692, 123.6098, 122.6662, 121.7379, 120.8245, 119.9257, 119.0411, 118.1704, 117.3133, 116.4695, 115.6386, 114.8204, 114.0145, 113.2208, 112.4389, 111.6685, 110.9095, 110.1615, 109.4244, 108.6978, 107.9817, 107.2756, 106.5796, 105.8932, 105.2164, 104.5489, 103.8905, 103.2411, 102.6005, 101.9685, 101.3449, 100.7296, 100.1223, 99.5231, 98.9316, 98.3477, 97.7713, 97.2023, 96.6405, 96.0858, 95.5380, 94.9970, 94.4627, 93.9350, 93.4138, 92.8988], dtype=float)

def madden_cfs_per_mw(nivel_alh_ft: float, metodo: str) -> float:
    tbl = {"Tabla de agua": _MAD_AGUA, "Modesto": _MAD_MOD, "Bill Shaw": _MAD_SHAW}
    return float(np.interp(nivel_alh_ft, _MAD_NV, tbl.get(metodo, _MAD_AGUA)))


# ── Constantes físicas del modelo de esclusajes (ConsumodeAguaEsclusas.xlsb) ─
# Fuente: hoja NeoPanamax (CC=Cocolí, AC=Agua Clara)
AC_NPX      = 26841.0       # m²  Área de cámara NPX
EQ_CC_m     = 18.407        # m   Nivel equivalente de equilibrio — Cocolí
EQ_AC_m     = 17.679        # m   Nivel equivalente de equilibrio — Agua Clara
FRAC_TINAS  = 0.60          # 60% de ahorro por tránsito con tinas activas
# Fuente: hoja Panamax
AC_PNX_REG  = 11132.878     # m²  Área cámara PNX Regular
AC_PNX_COR  = 10136.590     # m²  Área cámara PNX Corta
EQ_PM_ft    = 16.611        # ft  Nivel equiv. PedroMiguel (PNX)
EQ_GA_ft    = 17.830        # ft  Nivel equiv. Gatún (PNX)
CALIB_VPX   = 0.21407       # hm³ Vol/tránsito PNX regular @ H=87.5 ft (referencia histórica)
CALIB_H_REF = 87.5          # ft  Nivel de referencia de calibración PNX

# ── Consumo de agua por tránsito dependiente del nivel Gatún ────────────────
# Fuente: ConstantesConsumoEsclusas.xlsx · escenario NoSaving.
# Modelo lineal: volumen_hm3_por_tránsito = m * nivel_Gatún_ft + b
_NPX_NOSAVING_M = 0.009642629250000024
_NPX_NOSAVING_B = -0.3802567154100021
_PNX_NOSAVING_M = 0.006788340386202528
_PNX_NOSAVING_B = -0.38342433415759003

def _vol_lineal_por_nivel(H_ft: float, m: float, b: float, minimo: float = 0.001) -> float:
    """Volumen por tránsito (hm³) calculado con constantes oficiales y nivel Gatún."""
    try:
        return max(float(m) * float(H_ft) + float(b), minimo)
    except Exception:
        return minimo

def _npx_vol_base(H_ft: float) -> float:
    """Volumen base por tránsito NPX sin tinas (hm³), dependiente del nivel Gatún."""
    return _vol_lineal_por_nivel(H_ft, _NPX_NOSAVING_M, _NPX_NOSAVING_B)

def _pnx_vol_base(H_ft: float) -> float:
    """Volumen base por tránsito PNX regular (hm³), dependiente del nivel Gatún."""
    return _vol_lineal_por_nivel(H_ft, _PNX_NOSAVING_M, _PNX_NOSAVING_B)

def _pnx_ahorro_cc_per_transit(H_ft: float) -> float:
    """Ahorro por tránsito usando Cámara Corta vs Regular (hm³)."""
    # ΔAc × (EqPM + EqGA) × conversión ft→m (verif. vs xlsb: ≈0.01003 @ 87.5 ft)
    return (AC_PNX_REG - AC_PNX_COR) * (EQ_PM_ft + EQ_GA_ft) * 0.3048 * 1e-6

def area_desde_nivel_gat(nivel_ft: float, daily: bool = False) -> float:
    if daily and _DAILY_GAT_LOADED:
        return float(np.interp(nivel_ft, _NV_GAT_DAILY, _AR_GAT_DAILY))
    return float(np.interp(nivel_ft, _NV_GAT, _AR_GAT))

def area_desde_nivel_alh(nivel_ft: float, daily: bool = False) -> float:
    if daily and _DAILY_ALH_LOADED:
        return float(np.interp(nivel_ft, _NV_ALH_DAILY, _AR_ALH_DAILY))
    return float(np.interp(nivel_ft, _NV_ALH, _AR_ALH))

# ── Logo helper ──────────────────────────────────────────────────────────────
def _logo_b64(path: str):
    if os.path.exists(path):
        ext = path.rsplit(".", 1)[-1].lower()
        mime = "image/png" if ext == "png" else "image/jpeg"
        with open(path, "rb") as f:
            return mime, base64.b64encode(f.read()).decode()
    return None, None

def _img_tag(mime, b64, style=""):
    if b64:
        return f"<img src='data:{mime};base64,{b64}' style='{style}'/>"
    return ""

_logo_mime, _logo       = _logo_b64("LOGO_HIMH.jpg")
_logo_cp_mime, _logo_cp = _logo_b64("CP_RGB_p_Ver.jpg")

def f3u(hm3):
    return f"{hm3:.3f} hm³/d · {hm3/CFS2HM3:.0f} cfs · {hm3*HM3D2M3S:.1f} m³/s"


def fmt_sig(value, sig=3):
    """Formatea un número con cifras significativas y evita residuos binarios (p. ej. 27.0999999)."""
    try:
        x = float(value)
    except (TypeError, ValueError):
        return "—"
    if not np.isfinite(x):
        return "—"
    if x == 0:
        return "0." + ("0" * max(sig - 1, 0))
    decimals = sig - int(np.floor(np.log10(abs(x)))) - 1
    if decimals > 0:
        return f"{x:.{decimals}f}"
    return f"{round(x, decimals):.0f}"


def tbl(usos, total, nombre, dem_t):
    rows = []
    for nm, (h, c, _) in usos.items():
        rows.append({"Uso": nm, "hm³/día": round(h, 4), "cfs": round(c, 1),
                     "m³/s": round(c*CFS2M3S, 2),
                     f"% {nombre}": round(h/max(total,.001)*100, 1),
                     "% Sistema": round(h/max(dem_t,.001)*100, 1)})
    rows.append({"Uso": "TOTAL", "hm³/día": round(total, 4),
                 "cfs": round(total/CFS2HM3, 1), "m³/s": round(total*HM3D2M3S, 2),
                 f"% {nombre}": 100.0, "% Sistema": round(total/max(dem_t,.001)*100, 1)})
    return pd.DataFrame(rows)



# ── Evaporación desde archivos normalizados de Aquarius ──────────────────────
EVAP_AQUARIUS_FILES = {
    "mm_gat": "Evapo_Rate_Daily_Tank_CZL.csv",
    "mm_alh": "Evapo_Rate_Daily_Tank_PMG.csv",
    "hm3_gat": "Total_Storage_V_Evap_Gat_0_85_GAT.csv",
    "hm3_alh": "Total_Storage_V_Evap_Alha_0_85_MAD.csv",
}


def _buscar_archivo_data(nombre: str) -> str | None:
    """Busca un CSV en /data junto al app, en el cwd y como respaldo en la raíz."""
    base_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in globals() else os.getcwd()
    candidatos = [
        os.path.join(base_dir, "data", nombre),
        os.path.join(os.getcwd(), "data", nombre),
        os.path.join(base_dir, nombre),
        os.path.join(os.getcwd(), nombre),
    ]
    vistos = set()
    for path in candidatos:
        path_abs = os.path.abspath(path)
        if path_abs in vistos:
            continue
        vistos.add(path_abs)
        if os.path.isfile(path_abs):
            return path_abs
    return None


@st.cache_data(show_spinner=False)
def _leer_ultimo_valor_aquarius(path: str, mtime_ns: int) -> dict:
    """Lee el último valor numérico válido de un CSV normalizado por download_data.py."""
    del mtime_ns  # Solo se usa para invalidar la caché cuando cambia el archivo.
    df = pd.read_csv(path)
    if df.empty:
        raise ValueError("archivo vacío")

    columnas_l = {str(c).strip().lower(): c for c in df.columns}
    col_valor = next(
        (columnas_l[k] for k in ("valor_raw", "valor", "value", "result") if k in columnas_l),
        None,
    )
    if col_valor is None:
        # Respaldo: primera columna con al menos un valor numérico.
        for c in df.columns:
            serie_num = pd.to_numeric(df[c], errors="coerce")
            if serie_num.notna().any():
                col_valor = c
                break
    if col_valor is None:
        raise ValueError("no contiene una columna numérica reconocible")

    valores = pd.to_numeric(df[col_valor], errors="coerce")
    validos = df.loc[valores.notna()].copy()
    if validos.empty:
        raise ValueError("no contiene valores numéricos válidos")
    validos["_valor"] = valores.loc[validos.index].astype(float)
    validos = validos.loc[np.isfinite(validos["_valor"]) & (validos["_valor"] >= 0)].copy()
    if validos.empty:
        raise ValueError("no contiene valores no negativos válidos")

    col_fecha = next(
        (columnas_l[k] for k in ("fecha_inicio", "fecha_fin", "timestamp", "date", "fecha") if k in columnas_l),
        None,
    )
    fecha_dt = None
    fecha_txt = "N/D"
    if col_fecha is not None:
        validos["_fecha"] = pd.to_datetime(validos[col_fecha], errors="coerce")
        con_fecha = validos.loc[validos["_fecha"].notna()].sort_values("_fecha")
        if not con_fecha.empty:
            fila = con_fecha.iloc[-1]
            fecha_dt = fila["_fecha"]
            fecha_txt = fecha_dt.strftime("%d/%m/%Y %H:%M")
        else:
            fila = validos.iloc[-1]
            fecha_txt = str(fila.get(col_fecha, "N/D"))
    else:
        fila = validos.iloc[-1]

    return {
        "valor": float(fila["_valor"]),
        "fecha": fecha_txt,
        "fecha_dt": fecha_dt,
        "archivo": os.path.basename(path),
        "path": path,
    }


def _cargar_serie_evap_aquarius(nombre: str) -> dict:
    """Devuelve valor/fecha o un error legible sin detener el dashboard."""
    path = _buscar_archivo_data(nombre)
    if not path:
        return {"ok": False, "archivo": nombre, "error": "archivo no encontrado en la carpeta data"}
    try:
        stat = os.stat(path)
        info = _leer_ultimo_valor_aquarius(path, int(stat.st_mtime_ns))
        info["ok"] = True
        return info
    except Exception as exc:
        return {"ok": False, "archivo": nombre, "path": path, "error": str(exc)}


# ── Aportes observados desde Aquarius (Discharge AT) ─────────────────────────
APORTES_OBSERVADOS_FILES = {
    # Se acepta el nombre base y cualquier copia numerada, por ejemplo:
    # Discharge_AT_GAT_Diario(6).csv. Si hay varias copias, se usa la más reciente.
    "Gatún": ["Discharge_AT_GAT_Diario.csv", "Discharge_AT_GAT_Diario*.csv"],
    "Alhajuela": ["Discharge_AT_ALHA_Diario.csv", "Discharge_AT_ALHA_Diario*.csv"],
}


def _dirs_busqueda_data():
    """Directorios donde el app busca datos locales, sin duplicados."""
    base_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in globals() else os.getcwd()
    candidatos = [
        os.path.join(base_dir, "data"),
        os.path.join(os.getcwd(), "data"),
        base_dir,
        os.getcwd(),
    ]
    dirs, vistos = [], set()
    for folder in candidatos:
        folder_abs = os.path.abspath(folder)
        if folder_abs in vistos:
            continue
        vistos.add(folder_abs)
        if os.path.isdir(folder_abs):
            dirs.append(folder_abs)
    return dirs


def _buscar_archivo_data_multi(nombres):
    """Busca CSV locales y, si hay copias numeradas, usa el archivo más reciente."""
    encontrados = []
    for folder in _dirs_busqueda_data():
        for nombre in nombres:
            patron = os.path.join(folder, nombre)
            if any(ch in nombre for ch in "*?[]"):
                candidatos = glob.glob(patron)
            else:
                candidatos = [patron]
            for path in candidatos:
                try:
                    if os.path.isfile(path) and not os.path.basename(path).startswith("."):
                        encontrados.append(os.path.abspath(path))
                except Exception:
                    continue
    if not encontrados:
        return None
    encontrados = sorted(set(encontrados), key=lambda p: os.path.getmtime(p), reverse=True)
    return encontrados[0]


def _hoy_panama_normalizado() -> pd.Timestamp:
    """Fecha actual operativa en Panamá, igual que en el app DSS."""
    try:
        return pd.Timestamp.now(tz="America/Panama").tz_localize(None).normalize()
    except Exception:
        return pd.Timestamp.today().normalize()


def _ajustar_fecha_aporte_observado(df: pd.DataFrame, date_col: str = "fecha") -> pd.DataFrame:
    """Normaliza la fecha operativa y elimina registros posteriores a hoy Panamá.

    En `Discharge_AT_*_Diario`, Aquarius entrega el dato diario con sello
    00:00 del día siguiente. Por ejemplo, la fila 22/06 00:00 corresponde
    operativamente al aporte observado del 21/06. La conversión de fecha se
    hace al leer la serie; aquí solo se filtra por la fecha operativa final.
    """
    if df is None or df.empty or date_col not in df.columns:
        return df
    out = df.copy()
    out[date_col] = pd.to_datetime(out[date_col], errors="coerce").dt.normalize()
    hoy = _hoy_panama_normalizado()
    out = out.loc[out[date_col].notna() & (out[date_col] <= hoy)].copy()
    return out


@st.cache_data(show_spinner=False)
def _leer_aportes_observados_csv(path: str, mtime_ns: int) -> pd.DataFrame:
    """Lee una serie de aportes observados diarios y normaliza unidades.

    La serie `Discharge_AT_*_Diario` de Aquarius viene en m³/s. El app conserva
    ese valor como unidad fuente y calcula sus equivalentes en hm³/día y p³/s
    para visores, gráficas y tablas.
    """
    del mtime_ns
    df = pd.read_csv(path)
    cols_out = ["fecha", "cfs", "hm3_d", "m3s"]
    if df.empty:
        return pd.DataFrame(columns=cols_out)

    columnas_l = {str(c).strip().lower(): c for c in df.columns}
    col_valor = next(
        (columnas_l[k] for k in ("valor_raw", "valor", "value", "result", "discharge") if k in columnas_l),
        None,
    )
    if col_valor is None:
        for c in df.columns:
            serie_num = pd.to_numeric(df[c], errors="coerce")
            if serie_num.notna().any():
                col_valor = c
                break
    if col_valor is None:
        return pd.DataFrame(columns=cols_out)

    # En las series diarias `Discharge_AT_*_Diario`, Aquarius etiqueta el
    # cierre del intervalo a las 00:00 del día siguiente. Por eso se usa
    # preferentemente `fecha_fin` y, cuando el sello está exactamente a
    # medianoche, se resta un día: 22/06 00:00 => aporte operativo 21/06.
    col_fecha_inicio = columnas_l.get("fecha_inicio")
    col_fecha_fin = columnas_l.get("fecha_fin")
    col_fecha = col_fecha_fin or col_fecha_inicio or next(
        (columnas_l[k] for k in ("timestamp", "date", "fecha") if k in columnas_l),
        None,
    )
    out = df.copy()
    out["m3s"] = pd.to_numeric(out[col_valor], errors="coerce")
    if col_fecha is not None:
        fechas = pd.to_datetime(out[col_fecha], errors="coerce")
        try:
            mask_cierre_medianoche = (
                fechas.notna()
                & fechas.dt.hour.eq(0)
                & fechas.dt.minute.eq(0)
                & fechas.dt.second.eq(0)
            )
            fechas = fechas.mask(mask_cierre_medianoche, fechas - pd.Timedelta(days=1))
        except Exception:
            pass
        out["fecha"] = fechas
    else:
        out["fecha"] = pd.NaT

    out = out.loc[out["m3s"].notna() & np.isfinite(out["m3s"]) & (out["m3s"] >= 0)].copy()
    out = out.loc[out["fecha"].notna()].sort_values("fecha")
    if out.empty:
        return pd.DataFrame(columns=cols_out)

    # Luego del ajuste, 22/06 00:00 queda como fecha operativa 21/06.
    # Se filtran fechas futuras reales y, si hay duplicados del mismo día,
    # se conserva el último valor cargado para ese día operativo.
    out = _ajustar_fecha_aporte_observado(out, "fecha")
    if out.empty:
        return pd.DataFrame(columns=cols_out)
    out = (
        out.groupby("fecha", as_index=False)
           .agg(m3s=("m3s", "last"))
           .sort_values("fecha")
    )

    out["hm3_d"] = out["m3s"].astype(float) / HM3D2M3S
    out["cfs"] = out["m3s"].astype(float) * M3S2CFS
    return out[["fecha", "cfs", "hm3_d", "m3s"]].reset_index(drop=True)


def _cargar_aportes_observados() -> dict:
    """Carga los aportes observados de Gatún y Alhajuela, si existen."""
    series = {}
    for embalse, nombres in APORTES_OBSERVADOS_FILES.items():
        path = _buscar_archivo_data_multi(nombres)
        if not path:
            series[embalse] = {
                "ok": False,
                "archivo": " / ".join(nombres),
                "error": "archivo no encontrado en carpeta data o junto al app",
                "df": pd.DataFrame(columns=["fecha", "cfs", "hm3_d", "m3s"]),
            }
            continue
        try:
            stat = os.stat(path)
            df = _leer_aportes_observados_csv(path, int(stat.st_mtime_ns))
            series[embalse] = {
                "ok": not df.empty,
                "archivo": os.path.basename(path),
                "path": path,
                "df": df,
                "error": None if not df.empty else "sin datos válidos",
            }
        except Exception as exc:
            series[embalse] = {
                "ok": False,
                "archivo": os.path.basename(path),
                "path": path,
                "error": str(exc),
                "df": pd.DataFrame(columns=["fecha", "cfs", "hm3_d", "m3s"]),
            }
    return series


def _resumen_aportes_df(df: pd.DataFrame, dias_prom: int = 7) -> dict | None:
    """Resumen rápido para una serie de aportes observados.

    Presenta el caudal principal como volumen diario (hm³/d) y mantiene
    sus conversiones operativas en p³/s y m³/s.
    """
    if df is None or df.empty:
        return None
    d = df.sort_values("fecha").copy()
    ult = d.iloc[-1]
    ventana = d.tail(max(int(dias_prom), 1))
    ventana_30 = d.tail(30)
    return {
        "fecha": ult["fecha"],
        "ultimo_cfs": float(ult["cfs"]),
        "ultimo_hm3": float(ult["hm3_d"]),
        "ultimo_m3s": float(ult["m3s"]),
        "prom_cfs": float(ventana["cfs"].mean()),
        "prom_hm3": float(ventana["hm3_d"].mean()),
        "prom_m3s": float(ventana["m3s"].mean()),
        "prom30_cfs": float(ventana_30["cfs"].mean()),
        "prom30_hm3": float(ventana_30["hm3_d"].mean()),
        "prom30_m3s": float(ventana_30["m3s"].mean()),
        "n": int(len(ventana)),
        "n30": int(len(ventana_30)),
    }


# ═══ SIDEBAR ═══
# Logos: Canal de Panamá + HIMH
_sb_logos = ""
if _logo_cp:
    _sb_logos += _img_tag(_logo_cp_mime, _logo_cp, "max-width:110px;margin:0 6px;")
if _logo:
    _sb_logos += _img_tag(_logo_mime, _logo, "max-width:90px;margin:0 6px;")
if _sb_logos:
    st.sidebar.markdown(
        f"<div style='text-align:center;margin-bottom:8px;display:flex;"
        f"align-items:center;justify-content:center;gap:8px;'>{_sb_logos}</div>",
        unsafe_allow_html=True)
st.sidebar.markdown("## 💧 Demandas de Agua\nCanal de Panamá")
_contador_consecutivo = get_consecutive_counter()
st.sidebar.markdown(
    f"""
    <div style="display:flex; align-items:center; gap:8px; margin:2px 0 10px 0;">
        <span style="font-weight:800; color:#004b7a; font-size:1.02rem; letter-spacing:0.2px;">ACP-HIMH</span>
        <span style="display:inline-flex; align-items:center; gap:6px; background:#004b7a; color:#ffffff; border-radius:9px; padding:4px 10px; font-weight:800; font-size:0.86rem; line-height:1.1; box-shadow:0 1px 3px rgba(0,0,0,0.12);">
            <span style="font-size:0.92rem;">👁️</span>
            <span>{_contador_consecutivo:,}</span>
        </span>
    </div>
    """,
    unsafe_allow_html=True,
)
st.sidebar.markdown("---")

# ═══ BÚSQUEDA LOCAL UNIFICADA DE LAKEHOUSE ══════════════════════════════════
def _buscar_lakehouse_local():
    """Busca archivos LakeHouse*.xlsx junto al app y en el directorio de ejecución.
    Devuelve la lista ordenada por modificación reciente para que sidebar y pestaña usen la misma fuente.
    """
    try:
        import glob as _glob
        base_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in globals() else os.getcwd()
        patrones = [
            os.path.join(base_dir, "LakeHouse*.xlsx"),
            os.path.join(os.getcwd(), "LakeHouse*.xlsx"),
        ]
        candidatos = []
        for pat in patrones:
            candidatos.extend(_glob.glob(pat))
        candidatos = sorted(set(candidatos), key=lambda x: (os.path.getmtime(x), x), reverse=True)
        return candidatos
    except Exception:
        return []

# ═══ LECTURA RÁPIDA DE NIVELES DESDE LAKEHOUSE PARA EL SIDEBAR ═══════════════
# Solo alimenta los valores iniciales de Niveles Operativos.
# Después de cargados, el usuario puede modificarlos manualmente en el sidebar.
def _lkh_sidebar_source():
    """Devuelve una fuente LakeHouse local o subida previamente, sin alterar el resto del app."""
    try:
        up = st.session_state.get("lk", None)
        if up is not None:
            try:
                size = getattr(up, "size", None)
                name = getattr(up, "name", "LakeHouse_upload.xlsx")
                up.seek(0)
                return up, f"upload:{name}:{size}"
            except Exception:
                pass
    except Exception:
        pass

    try:
        candidatos = _buscar_lakehouse_local()
        if candidatos:
            # Si el usuario seleccionó un LakeHouse local en la pestaña 📂 Datos Lake House,
            # se respeta esa selección; si no, se usa el archivo local más reciente.
            seleccion_local = st.session_state.get("lkh_local_path_select")
            if seleccion_local and os.path.exists(str(seleccion_local)):
                path = str(seleccion_local)
            else:
                path = candidatos[0]
            try:
                stat = os.stat(path)
                return path, f"local:{os.path.abspath(path)}:{stat.st_mtime_ns}:{stat.st_size}"
            except Exception:
                return path, f"local:{os.path.abspath(path)}"
    except Exception:
        pass
    return None, None


def _fecha_lkh_texto(valor, formato="%d/%m/%Y"):
    """Formatea fechas LakeHouse, incluyendo seriales numéricos de Excel."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return "N/D"
    try:
        if isinstance(valor, (int, float, np.integer, np.floating)):
            numero = float(valor)
            if 20000 <= numero <= 80000:
                fecha = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=numero)
                return fecha.strftime(formato)
        fecha = pd.to_datetime(valor, errors="coerce")
        if pd.isna(fecha):
            return str(valor)
        return fecha.strftime(formato)
    except Exception:
        return str(valor)


@st.cache_data(show_spinner=False)
def _leer_ultimos_niveles_lkh(path_o_bytes, source_id):
    """Lee únicamente fecha, nivel Gatún y nivel Alhajuela del último registro válido."""
    try:
        import openpyxl
        from io import BytesIO

        src = BytesIO(path_o_bytes) if isinstance(path_o_bytes, (bytes, bytearray)) else path_o_bytes
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
        hojas = [x for x in wb.sheetnames if x not in ["Sheet1", "Para BalanceH"]]
        if not hojas:
            hojas = wb.sheetnames

        for hoja in hojas:
            ws = wb[hoja]
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            header_l = [str(c).strip().lower() if c is not None else "" for c in header]
            idx_f = next((i for i, c in enumerate(header_l) if "date" in c), None)
            idx_g = next((i for i, c in enumerate(header_l) if "gatel" in c), None)
            idx_a = next((i for i, c in enumerate(header_l) if "madel" in c), None)
            if idx_g is None and idx_a is None:
                continue

            ultimo = None
            for row in rows:
                vals = list(row)
                nv_g = vals[idx_g] if idx_g is not None and idx_g < len(vals) else None
                nv_a = vals[idx_a] if idx_a is not None and idx_a < len(vals) else None
                fecha = vals[idx_f] if idx_f is not None and idx_f < len(vals) else None
                try:
                    nv_g = float(nv_g) if nv_g is not None else None
                except Exception:
                    nv_g = None
                try:
                    nv_a = float(nv_a) if nv_a is not None else None
                except Exception:
                    nv_a = None
                if nv_g is not None or nv_a is not None:
                    ultimo = {"fecha": fecha, "gatun": nv_g, "alhajuela": nv_a, "hoja": hoja}
            if ultimo is not None:
                return ultimo
    except Exception:
        return None
    return None

def _aplicar_niveles_lkh_si_corresponde():
    src, sid = _lkh_sidebar_source()
    if not src or not sid:
        return None
    try:
        if hasattr(src, "getvalue"):
            src.seek(0)
            payload = src.getvalue()
        else:
            payload = src
        info = _leer_ultimos_niveles_lkh(payload, sid)
        if not info:
            return None

        # Cuando cambia la fuente LakeHouse, se actualizan los campos una sola vez.
        # Luego el usuario puede editarlos sin que el app los reescriba en cada rerun.
        if st.session_state.get("_niveles_lkh_source_id") != sid:
            if info.get("gatun") is not None:
                st.session_state["nivel_gat_op"] = max(55.0, min(93.0, round(float(info["gatun"]), 2)))
            if info.get("alhajuela") is not None:
                st.session_state["nivel_alh_op"] = max(160.0, min(260.0, round(float(info["alhajuela"]), 2)))
            st.session_state["_niveles_lkh_source_id"] = sid
            st.session_state["_niveles_lkh_info"] = info
        return info
    except Exception:
        return None


@st.cache_data(show_spinner=False)
def _leer_defaults_operativos_lkh(path_o_bytes, source_id, n_dias=5):
    """
    Lee los últimos N registros del LakeHouse y devuelve promedios operativos
    para inicializar los controles del sidebar. Los controles siguen siendo editables.
    """
    try:
        import openpyxl
        from io import BytesIO

        MCF_TO_CFS_LOCAL = 1_000_000.0 / 86400.0

        def _num(v):
            try:
                if v is None:
                    return None
                if isinstance(v, str) and v.strip().startswith("#"):
                    return None
                return float(v)
            except Exception:
                return None

        def _mean(vals):
            vals = [float(x) for x in vals if x is not None and np.isfinite(float(x))]
            return float(np.mean(vals)) if vals else None

        def _find(header_l, *tokens, exact=None):
            if exact:
                exact_l = [e.lower() for e in exact]
                for i, h in enumerate(header_l):
                    if h in exact_l:
                        return i
            for i, h in enumerate(header_l):
                if all(t.lower() in h for t in tokens):
                    return i
            return None

        src = BytesIO(path_o_bytes) if isinstance(path_o_bytes, (bytes, bytearray)) else path_o_bytes
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
        hojas = [x for x in wb.sheetnames if x not in ["Sheet1", "Para BalanceH"]] or wb.sheetnames

        for hoja in hojas:
            ws = wb[hoja]
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            header_l = [str(c).strip().lower() if c is not None else "" for c in header]

            idx = {
                "fecha": _find(header_l, "date"),
                "nv_g": _find(header_l, "gatel"),
                "nv_a": _find(header_l, "madel"),
                "n_g": _find(header_l, exact=["numlockgat"]),
                "n_p": _find(header_l, exact=["numlockpm"]),
                "n_a": _find(header_l, exact=["numlockac", "numlockacl"]),
                "n_c": _find(header_l, exact=["numlockccl"]),
                "mad_mwh": _find(header_l, exact=["madmwh"]),
                "gat_mwh": _find(header_l, exact=["gatmwh"]),
                "pot_a_mcf": _find(header_l, exact=["munic_mad"]),
                "pot_g_mcf": _find(header_l, exact=["munic_gat"]),
                "pot_a_hm3": _find(header_l, exact=["munic_mad_hm3"]),
                "pot_g_hm3": _find(header_l, exact=["munic_gat_hm3"]),
                "fug_a_mcf": _find(header_l, exact=["leak_mad"]),
                "fug_g_mcf": _find(header_l, exact=["leak_gat"]),
                "fug_a_hm3": _find(header_l, exact=["leak_mad_hm3"]),
                "fug_g_hm3": _find(header_l, exact=["leak_gat_hm3"]),
                # Vertidos LakeHouse en MPC/MCF por día.
                "vert_a_mcf": _find(header_l, exact=["madspill"]),
                "vert_g_mcf": _find(header_l, exact=["gatspill"]),
                "evap_g_mm": _find(header_l, exact=["evap_gatun_mm"]),
                "evap_a_mm": _find(header_l, exact=["evap_alaj_mm", "evap_ala_mm"]),
                "evap_g_hm3": _find(header_l, exact=["vol_evap_gat_hm3"]),
                "evap_a_hm3": _find(header_l, exact=["vol_evap_ala_hm3"]),
                "pnx_unit_hm3": _find(header_l, exact=["total pnx"]),
                "npx_unit_hm3": _find(header_l, exact=["total npx"]),
            }

            ult = []
            for row in rows:
                vals = list(row)
                rec = {}
                for k, i in idx.items():
                    rec[k] = vals[i] if i is not None and i < len(vals) else None
                if any(_num(rec.get(k)) is not None for k in rec if k != "fecha"):
                    ult.append(rec)
                    if len(ult) > max(int(n_dias), 1):
                        ult.pop(0)
            if not ult:
                continue

            n_pnx_vals, n_npx_vals = [], []
            for r in ult:
                ng, np_ = _num(r.get("n_g")), _num(r.get("n_p"))
                na, nc  = _num(r.get("n_a")), _num(r.get("n_c"))
                if ng is not None or np_ is not None:
                    n_pnx_vals.append(_mean([ng, np_]))
                if na is not None or nc is not None:
                    n_npx_vals.append(_mean([na, nc]))

            # Conversión de consumos MCF/d → cfs; hm³/d → cfs cuando existe columna directa.
            def _cfs_from(mcf_key, hm3_key):
                hm3 = _mean([_num(r.get(hm3_key)) for r in ult])
                if hm3 is not None and hm3 > 0:
                    return hm3 / CFS2HM3
                mcf = _mean([_num(r.get(mcf_key)) for r in ult])
                return mcf * MCF_TO_CFS_LOCAL if mcf is not None else None

            def _cfs_from_mcf_first(mcf_key, hm3_key, max_cfs=None):
                """Prioriza la columna MCF/MPC del LakeHouse y usa hm³/día solo como respaldo validado.

                Esto evita que el sidebar tome valores desajustados de columnas *_hm3
                cuando el LakeHouse trae también la columna base en MCF/MPC por día.
                """
                mcf = _mean([_num(r.get(mcf_key)) for r in ult])
                if mcf is not None and mcf >= 0:
                    cfs = mcf * MCF_TO_CFS_LOCAL
                    if max_cfs is None or cfs <= float(max_cfs):
                        return cfs
                hm3 = _mean([_num(r.get(hm3_key)) for r in ult])
                if hm3 is not None and hm3 >= 0:
                    cfs = hm3 / CFS2HM3
                    if max_cfs is None or cfs <= float(max_cfs):
                        return cfs
                return None

            def _cfs_from_mcf_only(mcf_key, max_cfs=None):
                """Convierte MPC/MCF por día del LakeHouse a cfs promedio diario."""
                mcf = _mean([_num(r.get(mcf_key)) for r in ult])
                if mcf is None or mcf < 0:
                    return None
                cfs = mcf * MCF_TO_CFS_LOCAL
                if max_cfs is not None and cfs > float(max_cfs):
                    return None
                return cfs

            out = {
                "hoja": hoja,
                "n_dias": int(n_dias),
                "fecha_ultimo": ult[-1].get("fecha"),
                "n_pnx": _mean(n_pnx_vals),
                "n_npx": _mean(n_npx_vals),
                "gm_mw": (_mean([_num(r.get("mad_mwh")) for r in ult]) or 0) / 24.0 if _mean([_num(r.get("mad_mwh")) for r in ult]) is not None else None,
                "gg_mw": (_mean([_num(r.get("gat_mwh")) for r in ult]) or 0) / 24.0 if _mean([_num(r.get("gat_mwh")) for r in ult]) is not None else None,
                "pot_alh_cfs": _cfs_from_mcf_first("pot_a_mcf", "pot_a_hm3", max_cfs=800),
                "pot_gat_cfs": _cfs_from_mcf_first("pot_g_mcf", "pot_g_hm3", max_cfs=600),
                "fug_alh_cfs": _cfs_from_mcf_first("fug_a_mcf", "fug_a_hm3", max_cfs=300),
                "fug_gat_cfs": _cfs_from_mcf_first("fug_g_mcf", "fug_g_hm3", max_cfs=400),
                "vert_alh_cfs": _cfs_from_mcf_only("vert_a_mcf", max_cfs=5000),
                "vert_gat_cfs": _cfs_from_mcf_only("vert_g_mcf", max_cfs=20000),
                "evap_gat_mm": _mean([_num(r.get("evap_g_mm")) for r in ult]),
                "evap_alh_mm": _mean([_num(r.get("evap_a_mm")) for r in ult]),
                "evap_gat_hm3_lkh": _mean([_num(r.get("evap_g_hm3")) for r in ult]),
                "evap_alh_hm3_lkh": _mean([_num(r.get("evap_a_hm3")) for r in ult]),
                "vp_lkh": _mean([_num(r.get("pnx_unit_hm3")) for r in ult]),
                "vn_lkh": _mean([_num(r.get("npx_unit_hm3")) for r in ult]),
            }
            return out
    except Exception:
        return None
    return None

def _aplicar_defaults_operativos_lkh_si_corresponde(n_dias=5):
    """Inicializa sliders/number_inputs con LakeHouse solo cuando cambia la fuente."""
    src, sid = _lkh_sidebar_source()
    if not src or not sid:
        return None
    try:
        if hasattr(src, "getvalue"):
            src.seek(0)
            payload = src.getvalue()
        else:
            payload = src
        sid_op = f"{sid}:defaults:{int(n_dias)}"
        info = _leer_defaults_operativos_lkh(payload, sid_op, int(n_dias))
        if not info:
            return None
        if st.session_state.get("_defaults_lkh_source_id") != sid_op:
            # Los vertidos se inicializan con el promedio LakeHouse del período elegido.
            # La evaporación no se inicializa desde LakeHouse; la controla la fuente activa
            # seleccionada en el bloque de evaporación (Manual o Aquarius).
            mapeo = {
                "n_pnx_lkh": (info.get("n_pnx"), 0.0, 40.0, 1),
                "n_npx_lkh": (info.get("n_npx"), 0.0, 20.0, 1),
                "gm_mw_lkh": (info.get("gm_mw"), 0.0, 36.0, 0),
                "gg_mw_lkh": (info.get("gg_mw"), 0.0, 30.0, 0),
                "pot_alh_lkh": (info.get("pot_alh_cfs"), 0.0, 800.0, 0),
                "pot_gat_lkh": (info.get("pot_gat_cfs"), 0.0, 600.0, 0),
                "fug_alh_lkh": (info.get("fug_alh_cfs"), 0.0, 300.0, 0),
                "fug_gat_lkh": (info.get("fug_gat_cfs"), 0.0, 400.0, 0),
                "v_fondo_lkh": (info.get("vert_alh_cfs"), 0.0, 5000.0, 0),
                "v_gatun_lkh": (info.get("vert_gat_cfs"), 0.0, 20000.0, 0),
            }
            for key, (val, mn, mx, nd) in mapeo.items():
                if val is None or pd.isna(val):
                    continue
                val = max(float(mn), min(float(mx), float(val)))
                st.session_state[key] = int(round(val)) if nd == 0 else round(val, nd)
            st.session_state["_defaults_lkh_source_id"] = sid_op
            st.session_state["_defaults_lkh_info"] = info
        return info
    except Exception:
        return None


@st.cache_data(show_spinner=False)
def _leer_balance_detallado_lkh(path_o_bytes, source_id, n_dias=5):
    """Lee promedios detallados de salidas por embalse desde LakeHouse.
    Alimenta los KPI superiores y el resumen de Balance; no altera los cálculos principales.
    """
    try:
        import openpyxl
        from io import BytesIO

        MCF_TO_CFS_LOCAL = 1_000_000.0 / 86400.0

        def _num(v):
            try:
                if v is None:
                    return None
                if isinstance(v, str) and v.strip().startswith("#"):
                    return None
                return float(v)
            except Exception:
                return None

        def _mean(vals):
            vals = [float(x) for x in vals if x is not None and np.isfinite(float(x))]
            return float(np.mean(vals)) if vals else None

        def _find(header_l, *tokens, exact=None):
            if exact:
                exact_l = [str(e).lower() for e in exact]
                for i, h in enumerate(header_l):
                    if h in exact_l:
                        return i
            for i, h in enumerate(header_l):
                if all(str(t).lower() in h for t in tokens):
                    return i
            return None

        src = BytesIO(path_o_bytes) if isinstance(path_o_bytes, (bytes, bytearray)) else path_o_bytes
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
        hojas = [x for x in wb.sheetnames if x not in ["Sheet1", "Para BalanceH"]] or wb.sheetnames

        for hoja in hojas:
            ws = wb[hoja]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                continue
            header_l = [str(c).strip().lower() if c is not None else "" for c in header]

            idx = {
                "fecha": _find(header_l, "date"),
                "gat_hm3": _find(header_l, exact=["gatlockhm3"]),
                "pm_hm3": _find(header_l, exact=["pmlockhm3"]),
                "acl_hm3": _find(header_l, exact=["aclockhm3"]),
                "ccl_hm3": _find(header_l, exact=["ccllockhm3"]),
                "gat_mcf": _find(header_l, "gatlockmcf"),
                "pm_mcf": _find(header_l, "pmlockmcf"),
                "acl_mcf": _find(header_l, "aclockmcf"),
                "ccl_mcf": _find(header_l, "ccllockmcf"),
                "gen_mad_hm3": _find(header_l, exact=["madhm3"]),
                "gen_gat_hm3": _find(header_l, exact=["gathm3"]),
                "pot_m_hm3": _find(header_l, exact=["munic_mad_hm3"]),
                "pot_g_hm3": _find(header_l, exact=["munic_gat_hm3"]),
                "pot_m_mcf": _find(header_l, exact=["munic_mad"]),
                "pot_g_mcf": _find(header_l, exact=["munic_gat"]),
                "fug_m_hm3": _find(header_l, exact=["leak_mad_hm3"]),
                "fug_g_hm3": _find(header_l, exact=["leak_gat_hm3"]),
                "fug_m_mcf": _find(header_l, exact=["leak_mad"]),
                "fug_g_mcf": _find(header_l, exact=["leak_gat"]),
                # madspill corresponde al volumen diario vertido en Madden,
                # expresado por LakeHouse en MPC/MCF (millones de pies³ por día).
                "vert_m_mcf": _find(header_l, exact=["madspill"]),
                "vert_g_mcf": _find(header_l, exact=["gatspill"]),
                "evap_m_hm3": _find(header_l, exact=["vol_evap_ala_hm3"]),
                "evap_g_hm3": _find(header_l, exact=["vol_evap_gat_hm3"]),
                # Totales y datos auxiliares para los KPI superiores.
                "total_consumo_hm3": _find(header_l, exact=["agua_consumida_ala_gat_hm3"]),
                "usos_hm3": _find(header_l, exact=["usos_hm3"]),
                "zz_ccl_hm3": _find(header_l, exact=["ccl_zz_flush"]),
                "zz_acl_hm3": _find(header_l, exact=["acl_zz_flush"]),
                "mad_mwh": _find(header_l, exact=["madmwh"]),
                "gat_mwh": _find(header_l, exact=["gatmwh"]),
                "n_g": _find(header_l, exact=["numlockgat"]),
                "n_p": _find(header_l, exact=["numlockpm"]),
                "n_a": _find(header_l, exact=["numlockac", "numlockacl"]),
                "n_c": _find(header_l, exact=["numlockccl"]),
            }

            ult = []
            for row in rows_iter:
                vals = list(row)
                rec = {}
                for k, i in idx.items():
                    rec[k] = vals[i] if i is not None and i < len(vals) else None
                if any(_num(rec.get(k)) is not None for k in rec if k != "fecha"):
                    ult.append(rec)
                    if len(ult) > max(int(n_dias), 1):
                        ult.pop(0)
            if not ult:
                continue

            def _hm3_component(hm3_keys=None, mcf_keys=None):
                hm3_keys = hm3_keys or []
                mcf_keys = mcf_keys or []
                if hm3_keys:
                    vals_hm3 = []
                    for r in ult:
                        vals = [_num(r.get(k)) for k in hm3_keys]
                        vals = [v for v in vals if v is not None]
                        if vals:
                            vals_hm3.append(sum(vals))
                    m = _mean(vals_hm3)
                    if m is not None:
                        return m
                if mcf_keys:
                    vals_mcf = []
                    for r in ult:
                        vals = [_num(r.get(k)) for k in mcf_keys]
                        vals = [v for v in vals if v is not None]
                        if vals:
                            vals_mcf.append(sum(vals))
                    mcf = _mean(vals_mcf)
                    if mcf is not None:
                        return mcf * MCF_TO_CFS_LOCAL * CFS2HM3
                return None

            def _hm3_component_mcf_first(mcf_keys=None, hm3_keys=None, max_hm3=None):
                """Primero MCF/MPC por día del LakeHouse; hm³/día solo como respaldo validado."""
                mcf_keys = mcf_keys or []
                hm3_keys = hm3_keys or []
                if mcf_keys:
                    vals_mcf = []
                    for r in ult:
                        vals = [_num(r.get(k)) for k in mcf_keys]
                        vals = [v for v in vals if v is not None]
                        if vals:
                            vals_mcf.append(sum(vals))
                    mcf = _mean(vals_mcf)
                    if mcf is not None:
                        hm3 = mcf * MCF_TO_CFS_LOCAL * CFS2HM3
                        if max_hm3 is None or hm3 <= float(max_hm3):
                            return hm3
                if hm3_keys:
                    vals_hm3 = []
                    for r in ult:
                        vals = [_num(r.get(k)) for k in hm3_keys]
                        vals = [v for v in vals if v is not None]
                        if vals:
                            vals_hm3.append(sum(vals))
                    hm3 = _mean(vals_hm3)
                    if hm3 is not None and (max_hm3 is None or hm3 <= float(max_hm3)):
                        return hm3
                return None

            detalle = [
                {"Embalse": "Alhajuela ", "Uso": "Generación Madden", "hm3": _hm3_component(["gen_mad_hm3"])},
                {"Embalse": "Alhajuela ", "Uso": "Potabilización", "hm3": _hm3_component_mcf_first(["pot_m_mcf"], ["pot_m_hm3"], max_hm3=3.0)},
                {"Embalse": "Alhajuela ", "Uso": "Fugas", "hm3": _hm3_component_mcf_first(["fug_m_mcf"], ["fug_m_hm3"], max_hm3=1.5)},
                {"Embalse": "Alhajuela ", "Uso": "Vertido Madden", "hm3": _hm3_component([], ["vert_m_mcf"])},
                {"Embalse": "Alhajuela ", "Uso": "Evaporación", "hm3": _hm3_component(["evap_m_hm3"])},
                {"Embalse": "Gatún", "Uso": "Esclusajes PNX", "hm3": _hm3_component(["gat_hm3", "pm_hm3"], ["gat_mcf", "pm_mcf"])},
                {"Embalse": "Gatún", "Uso": "Esclusajes NPX", "hm3": _hm3_component(["acl_hm3", "ccl_hm3"], ["acl_mcf", "ccl_mcf"])},
                {"Embalse": "Gatún", "Uso": "Generación Gatún", "hm3": _hm3_component(["gen_gat_hm3"])},
                {"Embalse": "Gatún", "Uso": "Potabilización", "hm3": _hm3_component_mcf_first(["pot_g_mcf"], ["pot_g_hm3"], max_hm3=3.0)},
                {"Embalse": "Gatún", "Uso": "Fugas", "hm3": _hm3_component_mcf_first(["fug_g_mcf"], ["fug_g_hm3"], max_hm3=1.5)},
                {"Embalse": "Gatún", "Uso": "Vertido Gatún", "hm3": _hm3_component([], ["vert_g_mcf"])},
                {"Embalse": "Gatún", "Uso": "Evaporación", "hm3": _hm3_component(["evap_g_hm3"])},
            ]
            if not any(x.get("hm3") is not None for x in detalle):
                continue

            def _mean_row_sum(keys):
                vals_rows = []
                for r in ult:
                    vals = [_num(r.get(k)) for k in keys]
                    vals = [v for v in vals if v is not None]
                    if vals:
                        vals_rows.append(sum(vals))
                return _mean(vals_rows)

            def _mean_row_pair(keys):
                vals_rows = []
                for r in ult:
                    vals = [_num(r.get(k)) for k in keys]
                    vals = [v for v in vals if v is not None]
                    if vals:
                        vals_rows.append(float(np.mean(vals)))
                return _mean(vals_rows)

            _mad_mwh_prom = _mean([_num(r.get("mad_mwh")) for r in ult])
            _gat_mwh_prom = _mean([_num(r.get("gat_mwh")) for r in ult])

            return {
                "hoja": hoja,
                "n_dias": int(n_dias),
                "fecha_ultimo": ult[-1].get("fecha"),
                "detalle": detalle,
                # Total oficial del LakeHouse (incluye evaporación cuando está disponible).
                "total_consumo_hm3": _mean([_num(r.get("total_consumo_hm3")) for r in ult]),
                "usos_hm3": _mean([_num(r.get("usos_hm3")) for r in ult]),
                "zz_flush_hm3": _mean_row_sum(["zz_ccl_hm3", "zz_acl_hm3"]),
                # Potencia media del período: MWh/día ÷ 24 = MW medios.
                "mad_mw": (_mad_mwh_prom / 24.0) if _mad_mwh_prom is not None else None,
                "gat_mw": (_gat_mwh_prom / 24.0) if _gat_mwh_prom is not None else None,
                # Tránsitos: promedio entre complejos para evitar duplicar buques.
                "n_pnx": _mean_row_pair(["n_g", "n_p"]),
                "n_npx": _mean_row_pair(["n_a", "n_c"]),
            }
    except Exception:
        return None
    return None


def _obtener_balance_detallado_lkh(n_dias=5):
    """Obtiene el resumen detallado del LakeHouse local o subido, respetando el período 1/5/7/10/30 días."""
    src, sid = _lkh_sidebar_source()
    if not src or not sid:
        return None
    try:
        if hasattr(src, "getvalue"):
            src.seek(0)
            payload = src.getvalue()
        else:
            payload = src
        sid_det = f"{sid}:balance_detalle:{int(n_dias)}"
        return _leer_balance_detallado_lkh(payload, sid_det, int(n_dias))
    except Exception:
        return None

_info_niveles_lkh = _aplicar_niveles_lkh_si_corresponde()
# El período seleccionado en la pestaña 📂 Datos Lake House también alimenta
# los valores iniciales/editables del sidebar y, por lo tanto, el balance principal.
# Como el sidebar se construye antes que las pestañas, se lee el valor guardado
# en session_state del último selector usado (1, 5, 7, 10 o 30 días).
try:
    _dias_defaults_lkh = int(st.session_state.get("dias_op", 5))
except Exception:
    _dias_defaults_lkh = 5
if _dias_defaults_lkh not in (1, 5, 7, 10, 30):
    _dias_defaults_lkh = 5
st.session_state.setdefault("dias_op", _dias_defaults_lkh)
_info_defaults_lkh = _aplicar_defaults_operativos_lkh_si_corresponde(_dias_defaults_lkh)
_info_balance_lkh = _obtener_balance_detallado_lkh(_dias_defaults_lkh)
# Período LakeHouse disponible para todo el app, incluso cuando la fuente de
# consumo por esclusaje no está configurada como 'Basado en LakeHouse'.
# La variable ya existe desde el inicio y aquí se actualiza con el período real.
_dias_lkh_balance = _dias_lkh_seguros(_dias_defaults_lkh)
# Valores de respaldo si no existe LakeHouse; se definen antes de crear los widgets.
st.session_state.setdefault("nivel_gat_op", 87.0)
st.session_state.setdefault("nivel_alh_op", 252.0)
st.session_state.setdefault("n_pnx_lkh", 28.0)
st.session_state.setdefault("n_npx_lkh", 11.0)
st.session_state.setdefault("gm_mw_lkh", 19)
st.session_state.setdefault("gg_mw_lkh", 0)
st.session_state.setdefault("pot_alh_lkh", 377)
st.session_state.setdefault("pot_gat_lkh", 264)
st.session_state.setdefault("fug_alh_lkh", 71)
st.session_state.setdefault("fug_gat_lkh", 159)
st.session_state.setdefault("v_fondo_lkh", 0)
st.session_state.setdefault("v_gatun_lkh", 0)
st.session_state.setdefault("evap_gat_mm_lkh", 4.0)
st.session_state.setdefault("evap_alh_mm_lkh", st.session_state.get("evap_gat_mm_lkh", 4.0))
st.session_state.setdefault("fuente_evap", "Aquarius · lámina (mm/día)")
st.session_state.setdefault("curva_gat", "Daily")
st.session_state.setdefault("curva_alh", "Daily")
st.session_state.setdefault("amg", "Calcular desde nivel (ft)")
st.session_state.setdefault("ama", "Calcular desde nivel (ft)")

# ── Blindaje de opciones persistentes de sesión ──────────────────────────────
def _validar_opcion_session(key: str, opciones: list[str], default: str) -> None:
    """Evita que una sesión vieja o una edición externa deje radios fuera de rango."""
    try:
        if st.session_state.get(key, default) not in opciones:
            st.session_state[key] = default
    except Exception:
        st.session_state[key] = default

_validar_opcion_session("fuente_evap", ["Manual", "Aquarius · lámina (mm/día)", "Aquarius · caudal/volumen"], "Aquarius · lámina (mm/día)")
_validar_opcion_session("curva_gat", ["Estándar", "Daily"], "Daily")
_validar_opcion_session("curva_alh", ["Estándar", "Daily"], "Daily")
_validar_opcion_session("amg", ["Calcular desde nivel (ft)", "Manual"], "Calcular desde nivel (ft)")
_validar_opcion_session("ama", ["Calcular desde nivel (ft)", "Manual"], "Calcular desde nivel (ft)")
_validar_opcion_session("fuente_consumo_escl", ["Basado en Nivel", "Basado en LakeHouse", "Manual"], "Basado en LakeHouse")
_validar_opcion_session("modo_balance_esclusajes", ["Manual sidebar", "Sidebar + ahorro", "Modelo físico base", "Modelo físico + ahorro"], "Sidebar + ahorro")
_validar_opcion_session("met_mad", ["Manual", "Tabla de agua", "Modesto", "Bill Shaw"], "Tabla de agua")

# ═══ NIVELES OPERATIVOS — fuente única para todo el app ═══════════════════════
st.sidebar.markdown("### 📍 Niveles Operativos")
def _on_sb_gat():
    v = st.session_state.get("nivel_gat_op", 87.0)
    daily = st.session_state.get("curva_gat", "Daily") == "Daily"
    mn = float(_NV_GAT_DAILY[0])  if (daily and _DAILY_GAT_LOADED) else 55.0
    mx = float(_NV_GAT_DAILY[-1]) if (daily and _DAILY_GAT_LOADED) else 93.0
    st.session_state["ae_gat"]      = max(mn, min(mx, round(v, 1)))
    st.session_state["ae_gat_fine"] = v
nivel_gat_op = st.sidebar.number_input(
    "Nivel Gatún (ft)", min_value=55.0, max_value=93.0, step=0.01, format="%.2f", key="nivel_gat_op",
    help="Nivel actual del lago Gatún — se carga inicialmente del último día LakeHouse y puede modificarse manualmente",
    on_change=_on_sb_gat)
def _on_sb_alh():
    v = st.session_state.get("nivel_alh_op", 252.0)
    daily = st.session_state.get("curva_alh", "Daily") == "Daily"
    mn = float(_NV_ALH_DAILY[0])  if (daily and _DAILY_ALH_LOADED) else 180.0
    mx = float(_NV_ALH_DAILY[-1]) if (daily and _DAILY_ALH_LOADED) else 260.0
    st.session_state["ae_alh"]      = max(mn, min(mx, round(v, 1)))
    st.session_state["ae_alh_fine"] = v
nivel_alh_op = st.sidebar.number_input(
    "Nivel Alhajuela (ft)", min_value=160.0, max_value=260.0, step=0.01, format="%.2f", key="nivel_alh_op",
    help="Nivel actual del lago Alhajuela — se carga inicialmente del último día LakeHouse y puede modificarse manualmente",
    on_change=_on_sb_alh)
st.sidebar.markdown("---")
if _info_defaults_lkh:
    _f_lkh = _info_defaults_lkh.get("fecha_ultimo", "")
    _f_lkh_txt = _fecha_lkh_texto(_f_lkh)
    st.sidebar.success(
        f"LakeHouse aplicado: últimos {_info_defaults_lkh.get('n_dias', 5)} días · "
        f"último registro {_f_lkh_txt}. Valores editables."
    )

st.sidebar.markdown("**🧮 Fuente de consumo para el balance principal**")
modo_balance_esclusajes = st.sidebar.radio(
    "Usar en el balance:",
    ["Manual sidebar", "Sidebar + ahorro", "Modelo físico base", "Modelo físico + ahorro"],
    index=1,
    key="modo_balance_esclusajes",
    help="Define qué consumo de esclusajes alimenta los balances principales del dashboard."
)

st.sidebar.markdown("### 🚢 Esclusajes")
n_pnx = st.sidebar.slider("Panamax (PNX) / día", min_value=0.0, max_value=40.0, step=0.5, format="%.1f", key="n_pnx_lkh")
n_npx = st.sidebar.slider("NeoPanamax (NPX) / día", min_value=0.0, max_value=20.0, step=0.5, format="%.1f", key="n_npx_lkh")
n_t   = n_pnx + n_npx

st.sidebar.markdown("### 📐 Consumo por esclusaje")
_vp_nivel = _pnx_vol_base(nivel_gat_op)
_vn_nivel = _npx_vol_base(nivel_gat_op)
fuente_consumo_escl = st.sidebar.radio(
    "Fuente vol/tránsito — Esclusajes",
    ["Basado en Nivel", "Basado en LakeHouse", "Manual"],
    index=1,
    horizontal=True,
    key="fuente_consumo_escl",
    help=(
        "Basado en Nivel: usa el Nivel Gatún de 📍 Niveles Operativos. "
        "Basado en LakeHouse: usa el consumo unitario promedio del período seleccionado en 📂 Datos Lake House (1, 5, 7, 10 o 30 días). "
        "Manual: permite ingresar valores directamente."
    )
)

if fuente_consumo_escl == "Basado en Nivel":
    modo = "hm³/escl"
    vp = _vp_nivel
    vn = _vn_nivel
    st.sidebar.caption(
        f"🔗 Calculado con Nivel Gatún operativo: **{nivel_gat_op:.2f} ft** "
        f"({'LakeHouse' if _info_niveles_lkh else 'valor manual/respaldo'})."
    )
    st.sidebar.caption(
        "Para modificar el consumo, cambie el Nivel Gatún arriba."
    )
elif fuente_consumo_escl == "Basado en LakeHouse":
    modo = "hm³/escl"
    _vp_lkh = (_info_defaults_lkh or {}).get("vp_lkh") if _info_defaults_lkh else None
    _vn_lkh = (_info_defaults_lkh or {}).get("vn_lkh") if _info_defaults_lkh else None
    vp = float(_vp_lkh) if _vp_lkh is not None and pd.notna(_vp_lkh) else 0.2173
    vn = float(_vn_lkh) if _vn_lkh is not None and pd.notna(_vn_lkh) else 0.4036
    _dias_lkh_vol = int((_info_defaults_lkh or {}).get("n_dias", st.session_state.get("dias_op", 5) or 5))
    _src_lkh_vol = f"LakeHouse últimos {_dias_lkh_vol} días" if (_vp_lkh is not None or _vn_lkh is not None) else "respaldo fijo"
    st.sidebar.caption(
        f"🔗 Valores {_src_lkh_vol}: PNX **{vp:.4f}** hm³/escl · NPX **{vn:.4f}** hm³/escl"
    )
else:  # Manual
    modo = st.sidebar.radio("Entrada manual", ["hm³/escl", "cfs equiv", "m³/s equiv"], horizontal=True)
    if modo == "hm³/escl":
        vp = st.sidebar.number_input("Vol PNX (hm³)", 0.05, 0.5, float(_vp_nivel), 0.001, format="%.4f")
        vn = st.sidebar.number_input("Vol NPX (hm³)", 0.1,  0.8, float(_vn_nivel), 0.001, format="%.4f")
    elif modo == "cfs equiv":
        vp_c = st.sidebar.number_input("PNX (cfs/escl)", 20.0, 300.0, float(_vp_nivel/CFS2HM3), 0.1)
        vn_c = st.sidebar.number_input("NPX (cfs/escl)", 50.0, 500.0, float(_vn_nivel/CFS2HM3), 0.1)
        vp = vp_c*CFS2HM3; vn = vn_c*CFS2HM3
    else:
        vp_m = st.sidebar.number_input("PNX (m³/s equiv)", 0.5, 10.0, float(_vp_nivel*HM3D2M3S), 0.01)
        vn_m = st.sidebar.number_input("NPX (m³/s equiv)", 1.0, 15.0, float(_vn_nivel*HM3D2M3S), 0.01)
        vp = vp_m/HM3D2M3S; vn = vn_m/HM3D2M3S

st.sidebar.caption(f"**PNX:** {vp:.3f} hm³ = {vp*HM3D2M3S:.2f} m³/s = {vp/CFS2HM3:.1f} cfs")
st.sidebar.caption(f"**NPX:** {vn:.3f} hm³ = {vn*HM3D2M3S:.2f} m³/s = {vn/CFS2HM3:.1f} cfs")

st.sidebar.markdown("### ⚡ Generación")
st.sidebar.markdown("**Método Madden:**")
metodo_madden = st.sidebar.radio(
    "Cálculo cfs/MW Madden",
    ["Manual", "Tabla de agua", "Modesto", "Bill Shaw"],
    index=1,
    horizontal=False, key="met_mad",
    help="Manual: factor fijo · Otros: cfs/MW según nivel Alhajuela (curvas ACP)")
if metodo_madden == "Manual":
    mw_madden = st.sidebar.number_input("Factor Madden (cfs/MW)", 50.0, 1200.0, 100.00, 0.01, format="%.2f")
    if mw_madden != 100.00:
        st.sidebar.warning(f"⚠️ Factor Madden modificado · {AHORA}")
    else:
        st.sidebar.caption("Factor inicial: 100 cfs/MW")
else:
    _niv_mad_ref = nivel_alh_op  # sincronizado con Niveles Operativos
    mw_madden = madden_cfs_per_mw(_niv_mad_ref, metodo_madden)
    st.sidebar.caption(f"📊 cfs/MW calculado: **{mw_madden:.2f}** @ {_niv_mad_ref:.1f} ft ({metodo_madden})")
st.sidebar.markdown("Factor Gatún (cfs/MW):")
mw_gatun  = st.sidebar.number_input("Factor Gatún (cfs/MW)", 100.0, 250.0, 200.00, 0.01, format="%.2f")
if mw_gatun != 200.00:
    st.sidebar.warning(f"⚠️ Factor Gatún modificado · {AHORA}")
else:
    st.sidebar.caption("Factor inicial: 200 cfs/MW")
gm_mw = st.sidebar.slider("Madden (MW)", min_value=0, max_value=36, key="gm_mw_lkh")
gg_mw = st.sidebar.slider("Gatún (MW)", min_value=0, max_value=30, key="gg_mw_lkh")

st.sidebar.markdown("### 🚰 Potable (cfs)")
pot_alh = st.sidebar.number_input("Alhajuela", min_value=0, max_value=800, step=1, key="pot_alh_lkh")
pot_gat = st.sidebar.number_input("Gatún", min_value=0, max_value=600, step=1, key="pot_gat_lkh")

st.sidebar.markdown("### 💨 Fugas (cfs)")
fug_alh = st.sidebar.number_input("Alhajuela ", min_value=0, max_value=300, step=1, key="fug_alh_lkh")
fug_gat = st.sidebar.number_input("Gatún ", min_value=0, max_value=400, step=1, key="fug_gat_lkh")

st.sidebar.markdown("### 🌊 Vertidos Alhajuela (cfs)")
v_fondo  = st.sidebar.number_input(
    "Fondo Madden", 0, 5000, step=1, key="v_fondo_lkh",
    help="Se inicializa con el promedio de `madspill` del LakeHouse y queda editable."
)
v_tambor = st.sidebar.number_input("Compuertas Tambor",  0, 30000,     0, 100)
v_libre  = st.sidebar.number_input("Libre (overflow)",   0, 20000,     0, 100)
st.sidebar.caption(
    f"LakeHouse `madspill`: **{v_fondo:,.0f} cfs** aplicado como vertido total inicial de Madden."
)

st.sidebar.markdown("### 🌊 Vertidos Gatún (cfs)")
v_gatun  = st.sidebar.number_input(
    "Vertido Gatún", 0, 20000, step=1, key="v_gatun_lkh",
    help="Se inicializa con el promedio de `gatspill` del LakeHouse y queda editable."
)
st.sidebar.caption(f"LakeHouse `gatspill`: **{v_gatun:,.0f} cfs** como valor inicial editable.")

st.sidebar.markdown("### 🔄 ZZ-Flush")
flush_cc = st.sidebar.number_input("Cocolí (hrs)",  0.0, 8.0, 0.0, 0.5)
flush_ac = st.sidebar.number_input("A.Clara (hrs)", 0.0, 8.0, 0.0, 0.5)
st.sidebar.caption(
    f"Caudal instantáneo de referencia ZZ-Flush: {ZZ_FLUSH_M3S:.1f} m³/s; "
    f"2 h ≈ {ZZ_FLUSH_M3S*2*3600/1e6:.2f} hm³."
)

# ── Evaporación con opción de área desde nivel ───────────────────────────────
st.sidebar.markdown("### 💾 Ahorro de Agua — Esclusajes")
st.sidebar.caption("Modelo físico · ConsumodeAguaEsclusas.xlsb")
nivel_modelo_ft = nivel_gat_op  # sincronizado con Niveles Operativos
_H_m = nivel_modelo_ft * 0.3048
_vn_fis = _npx_vol_base(nivel_modelo_ft)
_vp_fis = _pnx_vol_base(nivel_modelo_ft)
st.sidebar.caption(
    f"Vol/tránsito modelo: **NPX** {_vn_fis:.4f} hm³ · **PNX** {_vp_fis:.4f} hm³")

st.sidebar.markdown("**🌊 NPX — Tinas de ahorro**")
pct_tinas_cc = st.sidebar.slider("Tinas Cocolí (%)",     0, 100,  0, 5, key="ptcc")
pct_tinas_ac = st.sidebar.slider("Tinas Agua Clara (%)", 0, 100,  0, 5, key="ptac")

st.sidebar.markdown("**↔️ NPX — Turn Around**")
n_turnaround_npx = st.sidebar.number_input(
    "Turn Around NPX/día", 0.0, 10.0, 0.0, 1.0, key="turn_npx",
    help="Cantidad diaria de eventos Turn Around NPX a considerar en el ahorro."
)
usar_turnaround_npx = st.sidebar.checkbox(
    "Aplicar ahorro Turn Around NPX", value=False, key="usar_turn_npx",
    help="Basado en el workbook: el ahorro del Turn Around NPX equivale a ~5% del volumen sin tinas."
)

st.sidebar.markdown("**🚢 PNX — Eficiencia operativa**")
pct_cam_corta   = st.sidebar.slider("Cámaras Cortas (%)", 0, 100, 0, 1, key="pcc")
pct_crossfill   = st.sidebar.slider("CrossFilling (%)",   0, 100,  0, 5, key="pxf")

# ── Cálculos de ahorro (se usan en la pestaña Ahorro) ─────────────────────────
_V_CC = AC_NPX * max(_H_m - EQ_CC_m, 0) * 1e-6   # hm³ / tránsito lado CC
_V_AC = AC_NPX * max(_H_m - EQ_AC_m, 0) * 1e-6   # hm³ / tránsito lado AC

ahorro_tinas_cc  = n_npx * 0.5 * _V_CC * FRAC_TINAS * pct_tinas_cc / 100
ahorro_tinas_ac  = n_npx * 0.5 * _V_AC * FRAC_TINAS * pct_tinas_ac / 100
_sav_cc_tr       = _pnx_ahorro_cc_per_transit(nivel_modelo_ft)
ahorro_cam_corta = n_pnx * _sav_cc_tr * pct_cam_corta / 100
ahorro_xfill_tr  = (pct_crossfill/100) * AC_PNX_REG * EQ_PM_ft * 0.3048 * 1e-6 * 0.5
ahorro_xfill     = n_pnx * ahorro_xfill_tr

# Turn Around NPX
TURN_NPX_SAVING_PCT = 0.05  # 5% del volumen de Turn Around sin tinas, consistente con el workbook
turnaround_npx_base_tr_modelo = 2.0 * _vn_fis
turnaround_npx_ahorro_tr_modelo = turnaround_npx_base_tr_modelo * TURN_NPX_SAVING_PCT
ahorro_turnaround_npx_modelo = (
    n_turnaround_npx * turnaround_npx_ahorro_tr_modelo if usar_turnaround_npx else 0.0
)

turnaround_npx_base_tr_sidebar = 2.0 * vn
turnaround_npx_ahorro_tr_sidebar = turnaround_npx_base_tr_sidebar * TURN_NPX_SAVING_PCT
ahorro_turnaround_npx_sidebar = (
    n_turnaround_npx * turnaround_npx_ahorro_tr_sidebar if usar_turnaround_npx else 0.0
)

ahorro_total_esc = (
    ahorro_tinas_cc + ahorro_tinas_ac + ahorro_cam_corta + ahorro_xfill
    + ahorro_turnaround_npx_modelo
)

# Vol/tránsito efectivo con estrategias activas
frac_ahorro_npx = max(0.0,
    1.0
    - 0.5 * FRAC_TINAS * pct_tinas_cc/100
    - 0.5 * FRAC_TINAS * pct_tinas_ac/100
)
vn_efectivo = max(_vn_fis * frac_ahorro_npx, 0.001)
vp_efectivo = max(_vp_fis - _sav_cc_tr * pct_cam_corta/100 - ahorro_xfill_tr, 0.001)

# Variante híbrida: consumo manual del sidebar aplicando los mismos porcentajes/ahorros
vn_sidebar_ahorro = max(vn * frac_ahorro_npx, 0.001)
vp_sidebar_ahorro = max(vp - _sav_cc_tr * pct_cam_corta/100 - ahorro_xfill_tr, 0.001)

dem_escl_modelo         = n_npx * _vn_fis + n_pnx * _vp_fis
dem_escl_efectivo       = max(n_npx * vn_efectivo + n_pnx * vp_efectivo - ahorro_turnaround_npx_modelo, 0.0)
dem_escl_sidebar_ahorro = max(n_npx * vn_sidebar_ahorro + n_pnx * vp_sidebar_ahorro - ahorro_turnaround_npx_sidebar, 0.0)

st.sidebar.markdown("### ☀️ Evaporación")
fuente_evap = st.sidebar.radio(
    "Fuente de evaporación",
    ["Manual", "Aquarius · lámina (mm/día)", "Aquarius · caudal/volumen", "LakeHouse · promedio"],
    index=1,
    key="fuente_evap",
    help=(
        "Manual: ingresa la lámina por embalse. "
        "Aquarius · lámina: usa CZL para Gatún y PMG para Alhajuela y aplica hm³/d = mm/d × área(km²) × 0.001 × 0.85. "
        "Aquarius · caudal/volumen: usa directamente las series V Evap 0.85 en hm³/día. "
        "LakeHouse · promedio: toma el promedio de evaporación del LakeHouse para el período seleccionado (1, 5, 7, 10 o 30 días)."
    ),
)

# Los campos manuales permanecen visibles y guardados; en modo Aquarius quedan bloqueados.
_evap_manual_bloqueado = fuente_evap != "Manual"
evap_gat_mm_manual = st.sidebar.number_input(
    "Lámina manual Gatún (mm/día)",
    min_value=0.0, max_value=15.0, step=0.1, key="evap_gat_mm_lkh",
    disabled=_evap_manual_bloqueado,
    help="Se usa únicamente cuando la fuente seleccionada es Manual."
)
evap_alh_mm_manual = st.sidebar.number_input(
    "Lámina manual Alhajuela (mm/día)",
    min_value=0.0, max_value=15.0, step=0.1, key="evap_alh_mm_lkh",
    disabled=_evap_manual_bloqueado,
    help="Se usa únicamente cuando la fuente seleccionada es Manual."
)
if _evap_manual_bloqueado:
    st.sidebar.caption("Los valores manuales quedan guardados como respaldo, pero no alimentan el balance mientras Aquarius esté activo.")

_evap_aq_gat = None
_evap_aq_alh = None
_evap_modo_efectivo = fuente_evap
_evap_gat_hm3_directo = None
_evap_alh_hm3_directo = None
_evap_lkh_info = _info_defaults_lkh if isinstance(globals().get("_info_defaults_lkh"), dict) else {}
_evap_lkh_n_dias = _dias_lkh_seguros(_dias_defaults_lkh)

if fuente_evap == "Aquarius · lámina (mm/día)":
    _evap_aq_gat = _cargar_serie_evap_aquarius(EVAP_AQUARIUS_FILES["mm_gat"])
    _evap_aq_alh = _cargar_serie_evap_aquarius(EVAP_AQUARIUS_FILES["mm_alh"])
    if _evap_aq_gat.get("ok") and _evap_aq_alh.get("ok"):
        evap_gat_mm = float(_evap_aq_gat["valor"])
        evap_alh_mm = float(_evap_aq_alh["valor"])
    else:
        _evap_modo_efectivo = "Manual · respaldo"
        evap_gat_mm = evap_gat_mm_manual
        evap_alh_mm = evap_alh_mm_manual
        _faltantes = [x for x in (_evap_aq_gat, _evap_aq_alh) if not x.get("ok")]
        st.sidebar.warning(
            "No fue posible usar las dos láminas de Aquarius; se aplicaron los valores manuales guardados. "
            + " | ".join(f"{x.get('archivo')}: {x.get('error')}" for x in _faltantes)
        )
elif fuente_evap == "Aquarius · caudal/volumen":
    _evap_aq_gat = _cargar_serie_evap_aquarius(EVAP_AQUARIUS_FILES["hm3_gat"])
    _evap_aq_alh = _cargar_serie_evap_aquarius(EVAP_AQUARIUS_FILES["hm3_alh"])
    if _evap_aq_gat.get("ok") and _evap_aq_alh.get("ok"):
        _evap_gat_hm3_directo = float(_evap_aq_gat["valor"])
        _evap_alh_hm3_directo = float(_evap_aq_alh["valor"])
        # Se completan después de calcular las áreas, como láminas equivalentes.
        evap_gat_mm = 0.0
        evap_alh_mm = 0.0
    else:
        _evap_modo_efectivo = "Manual · respaldo"
        evap_gat_mm = evap_gat_mm_manual
        evap_alh_mm = evap_alh_mm_manual
        _faltantes = [x for x in (_evap_aq_gat, _evap_aq_alh) if not x.get("ok")]
        st.sidebar.warning(
            "No fue posible usar los dos volúmenes de Aquarius; se aplicaron los valores manuales guardados. "
            + " | ".join(f"{x.get('archivo')}: {x.get('error')}" for x in _faltantes)
        )
elif fuente_evap == "LakeHouse · promedio":
    def _evap_lkh_num(key):
        try:
            val = _evap_lkh_info.get(key)
            if val is None or pd.isna(val):
                return None
            val = float(val)
            return val if np.isfinite(val) and val >= 0 else None
        except Exception:
            return None

    _lkh_gat_hm3 = _evap_lkh_num("evap_gat_hm3_lkh")
    _lkh_alh_hm3 = _evap_lkh_num("evap_alh_hm3_lkh")
    _lkh_gat_mm = _evap_lkh_num("evap_gat_mm")
    _lkh_alh_mm = _evap_lkh_num("evap_alh_mm")

    if _lkh_gat_hm3 is not None and _lkh_alh_hm3 is not None:
        _evap_modo_efectivo = "LakeHouse · volumen promedio"
        _evap_gat_hm3_directo = _lkh_gat_hm3
        _evap_alh_hm3_directo = _lkh_alh_hm3
        # Se completan después de calcular las áreas, como láminas equivalentes.
        evap_gat_mm = 0.0
        evap_alh_mm = 0.0
    elif _lkh_gat_mm is not None and _lkh_alh_mm is not None:
        _evap_modo_efectivo = "LakeHouse · lámina promedio"
        evap_gat_mm = _lkh_gat_mm
        evap_alh_mm = _lkh_alh_mm
    else:
        _evap_modo_efectivo = "Manual · respaldo"
        evap_gat_mm = evap_gat_mm_manual
        evap_alh_mm = evap_alh_mm_manual
        st.sidebar.warning(
            "No fue posible obtener evaporación promedio del LakeHouse para ambos embalses; "
            "se aplicaron los valores manuales guardados."
        )
else:
    evap_gat_mm = evap_gat_mm_manual
    evap_alh_mm = evap_alh_mm_manual

st.sidebar.markdown("**Área espejo de embalse**")
# ── Selector de curva hipsométrica Gatún ─────────────────────────────────────
_curva_gat_sel = st.sidebar.radio(
    "Curva hipsométrica Gatún", ["Estándar", "Daily"], index=(1 if _DAILY_GAT_LOADED else 0), horizontal=True, key="curva_gat",
    help=(
        f"Estándar: {len(_NV_GAT):,} pts hardcoded (55–89 ft)\n"
        f"Daily: {len(_NV_GAT_DAILY):,} pts @ 0.01 ft "
        f"({_NV_GAT_DAILY[0]:.2f}–{_NV_GAT_DAILY[-1]:.2f} ft) — tabla oficial ACP. "
        "Opción recomendada y cargada por defecto."
    ))
_use_daily_gat = (_curva_gat_sel == "Daily")
area_modo_gat = st.sidebar.radio("Área Gatún", ["Calcular desde nivel (ft)", "Manual"],
                                  index=0, horizontal=True, key="amg",
                                  help="Por defecto calcula el área espejo desde el nivel operativo de Gatún. Use Manual solo para pruebas o respaldo.")
if area_modo_gat == "Manual":
    area_gat = st.sidebar.number_input("Área espejo Gatún (km²)", 0.0, 500.0, 425.0, 1.0)
    nivel_gat_ft = None
    st.sidebar.caption("Área manual utilizada por la fuente activa de evaporación.")
else:
    nivel_gat_ft = nivel_gat_op  # sincronizado con Niveles Operativos
    area_gat = area_desde_nivel_gat(nivel_gat_ft, daily=_use_daily_gat)
    st.sidebar.caption(f"📐 Área calculada: **{area_gat:.4f} km²** @ {nivel_gat_ft:.2f} ft  ({'Daily' if _use_daily_gat else 'Estándar'})")
    if (not _use_daily_gat) and not (_NV_GAT[0] <= nivel_gat_ft <= _NV_GAT[-1]):
        st.sidebar.warning("Nivel Gatún fuera del rango de curva estándar; use Daily para evitar extrapolación/clamp.")

# ── Selector de curva hipsométrica Alhajuela ─────────────────────────────────
_curva_alh_sel = st.sidebar.radio(
    "Curva hipsométrica Alhajuela", ["Estándar", "Daily"], index=(1 if _DAILY_ALH_LOADED else 0), horizontal=True, key="curva_alh",
    help=(
        f"Estándar: {len(_NV_ALH):,} pts hardcoded (180–255 ft)\n"
        f"Daily: {len(_NV_ALH_DAILY):,} pts @ 0.01 ft "
        f"({_NV_ALH_DAILY[0]:.2f}–{_NV_ALH_DAILY[-1]:.2f} ft) — tabla oficial ACP. "
        "Opción recomendada y cargada por defecto."
    ))
_use_daily_alh = (_curva_alh_sel == "Daily")
area_modo_alh = st.sidebar.radio("Área Alhajuela", ["Calcular desde nivel (ft)", "Manual"],
                                   index=0, horizontal=True, key="ama",
                                   help="Por defecto calcula el área espejo desde el nivel operativo de Alhajuela. Use Manual solo para pruebas o respaldo.")
if area_modo_alh == "Manual":
    area_alh = st.sidebar.number_input("Área espejo Alhajuela (km²)", 0.0, 100.0, 49.0, 1.0)
    nivel_alh_ft = None
    st.sidebar.caption("Área manual utilizada por la fuente activa de evaporación.")
else:
    nivel_alh_ft = nivel_alh_op  # sincronizado con Niveles Operativos
    area_alh = area_desde_nivel_alh(nivel_alh_ft, daily=_use_daily_alh)
    st.sidebar.caption(f"📐 Área calculada: **{area_alh:.4f} km²** @ {nivel_alh_ft:.2f} ft  ({'Daily' if _use_daily_alh else 'Estándar'})")
    if (not _use_daily_alh) and not (_NV_ALH[0] <= nivel_alh_ft <= _NV_ALH[-1]):
        st.sidebar.warning("Nivel Alhajuela fuera del rango de curva estándar; use Daily para evitar extrapolación/clamp.")

# Fuente única de evaporación aplicada a TODO el dashboard.
# Por lámina, el cálculo operativo queda:
# hm³/día = mm/día × área(km²) × 0.001 × 0.85
# La opción Aquarius volumen ya viene como V Evap 0.85 y no se multiplica nuevamente.
if _evap_modo_efectivo in ("Aquarius · caudal/volumen", "LakeHouse · volumen promedio"):
    evap_gat = max(float(_evap_gat_hm3_directo or 0.0), 0.0)
    evap_alh = max(float(_evap_alh_hm3_directo or 0.0), 0.0)
    # Lámina equivalente para documentar el valor directo ya corregido por 0.85.
    evap_gat_mm = evap_gat / (area_gat * 1e-3 * EVAP_COEF) if area_gat > 0 and EVAP_COEF > 0 else 0.0
    evap_alh_mm = evap_alh / (area_alh * 1e-3 * EVAP_COEF) if area_alh > 0 and EVAP_COEF > 0 else 0.0
elif _evap_modo_efectivo in ("Aquarius · lámina (mm/día)", "LakeHouse · lámina promedio"):
    evap_gat = max(float(evap_gat_mm), 0.0) * area_gat * 1e-3 * EVAP_COEF
    evap_alh = max(float(evap_alh_mm), 0.0) * area_alh * 1e-3 * EVAP_COEF
else:
    evap_gat = max(float(evap_gat_mm), 0.0) * area_gat * 1e-3 * EVAP_COEF
    evap_alh = max(float(evap_alh_mm), 0.0) * area_alh * 1e-3 * EVAP_COEF

evap_tot = evap_gat + evap_alh

if _evap_modo_efectivo == "Aquarius · lámina (mm/día)":
    evap_fuente_label = "Aquarius · lámina diaria × área × 0.85"
    evap_fuente_corta = "Aquarius mm"
    evap_detalle_gat = f"CZL · {_evap_aq_gat.get('fecha', 'N/D')}"
    evap_detalle_alh = f"PMG · {_evap_aq_alh.get('fecha', 'N/D')}"
elif _evap_modo_efectivo == "Aquarius · caudal/volumen":
    evap_fuente_label = "Aquarius · volumen V Evap 0.85"
    evap_fuente_corta = "Aquarius volumen"
    evap_detalle_gat = f"GAT · {_evap_aq_gat.get('fecha', 'N/D')}"
    evap_detalle_alh = f"MAD · {_evap_aq_alh.get('fecha', 'N/D')}"
elif _evap_modo_efectivo == "LakeHouse · volumen promedio":
    evap_fuente_label = f"LakeHouse · promedio de volumen ({_evap_lkh_n_dias} días)"
    evap_fuente_corta = "LakeHouse promedio"
    evap_detalle_gat = f"vol_evap_gat_hm3 · últimos {_evap_lkh_n_dias} días"
    evap_detalle_alh = f"vol_evap_ala_hm3 · últimos {_evap_lkh_n_dias} días"
elif _evap_modo_efectivo == "LakeHouse · lámina promedio":
    evap_fuente_label = f"LakeHouse · promedio de lámina ({_evap_lkh_n_dias} días) × área × 0.85"
    evap_fuente_corta = "LakeHouse mm"
    evap_detalle_gat = f"evap_gatun_mm · últimos {_evap_lkh_n_dias} días"
    evap_detalle_alh = f"evap_alaj_mm · últimos {_evap_lkh_n_dias} días"
elif _evap_modo_efectivo == "Manual · respaldo":
    evap_fuente_label = "Manual · respaldo por archivo Aquarius no disponible"
    evap_fuente_corta = "Manual respaldo"
    evap_detalle_gat = "Lámina manual guardada"
    evap_detalle_alh = "Lámina manual guardada"
else:
    evap_fuente_label = "Manual · lámina × área × 0.85"
    evap_fuente_corta = "Manual"
    evap_detalle_gat = "Lámina manual"
    evap_detalle_alh = "Lámina manual"

st.sidebar.markdown("**Resultado aplicado al balance**")
st.sidebar.caption(f"Fórmula por lámina: hm³/d = mm/d × área(km²) × 0.001 × {EVAP_COEF:.2f}.")
st.sidebar.caption(
    f"**Gatún:** {evap_gat:.4f} hm³/d · {evap_gat/CFS2HM3:.1f} cfs · "
    f"{evap_gat_mm:.2f} mm/día ({evap_detalle_gat})"
)
st.sidebar.caption(
    f"**Alhajuela:** {evap_alh:.4f} hm³/d · {evap_alh/CFS2HM3:.1f} cfs · "
    f"{evap_alh_mm:.2f} mm/día ({evap_detalle_alh})"
)
st.sidebar.caption(f"Fuente activa: **{evap_fuente_label}**")

st.sidebar.markdown("---")
unidad  = st.sidebar.radio("Unidad visual", ["hm³/día", "cfs", "m³/s"], horizontal=True)
u_label = unidad
u_cv    = 1 if unidad == "hm³/día" else (1/CFS2HM3 if unidad == "cfs" else HM3D2M3S)
st.sidebar.markdown("---")
st.sidebar.caption(f"📅 Sesión: {AHORA}")


# ═══ CÁLCULOS ═══
if modo_balance_esclusajes == "Manual sidebar":
    vp_balance = vp
    vn_balance = vn
    ahorro_turnaround_aplicado = 0.0
    balance_escl_label = "Manual sidebar"
elif modo_balance_esclusajes == "Sidebar + ahorro":
    vp_balance = vp_sidebar_ahorro
    vn_balance = vn_sidebar_ahorro
    ahorro_turnaround_aplicado = ahorro_turnaround_npx_sidebar
    balance_escl_label = "Sidebar + ahorro"
elif modo_balance_esclusajes == "Modelo físico base":
    vp_balance = _vp_fis
    vn_balance = _vn_fis
    ahorro_turnaround_aplicado = 0.0
    balance_escl_label = "Modelo físico base"
else:
    vp_balance = vp_efectivo
    vn_balance = vn_efectivo
    ahorro_turnaround_aplicado = ahorro_turnaround_npx_modelo
    balance_escl_label = "Modelo físico + ahorro"

dem_pnx       = n_pnx * vp_balance
dem_npx_bruto = n_npx * vn_balance
dem_npx       = max(dem_npx_bruto - ahorro_turnaround_aplicado, 0.0)
dem_escl      = dem_pnx + dem_npx
gen_alh   = gm_mw*mw_madden*CFS2HM3
gen_gat   = gg_mw*mw_gatun *CFS2HM3
gen_tot   = gen_alh+gen_gat
alh_pot   = pot_alh*CFS2HM3; gat_pot = pot_gat*CFS2HM3
alh_fug   = fug_alh*CFS2HM3; gat_fug = fug_gat*CFS2HM3
alh_vf    = v_fondo*CFS2HM3; alh_vt = v_tambor*CFS2HM3; alh_vl = v_libre*CFS2HM3
alh_vert  = alh_vf+alh_vt+alh_vl
gat_ver   = v_gatun*CFS2HM3
dem_flush = ZZ_FLUSH_M3S*(flush_cc+flush_ac)*3600/1e6
# La evaporación ya fue resuelta una sola vez según la fuente seleccionada en el sidebar.

alh_total = gen_alh+alh_pot+alh_fug+alh_vert+evap_alh
gat_total = gen_gat+gat_pot+gat_fug+gat_ver+dem_escl+dem_flush+evap_gat
dem_total = alh_total+gat_total

alh_usos = {
    "Generación Madden":  (gen_alh,  gm_mw*mw_madden,     COL["generacion"]),
    "Agua Potable":       (alh_pot,  pot_alh,              COL["potable"]),
    "Fugas":              (alh_fug,  fug_alh,              COL["fugas"]),
    "Vertido fondo":      (alh_vf,   v_fondo,              "#7f8c8d"),
    "Compuertas Tambor":  (alh_vt,   v_tambor,             COL["tambor"]),
    "Vertido libre":      (alh_vl,   v_libre,              COL["vertidos"]),
    "Evaporación":        (evap_alh, evap_alh/CFS2HM3,     COL["evap"]),
}
gat_usos = {
    "Esclusajes PNX":  (dem_pnx,  dem_pnx/CFS2HM3,         COL["pnx"]),
    "Esclusajes NPX":  (dem_npx,  dem_npx/CFS2HM3,         COL["npx"]),
    "ZZ-Flush":        (dem_flush, dem_flush/CFS2HM3,       COL["flush"]),
    "Generación Gatún":(gen_gat,  gg_mw*mw_gatun,          COL["gatgen"]),
    "Agua Potable":    (gat_pot,  pot_gat,                  COL["potable"]),
    "Fugas":           (gat_fug,  fug_gat,                  COL["fugas"]),
    "Vertido Gatún":   (gat_ver,  v_gatun,                  COL["vertidos"]),
    "Evaporación":     (evap_gat, evap_gat/CFS2HM3,         COL["evap"]),
}


# ═══ HEADER ═══
hdr_c1, hdr_c2, hdr_c3 = st.columns([1, 5, 1])
with hdr_c1:
    # Logo Canal de Panamá — izquierda
    _cp_tag = _img_tag(_logo_cp_mime, _logo_cp, "width:110px;margin-top:4px;")
    if _cp_tag:
        st.markdown(_cp_tag, unsafe_allow_html=True)
with hdr_c2:
    st.markdown(
        "<h1 style='color:#1a5276;margin-bottom:0;text-align:center;'>"
        "💧 Demandas de Agua por Embalse</h1>"
        "<p style='color:#5d6d7e;margin-top:-4px;text-align:center;'>"
        "Canal de Panamá · <b>HIMH — Sección de Hidrología</b> · Creado para HIMH por: JFRodriguez</p>",
        unsafe_allow_html=True)
with hdr_c3:
    # Logo HIMH — derecha
    _himh_tag = _img_tag(_logo_mime, _logo, "width:80px;margin-top:4px;float:right;")
    if _himh_tag:
        st.markdown(_himh_tag, unsafe_allow_html=True)

_fecha_lkh_encabezado = _fecha_lkh_texto(
    (_info_defaults_lkh or {}).get("fecha_ultimo") if _info_defaults_lkh else None
)
if _info_defaults_lkh:
    st.info(
        f"🗓️ **Último dato disponible del LakeHouse: {_fecha_lkh_encabezado}** · "
        f"promedio operativo aplicado: **últimos {_info_defaults_lkh.get('n_dias', _dias_defaults_lkh)} días**."
    )
else:
    st.warning("🗓️ LakeHouse no disponible; se mantienen los valores manuales/de respaldo del app.")

# ── Visor principal único: valores organizados por embalse ───────────────────
st.markdown("#### 📌 Variables operativas por embalse")
_src_col, _dias_col, _obs_col = st.columns([2.35, 1.0, 0.95])
with _src_col:
    fuente_kpis_superiores = st.radio(
        "Fuente de los valores",
        ["Calculado por app", "Promedio LakeHouse", "Comparar App vs LakeHouse"],
        index=1,
        horizontal=True,
        key="fuente_kpis_superiores",
        help=(
            "Cambia únicamente los valores mostrados en este visor. "
            "No modifica los cálculos del balance ni los controles del sidebar."
        ),
    )
with _dias_col:
    _dias_kpi_opciones = [1, 5, 7, 10, 30]
    try:
        _dias_kpi_actual = int(st.session_state.get("dias_op", 5))
    except Exception:
        _dias_kpi_actual = 5
    _dias_kpi_index = _dias_kpi_opciones.index(_dias_kpi_actual) if _dias_kpi_actual in _dias_kpi_opciones else 0
    dias_kpi_superiores = st.radio(
        "Promedio LakeHouse",
        options=_dias_kpi_opciones,
        index=_dias_kpi_index,
        horizontal=True,
        key="dias_op",
        help="Selecciona 1, 5, 7, 10 o 30 registros/días para los promedios del LakeHouse.",
    )
with _obs_col:
    mostrar_aportes_observados = st.checkbox(
        "Ver aportes obs.",
        value=False,
        key="mostrar_aportes_observados",
        help="Muestra un visor compacto con los aportes observados diarios de Aquarius (Discharge AT GAT/ALHA).",
    )

_detalle_kpi_lkh = (_info_balance_lkh or {}).get("detalle", [])
_zz_lkh = (_info_balance_lkh or {}).get("zz_flush_hm3")


def _buscar_detalle_lkh(embalse, uso=None, prefijo_uso=None):
    _valores = []
    for _item in _detalle_kpi_lkh:
        try:
            if str(_item.get("Embalse", "")) != str(embalse):
                continue
            _uso_item = str(_item.get("Uso", ""))
            _coincide = (_uso_item == uso) if uso is not None else _uso_item.startswith(str(prefijo_uso or ""))
            _valor = _item.get("hm3")
            if _coincide and _valor is not None and pd.notna(_valor):
                _valores.append(float(_valor))
        except Exception:
            continue
    return float(sum(_valores)) if _valores else None


_detalle_principal_lkh = {
    "alh_gen":  _buscar_detalle_lkh("Alhajuela ", uso="Generación Madden"),
    "alh_pot":  _buscar_detalle_lkh("Alhajuela ", uso="Potabilización"),
    "alh_fug":  _buscar_detalle_lkh("Alhajuela ", uso="Fugas"),
    "alh_ver":  _buscar_detalle_lkh("Alhajuela ", prefijo_uso="Vertido"),
    # La evaporación usa la fuente activa seleccionada en el sidebar.
    "gat_pnx":  _buscar_detalle_lkh("Gatún", uso="Esclusajes PNX"),
    "gat_npx":  _buscar_detalle_lkh("Gatún", uso="Esclusajes NPX"),
    "gat_gen":  _buscar_detalle_lkh("Gatún", uso="Generación Gatún"),
    "gat_pot":  _buscar_detalle_lkh("Gatún", uso="Potabilización"),
    "gat_fug":  _buscar_detalle_lkh("Gatún", uso="Fugas"),
    "gat_ver":  _buscar_detalle_lkh("Gatún", prefijo_uso="Vertido"),
    "gat_flush": float(_zz_lkh) if _zz_lkh is not None and pd.notna(_zz_lkh) else None,
}


def _total_promedio_lkh(componentes, evaporacion_app):
    """Suma el promedio LakeHouse por embalse y agrega la evaporación de la fuente activa."""
    _total = 0.0
    for _valor_lkh, _valor_app in componentes:
        if _valor_lkh is not None and pd.notna(_valor_lkh):
            _total += float(_valor_lkh)
        elif _valor_app is not None and pd.notna(_valor_app) and abs(float(_valor_app)) < 1e-12:
            # Un componente operativo igual a cero no invalida el total si no existe columna LakeHouse.
            _total += 0.0
        else:
            return None
    return _total + float(evaporacion_app)


_alh_total_lkh = _total_promedio_lkh(
    [
        (_detalle_principal_lkh["alh_gen"], gen_alh),
        (_detalle_principal_lkh["alh_pot"], alh_pot),
        (_detalle_principal_lkh["alh_fug"], alh_fug),
        (_detalle_principal_lkh["alh_ver"], alh_vert),
    ],
    evap_alh,
)
_gat_total_lkh = _total_promedio_lkh(
    [
        (_detalle_principal_lkh["gat_pnx"], dem_pnx),
        (_detalle_principal_lkh["gat_npx"], dem_npx),
        (_detalle_principal_lkh["gat_gen"], gen_gat),
        (_detalle_principal_lkh["gat_pot"], gat_pot),
        (_detalle_principal_lkh["gat_fug"], gat_fug),
        (_detalle_principal_lkh["gat_ver"], gat_ver),
        (_detalle_principal_lkh["gat_flush"], dem_flush),
    ],
    evap_gat,
)

_mad_mw_lkh = (_info_balance_lkh or {}).get("mad_mw")
_gat_mw_lkh = (_info_balance_lkh or {}).get("gat_mw")
_n_pnx_lkh = (_info_balance_lkh or {}).get("n_pnx")
_n_npx_lkh = (_info_balance_lkh or {}).get("n_npx")


def _nota_mw(valor, prefijo="Potencia"):
    if valor is None or pd.isna(valor):
        return None
    return f"{prefijo}: {fmt_sig(float(valor), 3)} MW"


_detalle_principal_app_alh = [
    {
        "etiqueta": "Total Alhajuela",
        "app": alh_total,
        "lkh": _alh_total_lkh,
        "solo_app": False,
        "fuente_lkh": f"LakeHouse + evap. {evap_fuente_corta} · {dias_kpi_superiores} días",
    },
    {
        "etiqueta": "Generación Madden",
        "app": gen_alh,
        "lkh": _detalle_principal_lkh["alh_gen"],
        "solo_app": False,
        "nota_app": _nota_mw(gm_mw),
        "nota_lkh": _nota_mw(_mad_mw_lkh, "Potencia media"),
    },
    {"etiqueta": "Potabilización", "app": alh_pot, "lkh": _detalle_principal_lkh["alh_pot"], "solo_app": False},
    {"etiqueta": "Fugas", "app": alh_fug, "lkh": _detalle_principal_lkh["alh_fug"], "solo_app": False},
    {"etiqueta": "Vertido Madden", "app": alh_vert, "lkh": _detalle_principal_lkh["alh_ver"], "solo_app": False},
    {"etiqueta": "Evaporación", "app": evap_alh, "lkh": None, "solo_app": True, "fuente_app": evap_fuente_label},
]

_detalle_principal_app_gat = [
    {
        "etiqueta": "Total Gatún",
        "app": gat_total,
        "lkh": _gat_total_lkh,
        "solo_app": False,
        "fuente_lkh": f"LakeHouse + evap. {evap_fuente_corta} · {dias_kpi_superiores} días",
    },
    {
        "etiqueta": "Generación Gatún",
        "app": gen_gat,
        "lkh": _detalle_principal_lkh["gat_gen"],
        "solo_app": False,
        "nota_app": _nota_mw(gg_mw),
        "nota_lkh": _nota_mw(_gat_mw_lkh, "Potencia media"),
    },
    {"etiqueta": "Potabilización", "app": gat_pot, "lkh": _detalle_principal_lkh["gat_pot"], "solo_app": False},
    {"etiqueta": "Fugas", "app": gat_fug, "lkh": _detalle_principal_lkh["gat_fug"], "solo_app": False},
    {"etiqueta": "Vertido Gatún", "app": gat_ver, "lkh": _detalle_principal_lkh["gat_ver"], "solo_app": False},
    {"etiqueta": "ZZ-Flush", "app": dem_flush, "lkh": _detalle_principal_lkh["gat_flush"], "solo_app": False},
    {"etiqueta": "Evaporación", "app": evap_gat, "lkh": None, "solo_app": True, "fuente_app": evap_fuente_label},
]


def _seleccionar_flujo_principal(valor_app, valor_lkh, solo_app=False, fuente_lkh=None, fuente_app=None):
    _app_ok = valor_app is not None and pd.notna(valor_app)
    if solo_app:
        return (float(valor_app) if _app_ok else None), (fuente_app or "App calculado"), None, "app"
    _lkh_ok = valor_lkh is not None and pd.notna(valor_lkh)
    _etiqueta_lkh = fuente_lkh or f"LakeHouse · {dias_kpi_superiores} días"

    if fuente_kpis_superiores == "Promedio LakeHouse":
        if _lkh_ok:
            return float(valor_lkh), _etiqueta_lkh, None, "lkh"
        return (float(valor_app) if _app_ok else None), "App · respaldo", "LakeHouse: N/D", "app"

    if fuente_kpis_superiores == "Comparar App vs LakeHouse":
        _comparacion = f"LakeHouse: {float(valor_lkh):.3f} hm³/d" if _lkh_ok else "LakeHouse: N/D"
        return (float(valor_app) if _app_ok else None), "App calculado", _comparacion, "app"

    return (float(valor_app) if _app_ok else None), "App calculado", None, "app"


def _equiv_esclusajes_txt(valor_hm3, prefijo="Eq. esclusajes"):
    """Texto compacto de equivalencia en esclusajes PNX/NPX para cualquier volumen diario."""
    try:
        h = float(valor_hm3)
        vp_ref = float(globals().get("vp_balance", globals().get("vp", 0.0)) or 0.0)
        vn_ref = float(globals().get("vn_balance", globals().get("vn", 0.0)) or 0.0)
        partes = []
        if vp_ref > 0:
            partes.append(f"{h / vp_ref:.1f} PNX")
        if vn_ref > 0:
            partes.append(f"{h / vn_ref:.1f} NPX")
        return f"{prefijo}: " + " · ".join(partes) if partes else ""
    except Exception:
        return ""


def _eed_txt(valor_hm3, prefijo="EED"):
    """Equivalente EED: 1 EED = 55 Mgal/d.

    Se muestra solamente el valor equivalente de EED para no mezclarlo
    visualmente con las conversiones hidrológicas p³/s y m³/s.
    """
    try:
        h = float(valor_hm3)
        if EED_HM3D <= 0:
            return ""
        return f"{prefijo}: {h / EED_HM3D:.1f}"
    except Exception:
        return ""


def _equiv_esclusajes_html(valor_hm3):
    txt = _equiv_esclusajes_txt(valor_hm3)
    eed = _eed_txt(valor_hm3)
    partes = []
    if txt:
        partes.append(f"🚢 {txt}")
    if eed:
        partes.append(f"💧 {eed}")
    if not partes:
        return ""
    return f'<div style="font-size:0.74rem;margin-top:4px;color:rgba(71,85,105,.86);">' + '<br>'.join(partes) + '</div>'


def _tarjeta_flujo_principal(
    etiqueta,
    valor_app,
    valor_lkh,
    clase_embalse,
    solo_app=False,
    nota_app=None,
    nota_lkh=None,
    fuente_lkh=None,
    fuente_app=None,
):
    _hm3, _fuente, _comparacion, _origen = _seleccionar_flujo_principal(
        valor_app,
        valor_lkh,
        solo_app=solo_app,
        fuente_lkh=fuente_lkh,
        fuente_app=fuente_app,
    )
    if _hm3 is None or pd.isna(_hm3):
        _valor_txt = "N/D"
        _conversion_txt = "p³/s: N/D · m³/s: N/D"
    else:
        _hm3 = float(_hm3)
        _valor_txt = f"{_hm3:.3f} hm³/d"
        _conversion_txt = f"{_hm3 / CFS2HM3:.1f} p³/s · {_hm3 * HM3D2M3S:.2f} m³/s"

    _nota = nota_lkh if _origen == "lkh" and nota_lkh else nota_app
    _nota_html = (
        f'<div style="font-size:0.78rem;margin-top:5px;color:rgba(71,85,105,.96);">{_nota}</div>'
        if _nota else ""
    )
    if fuente_kpis_superiores == "Comparar App vs LakeHouse" and nota_lkh and _comparacion and _comparacion != "LakeHouse: N/D":
        _comparacion = f"{_comparacion} · {nota_lkh}"
    _comparacion_html = f'<div class="flow-compare">{_comparacion}</div>' if _comparacion else ""
    _equiv_html = _equiv_esclusajes_html(_hm3) if (_hm3 is not None and not pd.isna(_hm3)) else ""

    st.markdown(
        f"""
        <div class="flow-card {clase_embalse}">
            <div class="flow-label">{etiqueta}</div>
            <div class="flow-value">{_valor_txt}</div>
            <div class="flow-conv">{_conversion_txt}</div>
            {_equiv_html}
            <div class="flow-source">{_fuente}</div>
            {_nota_html}
            {_comparacion_html}
        </div>
        """,
        unsafe_allow_html=True,
    )

def _tarjeta_aporte_observado(embalse, resumen, archivo):
    if resumen is None:
        st.markdown(
            f'''
            <div class="lkh-card">
                <div class="label">📈 Aporte observado {embalse}</div>
                <div class="value">N/D</div>
                <div class="sub">Archivo no disponible o sin datos válidos.</div>
            </div>
            ''',
            unsafe_allow_html=True,
        )
        return
    fecha_txt = pd.to_datetime(resumen["fecha"]).strftime("%d/%m/%Y")
    st.markdown(
        f'''
        <div class="lkh-card">
            <div class="label">📈 Aporte observado {embalse}</div>
            <div class="value">{resumen['ultimo_hm3']:.2f} hm³/d</div>
            <div class="sub">{resumen['ultimo_m3s']:.2f} m³/s · {resumen['ultimo_cfs']:,.0f} p³/s · último {fecha_txt}</div>
            <div class="sub">Prom. {resumen['n']}d: {resumen['prom_hm3']:.2f} hm³/d · {resumen['prom_m3s']:.2f} m³/s · {resumen['prom_cfs']:,.0f} p³/s</div>
            <div class="sub">Prom. {resumen['n30']}d: {resumen['prom30_hm3']:.2f} hm³/d · {resumen['prom30_m3s']:.2f} m³/s · {resumen['prom30_cfs']:,.0f} p³/s</div>
            <div class="sub">{archivo}</div>
        </div>
        ''',
        unsafe_allow_html=True,
    )




if mostrar_aportes_observados:
    _aportes_obs = _cargar_aportes_observados()
    st.markdown("##### 📈 Aportes observados — Aquarius Discharge AT")
    _obs_cols = st.columns(3)
    _res_gat = _resumen_aportes_df(_aportes_obs.get("Gatún", {}).get("df"), dias_prom=int(dias_kpi_superiores))
    _res_alh = _resumen_aportes_df(_aportes_obs.get("Alhajuela", {}).get("df"), dias_prom=int(dias_kpi_superiores))
    with _obs_cols[0]:
        _tarjeta_aporte_observado("Gatún", _res_gat, _aportes_obs.get("Gatún", {}).get("archivo", ""))
    with _obs_cols[1]:
        _tarjeta_aporte_observado("Alhajuela", _res_alh, _aportes_obs.get("Alhajuela", {}).get("archivo", ""))
    with _obs_cols[2]:
        if _res_gat and _res_alh:
            _total_ult_cfs = _res_gat["ultimo_cfs"] + _res_alh["ultimo_cfs"]
            _total_ult_hm3 = _res_gat["ultimo_hm3"] + _res_alh["ultimo_hm3"]
            _total_ult_m3s = _res_gat["ultimo_m3s"] + _res_alh["ultimo_m3s"]
            _total_p7_cfs = _res_gat["prom_cfs"] + _res_alh["prom_cfs"]
            _total_p7_hm3 = _res_gat["prom_hm3"] + _res_alh["prom_hm3"]
            _total_p7_m3s = _res_gat["prom_m3s"] + _res_alh["prom_m3s"]
            _total_p30_cfs = _res_gat["prom30_cfs"] + _res_alh["prom30_cfs"]
            _total_p30_hm3 = _res_gat["prom30_hm3"] + _res_alh["prom30_hm3"]
            _total_p30_m3s = _res_gat["prom30_m3s"] + _res_alh["prom30_m3s"]
            st.markdown(
                f'''
                <div class="lkh-card">
                    <div class="label">📈 Aporte observado total</div>
                    <div class="value">{_total_ult_hm3:.2f} hm³/d</div>
                    <div class="sub">{_total_ult_m3s:.2f} m³/s · {_total_ult_cfs:,.0f} p³/s · Gatún + Alhajuela</div>
                    <div class="sub">Prom. {dias_kpi_superiores}d: {_total_p7_hm3:.2f} hm³/d · {_total_p7_m3s:.2f} m³/s · {_total_p7_cfs:,.0f} p³/s</div>
                    <div class="sub">Prom. 30d: {_total_p30_hm3:.2f} hm³/d · {_total_p30_m3s:.2f} m³/s · {_total_p30_cfs:,.0f} p³/s</div>
                    <div class="sub">Use la pestaña 📈 Aportes observados para la gráfica completa.</div>
                </div>
                ''',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '''
                <div class="lkh-card">
                    <div class="label">📈 Aporte observado total</div>
                    <div class="value">N/D</div>
                    <div class="sub">Se requiere Gatún y Alhajuela válidos.</div>
                </div>
                ''',
                unsafe_allow_html=True,
            )


st.markdown("##### 🏔️ Alhajuela")
_cols_det_alh = st.columns(3)
for _i, _item in enumerate(_detalle_principal_app_alh):
    with _cols_det_alh[_i % len(_cols_det_alh)]:
        _tarjeta_flujo_principal(
            _item["etiqueta"],
            _item["app"],
            _item["lkh"],
            "flow-card-alh",
            solo_app=_item.get("solo_app", False),
            nota_app=_item.get("nota_app"),
            nota_lkh=_item.get("nota_lkh"),
            fuente_lkh=_item.get("fuente_lkh"),
            fuente_app=_item.get("fuente_app"),
        )

st.markdown("##### 🌊 Gatún")

# ── Esclusajes integrados en el mismo visor de Gatún ─────────────────────────
def _numero_valido(_valor):
    return _valor is not None and pd.notna(_valor)


def _resumen_esclusajes_seleccionado():
    """Devuelve cantidad y consumo PNX/NPX según la fuente visible."""
    _app = {
        "pnx_n": float(n_pnx),
        "npx_n": float(n_npx),
        "pnx_hm3": float(dem_pnx),
        "npx_hm3": float(dem_npx),
        "fuente": "Calculado por app",
    }

    _lkh_disponible = all(
        _numero_valido(_v)
        for _v in (
            _n_pnx_lkh,
            _n_npx_lkh,
            _detalle_principal_lkh["gat_pnx"],
            _detalle_principal_lkh["gat_npx"],
        )
    )
    _lkh = None
    if _lkh_disponible:
        _lkh = {
            "pnx_n": float(_n_pnx_lkh),
            "npx_n": float(_n_npx_lkh),
            "pnx_hm3": float(_detalle_principal_lkh["gat_pnx"]),
            "npx_hm3": float(_detalle_principal_lkh["gat_npx"]),
            "fuente": f"Promedio LakeHouse · {dias_kpi_superiores} días",
        }

    if fuente_kpis_superiores == "Promedio LakeHouse" and _lkh is not None:
        return _lkh, None
    if fuente_kpis_superiores == "Comparar App vs LakeHouse":
        return _app, _lkh
    return _app, None


_escl_vis, _escl_comparar = _resumen_esclusajes_seleccionado()
_escl_vis["total_n"] = _escl_vis["pnx_n"] + _escl_vis["npx_n"]
_escl_vis["total_hm3"] = _escl_vis["pnx_hm3"] + _escl_vis["npx_hm3"]

if _escl_comparar is not None:
    _escl_comparar["total_n"] = _escl_comparar["pnx_n"] + _escl_comparar["npx_n"]
    _escl_comparar["total_hm3"] = _escl_comparar["pnx_hm3"] + _escl_comparar["npx_hm3"]


def _tarjeta_esclusaje_integrada(etiqueta, cantidad, consumo_hm3, comparacion=None):
    """Tarjeta de esclusajes: cantidad diaria, volumen en hm³/d y equivalencias de caudal."""
    _cantidad_txt = f"{fmt_sig(float(cantidad), 3)}/día" if _numero_valido(cantidad) else "N/D"

    if _numero_valido(consumo_hm3):
        _consumo_hm3 = float(consumo_hm3)
        _volumen_txt = f"Consumo: {_consumo_hm3:.3f} hm³/d"
        _conversion_txt = (
            f"{_consumo_hm3 / CFS2HM3:.1f} p³/s · "
            f"{_consumo_hm3 * HM3D2M3S:.2f} m³/s"
        )
    else:
        _volumen_txt = "Consumo: N/D"
        _conversion_txt = "p³/s: N/D · m³/s: N/D"

    _equiv_html = _equiv_esclusajes_html(consumo_hm3) if _numero_valido(consumo_hm3) else ""
    _comparacion_html = ""
    if comparacion is not None:
        _cant_cmp, _hm3_cmp = comparacion
        if _numero_valido(_cant_cmp) and _numero_valido(_hm3_cmp):
            _hm3_cmp = float(_hm3_cmp)
            _comparacion_html = (
                '<div class="flow-compare">'
                f'LakeHouse: {fmt_sig(float(_cant_cmp), 3)}/día · '
                f'{_hm3_cmp:.3f} hm³/d · '
                f'{_hm3_cmp / CFS2HM3:.1f} p³/s · '
                f'{_hm3_cmp * HM3D2M3S:.2f} m³/s'
                '</div>'
            )
        else:
            _comparacion_html = '<div class="flow-compare">LakeHouse: N/D</div>'

    st.markdown(
        f"""
        <div class="flow-card flow-card-gat">
            <div class="flow-label">{etiqueta}</div>
            <div class="flow-value">{_cantidad_txt}</div>
            <div class="flow-conv"><b>{_volumen_txt}</b></div>
            <div class="flow-conv">{_conversion_txt}</div>
            {_equiv_html}
            {_comparacion_html}
        </div>
        """,
        unsafe_allow_html=True,
    )


# Un solo conjunto de tarjetas para Gatún: total, esclusajes y demás demandas.
_gat_items_unificados = [
    ("flujo", _detalle_principal_app_gat[0]),
    (
        "esclusaje",
        {
            "etiqueta": "Esclusajes PNX",
            "cantidad": _escl_vis["pnx_n"],
            "hm3": _escl_vis["pnx_hm3"],
            "comparacion": (
                (_escl_comparar["pnx_n"], _escl_comparar["pnx_hm3"])
                if _escl_comparar is not None else None
            ),
        },
    ),
    (
        "esclusaje",
        {
            "etiqueta": "Esclusajes NPX",
            "cantidad": _escl_vis["npx_n"],
            "hm3": _escl_vis["npx_hm3"],
            "comparacion": (
                (_escl_comparar["npx_n"], _escl_comparar["npx_hm3"])
                if _escl_comparar is not None else None
            ),
        },
    ),
    (
        "esclusaje",
        {
            "etiqueta": "Total de esclusajes",
            "cantidad": _escl_vis["total_n"],
            "hm3": _escl_vis["total_hm3"],
            "comparacion": (
                (_escl_comparar["total_n"], _escl_comparar["total_hm3"])
                if _escl_comparar is not None else None
            ),
        },
    ),
]
_gat_items_unificados.extend(("flujo", _item) for _item in _detalle_principal_app_gat[1:])

_cols_det_gat = st.columns(4)
for _i, (_tipo, _item) in enumerate(_gat_items_unificados):
    with _cols_det_gat[_i % len(_cols_det_gat)]:
        if _tipo == "esclusaje":
            _tarjeta_esclusaje_integrada(
                _item["etiqueta"],
                _item["cantidad"],
                _item["hm3"],
                comparacion=_item.get("comparacion"),
            )
        else:
            _tarjeta_flujo_principal(
                _item["etiqueta"],
                _item["app"],
                _item["lkh"],
                "flow-card-gat",
                solo_app=_item.get("solo_app", False),
                nota_app=_item.get("nota_app"),
                nota_lkh=_item.get("nota_lkh"),
                fuente_lkh=_item.get("fuente_lkh"),
            )


# Resumen pequeño del equivalente de todo el visor mostrado.
try:
    _alh_vis_hm3, _, _, _ = _seleccionar_flujo_principal(alh_total, _alh_total_lkh)
    _gat_vis_hm3, _, _, _ = _seleccionar_flujo_principal(gat_total, _gat_total_lkh)
    if _alh_vis_hm3 is not None and _gat_vis_hm3 is not None:
        _total_vis_hm3 = float(_alh_vis_hm3) + float(_gat_vis_hm3)
        _eq_total_txt = _equiv_esclusajes_txt(_total_vis_hm3, prefijo="Total mostrado equivale a")
        _eed_total_txt = _eed_txt(_total_vis_hm3, prefijo="EED total")
        st.caption(
            f"🚢 **Equivalente total de los ítems visibles:** {_total_vis_hm3:.3f} hm³/d · "
            f"{_total_vis_hm3/CFS2HM3:.1f} p³/s · {_total_vis_hm3 * HM3D2M3S:.2f} m³/s · "
            f"{_eq_total_txt} · {_eed_total_txt}."
        )
except Exception:
    pass


if _info_balance_lkh:
    _fecha_kpi_lkh = _fecha_lkh_texto(_info_balance_lkh.get("fecha_ultimo"))
    st.caption(
        f"LakeHouse: hoja **{_info_balance_lkh.get('hoja', '')}** · "
        f"promedio **{_info_balance_lkh.get('n_dias', dias_kpi_superiores)} días** · "
        f"último dato **{_fecha_kpi_lkh}**. "
        "Los vertidos Madden y Gatún se inicializan desde `madspill` y `gatspill`; "
        f"la evaporación mostrada y usada en cada embalse proviene de **{evap_fuente_label}**."
    )
else:
    st.caption("LakeHouse no disponible; el visor mantiene los valores calculados por el app.")


st.markdown("---")


# ═══ TABS ═══
tabs = st.tabs(["📊 Balance", "🏔️ Alhajuela", "🌊 Gatún",
                "🚢 Esclusajes", "⚡ Generación",
                "💾 Ahorro de Agua",
                "📐 Área Espejo", "🔄 Conversor", "📤 Exportar", "📂 Datos Lake House",
                "📘 Instructivo", "📈 Aportes observados"])


# ═══ TAB 0 — BALANCE ═══
with tabs[0]:
    b1, b2 = st.columns(2)
    with b1:
        st.subheader("Por embalse")
        fig_b1 = go.Figure(go.Bar(x=["Alhajuela","Gatún","Total"],
            y=[alh_total*u_cv, gat_total*u_cv, dem_total*u_cv],
            marker_color=[COL["alhajuela"],COL["gatun"],COL["total"]],
            text=[f"{alh_total*u_cv:.2f}",f"{gat_total*u_cv:.2f}",f"{dem_total*u_cv:.2f}"],
            textposition="auto"))
        fig_b1.update_layout(yaxis_title=u_label, template="plotly_white", height=400,
            margin=dict(l=50,r=20,t=20,b=50))
        st.plotly_chart(fig_b1, use_container_width=True)
    with b2:
        st.subheader("Por uso")
        todos = {"Esclusajes":dem_escl,"Potable":alh_pot+gat_pot,"Generación":gen_tot,
                 "Fugas":alh_fug+gat_fug,"Vertidos":alh_vert+gat_ver,
                 "ZZ-Flush":dem_flush,"Evaporación":evap_tot}
        tf     = {k:v for k,v in todos.items() if v > 0.001}
        cols_t = [COL["esclusas"],COL["potable"],COL["generacion"],COL["fugas"],
                  COL["vertidos"],COL["flush"],COL["evap"]]
        fig_b2 = go.Figure(go.Pie(labels=list(tf.keys()),
            values=[v*u_cv for v in tf.values()],
            marker_colors=cols_t[:len(tf)], hole=0.45,
            textinfo="percent+label", textposition="outside"))
        fig_b2.update_layout(height=400, template="plotly_white",
            margin=dict(l=10,r=10,t=20,b=10), showlegend=False)
        st.plotly_chart(fig_b2, use_container_width=True)

    gauge_cols = st.columns(6)
    gauge_data = [
        ("Esclusajes", dem_escl,            COL["esclusas"]),
        ("Potable",    alh_pot+gat_pot,      COL["potable"]),
        ("Generación", gen_tot,              COL["generacion"]),
        ("Fugas",      alh_fug+gat_fug,      COL["fugas"]),
        ("Vertidos",   alh_vert+gat_ver+dem_flush, COL["vertidos"]),
        ("Evaporación",evap_tot,             COL["evap"]),
    ]
    for col_g, (nm, val, cl) in zip(gauge_cols, gauge_data):
        with col_g:
            pct = val/max(dem_total,.001)*100
            fig_gauge = go.Figure(go.Indicator(mode="gauge+number", value=pct,
                title={"text": nm, "font":{"size":11}},
                number={"suffix":"%","font":{"size":18}},
                gauge={"axis":{"range":[0,100]},"bar":{"color":cl}}))
            fig_gauge.update_layout(height=160, margin=dict(l=10,r=10,t=35,b=5))
            st.plotly_chart(fig_gauge, use_container_width=True)

    st.markdown("---")
    st.subheader("☀️ Evaporación aplicada en el balance")
    evc1, evc2, evc3 = st.columns(3)
    evc1.metric(
        "Gatún",
        f"{evap_gat:.3f} hm³/d",
        delta=f"{evap_gat_mm:.2f} mm/día · {area_gat:.1f} km²",
        delta_color="off"
    )
    evc2.metric(
        "Alhajuela",
        f"{evap_alh:.3f} hm³/d",
        delta=f"{evap_alh_mm:.2f} mm/día · {area_alh:.1f} km²",
        delta_color="off"
    )
    evc3.metric(
        "Total evaporación",
        f"{evap_tot:.3f} hm³/d",
        delta=f"{evap_tot/CFS2HM3:.1f} cfs · {evap_tot*HM3D2M3S:.2f} m³/s",
        delta_color="off"
    )
    st.caption(
        f"Fuente activa: **{evap_fuente_label}**. "
        "Cuando la fuente entrega lámina, el volumen se calcula con el área espejo activa; "
        "cuando entrega volumen V Evap 0.85, ese hm³/día entra directamente al balance y la lámina mostrada es equivalente."
    )

    st.markdown("---")
    st.subheader("🚢 Consumo unitario de esclusajes aplicado en el balance")
    
    # Coherente con la fuente seleccionada para el balance principal
    vp_mostrar_balance = vp_balance
    vn_mostrar_balance = vn_balance

    bk1, bk2 = st.columns(2)
    bk1.metric(
        "PNX Unitario",
        f"{vp_mostrar_balance:.4f} hm³/escl",
        delta=f"{vp_mostrar_balance/CFS2HM3:.1f} cfs · {vp_mostrar_balance*HM3D2M3S:.2f} m³/s equiv.",
        delta_color="off"
    )
    bk2.metric(
        "NPX Unitario",
        f"{vn_mostrar_balance:.4f} hm³/escl",
        delta=f"{vn_mostrar_balance/CFS2HM3:.1f} cfs · {vn_mostrar_balance*HM3D2M3S:.2f} m³/s equiv.",
        delta_color="off"
    )

    bd1, bd2 = st.columns(2)
    bd1.metric(
        "Diferencia NPX - PNX",
        f"{(vn_mostrar_balance - vp_mostrar_balance):.4f} hm³/escl",
        delta=f"{abs(vn_mostrar_balance - vp_mostrar_balance)/max(vp_mostrar_balance, 0.0001)*100:.1f}% de diferencia",
        delta_color="off"
    )
    bd2.metric(
        "Consumo diario de esclusajes",
        f"{dem_escl:.3f} hm³/d",
        delta=f"{dem_escl/CFS2HM3:.1f} cfs · últimos {_dias_lkh_seguros()} días LakeHouse base",
        delta_color="off"
    )
    
    st.caption(f"Fuente del balance: **{balance_escl_label}** · Fuente base sidebar: **{fuente_consumo_escl}** · PNX {vp_mostrar_balance:.4f} hm³/escl · NPX {vn_mostrar_balance:.4f} hm³/escl")
    
    st.markdown("---")

    st.subheader("Tabla completa (hm³/día · cfs · m³/s)")
    rows = []
    all_usos = {**{f"[ALH] {k}":v for k,v in alh_usos.items()},
                **{f"[GAT] {k}":v for k,v in gat_usos.items()}}
    for nm,(h,cf,_) in all_usos.items():
        if h > 0.0001:
            rows.append({"Uso":nm,"hm³/día":round(h,4),"cfs":round(cf,1),
                         "m³/s":round(cf*CFS2M3S,2),"%":round(h/max(dem_total,.001)*100,1)})
    rows.append({"Uso":"TOTAL","hm³/día":round(dem_total,4),"cfs":round(dem_total/CFS2HM3,1),
                 "m³/s":round(dem_total*HM3D2M3S,2),"%":100.0})
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


# ═══ TAB 1 — ALHAJUELA ═══
with tabs[1]:
    st.subheader("🏔️ Embalse Alhajuela"); st.metric("Total", f3u(alh_total))
    c1,c2 = st.columns(2)
    with c1:
        af = {k:v[0] for k,v in alh_usos.items() if v[0]>0.001}
        if af:
            fig_a1 = go.Figure(go.Pie(labels=list(af.keys()),
                values=[v*u_cv for v in af.values()],
                marker_colors=[alh_usos[k][2] for k in af], hole=0.45, textinfo="percent+label"))
            fig_a1.update_layout(height=400, template="plotly_white",
                margin=dict(l=10,r=10,t=20,b=10), showlegend=False)
            st.plotly_chart(fig_a1, use_container_width=True)
    with c2:
        fig_a2 = go.Figure()
        for nm,(h,cf,cl) in alh_usos.items():
            if h>0.001:
                fig_a2.add_trace(go.Bar(y=[nm],x=[h*u_cv],orientation="h",marker_color=cl,
                    text=[f"{h*u_cv:.3f}"],textposition="auto",showlegend=False))
        fig_a2.update_layout(xaxis_title=u_label,template="plotly_white",height=400,
            margin=dict(l=10,r=20,t=20,b=50))
        st.plotly_chart(fig_a2, use_container_width=True)
    st.dataframe(tbl(alh_usos,alh_total,"Alhajuela",dem_total), use_container_width=True, hide_index=True)


# ═══ TAB 2 — GATÚN ═══
with tabs[2]:
    st.subheader("🌊 Embalse Gatún"); st.metric("Total", f3u(gat_total))
    c1,c2 = st.columns(2)
    with c1:
        gf = {k:v[0] for k,v in gat_usos.items() if v[0]>0.001}
        fig_g1 = go.Figure(go.Pie(labels=list(gf.keys()),
            values=[v*u_cv for v in gf.values()],
            marker_colors=[gat_usos[k][2] for k in gf], hole=0.45, textinfo="percent+label"))
        fig_g1.update_layout(height=400,template="plotly_white",
            margin=dict(l=10,r=10,t=20,b=10),showlegend=False)
        st.plotly_chart(fig_g1, use_container_width=True)
    with c2:
        fig_g2 = go.Figure()
        for nm,(h,cf,cl) in gat_usos.items():
            if h>0.001:
                fig_g2.add_trace(go.Bar(y=[nm],x=[h*u_cv],orientation="h",marker_color=cl,
                    text=[f"{h*u_cv:.3f}"],textposition="auto",showlegend=False))
        fig_g2.update_layout(xaxis_title=u_label,template="plotly_white",height=400,
            margin=dict(l=10,r=20,t=20,b=50))
        st.plotly_chart(fig_g2, use_container_width=True)
    st.dataframe(tbl(gat_usos,gat_total,"Gatún",dem_total), use_container_width=True, hide_index=True)


# ═══ TAB 3 — ESCLUSAJES ═══
with tabs[3]:
    st.subheader("🚢 Dashboard de Esclusajes")

    # Tránsitos: se muestran por separado y el total se limita a 3 cifras significativas.
    tr1, tr2, tr3 = st.columns(3)
    tr1.metric("Total de tránsitos", f"{fmt_sig(n_t, 3)}/día")
    tr2.metric("Panamax (PNX)", f"{fmt_sig(n_pnx, 3)}/día")
    tr3.metric("NeoPanamax (NPX)", f"{fmt_sig(n_npx, 3)}/día")

    ek1, ek2, ek3 = st.columns(3)
    ek1.metric("Consumo total", f3u(dem_escl))
    ek2.metric("% de demanda", f"{dem_escl/max(dem_total,.001)*100:.1f}%")
    ek3.metric("Volumen promedio/tránsito", f"{dem_escl/max(n_t,1):.3f} hm³")

    # Consumo unitario - controlado desde sidebar
    st.markdown("---")
    st.markdown("##### 📊 Consumo unitario usado en el balance principal")
    
    # Coherente con la fuente seleccionada para el balance principal
    vp_mostrar = vp_balance
    vn_mostrar = vn_balance

    sk1, sk2, sk3 = st.columns(3)
    sk1.metric("PNX Unitario", f"{vp_mostrar:.4f} hm³/escl ({vp_mostrar/CFS2HM3:.1f} cfs)")
    sk2.metric("NPX Unitario", f"{vn_mostrar:.4f} hm³/escl ({vn_mostrar/CFS2HM3:.1f} cfs)")
    sk3.metric("Diferencia", f"{abs(vp_mostrar - vn_mostrar):.4f} hm³", 
               delta=f"{'NPX usa más' if vn_mostrar > vp_mostrar else 'PNX usa más'}")
    
    st.caption(f"Fuente del balance: **{balance_escl_label}** · Fuente base sidebar: **{fuente_consumo_escl}**")
    
    st.markdown("---")

    ec1,ec2 = st.columns(2)
    with ec1:
        fig_e1 = go.Figure(go.Bar(x=["Panamax","Neopanamax","Total"],
            y=[dem_pnx*u_cv,dem_npx*u_cv,dem_escl*u_cv],
            marker_color=[COL["pnx"],COL["npx"],COL["esclusas"]],
            text=[f"{dem_pnx*u_cv:.2f}",f"{dem_npx*u_cv:.2f}",f"{dem_escl*u_cv:.2f}"],
            textposition="auto"))
        fig_e1.update_layout(yaxis_title=u_label,template="plotly_white",height=380,
            margin=dict(l=50,r=20,t=20,b=50))
        st.plotly_chart(fig_e1, use_container_width=True)
    with ec2:
        fig_e2 = go.Figure(go.Pie(labels=["Panamax","Neopanamax"],values=[dem_pnx,dem_npx],
            marker_colors=[COL["pnx"],COL["npx"]],hole=0.45,
            textinfo="percent+label+value",
            texttemplate="%{label}<br>%{percent}<br>%{value:.2f} hm³/d"))
        fig_e2.update_layout(height=380,template="plotly_white",
            margin=dict(l=10,r=10,t=20,b=10),showlegend=False)
        st.plotly_chart(fig_e2, use_container_width=True)

    st.subheader("Detalle (3 unidades)")
    ed = []
    for tipo,n,v,th in [("Panamax",n_pnx,vp_balance,dem_pnx),("Neopanamax",n_npx,vn_balance,dem_npx)]:
        ed.append({"Tipo":tipo,"N/día":n,
            "hm³/escl":round(v,3),"cfs/escl":round(v/CFS2HM3,1),"m³/s/escl":round(v*HM3D2M3S,2),
            "hm³/día":round(th,2),"cfs":round(th/CFS2HM3,0),"m³/s":round(th*HM3D2M3S,1)})
    ed.append({"Tipo":"TOTAL","N/día":n_t,
        "hm³/escl":round(dem_escl/max(n_t,1),3),"cfs/escl":round(dem_escl/max(n_t,1)/CFS2HM3,1),
        "m³/s/escl":round(dem_escl/max(n_t,1)*HM3D2M3S,2),
        "hm³/día":round(dem_escl,2),"cfs":round(dem_escl/CFS2HM3,0),"m³/s":round(dem_escl*HM3D2M3S,1)})
    st.dataframe(pd.DataFrame(ed), use_container_width=True, hide_index=True)

    st.subheader("Proyección acumulada")
    pr1,pr2,pr3 = st.columns(3)
    pr1.metric("Diario",     f"{dem_escl:.2f} hm³ · {dem_escl/CFS2HM3:.0f} cfs")
    pr2.metric("Mensual (30d)", f"{dem_escl*30:.1f} hm³")
    pr3.metric("Anual (365d)",  f"{dem_escl*365:.0f} hm³")


# ═══ TAB 4 — GENERACIÓN ═══
with tabs[4]:
    st.subheader("⚡ Dashboard de Hidrogeneración")
    hk1,hk2,hk3,hk4 = st.columns(4)
    hk1.metric("Madden",    f"{gm_mw} MW")
    hk2.metric("Gatún",     f"{gg_mw} MW")
    hk3.metric("Total",     f"{gm_mw+gg_mw} MW")
    hk4.metric("Agua usada",f3u(gen_tot))

    hc1,hc2 = st.columns(2)
    with hc1:
        fig_h1 = go.Figure(go.Bar(x=["Madden","Gatún","Total"],
            y=[gen_alh*u_cv,gen_gat*u_cv,gen_tot*u_cv],
            marker_color=[COL["generacion"],COL["gatgen"],COL["total"]],
            text=[f"{gen_alh*u_cv:.2f}",f"{gen_gat*u_cv:.2f}",f"{gen_tot*u_cv:.2f}"],
            textposition="auto"))
        fig_h1.update_layout(yaxis_title=u_label,template="plotly_white",height=380,
            margin=dict(l=50,r=20,t=20,b=50))
        st.plotly_chart(fig_h1, use_container_width=True)
    with hc2:
        st.markdown(f"""
**Método activo Madden: {metodo_madden}** · Factor = **{mw_madden:.2f} cfs/MW**

| Central | cfs/MW | m³/s por MW | Método |
|---------|--------|-------------|--------|
| Madden | **{mw_madden:.2f}** | {mw_madden*CFS2M3S:.2f} | {metodo_madden} |
| Gatún  | **{mw_gatun:.2f}**  | {mw_gatun*CFS2M3S:.2f}  | Manual |

**Consumo actual:**

| Central | MW | cfs | m³/s | hm³/día |
|---------|-----|------|------|---------|
| Madden | {gm_mw} | {gm_mw*mw_madden:.1f} | {gm_mw*mw_madden*CFS2M3S:.1f} | {gen_alh:.3f} |
| Gatún  | {gg_mw} | {gg_mw*mw_gatun:.1f}  | {gg_mw*mw_gatun*CFS2M3S:.1f}  | {gen_gat:.3f} |
| **Total** | **{gm_mw+gg_mw}** | **{gm_mw*mw_madden+gg_mw*mw_gatun:.1f}** | **{(gm_mw*mw_madden+gg_mw*mw_gatun)*CFS2M3S:.1f}** | **{gen_tot:.3f}** |
        """)
    st.metric("% del sistema",f"{gen_tot/max(dem_total,.001)*100:.1f}%")

    # ── Comparación de métodos Madden ────────────────────────────────────
    st.markdown("---")
    st.subheader("📊 Comparación de métodos — Hidrogeneración Madden")
    _niv_ref_mad = nivel_alh_op
    _met_col = st.columns(4)
    _metodos_all = ["Manual", "Tabla de agua", "Modesto", "Bill Shaw"]
    _cfsmw_all   = [100.0,
                    madden_cfs_per_mw(_niv_ref_mad, "Tabla de agua"),
                    madden_cfs_per_mw(_niv_ref_mad, "Modesto"),
                    madden_cfs_per_mw(_niv_ref_mad, "Bill Shaw")]
    _gen_all_hm3 = [gm_mw * c * CFS2HM3 for c in _cfsmw_all]
    for _mc, _mt, _cf, _gh in zip(_met_col, _metodos_all, _cfsmw_all, _gen_all_hm3):
        _delta = f"{'▶ ACTIVO' if _mt==metodo_madden else ''}"
        _mc.metric(f"{'✅ ' if _mt==metodo_madden else ''}{_mt}",
                   f"{_cf:.1f} cfs/MW",
                   delta=f"{_gh:.3f} hm³/d | {_gh/CFS2HM3:.0f} cfs")

    _fig_met = go.Figure()
    _nv_range = np.arange(180, 261, 1, dtype=float)
    for _mt, _clr in [("Tabla de agua","#2980b9"),("Modesto","#27ae60"),("Bill Shaw","#e67e22")]:
        _cf_rng = [madden_cfs_per_mw(n, _mt) for n in _nv_range]
        _fig_met.add_trace(go.Scatter(x=_nv_range, y=_cf_rng,
            name=_mt, line=dict(color=_clr, width=2)))
    _fig_met.add_hline(y=100.0, line_dash="dot", line_color="#7f8c8d",
        annotation_text="Manual (100 cfs/MW fijo)")
    if metodo_madden != "Manual":
        _fig_met.add_vline(x=_niv_ref_mad, line_dash="dash", line_color="red",
            annotation_text=f"Nivel actual {_niv_ref_mad:.1f} ft")
    _fig_met.update_layout(
        xaxis_title="Nivel Alhajuela (ft)", yaxis_title="cfs/MW",
        template="plotly_white", height=380,
        hovermode="x unified", margin=dict(l=50,r=20,t=30,b=50),
        legend=dict(orientation="h", y=1.08))
    st.plotly_chart(_fig_met, use_container_width=True)

    # Tabla comparativa a niveles clave
    _nv_key = [215, 220, 225, 230, 235, 240, 245, 250, 252, 255, 260]
    _cmp_rows = []
    for _n in _nv_key:
        row = {"Nivel Alh (ft)": _n}
        for _mt, _col in [("Manual",None),("Tabla de agua","#2980b9"),
                          ("Modesto","#27ae60"),("Bill Shaw","#e67e22")]:
            _cf = 100.0 if _mt=="Manual" else madden_cfs_per_mw(float(_n), _mt)
            row[f"{_mt} (cfs/MW)"] = round(_cf, 1)
            row[f"{_mt} hm³/d @{gm_mw}MW"] = round(gm_mw * _cf * CFS2HM3, 3)
        _cmp_rows.append(row)
    st.dataframe(pd.DataFrame(_cmp_rows), use_container_width=True, hide_index=True)
    st.caption("📊 Fuente: Tablas_Hidrogeneracion_Madden_Alhajuela.xlsx · 66 niveles (190–255 ft) · datos oficiales ACP")
    if mw_madden!=100.00 and metodo_madden=="Manual":
        st.warning(f"⚠️ Factor Manual modificado a {mw_madden:.2f} cfs/MW (inicial: 100.0)")


# ═══ TAB 5 — AHORRO DE AGUA ═══
with tabs[5]:
    st.subheader("💾 Dashboard de Ahorro de Agua en Esclusajes")
    st.markdown(
        "Modelo físico basado en **ConsumodeAguaEsclusas.xlsb** · "
        f"Nivel de referencia: **{nivel_modelo_ft:.2f} ft** ({nivel_modelo_ft*0.3048:.3f} m)")

    # ── KPIs ──────────────────────────────────────────────────────────────────
    ah1,ah2,ah3,ah4,ah5,ah6 = st.columns(6)
    ah1.metric("Ahorro total", f"{ahorro_total_esc:.3f} hm³/d",
               delta=f"{ahorro_total_esc*365:.0f} hm³/año")
    ah2.metric("Tinas NPX (CC+AC)",
               f"{(ahorro_tinas_cc+ahorro_tinas_ac):.3f} hm³/d",
               delta=f"{(ahorro_tinas_cc+ahorro_tinas_ac)/CFS2HM3:.0f} cfs")
    ah3.metric("Turn Around NPX",
               f"{ahorro_turnaround_npx_modelo:.3f} hm³/d",
               delta=f"{ahorro_turnaround_npx_modelo/CFS2HM3:.0f} cfs")
    ah4.metric("Cámaras Cortas PNX",
               f"{ahorro_cam_corta:.3f} hm³/d",
               delta=f"{ahorro_cam_corta/CFS2HM3:.0f} cfs")
    ah5.metric("CrossFilling PNX",
               f"{ahorro_xfill:.3f} hm³/d",
               delta=f"{ahorro_xfill/CFS2HM3:.0f} cfs")
    _equiv_transitos = ahorro_total_esc / max(_vn_fis, 0.001)
    ah6.metric("Tránsitos equiv. ahorrados", f"{_equiv_transitos:.1f}/d")

    st.markdown("---")

    # ── Gráfico comparativo: Base vs Efectivo ──────────────────────────────────
    col_ah1, col_ah2 = st.columns(2)

    with col_ah1:
        st.markdown("#### Volumen por tránsito según nivel Gatún")

        # Esta gráfica debe responder al nivel operativo actual.  En lugar de
        # mostrar cuatro barras aisladas, se dibuja la curva por nivel y se
        # marca el nivel actual; así se ve claramente si el cálculo cambia al
        # mover el nivel y si las medidas de ahorro están activas.
        _nv_cmp_min = max(75.0, float(nivel_modelo_ft) - 7.0)
        _nv_cmp_max = min(90.0, float(nivel_modelo_ft) + 7.0)
        if _nv_cmp_max <= _nv_cmp_min + 1.0:
            _nv_cmp_min, _nv_cmp_max = 75.0, 90.0
        _nv_cmp = np.linspace(_nv_cmp_min, _nv_cmp_max, 140)

        _npx_base_curve = np.array([_npx_vol_base(n) for n in _nv_cmp], dtype=float)
        _pnx_base_curve = np.array([_pnx_vol_base(n) for n in _nv_cmp], dtype=float)
        _npx_eff_curve = np.maximum(_npx_base_curve * frac_ahorro_npx, 0.001)
        _sav_cc_curve = np.array([_pnx_ahorro_cc_per_transit(n) for n in _nv_cmp], dtype=float)
        _pnx_eff_curve = np.maximum(
            _pnx_base_curve - _sav_cc_curve * pct_cam_corta / 100.0 - ahorro_xfill_tr,
            0.001,
        )

        fig_cmp = go.Figure()
        fig_cmp.add_trace(go.Scatter(
            x=_nv_cmp, y=_npx_base_curve, mode="lines", name="NPX base",
            line=dict(color=COL["npx"], width=3),
            hovertemplate="Nivel %{x:.2f} ft<br>NPX base %{y:.4f} hm³/tránsito<extra></extra>",
        ))
        fig_cmp.add_trace(go.Scatter(
            x=_nv_cmp, y=_npx_eff_curve, mode="lines", name="NPX efectivo",
            line=dict(color=COL["flush"], width=3, dash="dash"),
            hovertemplate="Nivel %{x:.2f} ft<br>NPX efectivo %{y:.4f} hm³/tránsito<extra></extra>",
        ))
        fig_cmp.add_trace(go.Scatter(
            x=_nv_cmp, y=_pnx_base_curve, mode="lines", name="PNX base",
            line=dict(color=COL["pnx"], width=3),
            hovertemplate="Nivel %{x:.2f} ft<br>PNX base %{y:.4f} hm³/tránsito<extra></extra>",
        ))
        fig_cmp.add_trace(go.Scatter(
            x=_nv_cmp, y=_pnx_eff_curve, mode="lines", name="PNX efectivo",
            line=dict(color=COL["esclusas"], width=3, dash="dash"),
            hovertemplate="Nivel %{x:.2f} ft<br>PNX efectivo %{y:.4f} hm³/tránsito<extra></extra>",
        ))

        _puntos_actuales = [
            ("NPX base actual", _vn_fis, COL["npx"]),
            ("NPX efectivo actual", vn_efectivo, COL["flush"]),
            ("PNX base actual", _vp_fis, COL["pnx"]),
            ("PNX efectivo actual", vp_efectivo, COL["esclusas"]),
        ]
        for _nom, _val, _clr in _puntos_actuales:
            fig_cmp.add_trace(go.Scatter(
                x=[nivel_modelo_ft], y=[_val], mode="markers",
                name=_nom, marker=dict(color=_clr, size=9, line=dict(width=1, color="white")),
                showlegend=False,
                hovertemplate=f"{_nom}<br>Nivel {nivel_modelo_ft:.2f} ft<br>{_val:.4f} hm³/tránsito<extra></extra>",
            ))

        fig_cmp.add_vline(
            x=nivel_modelo_ft, line_dash="dot", line_color=COL["total"],
            annotation_text=f"Nivel actual {nivel_modelo_ft:.2f} ft",
            annotation_position="top right",
        )
        fig_cmp.update_layout(
            xaxis_title="Nivel Gatún (ft)", yaxis_title="hm³/tránsito",
            template="plotly_white", height=380, hovermode="x unified",
            margin=dict(l=50, r=20, t=20, b=55),
            legend=dict(orientation="h", y=1.10, x=0),
        )
        st.plotly_chart(fig_cmp, use_container_width=True)

        _ah_npxt = max(_vn_fis - vn_efectivo, 0.0)
        _ah_pnxt = max(_vp_fis - vp_efectivo, 0.0)
        st.caption(
            f"Actual @ {nivel_modelo_ft:.2f} ft · NPX: base {_vn_fis:.4f}, efectivo {vn_efectivo:.4f}, "
            f"ahorro {_ah_npxt:.4f} hm³/tránsito · PNX: base {_vp_fis:.4f}, efectivo {vp_efectivo:.4f}, "
            f"ahorro {_ah_pnxt:.4f} hm³/tránsito."
        )
        if (pct_tinas_cc == 0 and pct_tinas_ac == 0 and pct_cam_corta == 0 and pct_crossfill == 0
                and not usar_turnaround_npx):
            st.info("Base y efectivo coinciden porque no hay tinas, cámaras cortas, crossfilling ni Turn Around activos.")

    with col_ah2:
        st.markdown("#### Ahorro diario por mecanismo (hm³/d)")
        mec_lbl = ["Tinas Cocolí\n(NPX)", "Tinas A.Clara\n(NPX)",
                   "Turn Around\n(NPX)", "Cámaras Cortas\n(PNX)", "CrossFilling\n(PNX)"]
        mec_val = [ahorro_tinas_cc, ahorro_tinas_ac, ahorro_turnaround_npx_modelo, ahorro_cam_corta, ahorro_xfill]
        mec_clr = [COL["npx"], COL["flush"], COL["gatun"], COL["pnx"], COL["esclusas"]]
        fig_mec = go.Figure(go.Bar(
            x=mec_lbl, y=mec_val, marker_color=mec_clr,
            text=[f"{v:.4f}" for v in mec_val], textposition="auto"))
        fig_mec.add_hline(y=ahorro_total_esc, line_dash="dash",
                          line_color=COL["total"],
                          annotation_text=f"Total: {ahorro_total_esc:.3f} hm³/d")
        fig_mec.update_layout(
            yaxis_title="hm³/día", template="plotly_white",
            height=380, margin=dict(l=50,r=20,t=30,b=60))
        st.plotly_chart(fig_mec, use_container_width=True)

    st.markdown("---")

    # ── Sensibilidad al nivel del lago ─────────────────────────────────────────
    st.markdown("#### Sensibilidad del ahorro al nivel del lago Gatún")
    _nv_sens = np.linspace(75, 89, 80)
    _aho_cc  = [n_npx * 0.5 * (AC_NPX * max(n*0.3048 - EQ_CC_m, 0) * 1e-6)
                * FRAC_TINAS * pct_tinas_cc/100 for n in _nv_sens]
    _aho_ac  = [n_npx * 0.5 * (AC_NPX * max(n*0.3048 - EQ_AC_m, 0) * 1e-6)
                * FRAC_TINAS * pct_tinas_ac/100 for n in _nv_sens]
    _aho_c   = [n_pnx * _pnx_ahorro_cc_per_transit(n) * pct_cam_corta/100  for n in _nv_sens]
    _aho_xf  = [n_pnx * (pct_crossfill/100) * AC_PNX_REG * EQ_PM_ft * 0.3048 * 1e-6 * 0.5
                for n in _nv_sens]
    _aho_ta  = [n_turnaround_npx * (2.0 * _npx_vol_base(n)) * TURN_NPX_SAVING_PCT if usar_turnaround_npx else 0.0
                for n in _nv_sens]
    _aho_tot = [a+b+c+d+e for a,b,c,d,e in zip(_aho_cc, _aho_ac, _aho_c, _aho_xf, _aho_ta)]

    fig_sen = go.Figure()
    fig_sen.add_trace(go.Scatter(x=_nv_sens, y=_aho_cc, name="Tinas Cocolí",
        stackgroup="one", line=dict(color=COL["npx"]),   fillcolor="rgba(22,160,133,0.55)"))
    fig_sen.add_trace(go.Scatter(x=_nv_sens, y=_aho_ac, name="Tinas A.Clara",
        stackgroup="one", line=dict(color=COL["flush"]),  fillcolor="rgba(26,188,156,0.55)"))
    fig_sen.add_trace(go.Scatter(x=_nv_sens, y=_aho_ta, name="Turn Around NPX",
        stackgroup="one", line=dict(color=COL["gatun"]),  fillcolor="rgba(26,82,118,0.45)"))
    fig_sen.add_trace(go.Scatter(x=_nv_sens, y=_aho_c,  name="Cámaras Cortas",
        stackgroup="one", line=dict(color=COL["pnx"]),    fillcolor="rgba(44,62,80,0.55)"))
    fig_sen.add_trace(go.Scatter(x=_nv_sens, y=_aho_xf, name="CrossFilling",
        stackgroup="one", line=dict(color=COL["esclusas"]),fillcolor="rgba(41,128,185,0.55)"))
    fig_sen.add_vline(x=nivel_modelo_ft, line_dash="dot", line_color="red",
        annotation_text=f"Nivel actual\n{nivel_modelo_ft:.1f} ft", annotation_position="top right")
    fig_sen.update_layout(
        xaxis_title="Nivel lago Gatún (ft)", yaxis_title="Ahorro (hm³/d)",
        template="plotly_white", height=380, hovermode="x unified",
        margin=dict(l=50,r=20,t=20,b=50))
    st.plotly_chart(fig_sen, use_container_width=True)

    # ── Consumo diario total: Dashboard vs Modelo físico ──────────────────────
    st.markdown("---")
    st.markdown("#### Comparación consumo diario de esclusajes")
    cmp_cols = st.columns(4)
    cmp_cols[0].metric("Modelo físico base\n(nivel actual)",
                       f"{dem_escl_modelo:.3f} hm³/d",
                       delta=f"{dem_escl_modelo/CFS2HM3:.0f} cfs")
    cmp_cols[1].metric("Modelo físico + ahorro",
                       f"{dem_escl_efectivo:.3f} hm³/d",
                       delta=f"−{(dem_escl_modelo-dem_escl_efectivo):.3f} hm³/d vs base")
    cmp_cols[2].metric(f"Balance seleccionado\n({balance_escl_label})",
                       f"{dem_escl:.3f} hm³/d",
                       delta=f"{dem_escl/CFS2HM3:.0f} cfs")
    _dif = dem_escl - dem_escl_efectivo
    cmp_cols[3].metric("Potencial ahorro adicional",
                       f"{max(_dif,0):.3f} hm³/d",
                       delta=f"{max(_dif,0)*365:.0f} hm³/año")

    # ── Tabla de parámetros del modelo ────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### Parámetros del modelo físico (ConsumodeAguaEsclusas.xlsb)")
    tbl_fis = pd.DataFrame([
        {"Parámetro":"Nivel lago Gatún","Valor":f"{nivel_modelo_ft:.2f} ft = {nivel_modelo_ft*0.3048:.3f} m","Fuente":"Input"},
        {"Parámetro":"Área cámara NPX","Valor":f"{AC_NPX:,.0f} m²","Fuente":"Hoja NeoPanamax"},
        {"Parámetro":"Nivel equiv. Cocolí (NPX)","Valor":f"{EQ_CC_m:.3f} m","Fuente":"Hoja NeoPanamax"},
        {"Parámetro":"Nivel equiv. Agua Clara (NPX)","Valor":f"{EQ_AC_m:.3f} m","Fuente":"Hoja NeoPanamax"},
        {"Parámetro":"Ahorro con tinas (frac/tránsito)","Valor":f"{FRAC_TINAS*100:.0f}%","Fuente":"Hoja NeoPanamax"},
        {"Parámetro":"Vol/tránsito NPX base","Valor":f"{_vn_fis:.4f} hm³","Fuente":"Constantes NoSaving por nivel"},
        {"Parámetro":"Vol/tránsito NPX con tinas","Valor":f"{vn_efectivo:.4f} hm³","Fuente":"Cálculo físico"},
        {"Parámetro":"Turn Around NPX/día","Valor":f"{n_turnaround_npx:.1f}","Fuente":"Sidebar"},
        {"Parámetro":"Ahorro Turn Around NPX por evento","Valor":f"{turnaround_npx_ahorro_tr_modelo:.5f} hm³","Fuente":"Workbook / cálculo"},
        {"Parámetro":"Ahorro Turn Around NPX total","Valor":f"{ahorro_turnaround_npx_modelo:.4f} hm³/d","Fuente":"Cálculo físico"},
        {"Parámetro":"Área cámara PNX Regular","Valor":f"{AC_PNX_REG:,.1f} m²","Fuente":"Hoja Panamax"},
        {"Parámetro":"Área cámara PNX Corta","Valor":f"{AC_PNX_COR:,.1f} m²","Fuente":"Hoja Panamax"},
        {"Parámetro":"Nivel equiv. PedroMiguel","Valor":f"{EQ_PM_ft:.3f} ft","Fuente":"Hoja Panamax"},
        {"Parámetro":"Nivel equiv. Gatún (PNX)","Valor":f"{EQ_GA_ft:.3f} ft","Fuente":"Hoja Panamax"},
        {"Parámetro":"Vol/tránsito PNX base","Valor":f"{_vp_fis:.4f} hm³","Fuente":"Constantes NoSaving por nivel"},
        {"Parámetro":"Ahorro/tránsito Cámara Corta","Valor":f"{_sav_cc_tr:.5f} hm³","Fuente":"Cálculo físico"},
        {"Parámetro":"% Tinas Cocolí activas","Valor":f"{pct_tinas_cc}%","Fuente":"Sidebar"},
        {"Parámetro":"% Tinas Agua Clara activas","Valor":f"{pct_tinas_ac}%","Fuente":"Sidebar"},
        {"Parámetro":"% Cámaras Cortas activas","Valor":f"{pct_cam_corta}%","Fuente":"Sidebar"},
        {"Parámetro":"% CrossFilling activo","Valor":f"{pct_crossfill}%","Fuente":"Sidebar"},
    ])
    st.dataframe(tbl_fis, use_container_width=True, hide_index=True)


# ═══ TAB 6 — ÁREA ESPEJO ═══
with tabs[6]:
    st.subheader("📐 Área Espejo · Evaporación por Nivel")
    st.caption(
        f"Fuente activa: **{evap_fuente_label}**. Los niveles se ajustan desde **📍 Niveles Operativos**. "
        "Para fuentes en mm: Vol aplicado al balance (hm³/d) = Lámina (mm/d) × Área (km²) × 10⁻³ × 0.85. "
        "En modo Aquarius volumen, las láminas mostradas son equivalentes al volumen directo seleccionado."
    )

    # ── Métricas principales ──────────────────────────────────────────────────
    _nv_g   = nivel_gat_op
    _nv_a   = nivel_alh_op
    _ar_g_d = area_desde_nivel_gat(_nv_g, daily=True)
    _ar_g_s = area_desde_nivel_gat(_nv_g, daily=False)
    _ar_a_d = area_desde_nivel_alh(_nv_a, daily=True)
    _ar_a_s = area_desde_nivel_alh(_nv_a, daily=False)
    # Misma fórmula usada por el balance operativo.
    # Antes esta vista mostraba el volumen bruto sin 0.85, por eso no coincidía con
    # "Resultado aplicado al balance". Aquí se muestra el valor aplicado.
    _ev_g_d_bruto = evap_gat_mm * _ar_g_d * 1e-3
    _ev_g_s_bruto = evap_gat_mm * _ar_g_s * 1e-3
    _ev_a_d_bruto = evap_alh_mm * _ar_a_d * 1e-3
    _ev_a_s_bruto = evap_alh_mm * _ar_a_s * 1e-3
    _ev_g_d = _ev_g_d_bruto * EVAP_COEF
    _ev_g_s = _ev_g_s_bruto * EVAP_COEF
    _ev_a_d = _ev_a_d_bruto * EVAP_COEF
    _ev_a_s = _ev_a_s_bruto * EVAP_COEF

    _kpi = st.columns(6)
    _kpi[0].metric("Nivel Gatún",    f"{_nv_g:.2f} ft",
                   delta=f"{_nv_g*0.3048:.3f} m")
    _kpi[1].metric("Área Gatún Daily",   f"{_ar_g_d:.2f} km²",
                   delta=f"{_ar_g_d-_ar_g_s:+.2f} vs Estándar")
    _kpi[2].metric("Evap Gatún Daily aplicada",   f"{_ev_g_d:.4f} hm³/d",
                   delta=f"{(_ev_g_d-_ev_g_s)*1_000_000:+,.0f} m³/d vs Estándar")
    _kpi[3].metric("Nivel Alhajuela", f"{_nv_a:.2f} ft",
                   delta=f"{_nv_a*0.3048:.3f} m")
    _kpi[4].metric("Área Alhajuela Daily",  f"{_ar_a_d:.4f} km²",
                   delta=f"{_ar_a_d-_ar_a_s:+.4f} vs Estándar")
    _kpi[5].metric("Evap Alhajuela Daily aplicada",  f"{_ev_a_d:.5f} hm³/d",
                   delta=f"{(_ev_a_d-_ev_a_s)*1_000_000:+,.0f} m³/d vs Estándar")

    st.markdown("---")

    # ── Curvas hipsométricas (solo visualización) ─────────────────────────────
    _ae1, _ae2 = st.columns(2)

    with _ae1:
        st.markdown("#### 🌊 Lago Gatún")
        _gat_min_ft = float(_NV_GAT_DAILY[0])  if _use_daily_gat else 55.0
        _gat_max_ft = float(_NV_GAT_DAILY[-1]) if _use_daily_gat else 89.0
        st.caption(
            f"Curva activa: **{'Daily' if _use_daily_gat else 'Estándar'}** · "
            f"{len(_NV_GAT_DAILY) if _use_daily_gat else len(_NV_GAT):,} pts · "
            f"rango {_gat_min_ft:.2f}–{_gat_max_ft:.2f} ft")
        _ar_g_act = area_desde_nivel_gat(_nv_g, daily=_use_daily_gat)
        _ref_nfs_g = area_desde_nivel_gat(87.0, daily=_use_daily_gat)
        st.metric("Área espejo Gatún", f"{_ar_g_act:.4f} km²",
                  delta=f"{_ar_g_act-_ref_nfs_g:+.4f} km² vs NFS 87 ft")
        _ev_g_bruto = evap_gat_mm * _ar_g_act * 1e-3
        _ev_g_act = _ev_g_bruto * EVAP_COEF
        _ev_g_cfs = _ev_g_act / CFS2HM3
        _ev_g_m3s = _ev_g_act * HM3D2M3S
        st.info(
            f"**Lámina:** {evap_gat_mm} mm/d  ·  "
            f"**Área:** {_ar_g_act:.4f} km²  ·  "
            f"**Vol aplicado al balance:** {_ev_g_act:.5f} hm³/d  ({_ev_g_cfs:.1f} p³/s  ·  {_ev_g_m3s:.2f} m³/s)  ·  "
            f"coef. {EVAP_COEF:.2f}")

        _nv_rng_g  = _NV_GAT_DAILY if _use_daily_gat else np.linspace(_gat_min_ft, _gat_max_ft, 200)
        _ar_rng_g  = _AR_GAT_DAILY if _use_daily_gat else [area_desde_nivel_gat(n) for n in _nv_rng_g]
        _fig_cg = go.Figure()
        _fig_cg.add_trace(go.Scatter(x=_nv_rng_g, y=_ar_rng_g, mode="lines",
            line=dict(color=COL["gatun"], width=2),
            name=f"{'Daily '+str(len(_NV_GAT_DAILY))+' pts' if _use_daily_gat else 'Estándar'}"))
        _fig_cg.add_trace(go.Scatter(x=[_nv_g], y=[_ar_g_act], mode="markers",
            marker=dict(color="red", size=14, symbol="star"),
            name=f"Nivel actual: {_nv_g:.2f} ft → {_ar_g_act:.4f} km²"))
        _fig_cg.update_layout(
            xaxis_title="Nivel (ft)", yaxis_title="Área (km²)",
            template="plotly_white", height=340, margin=dict(l=50,r=20,t=20,b=50))
        st.plotly_chart(_fig_cg, use_container_width=True)

    with _ae2:
        st.markdown("#### 🏔️ Lago Alhajuela")
        _alh_min_ft = float(_NV_ALH_DAILY[0])  if _use_daily_alh else 180.0
        _alh_max_ft = float(_NV_ALH_DAILY[-1]) if _use_daily_alh else 255.0
        st.caption(
            f"Curva activa: **{'Daily' if _use_daily_alh else 'Estándar'}** · "
            f"{len(_NV_ALH_DAILY) if _use_daily_alh else len(_NV_ALH):,} pts · "
            f"rango {_alh_min_ft:.2f}–{_alh_max_ft:.2f} ft")
        _ar_a_act = area_desde_nivel_alh(_nv_a, daily=_use_daily_alh)
        _ref_nfs_a = area_desde_nivel_alh(252.0, daily=_use_daily_alh)
        st.metric("Área espejo Alhajuela", f"{_ar_a_act:.4f} km²",
                  delta=f"{_ar_a_act-_ref_nfs_a:+.4f} km² vs NFS 252 ft")
        _ev_a_bruto = evap_alh_mm * _ar_a_act * 1e-3
        _ev_a_act = _ev_a_bruto * EVAP_COEF
        _ev_a_cfs = _ev_a_act / CFS2HM3
        _ev_a_m3s = _ev_a_act * HM3D2M3S
        st.info(
            f"**Lámina:** {evap_alh_mm} mm/d  ·  "
            f"**Área:** {_ar_a_act:.4f} km²  ·  "
            f"**Vol aplicado al balance:** {_ev_a_act:.5f} hm³/d  ({_ev_a_cfs:.1f} p³/s  ·  {_ev_a_m3s:.2f} m³/s)  ·  "
            f"coef. {EVAP_COEF:.2f}")

        _nv_rng_a  = _NV_ALH_DAILY if _use_daily_alh else np.linspace(_alh_min_ft, _alh_max_ft, 200)
        _ar_rng_a  = _AR_ALH_DAILY if _use_daily_alh else [area_desde_nivel_alh(n) for n in _nv_rng_a]
        _fig_ca = go.Figure()
        _fig_ca.add_trace(go.Scatter(x=_nv_rng_a, y=_ar_rng_a, mode="lines",
            line=dict(color=COL["alhajuela"], width=2),
            name=f"{'Daily '+str(len(_NV_ALH_DAILY))+' pts' if _use_daily_alh else 'Estándar'}"))
        _fig_ca.add_trace(go.Scatter(x=[_nv_a], y=[_ar_a_act], mode="markers",
            marker=dict(color="red", size=14, symbol="star"),
            name=f"Nivel actual: {_nv_a:.2f} ft → {_ar_a_act:.4f} km²"))
        _fig_ca.update_layout(
            xaxis_title="Nivel (ft)", yaxis_title="Área (km²)",
            template="plotly_white", height=340, margin=dict(l=50,r=20,t=20,b=50))
        st.plotly_chart(_fig_ca, use_container_width=True)

    st.markdown("---")

    # ── Tabla Estándar vs Daily ───────────────────────────────────────────────
    st.subheader("🔄 Comparación Estándar vs Daily")
    _tc1, _tc2 = st.columns(2)

    with _tc1:
        st.markdown("**🌊 Gatún** — Área y Evaporación por nivel")
        _comp_g = []
        for _n in np.round(np.arange(max(77.0,_nv_g-5), min(93.0,_nv_g+5.1), 0.5), 2):
            _as = area_desde_nivel_gat(_n, daily=False)
            _ad = area_desde_nivel_gat(_n, daily=True)
            _es = round(evap_gat_mm * _as * 1e-3 * EVAP_COEF, 5)
            _ed = round(evap_gat_mm * _ad * 1e-3 * EVAP_COEF, 5)
            _comp_g.append({
                "Nivel (ft)": round(_n,2),
                "Área Std (km²)": round(_as,4),
                "Área Daily (km²)": round(_ad,4),
                "ΔÁrea": round(_ad-_as,4),
                f"Evap aplicada Std (hm³/d)": _es,
                f"Evap aplicada Daily (hm³/d)": _ed,
                "ΔEvap (hm³/d)": round(_ed-_es,5),
            })
        _df_cg = pd.DataFrame(_comp_g)
        st.dataframe(_df_cg, use_container_width=True, hide_index=True, height=340)

    with _tc2:
        st.markdown("**🏔️ Alhajuela** — Área y Evaporación por nivel")
        _comp_a = []
        for _n in np.round(np.arange(max(160.0,_nv_a-10), min(260.0,_nv_a+10.1), 1.0), 2):
            _as = area_desde_nivel_alh(_n, daily=False)
            _ad = area_desde_nivel_alh(_n, daily=True)
            _es = round(evap_alh_mm * _as * 1e-3 * EVAP_COEF, 5)
            _ed = round(evap_alh_mm * _ad * 1e-3 * EVAP_COEF, 5)
            _comp_a.append({
                "Nivel (ft)": round(_n,2),
                "Área Std (km²)": round(_as,4),
                "Área Daily (km²)": round(_ad,4),
                "ΔÁrea": round(_ad-_as,4),
                f"Evap aplicada Std (hm³/d)": _es,
                f"Evap aplicada Daily (hm³/d)": _ed,
                "ΔEvap (hm³/d)": round(_ed-_es,5),
            })
        _df_ca = pd.DataFrame(_comp_a)
        st.dataframe(_df_ca, use_container_width=True, hide_index=True, height=340)

    st.markdown("---")

    # ── Auditoría vs Referencia ACP ───────────────────────────────────────────
    st.subheader("🔍 Auditoría — Verificación vs referencia ACP")
    st.caption(
        "Referencia ACP: Vol_Gatún = Lámina_CZL(mm) × Área_Gatún(km²) × 10⁻³  |  "
        "Vol_Alhajuela = Lámina_PMG(mm) × Área_Alhajuela(km²) × 10⁻³.  "
        "El balance operativo del app aplica adicionalmente el coeficiente 0.85.")

    _AUD_DATA = [
        (3.390,3.350,441.29,51.76,1.26,0.15),(4.571,2.740,441.29,52.54,1.37,0.12),
        (3.281,2.070,441.29,52.56,1.00,0.09),(4.418,2.570,441.29,52.27,1.31,0.11),
        (5.115,4.420,441.29,52.03,1.79,0.20),(5.540,5.040,441.29,51.84,1.98,0.22),
        (5.705,4.380,441.29,51.75,1.89,0.19),(5.266,5.180,441.29,51.61,1.96,0.23),
        (4.726,4.330,441.29,51.48,1.70,0.19),(4.521,4.570,441.29,51.39,1.70,0.20),
        (3.738,3.780,441.29,51.56,1.41,0.17),(4.089,4.920,441.29,51.57,1.69,0.22),
        (1.422,2.220,441.29,52.29,0.68,0.10),(4.740,4.790,441.29,52.20,1.79,0.21),
        (1.500,0.960,441.29,52.36,0.46,0.04),(2.845,1.880,441.29,52.28,0.89,0.08),
        (4.751,2.820,442.22,52.31,1.42,0.13),(4.775,4.630,441.29,52.24,1.76,0.21),
        (5.258,2.680,441.29,52.05,1.49,0.12),(4.844,4.060,441.29,51.93,1.67,0.18),
        (5.110,5.610,441.29,51.83,2.01,0.25),(5.408,4.760,441.29,51.74,1.91,0.21),
        (4.354,5.100,441.29,51.66,1.77,0.22),(5.193,5.240,441.29,51.58,1.96,0.23),
        (5.884,5.300,441.29,51.51,2.10,0.23),(4.036,3.870,441.29,51.46,1.48,0.17),
        (5.131,5.500,441.29,51.47,1.99,0.24),(4.840,5.750,441.29,51.55,1.99,0.25),
        (5.871,6.660,441.29,51.50,2.35,0.29),(4.800,4.740,441.29,51.44,1.79,0.21),
        (4.694,3.660,441.29,51.40,1.57,0.16),(4.418,4.090,441.29,51.31,1.60,0.18),
        (4.570,4.650,441.29,51.59,1.73,0.20),(5.530,5.810,441.29,51.68,2.13,0.26),
        (4.478,2.510,441.29,51.65,1.31,0.11),(4.231,3.090,441.29,51.61,1.37,0.14),
        (2.500,1.660,441.29,51.75,0.78,0.07),(3.340,2.340,441.29,52.15,1.07,0.10),
        (4.795,4.240,441.29,52.23,1.69,0.19),(4.420,3.180,441.29,52.12,1.43,0.14),
        (5.100,4.470,441.29,51.98,1.79,0.20),(5.186,5.210,441.29,51.85,1.95,0.23),
        (5.720,5.130,441.29,51.75,2.03,0.23),(5.492,5.700,441.29,51.66,2.10,0.25),
        (5.533,5.270,441.29,51.61,2.03,0.23),(3.910,4.730,441.29,51.56,1.62,0.21),
        (5.254,5.710,441.29,51.52,2.06,0.25),(4.505,3.910,441.29,51.93,1.58,0.17),
        (5.038,5.490,441.29,52.06,1.97,0.24),(2.664,2.060,441.29,52.19,0.89,0.09),
        (2.049,2.430,441.29,52.32,0.84,0.11),(3.229,2.670,441.29,52.17,1.11,0.12),
        (4.573,4.970,441.29,52.05,1.79,0.22),(4.083,2.320,441.29,52.23,1.20,0.10),
        (5.513,4.560,441.29,52.27,1.89,0.20),(5.843,5.450,441.29,52.14,2.12,0.24),
        (5.024,3.860,441.29,52.02,1.67,0.17),(4.505,3.560,441.29,51.91,1.51,0.16),
        (5.490,4.500,442.22,51.80,1.88,0.20),(5.215,5.710,441.29,51.70,2.05,0.25),
        (5.740,5.620,441.29,51.61,2.13,0.25),(5.735,6.090,441.29,51.53,2.22,0.27),
        (5.855,5.450,441.29,51.46,2.12,0.24),(3.448,3.460,441.29,51.40,1.30,0.15),
        (5.313,5.470,441.29,51.35,2.02,0.24),(5.050,5.200,441.29,51.30,1.92,0.23),
        (5.100,5.990,442.22,51.26,2.08,0.26),(5.270,7.850,441.29,51.20,2.46,0.34),
        (5.309,8.490,441.29,51.12,2.59,0.37),(5.866,4.520,441.29,51.05,1.95,0.20),
        (5.810,4.650,441.29,50.97,1.96,0.20),(6.284,5.080,441.29,50.88,2.13,0.22),
        (6.892,5.300,441.29,50.79,2.29,0.23),(7.500,5.750,438.50,50.70,2.47,0.25),
        (5.866,6.830,438.50,50.60,2.37,0.29),(7.569,6.100,438.50,50.50,2.55,0.26),
        (6.020,5.620,437.57,50.39,2.16,0.24),(5.866,5.630,437.57,50.28,2.14,0.24),
        (3.400,2.010,437.57,50.20,1.01,0.09),(4.900,4.190,437.57,50.13,1.69,0.18),
        (5.500,4.400,437.57,50.03,1.84,0.19),(6.350,5.860,437.57,49.92,2.27,0.25),
        (7.137,4.200,436.64,49.80,2.10,0.18),(7.100,5.350,436.64,49.68,2.31,0.23),
        (6.680,5.780,436.64,49.56,2.31,0.24),(6.690,5.090,436.64,49.43,2.19,0.21),
        (3.000,5.590,436.64,49.31,1.59,0.23),(6.850,6.500,435.72,49.19,2.47,0.27),
        (6.477,5.440,435.72,49.08,2.21,0.23),(5.372,2.520,435.72,48.92,1.46,0.10),
        (5.013,2.590,434.79,48.82,1.40,0.11),(1.360,2.420,434.79,48.72,0.70,0.10),
        (5.562,4.890,434.79,48.57,1.93,0.20),(5.411,4.690,433.86,48.47,1.86,0.19),
        (4.873,4.740,433.86,48.35,1.77,0.19),(4.521,4.040,432.93,48.20,1.58,0.17),
        (5.500,4.130,432.93,48.09,1.77,0.17),(4.594,4.140,432.93,47.95,1.61,0.17),
        (6.800,5.980,432.93,47.82,2.35,0.24),(7.033,3.760,432.00,47.68,1.98,0.15),
        (3.000,2.910,432.00,47.62,1.09,0.12),(5.800,3.640,432.00,47.54,1.73,0.15),
        (6.500,5.250,431.07,47.41,2.15,0.21),(7.874,7.200,431.07,47.25,2.76,0.29),
        (7.975,8.000,431.07,47.09,2.93,0.32),(6.452,4.490,430.14,46.91,2.00,0.18),
        (7.082,4.250,430.14,46.74,2.07,0.17),(5.553,4.920,430.14,46.55,1.91,0.19),
        (6.263,2.520,429.21,46.35,1.60,0.10),
    ]
    import numpy as _np_aud
    _aud_rows = []
    for _czl, _pmg, _ag_ref, _aa_ref, _vg_ref, _va_ref in _AUD_DATA:
        _vg_calc = _czl * _ag_ref * 1e-3
        _va_calc = _pmg * _aa_ref * 1e-3
        _aud_rows.append({
            "Lám CZL (mm)":    _czl,
            "Lám PMG (mm)":    _pmg,
            "Área Gat (km²)":  _ag_ref,
            "Área Alh (km²)":  _aa_ref,
            "Vol Gat Ref (hm³)":  _vg_ref,
            "Vol Gat Calc":  round(_vg_calc, 3),
            "Err Gat %":     round((_vg_calc-_vg_ref)/_vg_ref*100, 1) if _vg_ref else 0,
            "Vol Alh Ref (hm³)":  _va_ref,
            "Vol Alh Calc":  round(_va_calc, 3),
            "Err Alh %":     round((_va_calc-_va_ref)/_va_ref*100, 1) if _va_ref else 0,
        })
    _df_aud = pd.DataFrame(_aud_rows)
    _mean_eg = _df_aud["Err Gat %"].mean()
    _mean_ea = _df_aud["Err Alh %"].mean()
    _aud_c1, _aud_c2 = st.columns(2)
    _aud_c1.metric("Error promedio Gatún vs referencia",  f"{_mean_eg:+.1f}%")
    _aud_c2.metric("Error promedio Alhajuela vs referencia", f"{_mean_ea:+.1f}%")
    st.caption(
        "ℹ️ El error sistemático (~+19% Gatún, ~+18% Alh) refleja que la referencia usa "
        "lámina NET de lago (después de aplicar coeficiente de bandeja ~0.84), "
        "mientras que las columnas CZL/PMG son medición bruta de bandeja evaporimétrica. "
        "Para replicar exactamente la referencia: Vol = Lámina_banda × 0.84 × Área × 10⁻³")
    st.dataframe(_df_aud, use_container_width=True, hide_index=True, height=400)

# ═══ TAB 7 — CONVERSOR ═══
with tabs[7]:
    st.subheader("🔄 Conversor de unidades")
    cv1,cv2 = st.columns(2)
    with cv1:
        st.markdown("### Caudal")
        m1 = st.radio("Desde:",["cfs","m³/s","hm³/día"],horizontal=True,key="mq")
        v1 = st.number_input("Valor",0.0,999999.0,100.0,key="vq")
        if m1=="cfs":
            st.success(f"**{v1:.2f} cfs** = **{v1*CFS2M3S:.4f} m³/s** = **{v1*CFS2HM3:.4f} hm³/día**")
        elif m1=="m³/s":
            st.success(f"**{v1:.4f} m³/s** = **{v1*M3S2CFS:.2f} cfs** = **{v1*M3S2CFS*CFS2HM3:.4f} hm³/día**")
        else:
            st.success(f"**{v1:.4f} hm³/día** = **{v1/CFS2HM3:.2f} cfs** = **{v1*HM3D2M3S:.4f} m³/s**")
    with cv2:
        st.markdown("### Volumen")
        m2 = st.radio("Desde:",["hm³","MPC","Mgal"],horizontal=True,key="mv")
        st.caption("MPC = millones de pies cúbicos · Mgal = millones de galones US. Conversión en ambos sentidos: hm³ ↔ MPC ↔ Mgal.")
        v2 = st.number_input("Valor ",0.0,999999.0,1.0,key="vv")
        if m2=="hm³":
            st.success(f"**{v2:,.4f} hm³** = **{v2*HM3_TO_MPC:,.4f} MPC** = **{v2*HM3_TO_MGAL:,.2f} Mgal**")
        elif m2=="MPC":
            h = v2*MPC_TO_HM3
            st.success(f"**{v2:,.4f} MPC** = **{h:,.6f} hm³** = **{h*HM3_TO_MGAL:,.2f} Mgal**")
        else:
            h = v2*MGAL_TO_HM3
            st.success(f"**{v2:,.2f} Mgal** = **{h:,.6f} hm³** = **{h*HM3_TO_MPC:,.4f} MPC**")
    st.markdown("---")
    st.dataframe(pd.DataFrame([
        {"cfs":r,"m³/s":round(r*CFS2M3S,3),"hm³/día":round(r*CFS2HM3,4),"hm³/mes":round(r*CFS2HM3*30,2)}
        for r in [1,10,50,100,500,1000,2000,4000,5000]
    ]), use_container_width=True, hide_index=True)


# ═══ TAB 8 — EXPORTAR ═══
with tabs[8]:
    st.subheader("📤 Exportar compilado de usos del dashboard")
    st.markdown("Descarga el estado actual del dashboard en Excel con múltiples hojas.")

    # ── Construir Excel en memoria ────────────────────────────────────────────
    def build_export_excel() -> bytes:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:

            # Hoja 1: Resumen general
            resumen = pd.DataFrame([
                {"Parámetro":"Fecha de sesión","Valor":AHORA,"Unidad":""},
                {"Parámetro":"Unidad visual","Valor":unidad,"Unidad":""},
                {"Parámetro":"Fuente balance esclusajes","Valor":balance_escl_label,"Unidad":""},
                {"Parámetro":"Turn Around NPX/día","Valor":n_turnaround_npx,"Unidad":"eventos/día"},
                {"Parámetro":"Ahorro Turn Around aplicado","Valor":round(ahorro_turnaround_aplicado,4),"Unidad":"hm³/día"},
                {"Parámetro":"PNX/día","Valor":n_pnx,"Unidad":"esclusajes"},
                {"Parámetro":"NPX/día","Valor":n_npx,"Unidad":"esclusajes"},
                {"Parámetro":"Vol PNX usado en balance","Valor":round(vp_balance,4),"Unidad":"hm³/escl"},
                {"Parámetro":"Vol NPX usado en balance","Valor":round(vn_balance,4),"Unidad":"hm³/escl"},
                {"Parámetro":"Vol PNX base sidebar","Valor":round(vp,4),"Unidad":"hm³/escl"},
                {"Parámetro":"Vol NPX base sidebar","Valor":round(vn,4),"Unidad":"hm³/escl"},
                {"Parámetro":"Gen Madden","Valor":gm_mw,"Unidad":"MW"},
                {"Parámetro":"Gen Gatún","Valor":gg_mw,"Unidad":"MW"},
                {"Parámetro":"Factor Madden","Valor":mw_madden,"Unidad":"cfs/MW"},
                {"Parámetro":"Factor Gatún","Valor":mw_gatun,"Unidad":"cfs/MW"},
                {"Parámetro":"Potable Alhajuela","Valor":pot_alh,"Unidad":"cfs"},
                {"Parámetro":"Potable Gatún","Valor":pot_gat,"Unidad":"cfs"},
                {"Parámetro":"Fugas Alhajuela","Valor":fug_alh,"Unidad":"cfs"},
                {"Parámetro":"Fugas Gatún","Valor":fug_gat,"Unidad":"cfs"},
                {"Parámetro":"Vertido Fondo Madden","Valor":v_fondo,"Unidad":"cfs"},
                {"Parámetro":"Compuertas Tambor","Valor":v_tambor,"Unidad":"cfs"},
                {"Parámetro":"Vertido Libre","Valor":v_libre,"Unidad":"cfs"},
                {"Parámetro":"Vertido Gatún","Valor":v_gatun,"Unidad":"cfs"},
                {"Parámetro":"ZZ-Flush Cocolí","Valor":flush_cc,"Unidad":"hrs"},
                {"Parámetro":"ZZ-Flush A.Clara","Valor":flush_ac,"Unidad":"hrs"},
                {"Parámetro":"Fuente evaporación","Valor":evap_fuente_label,"Unidad":""},
                {"Parámetro":"Detalle evaporación Gatún","Valor":evap_detalle_gat,"Unidad":""},
                {"Parámetro":"Detalle evaporación Alhajuela","Valor":evap_detalle_alh,"Unidad":""},
                {"Parámetro":"Evap lámina Gatún","Valor":evap_gat_mm,"Unidad":"mm/día"},
                {"Parámetro":"Evap lámina Alhajuela","Valor":evap_alh_mm,"Unidad":"mm/día"},
                {"Parámetro":"Evap volumen Gatún aplicado","Valor":evap_gat,"Unidad":"hm³/día"},
                {"Parámetro":"Evap volumen Alhajuela aplicado","Valor":evap_alh,"Unidad":"hm³/día"},
                {"Parámetro":"Evap total aplicado","Valor":evap_tot,"Unidad":"hm³/día"},
                {"Parámetro":"Área espejo Gatún","Valor":round(area_gat,2),"Unidad":"km²"},
                {"Parámetro":"Área espejo Alhajuela","Valor":round(area_alh,2),"Unidad":"km²"},
                {"Parámetro":"Modo área Gatún","Valor":area_modo_gat,"Unidad":""},
                {"Parámetro":"Modo área Alhajuela","Valor":area_modo_alh,"Unidad":""},
            ])
            resumen.to_excel(writer, sheet_name="Parámetros", index=False)

            # Hoja 2: Demandas por embalse
            dem_rows = []
            for nm,(h,cf,_) in all_usos.items():
                dem_rows.append({"Uso":nm,
                    "hm³/día":round(h,4),"cfs":round(cf,1),"m³/s":round(cf*CFS2M3S,2),
                    "% Sistema":round(h/max(dem_total,.001)*100,2)})
            dem_rows.append({"Uso":"TOTAL SISTEMA",
                "hm³/día":round(dem_total,4),"cfs":round(dem_total/CFS2HM3,1),
                "m³/s":round(dem_total*HM3D2M3S,2),"% Sistema":100.0})
            pd.DataFrame(dem_rows).to_excel(writer, sheet_name="Demandas Sistema", index=False)

            # Hoja 3: Alhajuela detalle
            tbl(alh_usos, alh_total, "Alhajuela", dem_total).to_excel(
                writer, sheet_name="Alhajuela Detalle", index=False)

            # Hoja 4: Gatún detalle
            tbl(gat_usos, gat_total, "Gatún", dem_total).to_excel(
                writer, sheet_name="Gatún Detalle", index=False)

            # Hoja 5: Curva área espejo Gatún
            # Área espejo Gatún (nivel actual ±6 ft, paso 0.5)
            _exp_ae_rows_g = []
            for _en in np.round(np.arange(max(77.0, nivel_gat_op-6),
                                          min(93.0, nivel_gat_op+6.1), 0.5), 2):
                _ea = area_desde_nivel_gat(float(_en), daily=_use_daily_gat)
                _exp_ae_rows_g.append({
                    "Nivel Gatún (ft)": round(_en,2),
                    "Área Daily (km²)": round(_ea,4),
                    "Área Std (km²)":   round(area_desde_nivel_gat(float(_en),daily=False),4),
                    f"Evap aplicada {evap_gat_mm:.3f} mm × {EVAP_COEF:.2f} (hm³/d)": round(evap_gat_mm*_ea*1e-3*EVAP_COEF, 5),
                })
            pd.DataFrame(_exp_ae_rows_g).to_excel(
                writer, sheet_name="Área Espejo Gatún", index=False)

            # Área espejo Alhajuela (nivel actual ±10 ft, paso 1)
            _exp_ae_rows_a = []
            for _en in np.round(np.arange(max(160.0, nivel_alh_op-10),
                                          min(260.0, nivel_alh_op+10.1), 1.0), 2):
                _ea = area_desde_nivel_alh(float(_en), daily=_use_daily_alh)
                _exp_ae_rows_a.append({
                    "Nivel Alhajuela (ft)": round(_en,2),
                    "Área Daily (km²)":     round(_ea,4),
                    "Área Std (km²)":       round(area_desde_nivel_alh(float(_en),daily=False),4),
                    f"Evap aplicada {evap_alh_mm:.3f} mm × {EVAP_COEF:.2f} (hm³/d)": round(evap_alh_mm*_ea*1e-3*EVAP_COEF, 5),
                })
            pd.DataFrame(_exp_ae_rows_a).to_excel(
                writer, sheet_name="Área Espejo Alhajuela", index=False)

        buf.seek(0)
        return buf.read()

    # ── Vista previa ─────────────────────────────────────────────────────────
    exp_c1, exp_c2 = st.columns(2)
    with exp_c1:
        st.markdown("**📋 Resumen de demandas (sistema)**")
        rows_exp = []
        for nm,(h,cf,_) in all_usos.items():
            if h>0.0001:
                rows_exp.append({"Uso":nm,"hm³/día":round(h,4),
                    "cfs":round(cf,1),"m³/s":round(cf*CFS2M3S,2),
                    "%":round(h/max(dem_total,.001)*100,1)})
        rows_exp.append({"Uso":"TOTAL","hm³/día":round(dem_total,4),
            "cfs":round(dem_total/CFS2HM3,1),
            "m³/s":round(dem_total*HM3D2M3S,2),"%":100.0})
        st.dataframe(pd.DataFrame(rows_exp), use_container_width=True, hide_index=True)
    with exp_c2:
        st.markdown("**⚙️ Parámetros actuales**")
        params_prev = pd.DataFrame([
            {"Parámetro":"PNX/día",       "Valor":n_pnx},
            {"Parámetro":"NPX/día",       "Valor":n_npx},
            {"Parámetro":"Vol PNX balance (hm³/escl)", "Valor":round(vp_balance,4)},
            {"Parámetro":"Vol NPX balance (hm³/escl)", "Valor":round(vn_balance,4)},
            {"Parámetro":"Gen Madden (MW)","Valor":gm_mw},
            {"Parámetro":"Gen Gatún (MW)","Valor":gg_mw},
            {"Parámetro":"Área Gatún (km²)","Valor":round(area_gat,1)},
            {"Parámetro":"Área Alh (km²)","Valor":round(area_alh,1)},
            {"Parámetro":"Demanda total (hm³/d)","Valor":round(dem_total,3)},
            {"Parámetro":"Demanda total (cfs)",  "Valor":round(dem_total/CFS2HM3,1)},
            {"Parámetro":"Demanda total (m³/s)", "Valor":round(dem_total*HM3D2M3S,2)},
        ])
        st.dataframe(params_prev, use_container_width=True, hide_index=True)

    st.markdown("---")
    fname = f"demandas_ACP_{datetime.date.today().isoformat()}.xlsx"
    st.download_button(
        label="⬇️ Descargar compilado Excel (.xlsx)",
        data=build_export_excel(),
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )
    st.caption(f"El archivo incluye hojas: Parámetros · Demandas Sistema · "
               f"Alhajuela Detalle · Gatún Detalle · Área Espejo Gatún · Área Espejo Alhajuela")


# ═══ TAB 9 — DATOS LAKE HOUSE ═══
with tabs[9]:
    st.subheader("📂 Datos Lake House")

    @st.cache_data(show_spinner="Cargando LakeHouse...")
    def cargar_lkh(src, sh):
        df = pd.read_excel(src, sheet_name=sh)
        col_f = None
        for c in df.columns:
            if "date" in str(c).lower(): col_f=c; break
        if col_f is None: col_f=df.columns[1]
        df["fecha"] = pd.to_datetime(df[col_f], errors="coerce")
        df = df.dropna(subset=["fecha"]).sort_values("fecha").reset_index(drop=True)
        rn = {}
        for c in df.columns:
            cl = str(c).lower()
            if "madel" in cl:               rn[c]="nv_a"
            elif "gatel" in cl:             rn[c]="nv_g"
            elif cl=="numlockgat":          rn[c]="n_g"
            elif cl=="numlockpm":           rn[c]="n_p"
            elif cl in ("numlockac","numlockacl"): rn[c]="n_a"
            elif cl=="numlockccl":          rn[c]="n_c"
            elif cl=="gatlockhm3":          rn[c]="gat_hm3"
            elif cl=="pmlockhm3":           rn[c]="pm_hm3"
            elif cl=="aclockhm3":           rn[c]="acl_hm3"
            elif cl=="ccllockhm3":          rn[c]="ccl_hm3"
            elif cl=="gatlockporlockhec":   rn[c]="gat_unit_hm3"
            elif cl=="pmlockporlockhec":    rn[c]="pm_unit_hm3"
            elif cl in ("aclocporlockhec", "aclockporlockhec"): rn[c]="acl_unit_hm3"
            elif cl=="ccllockporlockhec":   rn[c]="ccl_unit_hm3"
            elif cl=="total pnx":           rn[c]="pnx_unit_hm3"
            elif cl=="total npx":           rn[c]="npx_unit_hm3"
            elif "gatlockmcf" in cl:        rn[c]="gat_mcf"
            elif "pmlockmcf" in cl:         rn[c]="pm_mcf"
            elif "aclockmcf" in cl:         rn[c]="acl_mcf"
            elif "ccllockmcf" in cl:        rn[c]="ccl_mcf"
            elif cl=="gatspill":            rn[c]="vert_g"
            elif cl=="madspill":            rn[c]="vert_m"
            elif cl=="munic_mad_hm3":       rn[c]="mun_m_hm3"
            elif cl=="munic_gat_hm3":       rn[c]="mun_g_hm3"
            elif cl=="munic_mad":           rn[c]="mun_m"
            elif cl=="munic_gat":           rn[c]="mun_g"
            elif cl=="leak_mad_hm3":        rn[c]="leak_m_hm3"
            elif cl=="leak_gat_hm3":        rn[c]="leak_g_hm3"
            elif cl=="leak_mad":            rn[c]="leak_m"
            elif cl=="leak_gat":            rn[c]="leak_g"
            elif cl=="evap_gatun_mm":       rn[c]="evap_gat_mm"
            elif cl=="evap_alaj_mm":        rn[c]="evap_alh_mm"
            elif cl=="vol_evap_gat_hm3":    rn[c]="evap_gat_hm3"
            elif cl=="vol_evap_ala_hm3":    rn[c]="evap_alh_hm3"
            elif cl=="saving_water_panamax":rn[c]="ahorro_pnx"
            elif cl=="total_saving_water_neo_hm3": rn[c]="ahorro_npx"
            elif cl=="cca_neo":              rn[c]="cca_neo_val"
            elif cl=="madhm3":              rn[c]="gen_mad_hm3"
            elif cl=="gathm3":              rn[c]="gen_gat_hm3"
            elif cl in ("madmwh",):         rn[c]="mad_mwh"
            elif cl=="gatmwh":              rn[c]="gat_mwh"
            elif "total todos" in cl and "hec" in cl: rn[c]="total_escl_hm3"
            elif cl=="capgat_hm3":          rn[c]="cap_gat_hm3"
            elif cl=="capmad_hm3":          rn[c]="cap_mad_hm3"
            elif cl=="usos_hm3":             rn[c]="usos_hm3"
        df = df.rename(columns=rn)
        for c in set(rn.values()):
            if c in df and isinstance(df[c], pd.Series):
                df[c]=pd.to_numeric(df[c],errors="coerce")
        if "gat_hm3" in df and "pm_hm3" in df:
            df["pnx_hm3"]=df["gat_hm3"].fillna(0)+df["pm_hm3"].fillna(0)
        if "acl_hm3" in df and "ccl_hm3" in df:
            df["npx_hm3"]=df["acl_hm3"].fillna(0)+df["ccl_hm3"].fillna(0)
        if "pnx_hm3" in df and "npx_hm3" in df:
            df["total_hm3"]=df["pnx_hm3"]+df["npx_hm3"]
        elif "gat_mcf" in df and "pm_mcf" in df:
            df["pnx_m"]=df["gat_mcf"].fillna(0)+df["pm_mcf"].fillna(0)
            if "acl_mcf" in df and "ccl_mcf" in df:
                df["npx_m"]=df["acl_mcf"].fillna(0)+df["ccl_mcf"].fillna(0)
            if "pnx_m" in df and "npx_m" in df:
                # Columnas *MCF* del LakeHouse son millones de pies³/día; primero se llevan a cfs promedio diario.
                df["pnx_hm3"]=df["pnx_m"]*MCF_TO_CFS*CFS2HM3
                df["npx_hm3"]=df["npx_m"]*MCF_TO_CFS*CFS2HM3
                df["total_hm3"]=df["pnx_hm3"]+df["npx_hm3"]

        # madspill = volumen diario vertido en Madden en MPC/MCF
        # (millones de pies³ por día). Se convierte directamente a hm³/día.
        if "vert_m" in df:
            df["vert_m_hm3"] = pd.to_numeric(df["vert_m"], errors="coerce") * MPC_TO_HM3
        # Tránsitos diarios por tipo de esclusaje:
        # se usa el PROMEDIO entre complejos y no la suma, para evitar duplicar el tránsito
        # cuando el mismo buque pasa por ambos complejos del sistema Panamax o NeoPanamax.
        if "n_g" in df and "n_p" in df:
            df["n_pnx_r"] = df[["n_g", "n_p"]].mean(axis=1, skipna=True)
        if "n_a" in df and "n_c" in df:
            df["n_npx_r"] = df[["n_a", "n_c"]].mean(axis=1, skipna=True)
        return df

    lf = _buscar_lakehouse_local(); dl = None

    # La pestaña siempre permite subir un LakeHouse, aun cuando exista uno local.
    # Prioridad de uso: archivo subido por el usuario → archivo local seleccionado → archivo local más reciente.
    fl = st.file_uploader(
        "Sube LakeHouse (xlsx) opcional",
        type=["xlsx"],
        key="lk",
        help=(
            "Si cargas un archivo aquí, el app lo usará con prioridad. "
            "Si no cargas nada, usará automáticamente el LakeHouse local detectado junto al app."
        ),
    )

    fuente_lkh = None
    fuente_txt = ""
    if fl is not None:
        fuente_lkh = fl
        fuente_txt = f"archivo subido: {getattr(fl, 'name', 'LakeHouse.xlsx')}"
    elif lf:
        if len(lf) > 1:
            sel_local = st.selectbox(
                "Archivo LakeHouse local detectado",
                options=lf,
                format_func=lambda x: os.path.basename(str(x)),
                key="lkh_local_path_select",
                help="Se usará este archivo local como base del LakeHouse."
            )
        else:
            sel_local = lf[0]
            st.session_state.setdefault("lkh_local_path_select", sel_local)
        fuente_lkh = sel_local
        fuente_txt = f"archivo local: {os.path.basename(str(sel_local))}"
    else:
        st.info("No se detectó un LakeHouse local. Puede subir uno en el selector anterior.")

    if fuente_lkh is not None:
        try:
            if hasattr(fuente_lkh, "seek"):
                fuente_lkh.seek(0)
            xls = pd.ExcelFile(fuente_lkh)
            hojas_validas = [x for x in xls.sheet_names if x not in ["Sheet1", "Para BalanceH"]] or xls.sheet_names
            hoja = st.selectbox("Hoja", hojas_validas, key="lkh_sheet") if len(hojas_validas) > 1 else hojas_validas[0]
            if hasattr(fuente_lkh, "seek"):
                fuente_lkh.seek(0)
            dl = cargar_lkh(fuente_lkh, hoja)
            st.success(
                f"✅ Usando {fuente_txt} · {len(dl):,} registros · "
                f"{dl['fecha'].min().date()} → {dl['fecha'].max().date()}"
            )
        except Exception as e:
            st.error(str(e))

    if dl is not None and len(dl)>0:
        total_dias = (dl["fecha"].max()-dl["fecha"].min()).days
        st.markdown("---")
        try:
            dias_sel = int(st.session_state.get("dias_op", 5))
        except Exception:
            dias_sel = 5
        if dias_sel not in (1, 5, 7, 10, 30):
            dias_sel = 5
        st.info(
            f"📅 Promedio activo: **últimos {dias_sel} días/registros**. "
            "Puede cambiarlo en la sección superior «Variables operativas por embalse»."
        )

        # Usar exactamente los últimos N registros/días disponibles, no una ventana inclusiva por fecha.
        # Esto evita que, por ejemplo, "5 días" incluya 6 registros al restar 5 días al último dato.
        def _ultimos_dias(df_base, n):
            return df_base.sort_values("fecha").tail(n).copy()

        dv = _ultimos_dias(dl, dias_sel)
        ultimo_reg = dl.sort_values("fecha").iloc[-1]
        fecha_ultimo = ultimo_reg["fecha"].date()
        st.caption(
            f"Mostrando: **{len(dv)} registros/días** · {dv['fecha'].min().date()} → {dv['fecha'].max().date()} · "
            f"último registro disponible: **{fecha_ultimo}**"
        )

        # ── Resumen rápido: métricas clave ──────────────────────────────────────
        st.markdown("---")
        lk1, lk2, lk3, lk4, lk5 = st.columns(5)
        if "nv_g" in dl and pd.notna(ultimo_reg.get("nv_g", np.nan)):
            lk1.metric("Nivel Gatún (último día)", f"{ultimo_reg['nv_g']:.2f} ft")
        if "nv_a" in dl and pd.notna(ultimo_reg.get("nv_a", np.nan)):
            lk2.metric("Nivel Alhajuela (último día)", f"{ultimo_reg['nv_a']:.2f} ft")
        if "n_pnx_r" in dv:
            lk3.metric(f"PNX/d ({dias_sel}d)", f"{dv['n_pnx_r'].mean():.1f}")
        if "n_npx_r" in dv:
            lk4.metric(f"NPX/d ({dias_sel}d)", f"{dv['n_npx_r'].mean():.1f}")

        # Evita duplicar la misma demanda de esclusajes cuando el LakeHouse trae
        # tanto total_hm3 calculado como total_escl_hm3 directo. Se muestra una sola tarjeta.
        _cons_escl = None
        _cons_escl_fuente = ""
        if "total_hm3" in dv and dv["total_hm3"].notna().sum() > 0:
            _cons_escl = float(pd.to_numeric(dv["total_hm3"], errors="coerce").mean())
            _cons_escl_fuente = "PNX + NPX"
        elif "total_escl_hm3" in dv and dv["total_escl_hm3"].notna().sum() > 0:
            _cons_escl = float(pd.to_numeric(dv["total_escl_hm3"], errors="coerce").mean())
            _cons_escl_fuente = "Total escl. LakeHouse"
        if _cons_escl is not None:
            lk5.metric(f"Consumo esclusajes ({dias_sel}d)", f"{_cons_escl:.2f} hm³/d")

        st.caption(
            "Los niveles corresponden al último día disponible; PNX y NPX son el promedio de tránsitos entre complejos, no la suma; "
            f"el consumo de esclusajes se muestra una sola vez y corresponde al promedio del período seleccionado"
            f"{(' · fuente: ' + _cons_escl_fuente) if _cons_escl_fuente else ''}."
        )

        # ── Promedios operativos por categoría ──────────────────────────────────
        st.markdown("---")
        st.subheader(f"📊 Promedios operativos — últimos {dias_sel} días")

        # Conversiones de flujo
        MCF_TO_CFS = 1_000_000.0 / 86400.0   # MCF (millones de pies³/día) → cfs

        def _hm3_series(df_base, col_hm3=None, col_mcf=None, prefer_mcf=False):
            """Devuelve una serie en hm³/día.

            Para fugas se puede priorizar MCF porque en algunos LakeHouse el campo *_hm3
            puede venir desajustado respecto al consumo diario usado por el balance.
            """
            if prefer_mcf and col_mcf and col_mcf in df_base and df_base[col_mcf].notna().sum() > 0:
                return pd.to_numeric(df_base[col_mcf], errors="coerce") * MCF_TO_CFS * CFS2HM3
            if col_hm3 and col_hm3 in df_base and df_base[col_hm3].notna().sum() > 0:
                return pd.to_numeric(df_base[col_hm3], errors="coerce")
            if col_mcf and col_mcf in df_base and df_base[col_mcf].notna().sum() > 0:
                return pd.to_numeric(df_base[col_mcf], errors="coerce") * MCF_TO_CFS * CFS2HM3
            return None

        def _prom_flow(series, src_unit):
            """Retorna dict con Promedio en hm³/d, cfs, m³/s a partir de la serie y unidad fuente."""
            m = pd.to_numeric(series, errors="coerce").mean()
            if src_unit == "hm3":
                hm3 = m
            elif src_unit == "mcf":
                hm3 = m * MCF_TO_CFS * CFS2HM3
            else:
                return {"Prom hm³/d":"—","Prom cfs":"—","Prom m³/s":"—"}
            cfs_v = hm3 / CFS2HM3
            m3s   = hm3 * HM3D2M3S
            return {"Prom hm³/d":round(hm3,3),"Prom cfs":round(cfs_v,1),"Prom m³/s":round(m3s,2)}

        def _flow_prom_from_hm3(hm3):
            if hm3 is None or pd.isna(hm3):
                return None
            return {
                "hm3": round(float(hm3), 3),
                "cfs": round(float(hm3) / CFS2HM3, 1),
                "m3s": round(float(hm3) * HM3D2M3S, 2),
            }

        def _serie_unitaria(df_base, unit_col=None, consumo_col=None, trans_col=None):
            """Devuelve serie hm³/escl; usa columna unitaria si existe, si no calcula consumo/tránsitos."""
            if unit_col and unit_col in df_base and df_base[unit_col].notna().sum() > 0:
                return pd.to_numeric(df_base[unit_col], errors="coerce")
            if consumo_col and trans_col and consumo_col in df_base and trans_col in df_base:
                cons = pd.to_numeric(df_base[consumo_col], errors="coerce")
                trans = pd.to_numeric(df_base[trans_col], errors="coerce")
                return cons.where(trans > 0) / trans.replace(0, np.nan)
            return None

        def _prom_unitario(df_base, unit_col=None, consumo_col=None, trans_col=None):
            s = _serie_unitaria(df_base, unit_col=unit_col, consumo_col=consumo_col, trans_col=trans_col)
            if s is None or s.notna().sum() == 0:
                return None
            hm3_escl = float(s.mean())
            return {
                "hm3_escl": round(hm3_escl, 4),
                "cfs_eq": round(hm3_escl / CFS2HM3, 1),
                "m3s_eq": round(hm3_escl * HM3D2M3S, 2),
            }

        def _promedio_componentes(df_base, componentes):
            series = []
            for comp in componentes:
                col_hm3, col_mcf = comp[0], comp[1]
                prefer_mcf = bool(comp[2]) if len(comp) >= 3 else False
                s = _hm3_series(df_base, col_hm3=col_hm3, col_mcf=col_mcf, prefer_mcf=prefer_mcf)
                if s is not None:
                    series.append(s.fillna(0))
            if not series:
                return None
            total = series[0].copy()
            for s in series[1:]:
                total = total.add(s, fill_value=0)
            return _flow_prom_from_hm3(total.mean())

        # --- Resumen visual de consumos del período seleccionado ---
        def _fmt_lkh(v, unit, nd=2):
            try:
                if v is None or pd.isna(v):
                    return "—"
                return f"{float(v):.{nd}f} {unit}"
            except Exception:
                return "—"

        def _lkh_card(label, value, sub=""):
            st.markdown(
                f"""
                <div class="lkh-card">
                    <div class="label">{label}</div>
                    <div class="value">{value}</div>
                    <div class="sub">{sub}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        def _mean_col(col):
            if col in dv and dv[col].notna().sum() > 0:
                return float(pd.to_numeric(dv[col], errors="coerce").mean())
            return None

        _escl_hm3 = _mean_col("total_hm3")
        if _escl_hm3 is None:
            _escl_hm3 = _mean_col("total_escl_hm3")
        _pot_prom = _promedio_componentes(dv, [("mun_m_hm3", "mun_m", True), ("mun_g_hm3", "mun_g", True)])
        _fug_prom = _promedio_componentes(dv, [("leak_m_hm3", "leak_m", True), ("leak_g_hm3", "leak_g", True)])
        # La evaporación mostrada corresponde a la fuente activa seleccionada en el sidebar.
        _evap_prom = _flow_prom_from_hm3(evap_tot)
        _gen_mwh = 0.0
        _gen_count = 0
        for _c in ["mad_mwh", "gat_mwh"]:
            if _c in dv and dv[_c].notna().sum() > 0:
                _gen_mwh += float(pd.to_numeric(dv[_c], errors="coerce").mean())
                _gen_count += 1
        _gen_mw = (_gen_mwh / 24.0) if _gen_count > 0 else None

        st.markdown("#### 📌 Resumen de consumos promedio del período seleccionado")
        cA, cB, cC, cD, cE = st.columns(5)
        with cA:
            _lkh_card("🚢 Esclusajes", _fmt_lkh(_escl_hm3, "hm³/d", 2), f"{_fmt_lkh((_escl_hm3 / CFS2HM3) if _escl_hm3 is not None else None, 'cfs', 0)} · últimos {dias_sel} días")
        with cB:
            _lkh_card("🚰 Potabilización", _fmt_lkh(_pot_prom['hm3'] if _pot_prom else None, "hm³/d", 2), f"{_fmt_lkh(_pot_prom['cfs'] if _pot_prom else None, 'cfs', 0)} · promedio")
        with cC:
            _lkh_card("💨 Fugas", _fmt_lkh(_fug_prom['hm3'] if _fug_prom else None, "hm³/d", 2), f"{_fmt_lkh(_fug_prom['cfs'] if _fug_prom else None, 'cfs', 0)} · promedio")
        with cD:
            _lkh_card("⚡ Generación", _fmt_lkh(_gen_mw, "MW", 1), "MWh/d ÷ 24 · promedio horario")
        with cE:
            _lkh_card("☀️ Evaporación", _fmt_lkh(_evap_prom['hm3'] if _evap_prom else None, "hm³/d", 2), f"{_fmt_lkh(_evap_prom['cfs'] if _evap_prom else None, 'cfs', 0)} · {evap_fuente_corta}")
        st.caption(
            f"Esclusajes, potabilización, fugas y generación se calculan con los últimos {dias_sel} registros/días seleccionados del LakeHouse. "
            f"La evaporación usa la fuente activa **{evap_fuente_label}** y se aplica por embalse."
        )

        # --- Salidas de agua por embalse desde LakeHouse + evaporación de la fuente activa ---
        st.markdown("#### 🏔️🌊 Salidas de agua por embalse — LakeHouse + evaporación activa")

        def _agregar_salida_embalse(rows, embalse, salida, componentes):
            prom = _promedio_componentes(dv, componentes)
            if prom is None:
                return
            rows.append({
                "Embalse": embalse,
                "Salida / uso de agua": salida,
                "Prom hm³/d": prom["hm3"],
                "Prom cfs": prom["cfs"],
                "Prom m³/s": prom["m3s"],
                "Fuente": "LakeHouse",
            })

        def _agregar_salida_embalse_app(rows, embalse, salida, hm3_val):
            prom = _flow_prom_from_hm3(hm3_val)
            if prom is None:
                return
            rows.append({
                "Embalse": embalse,
                "Salida / uso de agua": salida,
                "Prom hm³/d": prom["hm3"],
                "Prom cfs": prom["cfs"],
                "Prom m³/s": prom["m3s"],
                "Fuente": evap_fuente_label,
            })

        salidas_embalse_rows = []
        # Alhajuela
        _agregar_salida_embalse(salidas_embalse_rows, "Alhajuela", "Generación Madden", [("gen_mad_hm3", None)])
        _agregar_salida_embalse(salidas_embalse_rows, "Alhajuela", "Potabilización", [("mun_m_hm3", "mun_m", True)])
        _agregar_salida_embalse(salidas_embalse_rows, "Alhajuela", "Fugas", [("leak_m_hm3", "leak_m", True)])
        _agregar_salida_embalse(salidas_embalse_rows, "Alhajuela", "Vertido Madden", [("vert_m_hm3", None)])
        _agregar_salida_embalse_app(salidas_embalse_rows, "Alhajuela", "Evaporación", evap_alh)

        # Gatún: si existen PNX/NPX separados se muestran por separado; si no, se usa total de esclusajes.
        _hay_pnx_npx_lkh = any(c in dv and dv[c].notna().sum() > 0 for c in ["pnx_hm3", "npx_hm3"])
        if _hay_pnx_npx_lkh:
            _agregar_salida_embalse(salidas_embalse_rows, "Gatún", "Esclusajes PNX", [("pnx_hm3", None)])
            _agregar_salida_embalse(salidas_embalse_rows, "Gatún", "Esclusajes NPX", [("npx_hm3", None)])
        else:
            _col_total_escl = "total_hm3" if ("total_hm3" in dv and dv["total_hm3"].notna().sum() > 0) else "total_escl_hm3"
            _agregar_salida_embalse(salidas_embalse_rows, "Gatún", "Esclusajes total", [(_col_total_escl, None)])
        _agregar_salida_embalse(salidas_embalse_rows, "Gatún", "Generación Gatún", [("gen_gat_hm3", None)])
        _agregar_salida_embalse(salidas_embalse_rows, "Gatún", "Potabilización", [("mun_g_hm3", "mun_g", True)])
        _agregar_salida_embalse(salidas_embalse_rows, "Gatún", "Fugas", [("leak_g_hm3", "leak_g", True)])
        _agregar_salida_embalse(salidas_embalse_rows, "Gatún", "Vertido Gatún", [(None, "vert_g")])
        _agregar_salida_embalse_app(salidas_embalse_rows, "Gatún", "Evaporación", evap_gat)

        if salidas_embalse_rows:
            df_sal_emb = pd.DataFrame(salidas_embalse_rows)
            tot_emb = df_sal_emb.groupby("Embalse", as_index=False)[["Prom hm³/d", "Prom cfs", "Prom m³/s"]].sum()
            tot_sistema = {
                "Prom hm³/d": float(tot_emb["Prom hm³/d"].sum()),
                "Prom cfs": float(tot_emb["Prom cfs"].sum()),
                "Prom m³/s": float(tot_emb["Prom m³/s"].sum()),
            }

            e1, e2, e3 = st.columns(3)
            _tot_alh = tot_emb.loc[tot_emb["Embalse"].eq("Alhajuela")]
            _tot_gat = tot_emb.loc[tot_emb["Embalse"].eq("Gatún")]
            with e1:
                _lkh_card(
                    "🏔️ Total Alhajuela",
                    _fmt_lkh(_tot_alh["Prom hm³/d"].iloc[0] if not _tot_alh.empty else None, "hm³/d", 2),
                    f"{_fmt_lkh(_tot_alh['Prom cfs'].iloc[0] if not _tot_alh.empty else None, 'cfs', 0)} · LakeHouse + evap. activa",
                )
            with e2:
                _lkh_card(
                    "🌊 Total Gatún",
                    _fmt_lkh(_tot_gat["Prom hm³/d"].iloc[0] if not _tot_gat.empty else None, "hm³/d", 2),
                    f"{_fmt_lkh(_tot_gat['Prom cfs'].iloc[0] if not _tot_gat.empty else None, 'cfs', 0)} · LakeHouse + evap. activa",
                )
            with e3:
                _lkh_card(
                    "💧 Total sistema",
                    _fmt_lkh(tot_sistema["Prom hm³/d"], "hm³/d", 2),
                    f"{_fmt_lkh(tot_sistema['Prom cfs'], 'cfs', 0)} · suma LakeHouse + evap. activa",
                )

            st.dataframe(df_sal_emb, use_container_width=True, hide_index=True)
            fig_sal_emb = go.Figure()
            for _embalse in df_sal_emb["Embalse"].dropna().unique():
                _dfe = df_sal_emb[df_sal_emb["Embalse"] == _embalse]
                fig_sal_emb.add_bar(
                    name=_embalse,
                    x=_dfe["Salida / uso de agua"],
                    y=_dfe["Prom hm³/d"],
                    text=[f"{v:.2f}" for v in _dfe["Prom hm³/d"]],
                    textposition="outside",
                )
            fig_sal_emb.update_layout(
                barmode="group",
                height=420,
                title=f"Salidas promedio por embalse — últimos {dias_sel} días",
                yaxis_title="hm³/día",
                xaxis_title="Uso / salida de agua",
                margin=dict(t=70, b=90),
            )
            st.plotly_chart(fig_sal_emb, use_container_width=True)
            st.caption(f"La tabla separa las salidas por embalse. La evaporación corresponde a **{evap_fuente_label}**.")
        else:
            st.info("No se encontraron columnas suficientes en el LakeHouse para separar las salidas por embalse.")

        # --- Potabilización y fugas comparativo 1/5/7/10/30 días ---
        st.markdown("#### 🚰 Potabilización y fugas — promedios 1, 5, 7, 10 y 30 días")
        comp_def = [
            ("Potable Alhajuela", [("mun_m_hm3", "mun_m", True)]),
            ("Potable Gatún", [("mun_g_hm3", "mun_g", True)]),
            ("Potable total", [("mun_m_hm3", "mun_m", True), ("mun_g_hm3", "mun_g", True)]),
            ("Fugas Alhajuela", [("leak_m_hm3", "leak_m", True)]),
            ("Fugas Gatún", [("leak_g_hm3", "leak_g", True)]),
            ("Fugas total", [("leak_m_hm3", "leak_m", True), ("leak_g_hm3", "leak_g", True)]),
        ]
        comp_rows = []
        for etiqueta, componentes in comp_def:
            fila = {"Parámetro": etiqueta}
            tiene_dato = False
            for n_comp in [1, 5, 7, 10, 30]:
                dcomp = _ultimos_dias(dl, n_comp)
                prom = _promedio_componentes(dcomp, componentes)
                if prom is None:
                    fila[f"{n_comp}d hm³/d"] = "—"
                    fila[f"{n_comp}d cfs"] = "—"
                    fila[f"{n_comp}d m³/s"] = "—"
                else:
                    tiene_dato = True
                    fila[f"{n_comp}d hm³/d"] = prom["hm3"]
                    fila[f"{n_comp}d cfs"] = prom["cfs"]
                    fila[f"{n_comp}d m³/s"] = prom["m3s"]
            if tiene_dato:
                comp_rows.append(fila)
        if comp_rows:
            st.dataframe(pd.DataFrame(comp_rows), use_container_width=True, hide_index=True)
            st.caption("Promedios calculados con los últimos 1, 5, 7, 10 y 30 registros/días disponibles. En fugas se priorizan MCF (`leak_mad` / `leak_gat`) para evitar desajustes del campo *_hm3.")
        else:
            st.info("No se encontraron columnas de potabilización o fugas en la hoja seleccionada.")

        # --- Esclusajes por complejo ---
        st.markdown("#### 🚢 Esclusajes por complejo")
        escl_rows = []
        escl_items = [
            ("n_pnx_r","Panamax (promedio PNX)","count"),
            ("n_g","  ↳ Gatún","count"),("n_p","  ↳ Pedro Miguel","count"),
            ("n_npx_r","Neopanamax (promedio NPX)","count"),
            ("n_a","  ↳ Agua Clara","count"),("n_c","  ↳ Cocolí","count"),
            ("pnx_hm3","Consumo agua PNX","hm3"),
            ("npx_hm3","Consumo agua NPX","hm3"),
            ("total_hm3","Consumo total esclusajes","hm3"),
        ]
        for col_name,label,src in escl_items:
            if col_name in dv and dv[col_name].notna().sum()>0:
                v = dv[col_name]
                if src == "count":
                    escl_rows.append({"Parámetro":label,
                        "Promedio":round(v.mean(),1),"Unidad":"escl/día",
                        "Prom hm³/d":"—", "Prom cfs":"—", "Prom m³/s":"—"})
                else:
                    row = {"Parámetro":label, "Promedio":"—", "Unidad":"hm³/d"}
                    row.update(_prom_flow(v, src))
                    escl_rows.append(row)
        if escl_rows:
            st.dataframe(pd.DataFrame(escl_rows).fillna("—"), use_container_width=True, hide_index=True)
            st.caption("PNX y NPX se muestran como promedio de tránsitos entre complejos. Los consumos de agua se mantienen como volumen total diario del sistema.")

        # --- Consumo unitario por esclusaje ---
        st.markdown("#### ⚓ Consumo unitario por esclusaje")
        unit_rows = []
        unit_items = [
            ("PNX total", "pnx_unit_hm3", "pnx_hm3", "n_pnx_r"),
            ("  ↳ Gatún", "gat_unit_hm3", "gat_hm3", "n_g"),
            ("  ↳ Pedro Miguel", "pm_unit_hm3", "pm_hm3", "n_p"),
            ("NPX total", "npx_unit_hm3", "npx_hm3", "n_npx_r"),
            ("  ↳ Agua Clara", "acl_unit_hm3", "acl_hm3", "n_a"),
            ("  ↳ Cocolí", "ccl_unit_hm3", "ccl_hm3", "n_c"),
        ]
        for label, unit_col, consumo_col, trans_col in unit_items:
            prom_u = _prom_unitario(dv, unit_col=unit_col, consumo_col=consumo_col, trans_col=trans_col)
            if prom_u is not None:
                unit_rows.append({
                    "Parámetro": label,
                    "Prom hm³/escl": prom_u["hm3_escl"],
                    "Equiv cfs/escl": prom_u["cfs_eq"],
                    "Equiv m³/s/escl": prom_u["m3s_eq"],
                })
        # Total ponderado usando el consumo total diario dividido entre los tránsitos promedio PNX+NPX.
        if "total_hm3" in dv and {"n_pnx_r", "n_npx_r"}.issubset(dv.columns):
            trans_total = pd.to_numeric(dv["n_pnx_r"], errors="coerce").fillna(0) + pd.to_numeric(dv["n_npx_r"], errors="coerce").fillna(0)
            serie_total_unit = pd.to_numeric(dv["total_hm3"], errors="coerce").where(trans_total > 0) / trans_total.replace(0, np.nan)
            if serie_total_unit.notna().sum() > 0:
                hm3_escl_total = float(serie_total_unit.mean())
                unit_rows.append({
                    "Parámetro": "Total ponderado",
                    "Prom hm³/escl": round(hm3_escl_total, 4),
                    "Equiv cfs/escl": round(hm3_escl_total / CFS2HM3, 1),
                    "Equiv m³/s/escl": round(hm3_escl_total * HM3D2M3S, 2),
                })
        if unit_rows:
            st.dataframe(pd.DataFrame(unit_rows), use_container_width=True, hide_index=True)
            st.caption("El consumo unitario se toma directamente de las columnas *PorLockHec/TOTAL PNX/TOTAL NPX* cuando existen; de lo contrario se calcula como consumo diario dividido entre el promedio de tránsitos.")
        else:
            st.info("No se encontraron columnas suficientes para calcular consumo unitario por esclusaje.")

        # --- Salidas de agua ---
        st.markdown("#### 💧 Salidas de agua")
        sal_rows = []

        def _add_sal_row(label, componentes, fuente="LakeHouse"):
            prom = _promedio_componentes(dv, componentes)
            if prom is None:
                return
            sal_rows.append({
                "Parámetro": label,
                "Fuente": fuente,
                "Prom hm³/d": prom["hm3"],
                "Prom cfs": prom["cfs"],
                "Prom m³/s": prom["m3s"],
            })

        _add_sal_row("Usos totales sistema", [("usos_hm3", None)])
        # Potabilización y fugas priorizan MCF/MPC; los campos *_hm3 quedan como respaldo.
        _add_sal_row("Potable Alhajuela", [("mun_m_hm3", "mun_m", True)], "LakeHouse MCF/MPC")
        _add_sal_row("Potable Gatún", [("mun_g_hm3", "mun_g", True)], "LakeHouse MCF/MPC")
        _add_sal_row("Fugas Alhajuela", [("leak_m_hm3", "leak_m", True)], "LakeHouse MCF/MPC")
        _add_sal_row("Fugas Gatún", [("leak_g_hm3", "leak_g", True)], "LakeHouse MCF/MPC")
        _add_sal_row("Vertido Gatún", [(None, "vert_g")], "LakeHouse")
        _add_sal_row("Vertido Madden", [("vert_m_hm3", None)], "LakeHouse cfs")

        # Evaporación: usa la fuente activa seleccionada en el sidebar.
        for _label, _hm3_app in [("Evaporación Gatún", evap_gat), ("Evaporación Alhajuela", evap_alh)]:
            _prom_app = _flow_prom_from_hm3(_hm3_app)
            if _prom_app is not None:
                sal_rows.append({
                    "Parámetro": _label,
                    "Fuente": evap_fuente_label,
                    "Prom hm³/d": _prom_app["hm3"],
                    "Prom cfs": _prom_app["cfs"],
                    "Prom m³/s": _prom_app["m3s"],
                })

        if sal_rows:
            st.dataframe(pd.DataFrame(sal_rows), use_container_width=True, hide_index=True)
            st.caption(f"Potabilización y fugas priorizan MCF/MPC y usan *_hm3 solo como respaldo. La evaporación usa **{evap_fuente_label}**. El vertido Madden `madspill` se interpreta como MPC/MCF por día.")

        # --- Generación ---
        st.markdown("#### ⚡ Generación hidroeléctrica")
        gen_rows = []
        energia_series = []
        gen_items = [
            ("gen_mad_hm3","Generación Madden","hm3"),
            ("gen_gat_hm3","Generación Gatún","hm3"),
            ("mad_mwh","Energía Madden","mwh"),
            ("gat_mwh","Energía Gatún","mwh"),
        ]
        for col_name,label,src in gen_items:
            if col_name in dv and dv[col_name].notna().sum()>0:
                v = pd.to_numeric(dv[col_name], errors="coerce")
                if src == "mwh":
                    prom_mwh_d = float(v.mean())
                    prom_mwh_h = prom_mwh_d / 24.0
                    gen_rows.append({
                        "Parámetro":label,
                        "Promedio":round(prom_mwh_d,2),
                        "Unidad":"MWh/d",
                        "Prom MWh/h":round(prom_mwh_h,2),
                        "MW promedio horario":round(prom_mwh_h,2),
                    })
                    gen_rows.append({
                        "Parámetro":label.replace("Energía", "Potencia promedio horaria"),
                        "Promedio":round(prom_mwh_h,2),
                        "Unidad":"MW",
                        "Prom MWh/h":round(prom_mwh_h,2),
                        "MW promedio horario":round(prom_mwh_h,2),
                    })
                    energia_series.append(v.fillna(0))
                else:
                    row = {"Parámetro":label}
                    row.update(_prom_flow(v, src))
                    gen_rows.append(row)
        if energia_series:
            total_mwh_d = energia_series[0].copy()
            for _s in energia_series[1:]:
                total_mwh_d = total_mwh_d.add(_s, fill_value=0)
            prom_total_mwh_d = float(total_mwh_d.mean())
            prom_total_mw = prom_total_mwh_d / 24.0
            gen_rows.append({
                "Parámetro":"Energía total",
                "Promedio":round(prom_total_mwh_d,2),
                "Unidad":"MWh/d",
                "Prom MWh/h":round(prom_total_mw,2),
                "MW promedio horario":round(prom_total_mw,2),
            })
            gen_rows.append({
                "Parámetro":"Potencia promedio horaria total",
                "Promedio":round(prom_total_mw,2),
                "Unidad":"MW",
                "Prom MWh/h":round(prom_total_mw,2),
                "MW promedio horario":round(prom_total_mw,2),
            })
        if gen_rows:
            st.dataframe(pd.DataFrame(gen_rows).fillna("—"), use_container_width=True, hide_index=True)
            st.caption("Para las columnas de energía del LakeHouse, el valor horario se calcula como MWh/d ÷ 24. Numéricamente, MWh/h equivale a MW promedio horario.")

        st.markdown("---")
        st.download_button("⬇️ Descargar período (CSV)",
            dv.to_csv(index=False).encode("utf-8"),
            f"lakehouse_{dias_sel}dias.csv","text/csv")
    else:
        st.info("Sube **LakeHouse_Data.xlsx** o **LakeHouse_NEW.xlsx**, o colócalo en la carpeta.")


# ═══ TAB 10 — INSTRUCTIVO ═══
with tabs[10]:
    st.subheader("📘 Instructivo operativo fácil")
    st.caption(
        "Siga el orden 1 → 6. Cada paso indica dónde entrar, qué revisar y qué resultado debe confirmar."
    )

    st.info(
        f"**Escenario mostrado:** Panamax {fmt_sig(n_pnx, 3)}/día + "
        f"NeoPanamax {fmt_sig(n_npx, 3)}/día = **{fmt_sig(n_t, 3)} tránsitos/día** · "
        f"Balance de esclusajes: **{balance_escl_label}**."
    )

    st.markdown("### 🚀 Inicio rápido en 6 pasos")
    paso1, paso2 = st.columns(2)
    with paso1:
        st.markdown("""
        #### 1️⃣ Cargue y confirme los datos
        Entre a **📂 Datos Lake House** y:
        1. Seleccione o cargue el archivo `.xlsx`.
        2. Confirme la hoja y la fecha del último registro.
        3. Elija el período operativo de **1, 5, 7, 10 o 30 días**.

        **Resultado esperado:** la app informa que LakeHouse fue aplicado y deja los valores editables.
        """)
    with paso2:
        st.markdown("""
        #### 2️⃣ Revise niveles y área espejo
        En el sidebar, abra **📍 Niveles Operativos** y confirme:
        - **Nivel Gatún (ft)**.
        - **Nivel Alhajuela (ft)**.

        En **☀️ Evaporación → Área espejo de embalse**, la configuración recomendada ya queda por defecto:
        - Curva hipsométrica **Daily** para Gatún y Alhajuela.
        - **Calcular desde nivel (ft)** para ambas áreas.

        Los niveles actualizan automáticamente el área del embalse, la evaporación y los cálculos que dependen del nivel.
        """)

    paso3, paso4 = st.columns(2)
    with paso3:
        st.markdown("""
        #### 3️⃣ Ingrese los tránsitos diarios
        En **🚢 Esclusajes**, ajuste:
        - **Panamax (PNX) / día**.
        - **NeoPanamax (NPX) / día**.

        El total aparece en el encabezado con **3 cifras significativas** y también se muestra separado por tipo de tránsito.
        """)
    with paso4:
        st.markdown("""
        #### 4️⃣ Defina cómo calcular esclusajes
        Revise dos controles distintos:
        - **Fuente vol/tránsito:** Nivel, LakeHouse o Manual. Por defecto queda **Basado en LakeHouse**.
        - **Usar en el balance:** Manual, Sidebar + ahorro, Modelo físico base o Modelo físico + ahorro.

        La app muestra PNX y NPX en hm³/esclusaje con equivalentes p³/s y m³/s. En el visor también se presenta EED como referencia operativa diaria.

        **Importante:** la primera opción define el volumen unitario; la segunda decide qué consumo entra al balance principal.
        """)

    paso5, paso6 = st.columns(2)
    with paso5:
        st.markdown("""
        #### 5️⃣ Ajuste las demás demandas
        Revise en el sidebar:
        - Generación Madden y Gatún. Madden queda por defecto con **Tabla de agua**.
        - Potabilización y fugas.
        - Vertidos, ZZ-Flush y evaporación.
        - Ahorros por tinas, cámara corta, crossfilling y Turn Around NPX.

        En evaporación, la fuente recomendada por defecto es **Aquarius · lámina (mm/día)**; las áreas se calculan con curva **Daily** y nivel operativo.
        """)
    with paso6:
        st.markdown("""
        #### 6️⃣ Valide y exporte
        Revise en este orden:
        1. **📊 Balance**: total y distribución por embalse.
        2. **🏔️ Alhajuela** y **🌊 Gatún**: componentes individuales.
        3. **🚢 Esclusajes**: tránsitos, volumen unitario, consumo y EED.
        4. **📈 Aportes observados**: último dato y promedios de Aquarius.
        5. **📤 Exportar**: descargue el Excel final.
        """)

    st.markdown("### 🎛️ Qué modifica cada control principal")
    controles = pd.DataFrame([
        {"Control": "Período 1/5/7/10/30 días", "Dónde": "Parte superior · Variables operativas por embalse", "Efecto": "Cambia los promedios LakeHouse mostrados arriba y los valores iniciales del sidebar."},
        {"Control": "Fuente de valores App/LakeHouse", "Dónde": "Parte superior · Variables operativas por embalse", "Efecto": "Permite ver el cálculo actual, el promedio LakeHouse o ambos sin modificar el balance."},
        {"Control": "Nivel Gatún", "Dónde": "Sidebar · Niveles Operativos", "Efecto": "Actualiza área Gatún, evaporación y consumos dependientes del nivel."},
        {"Control": "Nivel Alhajuela", "Dónde": "Sidebar · Niveles Operativos", "Efecto": "Actualiza área Alhajuela, evaporación y generación Madden dependiente del nivel."},
        {"Control": "Curva hipsométrica Daily", "Dónde": "Sidebar · Evaporación · Área espejo", "Efecto": "Queda activa por defecto en Gatún y Alhajuela; calcula área con tabla Daily por nivel."},
        {"Control": "Área calculada desde nivel", "Dónde": "Sidebar · Evaporación · Área espejo", "Efecto": "Queda activa por defecto; usa el nivel operativo para el área espejo. Manual queda como respaldo."},
        {"Control": "Panamax/NeoPanamax por día", "Dónde": "Sidebar · Esclusajes", "Efecto": "Cambia la cantidad de tránsitos y el consumo diario de esclusajes."},
        {"Control": "Fuente vol/tránsito", "Dónde": "Sidebar · Consumo por esclusaje", "Efecto": "Define el hm³ por tránsito PNX y NPX."},
        {"Control": "Usar en el balance", "Dónde": "Sidebar · Fuente de consumo", "Efecto": "Define si el balance usa valores del sidebar o del modelo físico, con o sin ahorro."},
        {"Control": "Fuente de evaporación", "Dónde": "Sidebar · Evaporación", "Efecto": "Por defecto usa Aquarius · lámina (mm/día). Manual y volumen V Evap 0.85 quedan como alternativas."},
        {"Control": "Ver aportes obs.", "Dónde": "Visor principal superior", "Efecto": "Muestra arriba el visor compacto de aportes observados de Aquarius para Gatún, Alhajuela y total."},
        {"Control": "Unidad visual", "Dónde": "Final del sidebar", "Efecto": "Solo cambia la presentación entre hm³/día, cfs y m³/s; no altera el cálculo base."},
    ])
    st.dataframe(controles, use_container_width=True, hide_index=True)

    st.markdown("### ✅ Estado actual del cálculo")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f"""
        <div class="lkh-card">
            <div class="label">LakeHouse</div>
            <div class="value">{_dias_lkh_seguros()} días</div>
            <div class="sub">Período operativo base guardado en la sesión.</div>
        </div>
        """, unsafe_allow_html=True)
    with c2:
        st.markdown(f"""
        <div class="lkh-card">
            <div class="label">Tránsitos diarios</div>
            <div class="value">{fmt_sig(n_t, 3)}</div>
            <div class="sub">Panamax {fmt_sig(n_pnx, 3)} · NeoPanamax {fmt_sig(n_npx, 3)}.</div>
        </div>
        """, unsafe_allow_html=True)
    with c3:
        st.markdown(f"""
        <div class="lkh-card">
            <div class="label">Balance esclusajes</div>
            <div class="value">{balance_escl_label}</div>
            <div class="sub">Modo que alimenta el balance principal.</div>
        </div>
        """, unsafe_allow_html=True)

    c4, c5, c6 = st.columns(3)
    with c4:
        st.markdown(f"""
        <div class="lkh-card">
            <div class="label">Evaporación Gatún</div>
            <div class="value">{evap_gat:.3f}</div>
            <div class="sub">hm³/d · {evap_gat_mm:.2f} mm/día · {evap_fuente_corta}.</div>
        </div>
        """, unsafe_allow_html=True)
    with c5:
        st.markdown(f"""
        <div class="lkh-card">
            <div class="label">Evaporación Alhajuela</div>
            <div class="value">{evap_alh:.3f}</div>
            <div class="sub">hm³/d · {evap_alh_mm:.2f} mm/día · {evap_fuente_corta}.</div>
        </div>
        """, unsafe_allow_html=True)
    with c6:
        st.markdown(f"""
        <div class="lkh-card">
            <div class="label">Unidad visual</div>
            <div class="value">{u_label}</div>
            <div class="sub">Solo cambia la forma de presentación.</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("### 🧩 Mapa rápido del dashboard")
    guia_tabs = pd.DataFrame([
        {"Pestaña": "📊 Balance", "Uso principal": "Validar demanda total, distribución por embalse y detalle operativo.", "Revise antes de reportar": "Total sistema, Alhajuela, Gatún y fuente de cada componente."},
        {"Pestaña": "🏔️ Alhajuela", "Uso principal": "Revisar salidas Madden/Alhajuela.", "Revise antes de reportar": "Generación Madden, potable, fugas, vertido fondo/tambor/libre y evaporación."},
        {"Pestaña": "🌊 Gatún", "Uso principal": "Revisar salidas de Gatún.", "Revise antes de reportar": "Esclusajes, generación Gatún, potable, fugas, vertido, ZZ-Flush y evaporación."},
        {"Pestaña": "🚢 Esclusajes", "Uso principal": "Comparar tránsitos Panamax/NeoPanamax y consumos unitarios.", "Revise antes de reportar": "Cantidad por tipo, total, fuente del consumo, hm³/escl, cfs equivalente y ahorro aplicado."},
        {"Pestaña": "⚡ Generación", "Uso principal": "Validar MW y conversión a caudal/volumen.", "Revise antes de reportar": "Factor cfs/MW, MW por planta y hm³/día."},
        {"Pestaña": "💾 Ahorro de Agua", "Uso principal": "Evaluar tinas, cámara corta, crossfilling y Turn Around NPX.", "Revise antes de reportar": "Ahorro total y modo de balance que lo está usando."},
        {"Pestaña": "📐 Área Espejo", "Uso principal": "Validar área por nivel y evaporación.", "Revise antes de reportar": "Daily por defecto, cálculo desde nivel, área km² y hm³/d aplicado al balance."},
        {"Pestaña": "📂 Datos Lake House", "Uso principal": "Confirmar datos reales recientes.", "Revise antes de reportar": "Archivo, hoja, fecha final, promedios 1/5/7/10/30 días y salidas por embalse."},
        {"Pestaña": "📤 Exportar", "Uso principal": "Generar respaldo en Excel.", "Revise antes de reportar": "Parámetros, demandas por embalse y hoja de áreas."},
        {"Pestaña": "📈 Aportes observados", "Uso principal": "Ver aportes diarios observados de Aquarius.", "Revise antes de reportar": "Último dato, promedio del período LakeHouse/30 días y coherencia Gatún + Alhajuela."},
    ])
    st.dataframe(guia_tabs, use_container_width=True, hide_index=True)

    st.markdown("### 📌 Reglas técnicas que debe recordar")
    reglas = pd.DataFrame([
        {"Elemento": "Niveles", "Criterio": "Se toman del último registro válido del LakeHouse y luego quedan editables en el sidebar."},
        {"Elemento": "Promedios 1/5/7/10/30 días", "Criterio": "Se calculan con los últimos N registros disponibles, no por ventana inclusiva de calendario."},
        {"Elemento": "Tránsitos", "Criterio": "PNX y NPX se promedian entre complejos para evitar duplicar el tránsito; el total se muestra con 3 cifras significativas y equivalentes operativos."},
        {"Elemento": "Potabilización y fugas", "Criterio": "Se priorizan columnas MCF/MPC del LakeHouse; *_hm3 se usa como respaldo."},
        {"Elemento": "Vertidos", "Criterio": "`madspill` y `gatspill` del LakeHouse se convierten de MPC/MCF por día a cfs y cargan los controles editables de Madden y Gatún."},
        {"Elemento": "EED", "Criterio": "Se muestra como equivalente diario de consumo. Es una referencia visual; no cambia el balance base."},
        {"Elemento": "Área espejo", "Criterio": "Por defecto usa curva hipsométrica Daily y área calculada desde nivel para Gatún y Alhajuela. Manual queda como respaldo."},
        {"Elemento": "Evaporación", "Criterio": "Por defecto usa Aquarius lámina (CZL/PMG) con área Daily calculada desde nivel: hm³/d = mm/d × área(km²) × 0.001 × 0.85. Aquarius volumen usa V Evap 0.85 directo."},
        {"Elemento": "Aportes observados", "Criterio": "El visor compacto se activa con Ver aportes obs.; la pestaña 📈 Aportes observados muestra series, último dato y promedios en hm³/d, m³/s y p³/s."},
        {"Elemento": "Conversor de volumen", "Criterio": "Permite hm³, MPC y Mgal. Mgal corresponde a millones de galones US."},
        {"Elemento": "Unidad visual", "Criterio": "Cambiar hm³/día, cfs o m³/s solo cambia la visualización; el cálculo base queda en hm³/día."},
    ])
    st.dataframe(reglas, use_container_width=True, hide_index=True)

    st.markdown("### 🧪 Lista de verificación antes de enviar")
    chk1, chk2, chk3 = st.columns(3)
    with chk1:
        st.markdown("""
        **Datos**
        - LakeHouse correcto.
        - Hoja correcta.
        - Último registro revisado.
        - Período 1/5/7/10/30 días confirmado.
        """)
    with chk2:
        st.markdown("""
        **Escenario**
        - Niveles revisados.
        - Panamax y NeoPanamax revisados.
        - Generación, potable y fugas revisadas.
        - Vertidos, ZZ-Flush y ahorros revisados.
        """)
    with chk3:
        st.markdown("""
        **Resultado**
        - Total de tránsitos correcto.
        - Balance total razonable.
        - Embalses separados.
        - Evaporación validada.
        - Área Daily desde nivel confirmada.
        - Aportes observados revisados si aplica.
        - Excel exportado.
        """)

    st.markdown("### 🛠️ Problemas comunes y solución rápida")
    problemas = pd.DataFrame([
        {"Situación": "No carga el LakeHouse", "Solución": "Verifique que sea .xlsx, que no esté abierto en Excel y que la hoja tenga columna de fecha."},
        {"Situación": "El sidebar cambió al abrir la app", "Solución": "La app toma LakeHouse como punto inicial. Después de cargarlo, todos los campos quedan editables."},
        {"Situación": "Cambio 1/5/7/10/30 días y niveles no cambian", "Solución": "Es normal: los niveles usan el último registro; los promedios operativos sí cambian con 1/5/7/10/30 días."},
        {"Situación": "El total de tránsitos no es el esperado", "Solución": "Revise por separado Panamax (PNX) y NeoPanamax (NPX) en el sidebar. El total es la suma de ambos."},
        {"Situación": "Potable o fugas se ven altos", "Solución": "Revise columnas MCF/MPC (`munic_*`, `leak_*`). La app las prioriza sobre *_hm3."},
        {"Situación": "Vertido fondo Madden no coincide", "Solución": "Verifique `madspill`: debe estar en MPC/MCF por día; la app lo convierte directamente a hm³/día."},
        {"Situación": "Evaporación no coincide", "Solución": "Revise la fuente activa, la curva Daily y que el área esté en Calcular desde nivel (ft). Manual y Aquarius mm aplican lámina × área × 0.001 × 0.85; Aquarius volumen usa V Evap 0.85 de GAT/MAD directamente."},
        {"Situación": "El área aparece diferente a la esperada", "Solución": "Confirme que la curva esté en Daily y que el modo sea Calcular desde nivel (ft). El área cambia automáticamente con el nivel operativo."},
        {"Situación": "No veo los aportes observados", "Solución": "Active Ver aportes obs. en el visor superior o entre a la pestaña 📈 Aportes observados; verifique los CSV Discharge_AT_GAT/ALHA."},
    ])
    st.dataframe(problemas, use_container_width=True, hide_index=True)

    st.success("Ruta recomendada: Datos LakeHouse → Niveles y área Daily → Tránsitos → Fuente de esclusajes → Evaporación → Aportes observados → Balance → Exportar.")


# ═══ TAB 11 — APORTES OBSERVADOS ═══
with tabs[11]:
    st.subheader("📈 Aportes observados — demandas")
    st.caption(
        "Carga automática desde `Discharge_AT_GAT_Diario.csv` y `Discharge_AT_ALHA_Diario.csv` "
        "ubicados en la carpeta `data` o junto al app. Los valores originales se leen en m³/s; el visor presenta hm³/d con conversión a m³/s y p³/s."
    )

    _aportes_tab = _cargar_aportes_observados()
    _gat_info = _aportes_tab.get("Gatún", {})
    _alh_info = _aportes_tab.get("Alhajuela", {})
    _gat_df = _gat_info.get("df", pd.DataFrame()).copy()
    _alh_df = _alh_info.get("df", pd.DataFrame()).copy()

    _aport_m1, _aport_m2, _aport_m3 = st.columns(3)
    _sum_gat = _resumen_aportes_df(_gat_df, dias_prom=int(st.session_state.get("dias_op", 5) or 5))
    _sum_alh = _resumen_aportes_df(_alh_df, dias_prom=int(st.session_state.get("dias_op", 5) or 5))
    with _aport_m1:
        if _sum_gat:
            st.metric(
                "Gatún último",
                f"{_sum_gat['ultimo_hm3']:.2f} hm³/d",
                delta=f"{_sum_gat['ultimo_m3s']:.2f} m³/s · {_sum_gat['ultimo_cfs']:,.0f} p³/s · prom {_sum_gat['n']}d {_sum_gat['prom_hm3']:.2f} hm³/d",
                delta_color="off",
            )
        else:
            st.metric("Gatún último", "N/D")
    with _aport_m2:
        if _sum_alh:
            st.metric(
                "Alhajuela último",
                f"{_sum_alh['ultimo_hm3']:.2f} hm³/d",
                delta=f"{_sum_alh['ultimo_m3s']:.2f} m³/s · {_sum_alh['ultimo_cfs']:,.0f} p³/s · prom {_sum_alh['n']}d {_sum_alh['prom_hm3']:.2f} hm³/d",
                delta_color="off",
            )
        else:
            st.metric("Alhajuela último", "N/D")
    with _aport_m3:
        if _sum_gat and _sum_alh:
            _total_cfs_u = _sum_gat["ultimo_cfs"] + _sum_alh["ultimo_cfs"]
            _total_hm3_u = _sum_gat["ultimo_hm3"] + _sum_alh["ultimo_hm3"]
            _total_m3s_u = _sum_gat["ultimo_m3s"] + _sum_alh["ultimo_m3s"]
            _total_prom_hm3 = _sum_gat["prom_hm3"] + _sum_alh["prom_hm3"]
            st.metric(
                "Total observado último",
                f"{_total_hm3_u:.2f} hm³/d",
                delta=f"{_total_m3s_u:.2f} m³/s · {_total_cfs_u:,.0f} p³/s · prom {_sum_gat['n']}d {_total_prom_hm3:.2f} hm³/d",
                delta_color="off",
            )
        else:
            st.metric("Total observado último", "N/D")

    _estado_archivos = []
    for _emb, _info in _aportes_tab.items():
        _estado_archivos.append({
            "Embalse": _emb,
            "Archivo": _info.get("archivo", ""),
            "Estado": "OK" if _info.get("ok") else f"N/D: {_info.get('error', '')}",
            "Registros": int(len(_info.get("df", pd.DataFrame()))),
        })
    st.dataframe(pd.DataFrame(_estado_archivos), use_container_width=True, hide_index=True)

    _hay_aportes = (not _gat_df.empty) or (not _alh_df.empty)
    if _hay_aportes:
        _ventana = st.radio(
            "Ventana gráfica",
            ["30 días", "90 días", "180 días", "Todo"],
            index=0,
            horizontal=True,
            key="ventana_aportes_observados",
        )

        _dfs = []
        if not _gat_df.empty:
            _d = _gat_df.copy(); _d["Embalse"] = "Gatún"; _dfs.append(_d)
        if not _alh_df.empty:
            _d = _alh_df.copy(); _d["Embalse"] = "Alhajuela"; _dfs.append(_d)
        _long = pd.concat(_dfs, ignore_index=True) if _dfs else pd.DataFrame()
        if _ventana != "Todo" and not _long.empty:
            _n_dias_graf = int(_ventana.split()[0])
            _fecha_max = _long["fecha"].max()
            _long = _long.loc[_long["fecha"] >= (_fecha_max - pd.Timedelta(days=_n_dias_graf))].copy()

        fig_ap = go.Figure()
        for _emb in ["Gatún", "Alhajuela"]:
            _de = _long.loc[_long["Embalse"].eq(_emb)].sort_values("fecha")
            if not _de.empty:
                fig_ap.add_trace(go.Scatter(
                    x=_de["fecha"], y=_de["hm3_d"], mode="lines+markers", name=_emb,
                    customdata=np.stack([_de["cfs"], _de["m3s"]], axis=-1),
                    hovertemplate="%{x|%d/%m/%Y}<br>%{y:.2f} hm³/d<br>%{customdata[1]:.2f} m³/s<br>%{customdata[0]:,.0f} p³/s<extra>" + _emb + "</extra>",
                ))

        if not _gat_df.empty and not _alh_df.empty:
            _mg = _gat_df[["fecha", "cfs", "hm3_d", "m3s"]].rename(columns={"cfs": "Gatún p³/s", "hm3_d": "Gatún hm³/d", "m3s": "Gatún m³/s"})
            _ma = _alh_df[["fecha", "cfs", "hm3_d", "m3s"]].rename(columns={"cfs": "Alhajuela p³/s", "hm3_d": "Alhajuela hm³/d", "m3s": "Alhajuela m³/s"})
            _mt = pd.merge(_mg, _ma, on="fecha", how="outer").sort_values("fecha")
            _mt["Total p³/s"] = _mt["Gatún p³/s"].fillna(0) + _mt["Alhajuela p³/s"].fillna(0)
            _mt["Total hm³/d"] = _mt["Gatún hm³/d"].fillna(0) + _mt["Alhajuela hm³/d"].fillna(0)
            _mt["Total m³/s"] = _mt["Gatún m³/s"].fillna(0) + _mt["Alhajuela m³/s"].fillna(0)
            if _ventana != "Todo":
                _mt = _mt.loc[_mt["fecha"] >= (_mt["fecha"].max() - pd.Timedelta(days=int(_ventana.split()[0])))].copy()
            fig_ap.add_trace(go.Scatter(
                x=_mt["fecha"], y=_mt["Total hm³/d"], mode="lines", name="Total",
                line=dict(dash="dash"),
                customdata=np.stack([_mt["Total p³/s"], _mt["Total m³/s"]], axis=-1),
                hovertemplate="%{x|%d/%m/%Y}<br>%{y:.2f} hm³/d<br>%{customdata[1]:.2f} m³/s<br>%{customdata[0]:,.0f} p³/s<extra>Total</extra>",
            ))

        fig_ap.update_layout(
            title="Aportes observados diarios — Aquarius Discharge AT",
            xaxis_title="Fecha",
            yaxis_title="hm³/d",
            template="plotly_white",
            height=480,
            margin=dict(l=50, r=20, t=70, b=50),
        )
        st.plotly_chart(fig_ap, use_container_width=True)

        # Tabla compacta para revisión operativa.
        if not _gat_df.empty or not _alh_df.empty:
            if not _gat_df.empty and not _alh_df.empty:
                _tabla = _mt.copy()
            elif not _gat_df.empty:
                _tabla = _gat_df[["fecha", "hm3_d", "cfs", "m3s"]].rename(columns={"hm3_d": "Gatún hm³/d", "cfs": "Gatún p³/s", "m3s": "Gatún m³/s"})
            else:
                _tabla = _alh_df[["fecha", "hm3_d", "cfs", "m3s"]].rename(columns={"hm3_d": "Alhajuela hm³/d", "cfs": "Alhajuela p³/s", "m3s": "Alhajuela m³/s"})
            _tabla = _tabla.sort_values("fecha").tail(30).copy()
            _tabla["Fecha"] = pd.to_datetime(_tabla["fecha"]).dt.strftime("%d/%m/%Y")
            _cols = ["Fecha"] + [c for c in _tabla.columns if c not in ("fecha", "Fecha")]
            st.markdown("#### Visor compacto — últimos 30 registros")
            st.dataframe(_tabla[_cols].round(3), use_container_width=True, hide_index=True)
    else:
        st.info("No se encontraron archivos de aportes observados. Coloque los CSV en `data` o junto al app.")

# ═══ FOOTER ═══
st.markdown("---")
ftr_c1, ftr_c2, ftr_c3 = st.columns([1, 6, 1])
with ftr_c1:
    _cp_f = _img_tag(_logo_cp_mime, _logo_cp, "width:65px;opacity:0.75;")
    if _cp_f:
        st.markdown(_cp_f, unsafe_allow_html=True)
with ftr_c2:
    st.markdown(
        "<div style='color:#aab7b8;font-size:0.85rem;padding-top:6px;text-align:center;'>"
        "💧 Demandas de Agua · Canal de Panamá · ACP · Elaborado para: HIMH — Sección de Hidrología<br>"
        f"Por: JFRodriguez · Sesión: {AHORA}</div>",
        unsafe_allow_html=True)
with ftr_c3:
    _himh_f = _img_tag(_logo_mime, _logo, "width:48px;opacity:0.75;float:right;")
    if _himh_f:
        st.markdown(_himh_f, unsafe_allow_html=True)
