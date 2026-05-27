"""
Calculadora Solar NOAA — Streamlit App  v2
Ecuaciones: NOAA/ESRL Solar Calculator (Spencer 1971, Iqbal 1983, Michalsky 1988)
Modos: Diario (24 h) | Anual (365 días)
Extras: mapa interactivo Folium, tabla en español, guía metodológica.
"""

import math
import json
import os
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from datetime import date, timedelta

# (sin dependencias de mapa externas)
_FOLIUM_OK = False

# ══════════════════════════════════════════════════════════════════════
#  1. MOTOR DE CÁLCULO SOLAR
# ══════════════════════════════════════════════════════════════════════

def _d2r(d): return d * math.pi / 180.0
def _r2d(r): return r * 180.0 / math.pi


def julian_day(year, month, day, hour_frac):
    if month <= 2:
        year -= 1; month += 12
    A = int(year / 100)
    B = 2 - A + int(A / 4)
    jd = int(365.25 * (year + 4716)) + int(30.6001 * (month + 1)) + day + B - 1524.5
    return jd + hour_frac / 24.0


def solar_position(year, month, day, hour_frac, lat_deg, lon_deg, tz_hrs):
    """Calcula todos los parámetros del NOAA Solar Calculator para un instante."""
    r = {}
    r["julian_day"]     = julian_day(year, month, day, hour_frac)
    r["julian_century"] = (r["julian_day"] - 2451545.0) / 36525.0
    T = r["julian_century"]

    # Geometría orbital
    L0 = (280.46646 + T * (36000.76983 + T * 0.0003032)) % 360
    r["geom_mean_long_sun"] = L0

    M_deg = (357.52911 + T * (35999.05029 - T * 0.0001537)) % 360
    r["geom_mean_anom_sun"] = M_deg
    M = _d2r(M_deg)

    e = 0.016708634 - T * (0.000042037 + T * 0.0000001267)
    r["eccent_earth_orbit"] = e

    C = (  math.sin(M)   * (1.914602 - T * (0.004817 + T * 0.000014))
         + math.sin(2*M) * (0.019993 - T * 0.000101)
         + math.sin(3*M) *  0.000289)
    r["sun_eq_of_ctr"]  = C
    r["sun_true_long"]  = L0 + C
    r["sun_true_anom"]  = M_deg + C
    r["sun_rad_vector"] = (1.000001018 * (1 - e**2)
                           / (1 + e * math.cos(_d2r(r["sun_true_anom"]))))

    sun_app_long = (r["sun_true_long"] - 0.00569
                    - 0.00478 * math.sin(_d2r(125.04 - 1934.136 * T)))
    r["sun_app_long"] = sun_app_long

    # Oblicuidad
    mean_obliq = 23 + (26 + (21.448 - T * (46.8150 + T * (0.00059 - T * 0.001813))) / 60) / 60
    r["mean_obliq_ecliptic"] = mean_obliq
    obliq_corr = mean_obliq + 0.00256 * math.cos(_d2r(125.04 - 1934.136 * T))
    r["obliq_corr"] = obliq_corr

    # Posición solar
    lam = _d2r(sun_app_long)
    eps = _d2r(obliq_corr)
    r["sun_rt_ascen"] = _r2d(math.atan2(math.cos(eps) * math.sin(lam), math.cos(lam)))
    r["sun_declin"]   = _r2d(math.asin(math.sin(eps) * math.sin(lam)))

    # Ecuación del tiempo
    var_y = math.tan(_d2r(obliq_corr / 2)) ** 2
    r["var_y"] = var_y
    L0r = _d2r(L0)
    r["eq_of_time"] = 4 * _r2d(
        var_y * math.sin(2 * L0r)
      - 2 * e * math.sin(M)
      + 4 * e * var_y * math.sin(M) * math.cos(2 * L0r)
      - 0.5 * var_y**2 * math.sin(4 * L0r)
      - 1.25 * e**2 * math.sin(2 * M))

    # Amanecer / atardecer
    lat    = _d2r(lat_deg)
    declin = _d2r(r["sun_declin"])
    cos_ha = (math.cos(_d2r(90.833)) / (math.cos(lat) * math.cos(declin))
              - math.tan(lat) * math.tan(declin))
    cos_ha = max(-1.0, min(1.0, cos_ha))
    r["ha_sunrise"]        = _r2d(math.acos(cos_ha))
    r["solar_noon"]        = (720 - 4 * lon_deg - r["eq_of_time"] + tz_hrs * 60) / 1440
    r["sunrise"]           = r["solar_noon"] - r["ha_sunrise"] * 4 / 1440
    r["sunset"]            = r["solar_noon"] + r["ha_sunrise"] * 4 / 1440
    r["sunlight_duration"] = 8 * r["ha_sunrise"]

    # Tiempo solar verdadero y ángulo horario
    tst = (hour_frac * 60 + r["eq_of_time"] + 4 * lon_deg - 60 * tz_hrs) % 1440
    r["true_solar_time"] = tst
    r["hour_angle"]      = tst / 4 + 180 if tst < 0 else tst / 4 - 180
    ha = r["hour_angle"]

    # Ángulos solares
    cos_sza = max(-1.0, min(1.0,
        math.sin(lat) * math.sin(declin)
      + math.cos(lat) * math.cos(declin) * math.cos(_d2r(ha))))
    sza = _r2d(math.acos(cos_sza))
    r["solar_zenith"]    = sza
    r["solar_elevation"] = 90.0 - sza

    # Refracción atmosférica
    el = r["solar_elevation"]
    if   el > 85:
        refr = 0.0
    elif el > 5:
        refr = (58.1 / math.tan(_d2r(el))
                - 0.07  / math.tan(_d2r(el))**3
                + 0.000086 / math.tan(_d2r(el))**5) / 3600
    elif el > -0.575:
        refr = (1735 + el * (-518.2 + el * (103.4 + el * (-12.79 + el * 0.711)))) / 3600
    else:
        refr = (-20.772 / math.tan(_d2r(el))) / 3600
    r["atm_refraction"]       = refr
    r["solar_elevation_corr"] = el + refr

    # Azimut
    sin_sza = math.sin(_d2r(sza))
    if abs(sin_sza) < 1e-10:
        r["solar_azimuth"] = 0.0
    elif ha > 0:
        r["solar_azimuth"] = (_r2d(math.acos(
            max(-1.0, min(1.0,
                (math.sin(lat) * cos_sza - math.sin(declin))
                / (math.cos(lat) * sin_sza)))
        )) + 180) % 360
    else:
        r["solar_azimuth"] = (540 - _r2d(math.acos(
            max(-1.0, min(1.0,
                (math.sin(lat) * cos_sza - math.sin(declin))
                / (math.cos(lat) * sin_sza)))
        ))) % 360
    return r


def day_series(year, month, day, lat, lon, tz, n_pts=241):
    hours = np.linspace(0, 24, n_pts, endpoint=False)
    rows  = []
    for h in hours:
        r = solar_position(year, month, day, h, lat, lon, tz)
        r["time_hrs"] = h
        rows.append(r)
    return pd.DataFrame(rows)


def year_series(year, hour_frac, lat, lon, tz):
    start  = date(year, 1, 1)
    n_days = 366 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 365
    rows   = []
    for d in range(n_days):
        dt = start + timedelta(days=d)
        r  = solar_position(dt.year, dt.month, dt.day, hour_frac, lat, lon, tz)
        r["date"] = dt
        r["doy"]  = d + 1
        rows.append(r)
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════
#  2. HELPERS Y CONSTANTES
# ══════════════════════════════════════════════════════════════════════

def frac_to_hm(frac):
    try:
        if math.isnan(frac) or not (0 <= frac <= 1):
            return "N/A"
    except TypeError:
        return "N/A"
    total_min = round(frac * 1440)
    h, m = divmod(total_min, 60)
    return f"{h:02d}:{m:02d}"


def hrs_to_hm(hrs):
    h = int(hrs); m = int(round((hrs - h) * 60))
    if m == 60: h += 1; m = 0
    return f"{h:02d}:{m:02d}"


def frac_to_hms(frac):
    """Fracción de día → HH:MM:SS (precisión de segundos)."""
    try:
        if math.isnan(frac) or not (0 <= frac <= 1): return "N/A"
    except TypeError:
        return "N/A"
    total_sec = round(frac * 86400)
    h, rem = divmod(total_sec, 3600)
    m, s   = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"

def dur_to_hms(minutes):
    """Minutos decimales → H:MM:SS (precisión de segundos)."""
    if not minutes: return ""
    total_sec = round(float(minutes) * 60)
    h, rem = divmod(total_sec, 3600)
    m, s   = divmod(rem, 60)
    return f"{h}:{m:02d}:{s:02d}"


# Mapa de nombres español (clave interna → etiqueta)
ES = {
    "date"                : "Fecha",
    "doy"                 : "Día del Año",
    "time_hrs"            : "Hora Local (h)",
    "julian_day"          : "Día Juliano",
    "julian_century"      : "Siglo Juliano",
    "geom_mean_long_sun"  : "Long. Media Geom. del Sol (°)",
    "geom_mean_anom_sun"  : "Anom. Media Geom. del Sol (°)",
    "eccent_earth_orbit"  : "Excentricidad de la Órbita",
    "sun_eq_of_ctr"       : "Ec. del Centro Solar",
    "sun_true_long"       : "Long. Verdadera del Sol (°)",
    "sun_true_anom"       : "Anom. Verdadera del Sol (°)",
    "sun_rad_vector"      : "Vector Radio Solar (UA)",
    "sun_app_long"        : "Long. Aparente del Sol (°)",
    "mean_obliq_ecliptic" : "Oblicuidad Media Eclíptica (°)",
    "obliq_corr"          : "Oblicuidad Corregida (°)",
    "sun_rt_ascen"        : "Ascensión Recta del Sol (°)",
    "sun_declin"          : "Declinación Solar (°)",
    "var_y"               : "Variable y",
    "eq_of_time"          : "Ecuación del Tiempo (min)",
    "ha_sunrise"          : "Áng. Horario Amanecer (°)",
    "solar_noon"          : "Mediodía Solar (HH:MM)",
    "sunrise"             : "Amanecer (HH:MM)",
    "sunset"              : "Atardecer (HH:MM)",
    "sunlight_duration"   : "Duración del Día (min)",
    "true_solar_time"     : "Tiempo Solar Verdadero (min)",
    "hour_angle"          : "Ángulo Horario (°)",
    "solar_zenith"        : "Ángulo Cenital Solar (°)",
    "solar_elevation"     : "Elevación Solar (°)",
    "atm_refraction"      : "Refracción Atmosférica (°)",
    "solar_elevation_corr": "Elevación Solar Corregida (°)",
    "solar_azimuth"       : "Azimut Solar (° hor. desde N)",
}

PALETTE = {
    "elevation": "#F5A623", "azimuth": "#4A90D9",
    "declin"   : "#7ED321", "sunrise": "#FF6B35",
    "zenith"   : "#9B59B6", "noon"   : "#E74C3C",
    "duration" : "#1ABC9C",
}

HORA_COLS = {ES["solar_noon"], ES["sunrise"], ES["sunset"]}

# Orden completo de columnas (igual al XLS NOAA)
ORDEN_NOAA = [
    "julian_day", "julian_century",
    "geom_mean_long_sun", "geom_mean_anom_sun", "eccent_earth_orbit",
    "sun_eq_of_ctr", "sun_true_long", "sun_true_anom", "sun_rad_vector",
    "sun_app_long", "mean_obliq_ecliptic", "obliq_corr",
    "sun_rt_ascen", "sun_declin", "var_y", "eq_of_time",
    "ha_sunrise", "solar_noon", "sunrise", "sunset", "sunlight_duration",
    "true_solar_time", "hour_angle", "solar_zenith", "solar_elevation",
    "atm_refraction", "solar_elevation_corr", "solar_azimuth",
]


def build_noaa_table(df, first_cols):
    """Construye DataFrame con columnas en español en orden NOAA."""
    out = df[first_cols].copy()
    for key in ORDEN_NOAA:
        if key in ("solar_noon", "sunrise", "sunset"):
            out[ES[key]] = df[key].apply(frac_to_hm)
        else:
            out[ES[key]] = df[key]
    out = out.rename(columns={c: ES.get(c, c) for c in first_cols})
    return out



# ══════════════════════════════════════════════════════════════════════
#  FORMATO ACP — Tabla Salida/Puesta del Sol
# ══════════════════════════════════════════════════════════════════════

MESES_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
            "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

MESES_ABREV = ["Ene","Feb","Mar","Abr","May","Jun",
               "Jul","Ago","Sep","Oct","Nov","Dic"]


def dur_to_hm(minutes):
    """Convierte minutos decimales → string 'H:MM' (ej. 11:47)."""
    if not minutes:
        return ""
    total = round(float(minutes))
    h, m  = divmod(total, 60)
    return f"{h}:{m:02d}"


def build_acp_table(year, lat, lon, tz):
    """Tabla 31×12: filas=días, columnas MultiIndex mes×(Salida,Puesta).
    Retorna (df_srss, df_dur):
      - df_srss : tabla Salida/Puesta (formato HH:MM)
      - df_dur  : tabla Duración del día (formato H:MM)
    """
    start  = date(year, 1, 1)
    n_days = 366 if (year%4==0 and (year%100!=0 or year%400==0)) else 365

    data_sp  = {(m, s): [""] * 31
                for m in MESES_ES for s in ("Salida", "Puesta")}
    data_dur = {m: [""] * 31 for m in MESES_ES}

    for d in range(n_days):
        dt  = start + timedelta(days=d)
        r   = solar_position(dt.year, dt.month, dt.day, 0, lat, lon, tz)
        mes = MESES_ES[dt.month - 1]
        idx = dt.day - 1
        data_sp[(mes, "Salida")][idx] = frac_to_hm(r["sunrise"])
        data_sp[(mes, "Puesta")][idx] = frac_to_hm(r["sunset"])
        data_dur[mes][idx]            = dur_to_hm(r["sunlight_duration"])

    cols_sp = pd.MultiIndex.from_product([MESES_ES, ["Salida", "Puesta"]])
    df_srss = pd.DataFrame(data_sp, columns=cols_sp)
    df_srss.index = range(1, 32)
    df_srss.index.name = "Día"

    df_dur = pd.DataFrame(data_dur, columns=MESES_ES)
    df_dur.index = range(1, 32)
    df_dur.index.name = "Día"

    return df_srss, df_dur


def lat_to_str(lat):
    """8.9500 → 'N 8° 57\'' """
    hemi = "N" if lat >= 0 else "S"
    lat  = abs(lat)
    deg  = int(lat)
    mins = round((lat - deg) * 60)
    if mins == 60:
        deg += 1; mins = 0
    return f"{hemi} {deg}° {mins}\'"


def lon_to_str(lon):
    """−79.5500 → 'O 79° 33\''"""
    hemi = "E" if lon >= 0 else "O"
    lon  = abs(lon)
    deg  = int(lon)
    mins = round((lon - deg) * 60)
    if mins == 60:
        deg += 1; mins = 0
    return f"{hemi} {deg}° {mins}\'"


def export_acp_xlsx(df_acp, df_dur, year, city_name, lat, lon, tz):
    """Genera bytes de un .xlsx con formato idéntico al PDF de la ACP."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import (Font, Alignment, PatternFill,
                                 Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Sol {year}"

    # ── Estilos ──────────────────────────────────────────
    NAVY   = "003366"
    GOLD   = "C8A951"
    WHITE  = "FFFFFF"
    LGRAY  = "F2F2F2"
    DGRAY  = "CCCCCC"

    def cell_style(cell, bold=False, align="center", bg=None,
                   fg="000000", size=10, border=False):
        cell.font      = Font(name="Arial", bold=bold, size=size, color=fg)
        cell.alignment = Alignment(horizontal=align, vertical="center",
                                   wrap_text=True)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if border:
            thin = Side(style="thin", color=DGRAY)
            cell.border = Border(left=thin, right=thin,
                                 top=thin, bottom=thin)

    # ── Fila 1: Título ────────────────────────────────────
    ws.merge_cells("A1:Y1")
    c = ws["A1"]
    c.value = f"SALIDA Y PUESTA DEL SOL, AÑO {year}"
    cell_style(c, bold=True, size=14, bg=NAVY, fg=WHITE)
    ws.row_dimensions[1].height = 22

    # ── Fila 2: Ciudad ────────────────────────────────────
    ws.merge_cells("A2:Y2")
    c = ws["A2"]
    c.value = city_name.upper()
    cell_style(c, bold=True, size=12, bg=NAVY, fg=GOLD)
    ws.row_dimensions[2].height = 18

    # ── Fila 3: Coordenadas ───────────────────────────────
    ws.merge_cells("A3:Y3")
    c = ws["A3"]
    tz_lbl = f"{abs(tz)}h al {'Oeste' if tz < 0 else 'Este'} de Greenwich"
    c.value = (f"Localización: {lat_to_str(lat)} ,  {lon_to_str(lon)}"
               f"      Zona: {'-' if tz < 0 else '+'}{tz_lbl}")
    cell_style(c, size=9, bg=NAVY, fg="AAAAAA", align="left")
    ws.row_dimensions[3].height = 14

    # ── Fila 4: Fuente ────────────────────────────────────
    ws.merge_cells("A4:Y4")
    c = ws["A4"]
    c.value = "Fuente: Ecuaciones NOAA/ESRL Solar Calculator  ·  Spencer (1971), Iqbal (1983)"
    cell_style(c, size=8, bg="001A40", fg="888888", align="left")
    ws.row_dimensions[4].height = 12

    # ── Fila 5: espacio ───────────────────────────────────
    ws.row_dimensions[5].height = 4

    # ── Fila 6: Encabezados de mes ────────────────────────
    col_offset = 2          # columna B = col 2
    for mi, mes in enumerate(MESES_ES):
        c1 = col_offset + mi * 2          # col inicio mes
        c2 = c1 + 1                       # col fin mes
        ws.merge_cells(start_row=6, start_column=c1,
                       end_row=6,   end_column=c2)
        c = ws.cell(row=6, column=c1, value=mes)
        cell_style(c, bold=True, size=9, bg=NAVY, fg=WHITE)
    ws.row_dimensions[6].height = 16

    # ── Fila 7: Salida / Puesta ───────────────────────────
    ws.cell(row=7, column=1, value="Día")
    cell_style(ws.cell(row=7, column=1), bold=True, size=9,
               bg=GOLD, fg=NAVY)
    for mi in range(12):
        for si, sub in enumerate(("Salida", "Puesta")):
            col = col_offset + mi * 2 + si
            c   = ws.cell(row=7, column=col, value=sub)
            cell_style(c, bold=True, size=8,
                       bg="C8A951" if si==0 else "E8D090", fg=NAVY)
    ws.row_dimensions[7].height = 14

    # ── Fila 8: h:m ──────────────────────────────────────
    ws.cell(row=8, column=1, value="")
    for col in range(col_offset, col_offset + 24):
        c = ws.cell(row=8, column=col, value="h : m")
        cell_style(c, size=7, bg=LGRAY, fg="555555")
    ws.row_dimensions[8].height = 11

    # ── Filas 9-39: datos ─────────────────────────────────
    for day_idx in range(31):
        row = 9 + day_idx
        # Número de día
        c = ws.cell(row=row, column=1, value=day_idx + 1)
        cell_style(c, bold=True, size=9,
                   bg=GOLD if (day_idx % 2 == 0) else "E8D090",
                   fg=NAVY, border=True)
        # Datos
        row_bg = LGRAY if day_idx % 2 == 0 else WHITE
        for mi, mes in enumerate(MESES_ES):
            for si, sub in enumerate(("Salida", "Puesta")):
                col  = col_offset + mi * 2 + si
                val  = df_acp.iloc[day_idx][(mes, sub)]
                c    = ws.cell(row=row, column=col, value=val)
                cell_style(c, size=9, bg=row_bg, border=True)
        ws.row_dimensions[row].height = 13

    # ── Anchos de columna ─────────────────────────────────
    ws.column_dimensions["A"].width = 4.5       # Día
    for mi in range(12):
        for si in range(2):
            col_ltr = get_column_letter(col_offset + mi * 2 + si)
            ws.column_dimensions[col_ltr].width = 6.5

    # ── Congelar encabezados ──────────────────────────────
    ws.freeze_panes = "B9"

    # ── Hoja 2: Duración del Día ─────────────────────────
    ws2 = wb.create_sheet(title=f"Duración {year}")

    # Título
    ws2.merge_cells(f"A1:{get_column_letter(13)}1")
    c = ws2["A1"]
    c.value = f"DURACIÓN DEL DÍA SOLAR, AÑO {year}"
    cell_style(c, bold=True, size=13, bg=NAVY, fg=WHITE)
    ws2.row_dimensions[1].height = 20

    # Ciudad + coords
    ws2.merge_cells(f"A2:{get_column_letter(13)}2")
    c = ws2["A2"]
    c.value = f"{city_name.upper()}    ·    {lat_to_str(lat)}, {lon_to_str(lon)}    ·    UTC{tz:+d}"
    cell_style(c, bold=False, size=9, bg=NAVY, fg=GOLD, align="left")
    ws2.row_dimensions[2].height = 13

    # Fuente
    ws2.merge_cells(f"A3:{get_column_letter(13)}3")
    c = ws2["A3"]
    c.value = "Fuente: Ecuaciones NOAA/ESRL Solar Calculator  ·  Spencer (1971), Iqbal (1983)"
    cell_style(c, size=7, bg="001A40", fg="888888", align="left")
    ws2.row_dimensions[3].height = 11

    ws2.row_dimensions[4].height = 4  # espacio

    # Encabezado de meses (fila 5)
    ws2.cell(row=5, column=1, value="Día")
    cell_style(ws2.cell(row=5, column=1), bold=True, size=9, bg=GOLD, fg=NAVY)
    for mi, mes in enumerate(MESES_ES):
        c = ws2.cell(row=5, column=mi+2, value=mes)
        cell_style(c, bold=True, size=9, bg=NAVY, fg=WHITE)
        ws2.column_dimensions[get_column_letter(mi+2)].width = 7.5
    ws2.row_dimensions[5].height = 15
    ws2.column_dimensions["A"].width = 4.5

    # Sub-encabezado h:mm (fila 6)
    ws2.cell(row=6, column=1, value="")
    for col in range(2, 14):
        c = ws2.cell(row=6, column=col, value="h:mm")
        cell_style(c, size=7, bg=LGRAY, fg="555555")
    ws2.row_dimensions[6].height = 10

    # Datos (filas 7-37)
    for day_idx in range(31):
        row = 7 + day_idx
        c = ws2.cell(row=row, column=1, value=day_idx+1)
        cell_style(c, bold=True, size=9,
                   bg=GOLD if day_idx%2==0 else "E8D090", fg=NAVY, border=True)
        row_bg = LGRAY if day_idx%2==0 else WHITE
        for mi, mes in enumerate(MESES_ES):
            val = df_dur.iloc[day_idx][mes] if day_idx < len(df_dur) else ""
            c   = ws2.cell(row=row, column=mi+2, value=val)
            cell_style(c, size=9, bg=row_bg, border=True)
        ws2.row_dimensions[row].height = 13

    ws2.freeze_panes = "B7"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════
#  3. PESTAÑA DE METODOLOGÍA
# ══════════════════════════════════════════════════════════════════════

def show_explanation():
    st.markdown("## 📖 Metodología — Calculadora Solar NOAA")
    st.markdown("""
Esta aplicación reproduce fielmente las ecuaciones publicadas por el
**NOAA Global Monitoring Laboratory** (GML/ESRL) en su
[Solar Calculator](https://gml.noaa.gov/grad/solcalc/).
La cadena de cálculo sigue a **Spencer (1971)**, **Iqbal (1983)** y
**Michalsky (1988)**, con un error menor a **±0.01°** para fechas entre 2000 y 2100.
""")

    st.markdown("### 🔢 Cadena de ecuaciones")
    steps = [
        ("**1. Día Juliano (JD)**",
         r"Convierte año/mes/día/hora a un número continuo de días desde el 1 enero 4713 a.C. "
         r"Incluye la fracción de día para la hora local."),
        ("**2. Siglo Juliano (T)**",
         r"T = (JD − 2 451 545) / 36 525  — normaliza el tiempo respecto al J2000.0."),
        ("**3. Longitud media geométrica del Sol (L₀)**",
         r"L₀ = 280.46646 + 36 000.76983·T + 0.0003032·T²  (°, módulo 360)."),
        ("**4. Anomalía media geométrica del Sol (M)**",
         r"M = 357.52911 + 35 999.05029·T − 0.0001537·T²  (°)."),
        ("**5. Excentricidad de la órbita terrestre (e)**",
         r"e = 0.016 708 634 − 0.000 042 037·T − 1.267·10⁻⁷·T²."),
        ("**6. Ecuación del centro solar (C)**",
         r"C = (1.914602−0.004817T−0.000014T²)sin M + (0.019993−0.000101T)sin 2M + 0.000289 sin 3M."),
        ("**7. Longitud verdadera del Sol (Θ) y Longitud aparente (λ)**",
         r"Θ = L₀ + C.  λ = Θ − 0.00569 − 0.00478·sin(125.04 − 1934.136·T).  "
         r"La corrección de la longitud aparente incorpora la *aberración* y la *nutación*."),
        ("**8. Oblicuidad de la eclíptica (ε)**",
         r"ε₀ = 23° 26' 21.448'' − 46.8150''·T − ...  "
         r"ε = ε₀ + 0.00256·cos(125.04 − 1934.136·T)."),
        ("**9. Declinación solar (δ)**",
         r"δ = arcsin(sin ε · sin λ).  Varía entre ±23.44° a lo largo del año."),
        ("**10. Ecuación del tiempo (E)**",
         r"E = 4·[y·sin 2L₀ − 2e·sin M + 4ey·sin M·cos 2L₀ − ½y²·sin 4L₀ − 1.25e²·sin 2M]  (min).  "
         r"Diferencia entre tiempo solar medio y tiempo solar verdadero."),
        ("**11. Ángulo horario al amanecer (ω₀)**",
         r"cos ω₀ = cos 90.833° / (cos φ · cos δ) − tan φ · tan δ.  "
         r"Los 90.833° incluyen la corrección por refracción al horizonte (~0.833°)."),
        ("**12. Mediodía Solar, Amanecer, Atardecer**",
         r"Mediodía = (720 − 4·Lon − E + TZ·60) / 1440 (fracción de día).  "
         r"Amanecer = Mediodía − ω₀·4/1440.  Atardecer = Mediodía + ω₀·4/1440."),
        ("**13. Tiempo Solar Verdadero (TSV)**",
         r"TSV = hora_local·60 + E + 4·Lon − 60·TZ  (min, módulo 1440)."),
        ("**14. Ángulo horario (H)**",
         r"H = TSV/4 − 180  si TSV ≥ 0;  H = TSV/4 + 180  si TSV < 0  (°)."),
        ("**15. Ángulo cenital (θ_z) y Elevación solar (α)**",
         r"cos θ_z = sin φ·sin δ + cos φ·cos δ·cos H.  α = 90° − θ_z."),
        ("**16. Refracción atmosférica**",
         r"Corrección de Bennett (1982) dependiente del ángulo de elevación: "
         r"5 regímenes desde elevación >85° hasta <−0.575°."),
        ("**17. Azimut solar (A)**",
         r"A = [arccos((sin φ·cos θ_z − sin δ)/(cos φ·sin θ_z)) + 180] mod 360  si H>0;  "
         r"A = [540 − arccos(...)] mod 360  si H≤0.  Medido en grados horarios desde el Norte."),
    ]

    for titulo, texto in steps:
        with st.expander(titulo, expanded=False):
            st.markdown(texto)

    st.markdown("### 📐 Parámetros de entrada")
    st.markdown("""
| Parámetro | Descripción |
|-----------|-------------|
| **Latitud** | Grados decimales. Norte (+), Sur (−). |
| **Longitud** | Grados decimales. Este (+), Oeste (−). |
| **Zona horaria** | Horas de desfase respecto a UTC. Panamá = −5. |
| **Fecha** (modo Diario) | Cualquier fecha entre 1900 y 2100. |
| **Año + Hora** (modo Anual) | Año completo; hora local fija para cada día. |
""")

    st.markdown("### 📏 Precisión y limitaciones")
    st.info(
        "**Precisión angular:** ±0.01° para 2000–2100, hasta ±0.01° fuera de ese rango.  \n"
        "**Zona polar:** cuando el Sol no sale o no se pone (latitudes extremas), "
        "el ángulo horario al amanecer se trunca a ±90° y los tiempos se marcan como N/A.  \n"
        "**Refracción:** la corrección de Bennett es una aproximación; el error es <0.1' "
        "para elevaciones >5° y aumenta hacia el horizonte.  \n"
        "**Ecuaciones:** no incluyen perturbaciones planetarias ni efectos de mareas; "
        "son suficientes para todas las aplicaciones de ingeniería solar."
    )

    st.markdown("### 🔗 Referencias")
    st.markdown("""
- Michalsky, J.J. (1988). *The Astronomical Almanac's algorithm for approximate solar position (1950–2050)*. Solar Energy, 40(3), 227–235.
- Spencer, J.W. (1971). *Fourier series representation of the position of the Sun*. Search, 2(5), 172.
- Iqbal, M. (1983). *An Introduction to Solar Radiation*. Academic Press, Toronto.
- Meeus, J. (1998). *Astronomical Algorithms*. 2nd ed. Willmann-Bell.
- NOAA/GML Solar Calculator: [gml.noaa.gov/grad/solcalc](https://gml.noaa.gov/grad/solcalc/)
""")


# ══════════════════════════════════════════════════════════════════════
#  4. CONFIGURACIÓN DE PÁGINA
# ══════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Calculadora Solar NOAA",
    page_icon="☀️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
/* Visores diarios — fondo claro, más grandes y legibles */
.metric-box{
    background: linear-gradient(180deg, #FFF7E6 0%, #FFFFFF 100%);
    border-radius: 20px;
    padding: 28px 24px;
    text-align: center;
    border-left: 8px solid #F5A623;
    border-top: 1px solid rgba(245,166,35,0.35);
    border-right: 1px solid rgba(245,166,35,0.18);
    border-bottom: 1px solid rgba(245,166,35,0.18);
    min-height: 155px;
    display: flex;
    flex-direction: column;
    justify-content: center;
    box-shadow: 0 10px 24px rgba(15, 23, 42, 0.12);
}
.metric-val{
    font-size: 2.35rem;
    font-weight: 900;
    color: #D97706;
    line-height: 1.05;
    letter-spacing: 0.02em;
}
.metric-lbl{
    font-size: 0.95rem;
    color: #334155;
    font-weight: 700;
    margin-top: 12px;
}
/* KPI anual — fondo claro y tamaño reforzado */
.kpi-anual{
    background: linear-gradient(180deg, #F8FAFC 0%, #FFFFFF 100%);
    border-radius: 20px;
    padding: 28px 22px 22px 22px;
    text-align: center;
    border-top: 8px solid #F5A623;
    border-left: 1px solid rgba(148,163,184,0.35);
    border-right: 1px solid rgba(148,163,184,0.35);
    border-bottom: 1px solid rgba(148,163,184,0.35);
    min-height: 155px;
    box-shadow: 0 10px 24px rgba(15, 23, 42, 0.12);
}
.kpi-main{
    font-size: 2.35rem;
    font-weight: 900;
    line-height: 1.05;
}
.kpi-sub {
    font-size: 1.20rem;
    font-weight: 800;
    margin-top: 8px;
}
.kpi-lbl {
    font-size: 0.92rem;
    color: #334155;
    font-weight: 700;
    margin-top: 12px;
    letter-spacing: 0.03em;
}
</style>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
#  5. SESSION STATE
# ══════════════════════════════════════════════════════════════════════

if "lat" not in st.session_state: st.session_state["lat"] =  8.9500
if "lon" not in st.session_state: st.session_state["lon"] = -79.5500

# ── Contador de visitas (persiste en archivo JSON) ────────────────
_COUNTER_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "visit_counter.json")

def _load_visits():
    try:
        with open(_COUNTER_FILE, "r") as _f:
            return json.load(_f).get("visits", 0)
    except Exception:
        return 0

def _save_visits(n):
    try:
        with open(_COUNTER_FILE, "w") as _f:
            json.dump({"visits": n}, _f)
    except Exception:
        pass

if "visit_counted" not in st.session_state:
    st.session_state["visit_counted"] = True
    _total_visits = _load_visits() + 1
    _save_visits(_total_visits)
else:
    _total_visits = _load_visits()

# ══════════════════════════════════════════════════════════════════════
#  6. SIDEBAR
# ══════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.title("☀️ Calculadora Solar")
    st.caption("NOAA/ESRL · Spencer 1971 · Iqbal 1983")
    st.divider()

    mode = st.radio("**Modo de cálculo**",
                    ["📅 Diario", "📆 Anual"],
                    help="**Diario:** perfil de 24 h para una fecha dada.\n"
                         "**Anual:** 365 días del año a una hora fija.")
    st.divider()

    st.markdown("**📍 Coordenadas**")
    ca, cb = st.columns(2)
    with ca:
        lat = st.number_input("Latitud (°N)",
                              value=st.session_state["lat"],
                              min_value=-90.0, max_value=90.0,
                              step=0.0001, format="%.4f", key="lat_input")
    with cb:
        lon = st.number_input("Longitud (°E)",
                              value=st.session_state["lon"],
                              min_value=-180.0, max_value=180.0,
                              step=0.0001, format="%.4f", key="lon_input")
    # Sincronizar session state con inputs manuales
    st.session_state["lat"] = lat
    st.session_state["lon"] = lon

    tz = st.number_input("Zona horaria (h desde UTC)",
                         value=-5, min_value=-12, max_value=14, step=1)
    st.divider()

    if mode == "📅 Diario":
        st.markdown("**📅 Fecha**")
        sel_date = st.date_input("Fecha", value=date.today())
        year = sel_date.year; month = sel_date.month; day = sel_date.day
    else:
        st.markdown("**📆 Año y hora fija**")
        year = int(st.number_input("Año", value=date.today().year,
                                   min_value=1900, max_value=2100, step=1))
        from datetime import time as _time
        _sel_time  = st.time_input(
            "Hora local (HH:MM)",
            value=_time(12, 0),
            step=60,          # paso = 1 minuto
            help="Hora del día para la cual se calcula la posición solar cada día del año.")
        local_hour = _sel_time.hour + _sel_time.minute / 60.0

    st.divider()
    st.caption("🔗 [NOAA Solar Calculator](https://gml.noaa.gov/grad/solcalc/)  \n"
               "Precisión ±0.01° · 2000–2100")


# ══════════════════════════════════════════════════════════════════════
#  7. MODO DIARIO
# ══════════════════════════════════════════════════════════════════════

if mode == "📅 Diario":
    st.title(f"☀️ Perfil Solar Diario — {sel_date.strftime('%d %b %Y')}")
    st.caption(f"Lat: **{lat:.4f}°**  ·  Lon: **{lon:.4f}°**  ·  TZ: **UTC{tz:+d}**")

    with st.spinner("Calculando…"):
        df_d = day_series(year, month, day, lat, lon, tz)

    noon_frac = df_d["solar_noon"].iloc[0]
    sr_frac   = df_d["sunrise"].iloc[0]
    ss_frac   = df_d["sunset"].iloc[0]
    dur_min   = df_d["sunlight_duration"].iloc[0]
    max_el    = df_d["solar_elevation_corr"].max()
    decl_deg  = df_d["sun_declin"].iloc[0]

    kpi_data = [
        ("🌅 Amanecer",        frac_to_hm(sr_frac),   PALETTE["sunrise"]),
        ("🌇 Atardecer",       frac_to_hm(ss_frac),   PALETTE["sunrise"]),
        ("⏱ Horas de Sol",     f"{dur_min/60:.2f} h", PALETTE["duration"]),
        ("🕛 Mediodía Solar",   frac_to_hm(noon_frac), PALETTE["noon"]),        
        ("🔆 Elevación Máx.",  f"{max_el:.1f}°",       PALETTE["elevation"]),
        ("🌐 Declinación",     f"{decl_deg:.3f}°",     PALETTE["declin"]),
    ]
    for col, (lbl, val, color) in zip(st.columns(6), kpi_data):
        with col:
            st.markdown(
                f'<div class="metric-box" style="border-left-color:{color}">'
                f'<div class="metric-val">{val}</div>'
                f'<div class="metric-lbl">{lbl}</div></div>',
                unsafe_allow_html=True)

    st.markdown("---")
    tab_d1, tab_d2, tab_d3, tab_d4 = st.tabs(
        ["📈 Trayectoria", "🧭 Diagrama Polar", "📋 Tabla Completa", "📖 Metodología"])

    # ── Trayectoria ─────────────────────────────────────
    with tab_d1:
        fig = go.Figure()
        df_sun = df_d[df_d["solar_elevation_corr"] > 0]
        fig.add_trace(go.Scatter(
            x=df_sun["time_hrs"], y=df_sun["solar_elevation_corr"],
            fill="tozeroy", fillcolor="rgba(245,166,35,0.15)",
            line=dict(color=PALETTE["elevation"], width=2.5),
            name="Elevación corregida"))
        fig.add_trace(go.Scatter(
            x=df_d["time_hrs"], y=df_d["solar_elevation"],
            line=dict(color=PALETTE["elevation"], width=1, dash="dot"),
            name="Sin refracción"))
        for xv, lbl, c in [(sr_frac*24,   "🌅 Amanecer",  PALETTE["sunrise"]),
                           (noon_frac*24, "🕛 Mediodía",  PALETTE["noon"]),
                           (ss_frac*24,   "🌇 Atardecer", PALETTE["sunrise"])]:
            fig.add_vline(x=xv, line_dash="dash", line_color=c,
                          annotation_text=lbl, annotation_position="top")
        fig.add_hline(y=0, line_color="gray", line_width=0.8)
        fig.update_layout(
            title=f"Elevación Solar — {sel_date.strftime('%d %b %Y')}",
            xaxis_title="Hora Local (h)", yaxis_title="Elevación (°)",
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            template="plotly_dark", height=420,
            xaxis=dict(tickmode="linear", tick0=0, dtick=2))
        st.plotly_chart(fig, use_container_width=True)

        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(
            x=df_d["time_hrs"], y=df_d["solar_zenith"],
            line=dict(color=PALETTE["zenith"], width=2), name="Ángulo Cenital"))
        fig2.add_hline(y=90, line_color="gray", line_dash="dash",
                       annotation_text="Horizonte (90°)")
        fig2.update_layout(
            title="Ángulo Cenital Solar",
            xaxis_title="Hora Local (h)", yaxis_title="Ángulo Cenital (°)",
            template="plotly_dark", height=280,
            xaxis=dict(tickmode="linear", tick0=0, dtick=2))
        st.plotly_chart(fig2, use_container_width=True)

    # ── Diagrama polar ───────────────────────────────────
    with tab_d2:
        df_pos = df_d[df_d["solar_elevation_corr"] > 0].copy()
        fig_pol = go.Figure()
        fig_pol.add_trace(go.Scatterpolar(
            r=90 - df_pos["solar_elevation_corr"],
            theta=df_pos["solar_azimuth"],
            mode="lines+markers",
            marker=dict(size=5, color=df_pos["time_hrs"],
                        colorscale="Plasma", showscale=True,
                        colorbar=dict(title="Hora (h)", thickness=12)),
            line=dict(color="rgba(245,166,35,0.5)", width=2)))
        fig_pol.update_layout(
            title="Diagrama Polar — Trayectoria Solar",
            polar=dict(
                radialaxis=dict(range=[0,90], tickvals=[0,30,60,90],
                                ticktext=["90°","60°","30°","0°"]),
                angularaxis=dict(direction="clockwise",
                                 tickvals=[0,45,90,135,180,225,270,315],
                                 ticktext=["N","NE","E","SE","S","SO","O","NO"])),
            template="plotly_dark", height=500)
        st.plotly_chart(fig_pol, use_container_width=True)

        fig_az = go.Figure()
        fig_az.add_trace(go.Scatter(
            x=df_pos["time_hrs"], y=df_pos["solar_azimuth"],
            line=dict(color=PALETTE["azimuth"], width=2)))
        fig_az.update_layout(
            title="Azimut Solar (° horario desde Norte)",
            xaxis_title="Hora Local (h)", yaxis_title="Azimut (°)",
            template="plotly_dark", height=280,
            xaxis=dict(tickmode="linear", tick0=0, dtick=2))
        st.plotly_chart(fig_az, use_container_width=True)

    # ── Tabla completa diario ────────────────────────────
    with tab_d3:
        df_td = build_noaa_table(df_d, ["time_hrs"])
        num_d = [c for c in df_td.columns
                 if c not in HORA_COLS and df_td[c].dtype == float]
        st.dataframe(
            df_td.style.format({c: "{:.6f}" for c in num_d}),
            use_container_width=True, height=520)
        st.download_button(
            "⬇️ Descargar CSV (formato NOAA)",
            data=df_td.to_csv(index=False).encode(),
            file_name=f"NOAA_Solar_day_{sel_date}.csv",
            mime="text/csv")

    with tab_d4:
        show_explanation()


# ══════════════════════════════════════════════════════════════════════
#  8. MODO ANUAL
# ══════════════════════════════════════════════════════════════════════

else:
    st.title(f"☀️ Ciclo Solar Anual — {year}")
    st.caption(f"Hora fija: **{hrs_to_hm(local_hour)}**  ·  "
               f"Lat: **{lat:.4f}°**  ·  Lon: **{lon:.4f}°**  ·  TZ: **UTC{tz:+d}**")

    with st.spinner("Calculando 365 días…"):
        df_y = year_series(year, local_hour, lat, lon, tz)

    idx_max_dur = df_y["sunlight_duration"].idxmax()
    idx_min_dur = df_y["sunlight_duration"].idxmin()
    idx_max_el  = df_y["solar_elevation_corr"].idxmax()

    # ── KPI anuales: valor principal + sub-valor + etiqueta ─────────
    # Duraciones con precisión de segundos
    _dur_max_hms   = dur_to_hms(df_y.loc[idx_max_dur, "sunlight_duration"])
    _dur_min_hms   = dur_to_hms(df_y.loc[idx_min_dur, "sunlight_duration"])
    _night_max_hms = dur_to_hms(1440 - df_y.loc[idx_min_dur, "sunlight_duration"])
    _night_min_hms = dur_to_hms(1440 - df_y.loc[idx_max_dur, "sunlight_duration"])
    _sr_max = frac_to_hms(df_y.loc[idx_max_dur, "sunrise"])
    _ss_max = frac_to_hms(df_y.loc[idx_max_dur, "sunset"])
    _sr_min = frac_to_hms(df_y.loc[idx_min_dur, "sunrise"])
    _ss_min = frac_to_hms(df_y.loc[idx_min_dur, "sunset"])

    # Fila 1: Día más largo / corto / Noche más larga / corta / Elevación
    _kpi_anual = [
        {
            "lbl"  : "☀️ Día más largo",
            "main" : _dur_max_hms,
            "sub"  : df_y.loc[idx_max_dur,"date"].strftime("%d %b"),
            "extra": f"↑ {_sr_max}  ↓ {_ss_max}",
            "color": "#00C9A7",
        },
        {
            "lbl"  : "🌑 Día más corto",
            "main" : _dur_min_hms,
            "sub"  : df_y.loc[idx_min_dur,"date"].strftime("%d %b"),
            "extra": f"↑ {_sr_min}  ↓ {_ss_min}",
            "color": "#4A90D9",
        },
        {
            "lbl"  : "🌙 Noche más larga",
            "main" : _night_max_hms,
            "sub"  : df_y.loc[idx_min_dur,"date"].strftime("%d %b"),
            "extra": "= 24 h − día más corto",
            "color": "#9B59B6",
        },
        {
            "lbl"  : "🌟 Noche más corta",
            "main" : _night_min_hms,
            "sub"  : df_y.loc[idx_max_dur,"date"].strftime("%d %b"),
            "extra": "= 24 h − día más largo",
            "color": "#F5A623",
        },
        {
            "lbl"  : "🔆 Elevación máx.",
            "main" : f"{df_y.loc[idx_max_el,'solar_elevation_corr']:.2f}°",
            "sub"  : df_y.loc[idx_max_el,"date"].strftime("%d %b"),
            "extra": "",
            "color": "#E74C3C",
        },
    ]
    for col, k in zip(st.columns(5), _kpi_anual):
        with col:
            _extra_html = (f'<div style="font-size:0.62rem;color:{k["color"]};'
                           f'opacity:0.85;margin-top:4px;">{k["extra"]}</div>'
                           if k.get("extra") else "")
            st.markdown(
                f"""<div class="kpi-anual" style="border-top-color:{k['color']}">
  <div class="kpi-main" style="color:{k['color']};font-size:1.55rem;">{k['main']}</div>
  <div class="kpi-sub"  style="color:{k['color']}bb;">{k['sub']}</div>
  <div class="kpi-lbl">{k['lbl']}</div>
  {_extra_html}
</div>""",
                unsafe_allow_html=True)



    # ── Eventos Astronómicos ─────────────────────────────────────────
    _idx_sol_v = df_y['sun_declin'].idxmax()
    _idx_sol_i = df_y['sun_declin'].idxmin()
    _idx_peri  = df_y['sun_rad_vector'].idxmin()
    _idx_afe   = df_y['sun_rad_vector'].idxmax()

    # Equinoccios: días donde la declinación cruza 0°
    _eq_prim = _eq_oton = None
    for _ei in range(1, len(df_y)):
        _dp = df_y.iloc[_ei-1]['sun_declin']
        _dc = df_y.iloc[_ei  ]['sun_declin']
        if _dp < 0 <= _dc and _eq_prim is None:
            _eq_prim = df_y.iloc[_ei]['date']
        elif _dp > 0 >= _dc and _eq_oton is None:
            _eq_oton = df_y.iloc[_ei]['date']

    # Ordenar todos los eventos por fecha para presentarlos cronológicamente
    _astro_events = sorted([
        {
            "icon" : "🔴",
            "name" : "Perihelio",
            "desc" : "Tierra más cerca del Sol",
            "date" : df_y.loc[_idx_peri, 'date'],
            "extra": f"R = {df_y.loc[_idx_peri,'sun_rad_vector']:.5f} UA",
            "color": "#FF6B35",
        },
        {
            "icon" : "🌸",
            "name" : "Equinoccio de Primavera",
            "desc" : "Día y noche iguales · Sol en Ecuador",
            "date" : _eq_prim,
            "extra": "δ ≈ 0° · Sol sale exactamente por el Este",
            "color": "#7ED321",
        },
        {
            "icon" : "☀️",
            "name" : "Solsticio de Verano",
            "desc" : "Día más largo del hemisferio N",
            "date" : df_y.loc[_idx_sol_v, 'date'],
            "extra": f"δ = {df_y.loc[_idx_sol_v,'sun_declin']:.3f}° (máx.)",
            "color": "#F5A623",
        },
        {
            "icon" : "🔵",
            "name" : "Afelio",
            "desc" : "Tierra más lejos del Sol",
            "date" : df_y.loc[_idx_afe, 'date'],
            "extra": f"R = {df_y.loc[_idx_afe,'sun_rad_vector']:.5f} UA",
            "color": "#4A90D9",
        },
        {
            "icon" : "🍂",
            "name" : "Equinoccio de Otoño",
            "desc" : "Día y noche iguales · Sol en Ecuador",
            "date" : _eq_oton,
            "extra": "δ ≈ 0° · Sol sale exactamente por el Este",
            "color": "#C8A951",
        },
        {
            "icon" : "❄️",
            "name" : "Solsticio de Invierno",
            "desc" : "Día más corto del hemisferio N",
            "date" : df_y.loc[_idx_sol_i, 'date'],
            "extra": f"δ = {df_y.loc[_idx_sol_i,'sun_declin']:.3f}° (mín.)",
            "color": "#9B59B6",
        },
    ], key=lambda x: x["date"])

    st.markdown(
        f"<div style='font-size:0.85rem;font-weight:600;color:#888;"
        f"letter-spacing:0.08em;text-transform:uppercase;"
        f"margin-bottom:6px;'>📅 Eventos Astronómicos {year}</div>",
        unsafe_allow_html=True)

    _ev_cols = st.columns(6)
    for _col, _ev in zip(_ev_cols, _astro_events):
        _dt_str = _ev["date"].strftime("%d %b") if _ev["date"] else "N/D"
        _yr_str = _ev["date"].strftime("%Y")     if _ev["date"] else ""
        with _col:
            st.markdown(f"""
<div class="kpi-anual" style="border-top-color:{_ev['color']};min-height:160px;">
  <div style="font-size:1.65rem;line-height:1.1;">{_ev['icon']}</div>
  <div class="kpi-main" style="color:{_ev['color']};">{_dt_str}</div>
  <div class="kpi-sub"  style="color:{_ev['color']}bb;">{_yr_str}</div>
  <div class="kpi-lbl">{_ev['name']}</div>
  <div style="font-size:0.64rem;color:{_ev['color']};margin-top:4px;line-height:1.35;">{_ev['extra']}</div>
</div>""", unsafe_allow_html=True)

    st.markdown("---")
    tab_y1, tab_y2, tab_y3, tab_y4, tab_y5, tab_y6, tab_y7 = st.tabs(
        ["📈 Elevación & Declinación", "🌅 Amanecer / Atardecer",
         "⏱ Horas de Sol", "📋 Tabla Completa", "📖 Metodología",
         "📄 Formato ACP", "🔭 Gráficas Especiales"])

    # ── Elevación & Declinación ──────────────────────────
    with tab_y1:
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=df_y["date"], y=df_y["solar_elevation_corr"],
            line=dict(color=PALETTE["elevation"], width=2.5),
            name=f"Elevación @ {hrs_to_hm(local_hour)}"))
        fig.add_trace(go.Scatter(
            x=df_y["date"], y=df_y["sun_declin"],
            line=dict(color=PALETTE["declin"], width=2, dash="dash"),
            name="Declinación", yaxis="y2"))
        for m2, d2, lbl2 in [(3,20,"Equinoccio"),(6,21,"Solsticio"),
                              (9,22,"Equinoccio"),(12,21,"Solsticio")]:
            try:
                x_ts = pd.Timestamp(date(year, m2, d2))
                fig.add_vline(x=x_ts.value, line_dash="dot", line_color="#555")
                fig.add_annotation(x=x_ts, y=1, yref="paper", text=lbl2,
                                   showarrow=False,
                                   font=dict(size=9, color="#aaa"),
                                   textangle=-90, xanchor="left")
            except ValueError:
                pass
        fig.add_hline(y=0, line_color="#444", line_width=0.8)
        fig.update_layout(
            title=f"Elevación Solar y Declinación — {year}",
            xaxis_title="Fecha", yaxis_title="Elevación (°)",
            yaxis2=dict(title="Declinación (°)", overlaying="y",
                        side="right", showgrid=False),
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            template="plotly_dark", height=430)
        st.plotly_chart(fig, use_container_width=True)

        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(
            x=df_y["date"], y=df_y["hour_angle"],
            line=dict(color=PALETTE["azimuth"], width=1.8),
            name="Ángulo Horario"))
        fig2.add_hline(y=0, line_color="#444", line_width=0.8)
        fig2.update_layout(
            title="Ángulo Horario Solar",
            xaxis_title="Fecha", yaxis_title="Ángulo Horario (°)",
            template="plotly_dark", height=270)
        st.plotly_chart(fig2, use_container_width=True)

    # ── Amanecer / Atardecer ─────────────────────────────
    with tab_y2:
        sr_hrs = df_y["sunrise"]    * 24
        ss_hrs = df_y["sunset"]     * 24
        sn_hrs = df_y["solar_noon"] * 24
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=df_y["date"], y=ss_hrs,
            fill="tonexty", line=dict(color="rgba(0,0,0,0)"), showlegend=False))
        fig.add_trace(go.Scatter(x=df_y["date"], y=sr_hrs,
            fill="tozeroy", line=dict(color="rgba(0,0,0,0)"),
            fillcolor="rgba(245,166,35,0.18)", showlegend=False))
        fig.add_trace(go.Scatter(x=df_y["date"], y=sr_hrs,
            line=dict(color=PALETTE["sunrise"], width=2), name="Amanecer"))
        fig.add_trace(go.Scatter(x=df_y["date"], y=ss_hrs,
            line=dict(color="#FF8C00", width=2), name="Atardecer"))
        fig.add_trace(go.Scatter(x=df_y["date"], y=sn_hrs,
            line=dict(color=PALETTE["noon"], width=1.5, dash="dash"),
            name="Mediodía Solar"))
        fig.update_layout(
            title=f"Amanecer / Atardecer / Mediodía Solar — {year}",
            xaxis_title="Fecha", yaxis_title="Hora Local",
            yaxis=dict(tickvals=list(range(0,25,2)),
                       ticktext=[f"{h:02d}:00" for h in range(0,25,2)]),
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            template="plotly_dark", height=430)
        st.plotly_chart(fig, use_container_width=True)

        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(x=df_y["date"], y=df_y["eq_of_time"],
            line=dict(color="#E91E63", width=2),
            fill="tozeroy", fillcolor="rgba(233,30,99,0.12)",
            name="Ecuación del Tiempo"))
        fig3.add_hline(y=0, line_color="#444")
        fig3.update_layout(title="Ecuación del Tiempo",
            xaxis_title="Fecha", yaxis_title="Minutos",
            template="plotly_dark", height=270)
        st.plotly_chart(fig3, use_container_width=True)

    # ── Horas de Sol ─────────────────────────────────────
    with tab_y3:
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=df_y["date"], y=df_y["sunlight_duration"] / 60,
            fill="tozeroy", fillcolor="rgba(26,188,156,0.2)",
            line=dict(color=PALETTE["duration"], width=2.5),
            name="Horas de Sol"))
        fig.add_hline(y=12, line_color="#555", line_dash="dot",
                      annotation_text="12 h", annotation_position="right")
        fig.update_layout(title=f"Duración del Día Solar — {year}",
            xaxis_title="Fecha", yaxis_title="Horas de Sol (h)",
            template="plotly_dark", height=400)
        st.plotly_chart(fig, use_container_width=True)

        fig4 = go.Figure()
        fig4.add_trace(go.Scatter(
            x=df_y["date"], y=df_y["sun_rad_vector"],
            line=dict(color="#BDC3C7", width=1.8),
            name="Radio Vector"))
        fig4.add_hline(y=1.0, line_color="#555", line_dash="dot",
                       annotation_text="1 UA", annotation_position="right")
        fig4.update_layout(title="Radio Vector Sol–Tierra (UA)",
            xaxis_title="Fecha", yaxis_title="UA",
            template="plotly_dark", height=260)
        st.plotly_chart(fig4, use_container_width=True)

    # ── Tabla completa anual ─────────────────────────────
    with tab_y4:
        df_ty = build_noaa_table(df_y, ["date", "doy"])
        # Insertar hora fija como segunda columna
        df_ty.insert(2, "Hora (hrs desde medianoche)", local_hour)
        num_y = [c for c in df_ty.columns
                 if c not in HORA_COLS | {"Fecha"}
                 and df_ty[c].dtype == float]
        st.dataframe(
            df_ty.style.format({c: "{:.6f}" for c in num_y}),
            use_container_width=True, height=520)
        st.download_button(
            "⬇️ Descargar CSV (formato NOAA)",
            data=df_ty.to_csv(index=False).encode(),
            file_name=f"NOAA_Solar_Calculations_year_{year}.csv",
            mime="text/csv")

    with tab_y5:
        show_explanation()

    with tab_y6:
        st.markdown("### 📄 Tabla Formato ACP — Salida y Puesta del Sol")
        st.caption(
            "Reproduce el formato de la tabla oficial publicada por la "
            "Autoridad del Canal de Panamá (ACP). "
            "Fuente de referencia: https://aa.usno.navy.mil/")

        c1, c2 = st.columns([2, 1])
        with c1:
            city_name = st.text_input(
                "Nombre de la ciudad / estación",
                value="Ciudad de Panamá, Panamá")
        with c2:
            st.markdown(f"""
**Coordenadas actuales:**  
Lat: `{lat:.4f}°` · Lon: `{lon:.4f}°` · TZ: `UTC{tz:+d}`
""")

        with st.spinner("Generando tablas…"):
            df_acp, df_dur = build_acp_table(year, lat, lon, tz)

        # ── Sub-tabs: Salida/Puesta  |  Duración ──────────
        st_acp, st_dur = st.tabs(["🌅 Salida y Puesta", "⏱ Duración del Día"])

        with st_acp:
            st.markdown("##### Salida y Puesta del Sol — vista previa")
            h1, h2 = st.columns(2)
            with h1:
                st.markdown("**Enero – Junio**")
                cols_h1 = pd.MultiIndex.from_product(
                    [MESES_ES[:6], ["Salida", "Puesta"]])
                st.dataframe(df_acp[cols_h1].replace("", "—"),
                             use_container_width=True, height=430)
            with h2:
                st.markdown("**Julio – Diciembre**")
                cols_h2 = pd.MultiIndex.from_product(
                    [MESES_ES[6:], ["Salida", "Puesta"]])
                st.dataframe(df_acp[cols_h2].replace("", "—"),
                             use_container_width=True, height=430)

        with st_dur:
            st.markdown("##### Duración del Día Solar (H:MM) — vista previa")
            st.caption("Formato H:MM — horas y minutos de luz solar por día.")

            # Tabla completa de duración
            h1d, h2d = st.columns(2)
            with h1d:
                st.markdown("**Enero – Junio**")
                st.dataframe(
                    df_dur[MESES_ES[:6]].replace("", "—"),
                    use_container_width=True, height=430)
            with h2d:
                st.markdown("**Julio – Diciembre**")
                st.dataframe(
                    df_dur[MESES_ES[6:]].replace("", "—"),
                    use_container_width=True, height=430)

            # Gráfico de duración media mensual
            st.markdown("---")
            st.markdown("**Duración media mensual**")
            def hm_to_min(s):
                try:
                    h, m = s.split(":")
                    return int(h)*60 + int(m)
                except Exception:
                    return None
            medias = []
            for mes in MESES_ES:
                vals = [hm_to_min(v) for v in df_dur[mes] if v]
                medias.append(round(sum(vals)/len(vals)/60, 2) if vals else 0)
            fig_dur = go.Figure(go.Bar(
                x=MESES_ABREV, y=medias,
                marker_color="#1ABC9C",
                text=[f"{v:.2f} h" for v in medias],
                textposition="outside"))
            fig_dur.add_hline(y=12, line_dash="dot", line_color="#555",
                              annotation_text="12 h")
            fig_dur.update_layout(
                title=f"Duración Media Mensual del Día Solar — {year}",
                xaxis_title="Mes", yaxis_title="Horas de Sol (h)",
                template="plotly_dark", height=320,
                yaxis=dict(range=[0, max(medias)*1.15]))
            st.plotly_chart(fig_dur, use_container_width=True)

        # ── Exportar ──────────────────────────────────────
        st.markdown("---")
        st.markdown("#### Exportar")
        dl1, dl2, dl3 = st.columns(3)

        # Excel con ambas hojas
        with dl1:
            xlsx_bytes = export_acp_xlsx(df_acp, df_dur, year,
                                         city_name, lat, lon, tz)
            st.download_button(
                label="⬇️ Excel completo (.xlsx)",
                data=xlsx_bytes,
                file_name=f"Salida_Puesta_Sol_{year}_{city_name.replace(' ','_').replace(',','')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
            st.caption("📋 2 hojas: Salida/Puesta + Duración del Día")

        # CSV Salida/Puesta
        with dl2:
            rows_csv = []
            for mes in MESES_ES:
                for d in range(31):
                    sr  = df_acp.iloc[d][(mes, "Salida")]
                    ss  = df_acp.iloc[d][(mes, "Puesta")]
                    dur = df_dur.iloc[d][mes]
                    if sr:
                        rows_csv.append({
                            "Año": year, "Mes": mes, "Día": d+1,
                            "Salida": sr, "Puesta": ss,
                            "Duración (H:MM)": dur})
            csv_bytes = pd.DataFrame(rows_csv).to_csv(index=False).encode()
            st.download_button(
                label="⬇️ CSV — Salida/Puesta/Duración",
                data=csv_bytes,
                file_name=f"Salida_Puesta_Duracion_{year}.csv",
                mime="text/csv",
                use_container_width=True)
            st.caption("Formato largo: Año, Mes, Día, Salida, Puesta, Duración.")

        # CSV solo duración (formato tabla)
        with dl3:
            csv_dur = df_dur.replace("", "").to_csv().encode()
            st.download_button(
                label="⬇️ CSV — Duración (tabla)",
                data=csv_dur,
                file_name=f"Duracion_Dia_Solar_{year}.csv",
                mime="text/csv",
                use_container_width=True)
            st.caption("Tabla 31×12 en H:MM — mismo formato que Salida/Puesta.")

        st.info(
            f"📍 **{city_name}** · "
            f"Lat: {lat_to_str(lat)} · Lon: {lon_to_str(lon)} · "
            f"Zona: UTC{tz:+d}  \n"
            "⚠️ Los tiempos pueden diferir ±1 min respecto a tablas USNO/ACP "
            "por diferencias de algoritmo y redondeo.", icon="ℹ️")

    # ── tab_y7: Gráficas Especiales ──────────────────────
    with tab_y7:
        st.markdown("### 🔭 Gráficas Especiales")
        st.caption(
            "Visualizaciones únicas no incluidas en las otras pestañas. "
            f"Año **{year}** · Lat: {lat:.4f}° · Lon: {lon:.4f}°")

        g1, g2, g3 = st.tabs(["🌀 Analema", "⚡ Irradiancia TOA", "🎯 Azimut al Mediodía"])

        # ─────────────────────────────────────────────────
        # G1  ANALEMA  (azimut vs elevación a hora fija)
        # ─────────────────────────────────────────────────
        with g1:
            st.markdown("#### 🌀 Analema Solar")
            st.info(
                "El **analema** es la figura que traza el Sol en el cielo si se "
                "fotografía a la **misma hora del reloj** cada día durante un año. "
                "Su forma de «8» o «lágrima» surge de dos efectos combinados:\n\n"
                "- 🔄 **Ecuación del tiempo** — la Tierra recorre su órbita elíptica "
                "a velocidad variable, haciendo que el Sol llegue antes o después.\n"
                "- 🌍 **Declinación solar** — el eje terrestre (inclinado 23.4°) hace "
                "que el Sol suba o baje con las estaciones.\n\n"
                "Cada punto representa un día del año a la hora fija seleccionada "
                "en el sidebar. El color indica el mes del año.", icon="🔭")

            df_ana = df_y[df_y["solar_elevation_corr"] > 0].copy()
            if df_ana.empty:
                st.warning("El Sol está bajo el horizonte a esta hora. "
                           "Cambia la hora local en el sidebar (prueba entre 07:00–17:00).")
            else:
                fig_ana = go.Figure()
                fig_ana.add_trace(go.Scatter(
                    x=df_ana["solar_azimuth"],
                    y=df_ana["solar_elevation_corr"],
                    mode="lines+markers",
                    marker=dict(
                        size=6,
                        color=df_ana["doy"],
                        colorscale="Plasma",
                        showscale=True,
                        colorbar=dict(
                            title="Día del Año", thickness=14,
                            tickvals=[1, 91, 182, 274, 365],
                            ticktext=["Ene", "Abr", "Jul", "Oct", "Dic"]),
                    ),
                    line=dict(color="rgba(245,166,35,0.3)", width=1),
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>"
                        "Azimut: %{x:.2f}°<br>"
                        "Elevación: %{y:.2f}°<extra></extra>"),
                    customdata=list(zip(
                        df_ana["date"].astype(str),
                        df_ana["doy"])),
                ))
                fig_ana.update_layout(
                    title=(f"Analema Solar — {year}  "
                           f"(hora fija {hrs_to_hm(local_hour)})"),
                    xaxis_title="Azimut Solar (° horario desde N)",
                    yaxis_title="Elevación Solar Corregida (°)",
                    template="plotly_dark",
                    height=500,
                )
                st.plotly_chart(fig_ana, use_container_width=True)

                ca, cb, cc = st.columns(3)
                imax_a = df_ana["solar_elevation_corr"].idxmax()
                imin_a = df_ana["solar_elevation_corr"].idxmin()
                ca.metric("Elevación máxima",
                          f"{df_ana.loc[imax_a,'solar_elevation_corr']:.2f}°",
                          df_ana.loc[imax_a,'date'].strftime('%d %b'))
                cb.metric("Elevación mínima",
                          f"{df_ana.loc[imin_a,'solar_elevation_corr']:.2f}°",
                          df_ana.loc[imin_a,'date'].strftime('%d %b'))
                cc.metric("Días con sol a esta hora", str(len(df_ana)))

        # ─────────────────────────────────────────────────
        # G2  IRRADIANCIA EXTRATERRESTRE RELATIVA  (1/R²)
        # ─────────────────────────────────────────────────
        with g2:
            st.markdown("#### ⚡ Irradiancia Solar Extraterrestre Relativa")
            st.info(
                "La **irradiancia solar en la cima de la atmósfera** (TOA) varía a lo "
                "largo del año porque la órbita terrestre es una **elipse**, no un "
                "círculo perfecto. Según la ley del cuadrado inverso:\n\n"
                "**E_TOA ∝ 1 / R²**\n\n"
                "donde R es el vector radio Sol–Tierra en UA. La Tierra se acerca más "
                "al Sol (**perihelio**, ~3 ene) y se aleja (**afelio**, ~4 jul), "
                "produciendo una variación de ≈ **±3.4 %** respecto a la constante "
                "solar (1361 W/m²). Esta variación es opuesta al ciclo estacional "
                "del hemisferio norte, que está controlado por la inclinación axial, "
                "no por la distancia.", icon="⚡")

            # Irradiancia relativa = 1/R² normalizada a la media
            r2     = df_y["sun_rad_vector"] ** 2
            e_rel  = (1.0 / r2) / (1.0 / r2).mean() * 100   # % de la media
            e_wm2  = 1361.0 / r2                              # W/m² (constante solar 1361)

            fig_toa = go.Figure()
            fig_toa.add_trace(go.Scatter(
                x=df_y["date"], y=e_wm2,
                mode="lines", fill="tozeroy",
                fillcolor="rgba(255,200,0,0.10)",
                line=dict(color="#FFD700", width=2.5),
                name="E_TOA (W/m²)",
                hovertemplate="%{x|%d %b}<br>E_TOA = %{y:.1f} W/m²<extra></extra>",
            ))
            fig_toa.add_hline(y=1361, line_dash="dot", line_color="#777",
                              annotation_text="Constante Solar 1361 W/m²",
                              annotation_position="right")
            # Marcar perihelio y afelio
            idx_peri = df_y["sun_rad_vector"].idxmin()
            idx_afe  = df_y["sun_rad_vector"].idxmax()
            for idx, sym, col, lbl in [
                (idx_peri, "star",         "#FF6B35", "Perihelio"),
                (idx_afe,  "star-open",    "#4A90D9", "Afelio"),
            ]:
                fig_toa.add_trace(go.Scatter(
                    x=[df_y.loc[idx, "date"]],
                    y=[e_wm2[idx]],
                    mode="markers+text",
                    marker=dict(size=14, color=col, symbol=sym),
                    text=[f"{lbl}<br>{df_y.loc[idx,'date'].strftime('%d %b')}<br>"
                          f"R={df_y.loc[idx,'sun_rad_vector']:.5f} UA"],
                    textposition="top center",
                    textfont=dict(color=col, size=9),
                    showlegend=False,
                ))
            fig_toa.update_layout(
                title=f"Irradiancia Solar Extraterrestre (TOA) — {year}",
                xaxis_title="Fecha",
                yaxis_title="Irradiancia (W/m²)",
                yaxis=dict(range=[1310, 1420]),
                template="plotly_dark",
                height=420,
            )
            st.plotly_chart(fig_toa, use_container_width=True)

            # Segunda gráfica: R² y variación porcentual
            fig_var = go.Figure()
            fig_var.add_trace(go.Scatter(
                x=df_y["date"], y=e_rel - 100,
                mode="lines", fill="tozeroy",
                fillcolor="rgba(255,107,53,0.12)",
                line=dict(color="#FF6B35", width=2),
                name="Variación respecto a la media (%)",
                hovertemplate="%{x|%d %b}<br>Δ = %{y:+.2f}%<extra></extra>",
            ))
            fig_var.add_hline(y=0, line_color="#555", line_width=1)
            fig_var.update_layout(
                title="Variación de Irradiancia respecto a la Media Anual (%)",
                xaxis_title="Fecha",
                yaxis_title="Variación (%)",
                template="plotly_dark", height=280,
            )
            st.plotly_chart(fig_var, use_container_width=True)

            ca, cb, cc = st.columns(3)
            ca.metric("Perihelio",
                      f"{e_wm2[idx_peri]:.1f} W/m²",
                      df_y.loc[idx_peri, 'date'].strftime('%d %b') + f"  R={df_y.loc[idx_peri,'sun_rad_vector']:.5f} UA")
            cb.metric("Afelio",
                      f"{e_wm2[idx_afe]:.1f} W/m²",
                      df_y.loc[idx_afe, 'date'].strftime('%d %b') + f"  R={df_y.loc[idx_afe,'sun_rad_vector']:.5f} UA")
            cc.metric("Variación pico a pico",
                      f"{e_wm2[idx_peri]-e_wm2[idx_afe]:.1f} W/m²",
                      f"±{(e_rel-100).abs().max():.2f}% de la media")

        # ─────────────────────────────────────────────────
        # G3  AZIMUT AL MEDIODÍA SOLAR  (pasos cenitales)
        # ─────────────────────────────────────────────────
        with g3:
            st.markdown("#### 🎯 Azimut del Sol al Mediodía Solar")
            st.info(
                "Al **mediodía solar** el Sol alcanza su punto más alto del día y "
                "cruza el meridiano del observador. La dirección (azimut) en ese "
                "momento revela si el Sol está al Norte o al Sur del cénit:\n\n"
                "- **Azimut ≈ 180°** → Sol al **Sur** (hemisferio norte normal)\n"
                "- **Azimut ≈ 0° / 360°** → Sol al **Norte** (trópico, verano boreal)\n"
                "- **Azimut = indeterminado** → Sol exactamente en el **cénit** "
                "(paso cenital: elevación = 90°)\n\n"
                f"Para esta latitud ({lat:.2f}°), cuando la declinación solar iguala "
                f"la latitud (~{lat:.1f}°), el Sol pasa por el cénit. "
                "En Panamá esto ocurre **2 veces por año** (hacia mayo y agosto), "
                "un fenómeno exclusivo de la zona tropical (|lat| ≤ 23.44°).",
                icon="🎯")

            # Calcular azimut al solar noon para cada día
            import math as _math
            noon_az = []
            noon_el = []
            for _, row in df_y.iterrows():
                r2 = solar_position(
                    row["date"].year, row["date"].month, row["date"].day,
                    row["solar_noon"] * 24,   # hora = mediodía solar exacto
                    lat, lon, tz)
                noon_az.append(r2["solar_azimuth"])
                noon_el.append(r2["solar_elevation_corr"])

            df_noon = df_y[["date","doy","sun_declin"]].copy()
            df_noon["noon_azimuth"]   = noon_az
            df_noon["noon_elevation"] = noon_el

            fig_az = go.Figure()
            # Color por declinación
            fig_az.add_trace(go.Scatter(
                x=df_noon["date"], y=df_noon["noon_azimuth"],
                mode="markers",
                marker=dict(
                    size=5,
                    color=df_noon["sun_declin"],
                    colorscale="RdBu_r",
                    cmin=-23.5, cmax=23.5,
                    showscale=True,
                    colorbar=dict(title="Declinación (°)", thickness=14),
                ),
                hovertemplate=(
                    "%{x|%d %b}<br>"
                    "Azimut mediodía: %{y:.2f}°<br>"
                    "Declinación: %{customdata:.3f}°<extra></extra>"),
                customdata=df_noon["sun_declin"],
            ))
            # Línea 180° y 0°
            fig_az.add_hline(y=180, line_color="#888", line_dash="dot",
                             annotation_text="Sur (180°)", annotation_position="right")
            fig_az.add_hline(y=0, line_color="#888", line_dash="dot",
                             annotation_text="Norte (0°/360°)", annotation_position="right")
            # Marcar pasos cenitales si |lat| ≤ 23.44
            if abs(lat) <= 23.44:
                # Buscar días donde declinación cruza la latitud
                prev_sign = None
                for _, rn in df_noon.iterrows():
                    diff = rn["sun_declin"] - lat
                    sign = 1 if diff >= 0 else -1
                    if prev_sign is not None and sign != prev_sign:
                        fig_az.add_vline(
                            x=pd.Timestamp(rn["date"]).value,
                            line_dash="dash", line_color="#F5A623", line_width=1.5)
                        fig_az.add_annotation(
                            x=pd.Timestamp(rn["date"]), y=0.5, yref="paper",
                            text=f"Paso cenital<br>{rn['date'].strftime('%d %b')}",
                            showarrow=False,
                            font=dict(size=9, color="#F5A623"),
                            textangle=-90, xanchor="left")
                    prev_sign = sign

            fig_az.update_layout(
                title=f"Azimut Solar al Mediodía Solar — {year}",
                xaxis_title="Fecha",
                yaxis_title="Azimut (°)",
                yaxis=dict(range=[-5, 365]),
                template="plotly_dark",
                height=420,
            )
            st.plotly_chart(fig_az, use_container_width=True)

            # Elevación al mediodía
            fig_el = go.Figure()
            fig_el.add_trace(go.Scatter(
                x=df_noon["date"], y=df_noon["noon_elevation"],
                mode="lines", fill="tozeroy",
                fillcolor="rgba(245,166,35,0.12)",
                line=dict(color=PALETTE["elevation"], width=2),
                hovertemplate="%{x|%d %b}<br>Elevación: %{y:.2f}°<extra></extra>",
            ))
            fig_el.add_hline(y=90, line_color="#F5A623", line_dash="dot",
                             annotation_text="Cénit (90°)", annotation_position="right")
            fig_el.update_layout(
                title="Elevación Solar al Mediodía Solar (altura máxima diaria)",
                xaxis_title="Fecha", yaxis_title="Elevación (°)",
                template="plotly_dark", height=280,
            )
            st.plotly_chart(fig_el, use_container_width=True)

            imax_n = df_noon["noon_elevation"].idxmax()
            imin_n = df_noon["noon_elevation"].idxmin()
            ca2, cb2, cc2 = st.columns(3)
            ca2.metric("Elevación máx. al mediodía",
                       f"{df_noon.loc[imax_n,'noon_elevation']:.2f}°",
                       df_noon.loc[imax_n,'date'].strftime('%d %b'))
            cb2.metric("Elevación mín. al mediodía",
                       f"{df_noon.loc[imin_n,'noon_elevation']:.2f}°",
                       df_noon.loc[imin_n,'date'].strftime('%d %b'))
            cc2.metric("Pasos cenitales",
                       "2 por año" if abs(lat) <= 23.44 else "Ninguno (|lat| > 23.44°)")



# ── Footer ──────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(f"""
<div style="
    display:flex; align-items:center; justify-content:space-between;
    flex-wrap:wrap; gap:10px;
    background:linear-gradient(90deg,#0A1628 0%,#16213E 60%,#0A1628 100%);
    border-radius:12px; padding:14px 22px;
    border-top:3px solid #003E69; border-bottom:1px solid #333;">

  <!-- Izquierda: créditos ACP -->
  <div style="display:flex;flex-direction:column;gap:2px;">
    <span style="font-size:0.95rem;font-weight:700;color:#C8A951;
                 letter-spacing:0.05em;">
      ☀️ Calculadora Solar NOAA
    </span>
    <span style="font-size:0.78rem;color:#aaa;">
      Elaborado para
      <strong style="color:#C8A951;">HIMH</strong>
      &nbsp;·&nbsp; por
      <strong style="color:#4A90D9;">JF</strong>
      &nbsp;·&nbsp;
      <a href="https://gml.noaa.gov/grad/solcalc/"
         style="color:#888;text-decoration:none;" target="_blank">
        gml.noaa.gov/grad/solcalc
      </a>
      &nbsp;·&nbsp; Precisión ±0.01° · 2000–2100
    </span>
  </div>

  <!-- Derecha: contador de visitas -->
  <div style="
      background:#0A1628; border:1px solid #003E69;
      border-radius:10px; padding:8px 18px; text-align:center;
      min-width:130px;">
    <div style="font-size:1.55rem;font-weight:800;
                color:#C8A951;line-height:1.1;">
      {_total_visits:,}
    </div>
    <div style="font-size:0.72rem;color:#888;
                letter-spacing:0.06em;text-transform:uppercase;">
      visitas totales
    </div>
  </div>

</div>
""", unsafe_allow_html=True)
